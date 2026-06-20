 #!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=============================================================================
SISTEMA INTEGRAL PCI - ANALISIS DE FALLAS EN PAVIMENTOS FLEXIBLES
Metodo PCI (ASTM D6433-18)
Modelo YOLOv11 Segmentacion
=============================================================================
Clases del modelo:
    0: BACHE    -> HUECOS
    1: GRIETA   -> GRIETAS LONGITUDINALES Y TRANSVERSALES
    2: PARCHE   -> PARCHEO
    3: PIEL DE COCODRILO
=============================================================================
"""

import cv2
import numpy as np
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from PIL import Image, ImageTk
from pathlib import Path
from threading import Thread
from datetime import datetime
import time
import copy
import json
import math
import io
import os
import shutil
import sys
import tempfile
import warnings
import traceback
import zipfile

warnings.filterwarnings('ignore')

# Dependencias opcionales
try:
    import pandas as pd
    PANDAS_OK = True
except ImportError:
    PANDAS_OK = False

try:
    from ultralytics import YOLO
    YOLO_OK = True
except ImportError:
    YOLO_OK = False

try:
    from skimage.morphology import skeletonize, remove_small_objects
    from skimage.filters import frangi
    SKIMAGE_OK = True
except ImportError:
    SKIMAGE_OK = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# =============================================================================
# CONFIGURACION POR DEFECTO
# =============================================================================

def _directorio_general_base():
    docs = Path.home() / "Documents"
    raiz = docs if docs.exists() else Path.home()
    return raiz / "Analisis_Pavimentos"


DIRECTORIO_GENERAL_BASE = _directorio_general_base()
DIRECTORIOS_GENERALES = {
    "modelos": DIRECTORIO_GENERAL_BASE / "Modelos",
    "imagenes": DIRECTORIO_GENERAL_BASE / "Imagenes",
    "resultados": DIRECTORIO_GENERAL_BASE / "Resultados",
    "proyectos": DIRECTORIO_GENERAL_BASE / "Proyectos",
}

for _dir_general in DIRECTORIOS_GENERALES.values():
    try:
        _dir_general.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass


def _resolver_directorio_dialogo(tipo, ruta_actual=None):
    destino = DIRECTORIOS_GENERALES.get(tipo, DIRECTORIO_GENERAL_BASE)
    raw = str(ruta_actual or "").strip()
    if raw:
        try:
            ruta = Path(raw).expanduser()
            if ruta.exists():
                return str(ruta if ruta.is_dir() else ruta.parent)
            if ruta.suffix and ruta.parent.exists():
                return str(ruta.parent)
        except Exception:
            pass
    try:
        destino.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    return str(destino)

CONFIG_DEFAULT = {
    # Rutas
    'ruta_modelo': '',
    'ruta_imagenes': str(DIRECTORIOS_GENERALES['imagenes']),
    'ruta_salida': str(DIRECTORIOS_GENERALES['resultados']),
    'ruta_proyectos': str(DIRECTORIOS_GENERALES['proyectos']),

    # Calibracion
    'ancho_via_real_m': 6.5,
    'borde_interno_m': 0.30,
    'borde_externo_m': 0.30,
    'usar_borde_berma_pci': True,

    # YOLO
    'confianza_min': 0.1,
    'iou_threshold': 0.45,

    # Clases
    'clases': {0: 'HUECOS', 1: 'GRIETAS LONGITUDINALES Y TRANSVERSALES', 2: 'PARCHEO', 3: 'PIEL DE COCODRILO'},

    # ---- TRAMO Y PROGRESIVAS ----
    'tramo': '',
    'progresiva_inicio': '000+000',
    'progresiva_incremento': 20,  # metros entre imagenes

    # ---- VISUALIZACION ----
    'text_size': 0.45,       # escala de texto en resultados
    'show_numeros': True,    # mostrar valores numericos en etiquetas
    'show_etiquetas': True,  # mostrar etiquetas de fallas
    'show_mallas': True,     # mostrar mallas/esqueleto/circulos piel cocodrilo
    'show_circulos': True,   # mostrar circulos inscritos piel cocodrilo
    'show_poligonos': True,  # mostrar poligonos piel cocodrilo

    # ---- FILTROS MINIMOS POR UNIDAD ----
    'min_diametro_hueco_mm': 50.0,
    'profundidad_asumida_huecos': 'media',
    'min_longitud_grieta_m': 0.05,
    'min_area_parche_m2': 0.01,
    'parche_ratio_leve_max': 0.08,
    'parche_ratio_moderado_max': 0.18,
    'min_area_piel_m2': 0.05,

    # ---- FUSION DE DETECCIONES SOLAPADAS ----
    'merge_iou_threshold': 0.10,
    'merge_distancia_max_px': 50,

    # Piel de cocodrilo - parametros de procesamiento
    'clahe_clip': 4.0,
    'clahe_tile': 8,
    'bilateral_d': 9,
    'bilateral_sigma_color': 75,
    'bilateral_sigma_space': 75,
    'block_size': 23,
    'C_umbral': 10,
    'usar_frangi': True,
    'kernel_apertura': 3,
    'kernel_cierre': 6,
    'iteraciones_cierre': 2,
    'usar_multiescala': True,
    'min_area_poligono': 300,
    'min_circularidad': 0.08,
    'min_vertices': 4,
    'max_vertices': 25,
    'min_radio_circulo': 8,
    'min_longitud_rama': 30,
    'min_area_objeto': 100,
    'max_gap_cierre': 20,
}

# =============================================================================
# ESTILOS UI
# =============================================================================

class EstiloUI:
    BG_DARK = "#e9edf1"
    BG_PANEL = "#f3f5f7"
    BG_CARD = "#e2e8ee"
    BG_INPUT = "#ffffff"
    BG_BUTTON = "#2563eb"
    BG_BUTTON_HOVER = "#1d4ed8"
    BG_BUTTON_SECONDARY = "#d7dee7"
    BG_SUCCESS = "#15803d"
    BG_WARNING = "#b45309"
    BG_ACCENT = "#cfd7e1"
    BORDER = "#c7d0da"

    FG_PRIMARY = "#1f2937"
    FG_SECONDARY = "#5f6b78"
    FG_ACCENT = "#315b86"
    FG_HIGHLIGHT = "#315b86"
    FG_LOG = "#166534"

    FONT_TITLE = ("Segoe UI", 16, "bold")
    FONT_SUBTITLE = ("Segoe UI", 11, "bold")
    FONT_BODY = ("Segoe UI", 10)
    FONT_SMALL = ("Segoe UI", 9)
    FONT_MONO = ("Consolas", 9)
    FONT_LABEL = ("Segoe UI", 9, "bold")

# =============================================================================
# MOTOR DE CALIBRACION
# =============================================================================

class Calibrador:
    """Maneja la calibracion pixel-mm mediante linea dibujada por el usuario."""

    def __init__(self, ancho_via_real_m=6.5):
        self.ancho_via_real_m = ancho_via_real_m
        self.px_por_mm = None
        self.linea_px = None
        self.angulo_eje_via = 90.0  # longitud en pixeles de la linea dibujada
        self.punto_ancho_inicio = None
        self.punto_ancho_fin = None

    def calibrar_con_linea(self, longitud_linea_px):
        """Calibra usando la longitud de linea dibujada."""
        if longitud_linea_px <= 0:
            return None
        self.set_linea_ancho(None, None)
        self.linea_px = longitud_linea_px
        self.px_por_mm = longitud_linea_px / (self.ancho_via_real_m * 1000.0)
        return self.px_por_mm

    def calibrar_con_ancho_imagen(self, ancho_imagen_px):
        """Calibracion usando ancho completo de imagen."""
        self.set_linea_ancho(None, None)
        self.linea_px = ancho_imagen_px
        self.px_por_mm = ancho_imagen_px / (self.ancho_via_real_m * 1000.0)
        return self.px_por_mm

    def calibrar_eje(self, punto_inicio, punto_fin):
        """Define el angulo base de la via para detectar longitudinal/transversal."""
        self.punto_eje_inicio = punto_inicio
        self.punto_eje_fin = punto_fin
        dx = punto_fin[0] - punto_inicio[0]
        dy = punto_fin[1] - punto_inicio[1]
        self.angulo_eje_via = math.degrees(math.atan2(-dy, dx))

    def set_linea_ancho(self, punto_inicio=None, punto_fin=None):
        self.punto_ancho_inicio = tuple(punto_inicio) if punto_inicio is not None else None
        self.punto_ancho_fin = tuple(punto_fin) if punto_fin is not None else None

    def px_a_mm(self, valor_px):
        if self.px_por_mm is None or self.px_por_mm == 0:
            return None
        return valor_px / self.px_por_mm

    def px_a_m(self, valor_px):
        mm = self.px_a_mm(valor_px)
        return mm / 1000.0 if mm is not None else None

    def area_px_a_m2(self, area_px):
        if self.px_por_mm is None or self.px_por_mm == 0:
            return None
        area_mm2 = area_px / (self.px_por_mm ** 2)
        return area_mm2 / 1e6  # mm2 a m2

    def es_longitudinal(self, angulo_grieta_grados):
        diff = abs(angulo_grieta_grados - self.angulo_eje_via) % 180
        if diff > 90: diff = 180 - diff
        return diff < 45


# =============================================================================
# VENTANA DE CALIBRACION CON OPENCV
# =============================================================================

class VentanaCalibracion:
    """Permite al usuario dibujar una linea sobre la imagen para calibrar."""

    def __init__(self, imagen_cv2, ancho_via_real_m=6.5):
        self.imagen = imagen_cv2.copy()
        self.ancho_via_real_m = ancho_via_real_m
        self.punto_inicio = None
        self.punto_fin = None
        self.dibujando = False
        self.longitud_px = None
        self.completado = False
        self.imagen_display = None

    def _mouse_callback(self, event, x, y, flags, param):
        if event == cv2.EVENT_LBUTTONDOWN:
            self.punto_inicio = (x, y)
            self.dibujando = True
        elif event == cv2.EVENT_MOUSEMOVE and self.dibujando:
            self.imagen_display = self.imagen.copy()
            cv2.line(self.imagen_display, self.punto_inicio, (x, y), (0, 255, 255), 2)
            # Mostrar longitud provisional
            dist = np.sqrt((x - self.punto_inicio[0])**2 + (y - self.punto_inicio[1])**2)
            mid = ((self.punto_inicio[0] + x) // 2, (self.punto_inicio[1] + y) // 2 - 15)
            cv2.putText(self.imagen_display, f"{dist:.0f} px", mid,
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255), 2)
        elif event == cv2.EVENT_LBUTTONUP:
            self.punto_fin = (x, y)
            self.dibujando = False
            self.longitud_px = np.sqrt(
                (self.punto_fin[0] - self.punto_inicio[0])**2 +
                (self.punto_fin[1] - self.punto_inicio[1])**2
            )
            self.imagen_display = self.imagen.copy()
            cv2.line(self.imagen_display, self.punto_inicio, self.punto_fin, (0, 255, 0), 3)
            cv2.putText(self.imagen_display,
                        f"Ancho via: {self.longitud_px:.0f} px = {self.ancho_via_real_m} m",
                        (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)
            cv2.putText(self.imagen_display,
                        "ENTER=Confirmar | R=Reintentar | ESC=Usar ancho imagen",
                        (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1)

    def ejecutar(self):
        """Abre ventana OpenCV para dibujar linea de calibracion."""
        # Redimensionar para pantalla
        h, w = self.imagen.shape[:2]
        max_dim = 1200
        scale = min(max_dim / w, max_dim / h, 1.0)
        if scale < 1.0:
            display_w = int(w * scale)
            display_h = int(h * scale)
            self.imagen = cv2.resize(self.imagen, (display_w, display_h))

        self.imagen_display = self.imagen.copy()
        cv2.putText(self.imagen_display,
                    f"Dibuje una linea que represente el ancho de la via ({self.ancho_via_real_m} m)",
                    (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255), 2)
        cv2.putText(self.imagen_display,
                    "Click izq: inicio | Soltar: fin | ENTER: confirmar | ESC: auto",
                    (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1)

        nombre_ventana = "CALIBRACION - Dibuje el ancho de la via"
        cv2.namedWindow(nombre_ventana, cv2.WINDOW_AUTOSIZE)
        cv2.setMouseCallback(nombre_ventana, self._mouse_callback)

        while True:
            cv2.imshow(nombre_ventana, self.imagen_display)
            key = cv2.waitKey(1) & 0xFF

            if key == 13:  # ENTER
                if self.longitud_px is not None and self.longitud_px > 10:
                    self.completado = True
                    break
            elif key == ord('r') or key == ord('R'):
                self.punto_inicio = None
                self.punto_fin = None
                self.longitud_px = None
                self.imagen_display = self.imagen.copy()
                cv2.putText(self.imagen_display,
                            f"Dibuje el ancho de la via ({self.ancho_via_real_m} m)",
                            (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255), 2)
            elif key == 27:  # ESC - usar ancho de imagen completo
                self.longitud_px = self.imagen.shape[1]  # ancho imagen
                self.completado = True
                break

        cv2.destroyWindow(nombre_ventana)
        # Re-escalar longitud si se redimensiono
        if scale < 1.0 and self.longitud_px is not None:
            self.longitud_px = self.longitud_px / scale

        return self.longitud_px


# =============================================================================
# FUSIONADOR DE MASCARAS SOLAPADAS
# =============================================================================

class FusionadorMascaras:
    """
    Fusiona mascaras YOLO de la misma clase que se solapan o estan
    suficientemente cerca (continuidad), para evitar contar la misma falla
    multiples veces.
    """

    @staticmethod
    def fusionar_por_clase(mascaras_por_clase, iou_threshold=0.10, distancia_max_px=50):
        """
        Recibe dict {cls_id: [(mask_resized, conf), ...]}
        Retorna dict {cls_id: [(mask_fusionada, conf_max), ...]}
        donde mascaras solapadas o cercanas se unifican en una sola.
        """
        resultado = {}
        for cls_id, lista in mascaras_por_clase.items():
            if len(lista) <= 1:
                resultado[cls_id] = lista
                continue

            # Crear mascara binaria por cada deteccion
            bins = []
            confs = []
            for mask, conf in lista:
                bins.append((mask > 0.5).astype(np.uint8))
                confs.append(conf)

            # Fusionar iterativamente: Union-Find simplificado
            n = len(bins)
            parent = list(range(n))

            def find(x):
                while parent[x] != x:
                    parent[x] = parent[parent[x]]
                    x = parent[x]
                return x

            def union(a, b):
                ra, rb = find(a), find(b)
                if ra != rb:
                    parent[rb] = ra

            for i in range(n):
                for j in range(i + 1, n):
                    if FusionadorMascaras._deben_fusionar(
                            bins[i], bins[j], iou_threshold, distancia_max_px):
                        union(i, j)

            # Agrupar por componente
            grupos = {}
            for i in range(n):
                r = find(i)
                if r not in grupos:
                    grupos[r] = []
                grupos[r].append(i)

            # Crear mascaras fusionadas
            fusionadas = []
            for indices in grupos.values():
                mask_union = np.zeros_like(bins[0], dtype=np.uint8)
                conf_max = 0.0
                for idx in indices:
                    mask_union = np.bitwise_or(mask_union, bins[idx])
                    conf_max = max(conf_max, confs[idx])
                fusionadas.append((mask_union.astype(np.float32), conf_max))

            resultado[cls_id] = fusionadas

        return resultado

    @staticmethod
    def _deben_fusionar(mask_a, mask_b, iou_threshold, distancia_max_px):
        """Determina si dos mascaras deben fusionarse."""
        # 1. Verificar solapamiento IoU
        intersection = np.sum(np.bitwise_and(mask_a, mask_b))
        union_area = np.sum(np.bitwise_or(mask_a, mask_b))
        if union_area > 0:
            iou = intersection / union_area
            if iou >= iou_threshold:
                return True

        # 2. Verificar si alguno esta contenido en el otro (solapamiento parcial)
        area_a = np.sum(mask_a)
        area_b = np.sum(mask_b)
        if area_a > 0 and intersection / area_a > 0.3:
            return True
        if area_b > 0 and intersection / area_b > 0.3:
            return True

        # 3. Verificar proximidad (continuidad): dilatar una y ver si toca la otra
        if distancia_max_px > 0:
            kernel = cv2.getStructuringElement(
                cv2.MORPH_ELLIPSE,
                (distancia_max_px, distancia_max_px)
            )
            mask_a_dil = cv2.dilate(mask_a, kernel, iterations=1)
            contacto = np.sum(np.bitwise_and(mask_a_dil, mask_b))
            if contacto > 0:
                return True

        return False


# =============================================================================
# PROCESADORES POR TIPO DE FALLA
# =============================================================================

class ProcesadorBaches:
    """Clase 0: HUECOS - Calculo de diametro y severidad segun ASTM D6433-18."""

    NOMBRE_PCI = "HUECOS"

    # Tabla de severidad segun ASTM D6433-18
    # Filas: rango de profundidad maxima (mm)
    # Columnas: rango de diametro promedio (mm)
    # Formato: TABLA_SEVERIDAD[(prof_min, prof_max)][(diam_min, diam_max)] = severidad
    TABLA_SEVERIDAD = {
        (13, 25):           {(100, 200): 'L', (200, 450): 'L', (450, 750): 'M'},
        (25, 50):           {(100, 200): 'M', (200, 450): 'M', (450, 750): 'H'},
        (50, float('inf')): {(100, 200): 'M', (200, 450): 'M', (450, 750): 'H'},
    }

    # Profundidad asumida por defecto (no medible con imagen 2D)
    # Opciones: 'baja' (13-25mm), 'media' (25-50mm), 'alta' (>50mm)
    PROFUNDIDAD_ASUMIDA = 'media'

    @staticmethod
    def procesar(mascara, calibrador, confianza, imagen_original, profundidad_asumida=None):
        """Procesa una mascara de bache y retorna metricas."""
        resultados = []
        mask_bin = (mascara > 0.5).astype(np.uint8)
        mask_full = mask_bin.copy()

        prof = profundidad_asumida or ProcesadorBachesPCI.PROFUNDIDAD_ASUMIDA

        contornos, _ = cv2.findContours(mask_bin * 255, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        for i, cnt in enumerate(contornos):
            area_px = cv2.contourArea(cnt)
            if area_px < 50:
                continue

            diametro_px = 2 * np.sqrt(area_px / np.pi)
            diametro_mm = calibrador.px_a_mm(diametro_px)
            area_m2 = calibrador.area_px_a_m2(area_px)

            M = cv2.moments(cnt)
            if M["m00"] == 0:
                continue
            cx = int(M["m10"] / M["m00"])
            cy = int(M["m01"] / M["m00"])

            mask_individual = np.zeros_like(mask_bin)
            cv2.fillPoly(mask_individual, [cnt], 1)

            severidad = ProcesadorBachesPCI._clasificar_severidad(diametro_mm, prof)

            resultados.append({
                'tipo': ProcesadorBachesPCI.NOMBRE_PCI,
                'id': i + 1,
                'confianza': confianza,
                'diametro_px': diametro_px,
                'diametro_mm': diametro_mm if diametro_mm else 0,
                'area_px': area_px,
                'area_m2': area_m2 if area_m2 else 0,
                'espesor_px': diametro_px,
                'espesor_mm': diametro_mm if diametro_mm else 0,
                'severidad': severidad,
                'profundidad_asumida': prof,
                'ubicacion_x': cx,
                'ubicacion_y': cy,
                'contorno': cnt,
                'mask': mask_individual,
                'mask_full': mask_full,
                'unidad': 'UNIDAD',
            })

        return resultados

    @staticmethod
    def _clasificar_severidad(diametro_mm, profundidad_asumida='media'):
        """
        Clasifica severidad de bache segun tabla ASTM D6433-18.

        La profundidad no es medible con imagen 2D, se asume un rango:
          'baja'  -> 13 a 25 mm
          'media' -> 25 a 50 mm
          'alta'  -> > 50 mm

        Tabla de severidad:
        +-----------------+------------+------------+------------+
        | Profundidad     | 100-200mm  | 200-450mm  | 450-750mm  |
        +-----------------+------------+------------+------------+
        | 13 a <=25 mm    |     L      |     L      |     M      |
        | >25 a <=50 mm   |     M      |     M      |     H      |
        | >50 mm          |     M      |     M      |     H      |
        +-----------------+------------+------------+------------+
        """
        if diametro_mm is None or diametro_mm <= 0:
            return 'M'  # D=0 o sin medida -> severidad Media

        # Seleccionar fila de profundidad
        if profundidad_asumida == 'baja':
            fila_prof = (13, 25)
        elif profundidad_asumida == 'alta':
            fila_prof = (50, float('inf'))
        else:  # 'media' por defecto
            fila_prof = (25, 50)

        # Buscar en la tabla
        rangos_diametro = ProcesadorBachesPCI.TABLA_SEVERIDAD.get(fila_prof, {})

        for (d_min, d_max), severidad in rangos_diametro.items():
            if d_min <= diametro_mm < d_max:
                return severidad

        # Fuera de rango de la tabla
        if diametro_mm < 100:
            # Bache muy pequeno, severidad baja
            return 'L'
        elif diametro_mm >= 750:
            # Bache muy grande, severidad alta
            if profundidad_asumida == 'baja':
                return 'M'
            else:
                return 'H'

        return 'L'

    @staticmethod
    def dibujar(imagen, resultados, text_size=0.45, show_numeros=True, show_etiquetas=True, show_circulos=True):
        vis = imagen.copy()
        for r in resultados:
            cnt = r['contorno']
            sev = r['severidad']
            color = {'L': (0, 255, 0), 'M': (0, 165, 255), 'H': (0, 0, 255)}[sev]

            cv2.drawContours(vis, [cnt], -1, color, 2)
            cx, cy = r['ubicacion_x'], r['ubicacion_y']
            if show_circulos:
                cv2.circle(vis, (cx, cy), int(r['diametro_px'] / 2), color, 2)

            if show_etiquetas:
                titulo = f"HUECO [{sev}]"
                detalle = f"D={r['diametro_mm']:.0f}mm" if show_numeros else ""
                texto_visible_dos_lineas(
                    vis,
                    titulo,
                    detalle,
                    (cx - 80, cy - 25),
                    escala_titulo=text_size,
                    escala_detalle=max(0.35, text_size - 0.03),
                    color=color,
                    grosor_titulo=2,
                    grosor_detalle=1,
                )

        return vis


class ProcesadorGrietas:
    """Clase 1: GRIETAS LONGITUDINALES Y TRANSVERSALES."""

    NOMBRE_PCI = "GRIETAS LONGITUDINALES Y TRANSVERSALES"

    @staticmethod
    def procesar(mascara, calibrador, confianza, imagen_original, merge_dist_px=30):
        """Procesa grietas con transformada de distancia."""
        resultados = []
        mask_bin = (mascara > 0.5).astype(np.uint8)
        mask_original = mask_bin.copy()

        # Unir segmentos cercanos de la misma orientacion, igual que en MTC.
        if merge_dist_px and merge_dist_px > 0:
            ks = int(2 * merge_dist_px + 1)
            k_union = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (ks, ks))
            mask_unida = cv2.dilate(mask_bin, k_union, iterations=1)
            mask_unida = cv2.morphologyEx(mask_unida, cv2.MORPH_CLOSE, k_union, iterations=1)
        else:
            mask_unida = mask_bin

        # Componentes conectados para separar grietas individuales
        num_labels, labels = cv2.connectedComponents(mask_unida)

        for label_id in range(1, num_labels):
            comp_mask = cv2.bitwise_and(mask_original, (labels == label_id).astype(np.uint8))
            area_px = np.sum(comp_mask)
            if area_px < 30:
                continue

            # Transformada de distancia para espesor
            dist_map = cv2.distanceTransform(comp_mask, cv2.DIST_L2, 5)
            max_dist = np.max(dist_map)

            # Punto de maximo espesor (toda la mascara)
            max_loc = np.unravel_index(np.argmax(dist_map), dist_map.shape)
            punto_max_global = (max_loc[1], max_loc[0])  # (x, y)

            # Espesor maximo de toda la mascara (diametro = 2 * radio maximo)
            espesor_total_px = 2 * max_dist
            espesor_total_mm = calibrador.px_a_mm(espesor_total_px)

            # Espesor de zona roja del mapa de calor (mayor precision)
            # En COLORMAP_JET, rojo corresponde a valores >= 75% del maximo
            umbral_rojo = 0.75 * max_dist
            zona_roja = (dist_map >= umbral_rojo).astype(np.uint8)
            if np.any(zona_roja):
                dist_roja = cv2.distanceTransform(zona_roja, cv2.DIST_L2, 5)
                espesor_px = 2 * np.max(dist_roja)  # ancho de la zona roja
                max_loc_roja = np.unravel_index(np.argmax(dist_roja), dist_roja.shape)
                punto_max = (max_loc_roja[1], max_loc_roja[0])
            else:
                espesor_px = espesor_total_px
                punto_max = punto_max_global
            espesor_mm = calibrador.px_a_mm(espesor_px)

            # Longitud de la grieta: usar esqueletizacion
            esq = None
            try:
                esq = skeletonize(comp_mask.astype(bool))
                longitud_px = np.sum(esq)
            except:
                # Fallback: perimetro / 2
                cnts, _ = cv2.findContours(comp_mask * 255, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                longitud_px = sum(cv2.arcLength(c, True) for c in cnts) / 2

            # Calcular direccion perpendicular en punto de maximo espesor
            angulo_perp = ProcesadorGrietasPCI._calcular_angulo_perpendicular(
                esq, punto_max, comp_mask)
            angulo_grieta = ProcesadorGrietasMTC._calcular_angulo(comp_mask)
            es_longitudinal = calibrador.es_longitudinal(angulo_grieta)

            longitud_m = calibrador.px_a_m(longitud_px)
            area_m2 = calibrador.area_px_a_m2(area_px)

            # Severidad PCI basada en espesor de zona naranja (mas preciso)
            severidad = ProcesadorGrietasPCI._clasificar_severidad(espesor_mm)

            resultados.append({
                'tipo': ProcesadorGrietasPCI.NOMBRE_PCI,
                'id': label_id,
                'confianza': confianza,
                'espesor_total_px': espesor_total_px,
                'espesor_total_mm': espesor_total_mm if espesor_total_mm else 0,
                'espesor_px': espesor_px,
                'espesor_mm': espesor_mm if espesor_mm else 0,
                'longitud_px': longitud_px,
                'longitud_m': longitud_m if longitud_m else 0,
                'area_px': area_px,
                'area_m2': area_m2 if area_m2 else 0,
                'severidad': severidad,
                'ubicacion_x': punto_max[0],
                'ubicacion_y': punto_max[1],
                'punto_max': punto_max,
                'punto_max_global': punto_max_global,
                'angulo_perp': angulo_perp,
                'angulo_grieta': angulo_grieta,
                'es_longitudinal': es_longitudinal,
                'distance_map': dist_map,
                'mask': comp_mask,
                'unidad': 'm',
            })

        return resultados

    @staticmethod
    def _clasificar_severidad(espesor_mm):
        if espesor_mm is None:
            return 'L'
        if espesor_mm < 10:
            return 'L'
        elif espesor_mm <= 75:
            return 'M'
        else:
            return 'H'

    @staticmethod
    def _calcular_angulo_perpendicular(esqueleto, punto_max, comp_mask):
        """Calcula el angulo perpendicular a la grieta en el punto de maximo espesor.

        Usa el esqueleto para determinar la direccion local de la grieta y
        devuelve el angulo perpendicular en radianes.
        """
        px, py = punto_max  # (x, y)

        if esqueleto is not None and np.any(esqueleto):
            # Buscar puntos del esqueleto cercanos al punto de maximo espesor
            esq_points = np.argwhere(esqueleto)  # (row, col) = (y, x)
            if len(esq_points) > 2:
                # Distancias al punto maximo
                dists = np.sqrt((esq_points[:, 1] - px)**2 + (esq_points[:, 0] - py)**2)
                # Radio de busqueda adaptivo: entre 10 y 50 px
                radio_busqueda = max(10, min(50, int(np.max(dists) * 0.1)))
                cercanos = esq_points[dists < radio_busqueda]

                if len(cercanos) >= 2:
                    # Ajustar linea por PCA a los puntos cercanos del esqueleto
                    coords = cercanos.astype(np.float64)
                    mean = np.mean(coords, axis=0)
                    coords_centered = coords - mean
                    # PCA: direccion principal
                    cov = np.cov(coords_centered.T)
                    eigenvalues, eigenvectors = np.linalg.eigh(cov)
                    # El eigenvector con mayor eigenvalue es la direccion del esqueleto
                    dir_esqueleto = eigenvectors[:, np.argmax(eigenvalues)]  # (dy, dx)
                    angulo_esq = np.arctan2(dir_esqueleto[0], dir_esqueleto[1])
                    # Perpendicular = +90 grados
                    return angulo_esq + np.pi / 2

        # Fallback: usar gradiente de la transformada de distancia
        # Si no hay esqueleto, asumir horizontal (angulo perpendicular = pi/2 = vertical)
        return np.pi / 2  # linea horizontal por defecto

    @staticmethod
    def dibujar(imagen, resultados, text_size=0.45, show_numeros=True, show_etiquetas=True):
        vis = imagen.copy()
        for r in resultados:
            mask = r.get('_mask_display', r['mask'])
            dist_map = r['distance_map']
            punto = r['punto_max']
            sev = r['severidad']
            color_sev = {'L': (0, 255, 0), 'M': (0, 165, 255), 'H': (0, 0, 255)}[sev]

            # Mapa de calor sobre la grieta
            max_dist = np.max(dist_map)
            if max_dist > 0:
                dist_norm = (dist_map / max_dist * 255).astype(np.uint8)
                heatmap = cv2.applyColorMap(dist_norm, cv2.COLORMAP_JET)
                mask_3ch = np.stack([mask, mask, mask], axis=-1)
                heatmap_masked = heatmap * mask_3ch
                alpha = 0.5
                vis = np.where(mask_3ch > 0,
                               cv2.addWeighted(vis, 1 - alpha, heatmap_masked, alpha, 0),
                               vis)

            # Linea perpendicular de medicion del espesor (ancho de zona naranja)
            radio_px = int(r['espesor_px'] / 2)
            radio_draw = max(radio_px, 3)
            angulo = r.get('angulo_perp', np.pi / 2)

            # Extremos de la linea de medicion perpendicular a la grieta
            dx = int(radio_draw * np.cos(angulo))
            dy = int(radio_draw * np.sin(angulo))
            p1 = (punto[0] - dx, punto[1] - dy)
            p2 = (punto[0] + dx, punto[1] + dy)

            # Linea de medicion del espesor
            cv2.line(vis, p1, p2, color_sev, 2)

            # Marcas de acotado en los extremos
            marca_len = 6
            mx = int(marca_len * np.cos(angulo + np.pi / 2))
            my = int(marca_len * np.sin(angulo + np.pi / 2))
            cv2.line(vis, (p1[0] - mx, p1[1] - my), (p1[0] + mx, p1[1] + my), color_sev, 2)
            cv2.line(vis, (p2[0] - mx, p2[1] - my), (p2[0] + mx, p2[1] + my), color_sev, 2)

            # Punto central
            cv2.circle(vis, punto, 3, color_sev, -1)

            # Etiqueta principal
            if show_etiquetas:
                tipo_label = str(r.get('tipo', ProcesadorGrietasPCI.NOMBRE_PCI))
                titulo = f"{tipo_label} [{sev}]"
                detalle = f"e={r['espesor_mm']:.1f}mm L={r['longitud_m']:.2f}m" if show_numeros else ""
                texto_visible_dos_lineas(
                    vis,
                    titulo,
                    detalle,
                    (punto[0] - 80, punto[1] - radio_draw - 20),
                    escala_titulo=text_size,
                    escala_detalle=max(0.35, text_size - 0.03),
                    color=(0, 255, 255),
                    grosor_titulo=2,
                    grosor_detalle=1,
                )

        return vis


class ProcesadorParches:
    """Clase 2: PARCHEO."""

    NOMBRE_PCI = "PARCHEO"

    @staticmethod
    def procesar(mascara, calibrador, confianza, imagen_original, params=None):
        resultados = []
        mask_bin = (mascara > 0.5).astype(np.uint8)

        contornos, _ = cv2.findContours(mask_bin * 255, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        for i, cnt in enumerate(contornos):
            area_px = cv2.contourArea(cnt)
            if area_px < 50:
                continue

            area_m2 = calibrador.area_px_a_m2(area_px)

            M = cv2.moments(cnt)
            if M["m00"] == 0:
                continue
            cx = int(M["m10"] / M["m00"])
            cy = int(M["m01"] / M["m00"])

            # Severidad basada en estado del parche
            # Analisis de textura dentro del parche para estimar condicion
            comp_mask = np.zeros(mascara.shape[:2], dtype=np.uint8)
            cv2.fillPoly(comp_mask, [cnt], 255)

            severidad, edges_int, ratio_fisuras = ProcesadorParchesPCI._clasificar_severidad(
                comp_mask, imagen_original, calibrador, params=params)

            resultados.append({
                'tipo': ProcesadorParchesPCI.NOMBRE_PCI,
                'id': i + 1,
                'confianza': confianza,
                'espesor_px': 0,
                'espesor_mm': 0,
                'diametro_px': 0,
                'diametro_mm': 0,
                'area_px': area_px,
                'area_m2': area_m2 if area_m2 else 0,
                'severidad': severidad,
                'ubicacion_x': cx,
                'ubicacion_y': cy,
                'contorno': cnt,
                'mascara_parche': comp_mask,
                'unidad': 'm2',
                '_fisuras_mask': edges_int,
                'ratio_fisuras': ratio_fisuras,
            })

        return resultados

    @staticmethod
    def _clasificar_severidad(mask_parche, imagen, calibrador, params=None):
        """Clasifica severidad del parche analizando fisuracion interna."""
        params = params or {}
        ratio_leve, ratio_moderado = normalizar_umbrales_parche(
            params.get("parche_ratio_leve_max", 0.08),
            params.get("parche_ratio_moderado_max", 0.18),
        )
        edges_masked, ratio = medir_relacion_fisuras_parche(mask_parche, imagen)
        severidad = clasificar_severidad_parche_por_ratio(
            ratio,
            metodo="PCI",
            ratio_leve_max=ratio_leve,
            ratio_moderado_max=ratio_moderado,
        )
        return severidad, edges_masked, ratio

    @staticmethod
    def dibujar(imagen, resultados, text_size=0.45, show_numeros=True, show_etiquetas=True):
        vis = imagen.copy()
        for r in resultados:
            cnt = r['contorno']
            sev = r['severidad']
            color = {'L': (0, 255, 0), 'M': (0, 165, 255), 'H': (0, 0, 255)}[sev]

            cv2.drawContours(vis, [cnt], -1, color, 2)

            cx, cy = r['ubicacion_x'], r['ubicacion_y']

            if show_etiquetas:
                titulo = f"PARCHEO [{sev}]"
                detalle = f"A={r['area_m2']:.2f}m2" if show_numeros else ""
                texto_visible_dos_lineas(
                    vis,
                    titulo,
                    detalle,
                    (cx - 60, cy - 20),
                    escala_titulo=text_size,
                    escala_detalle=max(0.35, text_size - 0.03),
                    color=color,
                    grosor_titulo=2,
                    grosor_detalle=1,
                )
        return vis


class ProcesadorPielCocodrilo:
    """Clase 3: PIEL DE COCODRILO - Analisis de malla."""

    NOMBRE_PCI = "PIEL DE COCODRILO"

    def __init__(self, params=None):
        self.params = params or CONFIG_DEFAULT

    def procesar(self, mascara, calibrador, confianza, imagen_original):
        resultados_finales = []
        mask_bin_full = (mascara > 0.5).astype(np.uint8) * 255

        h, w = imagen_original.shape[:2]
        if mask_bin_full.shape[:2] != (h, w):
            mask_bin_full = cv2.resize(mask_bin_full, (w, h))

        # Encontrar trozos desconectados
        contornos, _ = cv2.findContours(mask_bin_full, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        for i, cnt in enumerate(contornos):
            area_msc_px = cv2.contourArea(cnt)
            if area_msc_px < 50: continue

            x, y, w_box, h_box = cv2.boundingRect(cnt)
            # Expandir el bounding box para contexto
            margen = 30
            x1 = max(0, x - margen)
            y1 = max(0, y - margen)
            x2 = min(w, x + w_box + margen)
            y2 = min(h, y + h_box + margen)

            # Mascara aislada para ESTE trozo en la region recortada
            mask_r_crop = np.zeros((y2 - y1, x2 - x1), dtype=np.uint8)
            cnt_crop = cnt.copy() - [x1, y1]
            cv2.fillPoly(mask_r_crop, [cnt_crop], 255)

            # 1. Pipeline para este trozo (procesado solo en el recorte)
            roi_crop = imagen_original[y1:y2, x1:x2]
            roi_crop = cv2.bitwise_and(roi_crop, roi_crop, mask=mask_r_crop)
            gris_crop = cv2.cvtColor(roi_crop, cv2.COLOR_BGR2GRAY)
            limpia_crop, pasos_crop = self._pipeline_imagen(gris_crop, mask_r_crop)

            # 2. Esqueletizar (procesado solo en el recorte)
            if SKIMAGE_OK:
                esq_crop = skeletonize(limpia_crop.astype(bool))
                esq_crop = (esq_crop * 255).astype(np.uint8)
                if self.params.get('usar_refinamiento', True):
                    esq_crop = self._refinar_esqueleto(esq_crop)
                    esq_crop = self._cerrar_gaps(esq_crop)
            else:
                esq_crop = cv2.ximgproc.thinning(limpia_crop) if hasattr(cv2, 'ximgproc') else limpia_crop
            
            # Restaurar resultados a tamano completo para la compatibilidad posterior
            mask_r = np.zeros((h, w), dtype=np.uint8)
            cv2.fillPoly(mask_r, [cnt], 255)

            limpia = np.zeros((h, w), dtype=np.uint8)
            limpia[y1:y2, x1:x2] = limpia_crop

            esq = np.zeros((h, w), dtype=np.uint8)
            esq[y1:y2, x1:x2] = esq_crop

            pasos = {}
            for k, p_crop in pasos_crop.items():
                p_full = np.zeros((h, w), dtype=np.uint8)
                p_full[y1:y2, x1:x2] = p_crop
                pasos[k] = p_full

            # 3. Detectar poligonos (celdas de la malla)
            poligonos = self._detectar_poligonos(esq)

            # 4. Calcular circulos inscritos
            circulos = []
            for pol in poligonos:
                c = self._calcular_circulo_inscrito(pol['contorno_original'])
                if c:
                    circulos.append(c)

            # 5. Calcular diametros de celdas
            if circulos:
                diametros_px = [2 * c['radio'] for c in circulos]
                diametros_mm = [
                    float(calibrador.px_a_mm(diametro_px) or 0)
                    for diametro_px in diametros_px
                ]
                diametro_promedio_px = np.mean(diametros_px)
                diametro_promedio_m = calibrador.px_a_m(diametro_promedio_px)
                diametro_min_px = np.min(diametros_px)
                diametro_min_m = calibrador.px_a_m(diametro_min_px)
            else:
                diametros_mm = []
                diametro_promedio_px = 0
                diametro_promedio_m = 0
                diametro_min_px = 0
                diametro_min_m = 0

            # 6. Area afectada especifica de este trozo
            area_m2 = calibrador.area_px_a_m2(area_msc_px)

            # 7. Severidad basada en diametro PROMEDIO
            if not circulos:
                severidad = 'M'
            else:
                severidad = self._clasificar_severidad(diametro_promedio_m)

            # 8. Centroide del trozo
            cx, cy = 0, 0
            M = cv2.moments(cnt)
            if M["m00"] > 0:
                cx = int(M["m10"] / M["m00"])
                cy = int(M["m01"] / M["m00"])

            resultados_finales.append({
                'tipo': ProcesadorPielCocodrilo.NOMBRE_PCI,
                'id': i + 1,
                'confianza': confianza,
                'espesor_px': diametro_promedio_px,
                'espesor_mm': calibrador.px_a_mm(diametro_promedio_px) if diametro_promedio_px else 0,
                'diametro_px': diametro_promedio_px,
                'diametro_mm': calibrador.px_a_mm(diametro_promedio_px) if diametro_promedio_px else 0,
                'diametro_promedio_px': diametro_promedio_px,
                'diametro_promedio_m': diametro_promedio_m if diametro_promedio_m else 0,
                'diametro_promedio_mm': calibrador.px_a_mm(diametro_promedio_px) if diametro_promedio_px else 0,
                'diametro_min_px': diametro_min_px,
                'diametro_min_m': diametro_min_m if diametro_min_m else 0,
                'diametro_min_mm': calibrador.px_a_mm(diametro_min_px) if diametro_min_px else 0,
                'diametros_celdas_px': [float(d) for d in diametros_px] if circulos else [],
                'diametros_celdas_mm': diametros_mm,
                'area_px': area_msc_px,
                'area_m2': area_m2 if area_m2 else 0,
                'severidad': severidad,
                'ubicacion_x': cx,
                'ubicacion_y': cy,
                'poligonos': poligonos,
                'circulos': circulos,
                'esqueleto': esq,
                'mascara_roi': mask_r,
                'unidad': 'm2',
                'n_celdas': len(circulos),
                'pasos': pasos,
                'limpia': limpia,
            })

        return resultados_finales

    def _pipeline_imagen(self, gris, mascara_roi):
        pasos = {}

        # Contraste CLAHE
        clahe = cv2.createCLAHE(
            clipLimit=self.params.get('clahe_clip', 4.0),
            tileGridSize=(self.params.get('clahe_tile', 8), self.params.get('clahe_tile', 8))
        )
        mejorada = clahe.apply(gris)
        ecualizada = cv2.equalizeHist(mejorada)
        mejorada = cv2.addWeighted(mejorada, 0.7, ecualizada, 0.3, 0)
        pasos['contraste'] = mejorada

        # Frangi (opcional)
        if self.params.get('usar_frangi', True) and SKIMAGE_OK:
            try:
                norm = mejorada.astype(float) / 255.0
                fr = frangi(norm, scale_range=(1, 5), scale_step=1, black_ridges=True)
                frangi_img = (fr * 255).astype(np.uint8)
                mejorada = cv2.addWeighted(mejorada, 0.6, frangi_img, 0.4, 0)
                pasos['frangi'] = frangi_img
            except:
                pass

        # Suavizado bilateral
        suavizada = cv2.bilateralFilter(
            mejorada,
            self.params.get('bilateral_d', 9),
            self.params.get('bilateral_sigma_color', 75),
            self.params.get('bilateral_sigma_space', 75)
        )
        pasos['suavizado'] = suavizada

        # Umbralizacion
        if self.params.get('usar_multiescala', True):
            umbralizada = self._umbralizar_multiescala(suavizada)
        else:
            bs = self.params.get('block_size', 23)
            if bs % 2 == 0:
                bs += 1
            umbralizada = cv2.adaptiveThreshold(
                suavizada, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                cv2.THRESH_BINARY_INV, bs, self.params.get('C_umbral', 10)
            )
        pasos['umbralizada'] = umbralizada

        # Aplicar ROI
        if mascara_roi is not None:
            umbralizada = cv2.bitwise_and(umbralizada, umbralizada, mask=mascara_roi)

        # Morfologia
        k_ap = cv2.getStructuringElement(cv2.MORPH_ELLIPSE,
                                          (self.params.get('kernel_apertura', 3),) * 2)
        abierta = cv2.morphologyEx(umbralizada, cv2.MORPH_OPEN, k_ap)
        k_ci = cv2.getStructuringElement(cv2.MORPH_ELLIPSE,
                                          (self.params.get('kernel_cierre', 6),) * 2)
        cerrada = cv2.morphologyEx(abierta, cv2.MORPH_CLOSE, k_ci,
                                    iterations=self.params.get('iteraciones_cierre', 2))
        pasos['morfologia'] = cerrada

        # Limpiar objetos pequenos
        if SKIMAGE_OK:
            limpia = remove_small_objects(
                cerrada.astype(bool),
                min_size=self.params.get('min_area_objeto', 100)
            ).astype(np.uint8) * 255
        else:
            limpia = cerrada
        pasos['limpia'] = limpia

        return limpia, pasos

    def _umbralizar_multiescala(self, gris):
        escalas = [
            {'block_size': 15, 'C': 8, 'peso': 0.3},
            {'block_size': 23, 'C': 10, 'peso': 0.4},
            {'block_size': 33, 'C': 12, 'peso': 0.3},
        ]
        resultados = []
        for esc in escalas:
            bs = esc['block_size']
            if bs % 2 == 0:
                bs += 1
            u = cv2.adaptiveThreshold(gris, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                       cv2.THRESH_BINARY_INV, bs, esc['C'])
            resultados.append((u * esc['peso']).astype(np.float32))
        combinado = np.clip(np.sum(resultados, axis=0), 0, 255).astype(np.uint8)
        _, final = cv2.threshold(combinado, 127, 255, cv2.THRESH_BINARY)
        return final

    def _refinar_esqueleto(self, esqueleto):
        num_labels, labels = cv2.connectedComponents(esqueleto)
        refinado = np.zeros_like(esqueleto)
        min_long = self.params.get('min_longitud_rama', 30)
        for label in range(1, num_labels):
            comp = (labels == label).astype(np.uint8) * 255
            if np.sum(comp == 255) >= min_long:
                refinado = cv2.bitwise_or(refinado, comp)
        return refinado

    def _cerrar_gaps(self, esqueleto):
        gap = self.params.get('max_gap_cierre', 20)
        kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (gap, gap))
        cerrado = cv2.morphologyEx(esqueleto, cv2.MORPH_CLOSE, kernel, iterations=2)
        if SKIMAGE_OK:
            final = skeletonize(cerrado.astype(bool))
            return (final * 255).astype(np.uint8)
        return cerrado

    def _detectar_poligonos(self, esqueleto):
        kernel = np.ones((3, 3), np.uint8)
        esq_dil = cv2.dilate(esqueleto, kernel, iterations=1)
        contornos, _ = cv2.findContours(esq_dil, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

        poligonos = []
        min_area = self.params.get('min_area_poligono', 300)
        min_circ = self.params.get('min_circularidad', 0.08)
        min_vert = self.params.get('min_vertices', 4)
        max_vert = self.params.get('max_vertices', 25)

        for cnt in contornos:
            area = cv2.contourArea(cnt)
            if area < min_area:
                continue
            per = cv2.arcLength(cnt, True)
            if per == 0:
                continue
            circ = 4 * np.pi * area / (per ** 2)
            if circ < min_circ:
                continue
            approx = cv2.approxPolyDP(cnt, 3, True)
            nv = len(approx)
            if nv < min_vert or nv > max_vert:
                continue
            M = cv2.moments(cnt)
            if M["m00"] == 0:
                continue
            cx = int(M["m10"] / M["m00"])
            cy = int(M["m01"] / M["m00"])
            poligonos.append({
                'contorno': approx,
                'contorno_original': cnt,
                'area': area,
                'perimetro': per,
                'circularidad': circ,
                'vertices': nv,
                'centroide': (cx, cy),
            })
        return poligonos

    def _calcular_circulo_inscrito(self, contorno):
        x, y, w, h = cv2.boundingRect(contorno)
        margen = 10
        mascara = np.zeros((h + 2 * margen, w + 2 * margen), dtype=np.uint8)
        cnt_local = contorno - [x - margen, y - margen]
        cv2.fillPoly(mascara, [cnt_local], 255)
        dist = cv2.distanceTransform(mascara, cv2.DIST_L2, 5)
        _, radio_max, _, centro_local = cv2.minMaxLoc(dist)
        cx = int(centro_local[0] + x - margen)
        cy = int(centro_local[1] + y - margen)
        radio = int(radio_max * 0.85)
        min_r = self.params.get('min_radio_circulo', 8)
        if radio < min_r:
            return None
        return {'centro': (cx, cy), 'radio': radio, 'area_circulo': np.pi * radio ** 2}

    @staticmethod
    def _clasificar_severidad(diametro_m):
        if diametro_m is None or diametro_m == 0:
            return 'H'
        diametro_mm = float(diametro_m) * 1000.0
        if diametro_mm < 150.0:
            return 'H'
        if diametro_mm <= 350.0:
            return 'M'
        return 'L'

    @staticmethod
    def dibujar(imagen, resultados, text_size=0.45, show_numeros=True,
                show_mallas=True, show_circulos=True, show_poligonos=True,
                show_etiquetas=True):
        vis = imagen.copy()
        for r in resultados:
            sev = r['severidad']
            color_sev = {'L': (0, 255, 0), 'M': (0, 165, 255), 'H': (0, 0, 255)}[sev]

            # Contorno ROI
            if 'mascara_roi' in r:
                cnts, _ = cv2.findContours(r['mascara_roi'], cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                cv2.drawContours(vis, cnts, -1, (255, 100, 0), 2)

            # Esqueleto (malla)
            if show_mallas and 'esqueleto' in r:
                k = np.ones((2, 2), np.uint8)
                esq_vis = cv2.dilate(r['esqueleto'], k, iterations=1)
                vis[esq_vis == 255] = (0, 0, 255)

            # Poligonos
            if show_mallas and show_poligonos:
                for pol in r.get('poligonos', []):
                    cv2.drawContours(vis, [pol['contorno']], -1, (0, 255, 255), 2)

            # Circulos inscritos
            if show_circulos:
                for circ in r.get('circulos', []):
                    cv2.circle(vis, circ['centro'], circ['radio'], (255, 0, 255), 2)
                    cv2.circle(vis, circ['centro'], 3, (0, 255, 0), -1)

            # Etiqueta
            if show_etiquetas:
                cx, cy = r['ubicacion_x'], r['ubicacion_y']
                titulo = f"PIEL DE COCODRILO [{sev}]"
                detalle = f"A={r['area_m2']:.2f}m2" if show_numeros else ""
                texto_visible_dos_lineas(
                    vis,
                    titulo,
                    detalle,
                    (cx - 100, cy - 25),
                    escala_titulo=text_size,
                    escala_detalle=max(0.35, text_size - 0.03),
                    color=color_sev,
                    grosor_titulo=2,
                    grosor_detalle=1,
                )

        return vis


# Alias explicitos para preservar la implementacion PCI
# antes de que el archivo redefina procesadores con la variante MTC.
ProcesadorBachesPCI = ProcesadorBaches
ProcesadorGrietasPCI = ProcesadorGrietas
ProcesadorParchesPCI = ProcesadorParches
ProcesadorPielCocodriloPCI = ProcesadorPielCocodrilo


# =============================================================================
# EXPORTADOR EXCEL
# =============================================================================

class ExportadorExcel:
    """Exporta resultados consolidados a Excel."""

    @staticmethod
    def exportar(todos_resultados, ruta_salida, calibrador, tramo='', progresivas=None):
        """
        todos_resultados: dict {nombre_imagen: [lista_de_fallas]}
        tramo: str nombre del tramo
        progresivas: dict {nombre_imagen: str_progresiva} o None
        """
        if not PANDAS_OK:
            print("pandas no disponible, no se puede exportar Excel")
            return None

        datos = []
        for nombre_img, fallas in todos_resultados.items():
            for falla in fallas:
                if falla.get('excluida', False):
                    continue
                datos.append({
                    'Nombre_Imagen': nombre_img,
                    'ID': f"{Path(nombre_img).stem}_{falla['tipo'][:3]}_{falla['id']}",
                    'Tipo_Falla': falla['tipo'],
                    'Unidad': falla.get('unidad', '-'),
                    'Confianza_%': round(falla['confianza'] * 100, 2),
                    'Espesor_Total_px': round(falla.get('espesor_total_px', falla.get('espesor_px', 0)), 2),
                    'Espesor_Total_mm': round(falla.get('espesor_total_mm', falla.get('espesor_mm', 0)), 2),
                    'Espesor_px': round(falla.get('espesor_px', 0), 2),
                    'Espesor_mm': round(falla.get('espesor_mm', 0), 2),
                    'Longitud_px': round(falla.get('longitud_px', 0), 2),
                    'Longitud_m': round(falla.get('longitud_m', 0), 4),
                    'Diametro_mm': round(falla.get('diametro_mm', 0), 2),
                    'Diam_Min_Celda_mm': round(falla.get('diametro_min_mm', 0), 2),
                    'Diam_Prom_Celda_mm': round(falla.get('diametro_promedio_mm', 0), 2),
                    'N_Celdas': falla.get('n_celdas', 0),
                    'Area_px': round(falla.get('area_px', 0), 2),
                    'Area_m2': round(falla.get('area_m2', 0), 4),
                    'Cantidad_Huecos': 1 if falla['tipo'] == 'HUECOS' else 0,
                    'Severidad_PCI': falla.get('severidad', '-'),
                    'Prof_Asumida': falla.get('profundidad_asumida', '-'),
                    'Ubicacion_X': falla.get('ubicacion_x', 0),
                    'Ubicacion_Y': falla.get('ubicacion_y', 0),
                })

        if not datos:
            return None

        df = pd.DataFrame(datos)

        # Estadisticas por tipo
        stats_data = []
        for tipo in df['Tipo_Falla'].unique():
            sub = df[df['Tipo_Falla'] == tipo]
            unidad = sub['Unidad'].iloc[0] if len(sub) > 0 else '-'
            espesores = sub['Espesor_mm'].values
            longitudes = sub['Longitud_m'].values
            areas = sub['Area_m2'].values

            stat = {
                'Tipo_Falla': tipo,
                'Unidad': unidad,
                'Total': len(sub),
                'Severidad_L': len(sub[sub['Severidad_PCI'] == 'L']),
                'Severidad_M': len(sub[sub['Severidad_PCI'] == 'M']),
                'Severidad_H': len(sub[sub['Severidad_PCI'] == 'H']),
            }

            # Metricas segun tipo de falla
            if tipo == 'HUECOS':
                diametros = sub['Diametro_mm'].values
                stat['Diametro_Min_mm'] = round(np.min(diametros), 2) if len(diametros) > 0 else 0
                stat['Diametro_Max_mm'] = round(np.max(diametros), 2) if len(diametros) > 0 else 0
                stat['Diametro_Prom_mm'] = round(np.mean(diametros), 2) if len(diametros) > 0 else 0
            elif 'GRIETA' in tipo:
                # Espesor naranja (usado para severidad)
                stat['Espesor_Min_mm'] = round(np.min(espesores), 2) if len(espesores) > 0 else 0
                stat['Espesor_Max_mm'] = round(np.max(espesores), 2) if len(espesores) > 0 else 0
                stat['Espesor_Prom_mm'] = round(np.mean(espesores), 2) if len(espesores) > 0 else 0
                # Espesor total de mascara
                espesores_total = sub['Espesor_Total_mm'].values
                stat['Espesor_Total_Max_mm'] = round(np.max(espesores_total), 2) if len(espesores_total) > 0 else 0
                stat['Espesor_Total_Prom_mm'] = round(np.mean(espesores_total), 2) if len(espesores_total) > 0 else 0
                stat['Longitud_Total_m'] = round(np.sum(longitudes), 4)
                stat['Longitud_Min_m'] = round(np.min(longitudes), 4) if len(longitudes) > 0 else 0
                stat['Longitud_Max_m'] = round(np.max(longitudes), 4) if len(longitudes) > 0 else 0
                stat['Longitud_Prom_m'] = round(np.mean(longitudes), 4) if len(longitudes) > 0 else 0
            elif 'COCODRILO' in tipo:
                stat['Area_Total_m2'] = round(np.sum(areas), 4)
                stat['Area_Min_m2'] = round(np.min(areas), 4) if len(areas) > 0 else 0
                stat['Area_Max_m2'] = round(np.max(areas), 4) if len(areas) > 0 else 0
                stat['Area_Prom_m2'] = round(np.mean(areas), 4) if len(areas) > 0 else 0
                # Diametro promedio de celdas general
                diam_celdas = sub['Diam_Prom_Celda_mm'].values
                stat['Diam_Prom_Celda_mm'] = round(np.mean(diam_celdas), 2) if len(diam_celdas) > 0 else 0
                stat['Diam_Prom_Celda_Min_mm'] = round(np.min(diam_celdas), 2) if len(diam_celdas) > 0 else 0
                stat['Diam_Prom_Celda_Max_mm'] = round(np.max(diam_celdas), 2) if len(diam_celdas) > 0 else 0
                n_celdas = sub['N_Celdas'].values
                stat['N_Celdas_Total'] = int(np.sum(n_celdas))
                # Diametro promedio de celdas POR SEVERIDAD
                for sev_key in ['L', 'M', 'H']:
                    sub_sev = sub[sub['Severidad_PCI'] == sev_key]
                    if len(sub_sev) > 0:
                        d_sev = sub_sev['Diam_Prom_Celda_mm'].values
                        stat[f'Diam_Prom_{sev_key}_mm'] = round(np.mean(d_sev), 2)
                    else:
                        stat[f'Diam_Prom_{sev_key}_mm'] = 0
            else:  # PARCHEO
                stat['Area_Total_m2'] = round(np.sum(areas), 4)
                stat['Area_Min_m2'] = round(np.min(areas), 4) if len(areas) > 0 else 0
                stat['Area_Max_m2'] = round(np.max(areas), 4) if len(areas) > 0 else 0
                stat['Area_Prom_m2'] = round(np.mean(areas), 4) if len(areas) > 0 else 0

            stats_data.append(stat)

        df_stats = pd.DataFrame(stats_data)

        # Resumen por imagen
        resumen_data = []
        for nombre_img in df['Nombre_Imagen'].unique():
            sub_img = df[df['Nombre_Imagen'] == nombre_img]
            resumen_data.append({
                'Nombre_Imagen': nombre_img,
                'Total_Fallas': len(sub_img),
                'Espesor_Prom_mm': round(sub_img['Espesor_mm'].mean(), 3),
                'Espesor_Min_mm': round(sub_img['Espesor_mm'].min(), 3),
                'Espesor_Max_mm': round(sub_img['Espesor_mm'].max(), 3),
                'Longitud_Total_m': round(sub_img['Longitud_m'].sum(), 4),
                'Area_Total_m2': round(sub_img['Area_m2'].sum(), 4),
                'Confianza_Prom_%': round(sub_img['Confianza_%'].mean(), 2),
            })
        df_resumen = pd.DataFrame(resumen_data)

        # Guardar
        archivo = Path(ruta_salida) / "RESULTADOS_PCI_CONSOLIDADO.xlsx"

        try:
            with pd.ExcelWriter(str(archivo), engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Fallas_Detectadas', index=False)
                df_stats.to_excel(writer, sheet_name='Estadisticas', index=False)
                df_resumen.to_excel(writer, sheet_name='Resumen_por_Imagen', index=False)
                if OPENPYXL_OK:
                    wb = writer.book
                    for sn in wb.sheetnames:
                        ws = wb[sn]
                        hf = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
                        hn = Font(color="FFFFFF", bold=True, size=10)
                        for cell in ws[1]:
                            cell.fill = hf; cell.font = hn; cell.alignment = Alignment(horizontal="center")
                        for col in ws.columns:
                            ml = max(len(str(cell.value or "")) for cell in col)
                            ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml+3, 30)

            return str(archivo)
        except Exception as e:
            # Fallback CSV
            csv_file = Path(ruta_salida) / "RESULTADOS_PCI_CONSOLIDADO.csv"
            df.to_csv(str(csv_file), index=False)
            return str(csv_file)


# =============================================================================
# MOTOR PRINCIPAL PCI
# =============================================================================

class MotorPCI:
    """Motor principal que orquesta todo el analisis PCI."""

    def __init__(self, config=None):
        self.config = config or CONFIG_DEFAULT.copy()
        self.modelo = None
        self.modelo_cargado = False
        self.calibrador = Calibrador(self.config['ancho_via_real_m'])
        self.procesador_piel = ProcesadorPielCocodrilo(self.config)
        self.todos_resultados = {}  # {nombre_img: [fallas]}

    def cargar_modelo(self, ruta):
        try:
            if not YOLO_OK:
                return False, "ultralytics no instalado. pip install ultralytics"
            self.modelo = YOLO(ruta)
            self.modelo_cargado = True
            return True, f"Modelo cargado: {Path(ruta).name}"
        except Exception as e:
            self.modelo_cargado = False
            return False, f"Error: {e}"

    def calibrar_imagen(self, imagen_cv2, usar_gui=True):
        """Calibra usando la imagen. Usa ancho de imagen como fallback."""
        # Ya no se abre ventana OpenCV aqui. La calibracion visual
        # se hace desde _calibrar_unica_ahora() que usa VentanaCalibracionTk.
        self.calibrador.calibrar_con_ancho_imagen(imagen_cv2.shape[1])
        return True

    def procesar_imagen(self, ruta_imagen, callback_log=None, calibrar_gui=False):
        """Procesa una imagen completa con todas las clases.
        1. Agrupa detecciones YOLO por clase
        2. Fusiona mascaras solapadas/cercanas de la misma clase
        3. Procesa cada mascara fusionada
        4. Filtra fallas por tamano minimo segun su unidad
        """
        def log(msg):
            if callback_log:
                callback_log(msg)

        imagen = cv2.imread(str(ruta_imagen))
        if imagen is None:
            log(f"ERROR: No se pudo cargar {ruta_imagen}")
            return None

        alto, ancho = imagen.shape[:2]
        nombre = Path(ruta_imagen).name
        log(f"Procesando: {nombre} ({ancho}x{alto})")

        # Calibrar
        if calibrar_gui:
            self.calibrar_imagen(imagen, usar_gui=True)
        elif self.calibrador.px_por_mm is None:
            self.calibrador.calibrar_con_ancho_imagen(ancho)

        log(f"  Calibracion: {self.calibrador.px_por_mm:.4f} px/mm")

        # YOLO inference
        resultados = self.modelo(
            imagen,
            conf=self.config['confianza_min'],
            iou=self.config.get('iou_threshold', 0.45),
            verbose=False
        )
        resultado = resultados[0]

        if resultado.masks is None:
            log("  Sin detecciones")
            return {'imagen': imagen, 'fallas': [], 'nombre': nombre}

        # =====================================================================
        # PASO 1: Agrupar mascaras por clase
        # =====================================================================
        mascaras_por_clase = {}  # {cls_id: [(mask_resized, conf), ...]}
        n_detecciones_raw = len(resultado.boxes)

        for idx in range(n_detecciones_raw):
            cls_id = int(resultado.boxes.cls[idx])
            conf = float(resultado.boxes.conf[idx])
            mask_raw = resultado.masks.data[idx].cpu().numpy()
            mask_resized = cv2.resize(mask_raw, (ancho, alto))

            if cls_id not in mascaras_por_clase:
                mascaras_por_clase[cls_id] = []
            mascaras_por_clase[cls_id].append((mask_resized, conf))

        # Resumen de detecciones crudas
        for cls_id, lista in mascaras_por_clase.items():
            nombre_clase = self.config['clases'].get(cls_id, f'CLASE_{cls_id}')
            log(f"  YOLO crudo: {nombre_clase} x{len(lista)}")

        # =====================================================================
        # PASO 2: Fusionar mascaras solapadas/cercanas de la misma clase
        # =====================================================================
        iou_merge = self.config.get('merge_iou_threshold', 0.10)
        dist_merge = self.config.get('merge_distancia_max_px', 50)

        mascaras_fusionadas = FusionadorMascaras.fusionar_por_clase(
            mascaras_por_clase, iou_threshold=iou_merge, distancia_max_px=dist_merge)

        for cls_id, lista in mascaras_fusionadas.items():
            nombre_clase = self.config['clases'].get(cls_id, f'CLASE_{cls_id}')
            n_antes = len(mascaras_por_clase.get(cls_id, []))
            n_despues = len(lista)
            if n_antes != n_despues:
                log(f"  Fusion {nombre_clase}: {n_antes} -> {n_despues} (fusionadas {n_antes - n_despues})")

        # =====================================================================
        # PASO 3: Procesar cada mascara fusionada
        # =====================================================================
        fallas_raw = []

        for cls_id, lista_fusionada in mascaras_fusionadas.items():
            nombre_clase = self.config['clases'].get(cls_id, f'CLASE_{cls_id}')

            for mask_f, conf_f in lista_fusionada:
                try:
                    if cls_id == 0:  # BACHE -> HUECOS
                        prof = self.config.get('profundidad_asumida_huecos', 'media')
                        res = ProcesadorBachesPCI.procesar(
                            mask_f, self.calibrador, conf_f, imagen,
                            profundidad_asumida=prof)
                        fallas_raw.extend(res)

                    elif cls_id == 1:  # GRIETA
                        # Restar mascaras de otras clases para que grietas no
                        # interfieran en el calculo de espesor
                        mask_grieta_limpia = mask_f.copy()
                        for otro_cls, otra_lista in mascaras_fusionadas.items():
                            if otro_cls == 1:  # No restar grietas propias
                                continue
                            for otra_mask, _ in otra_lista:
                                otra_bin = (otra_mask > 0.5).astype(np.uint8)
                                mask_grieta_limpia = mask_grieta_limpia * (1 - otra_bin)
                        res = ProcesadorGrietasPCI.procesar(
                            mask_grieta_limpia, self.calibrador, conf_f, imagen)
                        fallas_raw.extend(res)

                    elif cls_id == 2:  # PARCHE
                        res = ProcesadorParchesPCI.procesar(
                            mask_f, self.calibrador, conf_f, imagen)
                        fallas_raw.extend(res)

                    elif cls_id == 3:  # PIEL DE COCODRILO
                        res = self.procesador_piel.procesar(
                            mask_f, self.calibrador, conf_f, imagen)
                        fallas_raw.extend(res)

                except Exception as e:
                    log(f"    ERROR procesando {nombre_clase}: {e}")
                    traceback.print_exc()

        # =====================================================================
        # PASO 4: Filtrar por tamano minimo segun unidad
        # =====================================================================
        min_diametro_hueco = self.config.get('min_diametro_hueco_mm', 50.0)
        min_longitud_grieta = self.config.get('min_longitud_grieta_m', 0.05)
        min_area_parche = self.config.get('min_area_parche_m2', 0.01)
        min_area_piel = self.config.get('min_area_piel_m2', 0.05)

        fallas = []
        descartadas = 0

        for f in fallas_raw:
            tipo = f['tipo']
            categoria = categorizar_tipo_falla(tipo)
            descartar = False
            razon = ""

            if categoria == 'hueco':
                # Filtro por diametro minimo (mm)
                d_mm = f.get('diametro_mm', 0)
                if d_mm < min_diametro_hueco:
                    descartar = True
                    razon = f"D={d_mm:.0f}mm < min {min_diametro_hueco:.0f}mm"

            elif categoria == 'grieta':
                # Filtro por longitud minima (m)
                l_m = f.get('longitud_m', 0)
                if l_m < min_longitud_grieta:
                    descartar = True
                    razon = f"L={l_m:.3f}m < min {min_longitud_grieta:.3f}m"

            elif categoria == 'parche':
                # Filtro por area minima (m2)
                a_m2 = f.get('area_m2', 0)
                if a_m2 < min_area_parche:
                    descartar = True
                    razon = f"A={a_m2:.4f}m2 < min {min_area_parche:.4f}m2"

            elif categoria == 'piel':
                # Filtro por area minima (m2)
                a_m2 = f.get('area_m2', 0)
                if a_m2 < min_area_piel:
                    descartar = True
                    razon = f"A={a_m2:.4f}m2 < min {min_area_piel:.4f}m2"

            if descartar:
                descartadas += 1
            else:
                fallas.append(f)

        # =====================================================================
        # PASO 5: Re-numerar IDs
        # =====================================================================
        conteo_tipo = {}
        for f in fallas:
            t = f['tipo']
            conteo_tipo[t] = conteo_tipo.get(t, 0) + 1
            f['id'] = conteo_tipo[t]

        # Log resumen
        if descartadas > 0:
            log(f"  Filtrado: {descartadas} fragmento(s) descartado(s) por tamano minimo")
        for tipo, n in conteo_tipo.items():
            log(f"  Final: {tipo} x{n}")
        log(f"  Total fallas validas: {len(fallas)}")

        # Guardar en resultados globales
        self.todos_resultados[nombre] = fallas

        return {'imagen': imagen, 'fallas': fallas, 'nombre': nombre}

    def dibujar_resultado(self, imagen, fallas, progresiva=None):
        """Dibuja todas las fallas sobre la imagen con opciones de visualizacion."""
        vis = imagen.copy()
        ts = self.config.get('text_size', 0.45)
        sn = self.config.get('show_numeros', True)
        se = self.config.get('show_etiquetas', True)
        sm = self.config.get('show_mallas', True)
        sc = self.config.get('show_circulos', True)
        sp = sm  # poligonos controlados por show_mallas

        # Separar por tipo usando una categoria comun para las variantes PCI/MTC
        fallas_activas = [f for f in fallas if not f.get('excluida', False)]
        huecos = [f for f in fallas_activas if categorizar_tipo_falla(f.get('tipo')) == 'hueco']
        grietas = [f for f in fallas_activas if categorizar_tipo_falla(f.get('tipo')) == 'grieta']
        parches = [f for f in fallas_activas if categorizar_tipo_falla(f.get('tipo')) == 'parche']
        piel = [f for f in fallas_activas if categorizar_tipo_falla(f.get('tipo')) == 'piel']

        # --- Recorte de grietas: no se superponen a otras fallas ---
        if grietas:
            h, w = imagen.shape[:2]
            mascara_otras = np.zeros((h, w), dtype=np.uint8)
            for f in huecos:
                if 'contorno' in f:
                    cv2.fillPoly(mascara_otras, [f['contorno']], 255)
            for f in parches:
                if 'contorno' in f:
                    cv2.fillPoly(mascara_otras, [f['contorno']], 255)
            for f in piel:
                if 'mascara_roi' in f:
                    mascara_otras = np.bitwise_or(mascara_otras, f['mascara_roi'])
            if np.any(mascara_otras > 0):
                for g in grietas:
                    if 'mask' in g:
                        g['_mask_display'] = g['mask'].copy() & (mascara_otras == 0).astype(np.uint8)

        if huecos:
            vis = ProcesadorBachesPCI.dibujar(
                vis, huecos,
                text_size=ts, show_numeros=sn,
                show_etiquetas=se, show_circulos=sc)
        if grietas:
            vis = ProcesadorGrietasPCI.dibujar(
                vis, grietas,
                text_size=ts, show_numeros=sn,
                show_etiquetas=se)
        if parches:
            vis = ProcesadorParchesPCI.dibujar(
                vis, parches,
                text_size=ts, show_numeros=sn,
                show_etiquetas=se)
        if piel:
            vis = ProcesadorPielCocodrilo.dibujar(vis, piel, text_size=ts, show_numeros=sn,
                                                   show_mallas=sm, show_circulos=sc,
                                                   show_poligonos=sp, show_etiquetas=se)

        # Leyenda
        y_offset = 30
        cv2.putText(vis, "ASTM D6433 | YOLO", (10, y_offset),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
        y_offset += 25

        resumen = {}
        for f in fallas_activas:
            key = f['tipo']
            if key not in resumen:
                resumen[key] = {'total': 0, 'L': 0, 'M': 0, 'H': 0}
            resumen[key]['total'] += 1
            sev = severidad_ui(f.get('severidad'))
            if sev in resumen[key]:
                resumen[key][sev] += 1

        for tipo, counts in resumen.items():
            text = f"{tipo}: {counts['total']} (L:{counts['L']} M:{counts['M']} H:{counts['H']})"
            cv2.putText(vis, text, (10, y_offset),
                        cv2.FONT_HERSHEY_SIMPLEX, ts, (200, 200, 200), 1)
            y_offset += 20

        return vis

    def exportar_excel(self, ruta_salida, tramo='', progresivas=None):
        return ExportadorExcel.exportar(
            self.todos_resultados, ruta_salida, self.calibrador,
            tramo=tramo, progresivas=progresivas)


# =============================================================================
# AUTO-CALIBRAR ZONA PIEL DE COCODRILO
# =============================================================================

class VentanaAutoCalibrarPiel(tk.Toplevel):
    """Ventana Tkinter para seleccionar ROI y auto-calibrar parametros de piel de cocodrilo."""

    def __init__(self, parent, imagen_cv2, titulo="Seleccionar zona"):
        super().__init__(parent)
        self.title(titulo)
        self.configure(bg=EstiloUI.BG_DARK)
        self.geometry("1200x750")
        self.transient(parent)
        self.grab_set()

        self.imagen_original = imagen_cv2.copy()
        self.resultado = None  # dict de params calculados
        self.roi_coords = None  # (x, y, w, h)

        self._zoom = 1.0
        self._pan_x = 0
        self._pan_y = 0
        self._drawing = False
        self._pt1 = None
        self._pt2 = None
        self._pan_start = None

        self._crear_ui()
        self.after(150, self._mostrar_imagen)
        self.protocol("WM_DELETE_WINDOW", self._cancelar)
        self.bind("<Escape>", lambda e: self._cancelar())
        self.wait_window()

    def _crear_ui(self):
        top = tk.Frame(self, bg=EstiloUI.BG_PANEL)
        top.pack(fill="x", padx=5, pady=5)
        tk.Label(top, text="Dibuje un rectangulo sobre la zona de piel de cocodrilo",
                 font=EstiloUI.FONT_SUBTITLE, bg=EstiloUI.BG_PANEL,
                 fg=EstiloUI.FG_HIGHLIGHT).pack(side="left", padx=10)
        self.lbl_info = tk.Label(top, text="Click izq: dibujar | Scroll: zoom | Click der: pan",
                                  font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                                  fg=EstiloUI.FG_SECONDARY)
        self.lbl_info.pack(side="right", padx=10)

        # Contenedor principal: canvas + panel de parÃ¡metros
        main = tk.Frame(self, bg=EstiloUI.BG_DARK)
        main.pack(fill="both", expand=True, padx=5, pady=5)

        self.canvas = tk.Canvas(main, bg=EstiloUI.BG_INPUT, highlightthickness=0)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.canvas.bind("<ButtonPress-1>", self._on_click)
        self.canvas.bind("<B1-Motion>", self._on_drag)
        self.canvas.bind("<ButtonRelease-1>", self._on_release)
        self.canvas.bind("<MouseWheel>", self._on_wheel)
        self.canvas.bind("<ButtonPress-3>", self._on_pan_start)
        self.canvas.bind("<B3-Motion>", self._on_pan_move)
        self.canvas.bind("<ButtonRelease-3>", self._on_pan_end)

        # Panel lateral de parÃ¡metros calculados
        self.panel_params = tk.Frame(main, bg=EstiloUI.BG_PANEL, width=280)
        self.panel_params.pack(side="right", fill="y", padx=(5,0))
        self.panel_params.pack_propagate(False)
        tk.Label(self.panel_params, text="PARAMETROS CALCULADOS",
                 font=EstiloUI.FONT_SUBTITLE, bg=EstiloUI.BG_PANEL,
                 fg=EstiloUI.FG_HIGHLIGHT).pack(pady=(10,5))
        self.lbl_estado = tk.Label(self.panel_params, text="Dibuje un rectangulo primero",
                                    font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                                    fg=EstiloUI.FG_SECONDARY, wraplength=250)
        self.lbl_estado.pack(pady=5)
        # Seleccion de Tipo de Pavimento
        tf = tk.Frame(self.panel_params, bg=EstiloUI.BG_PANEL)
        tf.pack(fill="x", padx=10, pady=(5,10))
        tk.Label(tf, text="Tipo Pavimento:", font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY).pack(side="left")
        self.cb_pavimento = ttk.Combobox(tf, values=["Asfalto Claro / Concreto", "Asfalto Medio / Estandar", "Asfalto Oscuro / Nuevo"],
                                         state="readonly", width=22)
        self.cb_pavimento.current(0)  # Claro/Concreto por defecto
        self.cb_pavimento.pack(side="right")
        # Labels para cada parÃ¡metro
        self._param_labels = {}
        params_info = [
            ('clahe_clip', 'CLAHE Clip'), ('clahe_tile', 'CLAHE Tile'),
            ('bilateral_d', 'Bilateral D'), ('bilateral_sigma_color', 'Sigma Color'),
            ('bilateral_sigma_space', 'Sigma Espacio'), ('block_size', 'Block Size'),
            ('C_umbral', 'Constante C'), ('kernel_apertura', 'Kernel Apert.'),
            ('kernel_cierre', 'Kernel Cierre')
        ]
        for key, label in params_info:
            f = tk.Frame(self.panel_params, bg=EstiloUI.BG_PANEL)
            f.pack(fill="x", padx=10, pady=1)
            tk.Label(f, text=label, font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                     fg=EstiloUI.FG_PRIMARY).pack(side="left")
            lv = tk.Label(f, text="--", font=(EstiloUI.FONT_SMALL[0], 9, "bold"),
                          bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT)
            lv.pack(side="right")
            self._param_labels[key] = lv

        # Barra inferior
        bot = tk.Frame(self, bg=EstiloUI.BG_PANEL)
        bot.pack(fill="x", padx=5, pady=3)
        tk.Button(bot, text="CALCULAR", font=EstiloUI.FONT_SUBTITLE,
                  bg="#f39c12", fg="white", relief="flat", bd=0,
                  padx=15, pady=6, command=self._calcular).pack(side="left", padx=5)
        self.btn_aplicar = tk.Button(bot, text="APLICAR", font=EstiloUI.FONT_SUBTITLE,
                  bg=EstiloUI.BG_SUCCESS, fg="white", relief="flat", bd=0,
                  padx=15, pady=6, command=self._aplicar, state="disabled")
        self.btn_aplicar.pack(side="left", padx=5)
        tk.Button(bot, text="REINICIAR", font=EstiloUI.FONT_SUBTITLE,
                  bg=EstiloUI.BG_BUTTON_SECONDARY, fg="white", relief="flat", bd=0,
                  padx=15, pady=6, command=self._reiniciar).pack(side="left", padx=5)
        self.lbl_zoom = tk.Label(bot, text="Zoom: 1.0x", font=EstiloUI.FONT_SMALL,
                                  bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_SECONDARY)
        self.lbl_zoom.pack(side="left", padx=15)
        tk.Button(bot, text="CANCELAR", font=EstiloUI.FONT_SUBTITLE,
                  bg=EstiloUI.BG_BUTTON, fg="white", relief="flat", bd=0,
                  padx=15, pady=6, command=self._cancelar).pack(side="right", padx=5)
        self.lbl_roi = tk.Label(bot, text="", font=EstiloUI.FONT_BODY,
                                 bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT)
        self.lbl_roi.pack(side="right", padx=10)

    def _mostrar_imagen(self):
        self.update_idletasks()
        cw = self.canvas.winfo_width() or 900
        ch = self.canvas.winfo_height() or 600
        h, w = self.imagen_original.shape[:2]
        self._zoom = min(cw / w, ch / h, 1.0)
        self._pan_x = (cw - w * self._zoom) / 2
        self._pan_y = (ch - h * self._zoom) / 2
        self._render()

    def _render(self):
        self.canvas.delete("all")
        img_src = getattr(self, '_imagen_preview', self.imagen_original)
        h, w = img_src.shape[:2]
        zw, zh = int(w * self._zoom), int(h * self._zoom)
        if zw < 1 or zh < 1: return
        img = cv2.resize(img_src, (zw, zh),
                         interpolation=cv2.INTER_NEAREST if self._zoom > 2 else cv2.INTER_LINEAR)
        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        if self._pt1 and self._pt2:
            p1 = (int(self._pt1[0] * self._zoom), int(self._pt1[1] * self._zoom))
            p2 = (int(self._pt2[0] * self._zoom), int(self._pt2[1] * self._zoom))
            cv2.rectangle(img_rgb, p1, p2, (0, 255, 0), 2)
        pil = Image.fromarray(img_rgb)
        self._tk_img = ImageTk.PhotoImage(pil)
        self.canvas.create_image(int(self._pan_x), int(self._pan_y),
                                  anchor="nw", image=self._tk_img)
        self.lbl_zoom.config(text=f"Zoom: {self._zoom:.1f}x")

    def _canvas_to_img(self, cx, cy):
        return (cx - self._pan_x) / self._zoom, (cy - self._pan_y) / self._zoom

    def _on_click(self, event):
        self._drawing = True
        self._pt1 = self._canvas_to_img(event.x, event.y)
        self._pt2 = self._pt1

    def _on_drag(self, event):
        if self._drawing:
            self._pt2 = self._canvas_to_img(event.x, event.y)
            w = abs(self._pt2[0] - self._pt1[0])
            h = abs(self._pt2[1] - self._pt1[1])
            self.lbl_roi.config(text=f"ROI: {int(w)}x{int(h)} px")
            self._render()

    def _on_release(self, event):
        self._drawing = False
        if self._pt1 and self._pt2:
            w = abs(self._pt2[0] - self._pt1[0])
            h = abs(self._pt2[1] - self._pt1[1])
            self.lbl_roi.config(text=f"ROI: {int(w)}x{int(h)} px - Click CALCULAR")

    def _on_wheel(self, event):
        factor = 1.15 if event.delta > 0 else 1 / 1.15
        old_zoom = self._zoom
        self._zoom = max(0.1, min(10.0, self._zoom * factor))
        self._pan_x = event.x - (event.x - self._pan_x) * (self._zoom / old_zoom)
        self._pan_y = event.y - (event.y - self._pan_y) * (self._zoom / old_zoom)
        self._render()

    def _on_pan_start(self, event):
        self._pan_start = (event.x, event.y)

    def _on_pan_move(self, event):
        if self._pan_start:
            self._pan_x += event.x - self._pan_start[0]
            self._pan_y += event.y - self._pan_start[1]
            self._pan_start = (event.x, event.y)
            self._render()

    def _on_pan_end(self, event):
        self._pan_start = None

    def _calcular(self):
        """Auto-calcula parametros del ROI dibujado y muestra resultados."""
        if not self._pt1 or not self._pt2:
            self.lbl_estado.config(text="Dibuje un rectangulo primero", fg=EstiloUI.FG_SECONDARY)
            return
        x1 = int(min(self._pt1[0], self._pt2[0]))
        y1 = int(min(self._pt1[1], self._pt2[1]))
        x2 = int(max(self._pt1[0], self._pt2[0]))
        y2 = int(max(self._pt1[1], self._pt2[1]))
        h, w = self.imagen_original.shape[:2]
        x1, y1 = max(0, x1), max(0, y1)
        x2, y2 = min(w, x2), min(h, y2)
        rw, rh = x2 - x1, y2 - y1
        if rw < 10 or rh < 10:
            self.lbl_estado.config(text="Rectangulo muy pequeno", fg="#e74c3c")
            return
        self.roi_coords = (x1, y1, rw, rh)
        roi_img = self.imagen_original[y1:y2, x1:x2]
        gris = cv2.cvtColor(roi_img, cv2.COLOR_BGR2GRAY)
        pixels = gris[gris > 0] if np.any(gris > 0) else gris.flatten()
        if len(pixels) < 100:
            self.lbl_estado.config(text="Zona con pocos pixeles", fg="#e74c3c")
            return
        media = np.mean(pixels)
        std = np.std(pixels)
        n_pixels = len(pixels)
        # Auto-calcular parÃ¡metros
        params = {}
        params['clahe_clip'] = 6.0 if std < 30 else (4.0 if std < 50 else 2.5)
        params['clahe_tile'] = 12 if n_pixels > 100000 else (8 if n_pixels > 30000 else 6)
        if std < 25:
            params['bilateral_d'], params['bilateral_sigma_color'], params['bilateral_sigma_space'] = 7, 50, 50
        elif std < 45:
            params['bilateral_d'], params['bilateral_sigma_color'], params['bilateral_sigma_space'] = 9, 75, 75
        else:
            params['bilateral_d'], params['bilateral_sigma_color'], params['bilateral_sigma_space'] = 11, 100, 100
        if media < 80:
            params['block_size'], params['C_umbral'] = 19, 7
        elif media < 140:
            params['block_size'], params['C_umbral'] = 23, 10
        else:
            params['block_size'], params['C_umbral'] = 27, 13
            
        # Ajuste heuristico por tipo de pavimento
        tipo_pav = self.cb_pavimento.get()
        if "Claro" in tipo_pav:
            # Pavimento claro/texturizado = necesita menos sensibilidad al ruido
            params['C_umbral'] += 4
            params['block_size'] += 4
        elif "Oscuro" in tipo_pav:
            # Pavimento oscuro = necesita mas sensibilidad
            params['C_umbral'] = max(2, params['C_umbral'] - 2)
            params['clahe_clip'] = min(8.0, params['clahe_clip'] + 2.0)
        params['kernel_apertura'] = 2 if std < 30 else 3
        params['kernel_cierre'] = 5 if std < 30 else 6
        self._params_calculados = params
        # Ejecutar pipeline con los params calculados para vista previa
        try:
            clahe = cv2.createCLAHE(clipLimit=params['clahe_clip'],
                tileGridSize=(params['clahe_tile'],)*2)
            mej = clahe.apply(gris)
            ecua = cv2.equalizeHist(mej)
            mej = cv2.addWeighted(mej, 0.7, ecua, 0.3, 0)
            # Integracion de Filtro Frangi para vista previa (igual que en motor principal)
            if SKIMAGE_OK:
                try:
                    norm = mej.astype(float)/255.0
                    fr = frangi(norm, scale_range=(1,5), scale_step=1, black_ridges=True)
                    fi = (fr*255).astype(np.uint8)
                    mej = cv2.addWeighted(mej, 0.6, fi, 0.4, 0)
                except: pass
            suav = cv2.bilateralFilter(mej, params['bilateral_d'],
                params['bilateral_sigma_color'], params['bilateral_sigma_space'])
            bs = params['block_size']
            if bs % 2 == 0: bs += 1
            umb = cv2.adaptiveThreshold(suav, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                cv2.THRESH_BINARY_INV, bs, params['C_umbral'])
            k_ap = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (params['kernel_apertura'],)*2)
            ab = cv2.morphologyEx(umb, cv2.MORPH_OPEN, k_ap)
            k_ci = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (params['kernel_cierre'],)*2)
            ce = cv2.morphologyEx(ab, cv2.MORPH_CLOSE, k_ci, iterations=2)
            # Limpiar la imagen original para la vista previa (sin fondo rojo)
            self._imagen_preview = self.imagen_original.copy()
            # Ya no se pinta el fondo ni los contornos rojos, solo el esqueleto y los poligonos
            # Detectar grietas que forman piel de cocodrilo (esqueleto + poligonos)
            n_poligonos = 0
            n_circulos = 0
            try:
                if SKIMAGE_OK and np.sum(ce > 0) > 50:
                    cleaned = remove_small_objects(ce.astype(bool), min_size=80).astype(np.uint8) * 255
                    esq = skeletonize(cleaned.astype(bool))
                    esq_u8 = (esq * 255).astype(np.uint8)
                    
                    if params.get('usar_refinamiento', True):
                        # 1. Refinar esqueleto (eliminar ramas cortas)
                        nl, lb = cv2.connectedComponents(esq_u8)
                        ref = np.zeros_like(esq_u8)
                        for l in range(1, nl):
                            c = (lb == l).astype(np.uint8)*255
                            if np.sum(c == 255) >= params.get('min_longitud_rama', 30):
                                ref = cv2.bitwise_or(ref, c)
                        esq_u8 = ref
                        
                        # 2. Cerrar gaps en el esqueleto
                        g = params.get('max_gap_cierre', 20)
                        k_cg = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (g, g))
                        ce_esq = cv2.morphologyEx(esq_u8, cv2.MORPH_CLOSE, k_cg, iterations=2)
                        esq_u8 = (skeletonize(ce_esq.astype(bool))*255).astype(np.uint8)

                    # Dibujar esqueleto en rojo (grietas) mÃ¡s grueso para que se vea bien
                    esq_coords = np.column_stack(np.where(esq_u8 > 0))
                    for (ey, ex) in esq_coords:
                        py, px = ey + y1, ex + x1
                        if 0 <= py < self._imagen_preview.shape[0] and 0 <= px < self._imagen_preview.shape[1]:
                            # Dibujar un pequeÃ±o cuadrado de 3x3 para que resalte
                            ymin, ymax = max(0, py-1), min(self._imagen_preview.shape[0], py+2)
                            xmin, xmax = max(0, px-1), min(self._imagen_preview.shape[1], px+2)
                            self._imagen_preview[ymin:ymax, xmin:xmax] = [0, 0, 255]
                    # Detectar poligonos cerrados
                    k_dil = np.ones((3, 3), np.uint8)
                    esq_dil = cv2.dilate(esq_u8, k_dil, iterations=1)
                    cnts_p, _ = cv2.findContours(esq_dil, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
                    for cp in cnts_p:
                        a = cv2.contourArea(cp)
                        if a < params.get('min_area_poligono', 300): continue
                        p = cv2.arcLength(cp, True)
                        if p == 0: continue
                        circ = 4 * np.pi * a / (p ** 2)
                        if circ < params.get('min_circularidad', 0.08): continue
                        approx = cv2.approxPolyDP(cp, 3, True)
                        nv = len(approx)
                        if nv < params.get('min_vertices', 4) or nv > params.get('max_vertices', 25): continue
                        n_poligonos += 1
                        # Dibujar poligono en amarillo
                        cp_offset = cp.copy()
                        cp_offset[:, :, 0] += x1
                        cp_offset[:, :, 1] += y1
                        cv2.drawContours(self._imagen_preview, [cp_offset], -1, (0, 255, 255), 1)
                        # Circulo inscrito
                        bx, by, bw, bh = cv2.boundingRect(cp)
                        m = 5
                        mask_c = np.zeros((bh+2*m, bw+2*m), dtype=np.uint8)
                        cv2.fillPoly(mask_c, [cp - [bx-m, by-m]], 255)
                        dist_c = cv2.distanceTransform(mask_c, cv2.DIST_L2, 5)
                        _, rm, _, cl = cv2.minMaxLoc(dist_c)
                        r_circ = int(rm * 0.85)
                        if r_circ >= 5:
                            n_circulos += 1
                            centro = (int(cl[0]+bx-m+x1), int(cl[1]+by-m+y1))
                            cv2.circle(self._imagen_preview, centro, r_circ, (255, 0, 255), 1)
            except Exception:
                pass
        except Exception:
            self._imagen_preview = self.imagen_original.copy()
            n_poligonos = 0
            n_circulos = 0
        self._render()
        # Mostrar en panel
        for key, lbl in self._param_labels.items():
            v = params.get(key, '--')
            lbl.config(text=f"{v:.1f}" if isinstance(v, float) else str(v))
        self.lbl_estado.config(
            text=f"ROI: {rw}x{rh}px | Media: {media:.0f} | Std: {std:.0f} | Grietas: {n_poligonos} poligonos, {n_circulos} celdas",
            fg=EstiloUI.BG_SUCCESS)
        self.btn_aplicar.config(state="normal")

    def _aplicar(self):
        """Acepta los parametros calculados."""
        if hasattr(self, '_params_calculados'):
            self.resultado = self._params_calculados
            self.destroy()

    def _reiniciar(self):
        self._pt1 = None
        self._pt2 = None
        self.roi_coords = None
        self._imagen_preview = self.imagen_original
        self.lbl_roi.config(text="")
        self.lbl_estado.config(text="Dibuje un rectangulo primero", fg=EstiloUI.FG_SECONDARY)
        self.btn_aplicar.config(state="disabled")
        for lbl in self._param_labels.values():
            lbl.config(text="--")
        self._render()

    def _cancelar(self):
        self.resultado = None
        self.destroy()


# =============================================================================
# FUSIONADOR DE MÃSCARAS
# =============================================================================
class FusionadorMascaras:
    @staticmethod
    def fusionar_por_clase(mascaras_por_clase, iou_threshold=0.10, distancia_max_px=50):
        resultado = {}
        for cls_id, lista in mascaras_por_clase.items():
            if len(lista) <= 1:
                resultado[cls_id] = lista; continue
            bins = [(m > 0.5).astype(np.uint8) for m, c in lista]
            confs = [c for m, c in lista]
            n = len(bins)
            parent = list(range(n))
            def find(x):
                while parent[x] != x: parent[x] = parent[parent[x]]; x = parent[x]
                return x
            def union(a, b):
                ra, rb = find(a), find(b)
                if ra != rb: parent[rb] = ra
            for i in range(n):
                for j in range(i+1, n):
                    if FusionadorMascaras._deben_fusionar(bins[i], bins[j], iou_threshold, distancia_max_px):
                        union(i, j)
            grupos = {}
            for i in range(n):
                r = find(i)
                if r not in grupos: grupos[r] = []
                grupos[r].append(i)
            fusionadas = []
            for indices in grupos.values():
                mask_union = np.zeros_like(bins[0], dtype=np.uint8)
                conf_max = 0.0
                for idx in indices:
                    mask_union = np.bitwise_or(mask_union, bins[idx])
                    conf_max = max(conf_max, confs[idx])
                fusionadas.append((mask_union.astype(np.float32), conf_max))
            resultado[cls_id] = fusionadas
        return resultado

    @staticmethod
    def _deben_fusionar(mask_a, mask_b, iou_threshold, distancia_max_px):
        inter = np.sum(np.bitwise_and(mask_a, mask_b))
        union_a = np.sum(np.bitwise_or(mask_a, mask_b))
        if union_a > 0 and inter / union_a >= iou_threshold: return True
        area_a, area_b = np.sum(mask_a), np.sum(mask_b)
        if area_a > 0 and inter / area_a > 0.3: return True
        if area_b > 0 and inter / area_b > 0.3: return True
        if distancia_max_px > 0:
            kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (distancia_max_px, distancia_max_px))
            mask_a_dil = cv2.dilate(mask_a, kernel, iterations=1)
            if np.sum(np.bitwise_and(mask_a_dil, mask_b)) > 0: return True
        return False


# =============================================================================
# TEXTO CON CONTORNO PARA MAYOR VISIBILIDAD
# =============================================================================
def texto_visible(img, texto, pos, escala=0.50, color=(0,255,255), grosor=2, factor=1.0):
    """Dibuja texto con color directo. factor multiplica escala y grosor."""
    x, y = pos
    esc = escala * factor
    gr = max(1, int(grosor * factor))
    cv2.putText(img, texto, (x, y), cv2.FONT_HERSHEY_SIMPLEX, esc, color, gr)


def texto_visible_dos_lineas(
    img,
    titulo,
    detalle,
    pos,
    escala_titulo=0.50,
    escala_detalle=0.45,
    color=(0, 255, 255),
    grosor_titulo=2,
    grosor_detalle=1,
    factor=1.0,
    interlineado=20,
):
    """Dibuja una etiqueta en dos lineas y ajusta su posicion dentro de la imagen."""
    titulo = str(titulo or "")
    detalle = str(detalle or "")
    if not titulo:
        return

    h, w = img.shape[:2]
    font = cv2.FONT_HERSHEY_SIMPLEX
    esc_titulo = escala_titulo * factor
    esc_detalle = escala_detalle * factor
    gr_titulo = max(1, int(grosor_titulo * factor))
    gr_detalle = max(1, int(grosor_detalle * factor))

    size_titulo, _ = cv2.getTextSize(titulo, font, esc_titulo, gr_titulo)
    if detalle:
        size_detalle, _ = cv2.getTextSize(detalle, font, esc_detalle, gr_detalle)
    else:
        size_detalle = (0, 0)

    ancho_max = max(size_titulo[0], size_detalle[0])
    salto = max(size_titulo[1] + 6, size_detalle[1] + 6, int(interlineado * factor))
    extra_y = salto if detalle else 0

    x = int(max(5, min(pos[0], w - ancho_max - 5)))
    y = int(pos[1])
    y_min = size_titulo[1] + 5
    y_max = h - extra_y - 5
    y = max(y_min, min(y, y_max))

    texto_visible(img, titulo, (x, y), escala_titulo, color, grosor_titulo, factor)
    if detalle:
        texto_visible(img, detalle, (x, y + salto), escala_detalle, color, grosor_detalle, factor)


def _normalizar_rectangulo(rect):
    if rect is None:
        return None
    x1, y1, x2, y2 = [int(v) for v in rect]
    if x2 < x1:
        x1, x2 = x2, x1
    if y2 < y1:
        y1, y2 = y2, y1
    return (x1, y1, x2, y2)


def _area_solape_rect(rect_a, rect_b, margen=0):
    rect_a = _normalizar_rectangulo(rect_a)
    rect_b = _normalizar_rectangulo(rect_b)
    if rect_a is None or rect_b is None:
        return 0

    ax1, ay1, ax2, ay2 = rect_a
    bx1, by1, bx2, by2 = rect_b
    ax1 -= margen
    ay1 -= margen
    ax2 += margen
    ay2 += margen
    bx1 -= margen
    by1 -= margen
    bx2 += margen
    by2 += margen

    inter_w = max(0, min(ax2, bx2) - max(ax1, bx1))
    inter_h = max(0, min(ay2, by2) - max(ay1, by1))
    return inter_w * inter_h


def _rectangulo_mascara(mask):
    if mask is None:
        return None
    mask_u8 = ((mask > 0) * 255).astype(np.uint8)
    pts = cv2.findNonZero(mask_u8)
    if pts is None:
        return None
    x, y, w, h = cv2.boundingRect(pts)
    return (x, y, x + w, y + h)


def _obtener_rectangulo_falla(resultado, margen=6):
    rect = None

    contorno = resultado.get('contorno')
    if contorno is not None and len(contorno) > 0:
        x, y, w, h = cv2.boundingRect(contorno)
        rect = (x, y, x + w, y + h)

    if rect is None:
        for key in ('mask', 'mascara_roi', 'mascara_parche'):
            rect = _rectangulo_mascara(resultado.get(key))
            if rect is not None:
                break

    if rect is None and resultado.get('circulos'):
        xs1, ys1, xs2, ys2 = [], [], [], []
        for circ in resultado.get('circulos', []):
            centro = circ.get('centro')
            radio = int(circ.get('radio', 0) or 0)
            if not centro:
                continue
            xs1.append(int(centro[0] - radio))
            ys1.append(int(centro[1] - radio))
            xs2.append(int(centro[0] + radio))
            ys2.append(int(centro[1] + radio))
        if xs1:
            rect = (min(xs1), min(ys1), max(xs2), max(ys2))

    if rect is None:
        cx = int(resultado.get('ubicacion_x', 0) or 0)
        cy = int(resultado.get('ubicacion_y', 0) or 0)
        radio = max(
            int((resultado.get('diametro_px', 0) or 0) / 2),
            int((resultado.get('espesor_px', 0) or 0) / 2),
            10,
        )
        rect = (cx - radio, cy - radio, cx + radio, cy + radio)

    x1, y1, x2, y2 = rect
    return (x1 - margen, y1 - margen, x2 + margen, y2 + margen)


def _medir_etiqueta_dos_lineas(
    titulo,
    detalle,
    escala_titulo=0.50,
    escala_detalle=0.45,
    grosor_titulo=2,
    grosor_detalle=1,
    factor=1.0,
    interlineado=20,
):
    font = cv2.FONT_HERSHEY_SIMPLEX
    esc_titulo = escala_titulo * factor
    esc_detalle = escala_detalle * factor
    gr_titulo = max(1, int(grosor_titulo * factor))
    gr_detalle = max(1, int(grosor_detalle * factor))

    size_titulo, base_titulo = cv2.getTextSize(str(titulo or ""), font, esc_titulo, gr_titulo)
    if detalle:
        size_detalle, base_detalle = cv2.getTextSize(str(detalle or ""), font, esc_detalle, gr_detalle)
    else:
        size_detalle, base_detalle = (0, 0), 0

    salto = max(size_titulo[1] + 6, size_detalle[1] + 6, int(interlineado * factor))
    alto = size_titulo[1] + base_titulo + (salto + base_detalle if detalle else 0) + 2
    ancho = max(size_titulo[0], size_detalle[0])
    return {
        'ancho': ancho,
        'alto': alto,
        'alto_titulo': size_titulo[1],
        'salto': salto,
    }


def _ajustar_caja_a_imagen(x, y, ancho, alto, w_img, h_img, margen=5):
    max_x = max(margen, w_img - ancho - margen)
    max_y = max(margen, h_img - alto - margen)
    x = int(max(margen, min(x, max_x)))
    y = int(max(margen, min(y, max_y)))
    return x, y


def texto_visible_dos_lineas_cercano(
    img,
    titulo,
    detalle,
    anchor,
    occupied_rects=None,
    avoid_rects=None,
    escala_titulo=0.50,
    escala_detalle=0.45,
    color=(0, 255, 255),
    grosor_titulo=2,
    grosor_detalle=1,
    factor=1.0,
    interlineado=20,
    separacion=8,
    padding=5,
):
    titulo = str(titulo or "")
    detalle = str(detalle or "")
    if not titulo:
        return None

    obstaculos = []
    for rect in avoid_rects or []:
        rect_n = _normalizar_rectangulo(rect)
        if rect_n is not None:
            obstaculos.append(rect_n)

    h_img, w_img = img.shape[:2]
    medida = _medir_etiqueta_dos_lineas(
        titulo,
        detalle,
        escala_titulo=escala_titulo,
        escala_detalle=escala_detalle,
        grosor_titulo=grosor_titulo,
        grosor_detalle=grosor_detalle,
        factor=factor,
        interlineado=interlineado,
    )
    ancho = medida['ancho']
    alto = medida['alto']

    ax = int(anchor[0])
    ay = int(anchor[1])
    if obstaculos:
        ox1 = min(rect[0] for rect in obstaculos)
        oy1 = min(rect[1] for rect in obstaculos)
        ox2 = max(rect[2] for rect in obstaculos)
        oy2 = max(rect[3] for rect in obstaculos)
    else:
        ox1 = ox2 = ax
        oy1 = oy2 = ay

    x = int((ox1 + ox2 - ancho) / 2)
    y_top = int(oy1 - separacion - alto)
    x, y_top = _ajustar_caja_a_imagen(x, y_top, ancho, alto, w_img, h_img, margen=padding)
    rect = (x - padding, y_top - padding, x + ancho + padding, y_top + alto + padding)

    baseline_y = y_top + medida['alto_titulo']
    texto_visible(img, titulo, (x, baseline_y), escala_titulo, color, grosor_titulo, factor)
    if detalle:
        texto_visible(img, detalle, (x, baseline_y + medida['salto']), escala_detalle, color, grosor_detalle, factor)

    return rect


def categorizar_tipo_falla(tipo):
    """Normaliza nombres de fallas entre variantes PCI/MTC."""
    tipo_txt = str(tipo or "").upper()
    if "COCODRILO" in tipo_txt:
        return "piel"
    if "PARCHE" in tipo_txt or "REPARACION" in tipo_txt or "PARCHADO" in tipo_txt:
        return "parche"
    if "GRIETA" in tipo_txt or "FISURA" in tipo_txt:
        return "grieta"
    if "HUECO" in tipo_txt or "BACHE" in tipo_txt:
        return "hueco"
    return "otro"


def severidad_ui(severidad):
    """Convierte severidades 1/2/3 o L/M/H a la representacion L/M/H usada en la UI."""
    mapa = {1: "L", 2: "M", 3: "H", "1": "L", "2": "M", "3": "H"}
    if severidad in mapa:
        return mapa[severidad]
    if isinstance(severidad, str):
        sev = severidad.strip().upper()
        if sev in ("L", "M", "H"):
            return sev
    return "-"


def normalizar_umbrales_parche(ratio_leve_max=0.08, ratio_moderado_max=0.18):
    """Asegura umbrales consistentes para clasificar severidad de parcheos."""
    try:
        ratio_leve = float(ratio_leve_max)
    except Exception:
        ratio_leve = 0.08
    try:
        ratio_moderado = float(ratio_moderado_max)
    except Exception:
        ratio_moderado = 0.18

    ratio_leve = max(0.0, min(1.0, ratio_leve))
    ratio_moderado = max(ratio_leve, min(1.0, ratio_moderado))
    return ratio_leve, ratio_moderado


def medir_relacion_fisuras_parche(mask_parche, imagen):
    """Calcula la relacion entre bordes/fisuras internas y area total del parche."""
    gris = cv2.cvtColor(imagen, cv2.COLOR_BGR2GRAY)
    roi = cv2.bitwise_and(gris, gris, mask=mask_parche)
    edges = cv2.Canny(roi, 50, 150)
    edges_masked = cv2.bitwise_and(edges, edges, mask=mask_parche)

    area_parche = int(np.sum(mask_parche > 0))
    area_fisuras = int(np.sum(edges_masked > 0))
    if area_parche <= 0:
        return edges_masked, 0.0
    return edges_masked, float(area_fisuras / area_parche)


def clasificar_severidad_parche_por_ratio(ratio, metodo="PCI", ratio_leve_max=0.08, ratio_moderado_max=0.18):
    """Clasifica severidad de parcheos desde la relacion de area fisurada."""
    ratio_leve, ratio_moderado = normalizar_umbrales_parche(ratio_leve_max, ratio_moderado_max)
    try:
        ratio_val = float(ratio)
    except Exception:
        ratio_val = 0.0
    ratio_val = max(0.0, ratio_val)

    if ratio_val <= ratio_leve:
        base = 1
    elif ratio_val <= ratio_moderado:
        base = 2
    else:
        base = 3

    if str(metodo or "").upper() == "PCI":
        return {1: "L", 2: "M", 3: "H"}[base]
    return base


# =============================================================================
# PROCESADOR BACHES (HUECOS) - CLASE 0
# =============================================================================
class ProcesadorBaches:
    NOMBRE_PCI = "HUECOS"
    NOMBRE_MTC = "BACHES (HUECOS)"

    @staticmethod
    def procesar(mascara, calibrador, confianza, imagen_original, params=None):
        resultados = []
        mask_bin = (mascara > 0.5).astype(np.uint8)
        contornos, _ = cv2.findContours(mask_bin * 255, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        for i, cnt in enumerate(contornos):
            area_px = cv2.contourArea(cnt)
            if area_px < 50: continue
            diametro_px = 2 * np.sqrt(area_px / np.pi)
            diametro_mm = calibrador.px_a_mm(diametro_px)
            diametro_m = (diametro_mm / 1000.0) if diametro_mm else 0
            area_m2 = calibrador.area_px_a_m2(area_px)
            M = cv2.moments(cnt)
            if M["m00"] == 0: continue
            cx, cy = int(M["m10"]/M["m00"]), int(M["m01"]/M["m00"])
            mask_ind = np.zeros_like(mask_bin); cv2.fillPoly(mask_ind, [cnt], 1)
            severidad = ProcesadorBaches._clasificar_severidad(diametro_m)
            resultados.append({
                'tipo': ProcesadorBaches.NOMBRE_MTC, 'id': i+1, 'confianza': confianza,
                'diametro_px': diametro_px, 'diametro_mm': diametro_mm or 0,
                'area_px': area_px, 'area_m2': area_m2 or 0,
                'espesor_px': diametro_px, 'espesor_mm': diametro_mm or 0,
                'longitud_px': 0, 'longitud_m': 0,
                'severidad': severidad, 'ubicacion_x': cx, 'ubicacion_y': cy,
                'contorno': cnt, 'mask': mask_ind, 'unidad': 'UNIDAD',
            })
        return resultados

    @staticmethod
    def _clasificar_severidad(diametro_m):
        if diametro_m is None or diametro_m <= 0: return 1
        if diametro_m < 0.2: return 1
        elif diametro_m <= 0.5: return 2
        else: return 3

    @staticmethod
    def dibujar(imagen, resultados, mostrar_mallas=True, mostrar_etiquetas=True, mostrar_numeros=True, escala_texto=1.0, mostrar_circulos=True, occupied_labels=None):
        vis = imagen.copy()
        colores = {1: (0,255,0), 2: (0,165,255), 3: (0,0,255)}
        etiquetas_ocupadas = occupied_labels if occupied_labels is not None else []
        for r in resultados:
            cnt, sev = r['contorno'], r['severidad']
            color = colores.get(sev, (255,255,255))
            cv2.drawContours(vis, [cnt], -1, color, 2)
            cx, cy = r['ubicacion_x'], r['ubicacion_y']
            radio = max(int(r['diametro_px']/2), 5)
            es_danio_puntual = str(r.get('tipo', '')) == "DAÑOS PUNTUALES"
            if mostrar_mallas and not es_danio_puntual:
                cv2.line(vis, (cx - radio, cy), (cx + radio, cy), color, 1)
                cv2.line(vis, (cx, cy - radio), (cx, cy + radio), color, 1)
            if mostrar_circulos and not es_danio_puntual:
                cv2.circle(vis, (cx, cy), radio, color, 2)
                cv2.circle(vis, (cx, cy), 3, (0, 255, 0), -1)
            if mostrar_etiquetas:
                tipo_label = str(r.get('tipo', ProcesadorBaches.NOMBRE_MTC))
                titulo = f"{tipo_label} G:{sev}"
                if mostrar_numeros:
                    if es_danio_puntual:
                        detalle = f"A={r.get('area_m2', 0):.3f}m2"
                    else:
                        detalle = f"D={r['diametro_mm']:.0f}mm ({r.get('diametro_mm',0)/1000:.2f}m)"
                else:
                    detalle = ""
                texto_visible_dos_lineas(
                    vis,
                    titulo,
                    detalle,
                    (cx - 80, cy - 25),
                    escala_titulo=0.50,
                    escala_detalle=0.45,
                    color=color,
                    grosor_titulo=2,
                    grosor_detalle=1,
                    factor=escala_texto,
                )
        return vis

    @staticmethod
    def generar_pasos(mascara, imagen_original):
        """Genera imÃ¡genes de pasos intermedios para baches"""
        pasos = {}
        mask_bin = (mascara > 0.5).astype(np.uint8) * 255
        pasos['mascara_binaria'] = mask_bin
        gris = cv2.cvtColor(imagen_original, cv2.COLOR_BGR2GRAY)
        roi = cv2.bitwise_and(gris, gris, mask=mask_bin)
        pasos['roi_gris'] = roi
        contornos, _ = cv2.findContours(mask_bin, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        vis_cnt = imagen_original.copy()
        for cnt in contornos:
            area = cv2.contourArea(cnt)
            if area < 50: continue
            cv2.drawContours(vis_cnt, [cnt], -1, (0, 255, 0), 2)
            diam = 2 * np.sqrt(area / np.pi)
            M = cv2.moments(cnt)
            if M["m00"] > 0:
                cx, cy = int(M["m10"]/M["m00"]), int(M["m01"]/M["m00"])
                cv2.circle(vis_cnt, (cx, cy), max(int(diam/2), 5), (0, 255, 255), 2)
        pasos['contornos_diametros'] = vis_cnt
        dist = cv2.distanceTransform(mask_bin, cv2.DIST_L2, 5)
        if np.max(dist) > 0:
            dn = (dist / np.max(dist) * 255).astype(np.uint8)
            pasos['distancia'] = cv2.applyColorMap(dn, cv2.COLORMAP_JET)
        return pasos

# =============================================================================
# PROCESADOR GRIETAS (FISURAS) - CLASE 1
# =============================================================================
class ProcesadorGrietas:
    NOMBRE_PCI = "GRIETAS LONGITUDINALES Y TRANSVERSALES"

    @staticmethod
    def procesar(mascara, calibrador, confianza, imagen_original, min_rama_px=30, merge_dist_px=30):
        resultados = []
        mask_bin = (mascara > 0.5).astype(np.uint8)
        mask_original = mask_bin.copy()
        # Cierre morfologico para unir fisuras cercanas del mismo tipo
        if merge_dist_px > 0:
            ks = 2 * merge_dist_px + 1  # kernel impar
            k_union = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (ks, ks))
            mask_unida = cv2.dilate(mask_bin, k_union, iterations=1)
            mask_unida = cv2.morphologyEx(mask_unida, cv2.MORPH_CLOSE, k_union, iterations=1)
        else:
            mask_unida = mask_bin
        # Usar la mascara expandida SOLO para determinar conectividad
        num_labels, labels = cv2.connectedComponents(mask_unida)
        for label_id in range(1, num_labels):
            # Intersectar con mascara original para obtener los pixeles reales
            comp_mask = cv2.bitwise_and(mask_original, (labels == label_id).astype(np.uint8))
            area_px = np.sum(comp_mask)
            if area_px < 30: continue
            r = ProcesadorGrietas._procesar_submask(comp_mask, calibrador, confianza, label_id)
            if r is not None:
                resultados.append(r)
        return resultados

    @staticmethod
    def _separar_ramas(comp_mask, min_rama_px=30):
        """Separa un componente conectado en ramas si tiene puntos de ramificacion.
        Retorna lista de sub-mascaras, una por rama."""
        if not SKIMAGE_OK:
            return [comp_mask]
        try:
            esq = skeletonize(comp_mask.astype(bool)).astype(np.uint8)
        except:
            return [comp_mask]
        if np.sum(esq) < 10:
            return [comp_mask]
        # Detectar branch points: pixeles del esqueleto con >= 3 vecinos en el esqueleto
        kernel_cross = np.array([[1,1,1],[1,10,1],[1,1,1]], dtype=np.uint8)
        vecinos = cv2.filter2D(esq, -1, kernel_cross)
        # branch_points: pixeles que son parte del esqueleto (valor >= 10) y tienen >= 3 vecinos
        branch_points = ((vecinos - 10 * esq) >= 3) & (esq > 0)
        if not np.any(branch_points):
            # Sin ramificaciones: retornar componente original
            return [comp_mask]
        # Eliminar branch points y una vecindad pequeÃ±a para separar bien las ramas
        bp_mask = branch_points.astype(np.uint8)
        k_dil = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
        bp_dilated = cv2.dilate(bp_mask, k_dil, iterations=1)
        esq_sin_bp = esq.copy()
        esq_sin_bp[bp_dilated > 0] = 0
        # Encontrar componentes conectados en el esqueleto sin branch points
        n_ramas, labels_ramas = cv2.connectedComponents(esq_sin_bp)
        if n_ramas <= 2:
            # Solo 1 rama real (o ninguna): retornar original
            return [comp_mask]
        # Para cada rama del esqueleto, expandir para recuperar la mascara real
        sub_masks = []
        # Dilatar cada rama del esqueleto para cubrir el ancho real de la grieta
        dist_map = cv2.distanceTransform(comp_mask, cv2.DIST_L2, 5)
        max_espesor = int(np.max(dist_map)) + 3
        k_expand = cv2.getStructuringElement(cv2.MORPH_ELLIPSE,
            (2 * max_espesor + 1, 2 * max_espesor + 1))
        for rama_id in range(1, n_ramas):
            rama_esq = (labels_ramas == rama_id).astype(np.uint8)
            # Filtrar ramas muy pequeÃ±as
            if np.sum(rama_esq) < min_rama_px:
                continue
            # Expandir el esqueleto de la rama para cubrir toda la grieta
            rama_expandida = cv2.dilate(rama_esq, k_expand, iterations=1)
            # Intersectar con la mascara original del componente
            sub_mask = cv2.bitwise_and(comp_mask, rama_expandida)
            if np.sum(sub_mask) < 30:
                continue
            sub_masks.append(sub_mask)
        # Si no se generaron sub-mascaras validas, retornar original
        if not sub_masks:
            return [comp_mask]
        # Resolver solapamientos: asignar pixeles solapados a la rama mas cercana
        if len(sub_masks) > 1:
            # Crear mapa de asignacion por distancia al esqueleto de cada rama
            asignacion = np.full(comp_mask.shape, -1, dtype=np.int32)
            dist_min = np.full(comp_mask.shape, np.inf, dtype=np.float64)
            for i, sub in enumerate(sub_masks):
                # Distancia desde cada pixel al esqueleto de esta rama
                rama_esq_i = (labels_ramas == (i + 1)).astype(np.uint8)
                # Necesitamos filtrar las ramas pequeÃ±as que se saltaron
                if np.sum(rama_esq_i) < min_rama_px:
                    continue
                inv = 1 - rama_esq_i
                if np.sum(inv) == 0:
                    continue
                d = cv2.distanceTransform(inv, cv2.DIST_L2, 5)
                mascara_comp = comp_mask > 0
                mejor = (d < dist_min) & mascara_comp
                dist_min[mejor] = d[mejor]
                asignacion[mejor] = i
            # Reconstruir sub_masks limpias sin solapamiento
            sub_masks_limpias = []
            for i in range(len(sub_masks)):
                clean = ((asignacion == i) & (comp_mask > 0)).astype(np.uint8)
                if np.sum(clean) >= 30:
                    sub_masks_limpias.append(clean)
            if sub_masks_limpias:
                sub_masks = sub_masks_limpias
        return sub_masks

    @staticmethod
    def _procesar_submask(comp_mask, calibrador, confianza, label_id):
        """Procesa una sub-mascara individual y retorna el diccionario de resultado."""
        area_px = np.sum(comp_mask)
        if area_px < 30:
            return None
        dist_map = cv2.distanceTransform(comp_mask, cv2.DIST_L2, 5)
        max_val = np.max(dist_map)
        # Punto de maximo espesor (toda la mascara)
        max_loc = np.unravel_index(np.argmax(dist_map), dist_map.shape)
        punto_max_global = (max_loc[1], max_loc[0])
        if max_val > 0:
            # Espesor de zona roja del mapa de calor (mayor precision)
            # En COLORMAP_JET, rojo corresponde a valores >= 75% del maximo
            umbral_rojo = 0.75 * max_val
            zona_roja = (dist_map >= umbral_rojo).astype(np.uint8)
            if np.any(zona_roja):
                dist_roja = cv2.distanceTransform(zona_roja, cv2.DIST_L2, 5)
                espesor_px = 2 * np.max(dist_roja)  # ancho de la zona roja
                max_loc_roja = np.unravel_index(np.argmax(dist_roja), dist_roja.shape)
                punto_max = (max_loc_roja[1], max_loc_roja[0])
                radio_max_px = int(np.max(dist_roja))
            else:
                espesor_px = 2 * max_val
                punto_max = punto_max_global
                radio_max_px = int(max_val)
        else:
            espesor_px = 0; punto_max = (0, 0); radio_max_px = 0
        espesor_mm = calibrador.px_a_mm(espesor_px)
        try:
            if SKIMAGE_OK:
                esq = skeletonize(comp_mask.astype(bool))
                longitud_px = float(np.sum(esq))
            else:
                cnts, _ = cv2.findContours(comp_mask*255, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                longitud_px = sum(cv2.arcLength(c, True) for c in cnts) / 2
        except:
            cnts, _ = cv2.findContours(comp_mask*255, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            longitud_px = sum(cv2.arcLength(c, True) for c in cnts) / 2
        longitud_m = calibrador.px_a_m(longitud_px) or 0
        angulo = ProcesadorGrietas._calcular_angulo(comp_mask)
        es_long = calibrador.es_longitudinal(angulo)
        tipo_grieta = "FISURA LONGITUDINAL" if es_long else "FISURA TRANSVERSAL"
        severidad = ProcesadorGrietas._clasificar_severidad(espesor_mm)
        area_mtc = ProcesadorGrietas._calcular_area_mtc(longitud_m, severidad, es_long)
        return {
            'tipo': tipo_grieta, 'id': label_id, 'confianza': confianza,
            'espesor_px': espesor_px, 'espesor_mm': espesor_mm or 0,
            'radio_max_px': radio_max_px,
            'longitud_px': longitud_px, 'longitud_m': longitud_m,
            'diametro_px': 0, 'diametro_mm': 0,
            'area_px': area_px, 'area_m2': area_mtc,
            'severidad': severidad, 'es_longitudinal': es_long,
            'angulo': angulo,
            'ubicacion_x': punto_max[0], 'ubicacion_y': punto_max[1],
            'punto_max': punto_max, 'distance_map': dist_map,
            'mask': comp_mask, 'unidad': 'm2',
        }

    @staticmethod
    def _calcular_angulo(mask):
        coords = np.column_stack(np.where(mask > 0))
        if len(coords) < 5: return 0
        try:
            mean = np.mean(coords, axis=0)
            cov = np.cov((coords - mean).T)
            eigenvalues, eigenvectors = np.linalg.eigh(cov)
            principal = eigenvectors[:, np.argmax(eigenvalues)]
            return math.degrees(math.atan2(-principal[0], principal[1]))
        except: return 0

    @staticmethod
    def _clasificar_severidad(espesor_mm):
        if espesor_mm is None or espesor_mm <= 0: return 1
        if espesor_mm <= 1.0: return 1
        elif espesor_mm <= 3.0: return 2
        else: return 3

    @staticmethod
    def _calcular_area_mtc(longitud_m, severidad, es_longitudinal):
        """Area MTC = Longitud x Factor segun severidad
        Fisuras longitudinales y transversales:
        Sev 1: x 0.10m | Sev 2: x 0.30m | Sev 3: x 0.50m"""
        if longitud_m is None or longitud_m <= 0: return 0.0
        anchos = {1: 0.10, 2: 0.30, 3: 0.50}
        return longitud_m * anchos.get(severidad, 0.10)

    @staticmethod
    def dibujar(imagen, resultados, mostrar_etiquetas=True, mostrar_numeros=True, escala_texto=1.0, occupied_labels=None):
        vis = imagen.copy()
        colores_sev = {1: (0, 255, 0), 2: (0, 165, 255), 3: (0, 0, 255)}  # Verde, Naranja, Rojo
        etiquetas_ocupadas = occupied_labels if occupied_labels is not None else []
        for r in resultados:
            mask, dist_map = r['mask'], r['distance_map']
            punto, sev = r['punto_max'], r['severidad']
            color_sev = colores_sev.get(sev, (0, 165, 255))
            max_val = np.max(dist_map)
            if max_val > 0:
                dist_norm = (dist_map / max_val * 255).astype(np.uint8)
                heatmap = cv2.applyColorMap(dist_norm, cv2.COLORMAP_JET)
                mask_3ch = np.stack([mask]*3, axis=-1)
                vis = np.where(mask_3ch > 0,
                               cv2.addWeighted(vis, 0.5, heatmap * mask_3ch, 0.5, 0), vis)
            # Radio del espesor (zona naranja = 75% del max)
            radio_real = r.get('radio_max_px', max(int(r['espesor_px'] / 2), 3))
            radio_real = max(radio_real, 3)
            # Color naranja fijo para el sÃ­mbolo del espesor
            color_bar = (0, 165, 255)  # Naranja BGR
            # Barra centrada en el punto mÃ¡ximo (rojo), abarcando de naranja a naranja
            punto_bar = punto
            # Ãngulo: barra cruza el ancho (perpendicular a la grieta), ticks paralelos a la grieta
            angulo_grieta = r.get('angulo', 0)
            # DirecciÃ³n perpendicular a la grieta (cruza el espesor)
            ang_perp_rad = math.radians(angulo_grieta) + math.pi / 2
            dx = int(radio_real * math.cos(ang_perp_rad))
            dy = int(radio_real * math.sin(ang_perp_rad))
            p1 = (punto_bar[0] - dx, punto_bar[1] - dy)
            p2 = (punto_bar[0] + dx, punto_bar[1] + dy)
            # Barra "I" cruza el espesor en color naranja
            cv2.line(vis, p1, p2, color_bar, 3)
            # Ticks en extremos: paralelos a la grieta
            tick_len = 6
            ang_grieta_rad = math.radians(angulo_grieta)
            tdx = int(tick_len * math.cos(ang_grieta_rad))
            tdy = int(tick_len * math.sin(ang_grieta_rad))
            cv2.line(vis, (p1[0]-tdx, p1[1]-tdy), (p1[0]+tdx, p1[1]+tdy), color_bar, 2)
            cv2.line(vis, (p2[0]-tdx, p2[1]-tdy), (p2[0]+tdx, p2[1]+tdy), color_bar, 2)
            if mostrar_etiquetas:
                tipo_c = "FISURA LONGITUDINAL" if r.get('es_longitudinal') else "FISURA TRANSVERSAL"
                titulo = f"{tipo_c} [G:{sev}]"
                detalle = f"e={r['espesor_mm']:.1f}mm L={r['longitud_m']:.2f}m" if mostrar_numeros else ""
                texto_visible_dos_lineas(
                    vis,
                    titulo,
                    detalle,
                    (punto[0] - 80, punto[1] - radio_real - 20),
                    escala_titulo=0.45,
                    escala_detalle=max(0.35, 0.45 - 0.03),
                    color=(0, 255, 255),
                    grosor_titulo=2,
                    grosor_detalle=1,
                    factor=escala_texto,
                )
        return vis

    @staticmethod
    def generar_pasos(mascara, imagen_original, calibrador):
        pasos = {}
        mask_bin = (mascara > 0.5).astype(np.uint8)
        pasos['mascara_binaria'] = mask_bin * 255
        dist = cv2.distanceTransform(mask_bin, cv2.DIST_L2, 5)
        if np.max(dist) > 0:
            dn = (dist / np.max(dist) * 255).astype(np.uint8)
            pasos['mapa_distancia'] = cv2.applyColorMap(dn, cv2.COLORMAP_JET)
        if SKIMAGE_OK:
            esq = skeletonize(mask_bin.astype(bool))
            esq_vis = np.zeros_like(imagen_original)
            esq_vis[esq] = (0, 255, 0)
            base = imagen_original.copy()
            base[esq] = (0, 255, 0)
            pasos['esqueleto'] = base
        heatmap_full = imagen_original.copy()
        if np.max(dist) > 0:
            dn2 = (dist / np.max(dist) * 255).astype(np.uint8)
            hm = cv2.applyColorMap(dn2, cv2.COLORMAP_JET)
            m3 = np.stack([mask_bin]*3, axis=-1)
            heatmap_full = np.where(m3 > 0, cv2.addWeighted(heatmap_full, 0.4, hm * m3, 0.6, 0), heatmap_full)
        pasos['mapa_calor'] = heatmap_full
        return pasos

# =============================================================================
# PROCESADOR PARCHES - CLASE 2
# =============================================================================
class ProcesadorParches:
    NOMBRE_PCI = "PARCHEO"
    NOMBRE_MTC = "REPARACIONES O PARCHADOS"

    @staticmethod
    def procesar(mascara, calibrador, confianza, imagen_original, params=None):
        resultados = []
        mask_bin = (mascara > 0.5).astype(np.uint8)
        contornos, _ = cv2.findContours(mask_bin*255, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        for i, cnt in enumerate(contornos):
            area_px = cv2.contourArea(cnt)
            if area_px < 50: continue
            M = cv2.moments(cnt)
            if M["m00"] == 0: continue
            cx, cy = int(M["m10"]/M["m00"]), int(M["m01"]/M["m00"])
            area_m2 = calibrador.area_px_a_m2(area_px) or 0
            comp_mask = np.zeros(mascara.shape[:2], dtype=np.uint8)
            cv2.fillPoly(comp_mask, [cnt], 255)
            
            # Severidad basada en estado del parche
            severidad, edges_int, ratio_fisuras = ProcesadorParches._clasificar_severidad(
                comp_mask, imagen_original, params=params)

            resultados.append({
                'tipo': ProcesadorParches.NOMBRE_MTC, 'id': i+1, 'confianza': confianza,
                'espesor_px': 0, 'espesor_mm': 0,
                'diametro_px': 0, 'diametro_mm': 0,
                'longitud_px': 0, 'longitud_m': 0,
                'ancho_px': 0, 'ancho_m': 0,
                'area_px': area_px, 'area_m2': area_m2,
                'severidad': severidad, 'ubicacion_x': cx, 'ubicacion_y': cy,
                'contorno': cnt, 'unidad': 'm2',
                'esqueleto': edges_int, 'mascara_parche': comp_mask,
                'ratio_fisuras': ratio_fisuras,
            })
        return resultados

    @staticmethod
    def _clasificar_severidad(mask_parche, imagen, params=None):
        """Clasifica severidad del parche analizando fisuracion interna."""
        params = params or {}
        ratio_leve, ratio_moderado = normalizar_umbrales_parche(
            params.get("parche_ratio_leve_max", 0.08),
            params.get("parche_ratio_moderado_max", 0.18),
        )
        edges_masked, ratio = medir_relacion_fisuras_parche(mask_parche, imagen)
        severidad = clasificar_severidad_parche_por_ratio(
            ratio,
            metodo="MTC",
            ratio_leve_max=ratio_leve,
            ratio_moderado_max=ratio_moderado,
        )
        return severidad, edges_masked, ratio

    @staticmethod
    def dibujar(imagen, resultados, mostrar_mallas=True, mostrar_etiquetas=True, mostrar_numeros=True, escala_texto=1.0, mostrar_circulos=True, occupied_labels=None):
        vis = imagen.copy()
        colores = {1: (0,255,0), 2: (0,165,255), 3: (0,0,255)}
        etiquetas_ocupadas = occupied_labels if occupied_labels is not None else []
        for r in resultados:
            cnt, sev = r['contorno'], r['severidad']
            color = colores.get(sev, (255,255,255))
            cv2.drawContours(vis, [cnt], -1, color, 2)
            if mostrar_circulos:
                for circ in r.get('circulos', []):
                    cv2.circle(vis, circ['centro'], circ['radio'], (255, 0, 255), 1)
                    cv2.circle(vis, circ['centro'], 2, (0, 255, 0), -1)
            if mostrar_etiquetas:
                cx, cy = r['ubicacion_x'], r['ubicacion_y']
                detalle = f"A={r['area_m2']:.2f}m2" if mostrar_numeros else ""
                texto_visible_dos_lineas(
                    vis,
                    f"REPARACIONES O PARCHADOS G:{sev}",
                    detalle,
                    (cx - 60, cy - 20),
                    escala_titulo=0.50,
                    escala_detalle=0.45,
                    color=color,
                    grosor_titulo=2,
                    grosor_detalle=1,
                    factor=escala_texto,
                )
        return vis

    @staticmethod
    def generar_pasos(mascara, imagen_original, params=None):
        pasos = {}
        mask_bin = (mascara > 0.5).astype(np.uint8) * 255
        pasos['mascara_binaria'] = mask_bin
        
        roi_color = cv2.bitwise_and(imagen_original, imagen_original, mask=mask_bin)
        pasos['roi_color'] = roi_color
        
        gris = cv2.cvtColor(imagen_original, cv2.COLOR_BGR2GRAY)
        roi_gris = cv2.bitwise_and(gris, gris, mask=mask_bin)
        pasos['roi_gris'] = roi_gris
        
        edges = cv2.Canny(roi_gris, 50, 150)
        edges_masked = cv2.bitwise_and(edges, edges, mask=mask_bin)
        
        # Para visualizaciÃ³n en pasos
        edges_vis = cv2.cvtColor(edges_masked, cv2.COLOR_GRAY2BGR)
        edges_vis[edges_masked > 0] = (0, 0, 255)  # Bordes en rojo
        pasos['bordes_canny'] = edges_vis
        
        # Superponer bordes sobre la imagen original recortada para vis final de pasos
        resultado_mix = roi_color.copy()
        resultado_mix[edges_masked > 0] = (0, 0, 255)
        pasos['resultado_final'] = resultado_mix
        
        return pasos

# =============================================================================
# EXPORTADOR EXCEL
# =============================================================================
class ExportadorExcel:
    @staticmethod
    def exportar(todos_resultados, ruta_salida, calibrador, nombre_archivo="RESULTADOS_MTC_CONSOLIDADO.xlsx"):
        if not PANDAS_OK: return None
        datos = []
        for nombre_img, fallas in todos_resultados.items():
            for f in fallas:
                datos.append({
                    'Nombre_Imagen': nombre_img,
                    'ID': f"{Path(nombre_img).stem}_{f['tipo'][:3]}_{f['id']}",
                    'Tipo_Falla': f['tipo'], 'Confianza_%': round(f['confianza']*100, 2),
                    'Unidad_Medida': f.get('unidad', '-'),
                    'Espesor_px': round(f.get('espesor_px', 0), 2),
                    'Espesor_mm': round(f.get('espesor_mm', 0), 2),
                    'Diametro_px': round(f.get('diametro_px', 0), 2),
                    'Diametro_mm': round(f.get('diametro_mm', 0), 2),
                    'Longitud_px': round(f.get('longitud_px', 0), 2),
                    'Longitud_m': round(f.get('longitud_m', 0), 4),
                    'Area_px': round(f.get('area_px', 0), 2),
                    'Area_m2': round(f.get('area_m2', 0), 4),
                    'Conteo': 1,
                    'Severidad_MTC': f.get('severidad', '-'),
                    'Ubicacion_X': f.get('ubicacion_x', 0),
                    'Ubicacion_Y': f.get('ubicacion_y', 0),
                })
        if not datos: return None
        df = pd.DataFrame(datos)
        # EstadÃ­sticas
        stats = []
        for tipo in df['Tipo_Falla'].unique():
            sub = df[df['Tipo_Falla'] == tipo]
            s = {'Tipo_Falla': tipo, 'Unidad_Medida': sub['Unidad_Medida'].iloc[0],
                 'Total': len(sub), 'Sev_1_Leve': len(sub[sub['Severidad_MTC']==1]),
                 'Sev_2_Moderado': len(sub[sub['Severidad_MTC']==2]),
                 'Sev_3_Severo': len(sub[sub['Severidad_MTC']==3])}
            for col, label in [('Espesor_mm','Espesor'), ('Longitud_m','Longitud'), ('Area_m2','Area')]:
                vals = sub[col].values
                nz = vals[vals > 0]
                if len(nz) > 0:
                    s[f'{label}_Min'] = round(np.min(nz), 4)
                    s[f'{label}_Max'] = round(np.max(nz), 4)
                    s[f'{label}_Prom'] = round(np.mean(nz), 4)
                    s[f'{label}_Total'] = round(np.sum(nz), 4)
            stats.append(s)
        df_stats = pd.DataFrame(stats)
        # Resumen por imagen
        resumen = df.groupby('Nombre_Imagen').agg({
            'ID': 'count', 'Espesor_mm': ['mean', 'min', 'max'],
            'Longitud_m': 'sum', 'Area_m2': 'sum', 'Confianza_%': 'mean'
        }).round(3)
        resumen.columns = ['Total_Fallas', 'Espesor_Prom_mm', 'Espesor_Min_mm',
                           'Espesor_Max_mm', 'Longitud_Total_m', 'Area_Total_m2', 'Confianza_Prom_%']
        archivo = Path(ruta_salida) / nombre_archivo
        try:
            with pd.ExcelWriter(str(archivo), engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Fallas_Detectadas', index=False)
                df_stats.to_excel(writer, sheet_name='Estadisticas_MTC', index=False)
                resumen.to_excel(writer, sheet_name='Resumen_por_Imagen')
                if OPENPYXL_OK:
                    wb = writer.book
                    for sn in wb.sheetnames:
                        ws = wb[sn]
                        hf = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
                        hn = Font(color="FFFFFF", bold=True, size=10)
                        for cell in ws[1]:
                            cell.fill = hf; cell.font = hn; cell.alignment = Alignment(horizontal="center")
                        for col in ws.columns:
                            ml = max(len(str(cell.value or "")) for cell in col)
                            ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml+3, 30)
            return str(archivo)
        except Exception as e:
            csv_f = Path(ruta_salida) / nombre_archivo.replace('.xlsx', '.csv')
            df.to_csv(str(csv_f), index=False); return str(csv_f)

# =============================================================================
# MOTOR PRINCIPAL MTC
# =============================================================================
class MotorMTC:
    def __init__(self, config=None):
        self.config = config or copy.deepcopy(CONFIG_DEFAULT)
        self.modelo = None; self.modelo_cargado = False
        self.calibrador = Calibrador(self.config['ancho_via_real_m'])
        self.procesador_piel = ProcesadorPielCocodrilo(self.config)
        self.todos_resultados = {}  # {nombre_img: [fallas]}
        self.todos_pasos = {}       # {nombre_img: {clase: pasos_dict}}

    def cargar_modelo(self, ruta):
        try:
            if not YOLO_OK: return False, "ultralytics no instalado"
            self.modelo = YOLO(ruta); self.modelo_cargado = True
            return True, f"Modelo cargado: {Path(ruta).name}"
        except Exception as e:
            self.modelo_cargado = False; return False, f"Error: {e}"

    def calibrar_imagen(self, imagen_cv2, usar_gui=True, parent=None):
        if usar_gui:
            vent = VentanaCalibracion(imagen_cv2, self.config['ancho_via_real_m'], parent=parent)
            longitud, angulo_eje, pt_ini, pt_fin = vent.ejecutar()
            if longitud: self.calibrador.calibrar_con_linea(longitud)
            if pt_ini and pt_fin: self.calibrador.set_eje_via(pt_ini, pt_fin)
            else: self.calibrador.angulo_eje_via = angulo_eje
            return True
        self.calibrador.calibrar_con_ancho_imagen(imagen_cv2.shape[1]); return True

    def procesar_imagen(self, ruta_imagen, callback_log=None, calibrar_gui=False, parent=None):
        def log(msg):
            if callback_log: callback_log(msg)
        imagen = cv2.imread(str(ruta_imagen))
        if imagen is None: log(f"ERROR: No se pudo cargar {ruta_imagen}"); return None
        alto, ancho = imagen.shape[:2]
        nombre = Path(ruta_imagen).name
        log(f"Procesando: {nombre} ({ancho}x{alto} px)")
        if calibrar_gui: self.calibrar_imagen(imagen, usar_gui=True, parent=parent)
        elif self.calibrador.px_por_mm is None: self.calibrador.calibrar_con_ancho_imagen(ancho)
        log(f"  Calibracion: {self.calibrador.px_por_mm:.4f} px/mm | Eje: {self.calibrador.angulo_eje_via:.1f} grados")

        resultados = self.modelo(imagen, conf=self.config['confianza_min'],
                                  iou=self.config.get('iou_threshold', 0.45), verbose=False)
        resultado = resultados[0]
        if resultado.masks is None:
            log("  Sin detecciones")
            return {'imagen': imagen, 'fallas': [], 'nombre': nombre, 'ancho': ancho, 'alto': alto}

        mascaras_por_clase = {}
        for idx in range(len(resultado.boxes)):
            cls_id = int(resultado.boxes.cls[idx]); conf = float(resultado.boxes.conf[idx])
            mask_raw = resultado.masks.data[idx].cpu().numpy()
            mask_r = cv2.resize(mask_raw, (ancho, alto))
            if cls_id not in mascaras_por_clase: mascaras_por_clase[cls_id] = []
            mascaras_por_clase[cls_id].append((mask_r, conf))
        for cls_id, lista in mascaras_por_clase.items():
            log(f"  YOLO: {self.config['clases'].get(cls_id, f'CLS_{cls_id}')} x{len(lista)}")

        # Pre-clasificar fisuras (clase 1) por orientacion ANTES de fusionar
        # AsÃ­ solo se fusionan fisuras del mismo tipo (longitudinal con longitudinal, etc.)
        if 1 in mascaras_por_clase:
            lista_fisuras = mascaras_por_clase.pop(1)
            long_masks = []  # Fisuras longitudinales
            trans_masks = []  # Fisuras transversales
            for mask_r, conf in lista_fisuras:
                mask_bin = (mask_r > 0.5).astype(np.uint8)
                angulo = ProcesadorGrietas._calcular_angulo(mask_bin)
                es_long = self.calibrador.es_longitudinal(angulo)
                if es_long:
                    long_masks.append((mask_r, conf))
                else:
                    trans_masks.append((mask_r, conf))
            if long_masks:
                mascaras_por_clase['1_long'] = long_masks
            if trans_masks:
                mascaras_por_clase['1_trans'] = trans_masks
            n_l = len(long_masks)
            n_t = len(trans_masks)
            log(f"  Pre-clasificacion fisuras: {n_l} longitudinales, {n_t} transversales")

        mf = FusionadorMascaras.fusionar_por_clase(mascaras_por_clase,
            iou_threshold=self.config.get('merge_iou_threshold', 0.10),
            distancia_max_px=self.config.get('merge_distancia_max_px', 50))

        # Prioridad de mascaras: 3(piel) > 2(parche) > 0(bache) > 1(fisuras)
        # Sustraer regiones de mayor prioridad para evitar superposiciones
        prioridad = [3, 2, 0, '1_long', '1_trans']  # de mayor a menor prioridad
        mascara_ocupada = np.zeros((alto, ancho), dtype=np.uint8)
        for cls_pri in prioridad:
            if cls_pri not in mf: continue
            nuevas = []
            for mask_f, conf_f in mf[cls_pri]:
                mask_bin = (mask_f > 0.5).astype(np.uint8)
                # Restar zona ocupada por clases de mayor prioridad
                mask_bin[mascara_ocupada > 0] = 0
                if np.sum(mask_bin) < 30:
                    continue  # mascara demasiado pequena despues de recortar
                nuevas.append((mask_bin.astype(np.float32), conf_f))
            mf[cls_pri] = nuevas
            # Agregar esta clase a la mascara ocupada
            for mask_f, _ in nuevas:
                mascara_ocupada[(mask_f > 0.5)] = 1

        fallas_raw = []
        pasos_imagen = {}
        for cls_id, lista_f in mf.items():
            for midx, (mask_f, conf_f) in enumerate(lista_f):
                try:
                    if cls_id == 0:
                        fallas_raw.extend(ProcesadorBaches.procesar(mask_f, self.calibrador, conf_f, imagen))
                        pasos_imagen[f'bache_{midx}'] = ProcesadorBaches.generar_pasos(mask_f, imagen)
                    elif cls_id in (1, '1_long', '1_trans'):
                        merge_px = self.config.get('merge_fisuras_px', 30)
                        fallas_raw.extend(ProcesadorGrietas.procesar(mask_f, self.calibrador, conf_f, imagen, merge_dist_px=merge_px))
                        tipo_lbl = 'long' if cls_id == '1_long' else ('trans' if cls_id == '1_trans' else 'grieta')
                        pasos_imagen[f'grieta_{tipo_lbl}_{midx}'] = ProcesadorGrietas.generar_pasos(mask_f, imagen, self.calibrador)
                    elif cls_id == 2:
                        fallas_raw.extend(ProcesadorParches.procesar(mask_f, self.calibrador, conf_f, imagen, self.config))
                        pasos_imagen[f'parche_{midx}'] = ProcesadorParches.generar_pasos(mask_f, imagen, self.config)
                    elif cls_id == 3:
                        fallas_raw.extend(self.procesador_piel.procesar(mask_f, self.calibrador, conf_f, imagen))
                        pasos_imagen[f'piel_{midx}'] = self.procesador_piel.generar_pasos(mask_f, imagen)
                except Exception as e:
                    log(f"    ERROR clase {cls_id}: {e}"); traceback.print_exc()

        # Filtrar por tamaÃ±o mÃ­nimo (segÃºn unidad)
        fallas = []
        for f in fallas_raw:
            t = f['tipo']; skip = False
            if ('BACHE' in t or 'HUECO' in t) and self.config.get('filtrar_baches', True):
                if f.get('diametro_mm', 0) < self.config.get('min_diametro_hueco_mm', 50): skip = True
            elif 'FISURA' in t and self.config.get('filtrar_grietas', True):
                if f.get('longitud_m', 0) < self.config.get('min_longitud_grieta_m', 0.05): skip = True
            elif ('REPARACION' in t or 'PARCHADO' in t) and self.config.get('filtrar_parches', True):
                if f.get('area_m2', 0) < self.config.get('min_area_parche_m2', 0.01): skip = True
            elif 'COCODRILO' in t and self.config.get('filtrar_piel', True):
                if f.get('area_m2', 0) < self.config.get('min_area_piel_m2', 0.05): skip = True
            if not skip: fallas.append(f)

        conteo = {}
        for f in fallas:
            t = f['tipo']; conteo[t] = conteo.get(t, 0) + 1; f['id'] = conteo[t]
        for t, n in conteo.items(): log(f"  Final: {t} x{n}")
        # Log detallado con confianza por cada falla
        for f in fallas:
            conf_pct = f['confianza'] * 100
            t = f['tipo']
            detalles = f"  â†’ {t} #{f['id']} | Conf: {conf_pct:.1f}% | Sev: {f['severidad']}"
            if 'BACHE' in t or 'HUECO' in t:
                detalles += f" | D={f.get('diametro_mm',0):.0f}mm"
            elif 'FISURA' in t:
                detalles += f" | e={f.get('espesor_mm',0):.1f}mm L={f.get('longitud_m',0):.2f}m A={f.get('area_m2',0):.3f}m2"
            elif 'REPARACION' in t or 'PARCHADO' in t:
                detalles += f" | A={f.get('area_m2',0):.3f}m2"
            elif 'COCODRILO' in t:
                detalles += f" | A={f.get('area_m2',0):.3f}m2 D={f.get('diametro_mm',0)/1000:.3f}m"
            log(detalles)
        log(f"  Total fallas: {len(fallas)} (filtradas: {len(fallas_raw)-len(fallas)})")

        self.todos_resultados[nombre] = fallas
        self.todos_pasos[nombre] = pasos_imagen
        return {'imagen': imagen, 'fallas': fallas, 'nombre': nombre,
                'ancho': ancho, 'alto': alto, 'pasos': pasos_imagen}

    def dibujar_resultado(self, imagen, fallas, mostrar_mallas=True, mostrar_etiquetas=True, mostrar_numeros=True, escala_texto=1.0, mostrar_circulos=True):
        vis = imagen.copy()
        et = escala_texto
        etiquetas_ocupadas = None
        if mostrar_etiquetas:
            lineas_panel = 1 + len({f.get('tipo', '') for f in fallas})
            alto_panel = int(18 + lineas_panel * 22 * et)
            ancho_panel = min(vis.shape[1] - 6, 430)
            etiquetas_ocupadas = [(6, 6, ancho_panel, alto_panel)]
        ba = [f for f in fallas if 'BACHE' in f['tipo'] or 'HUECO' in f['tipo'] or f.get('tipo') == 'DAÑOS PUNTUALES']
        fi = [f for f in fallas if 'FISURA' in f['tipo']]
        pa = [f for f in fallas if 'REPARACION' in f['tipo'] or 'PARCHADO' in f['tipo']]
        pi = [f for f in fallas if 'COCODRILO' in f['tipo']]
        if ba: vis = ProcesadorBaches.dibujar(vis, ba, mostrar_mallas=mostrar_mallas, mostrar_etiquetas=mostrar_etiquetas, mostrar_numeros=mostrar_numeros, escala_texto=et, mostrar_circulos=mostrar_circulos, occupied_labels=etiquetas_ocupadas)
        if fi: vis = ProcesadorGrietas.dibujar(vis, fi, mostrar_etiquetas=mostrar_etiquetas, mostrar_numeros=mostrar_numeros, escala_texto=et, occupied_labels=etiquetas_ocupadas)
        if pa: vis = ProcesadorParches.dibujar(vis, pa, mostrar_mallas=mostrar_mallas, mostrar_etiquetas=mostrar_etiquetas, mostrar_numeros=mostrar_numeros, escala_texto=et, mostrar_circulos=mostrar_circulos, occupied_labels=etiquetas_ocupadas)
        if pi: vis = ProcesadorPielCocodriloMTC.dibujar(vis, pi, mostrar_mallas=mostrar_mallas, mostrar_etiquetas=mostrar_etiquetas, mostrar_numeros=mostrar_numeros, escala_texto=et, mostrar_circulos=mostrar_circulos, occupied_labels=etiquetas_ocupadas)
        if mostrar_etiquetas:
            h, w = vis.shape[:2]
            y = 30
            texto_visible(vis, f"MTC 2018 | YOLOv11 | {w}x{h}px", (10, y), 0.60, (255,255,255), 2, factor=et); y += int(25 * et)
            resumen = {}
            for f in fallas:
                k = f['tipo']
                if k not in resumen: resumen[k] = {'total': 0, 1: 0, 2: 0, 3: 0}
                resumen[k]['total'] += 1; s = f['severidad']
                if s in resumen[k]: resumen[k][s] += 1
            for t, c in resumen.items():
                texto_visible(vis, f"{t}: {c['total']} (G1:{c[1]} G2:{c[2]} G3:{c[3]})",
                            (10, y), 0.48, (200,200,200), 1, factor=et); y += int(22 * et)
        return vis

    def generar_mosaico_pasos(self, nombre_imagen, imagen_original):
        """Genera un mosaico de todos los pasos intermedios"""
        pasos = self.todos_pasos.get(nombre_imagen, {})
        if not pasos: return None
        imgs = []; labels = []
        for cls_key, pasos_cls in pasos.items():
            for paso_nombre, paso_img in pasos_cls.items():
                if paso_img is None: continue
                if len(paso_img.shape) == 2:
                    paso_img = cv2.cvtColor(paso_img, cv2.COLOR_GRAY2BGR)
                imgs.append(paso_img); labels.append(f"{cls_key}: {paso_nombre}")
        if not imgs: return None
        th, tw = 400, 500
        imgs_r = []
        for img, lab in zip(imgs, labels):
            r = cv2.resize(img, (tw, th))
            cv2.putText(r, lab, (5, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0,255,255), 1)
            imgs_r.append(r)
        ncols = min(3, len(imgs_r)); nrows = math.ceil(len(imgs_r) / ncols)
        while len(imgs_r) < ncols * nrows:
            imgs_r.append(np.zeros((th, tw, 3), dtype=np.uint8))
        filas = []
        for row in range(nrows):
            fila = np.hstack(imgs_r[row*ncols:(row+1)*ncols])
            filas.append(fila)
        return np.vstack(filas)

    def exportar_excel(self, ruta_salida, nombre="RESULTADOS_MTC_CONSOLIDADO.xlsx"):
        return ExportadorExcel.exportar(self.todos_resultados, ruta_salida, self.calibrador, nombre)

# =============================================================================
# INTERFAZ GRÃFICA PRINCIPAL
# =============================================================================
class AplicacionMTC(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Analisis de Pavimentos | MTC 2018 | YOLOv11")
        self.configure(bg=EstiloUI.BG_DARK)
        try: self.state('zoomed')
        except: self.geometry("1400x800")
        self.minsize(1280, 720)
        self.motor = MotorMTC()
        self.imagenes_cargadas = []; self.imagen_actual_idx = -1
        self.resultado_actual = None; self.vista_actual = "original"
        self.procesando = False; self.calibrar_cada_imagen = False
        self._detener_flag = False; self._calibracion_unica_guardada = None
        self._cronometro_activo = False; self._cronometro_after_id = None
        self._resultados_batch = {}; self.vars = {}
        self._cv2_actual = None; self._cv2_original_full = None; self._imagen_tk = None
        self._zoom_level = 1.0; self._zoom_min = 0.1; self._zoom_max = 10.0
        self._pan_offset_x = 0; self._pan_offset_y = 0
        self._pan_dragging = False; self._pan_start_x = 0; self._pan_start_y = 0
        self._crear_estilos(); self._crear_interfaz(); self._verificar_dependencias()

    def _crear_estilos(self):
        style = ttk.Style(); style.theme_use('clam')
        style.configure("Accent.TButton", background=EstiloUI.BG_BUTTON,
                         foreground="white", font=EstiloUI.FONT_SUBTITLE, padding=(10, 5))
        style.map("Accent.TButton", background=[('active', EstiloUI.BG_BUTTON_HOVER)])
        style.configure("Secondary.TButton", background=EstiloUI.BG_BUTTON_SECONDARY,
                         foreground=EstiloUI.FG_PRIMARY, font=EstiloUI.FONT_BODY, padding=(7, 3))
        style.map("Secondary.TButton",
                  background=[('active', EstiloUI.BG_ACCENT)],
                  foreground=[('active', EstiloUI.FG_PRIMARY)])
        style.configure("View.TNotebook", background=EstiloUI.BG_DARK, borderwidth=0)
        style.configure("View.TNotebook.Tab", background=EstiloUI.BG_BUTTON_SECONDARY,
                        foreground=EstiloUI.FG_PRIMARY, padding=(14, 8), font=EstiloUI.FONT_BODY)
        style.map("View.TNotebook.Tab",
                  background=[('selected', EstiloUI.BG_PANEL), ('active', EstiloUI.BG_ACCENT)],
                  foreground=[('selected', EstiloUI.FG_ACCENT), ('active', EstiloUI.FG_PRIMARY)])

    def _crear_interfaz(self):
        self.grid_rowconfigure(1, weight=1); self.grid_columnconfigure(0, weight=1)
        self._crear_barra_superior()
        # PanedWindow principal para paneles redimensionables
        self._paned = tk.PanedWindow(self, orient="horizontal", bg=EstiloUI.BG_DARK,
                                      sashwidth=6, sashrelief="raised",
                                      sashpad=1, opaqueresize=True)
        self._paned.grid(row=1, column=0, sticky="nsew")
        # Panel izquierdo (config)
        self._frame_config = tk.Frame(self._paned, bg=EstiloUI.BG_PANEL)
        self._paned.add(self._frame_config, minsize=200, width=300)
        self._crear_panel_config(self._frame_config)
        # Panel central (visor)
        self._frame_visor = tk.Frame(self._paned, bg=EstiloUI.BG_DARK)
        self._paned.add(self._frame_visor, minsize=400, stretch="always")
        self._crear_panel_visor(self._frame_visor)
        # Panel derecho (resultados)
        self._frame_resultados = tk.Frame(self._paned, bg=EstiloUI.BG_PANEL)
        self._paned.add(self._frame_resultados, minsize=180, width=260)
        self._crear_panel_resultados(self._frame_resultados)
        self._crear_barra_estado()

    def _crear_barra_superior(self):
        barra = tk.Frame(self, bg=EstiloUI.BG_PANEL, height=55)
        barra.grid(row=0, column=0, columnspan=3, sticky="ew"); barra.grid_propagate(False)
        tk.Label(barra, text="ANALISIS DE PAVIMENTOS | MTC 2018 | YOLOv11",
                 font=EstiloUI.FONT_TITLE, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY
                 ).pack(side="left", padx=15, pady=8)
        bf = tk.Frame(barra, bg=EstiloUI.BG_PANEL); bf.pack(side="right", padx=10, pady=8)
        # Botones principales (orden derecha a izquierda)
        ttk.Button(bf, text="Exportar Todo a Excel", style="Accent.TButton",
                   command=self._exportar_excel).pack(side="right", padx=3)
        ttk.Button(bf, text="Exportar Nuevo Excel", style="Accent.TButton",
                   command=self._exportar_nuevo_excel).pack(side="right", padx=3)
        ttk.Button(bf, text="Exportar Resultados Nuevos", style="Accent.TButton",
                   command=self._exportar_resultados_nuevos).pack(side="right", padx=3)
        ttk.Button(bf, text="Procesar Todo", style="Accent.TButton",
                   command=self._procesar_todo).pack(side="right", padx=3)
        ttk.Button(bf, text="Procesar Actual", style="Accent.TButton",
                   command=self._procesar_actual).pack(side="right", padx=3)
        self.lbl_tiempo = tk.Label(bf, text="", font=("Segoe UI", 10, "bold"),
                                    bg=EstiloUI.BG_PANEL, fg=EstiloUI.BG_SUCCESS)
        self.lbl_tiempo.pack(side="right", padx=6)
        self.btn_detener = ttk.Button(bf, text="Detener", style="Secondary.TButton",
                   command=self._detener_procesamiento, state="disabled")
        self.btn_detener.pack(side="right", padx=3)
        ttk.Button(bf, text="Cargar Imagenes", style="Secondary.TButton",
                   command=self._cargar_imagenes).pack(side="right", padx=3)
        ttk.Button(bf, text="Cargar Modelo", style="Secondary.TButton",
                   command=self._cargar_modelo).pack(side="right", padx=3)

    def _crear_panel_config(self, parent):
        parent.grid_rowconfigure(0, weight=1); parent.grid_columnconfigure(0, weight=1)
        canvas = tk.Canvas(parent, bg=EstiloUI.BG_PANEL, highlightthickness=0)
        scrollbar_v = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollbar_h = ttk.Scrollbar(parent, orient="horizontal", command=canvas.xview)
        sf = tk.Frame(canvas, bg=EstiloUI.BG_PANEL)
        sf.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=sf, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar_v.set, xscrollcommand=scrollbar_h.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar_v.grid(row=0, column=1, sticky="ns")
        scrollbar_h.grid(row=1, column=0, sticky="ew")
        def _mw_config(e):
            canvas.yview_scroll(int(-1*(e.delta/120)), "units")
        def _bind_mw_config(widget):
            widget.bind("<MouseWheel>", _mw_config)
            widget.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _mw_config))
            widget.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))
        _bind_mw_config(canvas)
        _bind_mw_config(sf)
        self._canvas_config = canvas
        self._sf_config = sf

        # GENERAL
        self._sec(sf, "CONFIGURACION GENERAL")
        self._sl(sf, 'ancho_via', "Ancho Via (m)", 3.0, 12.0, 6.5, resolution=0.1)
        self._sl(sf, 'confianza_min', "Confianza Min", 0.05, 0.95, 0.1, resolution=0.05)
        self._sl(sf, 'iou_threshold', "IoU Threshold", 0.1, 0.9, 0.45, resolution=0.05)


        # CALIBRACIÃ“N
        self._sec(sf, "CALIBRACION")
        self.vars['modo_calibracion'] = tk.StringVar(value="automatica")
        modos_cal = [
            ("Automatica (ancho imagen)", "automatica"),
            ("Calibrar cada imagen (OpenCV)", "cada_imagen"),
            ("Calibracion unica (1 ejemplar)", "unica"),
        ]
        for texto, valor in modos_cal:
            tk.Radiobutton(sf, text=texto, variable=self.vars['modo_calibracion'],
                           value=valor, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY,
                           selectcolor=EstiloUI.BG_DARK, activebackground=EstiloUI.BG_PANEL,
                           font=EstiloUI.FONT_SMALL).pack(anchor="w", padx=10, pady=1)
        btn_cal_frame = tk.Frame(sf, bg=EstiloUI.BG_PANEL); btn_cal_frame.pack(fill="x", padx=10, pady=3)
        ttk.Button(btn_cal_frame, text="Calibrar ahora (1 imagen)",
                   style="Secondary.TButton",
                   command=self._calibrar_unica).pack(fill="x")
        self.lbl_cal_estado = tk.Label(sf, text="Sin calibracion unica",
                                        font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                                        fg=EstiloUI.FG_SECONDARY)
        self.lbl_cal_estado.pack(anchor="w", padx=10, pady=1)

        # BOTONES PARA VENTANAS DE CONFIGURACIÃ“N AVANZADA
        self._sec(sf, "CONFIGURACION AVANZADA")
        btn_cfg = tk.Frame(sf, bg=EstiloUI.BG_PANEL); btn_cfg.pack(fill="x", padx=10, pady=2)
        for texto, cmd in [
            ("Filtros TamaÃ±o Minimo", self._abrir_filtros),
            ("Fusion Solapamientos", self._abrir_fusion),
            ("Piel de Cocodrilo", self._abrir_piel_config),
            ("Deteccion Poligonos", self._abrir_poligonos),
            ("Perfiles Rapidos", self._abrir_perfiles),
        ]:
            ttk.Button(btn_cfg, text=texto, style="Secondary.TButton",
                       command=cmd).pack(fill="x", pady=2)

        # Inicializar variables para ventanas modales
        self._init_vars_avanzadas(sf)

        # VISUALIZACIÃ“N DE MALLAS
        self._sec(sf, "VISUALIZACION")
        self.vars['mostrar_mallas'] = tk.BooleanVar(value=True)
        tk.Checkbutton(sf, text="Mostrar mallas/esqueleto",
                        variable=self.vars['mostrar_mallas'],
                        bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK,
                        font=EstiloUI.FONT_BODY,
                        command=self._refrescar_vista).pack(anchor="w", padx=10, pady=3)
        self.vars['mostrar_circulos'] = tk.BooleanVar(value=True)
        tk.Checkbutton(sf, text="Mostrar circulos",
                        variable=self.vars['mostrar_circulos'],
                        bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK,
                        font=EstiloUI.FONT_BODY,
                        command=self._refrescar_vista).pack(anchor="w", padx=10, pady=3)
        self.vars['mostrar_etiquetas'] = tk.BooleanVar(value=True)
        tk.Checkbutton(sf, text="Mostrar etiquetas",
                        variable=self.vars['mostrar_etiquetas'],
                        bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK,
                        font=EstiloUI.FONT_BODY,
                        command=self._refrescar_vista).pack(anchor="w", padx=10, pady=3)
        self.vars['mostrar_numeros'] = tk.BooleanVar(value=True)
        tk.Checkbutton(sf, text="Mostrar valores numericos",
                        variable=self.vars['mostrar_numeros'],
                        bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK,
                        font=EstiloUI.FONT_BODY,
                        command=self._refrescar_vista).pack(anchor="w", padx=10, pady=3)
        # Escala de texto se inicializa aqui pero se muestra en la barra de zoom
        self.vars['escala_texto'] = tk.DoubleVar(value=1.0)

    def _crear_panel_visor(self, parent):
        panel = tk.Frame(parent, bg=EstiloUI.BG_DARK)
        panel.pack(fill="both", expand=True, padx=2, pady=5)
        panel.grid_rowconfigure(2, weight=1); panel.grid_columnconfigure(0, weight=1)
        # NavegaciÃ³n
        nav = tk.Frame(panel, bg=EstiloUI.BG_CARD, height=38)
        nav.grid(row=0, column=0, sticky="ew", pady=(0,2))
        ttk.Button(nav, text="< Anterior", style="Secondary.TButton",
                   command=self._imagen_anterior).pack(side="left", padx=4, pady=4)
        ttk.Button(nav, text="Siguiente >", style="Secondary.TButton",
                   command=self._imagen_siguiente).pack(side="left", padx=4, pady=4)
        self.lbl_nav = tk.Label(nav, text="Sin imagenes", font=EstiloUI.FONT_BODY,
                                 bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_SECONDARY)
        self.lbl_nav.pack(side="left", padx=10)
        # Vistas: Original, Resultado, Pasos
        vf = tk.Frame(nav, bg=EstiloUI.BG_CARD); vf.pack(side="right", padx=8, pady=4)
        ttk.Button(vf, text="Original", style="Secondary.TButton",
                   command=lambda: self._cambiar_vista("original")).pack(side="left", padx=2)
        ttk.Button(vf, text="Resultado", style="Secondary.TButton",
                   command=lambda: self._cambiar_vista("resultado")).pack(side="left", padx=2)
        ttk.Button(vf, text="Pasos", style="Secondary.TButton",
                   command=lambda: self._cambiar_vista("pasos")).pack(side="left", padx=2)
        ttk.Button(vf, text="Guardar", style="Secondary.TButton",
                   command=self._guardar_resultado).pack(side="left", padx=2)
        # Zoom bar + datos de imagen
        zb = tk.Frame(panel, bg=EstiloUI.BG_PANEL, height=30)
        zb.grid(row=1, column=0, sticky="ew", pady=(0,2))
        zl = tk.Frame(zb, bg=EstiloUI.BG_PANEL); zl.pack(side="left", padx=5, pady=2)
        ttk.Button(zl, text="-", style="Secondary.TButton", command=self._zoom_out).pack(side="left", padx=1)
        self.lbl_zoom = tk.Label(zl, text="100%", font=EstiloUI.FONT_LABEL,
                                  bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT, width=6)
        self.lbl_zoom.pack(side="left", padx=3)
        ttk.Button(zl, text="+", style="Secondary.TButton", command=self._zoom_in).pack(side="left", padx=1)
        ttk.Button(zl, text="Ajustar", style="Secondary.TButton", command=self._zoom_fit).pack(side="left", padx=3)
        ttk.Button(zl, text="1:1", style="Secondary.TButton", command=self._zoom_100).pack(side="left", padx=3)
        # Separador visual
        tk.Frame(zl, bg=EstiloUI.FG_SECONDARY, width=1, height=20).pack(side="left", padx=6, fill="y")
        # Escala de texto en barra de zoom
        tk.Label(zl, text="Texto:", font=EstiloUI.FONT_SMALL,
                 bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY).pack(side="left", padx=(0,2))
        sl_txt = ttk.Scale(zl, from_=0.3, to=3.0, variable=self.vars.get('escala_texto', tk.DoubleVar(value=1.0)),
                           orient="horizontal", command=lambda _: self._on_escala_texto(), length=80)
        sl_txt.pack(side="left")
        self.lbl_escala_texto = tk.Label(zl, text="1.0x", font=EstiloUI.FONT_LABEL,
                                          bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT, width=4, anchor="e")
        self.lbl_escala_texto.pack(side="left", padx=(2,3))
        ttk.Button(zl, text="Reescribir", style="Secondary.TButton",
                   command=self._refrescar_vista).pack(side="left", padx=3)
        # Datos de imagen
        self.lbl_img_info = tk.Label(zb, text="", font=EstiloUI.FONT_SMALL,
                                      bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_SECONDARY)
        self.lbl_img_info.pack(side="left", padx=15)
        # Coordenadas del mouse
        self.lbl_pixel_coords = tk.Label(zb, text="X:-- Y:--", font=EstiloUI.FONT_MONO,
                                          bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT)
        self.lbl_pixel_coords.pack(side="right", padx=10)
        # Vista actual label
        self.lbl_vista = tk.Label(zb, text="[Original]", font=EstiloUI.FONT_LABEL,
                                   bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_ACCENT)
        self.lbl_vista.pack(side="right", padx=5)
        # Canvas
        self.canvas_imagen = tk.Canvas(panel, bg=EstiloUI.BG_INPUT, highlightthickness=0)
        self.canvas_imagen.grid(row=2, column=0, sticky="nsew")
        self.canvas_imagen.bind("<Configure>", self._on_canvas_resize)
        self.canvas_imagen.bind("<MouseWheel>", self._on_zoom_wheel)
        self.canvas_imagen.bind("<ButtonPress-2>", self._on_pan_start)
        self.canvas_imagen.bind("<B2-Motion>", self._on_pan_move)
        self.canvas_imagen.bind("<ButtonRelease-2>", self._on_pan_end)
        self.canvas_imagen.bind("<ButtonPress-3>", self._on_pan_start)
        self.canvas_imagen.bind("<B3-Motion>", self._on_pan_move)
        self.canvas_imagen.bind("<ButtonRelease-3>", self._on_pan_end)
        self.canvas_imagen.bind("<Motion>", self._on_mouse_move)
        self.canvas_imagen.bind("<ButtonPress-1>", self._on_click_falla)

    def _crear_panel_resultados(self, parent):
        parent.grid_rowconfigure(0, weight=1); parent.grid_columnconfigure(0, weight=1)
        panel = tk.Frame(parent, bg=EstiloUI.BG_PANEL)
        panel.grid(row=0, column=0, sticky="nsew")
        panel.grid_rowconfigure(3, weight=1); panel.grid_columnconfigure(0, weight=1)
        # Info imagen
        imf = tk.Frame(panel, bg=EstiloUI.BG_CARD)
        imf.grid(row=0, column=0, sticky="ew", padx=5, pady=(5,2))
        tk.Label(imf, text="DATOS IMAGEN", font=EstiloUI.FONT_SUBTITLE,
                 bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_ACCENT).pack(padx=10, pady=(4,2), anchor="w")
        self.img_data_labels = {}
        for key, lt in [("archivo","Archivo:"),("resolucion","Resolucion:"),("px_mm","px/mm:"),("eje","Eje via:")]:
            row = tk.Frame(imf, bg=EstiloUI.BG_CARD); row.pack(fill="x", padx=10, pady=0)
            tk.Label(row, text=lt, font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_SECONDARY).pack(side="left")
            lbl = tk.Label(row, text="--", font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_HIGHLIGHT)
            lbl.pack(side="right"); self.img_data_labels[key] = lbl
        tk.Frame(imf, bg=EstiloUI.BG_CARD, height=4).pack()
        # Lista de fallas individuales con checkboxes (seleccion individual)
        inf = tk.Frame(panel, bg=EstiloUI.BG_CARD)
        inf.grid(row=1, column=0, sticky="nsew", padx=5, pady=2)
        inf.grid_rowconfigure(1, weight=1); inf.grid_columnconfigure(0, weight=1)
        hdr = tk.Frame(inf, bg=EstiloUI.BG_CARD); hdr.pack(fill="x")
        tk.Label(hdr, text="FALLAS DETECTADAS", font=EstiloUI.FONT_SUBTITLE,
                 bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_ACCENT).pack(side="left", padx=10, pady=(4,2))
        self.lbl_fallas_total = tk.Label(hdr, text="0/0", font=EstiloUI.FONT_LABEL,
                                          bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_HIGHLIGHT)
        self.lbl_fallas_total.pack(side="right", padx=10, pady=(4,2))
        # Botones seleccionar/deseleccionar todo
        btn_f = tk.Frame(inf, bg=EstiloUI.BG_CARD); btn_f.pack(fill="x", padx=5, pady=2)
        ttk.Button(btn_f, text="Todas", style="Secondary.TButton",
                   command=lambda: self._sel_todas_fallas(True)).pack(side="left", padx=2)
        ttk.Button(btn_f, text="Ninguna", style="Secondary.TButton",
                   command=lambda: self._sel_todas_fallas(False)).pack(side="left", padx=2)
        tk.Label(btn_f, text="Click en imagen para toggle", font=EstiloUI.FONT_SMALL,
                 bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_SECONDARY).pack(side="right", padx=5)
        # Scrollable list de fallas
        lf_scroll = tk.Frame(inf, bg=EstiloUI.BG_CARD)
        lf_scroll.pack(fill="both", expand=True, padx=2, pady=2)
        self.canvas_fallas = tk.Canvas(lf_scroll, bg=EstiloUI.BG_DARK, highlightthickness=0, height=120)
        sb_fallas = ttk.Scrollbar(lf_scroll, orient="vertical", command=self.canvas_fallas.yview)
        self.frame_fallas = tk.Frame(self.canvas_fallas, bg=EstiloUI.BG_DARK)
        self.frame_fallas.bind("<Configure>",
            lambda e: self.canvas_fallas.configure(scrollregion=self.canvas_fallas.bbox("all")))
        self.canvas_fallas.create_window((0,0), window=self.frame_fallas, anchor="nw")
        self.canvas_fallas.configure(yscrollcommand=sb_fallas.set)
        self.canvas_fallas.pack(side="left", fill="both", expand=True)
        sb_fallas.pack(side="right", fill="y")
        # Mouse wheel scroll para la lista de fallas
        def _mw_fallas(e):
            self.canvas_fallas.yview_scroll(int(-1*(e.delta/120)), "units")
        self.canvas_fallas.bind("<MouseWheel>", _mw_fallas)
        self.frame_fallas.bind("<MouseWheel>", _mw_fallas)
        self.canvas_fallas.bind("<Enter>", lambda e: self.canvas_fallas.bind_all("<MouseWheel>", _mw_fallas))
        self.canvas_fallas.bind("<Leave>", lambda e: self.canvas_fallas.unbind_all("<MouseWheel>"))
        self._falla_vars = []  # [(BooleanVar, falla_dict)]
        self._fallas_excluidas = {}  # {img_name: set(falla_indices)}

        self._tiempo_proc = 0.0  # tiempo total de procesamiento
        # Severidad
        sv = tk.Frame(panel, bg=EstiloUI.BG_CARD)
        sv.grid(row=2, column=0, sticky="ew", padx=5, pady=2)
        tk.Label(sv, text="SEVERIDAD", font=EstiloUI.FONT_SUBTITLE,
                 bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_ACCENT).pack(padx=10, pady=(4,2), anchor="w")
        self.sev_labels = {}
        for key, lt, cl in [("s1","Sev 1 (Leve):","#00b894"),("s2","Sev 2 (Mod):","#fdcb6e"),("s3","Sev 3 (Sev):","#e94560")]:
            row = tk.Frame(sv, bg=EstiloUI.BG_CARD); row.pack(fill="x", padx=10, pady=0)
            tk.Label(row, text=lt, font=EstiloUI.FONT_BODY, bg=EstiloUI.BG_CARD, fg=cl).pack(side="left")
            lbl = tk.Label(row, text="0", font=EstiloUI.FONT_LABEL, bg=EstiloUI.BG_CARD, fg=cl)
            lbl.pack(side="right"); self.sev_labels[key] = lbl
        tk.Frame(sv, bg=EstiloUI.BG_CARD, height=4).pack()
        # Log
        lf = tk.Frame(panel, bg=EstiloUI.BG_PANEL)
        lf.grid(row=3, column=0, sticky="nsew", padx=5, pady=(0,5))
        lf.grid_rowconfigure(1, weight=1); lf.grid_columnconfigure(0, weight=1)
        tk.Label(lf, text="LOG", font=EstiloUI.FONT_SUBTITLE, bg=EstiloUI.BG_PANEL,
                 fg=EstiloUI.FG_ACCENT).grid(row=0, column=0, sticky="w", padx=5, pady=(3,1))
        self.log_text = scrolledtext.ScrolledText(
            lf,
            wrap=tk.WORD,
            font=EstiloUI.FONT_MONO,
            bg=EstiloUI.BG_INPUT,
            fg=EstiloUI.FG_LOG,
            insertbackground=EstiloUI.FG_LOG,
            relief="flat",
            height=12,
        )
        self.log_text.grid(row=1, column=0, sticky="nsew", padx=2, pady=2)
        ttk.Button(lf, text="Limpiar", style="Secondary.TButton",
                   command=lambda: self.log_text.delete(1.0, tk.END)).grid(row=2, column=0, sticky="e", padx=5, pady=2)

    def _crear_barra_estado(self):
        barra = tk.Frame(self, bg=EstiloUI.BG_CARD, height=26)
        barra.grid(row=2, column=0, columnspan=3, sticky="ew"); barra.grid_propagate(False)
        self.lbl_estado = tk.Label(barra, text="Listo", font=EstiloUI.FONT_SMALL,
                                    bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_SECONDARY)
        self.lbl_estado.pack(side="left", padx=10, pady=3)
        tk.Label(barra, text="Bach. Miguel Bernardino Quispe Arias  |  Bach. Briza Edith Catachura Aycaya",
                 font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_SECONDARY).pack(side="left", expand=True, pady=3)
        self.lbl_modelo = tk.Label(barra, text="Modelo: No cargado", font=EstiloUI.FONT_SMALL,
                                    bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_ACCENT)
        self.lbl_modelo.pack(side="right", padx=10, pady=3)

    # =========================================================================
    # HELPERS WIDGETS
    # =========================================================================
    def _sec(self, p, t):
        tk.Frame(p, bg=EstiloUI.BG_ACCENT, height=2).pack(fill="x", padx=10, pady=(10,0))
        tk.Label(p, text=t, font=EstiloUI.FONT_LABEL, bg=EstiloUI.BG_PANEL,
                 fg=EstiloUI.FG_ACCENT).pack(anchor="w", padx=10, pady=(2,4))

    def _sl(self, parent, key, label, from_, to, default, step=1, resolution=None):
        frame = tk.Frame(parent, bg=EstiloUI.BG_PANEL); frame.pack(fill="x", padx=10, pady=1)
        tk.Label(frame, text=label, font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                 fg=EstiloUI.FG_SECONDARY).pack(anchor="w")
        var = tk.DoubleVar(value=default); self.vars[key] = var
        row = tk.Frame(frame, bg=EstiloUI.BG_PANEL); row.pack(fill="x")
        res = resolution if resolution else step
        sl = tk.Scale(row, from_=from_, to=to, orient="horizontal", variable=var, resolution=res,
                      showvalue=False, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT,
                      troughcolor=EstiloUI.BG_DARK, highlightthickness=0,
                      activebackground=EstiloUI.BG_BUTTON, length=180, sliderlength=15)
        sl.pack(side="left", fill="x", expand=True)
        vl = tk.Label(row, text=str(default), font=EstiloUI.FONT_LABEL, bg=EstiloUI.BG_PANEL,
                       fg=EstiloUI.FG_HIGHLIGHT, width=7, anchor="e")
        vl.pack(side="right", padx=(4, 0))
        def upd(*a):
            v = var.get(); vl.config(text=f"{v:.2f}" if (resolution and resolution < 1) else str(int(v)))
        var.trace_add("write", upd)

    def _sync_params(self):
        fp = {'confianza_min','iou_threshold','clahe_clip','min_circularidad',
              'min_diametro_hueco_mm','min_longitud_grieta_m','min_area_parche_m2',
              'min_area_piel_m2','merge_iou_threshold'}
        for k, v in self.vars.items():
            if isinstance(v, tk.BooleanVar):
                self.motor.config[k] = v.get()
            elif k == 'ancho_via':
                self.motor.config['ancho_via_real_m'] = float(v.get())
                self.motor.calibrador.ancho_via_real_m = float(v.get())
            elif k in self.motor.config:
                val = v.get()
                self.motor.config[k] = float(val) if k in fp else int(val)
        self.motor.procesador_piel = ProcesadorPielCocodrilo(self.motor.config)
        modo_cal = self.vars.get('modo_calibracion', tk.StringVar(value="automatica")).get()
        self.calibrar_cada_imagen = (modo_cal == "cada_imagen")

    # =========================================================================
    # VARIABLES AVANZADAS (inicializadas pero no mostradas en panel principal)
    # =========================================================================
    def _init_vars_avanzadas(self, parent):
        """Inicializa variables que se muestran en ventanas modales"""
        # Filtros
        self.vars.setdefault('filtrar_baches', tk.BooleanVar(value=True))
        self.vars.setdefault('min_diametro_hueco_mm', tk.DoubleVar(value=50))
        self.vars.setdefault('filtrar_grietas', tk.BooleanVar(value=True))
        self.vars.setdefault('min_longitud_grieta_m', tk.DoubleVar(value=0.05))
        self.vars.setdefault('filtrar_parches', tk.BooleanVar(value=True))
        self.vars.setdefault('min_area_parche_m2', tk.DoubleVar(value=0.01))
        self.vars.setdefault('filtrar_piel', tk.BooleanVar(value=True))
        self.vars.setdefault('min_area_piel_m2', tk.DoubleVar(value=0.05))
        self.vars.setdefault('merge_iou_threshold', tk.DoubleVar(value=0.10))
        self.vars.setdefault('merge_distancia_max_px', tk.IntVar(value=50))
        self.vars.setdefault('merge_fisuras_px', tk.IntVar(value=30))
        # Piel de Cocodrilo
        for k, v in [('clahe_clip',4.0),('clahe_tile',8),('bilateral_d',9),
                      ('bilateral_sigma_color',75),('bilateral_sigma_space',75),
                      ('block_size',23),('C_umbral',10),('kernel_apertura',3),
                      ('kernel_cierre',6),('iteraciones_cierre',2)]:
            if isinstance(v, float):
                self.vars.setdefault(k, tk.DoubleVar(value=v))
            else:
                self.vars.setdefault(k, tk.IntVar(value=v))
        self.vars.setdefault('usar_frangi', tk.BooleanVar(value=True))
        self.vars.setdefault('usar_multiescala', tk.BooleanVar(value=True))
        self.vars.setdefault('usar_refinamiento', tk.BooleanVar(value=True))
        # Poligonos
        for k, v in [('min_area_poligono',300),('min_circularidad',0.08),
                      ('min_vertices',4),('max_vertices',25),('min_radio_circulo',8),
                      ('min_longitud_rama',30),('min_area_objeto',100),('max_gap_cierre',20)]:
            if isinstance(v, float):
                self.vars.setdefault(k, tk.DoubleVar(value=v))
            else:
                self.vars.setdefault(k, tk.IntVar(value=v))

    def _crear_ventana_modal(self, titulo, width=None, height=None):
        """Crea una ventana modal con estilo consistente (tamaÃ±o auto-ajustado al contenido)"""
        win = tk.Toplevel(self)
        win.title(titulo)
        win.configure(bg=EstiloUI.BG_PANEL)
        win.transient(self)
        win.grab_set()
        win.resizable(True, True)
        return win

    def _centrar_ventana_modal(self, win):
        """Centra una ventana modal en la ventana principal tras agregar el contenido"""
        win.update_idletasks()
        w = win.winfo_reqwidth()
        h = win.winfo_reqheight()
        x = self.winfo_x() + (self.winfo_width() - w) // 2
        y = self.winfo_y() + (self.winfo_height() - h) // 2
        win.geometry(f"+{x}+{y}")

    def _agregar_botones_aplicar(self, win, frame_parent):
        """Agrega botones Aplicar Actual / Aplicar Todas / Cerrar"""
        bf = tk.Frame(frame_parent, bg=EstiloUI.BG_PANEL)
        bf.pack(fill="x", padx=10, pady=(10,5), side="bottom")
        ttk.Button(bf, text="Aplicar a imagen actual", style="Accent.TButton",
                   command=lambda: self._aplicar_modal(win, solo_actual=True)).pack(fill="x", pady=2)
        ttk.Button(bf, text="Aplicar a todas las imagenes", style="Accent.TButton",
                   command=lambda: self._aplicar_modal(win, solo_actual=False)).pack(fill="x", pady=2)
        ttk.Button(bf, text="Cerrar", style="Secondary.TButton",
                   command=win.destroy).pack(fill="x", pady=2)
        self._centrar_ventana_modal(win)

    def _aplicar_modal(self, win, solo_actual=True):
        """Aplica config y guarda sin reprocesar"""
        self._sync_params()
        if not hasattr(self, '_config_avanzada_por_imagen'):
            self._config_avanzada_por_imagen = {}
            
        if solo_actual:
            if getattr(self, 'imagenes_cargadas', None):
                self._config_avanzada_por_imagen[self.imagen_actual_idx] = self.motor.config.copy()
                self._log("Config avanzada guardada SÃ“LO para la imagen actual.")
            else:
                self._log("No hay imagen actual cargada.")
        else:
            self._config_avanzada_por_imagen.clear()
            self._log("Config avanzada aplicada a TODAS las imÃ¡genes.")
        win.destroy()

    def _sl_modal(self, parent, key, label, from_, to, default, step=1, resolution=None):
        """Slider moderno para ventanas modales (usa self.vars existentes)"""
        f = tk.Frame(parent, bg=EstiloUI.BG_PANEL); f.pack(fill="x", padx=10, pady=2)
        # Label superior con nombre y valor
        hdr = tk.Frame(f, bg=EstiloUI.BG_PANEL); hdr.pack(fill="x")
        tk.Label(hdr, text=label, font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                 fg=EstiloUI.FG_PRIMARY).pack(side="left")
        var = self.vars[key]
        vl = tk.Label(hdr, text=str(default), font=("Segoe UI", 9, "bold"), bg=EstiloUI.BG_PANEL,
                       fg=EstiloUI.FG_HIGHLIGHT, anchor="e")
        vl.pack(side="right", padx=(4,0))
        # Slider con estilo moderno (colores consistentes con panel principal)
        res = resolution if resolution else step
        sl = tk.Scale(f, from_=from_, to=to, orient="horizontal", variable=var, resolution=res,
                      showvalue=False, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT,
                      troughcolor=EstiloUI.BG_DARK, highlightthickness=0,
                      activebackground=EstiloUI.BG_BUTTON, length=280, sliderlength=15,
                      bd=0, relief="flat")
        sl.pack(fill="x", pady=(0,2))
        def upd(*a):
            try:
                v = var.get(); vl.config(text=f"{v:.2f}" if (resolution and resolution < 1) else str(int(v)))
            except Exception: pass
        var.trace_add("write", upd); upd()

    def _abrir_filtros(self):
        win = self._crear_ventana_modal("Filtros TamaÃ±o MÃ­nimo", 480, 480)
        sf = tk.Frame(win, bg=EstiloUI.BG_PANEL); sf.pack(fill="both", expand=True)
        self._sec(sf, "FILTROS TAMAÃ‘O MINIMO")
        tk.Checkbutton(sf, text="Filtrar Baches por diametro", variable=self.vars['filtrar_baches'],
                        bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK,
                        font=EstiloUI.FONT_SMALL).pack(anchor="w", padx=10, pady=1)
        self._sl_modal(sf, 'min_diametro_hueco_mm', "Min Diam Bache (mm)", 0, 200, 50, step=10)
        tk.Checkbutton(sf, text="Filtrar Grietas por longitud", variable=self.vars['filtrar_grietas'],
                        bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK,
                        font=EstiloUI.FONT_SMALL).pack(anchor="w", padx=10, pady=1)
        self._sl_modal(sf, 'min_longitud_grieta_m', "Min Long Fisura (m)", 0.0, 1.0, 0.05, resolution=0.01)
        tk.Checkbutton(sf, text="Filtrar Parches por area", variable=self.vars['filtrar_parches'],
                        bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK,
                        font=EstiloUI.FONT_SMALL).pack(anchor="w", padx=10, pady=1)
        self._sl_modal(sf, 'min_area_parche_m2', "Min Area Parche (m2)", 0.0, 1.0, 0.01, resolution=0.005)
        tk.Checkbutton(sf, text="Filtrar P.Cocodrilo por area", variable=self.vars['filtrar_piel'],
                        bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK,
                        font=EstiloUI.FONT_SMALL).pack(anchor="w", padx=10, pady=1)
        self._sl_modal(sf, 'min_area_piel_m2', "Min Area P.Cocod (m2)", 0.0, 2.0, 0.05, resolution=0.01)
        self._sec(sf, "UNION FISURAS MISMO TIPO")
        self._sl_modal(sf, 'merge_fisuras_px', "Dist union fisuras (px)", 0, 100, 30, step=5)
        self._agregar_botones_aplicar(win, sf)

    def _abrir_fusion(self):
        win = self._crear_ventana_modal("FusiÃ³n Solapamientos", 480, 300)
        sf = tk.Frame(win, bg=EstiloUI.BG_PANEL); sf.pack(fill="both", expand=True)
        self._sec(sf, "FUSION SOLAPAMIENTOS")
        self._sl_modal(sf, 'merge_iou_threshold', "IoU Fusion", 0.0, 0.5, 0.10, resolution=0.02)
        self._sl_modal(sf, 'merge_distancia_max_px', "Dist Max Fusion (px)", 0, 200, 50, step=10)
        self._agregar_botones_aplicar(win, sf)

    def _abrir_piel_config(self):
        win = self._crear_ventana_modal("Piel de Cocodrilo - Parametros", 500, 750)
        sf = tk.Frame(win, bg=EstiloUI.BG_PANEL); sf.pack(fill="both", expand=True)
        self._sec(sf, "CALIBRAR CON MUESTRA")
        cal_f = tk.Frame(sf, bg=EstiloUI.BG_PANEL); cal_f.pack(fill="x", padx=10, pady=3)
        tk.Label(cal_f, text="Seleccione una imagen con piel de cocodrilo\npara auto-configurar los parametros:",
                 font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_SECONDARY,
                 justify="left").pack(anchor="w", pady=(0,3))
        self._lbl_piel_cal_estado = tk.Label(cal_f, text="Sin calibrar",
                                              font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                                              fg=EstiloUI.FG_SECONDARY)
        self._lbl_piel_cal_estado.pack(anchor="w", pady=1)
        ttk.Button(cal_f, text="Seleccionar imagen muestra", style="Accent.TButton",
                   command=lambda: self._calibrar_piel_con_muestra(win)).pack(fill="x", pady=2)
        self._sec(sf, "PIEL DE COCODRILO")
        self._sl_modal(sf, 'clahe_clip', "CLAHE Clip", 1.0, 8.0, 4.0, resolution=0.5)
        self._sl_modal(sf, 'clahe_tile', "CLAHE Tile", 4, 16, 8)
        self._sl_modal(sf, 'bilateral_d', "Bilateral D", 3, 15, 9)
        self._sl_modal(sf, 'bilateral_sigma_color', "Sigma Color", 20, 150, 75)
        self._sl_modal(sf, 'bilateral_sigma_space', "Sigma Espacio", 20, 150, 75)
        self._sl_modal(sf, 'block_size', "Block Size", 11, 51, 23, step=2)
        self._sl_modal(sf, 'C_umbral', "Constante C", 3, 25, 10)
        tk.Checkbutton(sf, text="Filtro Frangi", variable=self.vars['usar_frangi'],
                        bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK,
                        font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)
        tk.Checkbutton(sf, text="Refinar esq. y cerrar gaps", variable=self.vars['usar_refinamiento'],
                        bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK,
                        font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)
        tk.Checkbutton(sf, text="Multi-escala", variable=self.vars['usar_multiescala'],
                        bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK,
                        font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)
        self._sl_modal(sf, 'kernel_apertura', "Kernel Apertura", 2, 7, 3)
        self._sl_modal(sf, 'kernel_cierre', "Kernel Cierre", 3, 12, 6)
        self._sl_modal(sf, 'iteraciones_cierre', "Iter Cierre", 1, 5, 2)
        self._agregar_botones_aplicar(win, sf)

    def _calibrar_piel_con_muestra(self, parent_win):
        """Usa la imagen actual con ventana Tkinter para auto-calibrar piel de cocodrilo"""
        if not self.imagenes_cargadas:
            messagebox.showwarning("Error", "Cargue imagenes primero.", parent=parent_win)
            return
        ruta = self.imagenes_cargadas[self.imagen_actual_idx]
        imagen = cv2.imread(ruta)
        if imagen is None:
            messagebox.showwarning("Error", "No se pudo cargar la imagen actual.", parent=parent_win)
            return
        # Ventana con cÃ¡lculo automÃ¡tico integrado
        sel = VentanaSeleccionROI(parent_win, imagen, "Calibrar Piel de Cocodrilo")
        if sel.resultado is None:
            self._log("Calibracion cancelada.")
            return
        params = sel.resultado  # dict con params ya calculados
        # Actualizar las variables de la interfaz
        self.vars['clahe_clip'].set(params.get('clahe_clip', 4.0))
        self.vars['clahe_tile'].set(params.get('clahe_tile', 8))
        self.vars['bilateral_d'].set(params.get('bilateral_d', 9))
        self.vars['bilateral_sigma_color'].set(params.get('bilateral_sigma_color', 75))
        self.vars['bilateral_sigma_space'].set(params.get('bilateral_sigma_space', 75))
        self.vars['block_size'].set(params.get('block_size', 23))
        self.vars['C_umbral'].set(params.get('C_umbral', 10))
        self.vars['kernel_apertura'].set(params.get('kernel_apertura', 3))
        self.vars['kernel_cierre'].set(params.get('kernel_cierre', 6))
        # Actualizar procesador
        self.motor.procesador_piel.params.update(params)
        nombre = Path(ruta).name
        roi_info = f"{sel.roi_coords[2]}x{sel.roi_coords[3]}" if sel.roi_coords else "?"
        self._lbl_piel_cal_estado.config(
            text=f"Calibrado: {nombre} (ROI {roi_info})", fg=EstiloUI.BG_SUCCESS)
        self._log(f"Piel de Cocodrilo calibrada: {nombre}")
        self._log(f"  CLAHE: clip={params['clahe_clip']}, tile={params['clahe_tile']}")
        self._log(f"  Bilateral: d={params['bilateral_d']}, sigmaC={params['bilateral_sigma_color']}, sigmaS={params['bilateral_sigma_space']}")
        self._log(f"  Umbral: block={params['block_size']}, C={params['C_umbral']}")
        self._log(f"  Kernel: apertura={params['kernel_apertura']}, cierre={params['kernel_cierre']}")

    def _abrir_poligonos(self):
        win = self._crear_ventana_modal("DetecciÃ³n de PolÃ­gonos", 500, 600)
        sf = tk.Frame(win, bg=EstiloUI.BG_PANEL); sf.pack(fill="both", expand=True)
        self._sec(sf, "DETECCION POLIGONOS")
        self._sl_modal(sf, 'min_area_poligono', "Area Min Polig", 50, 1000, 300, step=50)
        self._sl_modal(sf, 'min_circularidad', "Circ Min", 0.01, 0.5, 0.08, resolution=0.01)
        self._sl_modal(sf, 'min_vertices', "Vert Min", 3, 8, 4)
        self._sl_modal(sf, 'max_vertices', "Vert Max", 10, 50, 25)
        self._sl_modal(sf, 'min_radio_circulo', "Radio Min", 3, 30, 8)
        self._sl_modal(sf, 'min_longitud_rama', "Long Min Rama", 10, 100, 30, step=5)
        self._sl_modal(sf, 'min_area_objeto', "Area Min Obj", 20, 500, 100, step=10)
        self._sl_modal(sf, 'max_gap_cierre', "Max Gap Cierre", 5, 50, 20)
        self._agregar_botones_aplicar(win, sf)

    def _abrir_perfiles(self):
        win = self._crear_ventana_modal("Perfiles RÃ¡pidos", 480, 420)
        sf = tk.Frame(win, bg=EstiloUI.BG_PANEL); sf.pack(fill="both", expand=True)
        self._sec(sf, "PERFILES RAPIDOS")
        pf = tk.Frame(sf, bg=EstiloUI.BG_PANEL); pf.pack(fill="x", padx=10, pady=2)
        tk.Label(pf, text="P. Cocodrilo:", font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                 fg=EstiloUI.FG_SECONDARY).pack(anchor="w")
        bf_piel = tk.Frame(pf, bg=EstiloUI.BG_PANEL); bf_piel.pack(fill="x", pady=2)
        for nombre, cmd in [("Alta Sens.", self._perfil_piel_alta),
                             ("Equilibrado", self._perfil_piel_equilibrado),
                             ("Alta Prec.", self._perfil_piel_alta_prec)]:
            ttk.Button(bf_piel, text=nombre, style="Secondary.TButton",
                       command=cmd).pack(side="left", padx=2, fill="x", expand=True)
        tk.Label(pf, text="Parches:", font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                 fg=EstiloUI.FG_SECONDARY).pack(anchor="w", pady=(5,0))
        bf_pa = tk.Frame(pf, bg=EstiloUI.BG_PANEL); bf_pa.pack(fill="x", pady=2)
        for nombre, cmd in [("Sin filtro", self._perfil_parche_sin_filtro),
                             ("Normal", self._perfil_parche_normal),
                             ("Estricto", self._perfil_parche_estricto)]:
            ttk.Button(bf_pa, text=nombre, style="Secondary.TButton",
                       command=cmd).pack(side="left", padx=2, fill="x", expand=True)
        tk.Label(pf, text="General:", font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                 fg=EstiloUI.FG_SECONDARY).pack(anchor="w", pady=(5,0))
        bf_gen = tk.Frame(pf, bg=EstiloUI.BG_PANEL); bf_gen.pack(fill="x", pady=2)
        for nombre, cmd in [("Alta Sens.", self._perfil_general_alta_sens),
                             ("Equilibrado", self._perfil_general_equilibrado),
                             ("Alta Prec.", self._perfil_general_alta_prec)]:
            ttk.Button(bf_gen, text=nombre, style="Secondary.TButton",
                       command=cmd).pack(side="left", padx=2, fill="x", expand=True)
        self._agregar_botones_aplicar(win, sf)

    # =========================================================================
    # PERFILES RÃPIDOS - PIEL DE COCODRILO
    # =========================================================================
    def _perfil_piel_alta(self):
        vals = {'clahe_clip': 4.5, 'clahe_tile': 6, 'bilateral_d': 7, 'bilateral_sigma_color': 60,
                'bilateral_sigma_space': 60, 'block_size': 19, 'C_umbral': 8, 'kernel_apertura': 2,
                'kernel_cierre': 7, 'iteraciones_cierre': 3, 'min_area_poligono': 150,
                'min_circularidad': 0.05, 'min_radio_circulo': 5, 'min_longitud_rama': 15,
                'min_area_objeto': 50, 'max_gap_cierre': 30, 'usar_frangi': True, 'usar_multiescala': True}
        self._aplicar_vals(vals); self._log("Perfil P.Cocodrilo: Alta Sensibilidad")

    def _perfil_piel_equilibrado(self):
        vals = {'clahe_clip': 4.0, 'clahe_tile': 8, 'bilateral_d': 9, 'bilateral_sigma_color': 75,
                'bilateral_sigma_space': 75, 'block_size': 23, 'C_umbral': 10, 'kernel_apertura': 3,
                'kernel_cierre': 6, 'iteraciones_cierre': 2, 'min_area_poligono': 300,
                'min_circularidad': 0.08, 'min_radio_circulo': 8, 'min_longitud_rama': 30,
                'min_area_objeto': 100, 'max_gap_cierre': 20, 'usar_frangi': True, 'usar_multiescala': True}
        self._aplicar_vals(vals); self._log("Perfil P.Cocodrilo: Equilibrado")

    def _perfil_piel_alta_prec(self):
        vals = {'clahe_clip': 3.0, 'clahe_tile': 10, 'bilateral_d': 11, 'bilateral_sigma_color': 90,
                'bilateral_sigma_space': 90, 'block_size': 31, 'C_umbral': 14, 'kernel_apertura': 4,
                'kernel_cierre': 5, 'iteraciones_cierre': 1, 'min_area_poligono': 500,
                'min_circularidad': 0.12, 'min_radio_circulo': 12, 'min_longitud_rama': 50,
                'min_area_objeto': 200, 'max_gap_cierre': 15, 'usar_frangi': False, 'usar_multiescala': False}
        self._aplicar_vals(vals); self._log("Perfil P.Cocodrilo: Alta Precision")

    # PERFILES PARCHES
    def _perfil_parche_sin_filtro(self):
        self._aplicar_vals({'min_area_parche_m2': 0, 'filtrar_parches': False})
        self._log("Perfil Parches: Sin filtro (todos)")
    def _perfil_parche_normal(self):
        self._aplicar_vals({'min_area_parche_m2': 0.01, 'filtrar_parches': True})
        self._log("Perfil Parches: Normal (>0.01 m2)")
    def _perfil_parche_estricto(self):
        self._aplicar_vals({'min_area_parche_m2': 0.05, 'filtrar_parches': True})
        self._log("Perfil Parches: Estricto (>0.05 m2)")

    # PERFILES GENERALES
    def _perfil_general_alta_sens(self):
        self._aplicar_vals({'confianza_min': 0.05, 'min_diametro_hueco_mm': 0,
                             'min_longitud_grieta_m': 0.01, 'min_area_parche_m2': 0.001,
                             'min_area_piel_m2': 0.01})
        self._log("Perfil General: Alta Sensibilidad")
    def _perfil_general_equilibrado(self):
        self._aplicar_vals({'confianza_min': 0.10, 'min_diametro_hueco_mm': 50,
                             'min_longitud_grieta_m': 0.05, 'min_area_parche_m2': 0.01,
                             'min_area_piel_m2': 0.05})
        self._log("Perfil General: Equilibrado")
    def _perfil_general_alta_prec(self):
        self._aplicar_vals({'confianza_min': 0.25, 'min_diametro_hueco_mm': 100,
                             'min_longitud_grieta_m': 0.10, 'min_area_parche_m2': 0.05,
                             'min_area_piel_m2': 0.10})
        self._log("Perfil General: Alta Precision")

    def _aplicar_vals(self, vals):
        for k, v in vals.items():
            if k in self.vars:
                self.vars[k].set(v)

    def _get_mallas(self):
        return self.vars.get('mostrar_mallas', tk.BooleanVar(value=True)).get()
    def _get_circulos(self):
        return self.vars.get('mostrar_circulos', tk.BooleanVar(value=True)).get()
    def _get_etiquetas(self):
        return self.vars.get('mostrar_etiquetas', tk.BooleanVar(value=True)).get()
    def _get_numeros(self):
        return self.vars.get('mostrar_numeros', tk.BooleanVar(value=True)).get()
    def _get_escala_texto(self):
        return self.vars.get('escala_texto', tk.DoubleVar(value=1.0)).get()
    def _on_escala_texto(self):
        val = self._get_escala_texto()
        if hasattr(self, 'lbl_escala_texto'):
            self.lbl_escala_texto.config(text=f"{val:.1f}x")

    def _refrescar_vista(self):
        """Refresca la vista actual con el estado de mostrar_mallas y fallas consideradas"""
        if self.resultado_actual and self.resultado_actual.get('fallas') and self.vista_actual == "resultado":
            fallas_filtradas = self._filtrar_fallas_por_considerar(self.resultado_actual['fallas'])
            vis = self.motor.dibujar_resultado(
                self.resultado_actual['imagen'], fallas_filtradas,
                mostrar_mallas=self._get_mallas(), mostrar_etiquetas=self._get_etiquetas(),
                mostrar_numeros=self._get_numeros(), escala_texto=self._get_escala_texto(), mostrar_circulos=self._get_circulos())
            self._mostrar_cv2(vis, reset_zoom=False)
            self._actualizar_conteos(fallas_filtradas)


    # =========================================================================
    # ZOOM & PAN
    # =========================================================================
    def _zoom_in(self): self._apply_zoom(self._zoom_level * 1.25)
    def _zoom_out(self): self._apply_zoom(self._zoom_level / 1.25)
    def _zoom_100(self): self._apply_zoom(1.0)

    def _zoom_fit(self):
        if self._cv2_original_full is None: return
        h, w = self._cv2_original_full.shape[:2]
        cw = max(self.canvas_imagen.winfo_width(), 100)
        ch = max(self.canvas_imagen.winfo_height(), 100)
        self._zoom_level = min(cw / w, ch / h)
        self._pan_offset_x = int((cw - w * self._zoom_level) / 2)
        self._pan_offset_y = int((ch - h * self._zoom_level) / 2)
        self._render_zoom()

    def _apply_zoom(self, nz, cx=None, cy=None):
        if self._cv2_original_full is None: return
        nz = max(self._zoom_min, min(self._zoom_max, nz))
        cw = max(self.canvas_imagen.winfo_width(), 100)
        ch = max(self.canvas_imagen.winfo_height(), 100)
        if cx is None: cx = cw / 2
        if cy is None: cy = ch / 2
        ix = (cx - self._pan_offset_x) / self._zoom_level
        iy = (cy - self._pan_offset_y) / self._zoom_level
        self._zoom_level = nz
        self._pan_offset_x = int(cx - ix * nz)
        self._pan_offset_y = int(cy - iy * nz)
        self._render_zoom()

    def _on_zoom_wheel(self, e):
        if self._cv2_original_full is None: return
        f = 1.15 if e.delta > 0 else 1.0 / 1.15
        self._apply_zoom(self._zoom_level * f, e.x, e.y)

    def _on_pan_start(self, e):
        self._pan_dragging = True
        self._pan_start_x = e.x
        self._pan_start_y = e.y

    def _on_pan_move(self, e):
        if not self._pan_dragging: return
        self._pan_offset_x += e.x - self._pan_start_x
        self._pan_offset_y += e.y - self._pan_start_y
        self._pan_start_x = e.x
        self._pan_start_y = e.y
        self._render_zoom()

    def _on_pan_end(self, e):
        self._pan_dragging = False

    def _on_canvas_resize(self, e=None):
        if self._cv2_original_full is not None:
            self._render_zoom()

    def _on_mouse_move(self, e):
        if self._cv2_original_full is None: return
        h, w = self._cv2_original_full.shape[:2]
        ix = int((e.x - self._pan_offset_x) / self._zoom_level)
        iy = int((e.y - self._pan_offset_y) / self._zoom_level)
        if 0 <= ix < w and 0 <= iy < h:
            b, g, r = self._cv2_original_full[iy, ix]
            hex_color = f"#{r:02x}{g:02x}{b:02x}"
            # Calcular color de texto legible (blanco o negro segun luminancia)
            lum = 0.299 * r + 0.587 * g + 0.114 * b
            fg_color = "#000000" if lum > 128 else "#ffffff"
            self.lbl_pixel_coords.config(
                text=f"RGB({r},{g},{b})", bg=hex_color, fg=fg_color)
        else:
            self.lbl_pixel_coords.config(
                text="X:-- Y:--", bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT)

    def _render_zoom(self):
        if self._cv2_original_full is None: return
        img = self._cv2_original_full
        ho, wo = img.shape[:2]
        cw = max(self.canvas_imagen.winfo_width(), 100)
        ch = max(self.canvas_imagen.winfo_height(), 100)
        ws = int(wo * self._zoom_level)
        hs = int(ho * self._zoom_level)
        xs = max(0, -self._pan_offset_x)
        ys = max(0, -self._pan_offset_y)
        xe = min(ws, cw - self._pan_offset_x)
        ye = min(hs, ch - self._pan_offset_y)
        if xe <= xs or ye <= ys:
            self.canvas_imagen.delete("all"); return
        x1 = max(0, int(xs / self._zoom_level))
        y1 = max(0, int(ys / self._zoom_level))
        x2 = min(wo, int(np.ceil(xe / self._zoom_level)))
        y2 = min(ho, int(np.ceil(ye / self._zoom_level)))
        if x2 <= x1 or y2 <= y1: return
        crop = img[y1:y2, x1:x2]
        cw2 = int((x2 - x1) * self._zoom_level)
        ch2 = int((y2 - y1) * self._zoom_level)
        if cw2 < 1 or ch2 < 1: return
        interp = cv2.INTER_NEAREST if self._zoom_level > 2 else cv2.INTER_LINEAR
        cs = cv2.resize(crop, (cw2, ch2), interpolation=interp)
        rgb = cv2.cvtColor(cs, cv2.COLOR_BGR2RGB)
        self._imagen_tk = ImageTk.PhotoImage(Image.fromarray(rgb))
        self.canvas_imagen.delete("all")
        self.canvas_imagen.create_image(
            max(0, self._pan_offset_x), max(0, self._pan_offset_y),
            anchor="nw", image=self._imagen_tk)
        self.lbl_zoom.config(text=f"{self._zoom_level * 100:.0f}%")

    def _mostrar_cv2(self, img, reset_zoom=True):
        self._cv2_actual = img
        self._cv2_original_full = img.copy()
        if reset_zoom:
            self._zoom_fit()
        else:
            self._render_zoom()
        # Actualizar info de imagen
        h, w = img.shape[:2]
        self.lbl_img_info.config(text=f"{w}x{h}px")

    # =========================================================================
    # ACCIONES PRINCIPALES
    # =========================================================================
    def _verificar_dependencias(self):
        deps = []
        if not YOLO_OK: deps.append("ultralytics")
        if not PANDAS_OK: deps.append("pandas")
        if not SKIMAGE_OK: deps.append("scikit-image")
        if not OPENPYXL_OK: deps.append("openpyxl")
        if deps:
            self._log(f"DEPENDENCIAS FALTANTES: {', '.join(deps)}")
            self._log(f"pip install {' '.join(deps)}")

    def _calibrar_unica(self, auto_procesar=False):
        """Calibrar con una sola imagen y aplicar a todas"""
        if not self.imagenes_cargadas:
            # Pedir imagen para calibrar
            ruta = filedialog.askopenfilename(title="Seleccionar imagen para calibrar",
                initialdir=self.motor.config['ruta_imagenes'],
                filetypes=[("Imagenes", "*.jpg *.jpeg *.png *.bmp *.tif *.tiff")])
            if not ruta: return
        else:
            # Usar la imagen actual
            ruta = self.imagenes_cargadas[self.imagen_actual_idx]
        imagen = cv2.imread(ruta)
        if imagen is None:
            self._log("ERROR: No se pudo cargar la imagen para calibrar"); return
        self._log(f"Calibrando con: {Path(ruta).name}")
        # Sincronizar ancho de via
        ancho_via = self.vars.get('ancho_via', tk.DoubleVar(value=6.5)).get()
        self.motor.calibrador.ancho_via_real_m = float(ancho_via)
        # Abrir ventana de calibraciÃ³n
        self.motor.calibrar_imagen(imagen, usar_gui=True, parent=self)
        if self.motor.calibrador.px_por_mm:
            px_mm = self.motor.calibrador.px_por_mm
            angulo = self.motor.calibrador.angulo_eje_via
            self._calibracion_unica_guardada = {
                'px_por_mm': px_mm,
                'angulo_eje_via': angulo
            }
            self.lbl_cal_estado.config(
                text=f"Calibrada: {px_mm:.4f} px/mm | Eje: {angulo:.1f}Â°",
                fg=EstiloUI.BG_SUCCESS)
            self._log(f"Calibracion unica guardada: {px_mm:.4f} px/mm | Eje: {angulo:.1f}Â°")
            # Forzar modo unica
            self.vars['modo_calibracion'].set("unica")
            # Auto-procesar si fue invocado desde _procesar_todo
            if auto_procesar:
                self._log("Calibracion confirmada, iniciando procesamiento...")
                self.after(100, self._procesar_todo)
            else:
                # Auto-procesar imagen actual tras calibracion standalone
                self._log("Calibracion confirmada, procesando imagen actual...")
                self.after(100, self._procesar_actual)
        else:
            self._log("Calibracion cancelada o fallida")

    def _detener_procesamiento(self):
        """Detiene el procesamiento batch"""
        if self.procesando:
            self._detener_flag = True
            self._log(">>> DETENIENDO procesamiento...")
            self._estado("Deteniendo...")
            self.btn_detener.config(state="disabled")

    def _cargar_modelo(self):
        ruta = filedialog.askopenfilename(title="Seleccionar modelo YOLO (.pt)",
            initialdir=str(Path(self.motor.config['ruta_modelo']).parent),
            filetypes=[("Modelos YOLO", "*.pt"), ("Todos", "*.*")])
        if not ruta: return
        self._log(f"Cargando modelo: {ruta}")
        self._estado("Cargando modelo...")
        def _load():
            ok, msg = self.motor.cargar_modelo(ruta)
            self.after(0, lambda: self._on_modelo(ok, msg, ruta))
        Thread(target=_load, daemon=True).start()

    def _on_modelo(self, ok, msg, ruta):
        self._log(msg)
        if ok:
            self.lbl_modelo.config(text=f"Modelo: {Path(ruta).name}", fg=EstiloUI.BG_SUCCESS)
            self._estado("Modelo cargado")
        else:
            self._estado("Error al cargar modelo")

    def _cargar_imagenes(self):
        rutas = filedialog.askopenfilenames(title="Seleccionar imagenes",
            initialdir=self.motor.config['ruta_imagenes'],
            filetypes=[("Imagenes", "*.jpg *.jpeg *.png *.bmp *.tif *.tiff"), ("Todos", "*.*")])
        if not rutas: return
        self.imagenes_cargadas = list(rutas)
        self.imagen_actual_idx = 0
        self._log(f"{len(self.imagenes_cargadas)} imagenes cargadas")
        self._mostrar_imagen_actual()
        self._actualizar_nav()

    def _imagen_anterior(self):
        if self.imagenes_cargadas and self.imagen_actual_idx > 0:
            self.imagen_actual_idx -= 1
            self._cargar_resultado_nav()
            self._actualizar_nav()

    def _imagen_siguiente(self):
        if self.imagenes_cargadas and self.imagen_actual_idx < len(self.imagenes_cargadas) - 1:
            self.imagen_actual_idx += 1
            self._cargar_resultado_nav()
            self._actualizar_nav()

    def _cargar_resultado_nav(self):
        idx = self.imagen_actual_idx
        r = self._resultados_batch.get(idx) if self._resultados_batch else None
        if r and r.get('fallas'):
            self.resultado_actual = r
            ff = self._filtrar_fallas_por_considerar(r['fallas'])
            self._poblar_lista_fallas(r['fallas'])
            if self.vista_actual == "resultado":
                self._mostrar_cv2(self.motor.dibujar_resultado(r['imagen'], ff, mostrar_mallas=self._get_mallas(), mostrar_etiquetas=self._get_etiquetas(), mostrar_numeros=self._get_numeros(), escala_texto=self._get_escala_texto(), mostrar_circulos=self._get_circulos()))
            elif self.vista_actual == "pasos":
                self._mostrar_pasos_vista()
            else:
                self._mostrar_cv2(r['imagen'])
            self._actualizar_conteos(ff)
            self._actualizar_img_data(r)
        else:
            self.resultado_actual = r
            self._mostrar_imagen_actual()
            self._reset_conteos()
            self._poblar_lista_fallas([])

    def _actualizar_nav(self):
        if self.imagenes_cargadas:
            n = Path(self.imagenes_cargadas[self.imagen_actual_idx]).name
            procesada = "OK" if self.imagen_actual_idx in self._resultados_batch else "--"
            self.lbl_nav.config(text=f"{self.imagen_actual_idx + 1}/{len(self.imagenes_cargadas)} | {n} [{procesada}]")
        else:
            self.lbl_nav.config(text="Sin imagenes")

    def _mostrar_imagen_actual(self):
        if not self.imagenes_cargadas: return
        img = cv2.imread(self.imagenes_cargadas[self.imagen_actual_idx])
        if img is not None:
            self._mostrar_cv2(img)
            self.vista_actual = "original"
            self.lbl_vista.config(text="[Original]")
            h, w = img.shape[:2]
            self.img_data_labels['archivo'].config(text=Path(self.imagenes_cargadas[self.imagen_actual_idx]).name)
            self.img_data_labels['resolucion'].config(text=f"{w}x{h} px")
            self.img_data_labels['px_mm'].config(text=f"{self.motor.calibrador.px_por_mm:.4f}" if self.motor.calibrador.px_por_mm else "--")
            self.img_data_labels['eje'].config(text=f"{self.motor.calibrador.angulo_eje_via:.1f} grados")

    def _cambiar_vista(self, vista):
        self.vista_actual = vista
        if vista == "original":
            self._mostrar_imagen_actual()
            self.lbl_vista.config(text="[Original]")
        elif vista == "resultado":
            if self.resultado_actual and self.resultado_actual.get('fallas'):
                ff = self._filtrar_fallas_por_considerar(self.resultado_actual['fallas'])
                self._mostrar_cv2(self.motor.dibujar_resultado(
                    self.resultado_actual['imagen'], ff,
                    mostrar_mallas=self._get_mallas(), mostrar_etiquetas=self._get_etiquetas(), mostrar_numeros=self._get_numeros(), escala_texto=self._get_escala_texto(), mostrar_circulos=self._get_circulos()))
                self.lbl_vista.config(text="[Resultado]")
                self._actualizar_conteos(ff)
            else:
                self._log("Procese la imagen primero")
        elif vista == "pasos":
            self._mostrar_pasos_vista()

    def _mostrar_pasos_vista(self):
        if not self.resultado_actual: self._log("Procese la imagen primero"); return
        nombre = self.resultado_actual.get('nombre', '')
        mosaico = self.motor.generar_mosaico_pasos(nombre, self.resultado_actual['imagen'])
        if mosaico is not None:
            self._mostrar_cv2(mosaico)
            self.lbl_vista.config(text="[Pasos]")
        else:
            self._log("No hay pasos intermedios disponibles")

    def _actualizar_img_data(self, resultado):
        if not resultado: return
        self.img_data_labels['archivo'].config(text=resultado.get('nombre', '--'))
        w = resultado.get('ancho', 0)
        h = resultado.get('alto', 0)
        self.img_data_labels['resolucion'].config(text=f"{w}x{h} px")
        px = self.motor.calibrador.px_por_mm
        self.img_data_labels['px_mm'].config(text=f"{px:.4f}" if px else "--")
        self.img_data_labels['eje'].config(text=f"{self.motor.calibrador.angulo_eje_via:.1f} grados")

    def _filtrar_fallas_por_considerar(self, fallas, nombre=None):
        """Filtra fallas segÃºn selecciÃ³n individual por imagen"""
        if not fallas: return []
        if nombre is None:
            nombre = self.resultado_actual.get('nombre', '') if self.resultado_actual else ''
        excluidas = self._fallas_excluidas.get(nombre, set())
        return [f for i, f in enumerate(fallas) if i not in excluidas]

    def _poblar_lista_fallas(self, fallas):
        """Rellena la lista de checkboxes individuales de fallas"""
        for w in self.frame_fallas.winfo_children(): w.destroy()
        self._falla_vars = []
        nombre = self.resultado_actual.get('nombre', '') if self.resultado_actual else ''
        excluidas = self._fallas_excluidas.get(nombre, set())
        colores_sev = {1: "#00b894", 2: "#fdcb6e", 3: "#e94560"}
        for i, f in enumerate(fallas):
            var = tk.BooleanVar(value=(i not in excluidas))
            self._falla_vars.append((var, i, f))
            row = tk.Frame(self.frame_fallas, bg=EstiloUI.BG_INPUT)
            row.pack(fill="x", pady=1)
            sev = f.get('severidad', 1)
            col = colores_sev.get(sev, EstiloUI.FG_PRIMARY)
            cb = tk.Checkbutton(row, variable=var, bg=EstiloUI.BG_INPUT,
                                fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_BUTTON_SECONDARY,
                                activebackground=EstiloUI.BG_INPUT,
                                activeforeground=EstiloUI.FG_PRIMARY,
                                indicatoron=True, onvalue=True, offvalue=False,
                                command=lambda idx=i, v=var: self._toggle_falla(idx, v))
            cb.pack(side="left")
            t = f['tipo']
            if 'BACHE' in t or 'HUECO' in t:
                info = f"Bache D={f.get('diametro_mm',0):.0f}mm"
            elif 'LONGITUDINAL' in t:
                info = f"F.Long e={f.get('espesor_mm',0):.1f}mm L={f.get('longitud_m',0):.2f}m"
            elif 'TRANSVERSAL' in t:
                info = f"F.Trans e={f.get('espesor_mm',0):.1f}mm L={f.get('longitud_m',0):.2f}m"
            elif 'REPARACION' in t or 'PARCHADO' in t:
                info = f"Parche A={f.get('area_m2',0):.2f}m2"
            elif 'COCODRILO' in t:
                info = f"P.Coc D={f.get('diametro_mm',0):.0f}mm A={f.get('area_m2',0):.2f}m2"
                btn_config_indiv = ttk.Button(row, text="âš™ï¸", width=3,
                                            command=lambda idx=i: self._configurar_piel_falla(idx))
                btn_config_indiv.pack(side="right", padx=2)
            else:
                info = t[:20]
            lbl_text = f"#{i+1} {info} S{sev}"
            tk.Label(row, text=lbl_text, font=EstiloUI.FONT_SMALL,
                     bg=EstiloUI.BG_DARK, fg=col).pack(side="left", padx=2)
        total = len(fallas)
        sel = total - len(excluidas & set(range(total)))
        self.lbl_fallas_total.config(text=f"{sel}/{total}")

    def _toggle_falla(self, idx, var):
        """Toggle individual falla por checkbox"""
        nombre = self.resultado_actual.get('nombre', '') if self.resultado_actual else ''
        if nombre not in self._fallas_excluidas:
            self._fallas_excluidas[nombre] = set()
        if var.get():
            self._fallas_excluidas[nombre].discard(idx)
        else:
            self._fallas_excluidas[nombre].add(idx)
        self._redibujar_con_seleccion()

    def _on_click_falla(self, event):
        """Click en imagen para toggle la falla mÃ¡s cercana"""
        if self.vista_actual != "resultado": return
        if not self.resultado_actual or not self.resultado_actual.get('fallas'): return
        # Convertir coords canvas â†’ coords imagen real
        cx = self.canvas_imagen.canvasx(event.x)
        cy = self.canvas_imagen.canvasy(event.y)
        if not hasattr(self, '_cv2_original_full') or self._cv2_original_full is None: return
        h_img, w_img = self._cv2_original_full.shape[:2]
        cw = self.canvas_imagen.winfo_width()
        ch = self.canvas_imagen.winfo_height()
        # Calcular posiciÃ³n real en imagen considerando zoom y pan
        img_x = (cx - self._pan_offset_x) / self._zoom_level
        img_y = (cy - self._pan_offset_y) / self._zoom_level
        if img_x < 0 or img_y < 0: return
        # Buscar falla mÃ¡s cercana
        fallas = self.resultado_actual['fallas']
        mejor_idx = -1; mejor_dist = float('inf')
        for i, f in enumerate(fallas):
            fx, fy = f.get('ubicacion_x', 0), f.get('ubicacion_y', 0)
            dist = ((img_x - fx)**2 + (img_y - fy)**2)**0.5
            # Radio de tolerancia segÃºn tipo
            if 'BACHE' in f['tipo'] or 'HUECO' in f['tipo']:
                radio = max(f.get('diametro_px', 50) / 2, 30)
            elif 'FISURA' in f['tipo']:
                radio = max(f.get('espesor_px', 30), 40)
            else:
                radio = max(f.get('area_px', 900)**0.5, 40)
            if dist < radio and dist < mejor_dist:
                mejor_dist = dist; mejor_idx = i
        if mejor_idx == -1:
            # Buscar con radio mayor
            for i, f in enumerate(fallas):
                fx, fy = f.get('ubicacion_x', 0), f.get('ubicacion_y', 0)
                dist = ((img_x - fx)**2 + (img_y - fy)**2)**0.5
                if dist < 80 and dist < mejor_dist:
                    mejor_dist = dist; mejor_idx = i
        if mejor_idx >= 0:
            nombre = self.resultado_actual.get('nombre', '')
            if nombre not in self._fallas_excluidas:
                self._fallas_excluidas[nombre] = set()
            if mejor_idx in self._fallas_excluidas[nombre]:
                self._fallas_excluidas[nombre].discard(mejor_idx)
            else:
                self._fallas_excluidas[nombre].add(mejor_idx)
            # Actualizar checkbox en lista
            for var, idx, f in self._falla_vars:
                if idx == mejor_idx:
                    var.set(mejor_idx not in self._fallas_excluidas[nombre])
                    break
            self._redibujar_con_seleccion()

    def _sel_todas_fallas(self, seleccionar):
        """Seleccionar o deseleccionar todas las fallas"""
        if not self.resultado_actual or not self.resultado_actual.get('fallas'): return
        nombre = self.resultado_actual.get('nombre', '')
        if seleccionar:
            self._fallas_excluidas[nombre] = set()
        else:
            self._fallas_excluidas[nombre] = set(range(len(self.resultado_actual['fallas'])))
        for var, idx, f in self._falla_vars:
            var.set(seleccionar)
        self._redibujar_con_seleccion()

    def _configurar_piel_falla(self, idx):
        """Abre una ventana para configurar la deteccion de piel de cocodrilo seleccionada e iterarla."""
        if not self.resultado_actual or 'fallas' not in self.resultado_actual: return
        falla = self.resultado_actual['fallas'][idx]
        
        win = self._crear_ventana_modal(f"Configurar Piel Cocodrilo #{idx+1}", 500, 500)
        sf = tk.Frame(win, bg=EstiloUI.BG_PANEL); sf.pack(fill="both", expand=True)
        self._sec(sf, f"CONFIGURACION INDIVIDUAL - FALLA #{idx+1}")
        
        local_vars = {}
        base_cfg = falla.get('config_personalizada', self.motor.config.copy())
        
        for k, v in [('clahe_clip',4.0),('clahe_tile',8),('bilateral_d',9),
                      ('bilateral_sigma_color',75),('bilateral_sigma_space',75),
                      ('block_size',23),('C_umbral',10),('kernel_apertura',3),
                      ('kernel_cierre',6),('iteraciones_cierre',2)]:
            val = base_cfg.get(k, base_cfg.get(k, v))
            local_vars[k] = tk.DoubleVar(value=val) if isinstance(v, float) else tk.IntVar(value=val)
            
        local_vars['usar_frangi'] = tk.BooleanVar(value=base_cfg.get('usar_frangi', True))
        local_vars['usar_multiescala'] = tk.BooleanVar(value=base_cfg.get('usar_multiescala', True))
        local_vars['usar_refinamiento'] = tk.BooleanVar(value=base_cfg.get('usar_refinamiento', True))

        def _sl_local(parent, key, label, from_, to, default, step=1, resolution=None):
            f = tk.Frame(parent, bg=EstiloUI.BG_PANEL); f.pack(fill="x", padx=10, pady=2)
            hdr = tk.Frame(f, bg=EstiloUI.BG_PANEL); hdr.pack(fill="x")
            tk.Label(hdr, text=label, font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY).pack(side="left")
            var = local_vars[key]
            vl = tk.Label(hdr, text=str(default), font=("Segoe UI", 9, "bold"), bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT, anchor="e")
            vl.pack(side="right", padx=(4,0))
            res = resolution if resolution else step
            sl = tk.Scale(f, from_=from_, to=to, orient="horizontal", variable=var, resolution=res,
                          showvalue=False, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT,
                          troughcolor=EstiloUI.BG_DARK, highlightthickness=0,
                          activebackground=EstiloUI.BG_BUTTON, length=280, sliderlength=15, bd=0, relief="flat")
            sl.pack(fill="x", pady=(0,2))
            def upd(*a):
                try:
                    v = var.get(); vl.config(text=f"{v:.2f}" if (resolution and resolution < 1) else str(int(v)))
                except Exception: pass
            var.trace_add("write", upd); upd()

        _sl_local(sf, 'clahe_clip', "CLAHE Clip", 1.0, 8.0, local_vars['clahe_clip'].get(), resolution=0.5)
        _sl_local(sf, 'block_size', "Block Size", 11, 51, local_vars['block_size'].get(), step=2)
        _sl_local(sf, 'C_umbral', "Constante C", 3, 25, local_vars['C_umbral'].get())
        _sl_local(sf, 'kernel_cierre', "Kernel Cierre", 3, 12, local_vars['kernel_cierre'].get())
        _sl_local(sf, 'iteraciones_cierre', "Iter Cierre", 1, 5, local_vars['iteraciones_cierre'].get())
        
        tk.Checkbutton(sf, text="Filtro Frangi", variable=local_vars['usar_frangi'],
                        bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK,
                        font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)
        tk.Checkbutton(sf, text="Refinar esq. y cerrar gaps", variable=local_vars['usar_refinamiento'],
                        bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK,
                        font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)
        tk.Checkbutton(sf, text="Multi-escala", variable=local_vars['usar_multiescala'],
                        bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK,
                        font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)

        bf = tk.Frame(sf, bg=EstiloUI.BG_PANEL)
        bf.pack(fill="x", padx=10, pady=(10,5), side="bottom")
        ttk.Button(bf, text="Aplicar Cambios a esta DetecciÃ³n", style="Accent.TButton",
                   command=lambda: self._aplicar_piel_individual(win, idx, local_vars)).pack(fill="x", pady=2)
        ttk.Button(bf, text="Cancelar", style="Secondary.TButton",
                   command=win.destroy).pack(fill="x", pady=2)
        self._centrar_ventana_modal(win)

    def _aplicar_piel_individual(self, win, idx, local_vars):
        if not self.resultado_actual or 'fallas' not in self.resultado_actual: return
        falla_vieja = self.resultado_actual['fallas'][idx]
        if 'mascara_roi' not in falla_vieja:
            self._log("Mascara no disponible. Procese de nuevo la imagen.")
            win.destroy(); return
            
        custom_cfg = self.motor.config.copy()
        for k, var in local_vars.items():
            if isinstance(var, tk.BooleanVar): custom_cfg[k] = var.get()
            elif isinstance(var, tk.DoubleVar): custom_cfg[k] = float(var.get())
            else: custom_cfg[k] = int(var.get())
            
        custom_cfg['usar_frangi'] = local_vars['usar_frangi'].get()
        custom_cfg['usar_multiescala'] = local_vars['usar_multiescala'].get()
        custom_cfg['usar_refinamiento'] = local_vars['usar_refinamiento'].get()

        proc = ProcesadorPielCocodrilo(custom_cfg)
        self._estado(f"Reprocesando falla #{idx+1}...")
        
        def _reproc():
            try:
                nuevo_res = proc.procesar(falla_vieja['mascara_roi'], self.motor.calibrador, falla_vieja['confianza'], self.resultado_actual['imagen'])
                if nuevo_res:
                    n_falla = nuevo_res[0]
                    n_falla['config_personalizada'] = custom_cfg
                    if 'id' in falla_vieja: n_falla['id'] = falla_vieja['id']
                    self.resultado_actual['fallas'][idx] = n_falla
                    self.after(0, lambda: self._on_reprocesado(idx, win, "Falla reprocesada con exito."))
                else:
                    self.after(0, lambda: self._on_reprocesado(idx, win, "Reprocesamiento no arrojÃ³ resultados."))
            except Exception as e:
                self.after(0, lambda e=e: self._on_reprocesado(idx, win, f"Error: {str(e)}"))
                
        from threading import Thread
        Thread(target=_reproc, daemon=True).start()
        
    def _on_reprocesado(self, idx, win, msg):
        self._log(msg)
        ff = self._filtrar_fallas_por_considerar(self.resultado_actual['fallas'])
        self._poblar_lista_fallas(self.resultado_actual['fallas'])
        self._actualizar_conteos(ff)
        self._cambiar_vista("resultado")
        self._estado("Listo")
        win.destroy()

    def _configurar_piel_falla(self, idx):
        """Abre una ventana para configurar la deteccion de piel de cocodrilo seleccionada e iterarla."""
        if not self.resultado_actual or 'fallas' not in self.resultado_actual: return
        falla = self.resultado_actual['fallas'][idx]
        
        win = self._crear_ventana_modal(f"Configurar Piel Cocodrilo #{idx+1}", 500, 500)
        sf = tk.Frame(win, bg=EstiloUI.BG_PANEL); sf.pack(fill="both", expand=True)
        self._sec(sf, f"CONFIGURACION INDIVIDUAL - FALLA #{idx+1}")
        
        local_vars = {}
        base_cfg = falla.get('config_personalizada', self.motor.config.copy())
        
        for k, v in [('clahe_clip',4.0),('clahe_tile',8),('bilateral_d',9),
                      ('bilateral_sigma_color',75),('bilateral_sigma_space',75),
                      ('block_size',23),('C_umbral',10),('kernel_apertura',3),
                      ('kernel_cierre',6),('iteraciones_cierre',2)]:
            val = base_cfg.get(k, base_cfg.get(k, v))
            local_vars[k] = tk.DoubleVar(value=val) if isinstance(v, float) else tk.IntVar(value=val)
            
        local_vars['usar_frangi'] = tk.BooleanVar(value=base_cfg.get('usar_frangi', True))
        local_vars['usar_multiescala'] = tk.BooleanVar(value=base_cfg.get('usar_multiescala', True))
        local_vars['usar_refinamiento'] = tk.BooleanVar(value=base_cfg.get('usar_refinamiento', True))

        def _sl_local(parent, key, label, from_, to, default, step=1, resolution=None):
            f = tk.Frame(parent, bg=EstiloUI.BG_PANEL); f.pack(fill="x", padx=10, pady=2)
            hdr = tk.Frame(f, bg=EstiloUI.BG_PANEL); hdr.pack(fill="x")
            tk.Label(hdr, text=label, font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY).pack(side="left")
            var = local_vars[key]
            vl = tk.Label(hdr, text=str(default), font=("Segoe UI", 9, "bold"), bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT, anchor="e")
            vl.pack(side="right", padx=(4,0))
            res = resolution if resolution else step
            sl = tk.Scale(f, from_=from_, to=to, orient="horizontal", variable=var, resolution=res,
                          showvalue=False, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT,
                          troughcolor=EstiloUI.BG_DARK, highlightthickness=0,
                          activebackground=EstiloUI.BG_BUTTON, length=280, sliderlength=15, bd=0, relief="flat")
            sl.pack(fill="x", pady=(0,2))
            def upd(*a):
                try:
                    v = var.get(); vl.config(text=f"{v:.2f}" if (resolution and resolution < 1) else str(int(v)))
                except Exception: pass
            var.trace_add("write", upd); upd()

        _sl_local(sf, 'clahe_clip', "CLAHE Clip", 1.0, 8.0, local_vars['clahe_clip'].get(), resolution=0.5)
        _sl_local(sf, 'block_size', "Block Size", 11, 51, local_vars['block_size'].get(), step=2)
        _sl_local(sf, 'C_umbral', "Constante C", 3, 25, local_vars['C_umbral'].get())
        _sl_local(sf, 'kernel_cierre', "Kernel Cierre", 3, 12, local_vars['kernel_cierre'].get())
        _sl_local(sf, 'iteraciones_cierre', "Iter Cierre", 1, 5, local_vars['iteraciones_cierre'].get())
        
        tk.Checkbutton(sf, text="Filtro Frangi", variable=local_vars['usar_frangi'],
                        bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK,
                        font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)
        tk.Checkbutton(sf, text="Refinar esq. y cerrar gaps", variable=local_vars['usar_refinamiento'],
                        bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK,
                        font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)
        tk.Checkbutton(sf, text="Multi-escala", variable=local_vars['usar_multiescala'],
                        bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK,
                        font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)

        bf = tk.Frame(sf, bg=EstiloUI.BG_PANEL)
        bf.pack(fill="x", padx=10, pady=(10,5), side="bottom")
        ttk.Button(bf, text="Aplicar Cambios a esta DetecciÃ³n", style="Accent.TButton",
                   command=lambda: self._aplicar_piel_individual(win, idx, local_vars)).pack(fill="x", pady=2)
        ttk.Button(bf, text="Cancelar", style="Secondary.TButton",
                   command=win.destroy).pack(fill="x", pady=2)
        self._centrar_ventana_modal(win)

    def _aplicar_piel_individual(self, win, idx, local_vars):
        if not self.resultado_actual or 'fallas' not in self.resultado_actual: return
        falla_vieja = self.resultado_actual['fallas'][idx]
        if 'mascara_roi' not in falla_vieja:
            self._log("Mascara no disponible. Procese de nuevo la imagen.")
            win.destroy(); return
            
        custom_cfg = self.motor.config.copy()
        for k, var in local_vars.items():
            if isinstance(var, tk.BooleanVar): custom_cfg[k] = var.get()
            elif isinstance(var, tk.DoubleVar): custom_cfg[k] = float(var.get())
            else: custom_cfg[k] = int(var.get())
            
        custom_cfg['usar_frangi'] = local_vars['usar_frangi'].get()
        custom_cfg['usar_multiescala'] = local_vars['usar_multiescala'].get()
        custom_cfg['usar_refinamiento'] = local_vars['usar_refinamiento'].get()

        proc = ProcesadorPielCocodrilo(custom_cfg)
        self._estado(f"Reprocesando falla #{idx+1}...")
        
        def _reproc():
            try:
                nuevo_res = proc.procesar(falla_vieja['mascara_roi'], self.motor.calibrador, falla_vieja['confianza'], self.resultado_actual['imagen'])
                if nuevo_res:
                    n_falla = nuevo_res[0]
                    n_falla['config_personalizada'] = custom_cfg
                    if 'id' in falla_vieja: n_falla['id'] = falla_vieja['id']
                    self.resultado_actual['fallas'][idx] = n_falla
                    self.after(0, lambda: self._on_reprocesado(idx, win, "Falla reprocesada con exito."))
                else:
                    self.after(0, lambda: self._on_reprocesado(idx, win, "Reprocesamiento no arrojÃ³ resultados."))
            except Exception as e:
                self.after(0, lambda e=e: self._on_reprocesado(idx, win, f"Error: {str(e)}"))
                
        from threading import Thread
        Thread(target=_reproc, daemon=True).start()
        
    def _on_reprocesado(self, idx, win, msg):
        self._log(msg)
        ff = self._filtrar_fallas_por_considerar(self.resultado_actual['fallas'])
        self._poblar_lista_fallas(self.resultado_actual['fallas'])
        self._actualizar_conteos(ff)
        self._cambiar_vista("resultado")
        self._estado("Listo")
        win.destroy()

    def _redibujar_con_seleccion(self):
        """Redibujar imagen con fallas seleccionadas"""
        if not self.resultado_actual: return
        ff = self._filtrar_fallas_por_considerar(self.resultado_actual['fallas'])
        if self.vista_actual == "resultado":
            mm = self._get_mallas()
            vis = self.motor.dibujar_resultado(self.resultado_actual['imagen'], ff,
                                                mostrar_mallas=mm, mostrar_etiquetas=self._get_etiquetas(), mostrar_numeros=self._get_numeros(), escala_texto=self._get_escala_texto(), mostrar_circulos=self._get_circulos())
            self._mostrar_cv2(vis, reset_zoom=False)
        self._actualizar_conteos(ff)
        # Actualizar total label
        total = len(self.resultado_actual['fallas'])
        sel = len(ff)
        self.lbl_fallas_total.config(text=f"{sel}/{total}")

    def _aplicar_filtro_fallas(self):
        """Compatibilidad - redibujar con seleccion actual"""
        self._redibujar_con_seleccion()

    def _actualizar_conteos(self, fallas):
        s1 = sum(1 for f in fallas if f['severidad'] == 1)
        s2 = sum(1 for f in fallas if f['severidad'] == 2)
        s3 = sum(1 for f in fallas if f['severidad'] == 3)
        self.sev_labels['s1'].config(text=str(s1))
        self.sev_labels['s2'].config(text=str(s2))
        self.sev_labels['s3'].config(text=str(s3))

    def _reset_conteos(self):
        self.lbl_fallas_total.config(text="0/0")
        for k in self.sev_labels:
            self.sev_labels[k].config(text="0")
        self._poblar_lista_fallas([])

    def _redibujar_con_seleccion(self):
        """Redibujar imagen con fallas seleccionadas"""
        if not self.resultado_actual: return
        ff = self._filtrar_fallas_por_considerar(self.resultado_actual['fallas'])
        if self.vista_actual == "resultado":
            mm = self._get_mallas()
            vis = self.motor.dibujar_resultado(self.resultado_actual['imagen'], ff,
                                                mostrar_mallas=mm, mostrar_etiquetas=self._get_etiquetas(), mostrar_numeros=self._get_numeros(), escala_texto=self._get_escala_texto(), mostrar_circulos=self._get_circulos())
            self._mostrar_cv2(vis, reset_zoom=False)
        self._actualizar_conteos(ff)
        # Actualizar total label
        total = len(self.resultado_actual['fallas'])
        sel = len(ff)
        self.lbl_fallas_total.config(text=f"{sel}/{total}")

    def _aplicar_filtro_fallas(self):
        """Compatibilidad - redibujar con seleccion actual"""
        self._redibujar_con_seleccion()

    def _actualizar_conteos(self, fallas):
        s1 = sum(1 for f in fallas if f['severidad'] == 1)
        s2 = sum(1 for f in fallas if f['severidad'] == 2)
        s3 = sum(1 for f in fallas if f['severidad'] == 3)
        self.sev_labels['s1'].config(text=str(s1))
        self.sev_labels['s2'].config(text=str(s2))
        self.sev_labels['s3'].config(text=str(s3))

    def _reset_conteos(self):
        self.lbl_fallas_total.config(text="0/0")
        for k in self.sev_labels:
            self.sev_labels[k].config(text="0")
        self._poblar_lista_fallas([])

    # =========================================================================
    # PROCESAMIENTO
    # =========================================================================
    def _iniciar_cronometro(self):
        """Inicia un cronometro visual que actualiza lbl_tiempo cada 100ms"""
        self._t_inicio = time.time()
        self._cronometro_activo = True
        self._actualizar_cronometro()

    def _actualizar_cronometro(self):
        """Actualiza el cronometro en pantalla mientras se procesa"""
        if not self._cronometro_activo: return
        elapsed = time.time() - self._t_inicio
        if elapsed < 60:
            txt = f"{elapsed:.1f}s"
        else:
            mins = int(elapsed // 60); secs = elapsed % 60
            txt = f"{mins}m {secs:.1f}s"
        self.lbl_tiempo.config(text=f"â± {txt}")
        self._cronometro_after_id = self.after(100, self._actualizar_cronometro)

    def _detener_cronometro(self):
        """Detiene el cronometro visual"""
        self._cronometro_activo = False
        if hasattr(self, '_cronometro_after_id') and self._cronometro_after_id:
            self.after_cancel(self._cronometro_after_id)
            self._cronometro_after_id = None

    def _procesar_actual(self):
        if not self.imagenes_cargadas:
            messagebox.showinfo("Info", "Cargue imagenes primero"); return
        if not self.motor.modelo_cargado:
            messagebox.showinfo("Info", "Cargue modelo YOLO primero"); return
        if self.procesando: return
        self.procesando = True
        self._sync_params()
        ruta = self.imagenes_cargadas[self.imagen_actual_idx]
        if self.calibrar_cada_imagen:
            self.motor.calibrador.px_por_mm = None
        self._log(f"\n{'=' * 50}\nProcesando: {Path(ruta).name}")
        self._estado(f"Procesando {Path(ruta).name}...")
        self._iniciar_cronometro()

        def _proc():
            # Cargar config especifica de esta imagen si existe
            config_global = self.motor.config.copy()
            if hasattr(self, '_config_avanzada_por_imagen') and self.imagen_actual_idx in self._config_avanzada_por_imagen:
                self.motor.config.update(self._config_avanzada_por_imagen[self.imagen_actual_idx])
                self.motor.procesador_piel = ProcesadorPielCocodrilo(self.motor.config)
                self._log_safe("Usando configuracion avanzada personalizada para esta imagen.")
                
            r = self.motor.procesar_imagen(ruta, callback_log=self._log_safe,
                                            calibrar_gui=self.calibrar_cada_imagen, parent=self)
                                            
            # Restaurar config global
            self.motor.config = config_global
            self.motor.procesador_piel = ProcesadorPielCocodrilo(self.motor.config)
            
            self.after(0, lambda: self._on_procesado(r))
        Thread(target=_proc, daemon=True).start()

    def _on_procesado(self, r):
        self.procesando = False
        self._detener_cronometro()
        t_elapsed = time.time() - getattr(self, '_t_inicio', time.time())
        self._tiempo_proc = t_elapsed
        if r is None:
            self._log("Error en procesamiento"); self._estado("Error"); return
        self.resultado_actual = r
        self._resultados_batch[self.imagen_actual_idx] = r
        self._log(f"  Tiempo: {t_elapsed:.2f}s")
        self.lbl_tiempo.config(text=f"â± {t_elapsed:.2f}s")
        if r['fallas']:
            ff = self._filtrar_fallas_por_considerar(r['fallas'])
            vis = self.motor.dibujar_resultado(r['imagen'], ff, mostrar_mallas=self._get_mallas(), mostrar_etiquetas=self._get_etiquetas(), mostrar_numeros=self._get_numeros(), escala_texto=self._get_escala_texto(), mostrar_circulos=self._get_circulos())
            self._mostrar_cv2(vis)
            self.vista_actual = "resultado"
            self.lbl_vista.config(text="[Resultado]")
            self._poblar_lista_fallas(r['fallas'])
            self._actualizar_conteos(ff)
            self._actualizar_img_data(r)
        else:
            self._log("Sin fallas detectadas")
            self._reset_conteos()
        self._actualizar_nav()
        self._estado(f"Completado ({t_elapsed:.2f}s)")

    def _procesar_todo(self):
        if not self.imagenes_cargadas:
            messagebox.showinfo("Info", "Cargue imagenes primero"); return
        if not self.motor.modelo_cargado:
            messagebox.showinfo("Info", "Cargue modelo YOLO primero"); return
        if self.procesando: return
        # Verificar calibracion unica si esta seleccionada
        modo_cal = self.vars.get('modo_calibracion', tk.StringVar(value="automatica")).get()
        if modo_cal == "unica" and not self._calibracion_unica_guardada:
            resp = messagebox.askyesno("Calibracion unica",
                "No hay calibracion unica guardada.\nÂ¿Desea calibrar ahora con una imagen?")
            if resp:
                self._calibrar_unica(auto_procesar=True)
                return  # _procesar_todo se llamarÃ¡ automÃ¡ticamente tras calibrar
            else: return
        # Preguntar carpeta de salida
        carpeta = filedialog.askdirectory(title="Seleccionar carpeta de salida para resultados",
                                           initialdir=self.motor.config['ruta_salida'])
        if not carpeta: return
        self._carpeta_batch = carpeta
        self.procesando = True
        self._detener_flag = False
        self.btn_detener.config(state="normal")
        self._sync_params()
        self._iniciar_cronometro()
        Path(carpeta).mkdir(parents=True, exist_ok=True)

        def _proc_all():
            t_batch_inicio = time.time()
            total = len(self.imagenes_cargadas)
            self._resultados_batch = {}
            cr = os.path.join(carpeta, "Resultados"); Path(cr).mkdir(exist_ok=True)
            cp = os.path.join(carpeta, "Pasos"); Path(cp).mkdir(exist_ok=True)
            # Aplicar calibracion unica si corresponde
            if modo_cal == "unica" and self._calibracion_unica_guardada:
                self.motor.calibrador.px_por_mm = self._calibracion_unica_guardada['px_por_mm']
                self.motor.calibrador.angulo_eje_via = self._calibracion_unica_guardada['angulo_eje_via']
                self._log_safe(f"Usando calibracion unica: {self.motor.calibrador.px_por_mm:.4f} px/mm | Eje: {self.motor.calibrador.angulo_eje_via:.1f}Â°")
            procesadas = 0
            for i, ruta in enumerate(self.imagenes_cargadas):
                # Verificar flag de detener
                if self._detener_flag:
                    self._log_safe(f"\n>>> DETENIDO por usuario en imagen {i}/{total}")
                    break
                nombre = Path(ruta).name
                stem = Path(ruta).stem
                t_img = time.time()
                self._log_safe(f"\n[{i + 1}/{total}] {nombre}")
                self.after(0, lambda n=nombre, ii=i: self._estado(
                    f"Procesando {n} ({ii + 1}/{total})"))
                # Calibracion segun modo
                cal_gui = False
                if modo_cal == "cada_imagen":
                    self.motor.calibrador.px_por_mm = None
                    cal_gui = True
                elif modo_cal == "automatica":
                    self.motor.calibrador.px_por_mm = None  # recalibrar por ancho
                    
                # Aplicar config especifica de la imagen si la tiene
                config_global = self.motor.config.copy()
                if hasattr(self, '_config_avanzada_por_imagen') and i in self._config_avanzada_por_imagen:
                    self.motor.config.update(self._config_avanzada_por_imagen[i])
                    self.motor.procesador_piel = ProcesadorPielCocodrilo(self.motor.config)
                    self._log_safe("Usando config avanzada personalizada para esta imagen.")
                    
                # modo "unica" mantiene la calibracion guardada
                r = self.motor.procesar_imagen(ruta, callback_log=self._log_safe,
                                                calibrar_gui=cal_gui, parent=self)
                                                
                # Restaurar config global
                self.motor.config = config_global
                self.motor.procesador_piel = ProcesadorPielCocodrilo(self.motor.config)
                
                if r:
                    self._resultados_batch[i] = r
                    procesadas += 1
                    t_img_elapsed = time.time() - t_img
                    self._log_safe(f"  Tiempo imagen: {t_img_elapsed:.2f}s")
                    if r.get('fallas'):
                        vis = self.motor.dibujar_resultado(r['imagen'], r['fallas'], mostrar_mallas=self._get_mallas(), mostrar_etiquetas=self._get_etiquetas(), mostrar_numeros=self._get_numeros(), escala_texto=self._get_escala_texto(), mostrar_circulos=self._get_circulos())
                        cv2.imwrite(os.path.join(cr, f"MTC_{stem}.png"), vis)
                        mosaico = self.motor.generar_mosaico_pasos(r['nombre'], r['imagen'])
                        if mosaico is not None:
                            cv2.imwrite(os.path.join(cp, f"PASOS_{stem}.png"), mosaico)
                        pasos = self.motor.todos_pasos.get(r['nombre'], {})
                        for cls_key, pasos_cls in pasos.items():
                            for paso_nombre, paso_img in pasos_cls.items():
                                if paso_img is None: continue
                                img_out = paso_img
                                if len(img_out.shape) == 2:
                                    img_out = cv2.cvtColor(img_out, cv2.COLOR_GRAY2BGR)
                                cv2.imwrite(os.path.join(cp, f"{stem}_{cls_key}_{paso_nombre}.png"), img_out)
            # Exportar Excel
            if self._resultados_batch:
                excel = ExportadorExcel.exportar(self.motor.todos_resultados, carpeta,
                                                 self.motor.calibrador, "RESULTADOS_MTC_CONSOLIDADO.xlsx")
            else: excel = None
            t_batch_total = time.time() - t_batch_inicio
            self._tiempo_proc = t_batch_total
            estado = "DETENIDO" if self._detener_flag else "COMPLETADO"
            self._log_safe(f"\n{'=' * 50}")
            self._log_safe(f"BATCH {estado}: {procesadas}/{total} imagenes procesadas")
            self._log_safe(f"Tiempo total: {t_batch_total:.2f}s ({t_batch_total/max(procesadas,1):.2f}s/img)")
            self._log_safe(f"Resultados guardados en: {cr}")
            self._log_safe(f"Pasos guardados en: {cp}")
            if excel: self._log_safe(f"Excel: {excel}")
            self._log_safe("Navegue con Anterior/Siguiente para verificar")
            self.after(0, lambda t=t_batch_total: self._finalizar_batch(t))
        Thread(target=_proc_all, daemon=True).start()

    def _finalizar_batch(self, t_total=0):
        self.procesando = False
        self._detener_cronometro()
        self._detener_flag = False
        self.btn_detener.config(state="disabled")
        n = len(self._resultados_batch)
        estado_txt = f"Batch {'detenido' if n < len(self.imagenes_cargadas) else 'completado'}: {n} imagenes"
        if t_total > 0:
            mins = int(t_total // 60); secs = t_total % 60
            tiempo_txt = f"{mins}m {secs:.1f}s" if mins > 0 else f"{secs:.1f}s"
            estado_txt += f" ({tiempo_txt})"
            self.lbl_tiempo.config(text=f"â± {tiempo_txt}")
        self._estado(estado_txt)
        self._actualizar_nav()
        # Mostrar primera imagen con resultados
        for i in range(len(self.imagenes_cargadas)):
            r = self._resultados_batch.get(i)
            if r and r.get('fallas'):
                self.imagen_actual_idx = i
                self.resultado_actual = r
                self._actualizar_nav()
                ff = self._filtrar_fallas_por_considerar(r['fallas'])
                self._poblar_lista_fallas(r['fallas'])
                self._mostrar_cv2(self.motor.dibujar_resultado(r['imagen'], ff, mostrar_mallas=self._get_mallas(), mostrar_etiquetas=self._get_etiquetas(), mostrar_numeros=self._get_numeros(), escala_texto=self._get_escala_texto(), mostrar_circulos=self._get_circulos()))
                self.vista_actual = "resultado"
                self.lbl_vista.config(text="[Resultado]")
                self._actualizar_conteos(ff)
                self._actualizar_img_data(r)
                break

    # =========================================================================
    # EXPORTACIONES
    # =========================================================================
    def _get_resultados_filtrados(self):
        """Obtiene todos_resultados filtrados segun checkboxes de CONSIDERAR"""
        filtrados = {}
        for nombre, fallas in self.motor.todos_resultados.items():
            ff = self._filtrar_fallas_por_considerar(fallas, nombre)
            if ff: filtrados[nombre] = ff
        return filtrados

    def _exportar_excel(self):
        """Exportar Todo a Excel - usa resultados filtrados por CONSIDERAR"""
        if not self.motor.todos_resultados:
            messagebox.showinfo("Info", "No hay resultados. Procese imagenes primero."); return
        carpeta = filedialog.askdirectory(title="Carpeta para Excel",
                                           initialdir=self.motor.config['ruta_salida'])
        if not carpeta: return
        Path(carpeta).mkdir(parents=True, exist_ok=True)
        filtrados = self._get_resultados_filtrados()
        path = ExportadorExcel.exportar(filtrados, carpeta, self.motor.calibrador,
                                         "RESULTADOS_MTC_CONSOLIDADO.xlsx")
        if path:
            self._log(f"Excel exportado: {path} ({sum(len(v) for v in filtrados.values())} fallas consideradas)")
            messagebox.showinfo("Exportado", f"Guardado:\n{path}")
        else:
            self._log("Error al exportar Excel")

    def _exportar_nuevo_excel(self):
        """Exportar Nuevo Excel - genera archivo con timestamp, filtrado por CONSIDERAR"""
        if not self.motor.todos_resultados:
            messagebox.showinfo("Info", "No hay resultados. Procese imagenes primero."); return
        carpeta = filedialog.askdirectory(title="Carpeta para Nuevo Excel",
                                           initialdir=self.motor.config['ruta_salida'])
        if not carpeta: return
        Path(carpeta).mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre = f"MTC_NUEVO_{ts}.xlsx"
        filtrados = self._get_resultados_filtrados()
        path = ExportadorExcel.exportar(filtrados, carpeta, self.motor.calibrador, nombre)
        if path:
            self._log(f"Nuevo Excel exportado: {path} ({sum(len(v) for v in filtrados.values())} fallas consideradas)")
            messagebox.showinfo("Exportado", f"Nuevo Excel guardado:\n{path}")
        else:
            self._log("Error al exportar Nuevo Excel")

    def _exportar_resultados_nuevos(self):
        """Exportar TODOS los resultados (corregidos y no corregidos) con fallas excluidas aplicadas"""
        if not self.motor.todos_resultados:
            messagebox.showinfo("Info", "No hay resultados. Procese imagenes primero."); return
        if not self._resultados_batch:
            messagebox.showinfo("Info", "No hay resultados batch. Procese con 'Procesar Todo' primero."); return
        carpeta = filedialog.askdirectory(title="Seleccionar carpeta para exportar todos los resultados",
                                           initialdir=self.motor.config['ruta_salida'])
        if not carpeta: return
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        carpeta_export = os.path.join(carpeta, f"Resultados_Nuevos_{ts}")
        Path(carpeta_export).mkdir(parents=True, exist_ok=True)
        cr = os.path.join(carpeta_export, "Resultados"); Path(cr).mkdir(exist_ok=True)
        cp = os.path.join(carpeta_export, "Pasos"); Path(cp).mkdir(exist_ok=True)
        total_imgs = len(self._resultados_batch)
        self._log(f"\nExportando todos los resultados ({total_imgs} imagenes)...")
        mm = self._get_mallas()
        mc = self._get_circulos()
        et = self._get_etiquetas()
        nm = self._get_numeros()
        esc = self._get_escala_texto()
        n_fallas = 0; n_excluidas_total = 0
        filtrados_excel = {}
        for i, ruta in enumerate(self.imagenes_cargadas):
            r = self._resultados_batch.get(i)
            if not r or not r.get('fallas'): continue
            nombre = r.get('nombre', Path(ruta).stem)
            stem = Path(ruta).stem
            # Aplicar exclusiones si existen
            ff = self._filtrar_fallas_por_considerar(r['fallas'], nombre)
            filtrados_excel[nombre] = ff
            n_fallas += len(ff)
            excluidas = self._fallas_excluidas.get(nombre, set())
            n_excl = len(excluidas)
            n_excluidas_total += n_excl
            mod = " (corregida)" if n_excl > 0 else ""
            self._log(f"  {stem}: {len(ff)} fallas{mod}{f' ({n_excl} excluidas)' if n_excl else ''}")
            # Imagen resultado con fallas filtradas y config visual actual
            vis = self.motor.dibujar_resultado(r['imagen'], ff,
                                                mostrar_mallas=mm, mostrar_etiquetas=et,
                                                mostrar_numeros=nm, escala_texto=esc, mostrar_circulos=mc)
            cv2.imwrite(os.path.join(cr, f"MTC_{stem}.png"), vis)
            # Mosaico de pasos
            mosaico = self.motor.generar_mosaico_pasos(nombre, r['imagen'])
            if mosaico is not None:
                cv2.imwrite(os.path.join(cp, f"PASOS_{stem}.png"), mosaico)
            # Pasos individuales
            pasos = self.motor.todos_pasos.get(nombre, {})
            for cls_key, pasos_cls in pasos.items():
                for paso_nombre, paso_img in pasos_cls.items():
                    if paso_img is None: continue
                    img_out = paso_img
                    if len(img_out.shape) == 2:
                        img_out = cv2.cvtColor(img_out, cv2.COLOR_GRAY2BGR)
                    cv2.imwrite(os.path.join(cp, f"{stem}_{cls_key}_{paso_nombre}.png"), img_out)
        # Excel con todas las fallas filtradas
        excel_path = None
        if filtrados_excel:
            excel_path = ExportadorExcel.exportar(filtrados_excel, carpeta_export,
                                                    self.motor.calibrador,
                                                    f"RESULTADOS_NUEVOS_{ts}.xlsx")
        imgs_corregidas = sum(1 for n in self._fallas_excluidas if len(self._fallas_excluidas[n]) > 0)
        self._log(f"{'=' * 50}")
        self._log(f"EXPORTACION COMPLETADA:")
        self._log(f"  Total imagenes: {total_imgs} ({imgs_corregidas} corregidas)")
        self._log(f"  Fallas consideradas: {n_fallas} ({n_excluidas_total} excluidas)")
        self._log(f"  Resultados: {cr}")
        self._log(f"  Pasos: {cp}")
        if excel_path: self._log(f"  Excel: {excel_path}")
        self._log(f"  Carpeta: {carpeta_export}")
        messagebox.showinfo("Exportado",
            f"Todos los resultados exportados:\n\n"
            f"Imagenes: {total_imgs} ({imgs_corregidas} corregidas)\n"
            f"Fallas consideradas: {n_fallas}\n"
            f"Fallas excluidas: {n_excluidas_total}\n\n"
            f"Carpeta: {carpeta_export}")

    def _guardar_resultado(self):
        if self._cv2_actual is None: return
        ruta = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[("PNG (sin perdida)", "*.png"), ("JPEG (alta calidad)", "*.jpg"), ("Todos", "*.*")],
            initialdir=self.motor.config['ruta_salida'])
        if not ruta: return
        if ruta.lower().endswith(('.jpg', '.jpeg')):
            cv2.imwrite(ruta, self._cv2_actual, [cv2.IMWRITE_JPEG_QUALITY, 98])
        else:
            cv2.imwrite(ruta, self._cv2_actual)
        self._log(f"Guardado: {ruta}")

    # =========================================================================
    # UTILIDADES
    # =========================================================================
    def _log(self, msg):
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)


class VentanaCalibracionTk(tk.Toplevel):
    """Ventana Tkinter para calibracion con zoom y herramientas amigables."""

    def __init__(self, parent, imagen_cv2, ancho_via_real_m=6.5):
        super().__init__(parent)
        self.title("Calibracion - Dibuje el ancho de la via")
        self.configure(bg=EstiloUI.BG_DARK)
        self.geometry("1100x750")
        self.transient(parent)
        self.grab_set()

        self.imagen_original = imagen_cv2.copy()
        self.ancho_via_real_m = ancho_via_real_m
        self.longitud_px = None
        self.completado = False
        self.resultado = None

        self._zoom = 1.0
        self._pan_x = 0
        self._pan_y = 0
        self._drawing = False
        self._pt1 = None
        self._pt2 = None
        self._pan_start = None

        self._crear_ui()
        self.after(150, self._mostrar_imagen)

        self.protocol("WM_DELETE_WINDOW", self._usar_ancho_imagen)
        self.bind("<Escape>", lambda e: self._usar_ancho_imagen())
        self.wait_window()

    def _crear_ui(self):
        # Barra superior
        top = tk.Frame(self, bg=EstiloUI.BG_PANEL)
        top.pack(fill="x", padx=5, pady=5)

        tk.Label(top, text=f"[ANCHO VÃA] Dibuje lÃ­nea transversal ({self.ancho_via_real_m}m)",
                 font=EstiloUI.FONT_SUBTITLE, bg=EstiloUI.BG_PANEL,
                 fg=EstiloUI.FG_HIGHLIGHT).pack(side="left", padx=10)

        self.lbl_info = tk.Label(top, text="Click izq: dibujar | Scroll: zoom | Click der: pan",
                                  font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                                  fg=EstiloUI.FG_SECONDARY)
        self.lbl_info.pack(side="right", padx=10)

        # Canvas
        self.canvas = tk.Canvas(self, bg=EstiloUI.BG_INPUT, highlightthickness=0)
        self.canvas.pack(fill="both", expand=True, padx=5, pady=5)

        self.canvas.bind("<ButtonPress-1>", self._on_click)
        self.canvas.bind("<B1-Motion>", self._on_drag)
        self.canvas.bind("<ButtonRelease-1>", self._on_release)
        self.canvas.bind("<MouseWheel>", self._on_wheel)
        self.canvas.bind("<ButtonPress-3>", self._on_pan_start)
        self.canvas.bind("<B3-Motion>", self._on_pan_move)
        self.canvas.bind("<ButtonRelease-3>", self._on_pan_end)

        # Barra inferior con botones estilizados
        bot = tk.Frame(self, bg=EstiloUI.BG_PANEL)
        bot.pack(fill="x", padx=5, pady=3)

        btn_confirmar = tk.Button(bot, text="CONFIRMAR",
                                   font=EstiloUI.FONT_SUBTITLE, bg=EstiloUI.BG_SUCCESS,
                                   fg="white", relief="flat", bd=0, padx=15, pady=6,
                                   command=self._confirmar)
        btn_confirmar.pack(side="left", padx=5)

        btn_reiniciar = tk.Button(bot, text="REINTENTAR",
                                   font=EstiloUI.FONT_SUBTITLE, bg=EstiloUI.BG_BUTTON_SECONDARY,
                                   fg="white", relief="flat", bd=0, padx=15, pady=6,
                                   command=self._reiniciar)
        btn_reiniciar.pack(side="left", padx=5)

        self.lbl_zoom = tk.Label(bot, text="Zoom: 1.0x",
                                  font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                                  fg=EstiloUI.FG_SECONDARY)
        self.lbl_zoom.pack(side="left", padx=15)

        btn_cerrar = tk.Button(bot, text="CERRAR / AUTO",
                                font=EstiloUI.FONT_SUBTITLE, bg=EstiloUI.BG_BUTTON,
                                fg="white", relief="flat", bd=0, padx=15, pady=6,
                                command=self._usar_ancho_imagen)
        btn_cerrar.pack(side="right", padx=5)

        self.lbl_dist = tk.Label(bot, text="",
                                  font=EstiloUI.FONT_BODY, bg=EstiloUI.BG_PANEL,
                                  fg=EstiloUI.FG_HIGHLIGHT)
        self.lbl_dist.pack(side="right", padx=10)

    def _mostrar_imagen(self):
        self.update_idletasks()
        cw = self.canvas.winfo_width() or 1000
        ch = self.canvas.winfo_height() or 600
        h, w = self.imagen_original.shape[:2]

        # Auto-fit zoom
        self._zoom = min(cw / w, ch / h, 1.0)
        self._pan_x = (cw - w * self._zoom) / 2
        self._pan_y = (ch - h * self._zoom) / 2
        self._render()

    def _render(self):
        self.canvas.delete("all")
        h, w = self.imagen_original.shape[:2]
        zw, zh = int(w * self._zoom), int(h * self._zoom)
        if zw < 1 or zh < 1:
            return

        img = cv2.resize(self.imagen_original, (zw, zh),
                         interpolation=cv2.INTER_NEAREST if self._zoom > 2 else cv2.INTER_LINEAR)
        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

        # Dibujar linea si existe
        if self._pt1 and self._pt2:
            p1 = (int(self._pt1[0] * self._zoom), int(self._pt1[1] * self._zoom))
            p2 = (int(self._pt2[0] * self._zoom), int(self._pt2[1] * self._zoom))
            cv2.line(img_rgb, p1, p2, (0, 255, 0), 3)
            mid = ((p1[0] + p2[0]) // 2, (p1[1] + p2[1]) // 2 - 15)
            dist = np.sqrt((self._pt2[0] - self._pt1[0])**2 + (self._pt2[1] - self._pt1[1])**2)
            cv2.putText(img_rgb, f"{dist:.0f} px = {self.ancho_via_real_m}m", mid,
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)

        pil = Image.fromarray(img_rgb)
        self._tk_img = ImageTk.PhotoImage(pil)
        self.canvas.create_image(int(self._pan_x), int(self._pan_y),
                                  anchor="nw", image=self._tk_img)
        self.lbl_zoom.config(text=f"Zoom: {self._zoom:.1f}x")

    def _canvas_to_img(self, cx, cy):
        ix = (cx - self._pan_x) / self._zoom
        iy = (cy - self._pan_y) / self._zoom
        return ix, iy

    def _on_click(self, event):
        self._drawing = True
        self._pt1 = self._canvas_to_img(event.x, event.y)
        self._pt2 = self._pt1

    def _on_drag(self, event):
        if self._drawing:
            self._pt2 = self._canvas_to_img(event.x, event.y)
            dist = np.sqrt((self._pt2[0] - self._pt1[0])**2 + (self._pt2[1] - self._pt1[1])**2)
            self.lbl_dist.config(text=f"Distancia: {dist:.0f} px")
            self._render()

    def _on_release(self, event):
        self._drawing = False
        if self._pt1 and self._pt2:
            dist = np.sqrt((self._pt2[0] - self._pt1[0])**2 + (self._pt2[1] - self._pt1[1])**2)
            self.lbl_dist.config(text=f"Distancia: {dist:.0f} px = {self.ancho_via_real_m}m")
            self.lbl_info.config(text="Presione 'Confirmar' para aceptar o 'Reiniciar'")

    def _on_wheel(self, event):
        factor = 1.15 if event.delta > 0 else 1 / 1.15
        old_zoom = self._zoom
        self._zoom = max(0.1, min(10.0, self._zoom * factor))
        # Zoom centrado en cursor
        self._pan_x = event.x - (event.x - self._pan_x) * (self._zoom / old_zoom)
        self._pan_y = event.y - (event.y - self._pan_y) * (self._zoom / old_zoom)
        self._render()

    def _on_pan_start(self, event):
        self._pan_start = (event.x, event.y)

    def _on_pan_move(self, event):
        if self._pan_start:
            dx = event.x - self._pan_start[0]
            dy = event.y - self._pan_start[1]
            self._pan_x += dx
            self._pan_y += dy
            self._pan_start = (event.x, event.y)
            self._render()

    def _on_pan_end(self, event):
        self._pan_start = None

    def _confirmar(self):
        if self._pt1 and self._pt2:
            dist = np.sqrt((self._pt2[0] - self._pt1[0])**2 + (self._pt2[1] - self._pt1[1])**2)
            if dist > 10:
                self.longitud_px = dist
                self.completado = True
                self.resultado = self.longitud_px
                self.destroy()

    def _reiniciar(self):
        self._pt1 = None
        self._pt2 = None
        self.lbl_dist.config(text="Distancia: -- px")
        self.lbl_info.config(text="Click izq: dibujar | Rueda: zoom | Click der: pan")
        self._render()

    def _usar_ancho_imagen(self):
        self.longitud_px = self.imagen_original.shape[1]
        self.completado = True
        self.resultado = self.longitud_px
        self.destroy()


# =============================================================================
# INTERFAZ GRAFICA PRINCIPAL
# =============================================================================

class AplicacionPCI(tk.Tk):
    """Ventana principal del sistema PCI."""

    def __init__(self):
        super().__init__()

        self.title("ASTM D6433 | YOLO")
        self.configure(bg=EstiloUI.BG_DARK)
        try:
            self.state('zoomed')
        except:
            self.geometry("1400x800")
        self.minsize(1280, 720)

        self.motor = MotorPCI()
        self.imagenes_cargadas = []
        self.imagen_actual_idx = -1
        self.resultado_actual = None
        self.vista_actual = "original"
        self.procesando = False
        self._detener_flag = False
        self.calibrar_cada_imagen = False
        self._resultados_batch = {}
        self._config_por_imagen = {}  # {idx: {key: valor, ...}} config avanzada individual por imagen
        self._config_global_base = {}  # snapshot de config global (se usa para imÃ¡genes sin config individual)
        self._tiempo_inicio = None
        self._crono_after_id = None
        self._crono_elapsed = 0.0
        self._carpeta_salida_batch = None
        self._falla_vars = []

        self.vars = {}
        self._cv2_actual = None
        self._cv2_original_full = None
        self._imagen_tk = None

        # Zoom y pan
        self._zoom_level = 1.0
        self._zoom_min = 0.1
        self._zoom_max = 10.0
        self._pan_offset_x = 0
        self._pan_offset_y = 0
        self._pan_dragging = False
        self._pan_start_x = 0
        self._pan_start_y = 0

        self._crear_estilos()
        self._crear_interfaz()
        self._verificar_dependencias()

    def _crear_estilos(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Dark.TFrame", background=EstiloUI.BG_DARK)
        style.configure("Panel.TFrame", background=EstiloUI.BG_PANEL)
        style.configure("Card.TFrame", background=EstiloUI.BG_CARD)
        style.configure("Accent.TButton", background=EstiloUI.BG_BUTTON,
                         foreground="white", font=EstiloUI.FONT_SUBTITLE, padding=(12, 6))
        style.map("Accent.TButton", background=[('active', EstiloUI.BG_BUTTON_HOVER)])
        style.configure("Secondary.TButton", background=EstiloUI.BG_BUTTON_SECONDARY,
                         foreground=EstiloUI.FG_PRIMARY, font=EstiloUI.FONT_BODY, padding=(8, 4))
        style.map("Secondary.TButton", background=[('active', EstiloUI.BG_ACCENT)])

    def _crear_interfaz(self):
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self._crear_barra_superior()

        # --- PanedWindow principal (paneles redimensionables) ---
        self.paned = tk.PanedWindow(self, orient=tk.HORIZONTAL,
                                     bg=EstiloUI.BG_DARK, sashwidth=6,
                                     sashrelief=tk.RAISED)
        self.paned.grid(row=1, column=0, sticky="nsew")

        # Panel izquierdo (config)
        self.frame_config = tk.Frame(self.paned, bg=EstiloUI.BG_PANEL, width=300)
        self.paned.add(self.frame_config, minsize=250, width=300)

        # Panel central (visor)
        self.frame_visor = tk.Frame(self.paned, bg=EstiloUI.BG_DARK)
        self.paned.add(self.frame_visor, minsize=400)

        # Panel derecho (resultados + log)
        self.frame_resultados = tk.Frame(self.paned, bg=EstiloUI.BG_PANEL, width=320)
        self.paned.add(self.frame_resultados, minsize=280, width=320)

        self._crear_panel_config(self.frame_config)
        self._crear_panel_visor(self.frame_visor)
        self._crear_panel_resultados(self.frame_resultados)

        # --- Barra inferior: estado + creditos ---
        self._crear_barra_estado()

    def _crear_barra_superior(self):
        barra = tk.Frame(self, bg=EstiloUI.BG_PANEL, height=55)
        barra.grid(row=0, column=0, sticky="ew")
        barra.grid_propagate(False)

        tk.Label(barra, text="ASTM D6433 | YOLO", font=EstiloUI.FONT_TITLE,
                 bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY).pack(side="left", padx=15, pady=8)

        btn_frame = tk.Frame(barra, bg=EstiloUI.BG_PANEL)
        btn_frame.pack(side="right", padx=10, pady=8)

        # Orden visual (izq->der): Modelo | Imagenes | Detener | ProcActual | ProcTodo | ExpResultNuevos | ExpNuevoExcel | ExpTodoExcel
        # pack(side=right) -> se definen en orden inverso
        ttk.Button(btn_frame, text="Exportar Todo Excel",
                   style="Accent.TButton",
                   command=self._exportar_excel).pack(side="right", padx=3)
        ttk.Button(btn_frame, text="Exportar Nuevo Excel",
                   style="Accent.TButton",
                   command=self._exportar_nuevo_excel).pack(side="right", padx=3)
        ttk.Button(btn_frame, text="Exportar Resultados Nuevos",
                   style="Accent.TButton",
                   command=self._exportar_resultados_nuevos).pack(side="right", padx=3)
        self.btn_procesar_todo = ttk.Button(btn_frame, text="Procesar Todo",
                                             style="Accent.TButton",
                                             command=self._procesar_todo)
        self.btn_procesar_todo.pack(side="right", padx=3)
        self.btn_procesar = ttk.Button(btn_frame, text="Procesar Actual",
                                        style="Accent.TButton",
                                        command=self._procesar_actual)
        self.btn_procesar.pack(side="right", padx=3)
        # Cronometro en vivo entre Detener y Procesar Actual
        self.lbl_cronometro = tk.Label(btn_frame, text="", font=("Consolas", 10, "bold"),
                                        bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT, width=8)
        self.lbl_cronometro.pack(side="right", padx=3)
        self.btn_detener = ttk.Button(btn_frame, text="Detener",
                                       style="Secondary.TButton",
                                       command=self._detener_procesamiento,
                                       state="disabled")
        self.btn_detener.pack(side="right", padx=3)
        ttk.Button(btn_frame, text="Cargar Imagenes",
                   style="Secondary.TButton",
                   command=self._cargar_imagenes).pack(side="right", padx=3)
        ttk.Button(btn_frame, text="Cargar Modelo",
                   style="Secondary.TButton",
                   command=self._cargar_modelo).pack(side="right", padx=3)

    def _crear_panel_config(self, parent):
        parent.grid_rowconfigure(0, weight=1)
        parent.grid_columnconfigure(0, weight=1)

        # Canvas con scroll vertical Y horizontal
        canvas = tk.Canvas(parent, bg=EstiloUI.BG_PANEL, highlightthickness=0)
        v_scroll = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        h_scroll = ttk.Scrollbar(parent, orient="horizontal", command=canvas.xview)
        scroll_frame = tk.Frame(canvas, bg=EstiloUI.BG_PANEL)
        scroll_frame.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

        canvas.grid(row=0, column=0, sticky="nsew")
        v_scroll.grid(row=0, column=1, sticky="ns")
        h_scroll.grid(row=1, column=0, sticky="ew")

        # Scroll con rueda del mouse solo cuando el cursor esta sobre el panel config
        def _on_enter(e):
            canvas.bind_all("<MouseWheel>", lambda ev: canvas.yview_scroll(int(-1*(ev.delta/120)), "units"))
        def _on_leave(e):
            canvas.unbind_all("<MouseWheel>")
        canvas.bind("<Enter>", _on_enter)
        canvas.bind("<Leave>", _on_leave)

        # === CONFIGURACION GENERAL ===
        self._seccion_titulo(scroll_frame, "CONFIGURACION GENERAL")

        self._crear_slider(scroll_frame, 'ancho_via', "Ancho Via (m)", 3.0, 12.0, 6.5, resolution=0.1)
        self._crear_slider(scroll_frame, 'confianza_min', "Confianza Min", 0.05, 0.95, 0.1, resolution=0.05)
        self._crear_slider(scroll_frame, 'iou_threshold', "IoU Threshold", 0.1, 0.9, 0.45, resolution=0.05)

        # === CALIBRACION ===
        self._seccion_titulo(scroll_frame, "CALIBRACION")

        self.vars['modo_calibracion'] = tk.StringVar(value='automatica')
        for val, txt in [
            ('automatica', "Automatica (ancho imagen)"),
            ('cada_imagen', "Calibrar cada imagen (OpenCV)"),
            ('unica', "Calibracion unica (1 ejemplar)"),
        ]:
            tk.Radiobutton(scroll_frame, text=txt, variable=self.vars['modo_calibracion'],
                           value=val, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY,
                           selectcolor=EstiloUI.BG_DARK, activebackground=EstiloUI.BG_PANEL,
                           font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)

        self.btn_calibrar_unica = ttk.Button(scroll_frame, text="Calibrar ahora (1 imagen)",
                                              style="Accent.TButton",
                                              command=self._calibrar_unica_ahora)
        self.btn_calibrar_unica.pack(fill="x", padx=20, pady=(2, 5))

        # Variables internas de compatibilidad (tramo/progresivas removidos de UI)
        self.vars['tramo'] = tk.StringVar(value='')
        self.vars['prog_km'] = tk.IntVar(value=0)
        self.vars['prog_m'] = tk.IntVar(value=0)
        self.vars['progresiva_inicio'] = tk.StringVar(value='000+000')
        self.vars['progresiva_incremento'] = tk.IntVar(value=20)

        # === CONFIGURACION AVANZADA ===
        self._seccion_titulo(scroll_frame, "CONFIGURACION AVANZADA")
        conf_avz_f = tk.Frame(scroll_frame, bg=EstiloUI.BG_PANEL)
        conf_avz_f.pack(fill="x", padx=10, pady=3)
        for sec_nombre, sec_id in [
            ("Filtros TamaÃ±o Minimo", "filtros"),
            ("Fusion Solapamientos", "fusion"),
            ("Piel de Cocodrilo", "piel"),
            ("Deteccion Poligonos", "poligonos"),
            ("Perfiles Rapidos", "perfiles"),
        ]:
            ttk.Button(conf_avz_f, text=sec_nombre, style="Secondary.TButton",
                       command=lambda s=sec_id: self._abrir_config_avanzada(seccion=s)).pack(
                fill="x", pady=1)

        # === VISUALIZACION (al final) ===
        self._seccion_titulo(scroll_frame, "VISUALIZACION")
        self.vars['text_size'] = tk.DoubleVar(value=0.45)

        for vkey, vtxt, vdef in [
            ('show_mallas', "Mostrar mallas/esqueleto", True),
            ('show_circulos', "Mostrar circulos", True),
            ('show_etiquetas', "Mostrar etiquetas", True),
            ('show_numeros', "Mostrar valores numericos", True),
        ]:
            self.vars[vkey] = tk.BooleanVar(value=vdef)
            tk.Checkbutton(scroll_frame, text=vtxt,
                            variable=self.vars[vkey],
                            bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY,
                            selectcolor=EstiloUI.BG_DARK, font=EstiloUI.FONT_BODY,
                            command=self._redibujar_resultado_actual).pack(
                anchor="w", padx=10, pady=1)

    def _crear_panel_visor(self, parent):
        parent.grid_rowconfigure(2, weight=1)
        parent.grid_columnconfigure(0, weight=1)

        # --- Fila 0: Navegacion ---
        nav = tk.Frame(parent, bg=EstiloUI.BG_CARD, height=38)
        nav.grid(row=0, column=0, sticky="ew", pady=(0, 2))

        ttk.Button(nav, text="< Anterior", style="Secondary.TButton",
                   command=self._imagen_anterior).pack(side="left", padx=5, pady=4)
        ttk.Button(nav, text="Siguiente >", style="Secondary.TButton",
                   command=self._imagen_siguiente).pack(side="left", padx=5, pady=4)

        self.lbl_nav = tk.Label(nav, text="Sin imagenes cargadas",
                                 font=EstiloUI.FONT_BODY, bg=EstiloUI.BG_CARD,
                                 fg=EstiloUI.FG_SECONDARY)
        self.lbl_nav.pack(side="left", padx=15)

        vista_frame = tk.Frame(nav, bg=EstiloUI.BG_CARD)
        vista_frame.pack(side="right", padx=10, pady=4)

        ttk.Button(vista_frame, text="Original",
                   style="Secondary.TButton",
                   command=lambda: self._cambiar_vista("original")).pack(side="left", padx=2)
        ttk.Button(vista_frame, text="Resultado",
                   style="Secondary.TButton",
                   command=lambda: self._cambiar_vista("resultado")).pack(side="left", padx=2)
        ttk.Button(vista_frame, text="Pasos",
                   style="Secondary.TButton",
                   command=lambda: self._cambiar_vista("pasos")).pack(side="left", padx=2)
        ttk.Button(vista_frame, text="Guardar",
                   style="Secondary.TButton",
                   command=self._guardar_resultado).pack(side="left", padx=2)

        # --- Fila 1: Barra de zoom e info ---
        zoom_bar = tk.Frame(parent, bg=EstiloUI.BG_PANEL, height=30)
        zoom_bar.grid(row=1, column=0, sticky="ew", pady=(0, 2))

        zoom_left = tk.Frame(zoom_bar, bg=EstiloUI.BG_PANEL)
        zoom_left.pack(side="left", padx=5, pady=2)

        ttk.Button(zoom_left, text="-", style="Secondary.TButton",
                   command=self._zoom_out).pack(side="left", padx=1)
        self.lbl_zoom = tk.Label(zoom_left, text="100%", font=EstiloUI.FONT_LABEL,
                                  bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT, width=6)
        self.lbl_zoom.pack(side="left", padx=3)
        ttk.Button(zoom_left, text="+", style="Secondary.TButton",
                   command=self._zoom_in).pack(side="left", padx=1)
        ttk.Button(zoom_left, text="Ajustar", style="Secondary.TButton",
                   command=self._zoom_fit).pack(side="left", padx=3)
        ttk.Button(zoom_left, text="1:1", style="Secondary.TButton",
                   command=self._zoom_100).pack(side="left", padx=1)

        # Separator
        tk.Label(zoom_left, text="|", font=EstiloUI.FONT_SMALL,
                 bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_SECONDARY).pack(side="left", padx=3)

        # Slider de texto para export (al lado de 1:1)
        tk.Label(zoom_left, text="Texto:", font=EstiloUI.FONT_SMALL,
                 bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_SECONDARY).pack(side="left", padx=(0, 2))
        self.text_size_slider = tk.Scale(zoom_left, from_=0.2, to=2.0, resolution=0.05,
                                          orient="horizontal", length=80,
                                          bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY,
                                          highlightthickness=0, troughcolor=EstiloUI.BG_DARK,
                                          showvalue=False)
        self.text_size_slider.set(self.vars.get('text_size', tk.DoubleVar(value=0.45)).get())
        self.text_size_slider.pack(side="left")
        self.lbl_text_size_val = tk.Label(zoom_left, text="0.45", font=EstiloUI.FONT_SMALL,
                                           bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT, width=4)
        self.lbl_text_size_val.pack(side="left", padx=(0, 3))
        self.text_size_slider.config(command=self._on_text_size_change)
        ttk.Button(zoom_left, text="Reescribir", style="Secondary.TButton",
                   command=self._redibujar_resultado).pack(side="left", padx=3)

        # Info excluir falla
        tk.Label(zoom_left, text=" | Click izq: Excluir/Incluir falla",
                 font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                 fg=EstiloUI.FG_SECONDARY).pack(side="left", padx=5)

        zoom_right = tk.Frame(zoom_bar, bg=EstiloUI.BG_PANEL)
        zoom_right.pack(side="right", padx=5, pady=2)

        # RGB label con fondo del color del pixel
        self.lbl_rgb = tk.Label(zoom_right, text="RGB(0,0,0)",
                                 font=("Segoe UI", 8, "bold"), bg="#000000",
                                 fg="#FFFFFF", padx=4, pady=1, relief="solid", borderwidth=1)
        self.lbl_rgb.pack(side="right", padx=(2, 5))

        self.lbl_img_info = tk.Label(zoom_right, text="--",
                                      font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                                      fg=EstiloUI.FG_SECONDARY)
        self.lbl_img_info.pack(side="right", padx=5)

        self.lbl_cursor_pos = tk.Label(zoom_right, text="X:-- Y:--",
                                        font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                                        fg=EstiloUI.FG_HIGHLIGHT, width=12)
        self.lbl_cursor_pos.pack(side="right", padx=5)

        # --- Fila 2: Canvas imagen ---
        self.canvas_imagen = tk.Canvas(parent, bg=EstiloUI.BG_INPUT, highlightthickness=0)
        self.canvas_imagen.grid(row=2, column=0, sticky="nsew")

        # Binds
        self.canvas_imagen.bind("<Configure>", self._on_canvas_resize)
        self.canvas_imagen.bind("<MouseWheel>", self._on_zoom_wheel)
        self.canvas_imagen.bind("<Button-4>", self._on_zoom_wheel_linux_up)
        self.canvas_imagen.bind("<Button-5>", self._on_zoom_wheel_linux_down)
        self.canvas_imagen.bind("<ButtonPress-2>", self._on_pan_start)
        self.canvas_imagen.bind("<B2-Motion>", self._on_pan_move)
        self.canvas_imagen.bind("<ButtonRelease-2>", self._on_pan_end)
        self.canvas_imagen.bind("<ButtonPress-3>", self._on_pan_start)
        self.canvas_imagen.bind("<B3-Motion>", self._on_pan_move)
        self.canvas_imagen.bind("<ButtonRelease-3>", self._on_pan_end)
        self.canvas_imagen.bind("<Motion>", self._on_mouse_move)
        # Click izquierdo para excluir/incluir fallas
        self.canvas_imagen.bind("<ButtonPress-1>", self._on_click_excluir_falla)

    def _crear_panel_resultados(self, parent):
        parent.grid_rowconfigure(3, weight=1)
        parent.grid_columnconfigure(0, weight=1)

        # --- Datos de imagen ---
        img_info_frame = tk.Frame(parent, bg=EstiloUI.BG_CARD)
        img_info_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=(5, 2))

        tk.Label(img_info_frame, text="DATOS DE IMAGEN", font=EstiloUI.FONT_SUBTITLE,
                 bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_ACCENT).pack(padx=10, pady=(6, 3), anchor="w")

        self.img_data_labels = {}
        for key, label_text in [
            ("archivo", "Archivo:"),
            ("dimensiones", "Dimensiones:"),
            ("pixeles_total", "Total Pixeles:"),
            ("canales", "Canales:"),
            ("profundidad", "Profundidad:"),
            ("tamano_archivo", "Tamano Archivo:"),
            ("resolucion_dpi", "px/mm (cal.):"),
        ]:
            row = tk.Frame(img_info_frame, bg=EstiloUI.BG_CARD)
            row.pack(fill="x", padx=10, pady=0)
            tk.Label(row, text=label_text, font=EstiloUI.FONT_SMALL,
                     bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_SECONDARY).pack(side="left")
            lbl = tk.Label(row, text="--", font=EstiloUI.FONT_SMALL,
                           bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_HIGHLIGHT)
            lbl.pack(side="right")
            self.img_data_labels[key] = lbl

        tk.Frame(img_info_frame, bg=EstiloUI.BG_CARD, height=4).pack()

        # --- FALLAS DETECTADAS (con checkboxes para incluir/excluir) ---
        fallas_frame = tk.Frame(parent, bg=EstiloUI.BG_CARD)
        fallas_frame.grid(row=1, column=0, sticky="ew", padx=5, pady=2)

        # Fila titulo + conteo
        fallas_header = tk.Frame(fallas_frame, bg=EstiloUI.BG_CARD)
        fallas_header.pack(fill="x", padx=10, pady=(6, 3))
        tk.Label(fallas_header, text="FALLAS DETECTADAS", font=EstiloUI.FONT_SUBTITLE,
                 bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_ACCENT).pack(side="left")
        self.lbl_fallas_conteo = tk.Label(fallas_header, text="0/0",
                                           font=EstiloUI.FONT_SUBTITLE,
                                           bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_HIGHLIGHT)
        self.lbl_fallas_conteo.pack(side="right")

        # Fila botones Todas/Ninguna + hint
        fallas_btn_row = tk.Frame(fallas_frame, bg=EstiloUI.BG_CARD)
        fallas_btn_row.pack(fill="x", padx=10, pady=(0, 3))
        btn_todas = tk.Button(fallas_btn_row, text="Todas",
                              font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                              fg=EstiloUI.FG_PRIMARY, relief="groove", bd=1,
                              command=self._seleccionar_todas_fallas)
        btn_todas.pack(side="left", padx=(0, 3), ipadx=8, ipady=1)
        btn_ninguna = tk.Button(fallas_btn_row, text="Ninguna",
                                font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                                fg=EstiloUI.FG_PRIMARY, relief="groove", bd=1,
                                command=self._deseleccionar_todas_fallas)
        btn_ninguna.pack(side="left", padx=(0, 5), ipadx=8, ipady=1)
        tk.Label(fallas_btn_row, text="Click en imagen para toggle",
                 font=("Segoe UI", 8, "italic"), bg=EstiloUI.BG_CARD,
                 fg=EstiloUI.FG_SECONDARY).pack(side="left", padx=5)

        # Contenedor scrolleable para lista de fallas individuales
        self.fallas_list_frame = tk.Frame(fallas_frame, bg=EstiloUI.BG_DARK)
        self.fallas_list_frame.pack(fill="both", expand=True, padx=5, pady=(3, 2))

        self.fallas_canvas = tk.Canvas(self.fallas_list_frame, bg=EstiloUI.BG_DARK,
                                        highlightthickness=0, height=150)
        fallas_scroll = ttk.Scrollbar(self.fallas_list_frame, orient="vertical",
                                       command=self.fallas_canvas.yview)
        self.fallas_inner = tk.Frame(self.fallas_canvas, bg=EstiloUI.BG_DARK)
        self.fallas_inner.bind("<Configure>",
            lambda e: self.fallas_canvas.configure(scrollregion=self.fallas_canvas.bbox("all")))
        self.fallas_canvas.create_window((0, 0), window=self.fallas_inner, anchor="nw")
        self.fallas_canvas.configure(yscrollcommand=fallas_scroll.set)
        self.fallas_canvas.pack(side="left", fill="both", expand=True)
        fallas_scroll.pack(side="right", fill="y")

        # Scroll con rueda del mouse
        def _fl_enter(e):
            self.fallas_canvas.bind_all("<MouseWheel>",
                lambda ev: self.fallas_canvas.yview_scroll(int(-1*(ev.delta/120)), "units"))
        def _fl_leave(e):
            self.fallas_canvas.unbind_all("<MouseWheel>")
        self.fallas_canvas.bind("<Enter>", _fl_enter)
        self.fallas_canvas.bind("<Leave>", _fl_leave)

        tk.Frame(fallas_frame, bg=EstiloUI.BG_CARD, height=3).pack()

        # resultado_labels dict (usado por _actualizar_conteos y _on_procesado)
        self.resultado_labels = {}

        # --- SEVERIDAD ---
        sev_frame = tk.Frame(parent, bg=EstiloUI.BG_CARD)
        sev_frame.grid(row=2, column=0, sticky="ew", padx=5, pady=2)

        tk.Label(sev_frame, text="SEVERIDAD", font=EstiloUI.FONT_SUBTITLE,
                 bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_ACCENT).pack(padx=10, pady=(6, 3), anchor="w")

        self.severidad_labels = {}
        for sev, color in [("L (Baja)", "#00cc44"), ("M (Media)", "#ff9900"), ("H (Alta)", "#ff2222")]:
            row = tk.Frame(sev_frame, bg=EstiloUI.BG_CARD)
            row.pack(fill="x", padx=10, pady=1)
            tk.Label(row, text=sev, font=EstiloUI.FONT_BODY,
                     bg=EstiloUI.BG_CARD, fg=color).pack(side="left")
            lbl = tk.Label(row, text="0", font=EstiloUI.FONT_LABEL,
                           bg=EstiloUI.BG_CARD, fg=color)
            lbl.pack(side="right")
            self.severidad_labels[sev[0]] = lbl

        tk.Frame(sev_frame, bg=EstiloUI.BG_CARD, height=3).pack()

        # --- Log ---
        log_frame = tk.Frame(parent, bg=EstiloUI.BG_PANEL)
        log_frame.grid(row=3, column=0, sticky="nsew", padx=5, pady=(0, 5))
        log_frame.grid_rowconfigure(1, weight=1)
        log_frame.grid_columnconfigure(0, weight=1)

        tk.Label(log_frame, text="LOG", font=EstiloUI.FONT_SUBTITLE,
                 bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_ACCENT).grid(
            row=0, column=0, sticky="w", padx=5, pady=(5, 2))

        self.log_text = scrolledtext.ScrolledText(
            log_frame,
            wrap=tk.WORD,
            font=EstiloUI.FONT_MONO,
            bg=EstiloUI.BG_INPUT,
            fg=EstiloUI.FG_LOG,
            insertbackground=EstiloUI.FG_LOG,
            relief="flat",
            height=15,
        )
        self.log_text.grid(row=1, column=0, sticky="nsew", padx=2, pady=2)

        ttk.Button(log_frame, text="Limpiar", style="Secondary.TButton",
                   command=lambda: self.log_text.delete(1.0, tk.END)).grid(
            row=2, column=0, sticky="e", padx=5, pady=3)

    def _crear_barra_estado(self):
        barra = tk.Frame(self, bg=EstiloUI.BG_CARD, height=42)
        barra.grid(row=2, column=0, sticky="ew")
        barra.grid_propagate(False)

        # Fila superior: estado + modelo + tiempo
        fila_sup = tk.Frame(barra, bg=EstiloUI.BG_CARD)
        fila_sup.pack(fill="x", padx=5, pady=(2, 0))

        self.lbl_estado = tk.Label(fila_sup, text="Listo - Cargue modelo e imagenes",
                                    font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_CARD,
                                    fg=EstiloUI.FG_SECONDARY)
        self.lbl_estado.pack(side="left", padx=5)

        self.lbl_tiempo = tk.Label(fila_sup, text="",
                                    font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_CARD,
                                    fg=EstiloUI.FG_HIGHLIGHT)
        self.lbl_tiempo.pack(side="left", padx=15)

        self.lbl_modelo = tk.Label(fila_sup, text="Modelo: No cargado",
                                    font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_CARD,
                                    fg=EstiloUI.FG_ACCENT)
        self.lbl_modelo.pack(side="right", padx=5)

        # Fila inferior: creditos bachilleres (centrado entre estado y modelo)
        tk.Label(fila_sup,
                 text="Bach. Miguel Bernardino Quispe Arias  |  Bach. Briza Edith Catachura Aycaya",
                 font=("Segoe UI", 8, "italic"), bg=EstiloUI.BG_CARD,
                 fg=EstiloUI.FG_SECONDARY).pack(side="left", expand=True)

    # =========================================================================
    # CRONOMETRO EN VIVO
    # =========================================================================

    def _iniciar_cronometro(self):
        """Inicia el cronometro en vivo en la barra superior."""
        self._tiempo_inicio = time.time()
        self._crono_elapsed = 0.0
        self._actualizar_cronometro()

    def _formato_tiempo(self, elapsed):
        """Formatea segundos a texto legible."""
        mins = int(elapsed // 60)
        secs = int(elapsed % 60)
        dec = int((elapsed % 1) * 10)
        if mins > 0:
            return f"{mins}:{secs:02d}.{dec}"
        else:
            return f"{secs}.{dec}s"

    def _actualizar_cronometro(self):
        """Actualiza el cronometro cada 100ms."""
        if self._tiempo_inicio and self.procesando:
            self._crono_elapsed = time.time() - self._tiempo_inicio
            self.lbl_cronometro.config(text=self._formato_tiempo(self._crono_elapsed))
            self._crono_after_id = self.after(100, self._actualizar_cronometro)

    def _detener_cronometro(self):
        """Detiene el cronometro y devuelve el tiempo final."""
        if self._crono_after_id:
            self.after_cancel(self._crono_after_id)
            self._crono_after_id = None
        if self._tiempo_inicio:
            self._crono_elapsed = time.time() - self._tiempo_inicio
            self.lbl_cronometro.config(text=self._formato_tiempo(self._crono_elapsed))
        return self._crono_elapsed

    def _on_error_procesamiento(self, contexto, error, detalle=None):
        """Restaura la UI si el hilo de procesamiento falla."""
        self.procesando = False
        t_elapsed = self._detener_cronometro()
        self._detener_flag = False
        if hasattr(self, 'btn_detener'):
            self.btn_detener.config(state="disabled")
        self._estado(contexto)
        if hasattr(self, 'lbl_tiempo'):
            self.lbl_tiempo.config(text=self._formato_tiempo(t_elapsed))
        self._log(f"{contexto}: {error}")
        if detalle:
            self._log(detalle.rstrip())

    # =========================================================================
    # WIDGETS AUXILIARES
    # =========================================================================

    def _seccion_titulo(self, parent, texto):
        tk.Frame(parent, bg=EstiloUI.BG_ACCENT, height=2).pack(fill="x", padx=10, pady=(10, 0))
        tk.Label(parent, text=texto, font=EstiloUI.FONT_LABEL,
                 bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_ACCENT).pack(anchor="w", padx=10, pady=(2, 4))

    def _crear_slider(self, parent, key, label, from_, to, default, step=1, resolution=None):
        frame = tk.Frame(parent, bg=EstiloUI.BG_PANEL)
        frame.pack(fill="x", padx=10, pady=1)

        top = tk.Frame(frame, bg=EstiloUI.BG_PANEL)
        top.pack(fill="x")
        tk.Label(top, text=label, font=EstiloUI.FONT_SMALL,
                 bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_SECONDARY).pack(side="left")

        # Usar valor: config individual de imagen > config global base > motor > default
        valor_inicial = self._resolver_valor_config(key, default)
        var = tk.DoubleVar(value=valor_inicial)
        self.vars[key] = var

        # Mostrar el valor inicial real (no el default hardcodeado)
        texto_inicial = f"{valor_inicial:.2f}" if (resolution and resolution < 1) else str(int(valor_inicial))
        val_label = tk.Label(top, text=texto_inicial, font=EstiloUI.FONT_LABEL,
                              bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT, width=6, anchor="e")
        val_label.pack(side="right")

        res = resolution if resolution else step
        slider = tk.Scale(frame, from_=from_, to=to, orient="horizontal",
                          variable=var, resolution=res, showvalue=False,
                          bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY,
                          troughcolor=EstiloUI.BG_DARK, highlightthickness=0,
                          activebackground=EstiloUI.BG_BUTTON,
                          length=200, sliderlength=15)
        slider.pack(fill="x")

        def update_label(*args):
            v = var.get()
            val_label.config(text=f"{v:.2f}" if (resolution and resolution < 1) else str(int(v)))
        var.trace_add("write", update_label)

    def _sync_params(self):
        """Sincroniza sliders con el motor."""
        float_params = {
            'confianza_min', 'iou_threshold', 'clahe_clip', 'min_circularidad',
            'min_diametro_hueco_mm', 'min_longitud_grieta_m',
            'min_area_parche_m2', 'min_area_piel_m2',
            'merge_iou_threshold', 'text_size',
        }
        string_params = {'tramo', 'progresiva_inicio'}
        skip_params = {'modo_calibracion'}  # handled separately

        for key, var in self.vars.items():
            if key in skip_params:
                continue
            if isinstance(var, tk.BooleanVar):
                self.motor.config[key] = var.get()
            elif isinstance(var, tk.StringVar):
                if key == 'profundidad_asumida_huecos':
                    val_str = var.get()
                    if 'baja' in val_str:
                        self.motor.config[key] = 'baja'
                    elif 'alta' in val_str:
                        self.motor.config[key] = 'alta'
                    else:
                        self.motor.config[key] = 'media'
                elif key in string_params:
                    self.motor.config[key] = var.get()
            elif isinstance(var, tk.IntVar):
                if key == 'progresiva_incremento':
                    self.motor.config[key] = var.get()
            elif key == 'ancho_via':
                self.motor.config['ancho_via_real_m'] = float(var.get())
                self.motor.calibrador.ancho_via_real_m = float(var.get())
            elif key in self.motor.config:
                val = var.get()
                if key in float_params:
                    self.motor.config[key] = float(val)
                else:
                    self.motor.config[key] = int(val)

        self.motor.procesador_piel = ProcesadorPielCocodrilo(self.motor.config)

        # Calibracion mode
        modo = self.vars.get('modo_calibracion', tk.StringVar(value='automatica')).get()
        self.calibrar_cada_imagen = (modo == 'cada_imagen')

    # Claves de configuracion avanzada (las que se guardan por imagen)
    _CLAVES_AVANZADAS = [
        'min_diametro_hueco_mm', 'profundidad_asumida_huecos',
        'min_longitud_grieta_m', 'min_area_parche_m2', 'min_area_piel_m2',
        'merge_iou_threshold', 'merge_distancia_max_px',
        'clahe_clip', 'clahe_tile', 'bilateral_d',
        'bilateral_sigma_color', 'bilateral_sigma_space',
        'block_size', 'C_umbral', 'usar_frangi', 'kernel_apertura',
        'kernel_cierre', 'iteraciones_cierre', 'usar_multiescala',
        'min_area_poligono', 'min_circularidad', 'min_vertices',
        'max_vertices', 'min_radio_circulo', 'min_longitud_rama',
        'min_area_objeto', 'max_gap_cierre',
    ]

    def _resolver_valor_config(self, key, default=None):
        """Resuelve el valor de una clave: individual de imagen > global base > motor > default."""
        idx = self.imagen_actual_idx
        if idx in self._config_por_imagen and key in self._config_por_imagen[idx]:
            return self._config_por_imagen[idx][key]
        if self._config_global_base and key in self._config_global_base:
            return self._config_global_base[key]
        if key in self.motor.config:
            return self.motor.config[key]
        return default

    def _snapshot_config_avanzada(self):
        """Devuelve una copia de los valores avanzados actuales del motor."""
        return {k: self.motor.config[k] for k in self._CLAVES_AVANZADAS if k in self.motor.config}

    def _guardar_config_imagen(self, idx=None):
        """Guarda la config avanzada actual SOLO para la imagen indicada."""
        if idx is None:
            idx = self.imagen_actual_idx
        if idx < 0:
            return
        self._config_por_imagen[idx] = self._snapshot_config_avanzada()

    def _guardar_config_global(self):
        """Guarda la config avanzada actual como base global (para todas las imagenes)."""
        self._config_global_base = self._snapshot_config_avanzada()

    def _restaurar_config_imagen(self, idx):
        """Restaura la config guardada para una imagen al motor.
        Si la imagen tiene config individual, usa esa. Si no, usa la global base."""
        if idx in self._config_por_imagen:
            src = self._config_por_imagen[idx]
        elif self._config_global_base:
            src = self._config_global_base
        else:
            return
        for k, v in src.items():
            self.motor.config[k] = v
        self.motor.procesador_piel = ProcesadorPielCocodrilo(self.motor.config)

    def _aplicar_config_individual_y_cerrar(self, win):
        """Guarda config avanzada SOLO para la imagen actual, sin reprocesar."""
        # Guardar la config global base antes de modificar el motor
        if not self._config_global_base:
            self._config_global_base = self._snapshot_config_avanzada()
        config_global_previa = self._snapshot_config_avanzada() if not self._config_global_base else dict(self._config_global_base)

        # Sincronizar sliders -> motor (temporalmente para esta imagen)
        self._sync_params()
        # Guardar snapshot individual para esta imagen
        self._guardar_config_imagen()

        # Restaurar la config global base al motor para que otras imagenes no se contaminen
        for k, v in config_global_previa.items():
            self.motor.config[k] = v
        self.motor.procesador_piel = ProcesadorPielCocodrilo(self.motor.config)

        self._log("Configuracion guardada para imagen actual. Ejecute 'Procesar' para aplicar.")
        win.destroy()

    def _aplicar_config_todas_y_cerrar(self, win):
        """Guarda config avanzada para TODAS las imagenes, sin reprocesar."""
        self._sync_params()
        # Esta es la nueva base global
        self._guardar_config_global()
        # Limpiar configs individuales: ahora todas usan la misma
        self._config_por_imagen.clear()
        self._log("Configuracion guardada para todas las imagenes. Ejecute 'Procesar' para aplicar.")
        win.destroy()

    def _auto_calibrar_piel(self):
        """Abre la ventana de auto-calibracion de zona para Piel de Cocodrilo."""
        if not self.imagenes_cargadas:
            messagebox.showinfo("Info", "Cargue imagenes primero.")
            return

        ruta = self.imagenes_cargadas[self.imagen_actual_idx]
        imagen = cv2.imread(str(ruta))
        if imagen is None:
            messagebox.showerror("Error", f"No se pudo cargar: {ruta}")
            return

        # La nueva VentanaSeleccionROI (renombrada a VentanaAutoCalibrarPiel aquÃ­)
        # No necesita params_actuales, sÃ³lo (parent, imagen_cv2)
        win = VentanaAutoCalibrarPiel(self, imagen)

        # El resultado se queda en win.resultado (que es el dict `_params_calculados`)
        if hasattr(win, 'resultado') and win.resultado:
            self._aplicar_perfil(win.resultado)
            self._log(f"Auto-calibracion Piel de Cocodrilo aplicada: {win.resultado}")
        else:
            self._log("Auto-calibracion cancelada o sin resultados.")

    # =========================================================================
    # PERFILES RAPIDOS
    # =========================================================================

    def _aplicar_perfil(self, valores):
        """Aplica un conjunto de valores a los sliders."""
        for key, val in valores.items():
            if key in self.vars:
                self.vars[key].set(val)

    # --- Perfiles Piel de Cocodrilo ---

    def _perfil_piel_estandar(self):
        self._aplicar_perfil({
            'clahe_clip': 3.0, 'clahe_tile': 8,
            'bilateral_d': 9, 'bilateral_sigma_color': 75, 'bilateral_sigma_space': 75,
            'block_size': 25, 'C_umbral': 12,
            'kernel_apertura': 3, 'kernel_cierre': 5, 'iteraciones_cierre': 2,
            'usar_frangi': False, 'usar_multiescala': False,
            'min_area_poligono': 300, 'min_circularidad': 0.08,
            'min_vertices': 4, 'max_vertices': 25,
            'min_radio_circulo': 8, 'min_longitud_rama': 30,
            'min_area_objeto': 100, 'max_gap_cierre': 20,
        })
        self._log("Perfil ESTANDAR (Piel Cocodrilo) aplicado")

    def _perfil_piel_agresivo(self):
        self._aplicar_perfil({
            'clahe_clip': 4.5, 'clahe_tile': 6,
            'bilateral_d': 7, 'bilateral_sigma_color': 60, 'bilateral_sigma_space': 60,
            'block_size': 19, 'C_umbral': 8,
            'kernel_apertura': 2, 'kernel_cierre': 7, 'iteraciones_cierre': 3,
            'usar_frangi': True, 'usar_multiescala': True,
            'min_area_poligono': 150, 'min_circularidad': 0.05,
            'min_vertices': 3, 'max_vertices': 35,
            'min_radio_circulo': 5, 'min_longitud_rama': 15,
            'min_area_objeto': 50, 'max_gap_cierre': 30,
        })
        self._log("Perfil AGRESIVO (Piel Cocodrilo) aplicado - Detecta mallas finas")

    def _perfil_piel_suave(self):
        self._aplicar_perfil({
            'clahe_clip': 2.5, 'clahe_tile': 10,
            'bilateral_d': 11, 'bilateral_sigma_color': 90, 'bilateral_sigma_space': 90,
            'block_size': 31, 'C_umbral': 15,
            'kernel_apertura': 4, 'kernel_cierre': 4, 'iteraciones_cierre': 1,
            'usar_frangi': False, 'usar_multiescala': False,
            'min_area_poligono': 500, 'min_circularidad': 0.12,
            'min_vertices': 5, 'max_vertices': 20,
            'min_radio_circulo': 12, 'min_longitud_rama': 50,
            'min_area_objeto': 200, 'max_gap_cierre': 15,
        })
        self._log("Perfil SUAVE (Piel Cocodrilo) aplicado - Solo mallas grandes y definidas")

    # --- Perfiles Deteccion General (Baches, Grietas, Parches) ---

    def _perfil_alta_sensibilidad(self):
        self._aplicar_perfil({
            'confianza_min': 0.05,
            'iou_threshold': 0.30,
        })
        self._log("Perfil ALTA SENSIBILIDAD aplicado - Confianza baja, detecta mas fallas (puede incluir falsos positivos)")

    def _perfil_balanceado(self):
        self._aplicar_perfil({
            'confianza_min': 0.10,
            'iou_threshold': 0.45,
        })
        self._log("Perfil BALANCEADO aplicado - Configuracion por defecto equilibrada")

    def _perfil_alta_precision(self):
        self._aplicar_perfil({
            'confianza_min': 0.30,
            'iou_threshold': 0.60,
        })
        self._log("Perfil ALTA PRECISION aplicado - Confianza alta, solo fallas claras (menos falsos positivos)")

    # =========================================================================
    # CONFIG AVANZADA (MODAL)
    # =========================================================================

    def _abrir_config_avanzada(self, seccion=None):
        """Abre ventana modal con parametros avanzados. Si seccion se especifica, solo muestra esa seccion."""
        titulos_seccion = {
            'filtros': 'Filtros TamaÃ±o Minimo',
            'fusion': 'Fusion Solapamientos',
            'piel': 'Piel de Cocodrilo',
            'poligonos': 'Deteccion Poligonos',
            'perfiles': 'Perfiles Rapidos',
        }
        titulo = titulos_seccion.get(seccion, "Configuracion Avanzada")

        win = tk.Toplevel(self)
        win.title(titulo)
        win.configure(bg=EstiloUI.BG_DARK)
        win.transient(self)
        win.grab_set()
        win.bind("<Escape>", lambda e: win.destroy())

        canvas = tk.Canvas(win, bg=EstiloUI.BG_PANEL, highlightthickness=0)
        vsb = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
        hsb = ttk.Scrollbar(win, orient="horizontal", command=canvas.xview)
        sf = tk.Frame(canvas, bg=EstiloUI.BG_PANEL)
        sf.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=sf, anchor="nw")
        canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        win.grid_rowconfigure(0, weight=1)
        win.grid_columnconfigure(0, weight=1)

        def _mw(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind("<MouseWheel>", _mw)

        # --- FILTROS TAMANO MINIMO ---
        if seccion in (None, 'filtros'):
            self._seccion_titulo(sf, "FILTROS TAMANO MINIMO")
            self._crear_slider(sf, 'min_diametro_hueco_mm',
                               "Min Diam. Hueco (mm)", 0, 200, 50, step=10)

            prof_frame = tk.Frame(sf, bg=EstiloUI.BG_PANEL)
            prof_frame.pack(fill="x", padx=10, pady=2)
            tk.Label(prof_frame, text="Prof. Asumida Huecos:",
                     font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                     fg=EstiloUI.FG_SECONDARY).pack(side="left")
            # Restaurar valor: individual > global base > motor > default
            p = self._resolver_valor_config('profundidad_asumida_huecos', 'media')
            prof_val = 'media (25-50mm)'
            if 'baja' in str(p): prof_val = 'baja (13-25mm)'
            elif 'alta' in str(p): prof_val = 'alta (>50mm)'
            self.vars['profundidad_asumida_huecos'] = tk.StringVar(value=prof_val)
            prof_combo = ttk.Combobox(prof_frame,
                                       textvariable=self.vars['profundidad_asumida_huecos'],
                                       values=['baja (13-25mm)', 'media (25-50mm)', 'alta (>50mm)'],
                                       state='readonly', width=16)
            prof_combo.pack(side="right")
            prof_combo.set(prof_val)

            self._crear_slider(sf, 'min_longitud_grieta_m',
                               "Min Long. Grieta (m)", 0.0, 1.0, 0.05, resolution=0.01)
            self._crear_slider(sf, 'min_area_parche_m2',
                               "Min Area Parche (m2)", 0.0, 1.0, 0.01, resolution=0.005)
            self._crear_slider(sf, 'min_area_piel_m2',
                               "Min Area P.Cocod (m2)", 0.0, 2.0, 0.05, resolution=0.01)

        # --- FUSION SOLAPAMIENTOS ---
        if seccion in (None, 'fusion'):
            self._seccion_titulo(sf, "FUSION SOLAPAMIENTOS")
            self._crear_slider(sf, 'merge_iou_threshold',
                               "IoU Fusion", 0.0, 0.5, 0.10, resolution=0.02)
            self._crear_slider(sf, 'merge_distancia_max_px',
                               "Dist. Max Fusion (px)", 0, 200, 50, step=10)

        # --- PIEL DE COCODRILO ---
        if seccion in (None, 'piel'):
            self._seccion_titulo(sf, "PIEL DE COCODRILO")
            self._crear_slider(sf, 'clahe_clip', "CLAHE Clip", 1.0, 8.0, 4.0, resolution=0.5)
            self._crear_slider(sf, 'clahe_tile', "CLAHE Tile", 4, 16, 8)
            self._crear_slider(sf, 'bilateral_d', "Bilateral D", 3, 15, 9)
            self._crear_slider(sf, 'bilateral_sigma_color', "Sigma Color", 20, 150, 75)
            self._crear_slider(sf, 'bilateral_sigma_space', "Sigma Espacio", 20, 150, 75)
            self._crear_slider(sf, 'block_size', "Block Size", 11, 51, 23, step=2)
            self._crear_slider(sf, 'C_umbral', "Constante C", 3, 25, 10)

            frangi_val = self._resolver_valor_config('usar_frangi', True)
            self.vars['usar_frangi'] = tk.BooleanVar(value=frangi_val)
            tk.Checkbutton(sf, text="Filtro Frangi",
                            variable=self.vars['usar_frangi'],
                            bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY,
                            selectcolor=EstiloUI.BG_DARK, font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)

            ref_val = self._resolver_valor_config('usar_refinamiento', True)
            self.vars['usar_refinamiento'] = tk.BooleanVar(value=ref_val)
            tk.Checkbutton(sf, text="Refinar esqueleto y cerrar gaps",
                            variable=self.vars['usar_refinamiento'],
                            bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY,
                            selectcolor=EstiloUI.BG_DARK, font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)

            multi_val = self._resolver_valor_config('usar_multiescala', True)
            self.vars['usar_multiescala'] = tk.BooleanVar(value=multi_val)
            tk.Checkbutton(sf, text="Multi-escala",
                            variable=self.vars['usar_multiescala'],
                            bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY,
                            selectcolor=EstiloUI.BG_DARK, font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)

            self._crear_slider(sf, 'kernel_apertura', "Kernel Apertura", 2, 7, 3)
            self._crear_slider(sf, 'kernel_cierre', "Kernel Cierre", 3, 12, 6)
            self._crear_slider(sf, 'iteraciones_cierre', "Iter. Cierre", 1, 5, 2)

            # Boton auto-calibrar zona
            tk.Frame(sf, bg=EstiloUI.BG_ACCENT, height=1).pack(fill="x", padx=10, pady=(8, 4))
            ttk.Button(sf, text="Auto-calibrar zona (seleccionar grietas)",
                       style="Accent.TButton",
                       command=self._auto_calibrar_piel).pack(fill="x", padx=10, pady=3)

        # --- DETECCION POLIGONOS ---
        if seccion in (None, 'poligonos'):
            self._seccion_titulo(sf, "DETECCION POLIGONOS")
            self._crear_slider(sf, 'min_area_poligono', "Area Min Poligono", 50, 1000, 300, step=50)
            self._crear_slider(sf, 'min_circularidad', "Circularidad Min", 0.01, 0.5, 0.08, resolution=0.01)
            self._crear_slider(sf, 'min_vertices', "Vertices Min", 3, 8, 4)
            self._crear_slider(sf, 'max_vertices', "Vertices Max", 10, 50, 25)
            self._crear_slider(sf, 'min_radio_circulo', "Radio Min Circulo", 3, 30, 8)
            self._crear_slider(sf, 'min_longitud_rama', "Long Min Rama", 10, 100, 30, step=5)
            self._crear_slider(sf, 'min_area_objeto', "Area Min Objeto", 20, 500, 100, step=10)
            self._crear_slider(sf, 'max_gap_cierre', "Max Gap Cierre", 5, 50, 20)

        # --- PERFILES RAPIDOS ---
        if seccion in (None, 'perfiles'):
            self._seccion_titulo(sf, "PERFILES RAPIDOS")
            pf = tk.Frame(sf, bg=EstiloUI.BG_PANEL)
            pf.pack(fill="x", padx=10, pady=5)
            for nombre, cmd in [
                ("Estandar", self._perfil_piel_estandar),
                ("Agresivo", self._perfil_piel_agresivo),
                ("Suave", self._perfil_piel_suave),
            ]:
                ttk.Button(pf, text=nombre, style="Secondary.TButton",
                           command=cmd).pack(fill="x", pady=1)

        # --- BOTONES APLICAR ---
        tk.Frame(sf, bg=EstiloUI.BG_ACCENT, height=2).pack(fill="x", padx=10, pady=(15, 5))

        btn_f = tk.Frame(sf, bg=EstiloUI.BG_PANEL)
        btn_f.pack(fill="x", padx=10, pady=5)
        ttk.Button(btn_f, text="Guardar para Imagen Actual y Cerrar", style="Accent.TButton",
                   command=lambda: self._aplicar_config_individual_y_cerrar(win)).pack(fill="x", pady=2)
        ttk.Button(btn_f, text="Guardar para Todas las Imagenes y Cerrar", style="Accent.TButton",
                   command=lambda: self._aplicar_config_todas_y_cerrar(win)).pack(fill="x", pady=2)

        # Auto-ajustar tamaÃ±o de la ventana al contenido
        win.update_idletasks()
        req_w = sf.winfo_reqwidth() + vsb.winfo_reqwidth() + 20
        req_h = sf.winfo_reqheight() + hsb.winfo_reqheight() + 20
        max_h = int(self.winfo_screenheight() * 0.8)
        max_w = int(self.winfo_screenwidth() * 0.8)
        w = min(req_w, max_w)
        h = min(req_h, max_h)
        x = self.winfo_x() + (self.winfo_width() - w) // 2
        y = self.winfo_y() + (self.winfo_height() - h) // 2
        win.geometry(f"{w}x{h}+{x}+{y}")

    # =========================================================================
    # EXCLUIR / INCLUIR FALLAS POR CLICK
    # =========================================================================

    def _on_click_excluir_falla(self, event):
        """Click izquierdo en vista resultado: excluir/incluir la falla bajo el cursor."""
        if self.vista_actual != "resultado" or not self.resultado_actual:
            return
        if not self.resultado_actual.get('fallas'):
            return

        # Coordenadas imagen real
        if self._cv2_original_full is None:
            return
        img_x = int((event.x - self._pan_offset_x) / self._zoom_level)
        img_y = int((event.y - self._pan_offset_y) / self._zoom_level)
        h, w = self._cv2_original_full.shape[:2]
        if img_x < 0 or img_y < 0 or img_x >= w or img_y >= h:
            return

        # Buscar falla bajo el cursor
        falla_encontrada = None
        for f in self.resultado_actual['fallas']:
            # Verificar si el punto esta dentro del contorno, mascara, o mascara_roi (piel cocodrilo)
            if 'contorno' in f:
                dist = cv2.pointPolygonTest(f['contorno'], (img_x, img_y), False)
                if dist >= 0:
                    falla_encontrada = f
                    break
            if 'mascara_roi' in f:
                mroi = f['mascara_roi']
                if 0 <= img_y < mroi.shape[0] and 0 <= img_x < mroi.shape[1]:
                    if mroi[img_y, img_x] > 0:
                        falla_encontrada = f
                        break
            if 'mask' in f:
                mask = f['mask']
                if 0 <= img_y < mask.shape[0] and 0 <= img_x < mask.shape[1]:
                    if mask[img_y, img_x] > 0:
                        falla_encontrada = f
                        break

        if falla_encontrada:
            # Toggle excluida
            excluida = falla_encontrada.get('excluida', False)
            falla_encontrada['excluida'] = not excluida
            estado = "EXCLUIDA" if not excluida else "INCLUIDA"
            self._log(f"Falla {falla_encontrada['tipo']} #{falla_encontrada['id']} -> {estado}")

            # Redibujar resultado
            vis = self.motor.dibujar_resultado(
                self.resultado_actual['imagen'],
                self.resultado_actual['fallas'],
                progresiva=self._progresiva_actual())
            self._mostrar_cv2(vis, reset_zoom=False)
            self._actualizar_conteos(self.resultado_actual['fallas'])

    # =========================================================================
    # DETENER PROCESAMIENTO
    # =========================================================================

    def _detener_procesamiento(self):
        """Detiene el procesamiento en curso."""
        if self.procesando:
            self._detener_flag = True
            self._log("Deteniendo procesamiento...")
            self._estado("Deteniendo...")

    # =========================================================================
    # REDIBUJAR / REPROCESAR / CALIBRAR UNICA
    # =========================================================================

    def _redibujar_resultado_actual(self):
        """Redibuja la imagen resultado con las opciones de visualizacion actuales."""
        if not self.resultado_actual or not self.resultado_actual.get('fallas'):
            return
        if self.vista_actual != "resultado":
            return
        self._sync_params()
        vis = self.motor.dibujar_resultado(
            self.resultado_actual['imagen'],
            self.resultado_actual['fallas'],
            progresiva=self._progresiva_actual())
        self._mostrar_cv2(vis, reset_zoom=False)

    def _on_text_size_change(self, val):
        """Actualiza el label del valor del slider de texto."""
        self.lbl_text_size_val.config(text=f"{float(val):.2f}")

    def _redibujar_resultado(self):
        """Redibuja el resultado con el tamaÃ±o de texto del slider."""
        nuevo_ts = self.text_size_slider.get()
        self.vars['text_size'].set(nuevo_ts)
        self._redibujar_resultado_actual()

    def _reprocesar_actual(self):
        """Reprocesa la imagen actual con los nuevos parametros avanzados."""
        if self.imagenes_cargadas and self.motor.modelo_cargado and not self.procesando:
            self._procesar_actual()

    def _calibrar_unica_ahora(self):
        """Abre calibracion Tk para una imagen y aplica a todas."""
        if not self.imagenes_cargadas:
            messagebox.showinfo("Info", "Cargue imagenes primero.")
            return
        ruta = self.imagenes_cargadas[self.imagen_actual_idx]
        imagen = cv2.imread(str(ruta))
        if imagen is None:
            messagebox.showerror("Error", f"No se pudo cargar: {ruta}")
            return

        ancho_via = float(self.vars.get('ancho_via', tk.DoubleVar(value=6.5)).get())
        # VentanaCalibracionTk.__init__ ya llama wait_window() internamente
        win = VentanaCalibracionTk(self, imagen, ancho_via_real_m=ancho_via)

        if win.completado and win.resultado:
            longitud_px = win.resultado
            self.motor.calibrador.calibrar_con_linea(longitud_px)
            px_por_mm = self.motor.calibrador.px_por_mm
            self._log(f"Calibracion unica aplicada: {px_por_mm:.4f} px/mm (linea={longitud_px:.1f}px)")
            self._log("Se usara esta calibracion para TODAS las imagenes.")
            # Forzar modo unica
            self.vars['modo_calibracion'].set('unica')
        else:
            self._log("Calibracion cancelada.")

    def _actualizar_lista_fallas(self):
        """Actualiza la lista de checkboxes de fallas individuales en el panel derecho."""
        # Limpiar lista anterior
        for w in self.fallas_inner.winfo_children():
            w.destroy()

        if not self.resultado_actual or not self.resultado_actual.get('fallas'):
            tk.Label(self.fallas_inner, text="Sin fallas", font=EstiloUI.FONT_SMALL,
                     bg=EstiloUI.BG_DARK, fg=EstiloUI.FG_SECONDARY).pack(anchor="w", padx=5)
            return

        fallas = self.resultado_actual['fallas']
        self._falla_vars = []

        for i, f in enumerate(fallas):
            var = tk.BooleanVar(value=not f.get('excluida', False))
            self._falla_vars.append((var, f))

            tipo = f['tipo']
            sev = severidad_ui(f.get('severidad'))
            categoria = categorizar_tipo_falla(tipo)
            color_map = {'L': '#00cc44', 'M': '#ff9900', 'H': '#ff2222'}
            color = color_map.get(sev, EstiloUI.FG_PRIMARY)

            # Nombre corto
            if categoria == 'hueco':
                txt = f"HUECO #{f['id']} D={f.get('diametro_mm', 0):.0f}mm [{sev}]"
            elif categoria == 'grieta':
                txt = f"GRIETA L/T #{f['id']} e={f.get('espesor_mm', 0):.1f}mm L={f.get('longitud_m', 0):.2f}m [{sev}]"
            elif categoria == 'parche':
                txt = f"PARCHE #{f['id']} A={f.get('area_m2', 0):.2f}m2 [{sev}]"
            elif categoria == 'piel':
                txt = f"P.COCOD #{f['id']} Dmin={f.get('diametro_min_mm', 0):.1f}mm A={f.get('area_m2', 0):.2f}m2 [{sev}]"
            else:
                txt = f"{tipo[:15]} #{f['id']} [{sev}]"

            row = tk.Frame(self.fallas_inner, bg=EstiloUI.BG_DARK)
            row.pack(fill="x", padx=2, pady=0)

            cb = tk.Checkbutton(row, text=txt, variable=var,
                                bg=EstiloUI.BG_DARK, fg=color,
                                selectcolor=EstiloUI.BG_DARK,
                                activebackground=EstiloUI.BG_DARK,
                                font=("Consolas", 8), anchor="w", justify="left",
                                command=lambda idx=i: self._toggle_falla_check(idx))
            cb.pack(side="left", fill="x", expand=True)

            if categoria == 'piel':
                btn_cfg = tk.Button(
                    row, text="\u2699",
                    font=("Segoe UI", 9, "bold"),
                    bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT,
                    activebackground=EstiloUI.BG_ACCENT,
                    activeforeground=EstiloUI.FG_PRIMARY,
                    relief="groove", bd=1, padx=6,
                    command=lambda idx=i: self._configurar_piel_falla(idx)
                )
                btn_cfg.pack(side="right", padx=(4, 0))

    def _toggle_falla_check(self, idx):
        """Toggle excluir/incluir falla desde checkbox en panel derecho."""
        if not self.resultado_actual or not self.resultado_actual.get('fallas'):
            return
        if idx >= len(self._falla_vars):
            return
        var, falla = self._falla_vars[idx]
        falla['excluida'] = not var.get()
        estado = "INCLUIDA" if var.get() else "EXCLUIDA"
        self._log(f"Falla {falla['tipo']} #{falla['id']} -> {estado}")

        # Redibujar y actualizar conteos + severidad
        self._sync_params()
        vis = self.motor.dibujar_resultado(
            self.resultado_actual['imagen'],
            self.resultado_actual['fallas'],
            progresiva=self._progresiva_actual())
        self._mostrar_cv2(vis, reset_zoom=False)
        self._actualizar_conteos(self.resultado_actual['fallas'])

    # =========================================================================
    # SELECCIONAR / DESELECCIONAR TODAS LAS FALLAS
    # =========================================================================

    def _seleccionar_todas_fallas(self):
        """Marca todas las fallas como incluidas."""
        if not self.resultado_actual or not self.resultado_actual.get('fallas'):
            return
        for f in self.resultado_actual['fallas']:
            f['excluida'] = False
        self._log("Todas las fallas INCLUIDAS")
        self._sync_params()
        vis = self.motor.dibujar_resultado(
            self.resultado_actual['imagen'],
            self.resultado_actual['fallas'],
            progresiva=self._progresiva_actual())
        self._mostrar_cv2(vis, reset_zoom=False)
        self._actualizar_conteos(self.resultado_actual['fallas'])

    def _deseleccionar_todas_fallas(self):
        """Marca todas las fallas como excluidas."""
        if not self.resultado_actual or not self.resultado_actual.get('fallas'):
            return
        for f in self.resultado_actual['fallas']:
            f['excluida'] = True
        self._log("Todas las fallas EXCLUIDAS")
        self._sync_params()
        vis = self.motor.dibujar_resultado(
            self.resultado_actual['imagen'],
            self.resultado_actual['fallas'],
            progresiva=self._progresiva_actual())
        self._mostrar_cv2(vis, reset_zoom=False)
        self._actualizar_conteos(self.resultado_actual['fallas'])

    # =========================================================================
    # CONFIGURACION INDIVIDUAL - PIEL DE COCODRILO
    # =========================================================================

    def _configurar_piel_falla(self, idx):
        """Abre una ventana para ajustar y reprocesar una deteccion de piel de cocodrilo."""
        if not self.resultado_actual or 'fallas' not in self.resultado_actual:
            return
        if idx < 0 or idx >= len(self.resultado_actual['fallas']):
            return

        falla = self.resultado_actual['fallas'][idx]
        if categorizar_tipo_falla(falla.get('tipo')) != 'piel':
            messagebox.showinfo("Info", "La configuracion individual solo aplica a piel de cocodrilo.")
            return

        if 'mascara_roi' not in falla:
            messagebox.showinfo("Info", "Esta deteccion no tiene mascara disponible para reprocesar.")
            return

        win = tk.Toplevel(self)
        win.title(f"Configurar Piel de Cocodrilo #{falla.get('id', idx + 1)}")
        win.configure(bg=EstiloUI.BG_DARK)
        win.transient(self)
        win.grab_set()
        win.bind("<Escape>", lambda e: win.destroy())

        sf = tk.Frame(win, bg=EstiloUI.BG_PANEL)
        sf.pack(fill="both", expand=True, padx=8, pady=8)
        self._seccion_titulo(sf, f"CONFIGURACION INDIVIDUAL - FALLA #{falla.get('id', idx + 1)}")

        local_vars = {}
        base_cfg = falla.get('config_personalizada', self.motor.config.copy())

        for k, v in [
            ('clahe_clip', 4.0), ('clahe_tile', 8), ('bilateral_d', 9),
            ('bilateral_sigma_color', 75), ('bilateral_sigma_space', 75),
            ('block_size', 23), ('C_umbral', 10), ('kernel_apertura', 3),
            ('kernel_cierre', 6), ('iteraciones_cierre', 2)
        ]:
            val = base_cfg.get(k, v)
            local_vars[k] = tk.DoubleVar(value=val) if isinstance(v, float) else tk.IntVar(value=val)

        local_vars['usar_frangi'] = tk.BooleanVar(value=base_cfg.get('usar_frangi', True))
        local_vars['usar_multiescala'] = tk.BooleanVar(value=base_cfg.get('usar_multiescala', True))
        local_vars['usar_refinamiento'] = tk.BooleanVar(value=base_cfg.get('usar_refinamiento', True))

        def _sl_local(parent, key, label, from_, to, step=1, resolution=None):
            frame = tk.Frame(parent, bg=EstiloUI.BG_PANEL)
            frame.pack(fill="x", padx=10, pady=2)
            hdr = tk.Frame(frame, bg=EstiloUI.BG_PANEL)
            hdr.pack(fill="x")
            tk.Label(hdr, text=label, font=EstiloUI.FONT_SMALL,
                     bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY).pack(side="left")
            val_lbl = tk.Label(hdr, font=("Segoe UI", 9, "bold"),
                               bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT)
            val_lbl.pack(side="right")
            var = local_vars[key]
            slider = tk.Scale(
                frame, from_=from_, to=to, orient="horizontal",
                variable=var, resolution=(resolution if resolution is not None else step),
                showvalue=False, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT,
                troughcolor=EstiloUI.BG_DARK, highlightthickness=0,
                activebackground=EstiloUI.BG_BUTTON, length=320,
                sliderlength=15, bd=0, relief="flat"
            )
            slider.pack(fill="x", pady=(0, 2))

            def _upd(*_args):
                value = var.get()
                if resolution is not None and resolution < 1:
                    val_lbl.config(text=f"{float(value):.2f}")
                else:
                    val_lbl.config(text=str(int(value)))

            var.trace_add("write", _upd)
            _upd()

        _sl_local(sf, 'clahe_clip', "CLAHE Clip", 1.0, 8.0, resolution=0.5)
        _sl_local(sf, 'block_size', "Block Size", 11, 51, step=2)
        _sl_local(sf, 'C_umbral', "Constante C", 3, 25)
        _sl_local(sf, 'kernel_cierre', "Kernel Cierre", 3, 12)
        _sl_local(sf, 'iteraciones_cierre', "Iter. Cierre", 1, 5)

        tk.Checkbutton(sf, text="Filtro Frangi", variable=local_vars['usar_frangi'],
                       bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY,
                       selectcolor=EstiloUI.BG_DARK,
                       font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)
        tk.Checkbutton(sf, text="Refinar esqueleto y cerrar gaps",
                       variable=local_vars['usar_refinamiento'],
                       bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY,
                       selectcolor=EstiloUI.BG_DARK,
                       font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)
        tk.Checkbutton(sf, text="Multi-escala", variable=local_vars['usar_multiescala'],
                       bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY,
                       selectcolor=EstiloUI.BG_DARK,
                       font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)

        btns = tk.Frame(sf, bg=EstiloUI.BG_PANEL)
        btns.pack(fill="x", padx=10, pady=(10, 2))
        ttk.Button(
            btns, text="Aplicar a esta deteccion", style="Accent.TButton",
            command=lambda: self._aplicar_piel_individual(win, idx, local_vars)
        ).pack(fill="x", pady=2)
        ttk.Button(
            btns, text="Cancelar", style="Secondary.TButton",
            command=win.destroy
        ).pack(fill="x", pady=2)

        win.update_idletasks()
        width = max(460, sf.winfo_reqwidth() + 20)
        height = min(int(self.winfo_screenheight() * 0.85), sf.winfo_reqheight() + 30)
        x = self.winfo_rootx() + max(20, (self.winfo_width() - width) // 2)
        y = self.winfo_rooty() + max(20, (self.winfo_height() - height) // 2)
        win.geometry(f"{width}x{height}+{x}+{y}")

    def _aplicar_piel_individual(self, win, idx, local_vars):
        """Reprocesa una deteccion individual de piel usando la configuracion elegida."""
        if not self.resultado_actual or 'fallas' not in self.resultado_actual:
            return
        falla_vieja = self.resultado_actual['fallas'][idx]
        if 'mascara_roi' not in falla_vieja:
            self._log("Mascara no disponible. Procese de nuevo la imagen.")
            win.destroy()
            return

        custom_cfg = self.motor.config.copy()
        for k, var in local_vars.items():
            if isinstance(var, tk.BooleanVar):
                custom_cfg[k] = var.get()
            elif isinstance(var, tk.DoubleVar):
                custom_cfg[k] = float(var.get())
            else:
                custom_cfg[k] = int(var.get())

        proc = ProcesadorPielCocodrilo(custom_cfg)
        self._estado(f"Reprocesando piel de cocodrilo #{falla_vieja.get('id', idx + 1)}...")

        def _reproc():
            try:
                nuevo_res = proc.procesar(
                    falla_vieja['mascara_roi'],
                    self.motor.calibrador,
                    falla_vieja['confianza'],
                    self.resultado_actual['imagen']
                )
                if nuevo_res:
                    n_falla = nuevo_res[0]
                    n_falla['config_personalizada'] = custom_cfg
                    if 'id' in falla_vieja:
                        n_falla['id'] = falla_vieja['id']
                    n_falla['excluida'] = falla_vieja.get('excluida', False)
                    self.resultado_actual['fallas'][idx] = n_falla
                    if hasattr(self, '_resultados_batch'):
                        self._resultados_batch[self.imagen_actual_idx] = self.resultado_actual
                    self.after(0, lambda: self._on_reprocesado_piel(win, "Falla reprocesada con exito."))
                else:
                    self.after(0, lambda: self._on_reprocesado_piel(win, "Reprocesamiento sin resultados."))
            except Exception as e:
                detalle = traceback.format_exc()
                self.after(0, lambda err=str(e), det=detalle: self._on_reprocesado_piel(win, f"Error: {err}", det))

        Thread(target=_reproc, daemon=True).start()

    def _on_reprocesado_piel(self, win, msg, detalle=None):
        """Actualiza UI luego de reprocesar una falla individual de piel."""
        self._log(msg)
        if detalle:
            self._log(detalle.rstrip())
        if self.resultado_actual and self.resultado_actual.get('fallas'):
            self._sync_params()
            vis = self.motor.dibujar_resultado(
                self.resultado_actual['imagen'],
                self.resultado_actual['fallas'],
                progresiva=self._progresiva_actual()
            )
            self._mostrar_cv2(vis, reset_zoom=False)
            self.vista_actual = "resultado"
            self._actualizar_conteos(self.resultado_actual['fallas'])
        self._estado("Listo")
        win.destroy()

    # =========================================================================
    # PROGRESIVAS - PREVIEW Y APLICAR
    # =========================================================================

    def _actualizar_preview_progresiva(self):
        """Actualiza la variable de progresiva desde los campos km y m."""
        try:
            km = int(self.vars['prog_km'].get())
        except (ValueError, tk.TclError):
            km = 0
        try:
            m = int(self.vars['prog_m'].get())
        except (ValueError, tk.TclError):
            m = 0
        prog_str = f"{km:03d}+{m:03d}"
        if hasattr(self, 'lbl_prog_preview'):
            self.lbl_prog_preview.config(text=prog_str)
        self.vars['progresiva_inicio'].set(prog_str)

    def _aplicar_progresiva_actual(self):
        """Asigna la progresiva actual a la imagen seleccionada."""
        if not self.imagenes_cargadas:
            self._log("No hay imagenes cargadas")
            return
        self._actualizar_preview_progresiva()
        prog = self.vars['progresiva_inicio'].get()
        idx = self.imagen_actual_idx
        if not hasattr(self, '_progresivas_asignadas'):
            self._progresivas_asignadas = {}
        self._progresivas_asignadas[idx] = prog
        self._log(f"Progresiva {prog} asignada a imagen {idx + 1}")
        if hasattr(self, 'lbl_prog_estado'):
            self.lbl_prog_estado.config(text=f"Imagen {idx + 1}: {prog}", fg=EstiloUI.FG_HIGHLIGHT)

    def _aplicar_progresiva_todas(self):
        """Asigna progresivas auto-incrementadas a todas las imagenes."""
        if not self.imagenes_cargadas:
            self._log("No hay imagenes cargadas")
            return
        self._actualizar_preview_progresiva()
        try:
            km = int(self.vars['prog_km'].get())
        except (ValueError, tk.TclError):
            km = 0
        try:
            m = int(self.vars['prog_m'].get())
        except (ValueError, tk.TclError):
            m = 0
        try:
            incremento = int(self.vars['progresiva_incremento'].get())
        except (ValueError, tk.TclError):
            incremento = 20

        if not hasattr(self, '_progresivas_asignadas'):
            self._progresivas_asignadas = {}

        total_m = km * 1000 + m
        for i in range(len(self.imagenes_cargadas)):
            p_km = total_m // 1000
            p_m = total_m % 1000
            prog_str = f"{p_km:03d}+{p_m:03d}"
            self._progresivas_asignadas[i] = prog_str
            total_m += incremento

        n = len(self.imagenes_cargadas)
        primera = self._progresivas_asignadas.get(0, '---')
        ultima = self._progresivas_asignadas.get(n - 1, '---')
        self._log(f"Progresivas asignadas a {n} imagenes: {primera} ... {ultima}")
        if hasattr(self, 'lbl_prog_estado'):
            self.lbl_prog_estado.config(
                text=f"{n} imgs: {primera} a {ultima}",
                fg=EstiloUI.FG_HIGHLIGHT)

    # =========================================================================
    # EXPORTAR RESULTADOS NUEVOS (imagenes + excel corregidos)
    # =========================================================================

    def _exportar_resultados_nuevos(self):
        """Exporta resultados corregidos: imagenes resultado, pasos y excel."""
        if not self.motor.todos_resultados and not self._resultados_batch:
            messagebox.showinfo("Info", "No hay resultados para exportar.")
            return

        carpeta = filedialog.askdirectory(title="Carpeta para Resultados Nuevos")
        if not carpeta:
            return

        self._sync_params()
        Path(carpeta).mkdir(parents=True, exist_ok=True)
        carpeta_res = os.path.join(carpeta, "Resultados")
        carpeta_pasos = os.path.join(carpeta, "Pasos_Intermedios")
        Path(carpeta_res).mkdir(parents=True, exist_ok=True)
        Path(carpeta_pasos).mkdir(parents=True, exist_ok=True)

        exportados = 0
        resultados_para_excel = {}

        # Usar resultados de batch o motor
        fuente = self._resultados_batch if self._resultados_batch else {}

        for idx, resultado in fuente.items():
            if not resultado or not resultado.get('fallas'):
                continue

            nombre = resultado.get('nombre', f'imagen_{idx}')
            fallas_activas = [f for f in resultado['fallas'] if not f.get('excluida', False)]

            if not fallas_activas:
                continue

            # Guardar en dict para excel
            resultados_para_excel[nombre] = fallas_activas

            # Redibujar con fallas activas
            vis = self.motor.dibujar_resultado(
                resultado['imagen'], resultado['fallas'])
            out_res = os.path.join(carpeta_res, f"PCI_{nombre}")
            cv2.imwrite(out_res, vis)

            # Pasos
            mosaico = self._generar_mosaico_pasos(resultado)
            if mosaico is not None:
                out_pasos = os.path.join(carpeta_pasos, f"PASOS_{nombre}")
                cv2.imwrite(out_pasos, mosaico)

            exportados += 1

        # Excel con fallas activas
        if resultados_para_excel:
            tramo, _ = self._get_export_params()
            excel_path = ExportadorExcel.exportar(
                resultados_para_excel, carpeta, self.motor.calibrador,
                tramo=tramo, progresivas=None)
        else:
            excel_path = None

        self._log(f"\nResultados Nuevos exportados: {exportados} imagenes")
        if excel_path:
            self._log(f"Excel: {excel_path}")
        self._log(f"Carpeta: {carpeta}")
        messagebox.showinfo("Exportado",
                            f"Exportados: {exportados} imagenes\n"
                            f"Resultados: {carpeta_res}\n"
                            f"Pasos: {carpeta_pasos}\n"
                            f"{'Excel: ' + excel_path if excel_path else ''}")

    # =========================================================================
    # PROGRESIVAS
    # =========================================================================

    def _calcular_progresiva(self, indice_imagen):
        """Calcula la progresiva para una imagen dada su indice. Formato: 000+000."""
        try:
            prog_str = self.vars.get('progresiva_inicio', tk.StringVar(value='000+000')).get()
            incremento = self.vars.get('progresiva_incremento', tk.IntVar(value=20)).get()

            # Parsear formato 000+000
            partes = prog_str.split('+')
            if len(partes) == 2:
                km = int(partes[0])
                m = int(partes[1])
            else:
                km, m = 0, 0

            total_m = km * 1000 + m + (indice_imagen * incremento)
            km_nuevo = total_m // 1000
            m_nuevo = total_m % 1000
            return f"{km_nuevo:03d}+{m_nuevo:03d}"
        except:
            return "000+000"

    def _get_export_params(self):
        """Obtiene tramo y dict de progresivas para exportar."""
        tramo = self.vars.get('tramo', tk.StringVar(value='')).get()
        return tramo, None

    def _progresiva_actual(self):
        """Obtiene la progresiva de la imagen actual (de batch o calculada)."""
        return None

    # =========================================================================
    # ACCIONES
    # =========================================================================

    def _verificar_dependencias(self):
        deps_faltantes = []
        if not YOLO_OK:
            deps_faltantes.append("ultralytics")
        if not PANDAS_OK:
            deps_faltantes.append("pandas")
        if not SKIMAGE_OK:
            deps_faltantes.append("scikit-image")
        if not OPENPYXL_OK:
            deps_faltantes.append("openpyxl")

        if deps_faltantes:
            self._log(f"DEPENDENCIAS FALTANTES: {', '.join(deps_faltantes)}")
            self._log(f"pip install {' '.join(deps_faltantes)}")

    def _cargar_modelo(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar modelo YOLO (.pt)",
            initialdir=str(Path(self.motor.config['ruta_modelo']).parent),
            filetypes=[("Modelos YOLO", "*.pt"), ("Todos", "*.*")]
        )
        if not ruta:
            return

        self._log(f"Cargando modelo: {ruta}")
        self._estado("Cargando modelo...")

        def _load():
            ok, msg = self.motor.cargar_modelo(ruta)
            self.after(0, lambda: self._on_modelo_cargado(ok, msg, ruta))

        Thread(target=_load, daemon=True).start()

    def _on_modelo_cargado(self, ok, msg, ruta):
        self._log(msg)
        if ok:
            self.lbl_modelo.config(text=f"Modelo: {Path(ruta).name}", fg=EstiloUI.BG_SUCCESS)
            self._estado("Modelo cargado correctamente")
        else:
            self._estado("Error al cargar modelo")

    def _cargar_imagenes(self):
        rutas = filedialog.askopenfilenames(
            title="Seleccionar imagenes de pavimento",
            initialdir=self.motor.config['ruta_imagenes'],
            filetypes=[("Imagenes", "*.jpg *.jpeg *.png *.bmp *.tif *.tiff"), ("Todos", "*.*")]
        )
        if not rutas:
            return

        self.imagenes_cargadas = list(rutas)
        self.imagen_actual_idx = 0
        self._log(f"{len(self.imagenes_cargadas)} imagenes cargadas")
        self._mostrar_imagen_actual()
        self._actualizar_nav()

    def _imagen_anterior(self):
        if self.imagenes_cargadas and self.imagen_actual_idx > 0:
            self.imagen_actual_idx -= 1
            self._cargar_resultado_navegacion()
            self._actualizar_nav()

    def _imagen_siguiente(self):
        if self.imagenes_cargadas and self.imagen_actual_idx < len(self.imagenes_cargadas) - 1:
            self.imagen_actual_idx += 1
            self._cargar_resultado_navegacion()
            self._actualizar_nav()

    def _cargar_resultado_navegacion(self):
        """Carga el resultado de la imagen actual si existe (de batch o individual)."""
        idx = self.imagen_actual_idx

        # Restaurar config avanzada guardada para esta imagen
        if idx in self._config_por_imagen:
            self._restaurar_config_imagen(idx)

        # Buscar resultado en batch
        resultado = None
        if hasattr(self, '_resultados_batch') and idx in self._resultados_batch:
            resultado = self._resultados_batch[idx]

        if resultado and resultado.get('fallas'):
            self.resultado_actual = resultado
            vis = self.motor.dibujar_resultado(
                resultado['imagen'], resultado['fallas'],
                progresiva=self._progresiva_actual())
            self._mostrar_cv2(vis)
            self.vista_actual = "resultado"
            self._actualizar_conteos(resultado['fallas'])
        else:
            self.resultado_actual = resultado  # puede ser None o sin fallas
            self._mostrar_imagen_actual()
            # Limpiar conteos
            if hasattr(self, 'lbl_fallas_conteo'):
                self.lbl_fallas_conteo.config(text="0/0")
            # Limpiar severidad
            if hasattr(self, 'severidad_labels'):
                for lbl in self.severidad_labels.values():
                    lbl.config(text="0")
            # Limpiar lista fallas
            self._actualizar_lista_fallas()

    def _actualizar_navegacion(self):
        """Alias para _actualizar_nav (usado por _finalizar_batch)."""
        self._actualizar_nav()

    def _actualizar_nav(self):
        if self.imagenes_cargadas:
            nombre = Path(self.imagenes_cargadas[self.imagen_actual_idx]).name
            self.lbl_nav.config(
                text=f"{self.imagen_actual_idx + 1}/{len(self.imagenes_cargadas)} | {nombre}")
        else:
            self.lbl_nav.config(text="Sin imagenes cargadas")

    def _mostrar_imagen_actual(self):
        if not self.imagenes_cargadas:
            return
        ruta = self.imagenes_cargadas[self.imagen_actual_idx]
        img = cv2.imread(ruta)
        if img is not None:
            self._mostrar_cv2(img)
            self.vista_actual = "original"

    def _mostrar_cv2(self, img_cv2, reset_zoom=True):
        """Almacena la imagen y renderiza con zoom actual."""
        self._cv2_actual = img_cv2
        self._cv2_original_full = img_cv2.copy()

        if reset_zoom:
            self._zoom_fit()
        else:
            self._render_zoom()

        # Actualizar info de imagen
        self._actualizar_info_imagen(img_cv2)

    def _actualizar_info_imagen(self, img_cv2):
        """Actualiza los labels de datos de imagen."""
        if img_cv2 is None:
            return

        h, w = img_cv2.shape[:2]
        canales = img_cv2.shape[2] if len(img_cv2.shape) == 3 else 1
        total_px = h * w
        profundidad = img_cv2.dtype

        # Tamano en memoria
        size_bytes = img_cv2.nbytes
        if size_bytes > 1e6:
            size_str = f"{size_bytes / 1e6:.1f} MB"
        else:
            size_str = f"{size_bytes / 1024:.0f} KB"

        # Nombre archivo
        nombre = "--"
        tamano_disco = "--"
        if self.imagenes_cargadas and 0 <= self.imagen_actual_idx < len(self.imagenes_cargadas):
            ruta = self.imagenes_cargadas[self.imagen_actual_idx]
            nombre = Path(ruta).name
            if len(nombre) > 22:
                nombre = nombre[:19] + "..."
            try:
                file_size = os.path.getsize(ruta)
                if file_size > 1e6:
                    tamano_disco = f"{file_size / 1e6:.1f} MB"
                else:
                    tamano_disco = f"{file_size / 1024:.0f} KB"
            except:
                tamano_disco = size_str

        px_mm = self.motor.calibrador.px_por_mm
        px_mm_str = f"{px_mm:.4f}" if px_mm else "--"

        self.img_data_labels['archivo'].config(text=nombre)
        self.img_data_labels['dimensiones'].config(text=f"{w} x {h} px")
        self.img_data_labels['pixeles_total'].config(text=f"{total_px:,}")
        self.img_data_labels['canales'].config(text=f"{canales} ({'BGR' if canales == 3 else 'Gris'})")
        self.img_data_labels['profundidad'].config(text=str(profundidad))
        self.img_data_labels['tamano_archivo'].config(text=tamano_disco)
        self.img_data_labels['resolucion_dpi'].config(text=px_mm_str)

    def _render_zoom(self):
        """Renderiza la imagen con el nivel de zoom y offset actuales."""
        if self._cv2_original_full is None:
            return

        img = self._cv2_original_full
        h_orig, w_orig = img.shape[:2]

        cw = max(self.canvas_imagen.winfo_width(), 100)
        ch = max(self.canvas_imagen.winfo_height(), 100)

        # Dimensiones de la imagen escalada
        w_scaled = int(w_orig * self._zoom_level)
        h_scaled = int(h_orig * self._zoom_level)

        # Region visible (crop de la imagen original en coords originales)
        # El offset es en coordenadas de la imagen escalada
        x_start_scaled = max(0, -self._pan_offset_x)
        y_start_scaled = max(0, -self._pan_offset_y)
        x_end_scaled = min(w_scaled, cw - self._pan_offset_x)
        y_end_scaled = min(h_scaled, ch - self._pan_offset_y)

        if x_end_scaled <= x_start_scaled or y_end_scaled <= y_start_scaled:
            self.canvas_imagen.delete("all")
            return

        # Convertir a coords originales
        x_start = x_start_scaled / self._zoom_level
        y_start = y_start_scaled / self._zoom_level
        x_end = x_end_scaled / self._zoom_level
        y_end = y_end_scaled / self._zoom_level

        # Crop de la imagen original
        x1 = max(0, int(x_start))
        y1 = max(0, int(y_start))
        x2 = min(w_orig, int(np.ceil(x_end)))
        y2 = min(h_orig, int(np.ceil(y_end)))

        if x2 <= x1 or y2 <= y1:
            return

        crop = img[y1:y2, x1:x2]

        # Escalar el crop
        crop_w = int((x2 - x1) * self._zoom_level)
        crop_h = int((y2 - y1) * self._zoom_level)
        if crop_w < 1 or crop_h < 1:
            return

        interp = cv2.INTER_NEAREST if self._zoom_level > 2.0 else cv2.INTER_LINEAR
        crop_scaled = cv2.resize(crop, (crop_w, crop_h), interpolation=interp)

        # Convertir a PIL
        img_rgb = cv2.cvtColor(crop_scaled, cv2.COLOR_BGR2RGB)
        img_pil = Image.fromarray(img_rgb)

        self._imagen_tk = ImageTk.PhotoImage(img_pil)
        self.canvas_imagen.delete("all")

        # Posicion en el canvas
        draw_x = max(0, self._pan_offset_x)
        draw_y = max(0, self._pan_offset_y)
        self.canvas_imagen.create_image(draw_x, draw_y, anchor="nw", image=self._imagen_tk)

        # Actualizar label de zoom
        self.lbl_zoom.config(text=f"{self._zoom_level * 100:.0f}%")

    # =========================================================================
    # ZOOM Y PAN
    # =========================================================================

    def _zoom_in(self):
        self._apply_zoom(self._zoom_level * 1.25)

    def _zoom_out(self):
        self._apply_zoom(self._zoom_level / 1.25)

    def _zoom_fit(self):
        """Ajusta zoom para que la imagen completa quepa en el canvas."""
        if self._cv2_original_full is None:
            return
        h, w = self._cv2_original_full.shape[:2]
        cw = max(self.canvas_imagen.winfo_width(), 100)
        ch = max(self.canvas_imagen.winfo_height(), 100)
        scale_x = cw / w
        scale_y = ch / h
        self._zoom_level = min(scale_x, scale_y)
        # Centrar
        w_scaled = w * self._zoom_level
        h_scaled = h * self._zoom_level
        self._pan_offset_x = int((cw - w_scaled) / 2)
        self._pan_offset_y = int((ch - h_scaled) / 2)
        self._render_zoom()

    def _zoom_100(self):
        """Zoom 1:1 (un pixel de imagen = un pixel de pantalla)."""
        if self._cv2_original_full is None:
            return
        h, w = self._cv2_original_full.shape[:2]
        cw = max(self.canvas_imagen.winfo_width(), 100)
        ch = max(self.canvas_imagen.winfo_height(), 100)
        self._zoom_level = 1.0
        self._pan_offset_x = int((cw - w) / 2)
        self._pan_offset_y = int((ch - h) / 2)
        self._render_zoom()

    def _apply_zoom(self, new_zoom, center_x=None, center_y=None):
        """Aplica zoom centrado en un punto del canvas."""
        if self._cv2_original_full is None:
            return

        new_zoom = max(self._zoom_min, min(self._zoom_max, new_zoom))

        if center_x is None or center_y is None:
            cw = max(self.canvas_imagen.winfo_width(), 100)
            ch = max(self.canvas_imagen.winfo_height(), 100)
            center_x = cw / 2
            center_y = ch / 2

        # Punto en la imagen original bajo el cursor
        img_x = (center_x - self._pan_offset_x) / self._zoom_level
        img_y = (center_y - self._pan_offset_y) / self._zoom_level

        old_zoom = self._zoom_level
        self._zoom_level = new_zoom

        # Ajustar offset para mantener el punto bajo el cursor
        self._pan_offset_x = int(center_x - img_x * self._zoom_level)
        self._pan_offset_y = int(center_y - img_y * self._zoom_level)

        self._render_zoom()

    def _on_zoom_wheel(self, event):
        """Zoom con rueda del mouse (Windows/Mac)."""
        if self._cv2_original_full is None:
            return
        factor = 1.15 if event.delta > 0 else 1 / 1.15
        self._apply_zoom(self._zoom_level * factor, event.x, event.y)

    def _on_zoom_wheel_linux_up(self, event):
        if self._cv2_original_full is None:
            return
        self._apply_zoom(self._zoom_level * 1.15, event.x, event.y)

    def _on_zoom_wheel_linux_down(self, event):
        if self._cv2_original_full is None:
            return
        self._apply_zoom(self._zoom_level / 1.15, event.x, event.y)

    def _on_pan_start(self, event):
        self._pan_dragging = True
        self._pan_start_x = event.x
        self._pan_start_y = event.y

    def _on_pan_move(self, event):
        if not self._pan_dragging:
            return
        dx = event.x - self._pan_start_x
        dy = event.y - self._pan_start_y
        self._pan_offset_x += dx
        self._pan_offset_y += dy
        self._pan_start_x = event.x
        self._pan_start_y = event.y
        self._render_zoom()

    def _on_pan_end(self, event):
        self._pan_dragging = False

    def _on_mouse_move(self, event):
        """Muestra coordenadas del pixel original bajo el cursor."""
        if self._cv2_original_full is None:
            self.lbl_cursor_pos.config(text="X:-- Y:--")
            if hasattr(self, 'lbl_rgb'):
                self.lbl_rgb.config(text="RGB(0,0,0)", bg="#000000", fg="#FFFFFF")
            return

        # Convertir coords del canvas a coords de la imagen original
        img_x = (event.x - self._pan_offset_x) / self._zoom_level
        img_y = (event.y - self._pan_offset_y) / self._zoom_level

        h, w = self._cv2_original_full.shape[:2]
        if 0 <= img_x < w and 0 <= img_y < h:
            ix, iy = int(img_x), int(img_y)
            pixel = self._cv2_original_full[iy, ix]
            if len(self._cv2_original_full.shape) == 3:
                b, g, r = int(pixel[0]), int(pixel[1]), int(pixel[2])
                self.lbl_cursor_pos.config(text=f"X:{ix} Y:{iy}")
                if hasattr(self, 'lbl_rgb'):
                    hex_color = f"#{r:02x}{g:02x}{b:02x}"
                    # Texto blanco o negro segun luminancia
                    lum = 0.299 * r + 0.587 * g + 0.114 * b
                    fg = "#000000" if lum > 128 else "#FFFFFF"
                    self.lbl_rgb.config(text=f"RGB({r},{g},{b})", bg=hex_color, fg=fg)
            else:
                v = int(pixel)
                self.lbl_cursor_pos.config(text=f"X:{ix} Y:{iy}")
                if hasattr(self, 'lbl_rgb'):
                    hex_color = f"#{v:02x}{v:02x}{v:02x}"
                    fg = "#000000" if v > 128 else "#FFFFFF"
                    self.lbl_rgb.config(text=f"RGB({v},{v},{v})", bg=hex_color, fg=fg)
        else:
            self.lbl_cursor_pos.config(text="X:-- Y:--")

    def _on_canvas_resize(self, event=None):
        """Re-renderiza al cambiar tamano del canvas."""
        if self._cv2_original_full is not None:
            # Solo re-fit si estamos cerca del fit
            self._render_zoom()

    def _cambiar_vista(self, vista):
        if vista == "original":
            self._mostrar_imagen_actual()
        elif vista == "resultado" and self.resultado_actual:
            vis = self.motor.dibujar_resultado(
                self.resultado_actual['imagen'],
                self.resultado_actual['fallas'],
                progresiva=self._progresiva_actual())
            self._mostrar_cv2(vis)
            self.vista_actual = "resultado"
        elif vista == "pasos" and self.resultado_actual:
            self._mostrar_pasos(self.resultado_actual)
        elif not self.resultado_actual:
            self._log("Primero procese la imagen")

    def _generar_mosaico_pasos(self, resultado):
        """Genera imagen de mosaico de pasos intermedios. Retorna imagen cv2 o None."""
        fallas = resultado.get('fallas', [])
        imagen = resultado.get('imagen')
        if not fallas or imagen is None:
            return None

        imgs_con_titulo = []

        for falla in fallas:
            tipo = falla.get('tipo', '')

            # --- PIEL DE COCODRILO: pipeline completo ---
            if 'COCODRILO' in tipo:
                pasos = falla.get('pasos', {})
                for nombre_paso, img_paso in [
                    ('1. Contraste CLAHE', pasos.get('contraste')),
                    ('2. Frangi', pasos.get('frangi')),
                    ('3. Suavizado Bilateral', pasos.get('suavizado')),
                    ('4. Umbralizacion', pasos.get('umbralizada')),
                    ('5. Morfologia', pasos.get('morfologia')),
                    ('6. Limpieza', pasos.get('limpia')),
                ]:
                    if img_paso is not None:
                        imgs_con_titulo.append((nombre_paso, img_paso))

                esq = falla.get('esqueleto')
                if esq is not None:
                    imgs_con_titulo.append(('7. Esqueleto', esq))

                roi = falla.get('mascara_roi')
                if roi is not None:
                    imgs_con_titulo.append(('8. Mascara ROI YOLO', roi))

            # --- GRIETAS: transformada de distancia ---
            elif 'GRIETA' in tipo:
                mask = falla.get('mask')
                dist_map = falla.get('distance_map')
                if mask is not None:
                    imgs_con_titulo.append((f'Grieta #{falla["id"]} - Mascara', (mask * 255).astype(np.uint8)))
                if dist_map is not None:
                    max_dist = np.max(dist_map)
                    if max_dist > 0:
                        dist_norm = (dist_map / max_dist * 255).astype(np.uint8)
                        dist_color = cv2.applyColorMap(dist_norm, cv2.COLORMAP_JET)

                        # Anotar linea de medicion del espesor de zona naranja
                        punto = falla.get('punto_max')
                        espesor_mm = falla.get('espesor_mm', 0)
                        espesor_total_mm = falla.get('espesor_total_mm', espesor_mm)
                        espesor_px = falla.get('espesor_px', 0)
                        angulo = falla.get('angulo_perp', np.pi / 2)
                        if punto is not None and espesor_px > 0:
                            radio = max(int(espesor_px / 2), 3)
                            dx = int(radio * np.cos(angulo))
                            dy = int(radio * np.sin(angulo))
                            p1 = (punto[0] - dx, punto[1] - dy)
                            p2 = (punto[0] + dx, punto[1] + dy)
                            cv2.line(dist_color, p1, p2, (255, 255, 255), 2)
                            cv2.circle(dist_color, punto, 3, (255, 255, 255), -1)
                            # Marcas de acotado
                            ml = 5
                            mx = int(ml * np.cos(angulo + np.pi / 2))
                            my = int(ml * np.sin(angulo + np.pi / 2))
                            cv2.line(dist_color, (p1[0]-mx, p1[1]-my), (p1[0]+mx, p1[1]+my), (255, 255, 255), 2)
                            cv2.line(dist_color, (p2[0]-mx, p2[1]-my), (p2[0]+mx, p2[1]+my), (255, 255, 255), 2)
                            # Mostrar ambos espesores
                            cv2.putText(dist_color, f"e Naranja: {espesor_mm:.1f}mm",
                                        (punto[0] + 10, punto[1] - 10),
                                        cv2.FONT_HERSHEY_SIMPLEX, 0.4, (255, 255, 255), 1)
                            cv2.putText(dist_color, f"e Total: {espesor_total_mm:.1f}mm",
                                        (punto[0] + 10, punto[1] + 15),
                                        cv2.FONT_HERSHEY_SIMPLEX, 0.35, (200, 200, 200), 1)

                        imgs_con_titulo.append((f'Grieta #{falla["id"]} - Espesor', dist_color))

            # --- BACHES: mascara ---
            elif tipo == 'HUECOS':
                mask = falla.get('mask')
                mask_full = falla.get('mask_full')
                if mask_full is not None and falla['id'] == 1:
                    imgs_con_titulo.append(('Huecos - Mascara YOLO', (mask_full * 255).astype(np.uint8)))
                if mask is not None:
                    imgs_con_titulo.append((f'Hueco #{falla["id"]} - Individual', (mask * 255).astype(np.uint8)))

            # --- PARCHES: mascara + fisuras internas ---
            elif tipo == 'PARCHEO':
                cnt = falla.get('contorno')
                if cnt is not None:
                    mask_vis = np.zeros(imagen.shape[:2], dtype=np.uint8)
                    cv2.fillPoly(mask_vis, [cnt], 255)
                    imgs_con_titulo.append((f'Parche #{falla["id"]} - Mascara', mask_vis))

                    # Mostrar fisuras internas detectadas
                    fisuras = falla.get('_fisuras_mask')
                    if fisuras is not None and np.any(fisuras > 0):
                        fis_vis = cv2.cvtColor(mask_vis, cv2.COLOR_GRAY2BGR)
                        fis_vis[fisuras > 0] = (0, 255, 255)  # Amarillo
                        imgs_con_titulo.append((f'Parche #{falla["id"]} - Grietas internas', fis_vis))

        if not imgs_con_titulo:
            return None

        # Construir mosaico con titulos
        h_ref, w_ref = imagen.shape[:2]
        cell_w = min(w_ref, 400)
        cell_h = min(h_ref, 300)

        celdas = []
        for titulo, img_paso in imgs_con_titulo:
            if len(img_paso.shape) == 2:
                img_c = cv2.cvtColor(img_paso, cv2.COLOR_GRAY2BGR)
            else:
                img_c = img_paso.copy()
            img_r = cv2.resize(img_c, (cell_w, cell_h))

            cv2.rectangle(img_r, (0, 0), (cell_w, 25), (30, 30, 30), -1)
            cv2.putText(img_r, titulo, (5, 18),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 255), 1)
            celdas.append(img_r)

        n_cols = min(3, len(celdas))
        n_rows = (len(celdas) + n_cols - 1) // n_cols

        while len(celdas) < n_rows * n_cols:
            celdas.append(np.zeros((cell_h, cell_w, 3), dtype=np.uint8))

        filas = []
        for r in range(n_rows):
            fila = np.hstack(celdas[r * n_cols:(r + 1) * n_cols])
            filas.append(fila)
        mosaico = np.vstack(filas)

        header = np.zeros((35, mosaico.shape[1], 3), dtype=np.uint8)
        cv2.putText(header, f"PASOS INTERMEDIOS - {resultado.get('nombre', '')}",
                    (10, 25), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
        mosaico = np.vstack([header, mosaico])

        return mosaico

    def _mostrar_pasos(self, resultado):
        """Genera y muestra mosaico de pasos intermedios de procesamiento."""
        mosaico = self._generar_mosaico_pasos(resultado)
        if mosaico is None:
            self._log("No hay pasos intermedios disponibles para esta imagen")
            return

        self._mostrar_cv2(mosaico)
        self.vista_actual = "pasos"
        self._log(f"Mostrando pasos intermedios")

    def _procesar_actual(self):
        if not self.imagenes_cargadas:
            messagebox.showinfo("Info", "Cargue imagenes primero")
            return
        if not self.motor.modelo_cargado:
            messagebox.showinfo("Info", "Cargue un modelo YOLO primero")
            return
        if self.procesando:
            return

        self.procesando = True
        self._iniciar_cronometro()
        self._sync_params()
        # Restaurar config individual si esta imagen tiene una guardada
        idx = self.imagen_actual_idx
        if idx in self._config_por_imagen:
            self._restaurar_config_imagen(idx)
        ruta = self.imagenes_cargadas[idx]
        nombre = Path(ruta).name

        self._log(f"\n{'='*50}")
        self._log(f"Procesando: {nombre}")
        self._estado(f"Procesando {nombre}...")

        modo = self.vars.get('modo_calibracion', tk.StringVar(value='automatica')).get()
        if modo == 'automatica':
            self.motor.calibrador.px_por_mm = None  # recalcular con ancho
        elif modo == 'cada_imagen':
            # Calibrar en el hilo principal con VentanaCalibracionTk
            self.motor.calibrador.px_por_mm = None
            imagen = cv2.imread(str(ruta))
            if imagen is not None:
                ancho_via = float(self.vars.get('ancho_via', tk.DoubleVar(value=6.5)).get())
                # VentanaCalibracionTk.__init__ ya llama wait_window() internamente
                win = VentanaCalibracionTk(self, imagen, ancho_via_real_m=ancho_via)
                if win.completado and win.resultado:
                    self.motor.calibrador.calibrar_con_linea(win.resultado)
                    self._log(f"Calibracion por imagen aplicada: {self.motor.calibrador.px_por_mm:.4f} px/mm")
                else:
                    self.motor.calibrador.calibrar_con_ancho_imagen(imagen.shape[1])
                    self._log("Calibracion cancelada, usando ancho de imagen.")
        # modo == 'unica': mantener calibracion existente

        def _proc():
            try:
                resultado = self.motor.procesar_imagen(
                    ruta,
                    callback_log=self._log_safe,
                    calibrar_gui=False
                )
            except Exception as e:
                detalle = traceback.format_exc()
                self.after(
                    0,
                    lambda err=str(e), det=detalle: self._on_error_procesamiento(
                        "Error en procesamiento", err, det)
                )
                return
            self.after(0, lambda r=resultado: self._on_procesado(r))

        Thread(target=_proc, daemon=True).start()

    def _procesar_todo(self):
        if not self.imagenes_cargadas:
            messagebox.showinfo("Info", "Cargue imagenes primero")
            return
        if not self.motor.modelo_cargado:
            messagebox.showinfo("Info", "Cargue un modelo YOLO primero")
            return
        if self.procesando:
            return

        carpeta = filedialog.askdirectory(
            title="Carpeta de salida",
            initialdir=self.motor.config['ruta_salida']
        )
        if not carpeta:
            return

        self.procesando = True
        self._detener_flag = False
        self._iniciar_cronometro()
        self.btn_detener.config(state="normal")
        self._sync_params()
        self._carpeta_salida_batch = carpeta
        Path(carpeta).mkdir(parents=True, exist_ok=True)

        # Capturar params de exportacion y calibracion antes del hilo
        tramo_export, _ = self._get_export_params()
        modo_cal = self.vars.get('modo_calibracion', tk.StringVar(value='automatica')).get()

        carpeta_resultados = os.path.join(carpeta, "Resultados")
        carpeta_pasos = os.path.join(carpeta, "Pasos_Intermedios")
        Path(carpeta_resultados).mkdir(parents=True, exist_ok=True)
        Path(carpeta_pasos).mkdir(parents=True, exist_ok=True)

        # Snapshot de config global base para restaurar entre imagenes
        config_base_batch = self._config_global_base if self._config_global_base else self._snapshot_config_avanzada()

        def _proc_all():
            try:
                total = len(self.imagenes_cargadas)
                procesadas = 0
                self._resultados_batch = {}

                for i, ruta in enumerate(self.imagenes_cargadas):
                    if self._detener_flag:
                        self._log_safe(f"\nDETENIDO por usuario en imagen {i+1}/{total}")
                        break

                    nombre = Path(ruta).name
                    self._log_safe(f"\n[{i+1}/{total}] {nombre}")
                    self.after(0, lambda n=nombre, ii=i: self._estado(f"Procesando {n} ({ii+1}/{total})"))

                    # Restaurar config: individual si existe, sino la global base
                    if i in self._config_por_imagen:
                        src = self._config_por_imagen[i]
                    else:
                        src = config_base_batch
                    for ck, cv in src.items():
                        self.motor.config[ck] = cv
                    self.motor.procesador_piel = ProcesadorPielCocodrilo(self.motor.config)

                    if modo_cal in ('automatica', 'cada_imagen'):
                        self.motor.calibrador.px_por_mm = None
                    # modo_cal == 'unica': mantener

                    resultado = self.motor.procesar_imagen(
                        ruta,
                        callback_log=self._log_safe,
                        calibrar_gui=False
                    )

                    if resultado:
                        self._resultados_batch[i] = resultado

                        if resultado['fallas']:
                            procesadas += 1
                            vis = self.motor.dibujar_resultado(
                                resultado['imagen'], resultado['fallas'])
                            out_resultado = os.path.join(carpeta_resultados, f"PCI_{nombre}")
                            cv2.imwrite(out_resultado, vis)

                            mosaico = self._generar_mosaico_pasos(resultado)
                            if mosaico is not None:
                                out_pasos = os.path.join(carpeta_pasos, f"PASOS_{nombre}")
                                cv2.imwrite(out_pasos, mosaico)

                # Tiempo total
                t_total = time.time() - self._tiempo_inicio
                t_str = f"{t_total:.1f}s" if t_total < 60 else f"{t_total/60:.1f}min"

                excel_path = ExportadorExcel.exportar(
                    self.motor.todos_resultados, carpeta, self.motor.calibrador,
                    tramo=tramo_export, progresivas=None)

                self._log_safe(f"\n{'='*50}")
                self._log_safe(f"COMPLETADO: {procesadas}/{total} imagenes con fallas")
                self._log_safe(f"Tiempo total: {t_str}")
                self._log_safe(f"  Resultados: {carpeta_resultados}")
                self._log_safe(f"  Pasos:      {carpeta_pasos}")
                if excel_path:
                    self._log_safe(f"  Excel:      {excel_path}")
                self._log_safe(f"Navegue entre imagenes para verificar.")

                self.after(0, lambda p=procesadas, tot=total, txt=t_str: self._finalizar_batch(p, tot, txt))
            except Exception as e:
                detalle = traceback.format_exc()
                self.after(
                    0,
                    lambda err=str(e), det=detalle: self._on_error_procesamiento(
                        "Error en procesamiento por lotes", err, det)
                )

        Thread(target=_proc_all, daemon=True).start()

    def _finalizar_batch(self, procesadas, total, t_str=""):
        """Finaliza el batch: muestra resultado de la primera imagen con fallas."""
        self.procesando = False
        t_elapsed = self._detener_cronometro()
        t_str = self._formato_tiempo(t_elapsed)
        self._detener_flag = False
        self.btn_detener.config(state="disabled")
        self._estado(f"Completado: {procesadas}/{total} - Navegue para verificar")
        if t_str:
            self.lbl_tiempo.config(text=t_str)

        encontrada = False
        for i in range(len(self.imagenes_cargadas)):
            resultado = self._resultados_batch.get(i)
            if resultado and resultado.get('fallas'):
                self.imagen_actual_idx = i
                self.resultado_actual = resultado
                self._actualizar_navegacion()

                vis = self.motor.dibujar_resultado(
                    resultado['imagen'], resultado['fallas'],
                    progresiva=self._progresiva_actual())
                self._mostrar_cv2(vis)
                self.vista_actual = "resultado"
                self._actualizar_conteos(resultado['fallas'])
                encontrada = True
                break

        if not encontrada:
            # Sin fallas en ninguna imagen
            self.resultado_actual = None
            if hasattr(self, 'lbl_fallas_conteo'):
                self.lbl_fallas_conteo.config(text="0/0")
            if hasattr(self, 'severidad_labels'):
                for lbl in self.severidad_labels.values():
                    lbl.config(text="0")
            self._actualizar_lista_fallas()

        # Forzar actualizacion visual de la lista de fallas
        self.update_idletasks()

    def _actualizar_conteos(self, fallas):
        """Actualiza los labels de conteo de fallas (excluyendo las marcadas)."""
        activas = [f for f in fallas if not f.get('excluida', False)]
        total = len(fallas)
        n_activas = len(activas)

        # Actualizar label de conteo en el header de FALLAS DETECTADAS
        if hasattr(self, 'lbl_fallas_conteo'):
            self.lbl_fallas_conteo.config(text=f"{n_activas}/{total}")

        # Severidad
        sev_counts = {'L': 0, 'M': 0, 'H': 0}
        for f in activas:
            s = severidad_ui(f.get('severidad'))
            if s in sev_counts:
                sev_counts[s] += 1
        if hasattr(self, 'severidad_labels'):
            for s_key, lbl in self.severidad_labels.items():
                lbl.config(text=str(sev_counts.get(s_key, 0)))

        # Actualizar lista de fallas con checkboxes
        self._actualizar_lista_fallas()

    def _on_procesado(self, resultado):
        self.procesando = False
        t_elapsed = self._detener_cronometro()
        t_str = self._formato_tiempo(t_elapsed)

        if resultado is None:
            self._log("Error en procesamiento")
            self._estado("Error")
            return

        self.resultado_actual = resultado

        if not hasattr(self, '_resultados_batch'):
            self._resultados_batch = {}
        self._resultados_batch[self.imagen_actual_idx] = resultado

        try:
            self._actualizar_navegacion()
            fallas = resultado.get('fallas', [])

            if fallas:
                vis = self.motor.dibujar_resultado(
                    resultado['imagen'], fallas,
                    progresiva=self._progresiva_actual())
                self._mostrar_cv2(vis)
                self.vista_actual = "resultado"
                self._actualizar_conteos(fallas)

                self._log(f"Total fallas: {len(fallas)} | Tiempo: {t_str}")
                for f in fallas:
                    sev = severidad_ui(f.get('severidad'))
                    self._log(f"  {f['tipo']} [{sev}] conf={f['confianza']:.1%}")
            else:
                imagen = resultado.get('imagen')
                if imagen is not None:
                    self._mostrar_cv2(imagen)
                self.vista_actual = "original"
                self._log(f"Sin fallas detectadas | Tiempo: {t_str}")
                if hasattr(self, 'lbl_fallas_conteo'):
                    self.lbl_fallas_conteo.config(text="0/0")
                if hasattr(self, 'severidad_labels'):
                    for lbl in self.severidad_labels.values():
                        lbl.config(text="0")
                self._actualizar_lista_fallas()

            self._estado("Procesamiento completado")
            self.lbl_tiempo.config(text=t_str)
            self.update_idletasks()
        except Exception as e:
            self._estado("Error al mostrar resultados")
            self.lbl_tiempo.config(text=t_str)
            self._log(f"Error al actualizar la interfaz: {e}")
            self._log(traceback.format_exc().rstrip())

    def _exportar_excel(self):
        """Exporta TODOS los resultados acumulados a Excel."""
        if not self.motor.todos_resultados:
            messagebox.showinfo("Info", "No hay resultados para exportar. Procese imagenes primero.")
            return

        carpeta = filedialog.askdirectory(title="Carpeta para Excel")
        if not carpeta:
            return

        tramo, _ = self._get_export_params()
        path = self.motor.exportar_excel(carpeta, tramo=tramo, progresivas=None)
        if path:
            self._log(f"Excel exportado (todos): {path}")
            messagebox.showinfo("Exportado", f"Archivo guardado:\n{path}")
        else:
            self._log("Error al exportar Excel")

    def _exportar_nuevo_excel(self):
        """Exporta solo la imagen actual procesada a un nuevo Excel independiente."""
        if not self.resultado_actual:
            messagebox.showinfo("Info", "Procese una imagen primero antes de exportar.")
            return

        nombre = self.resultado_actual.get('nombre', 'imagen')
        fallas = self.resultado_actual.get('fallas', [])

        if not fallas:
            messagebox.showinfo("Info", f"No hay fallas detectadas en {nombre}.")
            return

        carpeta = filedialog.askdirectory(title="Carpeta para Nuevo Excel")
        if not carpeta:
            return

        # Crear un dict temporal solo con la imagen actual
        resultados_individual = {nombre: fallas}

        # Generar nombre de archivo unico
        nombre_base = Path(nombre).stem
        archivo = Path(carpeta) / f"PCI_{nombre_base}.xlsx"

        # Exportar usando el mismo exportador
        tramo, _ = self._get_export_params()
        path = ExportadorExcel.exportar(
            resultados_individual, carpeta, self.motor.calibrador,
            tramo=tramo, progresivas=None)

        if path:
            # Renombrar al nombre individual
            try:
                archivo_consolidado = Path(path)
                if archivo_consolidado.exists() and archivo_consolidado != archivo:
                    archivo_consolidado.rename(archivo)
                    path = str(archivo)
            except:
                pass

            self._log(f"Nuevo Excel exportado: {path}")
            messagebox.showinfo("Exportado", f"Excel individual guardado:\n{path}\n\nImagen: {nombre}\nFallas: {len(fallas)}")
        else:
            self._log("Error al exportar Nuevo Excel")

    def _guardar_resultado(self):
        """Guarda la imagen de resultado Y los pasos intermedios de la imagen actual."""
        if self._cv2_actual is None and not self.resultado_actual:
            messagebox.showinfo("Info", "No hay resultados para guardar")
            return

        carpeta = filedialog.askdirectory(title="Carpeta para guardar imagenes")
        if not carpeta:
            return

        nombre = "imagen"
        if self.imagenes_cargadas and 0 <= self.imagen_actual_idx < len(self.imagenes_cargadas):
            nombre = Path(self.imagenes_cargadas[self.imagen_actual_idx]).stem

        guardados = []

        # --- Guardar imagen de RESULTADO ---
        if self.resultado_actual and self.resultado_actual.get('fallas'):
            vis = self.motor.dibujar_resultado(
                self.resultado_actual['imagen'],
                self.resultado_actual['fallas'],
                progresiva=self._progresiva_actual())
            ruta_res = os.path.join(carpeta, f"PCI_{nombre}.jpg")
            cv2.imwrite(ruta_res, vis)
            guardados.append(f"Resultado: PCI_{nombre}.jpg")

        # --- Guardar imagen de PASOS INTERMEDIOS ---
        if self.resultado_actual:
            mosaico = self._generar_mosaico_pasos(self.resultado_actual)
            if mosaico is not None:
                ruta_pasos = os.path.join(carpeta, f"PASOS_{nombre}.jpg")
                cv2.imwrite(ruta_pasos, mosaico)
                guardados.append(f"Pasos: PASOS_{nombre}.jpg")

        # --- Guardar vista actual si es diferente (ej: original) ---
        if not guardados and self._cv2_actual is not None:
            ruta_actual = os.path.join(carpeta, f"{self.vista_actual}_{nombre}.jpg")
            cv2.imwrite(ruta_actual, self._cv2_actual)
            guardados.append(f"Vista {self.vista_actual}: {Path(ruta_actual).name}")

        if guardados:
            resumen = "\n".join(guardados)
            self._log(f"Imagenes guardadas en {carpeta}:\n  " + "\n  ".join(guardados))
            messagebox.showinfo("Guardado", f"Archivos guardados:\n{resumen}\n\nEn: {carpeta}")
        else:
            self._log("No hay nada que guardar")

    # =========================================================================
    # UTILIDADES
    # =========================================================================

    def _log(self, msg):
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)

    def _log_safe(self, msg):
        self.after(0, lambda: self._log(msg))

    def _estado(self, msg):
        self.lbl_estado.config(text=msg)


class VentanaCalibracion:
    """Ventana de calibracion basada en Tkinter con zoom, pan y botones."""
    def __init__(self, imagen_cv2, ancho_via_real_m=6.5, parent=None, borde_interno_m=0.30, borde_externo_m=0.30, usar_borde_berma_pci=True):
        self.imagen_original = imagen_cv2.copy()
        self.ancho_via_real_m = ancho_via_real_m
        self.parent = parent
        self.punto_inicio_ancho = None; self.punto_fin_ancho = None
        self.longitud_px = None
        self.punto_inicio_eje = None; self.punto_fin_eje = None
        self.angulo_eje = 90.0
        self.modo = 'ancho'; self.dibujando = False
        self.completado = False
        self._zoom = 1.0; self._pan_x = 0; self._pan_y = 0
        self._drag_data = None
        self._resultado = None
        self.borde_interno_m = self._clamp_borde_m(borde_interno_m)
        self.borde_externo_m = self._clamp_borde_m(borde_externo_m)
        self.usar_borde_berma_pci = bool(usar_borde_berma_pci)
        self.var_borde_interno = None
        self.var_borde_externo = None
        self.var_usar_borde_berma_pci = None

    @staticmethod
    def _clamp_borde_m(valor):
        try:
            valor = float(valor)
        except (TypeError, ValueError):
            valor = 0.30
        return max(0.0, min(3.0, valor))

    def ejecutar(self):
        self.win = tk.Toplevel(self.parent) if self.parent else tk.Tk()
        self.win.title("CALIBRACION - Ancho de Via y Eje")
        self.win.configure(bg=EstiloUI.BG_DARK)
        self.win.protocol("WM_DELETE_WINDOW", self._cerrar_auto)
        # TamaÃ±o ventana
        h, w = self.imagen_original.shape[:2]
        self._zoom = min(900 / w, 600 / h, 1.0)
        win_w = min(int(w * self._zoom) + 40, 1200)
        win_h = min(int(h * self._zoom) + 250, 940)
        self.win.geometry(f"{win_w}x{win_h}")
        if self.parent:
            self.win.transient(self.parent)
            self.win.grab_set()
        # Frame superior: instrucciones
        top = tk.Frame(self.win, bg=EstiloUI.BG_PANEL); top.pack(fill="x")
        top_row = tk.Frame(top, bg=EstiloUI.BG_PANEL)
        top_row.pack(fill="x")
        self.lbl_modo = tk.Label(top_row, text="", font=("Segoe UI", 11, "bold"),
                                  bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT, pady=5)
        self.lbl_modo.pack(side="left", padx=10)
        self.lbl_info = tk.Label(top_row, text="", font=("Segoe UI", 9),
                                  bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_SECONDARY)
        self.lbl_info.pack(side="right", padx=10)
        tk.Label(
            top,
            text="Aplica a: PCI, MTC y VIZIR",
            font=("Segoe UI", 9, "bold"),
            bg=EstiloUI.BG_PANEL,
            fg=EstiloUI.FG_ACCENT,
        ).pack(anchor="w", padx=10, pady=(0, 4))
        # Canvas para imagen
        cf = tk.Frame(self.win, bg=EstiloUI.BG_INPUT); cf.pack(fill="both", expand=True, padx=5, pady=5)
        self.canvas = tk.Canvas(cf, bg=EstiloUI.BG_INPUT, highlightthickness=0, cursor="crosshair")
        self.canvas.pack(fill="both", expand=True)
        self.canvas.bind("<ButtonPress-1>", self._on_press)
        self.canvas.bind("<B1-Motion>", self._on_drag)
        self.canvas.bind("<ButtonRelease-1>", self._on_release)
        self.canvas.bind("<ButtonPress-3>", self._on_pan_start)
        self.canvas.bind("<B3-Motion>", self._on_pan_move)
        self.canvas.bind("<MouseWheel>", self._on_scroll)
        self.canvas.bind("<Button-4>", lambda e: self._on_scroll_linux(e, 1))
        self.canvas.bind("<Button-5>", lambda e: self._on_scroll_linux(e, -1))
        ctrl = tk.Frame(self.win, bg=EstiloUI.BG_PANEL)
        ctrl.pack(fill="x", padx=5, pady=(0, 4))
        tk.Label(
            ctrl,
            text="PCI / MTC / VIZIR | Zona de borde interno y berma desde cada extremo de la linea de ancho (0.00 a 3.00 m)",
            font=("Segoe UI", 9, "bold"),
            bg=EstiloUI.BG_PANEL,
            fg=EstiloUI.FG_ACCENT,
        ).pack(anchor="w", padx=10, pady=(4, 2))
        self.var_usar_borde_berma_pci = tk.BooleanVar(value=self.usar_borde_berma_pci)
        tk.Checkbutton(
            ctrl,
            text="Usar borde interno y berma en esta calibracion",
            variable=self.var_usar_borde_berma_pci,
            command=self._on_toggle_borde_berma,
            bg=EstiloUI.BG_PANEL,
            fg=EstiloUI.FG_PRIMARY,
            selectcolor=EstiloUI.BG_BUTTON_SECONDARY,
            activebackground=EstiloUI.BG_PANEL,
            activeforeground=EstiloUI.FG_PRIMARY,
            font=("Segoe UI", 9),
        ).pack(anchor="w", padx=10, pady=(0, 2))
        self.var_borde_interno = tk.DoubleVar(value=self.borde_interno_m)
        self.var_borde_externo = tk.DoubleVar(value=self.borde_externo_m)
        self._crear_slider_borde(ctrl, "Borde interno (m)", self.var_borde_interno)
        self._crear_slider_borde(ctrl, "Berma (m)", self.var_borde_externo)
        # Frame inferior: botones
        bot = tk.Frame(self.win, bg=EstiloUI.BG_PANEL, height=55); bot.pack(fill="x", side="bottom")
        bot.pack_propagate(False)
        estilo_btn = {"font": ("Segoe UI", 10, "bold"), "width": 16, "height": 1,
                      "relief": "flat", "cursor": "hand2", "bd": 0}
        self.btn_confirmar = tk.Button(bot, text="CONFIRMAR", bg="#00b894", fg="white",
                                        activebackground="#00a884", command=self._confirmar, **estilo_btn)
        self.btn_confirmar.pack(side="left", padx=10, pady=8)
        self.btn_reintentar = tk.Button(bot, text="REINTENTAR", bg="#0984e3", fg="white",
                                         activebackground="#0874c3", command=self._reintentar, **estilo_btn)
        self.btn_reintentar.pack(side="left", padx=5, pady=8)
        self.btn_cerrar = tk.Button(bot, text="CERRAR / AUTO", bg="#d63031", fg="white",
                                     activebackground="#c62020", command=self._cerrar_auto, **estilo_btn)
        self.btn_cerrar.pack(side="right", padx=10, pady=8)
        # Zoom info
        self.lbl_zoom = tk.Label(bot, text="Zoom: 1.0x", font=("Segoe UI", 8),
                                  bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_SECONDARY)
        self.lbl_zoom.pack(side="right", padx=5)
        # Inicializar
        self._actualizar_modo()
        self.win.after(50, self._redibujar)
        self.win.wait_window(self.win)
        return self.longitud_px, self.angulo_eje, self.punto_inicio_eje, self.punto_fin_eje

    def _crear_slider_borde(self, parent, texto, var):
        row = tk.Frame(parent, bg=EstiloUI.BG_PANEL)
        row.pack(fill="x", padx=10, pady=1)
        tk.Label(row, text=texto, font=("Segoe UI", 9), bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, width=18, anchor="w").pack(side="left")
        scale = tk.Scale(
            row,
            from_=0.00,
            to=3.00,
            resolution=0.01,
            orient="horizontal",
            variable=var,
            showvalue=False,
            length=230,
            bg=EstiloUI.BG_PANEL,
            fg=EstiloUI.FG_ACCENT,
            troughcolor=EstiloUI.BG_BUTTON_SECONDARY,
            highlightthickness=0,
            activebackground=EstiloUI.BG_BUTTON,
            command=lambda _v: self._on_borde_slider_change(),
        )
        scale.pack(side="left", fill="x", expand=True, padx=(2, 6))
        value_lbl = tk.Label(row, text=f"{float(var.get()):.2f} m", font=("Segoe UI", 9, "bold"), bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_ACCENT, width=8, anchor="e")
        value_lbl.pack(side="right")
        var.trace_add("write", lambda *_args, v=var, lbl=value_lbl: lbl.config(text=f"{self._clamp_borde_m(v.get()):.2f} m"))

    def _on_borde_slider_change(self):
        self.borde_interno_m = self._clamp_borde_m(self.var_borde_interno.get() if self.var_borde_interno is not None else self.borde_interno_m)
        self.borde_externo_m = self._clamp_borde_m(self.var_borde_externo.get() if self.var_borde_externo is not None else self.borde_externo_m)
        self._redibujar()
        self._actualizar_modo()

    def _on_toggle_borde_berma(self):
        self.usar_borde_berma_pci = bool(self.var_usar_borde_berma_pci.get()) if self.var_usar_borde_berma_pci is not None else self.usar_borde_berma_pci
        self._redibujar()
        self._actualizar_modo()

    def _actualizar_modo(self):
        if self.usar_borde_berma_pci:
            sufijo = f" | Borde:{self.borde_interno_m:.2f}m Berma:{self.borde_externo_m:.2f}m"
        else:
            sufijo = " | Borde/Berma: OFF"
        if self.modo == 'ancho':
            self.lbl_modo.config(text=f"[ANCHO VIA] Dibuje linea transversal ({self.ancho_via_real_m} m) | PCI, MTC y VIZIR")
            info = f"Distancia: {self.longitud_px:.0f} px" if self.longitud_px else "Click izq: dibujar | Scroll: zoom | Click der: mover"
            self.lbl_info.config(text=info + sufijo)
        elif self.modo == 'eje':
            self.lbl_modo.config(text="[EJE VIA] Dibuje linea en direccion del trafico | PCI, MTC y VIZIR")
            info = f"Angulo: {self.angulo_eje:.1f} grados" if self.punto_inicio_eje and self.punto_fin_eje else "Dibuje la direccion del eje de la via"
            self.lbl_info.config(text=info + sufijo)
        self.lbl_zoom.config(text=f"Zoom: {self._zoom:.1f}x")

    def _win_a_img(self, wx, wy):
        return (int((wx - self._pan_x) / self._zoom), int((wy - self._pan_y) / self._zoom))

    def _img_a_win(self, ix, iy):
        return (int(ix * self._zoom + self._pan_x), int(iy * self._zoom + self._pan_y))

    def _on_press(self, event):
        ix, iy = self._win_a_img(event.x, event.y)
        if self.modo == 'ancho':
            self.punto_inicio_ancho = (ix, iy)
            self.punto_fin_ancho = None
            self.dibujando = True
        elif self.modo == 'eje':
            self.punto_inicio_eje = (ix, iy)
            self.punto_fin_eje = None
            self.dibujando = True
        self._redibujar(); self._actualizar_modo()

    def _on_drag(self, event):
        ix, iy = self._win_a_img(event.x, event.y)
        if self.modo == 'ancho' and self.dibujando:
            self.punto_fin_ancho = (ix, iy)
            self.longitud_px = np.sqrt((ix - self.punto_inicio_ancho[0])**2 +
                                       (iy - self.punto_inicio_ancho[1])**2)
        elif self.modo == 'eje' and self.dibujando:
            self.punto_fin_eje = (ix, iy)
            dx = ix - self.punto_inicio_eje[0]; dy = iy - self.punto_inicio_eje[1]
            self.angulo_eje = math.degrees(math.atan2(-dy, dx))
        self._redibujar(); self._actualizar_modo()

    def _on_release(self, event):
        self.dibujando = False
        self._redibujar(); self._actualizar_modo()

    def _on_pan_start(self, event):
        self._drag_data = (event.x - self._pan_x, event.y - self._pan_y)

    def _on_pan_move(self, event):
        if self._drag_data:
            self._pan_x = event.x - self._drag_data[0]
            self._pan_y = event.y - self._drag_data[1]
            self._redibujar()

    def _on_scroll(self, event):
        factor = 1.15 if event.delta > 0 else 1/1.15
        old = self._zoom; self._zoom = max(0.15, min(8.0, self._zoom * factor))
        self._pan_x = int(event.x - (event.x - self._pan_x) * (self._zoom / old))
        self._pan_y = int(event.y - (event.y - self._pan_y) * (self._zoom / old))
        self._redibujar(); self._actualizar_modo()

    def _on_scroll_linux(self, event, direction):
        event.delta = direction * 120; self._on_scroll(event)

    def _redibujar(self):
        self.canvas.delete("all")
        h, w = self.imagen_original.shape[:2]
        nw, nh = max(int(w * self._zoom), 1), max(int(h * self._zoom), 1)
        resized = cv2.resize(self.imagen_original, (nw, nh), interpolation=cv2.INTER_LINEAR)
        rgb = cv2.cvtColor(resized, cv2.COLOR_BGR2RGB)
        self._pil_img = Image.fromarray(rgb)
        self._tk_img = ImageTk.PhotoImage(self._pil_img)
        self.canvas.create_image(self._pan_x, self._pan_y, anchor="nw", image=self._tk_img)
        # LÃ­nea ancho
        if self.punto_inicio_ancho and not self.punto_fin_ancho:
            self._dibujar_punto_control(self.punto_inicio_ancho, "#00ff00", "I")
        if self.punto_fin_ancho and not self.punto_inicio_ancho:
            self._dibujar_punto_control(self.punto_fin_ancho, "#00ff00", "F")
        if self.punto_inicio_ancho and self.punto_fin_ancho:
            x1, y1 = self._img_a_win(*self.punto_inicio_ancho)
            x2, y2 = self._img_a_win(*self.punto_fin_ancho)
            self.canvas.create_line(x1, y1, x2, y2, fill="#00ff00", width=2)
            mx, my = (x1+x2)//2, (y1+y2)//2 - 12
            self.canvas.create_text(mx, my, text=f"{self.longitud_px:.0f}px = {self.ancho_via_real_m}m",
                                     fill="#00ff00", font=("Segoe UI", 10, "bold"))
            if self.usar_borde_berma_pci:
                self._dibujar_zonas_borde_ancho()
            self._dibujar_punto_control(self.punto_inicio_ancho, "#00ff00", "I")
            self._dibujar_punto_control(self.punto_fin_ancho, "#00ff00", "F")
        # LÃ­nea eje
        if self.punto_inicio_eje and not self.punto_fin_eje:
            self._dibujar_punto_control(self.punto_inicio_eje, "#ff00ff", "I")
        if self.punto_fin_eje and not self.punto_inicio_eje:
            self._dibujar_punto_control(self.punto_fin_eje, "#ff00ff", "F")
        if self.punto_inicio_eje and self.punto_fin_eje:
            x1, y1 = self._img_a_win(*self.punto_inicio_eje)
            x2, y2 = self._img_a_win(*self.punto_fin_eje)
            self.canvas.create_line(x1, y1, x2, y2, fill="#ff00ff", width=3)
            mx, my = (x1+x2)//2, (y1+y2)//2 - 12
            self.canvas.create_text(mx, my, text=f"Eje: {self.angulo_eje:.1f} grados",
                                     fill="#ff00ff", font=("Segoe UI", 10, "bold"))
            self._dibujar_punto_control(self.punto_inicio_eje, "#ff00ff", "I")
            self._dibujar_punto_control(self.punto_fin_eje, "#ff00ff", "F")

    def _dibujar_punto_control(self, punto_img, color, etiqueta):
        x, y = self._img_a_win(*punto_img)
        radio = 6
        self.canvas.create_oval(
            x - radio, y - radio, x + radio, y + radio,
            fill=color, outline="white", width=1
        )
        self.canvas.create_text(x, y - 14, text=etiqueta, fill=color, font=("Segoe UI", 9, "bold"))

    def _dibujar_zonas_borde_ancho(self):
        if not (self.punto_inicio_ancho and self.punto_fin_ancho and self.longitud_px and self.longitud_px > 0 and self.ancho_via_real_m > 0):
            return
        ax, ay = self.punto_inicio_ancho
        bx, by = self.punto_fin_ancho
        ux = (bx - ax) / self.longitud_px
        uy = (by - ay) / self.longitud_px
        px_por_m = self.longitud_px / float(self.ancho_via_real_m)
        d_in = self.borde_interno_m * px_por_m
        d_out = self.borde_externo_m * px_por_m

        segmentos = [
            ((ax - ux * d_out, ay - uy * d_out), (ax + ux * d_in, ay + uy * d_in), "#ffd166"),
            ((bx - ux * d_in, by - uy * d_in), (bx + ux * d_out, by + uy * d_out), "#ff9f1c"),
        ]
        for p_ini, p_fin, color in segmentos:
            x1, y1 = self._img_a_win(*p_ini)
            x2, y2 = self._img_a_win(*p_fin)
            self.canvas.create_line(x1, y1, x2, y2, fill=color, width=6)
            self.canvas.create_oval(x1 - 3, y1 - 3, x1 + 3, y1 + 3, fill=color, outline="")
            self.canvas.create_oval(x2 - 3, y2 - 3, x2 + 3, y2 + 3, fill=color, outline="")

    def _confirmar(self):
        if self.modo == 'ancho' and self.longitud_px and self.longitud_px > 10:
            self.modo = 'eje'; self._actualizar_modo(); self._redibujar()
        elif self.modo == 'eje' and self.punto_inicio_eje and self.punto_fin_eje:
            self.completado = True; self.win.destroy()

    def _reintentar(self):
        if self.modo == 'ancho':
            self.punto_inicio_ancho = None; self.punto_fin_ancho = None; self.longitud_px = None
        elif self.modo == 'eje':
            self.punto_inicio_eje = None; self.punto_fin_eje = None
        self._redibujar(); self._actualizar_modo()

    def _cerrar_auto(self):
        if self.modo == 'ancho':
            h, w = self.imagen_original.shape[:2]
            self.longitud_px = w; self.modo = 'eje'
            self._actualizar_modo(); self._redibujar()
        elif self.modo == 'eje':
            self.angulo_eje = 90.0; self.completado = True; self.win.destroy()

# =============================================================================
# VENTANA SELECCIÃ“N DE ROI (RECTÃNGULO) CON ZOOM Y PAN
# =============================================================================
class VentanaSeleccionROI(tk.Toplevel):
    """Ventana Tkinter para seleccionar ROI y auto-calibrar parametros de piel de cocodrilo."""

    def __init__(self, parent, imagen_cv2, titulo="Seleccionar zona"):
        super().__init__(parent)
        self.title(titulo)
        self.configure(bg=EstiloUI.BG_DARK)
        self.geometry("1200x750")
        self.transient(parent)
        self.grab_set()

        self.imagen_original = imagen_cv2.copy()
        self.resultado = None  # dict de params calculados
        self.roi_coords = None  # (x, y, w, h)

        self._zoom = 1.0
        self._pan_x = 0
        self._pan_y = 0
        self._drawing = False
        self._pt1 = None
        self._pt2 = None
        self._pan_start = None

        self._crear_ui()
        self.after(150, self._mostrar_imagen)
        self.protocol("WM_DELETE_WINDOW", self._cancelar)
        self.bind("<Escape>", lambda e: self._cancelar())
        self.wait_window()

    def _crear_ui(self):
        top = tk.Frame(self, bg=EstiloUI.BG_PANEL)
        top.pack(fill="x", padx=5, pady=5)
        tk.Label(top, text="Dibuje un rectangulo sobre la zona de piel de cocodrilo",
                 font=EstiloUI.FONT_SUBTITLE, bg=EstiloUI.BG_PANEL,
                 fg=EstiloUI.FG_HIGHLIGHT).pack(side="left", padx=10)
        self.lbl_info = tk.Label(top, text="Click izq: dibujar | Scroll: zoom | Click der: pan",
                                  font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                                  fg=EstiloUI.FG_SECONDARY)
        self.lbl_info.pack(side="right", padx=10)

        # Contenedor principal: canvas + panel de parÃ¡metros
        main = tk.Frame(self, bg=EstiloUI.BG_DARK)
        main.pack(fill="both", expand=True, padx=5, pady=5)

        self.canvas = tk.Canvas(main, bg=EstiloUI.BG_INPUT, highlightthickness=0)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.canvas.bind("<ButtonPress-1>", self._on_click)
        self.canvas.bind("<B1-Motion>", self._on_drag)
        self.canvas.bind("<ButtonRelease-1>", self._on_release)
        self.canvas.bind("<MouseWheel>", self._on_wheel)
        self.canvas.bind("<ButtonPress-3>", self._on_pan_start)
        self.canvas.bind("<B3-Motion>", self._on_pan_move)
        self.canvas.bind("<ButtonRelease-3>", self._on_pan_end)

        # Panel lateral de parÃ¡metros calculados
        self.panel_params = tk.Frame(main, bg=EstiloUI.BG_PANEL, width=280)
        self.panel_params.pack(side="right", fill="y", padx=(5,0))
        self.panel_params.pack_propagate(False)
        tk.Label(self.panel_params, text="PARAMETROS CALCULADOS",
                 font=EstiloUI.FONT_SUBTITLE, bg=EstiloUI.BG_PANEL,
                 fg=EstiloUI.FG_HIGHLIGHT).pack(pady=(10,5))
        self.lbl_estado = tk.Label(self.panel_params, text="Dibuje un rectangulo primero",
                                    font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                                    fg=EstiloUI.FG_SECONDARY, wraplength=250)
        self.lbl_estado.pack(pady=5)
        # Seleccion de Tipo de Pavimento
        tf = tk.Frame(self.panel_params, bg=EstiloUI.BG_PANEL)
        tf.pack(fill="x", padx=10, pady=(5,10))
        tk.Label(tf, text="Tipo Pavimento:", font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY).pack(side="left")
        self.cb_pavimento = ttk.Combobox(tf, values=["Asfalto Claro / Concreto", "Asfalto Medio / Estandar", "Asfalto Oscuro / Nuevo"],
                                         state="readonly", width=22)
        self.cb_pavimento.current(0)  # Claro/Concreto por defecto
        self.cb_pavimento.pack(side="right")
        # Labels para cada parÃ¡metro
        self._param_labels = {}
        params_info = [
            ('clahe_clip', 'CLAHE Clip'), ('clahe_tile', 'CLAHE Tile'),
            ('bilateral_d', 'Bilateral D'), ('bilateral_sigma_color', 'Sigma Color'),
            ('bilateral_sigma_space', 'Sigma Espacio'), ('block_size', 'Block Size'),
            ('C_umbral', 'Constante C'), ('kernel_apertura', 'Kernel Apert.'),
            ('kernel_cierre', 'Kernel Cierre')
        ]
        for key, label in params_info:
            f = tk.Frame(self.panel_params, bg=EstiloUI.BG_PANEL)
            f.pack(fill="x", padx=10, pady=1)
            tk.Label(f, text=label, font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL,
                     fg=EstiloUI.FG_PRIMARY).pack(side="left")
            lv = tk.Label(f, text="--", font=(EstiloUI.FONT_SMALL[0], 9, "bold"),
                          bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT)
            lv.pack(side="right")
            self._param_labels[key] = lv

        # Barra inferior
        bot = tk.Frame(self, bg=EstiloUI.BG_PANEL)
        bot.pack(fill="x", padx=5, pady=3)
        tk.Button(bot, text="CALCULAR", font=EstiloUI.FONT_SUBTITLE,
                  bg="#f39c12", fg="white", relief="flat", bd=0,
                  padx=15, pady=6, command=self._calcular).pack(side="left", padx=5)
        self.btn_aplicar = tk.Button(bot, text="APLICAR", font=EstiloUI.FONT_SUBTITLE,
                  bg=EstiloUI.BG_SUCCESS, fg="white", relief="flat", bd=0,
                  padx=15, pady=6, command=self._aplicar, state="disabled")
        self.btn_aplicar.pack(side="left", padx=5)
        tk.Button(bot, text="REINICIAR", font=EstiloUI.FONT_SUBTITLE,
                  bg=EstiloUI.BG_BUTTON_SECONDARY, fg="white", relief="flat", bd=0,
                  padx=15, pady=6, command=self._reiniciar).pack(side="left", padx=5)
        self.lbl_zoom = tk.Label(bot, text="Zoom: 1.0x", font=EstiloUI.FONT_SMALL,
                                  bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_SECONDARY)
        self.lbl_zoom.pack(side="left", padx=15)
        tk.Button(bot, text="CANCELAR", font=EstiloUI.FONT_SUBTITLE,
                  bg=EstiloUI.BG_BUTTON, fg="white", relief="flat", bd=0,
                  padx=15, pady=6, command=self._cancelar).pack(side="right", padx=5)
        self.lbl_roi = tk.Label(bot, text="", font=EstiloUI.FONT_BODY,
                                 bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT)
        self.lbl_roi.pack(side="right", padx=10)

    def _mostrar_imagen(self):
        self.update_idletasks()
        cw = self.canvas.winfo_width() or 900
        ch = self.canvas.winfo_height() or 600
        h, w = self.imagen_original.shape[:2]
        self._zoom = min(cw / w, ch / h, 1.0)
        self._pan_x = (cw - w * self._zoom) / 2
        self._pan_y = (ch - h * self._zoom) / 2
        self._render()

    def _render(self):
        self.canvas.delete("all")
        img_src = getattr(self, '_imagen_preview', self.imagen_original)
        h, w = img_src.shape[:2]
        zw, zh = int(w * self._zoom), int(h * self._zoom)
        if zw < 1 or zh < 1: return
        img = cv2.resize(img_src, (zw, zh),
                         interpolation=cv2.INTER_NEAREST if self._zoom > 2 else cv2.INTER_LINEAR)
        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        if self._pt1 and self._pt2:
            p1 = (int(self._pt1[0] * self._zoom), int(self._pt1[1] * self._zoom))
            p2 = (int(self._pt2[0] * self._zoom), int(self._pt2[1] * self._zoom))
            cv2.rectangle(img_rgb, p1, p2, (0, 255, 0), 2)
        pil = Image.fromarray(img_rgb)
        self._tk_img = ImageTk.PhotoImage(pil)
        self.canvas.create_image(int(self._pan_x), int(self._pan_y),
                                  anchor="nw", image=self._tk_img)
        self.lbl_zoom.config(text=f"Zoom: {self._zoom:.1f}x")

    def _canvas_to_img(self, cx, cy):
        return (cx - self._pan_x) / self._zoom, (cy - self._pan_y) / self._zoom

    def _on_click(self, event):
        self._drawing = True
        self._pt1 = self._canvas_to_img(event.x, event.y)
        self._pt2 = self._pt1

    def _on_drag(self, event):
        if self._drawing:
            self._pt2 = self._canvas_to_img(event.x, event.y)
            w = abs(self._pt2[0] - self._pt1[0])
            h = abs(self._pt2[1] - self._pt1[1])
            self.lbl_roi.config(text=f"ROI: {int(w)}x{int(h)} px")
            self._render()

    def _on_release(self, event):
        self._drawing = False
        if self._pt1 and self._pt2:
            w = abs(self._pt2[0] - self._pt1[0])
            h = abs(self._pt2[1] - self._pt1[1])
            self.lbl_roi.config(text=f"ROI: {int(w)}x{int(h)} px - Click CALCULAR")

    def _on_wheel(self, event):
        factor = 1.15 if event.delta > 0 else 1 / 1.15
        old_zoom = self._zoom
        self._zoom = max(0.1, min(10.0, self._zoom * factor))
        self._pan_x = event.x - (event.x - self._pan_x) * (self._zoom / old_zoom)
        self._pan_y = event.y - (event.y - self._pan_y) * (self._zoom / old_zoom)
        self._render()

    def _on_pan_start(self, event):
        self._pan_start = (event.x, event.y)

    def _on_pan_move(self, event):
        if self._pan_start:
            self._pan_x += event.x - self._pan_start[0]
            self._pan_y += event.y - self._pan_start[1]
            self._pan_start = (event.x, event.y)
            self._render()

    def _on_pan_end(self, event):
        self._pan_start = None

    def _calcular(self):
        """Auto-calcula parametros del ROI dibujado y muestra resultados."""
        if not self._pt1 or not self._pt2:
            self.lbl_estado.config(text="Dibuje un rectangulo primero", fg=EstiloUI.FG_SECONDARY)
            return
        x1 = int(min(self._pt1[0], self._pt2[0]))
        y1 = int(min(self._pt1[1], self._pt2[1]))
        x2 = int(max(self._pt1[0], self._pt2[0]))
        y2 = int(max(self._pt1[1], self._pt2[1]))
        h, w = self.imagen_original.shape[:2]
        x1, y1 = max(0, x1), max(0, y1)
        x2, y2 = min(w, x2), min(h, y2)
        rw, rh = x2 - x1, y2 - y1
        if rw < 10 or rh < 10:
            self.lbl_estado.config(text="Rectangulo muy pequeno", fg="#e74c3c")
            return
        self.roi_coords = (x1, y1, rw, rh)
        roi_img = self.imagen_original[y1:y2, x1:x2]
        gris = cv2.cvtColor(roi_img, cv2.COLOR_BGR2GRAY)
        pixels = gris[gris > 0] if np.any(gris > 0) else gris.flatten()
        if len(pixels) < 100:
            self.lbl_estado.config(text="Zona con pocos pixeles", fg="#e74c3c")
            return
        media = np.mean(pixels)
        std = np.std(pixels)
        n_pixels = len(pixels)
        # Auto-calcular parÃ¡metros
        params = {}
        params['clahe_clip'] = 6.0 if std < 30 else (4.0 if std < 50 else 2.5)
        params['clahe_tile'] = 12 if n_pixels > 100000 else (8 if n_pixels > 30000 else 6)
        if std < 25:
            params['bilateral_d'], params['bilateral_sigma_color'], params['bilateral_sigma_space'] = 7, 50, 50
        elif std < 45:
            params['bilateral_d'], params['bilateral_sigma_color'], params['bilateral_sigma_space'] = 9, 75, 75
        else:
            params['bilateral_d'], params['bilateral_sigma_color'], params['bilateral_sigma_space'] = 11, 100, 100
        if media < 80:
            params['block_size'], params['C_umbral'] = 19, 7
        elif media < 140:
            params['block_size'], params['C_umbral'] = 23, 10
        else:
            params['block_size'], params['C_umbral'] = 27, 13
            
        # Ajuste heuristico por tipo de pavimento
        tipo_pav = self.cb_pavimento.get()
        if "Claro" in tipo_pav:
            # Pavimento claro/texturizado = necesita menos sensibilidad al ruido
            params['C_umbral'] += 4
            params['block_size'] += 4
        elif "Oscuro" in tipo_pav:
            # Pavimento oscuro = necesita mas sensibilidad
            params['C_umbral'] = max(2, params['C_umbral'] - 2)
            params['clahe_clip'] = min(8.0, params['clahe_clip'] + 2.0)
        params['kernel_apertura'] = 2 if std < 30 else 3
        params['kernel_cierre'] = 5 if std < 30 else 6
        self._params_calculados = params
        # Ejecutar pipeline con los params calculados para vista previa
        try:
            clahe = cv2.createCLAHE(clipLimit=params['clahe_clip'],
                tileGridSize=(params['clahe_tile'],)*2)
            mej = clahe.apply(gris)
            ecua = cv2.equalizeHist(mej)
            mej = cv2.addWeighted(mej, 0.7, ecua, 0.3, 0)
            # Integracion de Filtro Frangi para vista previa (igual que en motor principal)
            if SKIMAGE_OK:
                try:
                    norm = mej.astype(float)/255.0
                    fr = frangi(norm, scale_range=(1,5), scale_step=1, black_ridges=True)
                    fi = (fr*255).astype(np.uint8)
                    mej = cv2.addWeighted(mej, 0.6, fi, 0.4, 0)
                except: pass
            suav = cv2.bilateralFilter(mej, params['bilateral_d'],
                params['bilateral_sigma_color'], params['bilateral_sigma_space'])
            bs = params['block_size']
            if bs % 2 == 0: bs += 1
            umb = cv2.adaptiveThreshold(suav, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                cv2.THRESH_BINARY_INV, bs, params['C_umbral'])
            k_ap = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (params['kernel_apertura'],)*2)
            ab = cv2.morphologyEx(umb, cv2.MORPH_OPEN, k_ap)
            k_ci = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (params['kernel_cierre'],)*2)
            ce = cv2.morphologyEx(ab, cv2.MORPH_CLOSE, k_ci, iterations=2)
            # Limpiar la imagen original para la vista previa (sin fondo rojo)
            self._imagen_preview = self.imagen_original.copy()
            # Ya no se pinta el fondo ni los contornos rojos, solo el esqueleto y los poligonos
            # Detectar grietas que forman piel de cocodrilo (esqueleto + poligonos)
            n_poligonos = 0
            n_circulos = 0
            try:
                if SKIMAGE_OK and np.sum(ce > 0) > 50:
                    cleaned = remove_small_objects(ce.astype(bool), min_size=80).astype(np.uint8) * 255
                    esq = skeletonize(cleaned.astype(bool))
                    esq_u8 = (esq * 255).astype(np.uint8)
                    
                    if params.get('usar_refinamiento', True):
                        # 1. Refinar esqueleto (eliminar ramas cortas)
                        nl, lb = cv2.connectedComponents(esq_u8)
                        ref = np.zeros_like(esq_u8)
                        for l in range(1, nl):
                            c = (lb == l).astype(np.uint8)*255
                            if np.sum(c == 255) >= params.get('min_longitud_rama', 30):
                                ref = cv2.bitwise_or(ref, c)
                        esq_u8 = ref
                        
                        # 2. Cerrar gaps en el esqueleto
                        g = params.get('max_gap_cierre', 20)
                        k_cg = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (g, g))
                        ce_esq = cv2.morphologyEx(esq_u8, cv2.MORPH_CLOSE, k_cg, iterations=2)
                        esq_u8 = (skeletonize(ce_esq.astype(bool))*255).astype(np.uint8)

                    # Dibujar esqueleto en rojo (grietas) mÃ¡s grueso para que se vea bien
                    esq_coords = np.column_stack(np.where(esq_u8 > 0))
                    for (ey, ex) in esq_coords:
                        py, px = ey + y1, ex + x1
                        if 0 <= py < self._imagen_preview.shape[0] and 0 <= px < self._imagen_preview.shape[1]:
                            # Dibujar un pequeÃ±o cuadrado de 3x3 para que resalte
                            ymin, ymax = max(0, py-1), min(self._imagen_preview.shape[0], py+2)
                            xmin, xmax = max(0, px-1), min(self._imagen_preview.shape[1], px+2)
                            self._imagen_preview[ymin:ymax, xmin:xmax] = [0, 0, 255]
                    # Detectar poligonos cerrados
                    k_dil = np.ones((3, 3), np.uint8)
                    esq_dil = cv2.dilate(esq_u8, k_dil, iterations=1)
                    cnts_p, _ = cv2.findContours(esq_dil, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
                    for cp in cnts_p:
                        a = cv2.contourArea(cp)
                        if a < params.get('min_area_poligono', 300): continue
                        p = cv2.arcLength(cp, True)
                        if p == 0: continue
                        circ = 4 * np.pi * a / (p ** 2)
                        if circ < params.get('min_circularidad', 0.08): continue
                        approx = cv2.approxPolyDP(cp, 3, True)
                        nv = len(approx)
                        if nv < params.get('min_vertices', 4) or nv > params.get('max_vertices', 25): continue
                        n_poligonos += 1
                        # Dibujar poligono en amarillo
                        cp_offset = cp.copy()
                        cp_offset[:, :, 0] += x1
                        cp_offset[:, :, 1] += y1
                        cv2.drawContours(self._imagen_preview, [cp_offset], -1, (0, 255, 255), 1)
                        # Circulo inscrito
                        bx, by, bw, bh = cv2.boundingRect(cp)
                        m = 5
                        mask_c = np.zeros((bh+2*m, bw+2*m), dtype=np.uint8)
                        cv2.fillPoly(mask_c, [cp - [bx-m, by-m]], 255)
                        dist_c = cv2.distanceTransform(mask_c, cv2.DIST_L2, 5)
                        _, rm, _, cl = cv2.minMaxLoc(dist_c)
                        r_circ = int(rm * 0.85)
                        if r_circ >= 5:
                            n_circulos += 1
                            centro = (int(cl[0]+bx-m+x1), int(cl[1]+by-m+y1))
                            cv2.circle(self._imagen_preview, centro, r_circ, (255, 0, 255), 1)
            except Exception:
                pass
        except Exception:
            self._imagen_preview = self.imagen_original.copy()
            n_poligonos = 0
            n_circulos = 0
        self._render()
        # Mostrar en panel
        for key, lbl in self._param_labels.items():
            v = params.get(key, '--')
            lbl.config(text=f"{v:.1f}" if isinstance(v, float) else str(v))
        self.lbl_estado.config(
            text=f"ROI: {rw}x{rh}px | Media: {media:.0f} | Std: {std:.0f} | Grietas: {n_poligonos} poligonos, {n_circulos} celdas",
            fg=EstiloUI.BG_SUCCESS)
        self.btn_aplicar.config(state="normal")

    def _aplicar(self):
        """Acepta los parametros calculados."""
        if hasattr(self, '_params_calculados'):
            self.resultado = self._params_calculados
            self.destroy()

    def _reiniciar(self):
        self._pt1 = None
        self._pt2 = None
        self.roi_coords = None
        self._imagen_preview = self.imagen_original
        self.lbl_roi.config(text="")
        self.lbl_estado.config(text="Dibuje un rectangulo primero", fg=EstiloUI.FG_SECONDARY)
        self.btn_aplicar.config(state="disabled")
        for lbl in self._param_labels.values():
            lbl.config(text="--")
        self._render()

    def _cancelar(self):
        self.resultado = None
        self.destroy()


# =============================================================================
# FUSIONADOR DE MÃSCARAS
# =============================================================================
class FusionadorMascaras:
    @staticmethod
    def fusionar_por_clase(mascaras_por_clase, iou_threshold=0.10, distancia_max_px=50):
        resultado = {}
        for cls_id, lista in mascaras_por_clase.items():
            if len(lista) <= 1:
                resultado[cls_id] = lista; continue
            bins = [(m > 0.5).astype(np.uint8) for m, c in lista]
            confs = [c for m, c in lista]
            n = len(bins)
            parent = list(range(n))
            def find(x):
                while parent[x] != x: parent[x] = parent[parent[x]]; x = parent[x]
                return x
            def union(a, b):
                ra, rb = find(a), find(b)
                if ra != rb: parent[rb] = ra
            for i in range(n):
                for j in range(i+1, n):
                    if FusionadorMascaras._deben_fusionar(bins[i], bins[j], iou_threshold, distancia_max_px):
                        union(i, j)
            grupos = {}
            for i in range(n):
                r = find(i)
                if r not in grupos: grupos[r] = []
                grupos[r].append(i)
            fusionadas = []
            for indices in grupos.values():
                mask_union = np.zeros_like(bins[0], dtype=np.uint8)
                conf_max = 0.0
                for idx in indices:
                    mask_union = np.bitwise_or(mask_union, bins[idx])
                    conf_max = max(conf_max, confs[idx])
                fusionadas.append((mask_union.astype(np.float32), conf_max))
            resultado[cls_id] = fusionadas
        return resultado

    @staticmethod
    def _deben_fusionar(mask_a, mask_b, iou_threshold, distancia_max_px):
        inter = np.sum(np.bitwise_and(mask_a, mask_b))
        union_a = np.sum(np.bitwise_or(mask_a, mask_b))
        if union_a > 0 and inter / union_a >= iou_threshold: return True
        area_a, area_b = np.sum(mask_a), np.sum(mask_b)
        if area_a > 0 and inter / area_a > 0.3: return True
        if area_b > 0 and inter / area_b > 0.3: return True
        if distancia_max_px > 0:
            kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (distancia_max_px, distancia_max_px))
            mask_a_dil = cv2.dilate(mask_a, kernel, iterations=1)
            if np.sum(np.bitwise_and(mask_a_dil, mask_b)) > 0: return True
        return False


# =============================================================================
# TEXTO CON CONTORNO PARA MAYOR VISIBILIDAD
# =============================================================================
def texto_visible(img, texto, pos, escala=0.50, color=(0,255,255), grosor=2, factor=1.0):
    """Dibuja texto con color directo. factor multiplica escala y grosor."""
    x, y = pos
    esc = escala * factor
    gr = max(1, int(grosor * factor))
    cv2.putText(img, texto, (x, y), cv2.FONT_HERSHEY_SIMPLEX, esc, color, gr)


# =============================================================================
# PROCESADOR BACHES (HUECOS) - CLASE 0
# =============================================================================
class ProcesadorBaches:
    NOMBRE_MTC = "BACHES (HUECOS)"

    @staticmethod
    def procesar(mascara, calibrador, confianza, imagen_original, params=None):
        resultados = []
        mask_bin = (mascara > 0.5).astype(np.uint8)
        contornos, _ = cv2.findContours(mask_bin * 255, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        for i, cnt in enumerate(contornos):
            area_px = cv2.contourArea(cnt)
            if area_px < 50: continue
            diametro_px = 2 * np.sqrt(area_px / np.pi)
            diametro_mm = calibrador.px_a_mm(diametro_px)
            diametro_m = (diametro_mm / 1000.0) if diametro_mm else 0
            area_m2 = calibrador.area_px_a_m2(area_px)
            M = cv2.moments(cnt)
            if M["m00"] == 0: continue
            cx, cy = int(M["m10"]/M["m00"]), int(M["m01"]/M["m00"])
            mask_ind = np.zeros_like(mask_bin); cv2.fillPoly(mask_ind, [cnt], 1)
            severidad = ProcesadorBaches._clasificar_severidad(diametro_m)
            resultados.append({
                'tipo': ProcesadorBaches.NOMBRE_MTC, 'id': i+1, 'confianza': confianza,
                'diametro_px': diametro_px, 'diametro_mm': diametro_mm or 0,
                'area_px': area_px, 'area_m2': area_m2 or 0,
                'espesor_px': diametro_px, 'espesor_mm': diametro_mm or 0,
                'longitud_px': 0, 'longitud_m': 0,
                'severidad': severidad, 'ubicacion_x': cx, 'ubicacion_y': cy,
                'contorno': cnt, 'mask': mask_ind, 'unidad': 'UNIDAD',
            })
        return resultados

    @staticmethod
    def _clasificar_severidad(diametro_m):
        if diametro_m is None or diametro_m <= 0: return 1
        if diametro_m < 0.2: return 1
        elif diametro_m <= 0.5: return 2
        else: return 3

    @staticmethod
    def dibujar(imagen, resultados, mostrar_mallas=True, mostrar_etiquetas=True, mostrar_numeros=True, escala_texto=1.0, mostrar_circulos=True, occupied_labels=None):
        vis = imagen.copy()
        colores = {1: (0,255,0), 2: (0,165,255), 3: (0,0,255)}
        etiquetas_ocupadas = occupied_labels if occupied_labels is not None else []
        for r in resultados:
            cnt, sev = r['contorno'], r['severidad']
            color = colores.get(sev, (255,255,255))
            cv2.drawContours(vis, [cnt], -1, color, 2)
            cx, cy = r['ubicacion_x'], r['ubicacion_y']
            radio = max(int(r['diametro_px']/2), 5)
            if mostrar_mallas:
                cv2.line(vis, (cx - radio, cy), (cx + radio, cy), color, 1)
                cv2.line(vis, (cx, cy - radio), (cx, cy + radio), color, 1)
            if mostrar_circulos:
                cv2.circle(vis, (cx, cy), radio, color, 2)
                cv2.circle(vis, (cx, cy), 3, (0, 255, 0), -1)
            if mostrar_etiquetas:
                detalle = ""
                if mostrar_numeros:
                    detalle = f"D={r['diametro_mm']:.0f}mm ({r.get('diametro_mm',0)/1000:.2f}m)"
                texto_visible_dos_lineas(
                    vis,
                    f"BACHES (HUECOS) G:{sev}",
                    detalle,
                    (cx - 80, cy - 25),
                    escala_titulo=0.50,
                    escala_detalle=0.45,
                    color=color,
                    grosor_titulo=2,
                    grosor_detalle=1,
                    factor=escala_texto,
                )
        return vis

    @staticmethod
    def generar_pasos(mascara, imagen_original):
        """Genera imÃ¡genes de pasos intermedios para baches"""
        pasos = {}
        mask_bin = (mascara > 0.5).astype(np.uint8) * 255
        pasos['mascara_binaria'] = mask_bin
        gris = cv2.cvtColor(imagen_original, cv2.COLOR_BGR2GRAY)
        roi = cv2.bitwise_and(gris, gris, mask=mask_bin)
        pasos['roi_gris'] = roi
        contornos, _ = cv2.findContours(mask_bin, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        vis_cnt = imagen_original.copy()
        for cnt in contornos:
            area = cv2.contourArea(cnt)
            if area < 50: continue
            cv2.drawContours(vis_cnt, [cnt], -1, (0, 255, 0), 2)
            diam = 2 * np.sqrt(area / np.pi)
            M = cv2.moments(cnt)
            if M["m00"] > 0:
                cx, cy = int(M["m10"]/M["m00"]), int(M["m01"]/M["m00"])
                cv2.circle(vis_cnt, (cx, cy), max(int(diam/2), 5), (0, 255, 255), 2)
        pasos['contornos_diametros'] = vis_cnt
        dist = cv2.distanceTransform(mask_bin, cv2.DIST_L2, 5)
        if np.max(dist) > 0:
            dn = (dist / np.max(dist) * 255).astype(np.uint8)
            pasos['distancia'] = cv2.applyColorMap(dn, cv2.COLORMAP_JET)
        return pasos

# =============================================================================
# PROCESADOR GRIETAS (FISURAS) - CLASE 1
# =============================================================================
class ProcesadorGrietas:
    @staticmethod
    def procesar(mascara, calibrador, confianza, imagen_original, min_rama_px=30, merge_dist_px=30):
        resultados = []
        mask_bin = (mascara > 0.5).astype(np.uint8)
        mask_original = mask_bin.copy()
        # Cierre morfologico para unir fisuras cercanas del mismo tipo
        if merge_dist_px > 0:
            ks = 2 * merge_dist_px + 1  # kernel impar
            k_union = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (ks, ks))
            mask_unida = cv2.dilate(mask_bin, k_union, iterations=1)
            mask_unida = cv2.morphologyEx(mask_unida, cv2.MORPH_CLOSE, k_union, iterations=1)
        else:
            mask_unida = mask_bin
        # Usar la mascara expandida SOLO para determinar conectividad
        num_labels, labels = cv2.connectedComponents(mask_unida)
        for label_id in range(1, num_labels):
            # Intersectar con mascara original para obtener los pixeles reales
            comp_mask = cv2.bitwise_and(mask_original, (labels == label_id).astype(np.uint8))
            area_px = np.sum(comp_mask)
            if area_px < 30: continue
            r = ProcesadorGrietas._procesar_submask(comp_mask, calibrador, confianza, label_id)
            if r is not None:
                resultados.append(r)
        return resultados

    @staticmethod
    def _separar_ramas(comp_mask, min_rama_px=30):
        """Separa un componente conectado en ramas si tiene puntos de ramificacion.
        Retorna lista de sub-mascaras, una por rama."""
        if not SKIMAGE_OK:
            return [comp_mask]
        try:
            esq = skeletonize(comp_mask.astype(bool)).astype(np.uint8)
        except:
            return [comp_mask]
        if np.sum(esq) < 10:
            return [comp_mask]
        # Detectar branch points: pixeles del esqueleto con >= 3 vecinos en el esqueleto
        kernel_cross = np.array([[1,1,1],[1,10,1],[1,1,1]], dtype=np.uint8)
        vecinos = cv2.filter2D(esq, -1, kernel_cross)
        # branch_points: pixeles que son parte del esqueleto (valor >= 10) y tienen >= 3 vecinos
        branch_points = ((vecinos - 10 * esq) >= 3) & (esq > 0)
        if not np.any(branch_points):
            # Sin ramificaciones: retornar componente original
            return [comp_mask]
        # Eliminar branch points y una vecindad pequeÃ±a para separar bien las ramas
        bp_mask = branch_points.astype(np.uint8)
        k_dil = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
        bp_dilated = cv2.dilate(bp_mask, k_dil, iterations=1)
        esq_sin_bp = esq.copy()
        esq_sin_bp[bp_dilated > 0] = 0
        # Encontrar componentes conectados en el esqueleto sin branch points
        n_ramas, labels_ramas = cv2.connectedComponents(esq_sin_bp)
        if n_ramas <= 2:
            # Solo 1 rama real (o ninguna): retornar original
            return [comp_mask]
        # Para cada rama del esqueleto, expandir para recuperar la mascara real
        sub_masks = []
        # Dilatar cada rama del esqueleto para cubrir el ancho real de la grieta
        dist_map = cv2.distanceTransform(comp_mask, cv2.DIST_L2, 5)
        max_espesor = int(np.max(dist_map)) + 3
        k_expand = cv2.getStructuringElement(cv2.MORPH_ELLIPSE,
            (2 * max_espesor + 1, 2 * max_espesor + 1))
        for rama_id in range(1, n_ramas):
            rama_esq = (labels_ramas == rama_id).astype(np.uint8)
            # Filtrar ramas muy pequeÃ±as
            if np.sum(rama_esq) < min_rama_px:
                continue
            # Expandir el esqueleto de la rama para cubrir toda la grieta
            rama_expandida = cv2.dilate(rama_esq, k_expand, iterations=1)
            # Intersectar con la mascara original del componente
            sub_mask = cv2.bitwise_and(comp_mask, rama_expandida)
            if np.sum(sub_mask) < 30:
                continue
            sub_masks.append(sub_mask)
        # Si no se generaron sub-mascaras validas, retornar original
        if not sub_masks:
            return [comp_mask]
        # Resolver solapamientos: asignar pixeles solapados a la rama mas cercana
        if len(sub_masks) > 1:
            # Crear mapa de asignacion por distancia al esqueleto de cada rama
            asignacion = np.full(comp_mask.shape, -1, dtype=np.int32)
            dist_min = np.full(comp_mask.shape, np.inf, dtype=np.float64)
            for i, sub in enumerate(sub_masks):
                # Distancia desde cada pixel al esqueleto de esta rama
                rama_esq_i = (labels_ramas == (i + 1)).astype(np.uint8)
                # Necesitamos filtrar las ramas pequeÃ±as que se saltaron
                if np.sum(rama_esq_i) < min_rama_px:
                    continue
                inv = 1 - rama_esq_i
                if np.sum(inv) == 0:
                    continue
                d = cv2.distanceTransform(inv, cv2.DIST_L2, 5)
                mascara_comp = comp_mask > 0
                mejor = (d < dist_min) & mascara_comp
                dist_min[mejor] = d[mejor]
                asignacion[mejor] = i
            # Reconstruir sub_masks limpias sin solapamiento
            sub_masks_limpias = []
            for i in range(len(sub_masks)):
                clean = ((asignacion == i) & (comp_mask > 0)).astype(np.uint8)
                if np.sum(clean) >= 30:
                    sub_masks_limpias.append(clean)
            if sub_masks_limpias:
                sub_masks = sub_masks_limpias
        return sub_masks

    @staticmethod
    def _procesar_submask(comp_mask, calibrador, confianza, label_id):
        """Procesa una sub-mascara individual y retorna el diccionario de resultado."""
        area_px = np.sum(comp_mask)
        if area_px < 30:
            return None
        dist_map = cv2.distanceTransform(comp_mask, cv2.DIST_L2, 5)
        max_val = np.max(dist_map)
        # Punto de maximo espesor (toda la mascara)
        max_loc = np.unravel_index(np.argmax(dist_map), dist_map.shape)
        punto_max_global = (max_loc[1], max_loc[0])
        if max_val > 0:
            # Espesor de zona roja del mapa de calor (mayor precision)
            # En COLORMAP_JET, rojo corresponde a valores >= 75% del maximo
            umbral_rojo = 0.75 * max_val
            zona_roja = (dist_map >= umbral_rojo).astype(np.uint8)
            if np.any(zona_roja):
                dist_roja = cv2.distanceTransform(zona_roja, cv2.DIST_L2, 5)
                espesor_px = 2 * np.max(dist_roja)  # ancho de la zona roja
                max_loc_roja = np.unravel_index(np.argmax(dist_roja), dist_roja.shape)
                punto_max = (max_loc_roja[1], max_loc_roja[0])
                radio_max_px = int(np.max(dist_roja))
            else:
                espesor_px = 2 * max_val
                punto_max = punto_max_global
                radio_max_px = int(max_val)
        else:
            espesor_px = 0; punto_max = (0, 0); radio_max_px = 0
        espesor_mm = calibrador.px_a_mm(espesor_px)
        try:
            if SKIMAGE_OK:
                esq = skeletonize(comp_mask.astype(bool))
                longitud_px = float(np.sum(esq))
            else:
                cnts, _ = cv2.findContours(comp_mask*255, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                longitud_px = sum(cv2.arcLength(c, True) for c in cnts) / 2
        except:
            cnts, _ = cv2.findContours(comp_mask*255, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            longitud_px = sum(cv2.arcLength(c, True) for c in cnts) / 2
        longitud_m = calibrador.px_a_m(longitud_px) or 0
        angulo = ProcesadorGrietas._calcular_angulo(comp_mask)
        es_long = calibrador.es_longitudinal(angulo)
        tipo_grieta = "FISURA LONGITUDINAL" if es_long else "FISURA TRANSVERSAL"
        severidad = ProcesadorGrietas._clasificar_severidad(espesor_mm)
        area_mtc = ProcesadorGrietas._calcular_area_mtc(longitud_m, severidad, es_long)
        return {
            'tipo': tipo_grieta, 'id': label_id, 'confianza': confianza,
            'espesor_px': espesor_px, 'espesor_mm': espesor_mm or 0,
            'radio_max_px': radio_max_px,
            'longitud_px': longitud_px, 'longitud_m': longitud_m,
            'diametro_px': 0, 'diametro_mm': 0,
            'area_px': area_px, 'area_m2': area_mtc,
            'severidad': severidad, 'es_longitudinal': es_long,
            'angulo': angulo,
            'ubicacion_x': punto_max[0], 'ubicacion_y': punto_max[1],
            'punto_max': punto_max, 'distance_map': dist_map,
            'mask': comp_mask, 'unidad': 'm2',
        }

    @staticmethod
    def _calcular_angulo(mask):
        coords = np.column_stack(np.where(mask > 0))
        if len(coords) < 5: return 0
        try:
            mean = np.mean(coords, axis=0)
            cov = np.cov((coords - mean).T)
            eigenvalues, eigenvectors = np.linalg.eigh(cov)
            principal = eigenvectors[:, np.argmax(eigenvalues)]
            return math.degrees(math.atan2(-principal[0], principal[1]))
        except: return 0

    @staticmethod
    def _clasificar_severidad(espesor_mm):
        if espesor_mm is None or espesor_mm <= 0: return 1
        if espesor_mm <= 1.0: return 1
        elif espesor_mm <= 3.0: return 2
        else: return 3

    @staticmethod
    def _calcular_area_mtc(longitud_m, severidad, es_longitudinal):
        """Area MTC = Longitud x Factor segun severidad
        Fisuras longitudinales y transversales:
        Sev 1: x 0.10m | Sev 2: x 0.30m | Sev 3: x 0.50m"""
        if longitud_m is None or longitud_m <= 0: return 0.0
        anchos = {1: 0.10, 2: 0.30, 3: 0.50}
        return longitud_m * anchos.get(severidad, 0.10)

    @staticmethod
    def dibujar(imagen, resultados, mostrar_etiquetas=True, mostrar_numeros=True, escala_texto=1.0, occupied_labels=None):
        vis = imagen.copy()
        colores_sev = {1: (0, 255, 0), 2: (0, 165, 255), 3: (0, 0, 255)}  # Verde, Naranja, Rojo
        etiquetas_ocupadas = occupied_labels if occupied_labels is not None else []
        for r in resultados:
            mask, dist_map = r['mask'], r['distance_map']
            punto, sev = r['punto_max'], r['severidad']
            color_sev = colores_sev.get(sev, (0, 165, 255))
            max_val = np.max(dist_map)
            if max_val > 0:
                dist_norm = (dist_map / max_val * 255).astype(np.uint8)
                heatmap = cv2.applyColorMap(dist_norm, cv2.COLORMAP_JET)
                mask_3ch = np.stack([mask]*3, axis=-1)
                vis = np.where(mask_3ch > 0,
                               cv2.addWeighted(vis, 0.5, heatmap * mask_3ch, 0.5, 0), vis)
            # Radio del espesor (zona naranja = 75% del max)
            radio_real = r.get('radio_max_px', max(int(r['espesor_px'] / 2), 3))
            radio_real = max(radio_real, 3)
            # Color naranja fijo para el sÃ­mbolo del espesor
            color_bar = (0, 165, 255)  # Naranja BGR
            # Barra centrada en el punto mÃ¡ximo (rojo), abarcando de naranja a naranja
            punto_bar = punto
            # Ãngulo: barra cruza el ancho (perpendicular a la grieta), ticks paralelos a la grieta
            angulo_grieta = r.get('angulo', 0)
            # DirecciÃ³n perpendicular a la grieta (cruza el espesor)
            ang_perp_rad = math.radians(angulo_grieta) + math.pi / 2
            dx = int(radio_real * math.cos(ang_perp_rad))
            dy = int(radio_real * math.sin(ang_perp_rad))
            p1 = (punto_bar[0] - dx, punto_bar[1] - dy)
            p2 = (punto_bar[0] + dx, punto_bar[1] + dy)
            # Barra "I" cruza el espesor en color naranja
            cv2.line(vis, p1, p2, color_bar, 3)
            # Ticks en extremos: paralelos a la grieta
            tick_len = 6
            ang_grieta_rad = math.radians(angulo_grieta)
            tdx = int(tick_len * math.cos(ang_grieta_rad))
            tdy = int(tick_len * math.sin(ang_grieta_rad))
            cv2.line(vis, (p1[0]-tdx, p1[1]-tdy), (p1[0]+tdx, p1[1]+tdy), color_bar, 2)
            cv2.line(vis, (p2[0]-tdx, p2[1]-tdy), (p2[0]+tdx, p2[1]+tdy), color_bar, 2)
            if mostrar_etiquetas:
                tipo_c = "FISURA LONGITUDINAL" if r.get('es_longitudinal') else "FISURA TRANSVERSAL"
                titulo = f"{tipo_c} [G:{sev}]"
                detalle = ""
                if mostrar_numeros:
                    detalle = f"e={r['espesor_mm']:.1f}mm L={r['longitud_m']:.2f}m"
                texto_visible_dos_lineas(
                    vis,
                    titulo,
                    detalle,
                    (punto[0] - 80, punto[1] - radio_real - 20),
                    escala_titulo=0.45,
                    escala_detalle=max(0.35, 0.45 - 0.03),
                    color=(0, 255, 255),
                    grosor_titulo=2,
                    grosor_detalle=1,
                    factor=escala_texto,
                )
        return vis

    @staticmethod
    def generar_pasos(mascara, imagen_original, calibrador):
        pasos = {}
        mask_bin = (mascara > 0.5).astype(np.uint8)
        pasos['mascara_binaria'] = mask_bin * 255
        dist = cv2.distanceTransform(mask_bin, cv2.DIST_L2, 5)
        if np.max(dist) > 0:
            dn = (dist / np.max(dist) * 255).astype(np.uint8)
            pasos['mapa_distancia'] = cv2.applyColorMap(dn, cv2.COLORMAP_JET)
        if SKIMAGE_OK:
            esq = skeletonize(mask_bin.astype(bool))
            esq_vis = np.zeros_like(imagen_original)
            esq_vis[esq] = (0, 255, 0)
            base = imagen_original.copy()
            base[esq] = (0, 255, 0)
            pasos['esqueleto'] = base
        heatmap_full = imagen_original.copy()
        if np.max(dist) > 0:
            dn2 = (dist / np.max(dist) * 255).astype(np.uint8)
            hm = cv2.applyColorMap(dn2, cv2.COLORMAP_JET)
            m3 = np.stack([mask_bin]*3, axis=-1)
            heatmap_full = np.where(m3 > 0, cv2.addWeighted(heatmap_full, 0.4, hm * m3, 0.6, 0), heatmap_full)
        pasos['mapa_calor'] = heatmap_full
        return pasos

# =============================================================================
# PROCESADOR PARCHES - CLASE 2
# =============================================================================
class ProcesadorParches:
    NOMBRE_MTC = "REPARACIONES O PARCHADOS"

    @staticmethod
    def procesar(mascara, calibrador, confianza, imagen_original, params=None):
        resultados = []
        mask_bin = (mascara > 0.5).astype(np.uint8)
        contornos, _ = cv2.findContours(mask_bin*255, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        for i, cnt in enumerate(contornos):
            area_px = cv2.contourArea(cnt)
            if area_px < 50: continue
            M = cv2.moments(cnt)
            if M["m00"] == 0: continue
            cx, cy = int(M["m10"]/M["m00"]), int(M["m01"]/M["m00"])
            area_m2 = calibrador.area_px_a_m2(area_px) or 0
            comp_mask = np.zeros(mascara.shape[:2], dtype=np.uint8)
            cv2.fillPoly(comp_mask, [cnt], 255)
            
            # Severidad basada en estado del parche
            severidad, edges_int, ratio_fisuras = ProcesadorParches._clasificar_severidad(
                comp_mask, imagen_original, params=params)

            resultados.append({
                'tipo': ProcesadorParches.NOMBRE_MTC, 'id': i+1, 'confianza': confianza,
                'espesor_px': 0, 'espesor_mm': 0,
                'diametro_px': 0, 'diametro_mm': 0,
                'longitud_px': 0, 'longitud_m': 0,
                'ancho_px': 0, 'ancho_m': 0,
                'area_px': area_px, 'area_m2': area_m2,
                'severidad': severidad, 'ubicacion_x': cx, 'ubicacion_y': cy,
                'contorno': cnt, 'unidad': 'm2',
                'esqueleto': edges_int, 'mascara_parche': comp_mask,
                'ratio_fisuras': ratio_fisuras,
            })
        return resultados

    @staticmethod
    def _clasificar_severidad(mask_parche, imagen, params=None):
        """Clasifica severidad del parche analizando fisuracion interna."""
        params = params or {}
        ratio_leve, ratio_moderado = normalizar_umbrales_parche(
            params.get("parche_ratio_leve_max", 0.08),
            params.get("parche_ratio_moderado_max", 0.18),
        )
        edges_masked, ratio = medir_relacion_fisuras_parche(mask_parche, imagen)
        severidad = clasificar_severidad_parche_por_ratio(
            ratio,
            metodo="MTC",
            ratio_leve_max=ratio_leve,
            ratio_moderado_max=ratio_moderado,
        )
        return severidad, edges_masked, ratio

    @staticmethod
    def dibujar(imagen, resultados, mostrar_mallas=True, mostrar_etiquetas=True, mostrar_numeros=True, escala_texto=1.0, mostrar_circulos=True, occupied_labels=None):
        vis = imagen.copy()
        colores = {1: (0,255,0), 2: (0,165,255), 3: (0,0,255)}
        etiquetas_ocupadas = occupied_labels if occupied_labels is not None else []
        for r in resultados:
            cnt, sev = r['contorno'], r['severidad']
            color = colores.get(sev, (255,255,255))
            cv2.drawContours(vis, [cnt], -1, color, 2)
            if mostrar_circulos:
                for circ in r.get('circulos', []):
                    cv2.circle(vis, circ['centro'], circ['radio'], (255, 0, 255), 1)
                    cv2.circle(vis, circ['centro'], 2, (0, 255, 0), -1)
            if mostrar_etiquetas:
                cx, cy = r['ubicacion_x'], r['ubicacion_y']
                detalle = f"A={r['area_m2']:.2f}m2" if mostrar_numeros else ""
                texto_visible_dos_lineas(
                    vis,
                    f"REPARACIONES O PARCHADOS G:{sev}",
                    detalle,
                    (cx - 60, cy - 20),
                    escala_titulo=0.50,
                    escala_detalle=0.45,
                    color=color,
                    grosor_titulo=2,
                    grosor_detalle=1,
                    factor=escala_texto,
                )
        return vis

    @staticmethod
    def generar_pasos(mascara, imagen_original, params=None):
        pasos = {}
        mask_bin = (mascara > 0.5).astype(np.uint8) * 255
        pasos['mascara_binaria'] = mask_bin
        
        roi_color = cv2.bitwise_and(imagen_original, imagen_original, mask=mask_bin)
        pasos['roi_color'] = roi_color
        
        gris = cv2.cvtColor(imagen_original, cv2.COLOR_BGR2GRAY)
        roi_gris = cv2.bitwise_and(gris, gris, mask=mask_bin)
        pasos['roi_gris'] = roi_gris
        
        edges = cv2.Canny(roi_gris, 50, 150)
        edges_masked = cv2.bitwise_and(edges, edges, mask=mask_bin)
        
        # Para visualizaciÃ³n en pasos
        edges_vis = cv2.cvtColor(edges_masked, cv2.COLOR_GRAY2BGR)
        edges_vis[edges_masked > 0] = (0, 0, 255)  # Bordes en rojo
        pasos['bordes_canny'] = edges_vis
        
        # Superponer bordes sobre la imagen original recortada para vis final de pasos
        resultado_mix = roi_color.copy()
        resultado_mix[edges_masked > 0] = (0, 0, 255)
        pasos['resultado_final'] = resultado_mix
        
        return pasos

# =============================================================================
# PROCESADOR PIEL DE COCODRILO - CLASE 3
# =============================================================================
class ProcesadorPielCocodrilo:
    NOMBRE_MTC = "PIEL DE COCODRILO"

    def __init__(self, params=None):
        self.params = params or CONFIG_DEFAULT

    def procesar(self, mascara, calibrador, confianza, imagen_original):
        resultados = []
        mask_bin_full = (mascara > 0.5).astype(np.uint8)
        
        # Encontrar todas las piezas disjuntas (trozos separados) de piel de cocodrilo
        contornos, _ = cv2.findContours(mask_bin_full, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        for i, cnt in enumerate(contornos):
            area_msc_px = cv2.contourArea(cnt)
            # Ignorar trozos enanos
            if area_msc_px < 50: continue
            
            # Crear una mÃ¡scara aislada solo para ESTE trozo de piel de cocodrilo
            h, w = imagen_original.shape[:2]
            mask_r = np.zeros((h, w), dtype=np.uint8)
            cv2.fillPoly(mask_r, [cnt], 255)
            
            # 1. Pipeline para este trozo
            roi = cv2.bitwise_and(imagen_original, imagen_original, mask=mask_r)
            gris = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
            limpia, pasos = self._pipeline_imagen(gris, mask_r)
            
            # 2. EsqueletizaciÃ³n especÃ­fica de este trozo
            if SKIMAGE_OK:
                esq = skeletonize(limpia.astype(bool))
                esq = (esq * 255).astype(np.uint8)
                if self.params.get('usar_refinamiento', True):
                    esq = self._refinar_esqueleto(esq)
                    esq = self._cerrar_gaps(esq)
            else: 
                esq = limpia
                
            # 3. Detectar polÃ­gonos y cÃ­rculos en el trozo
            poligonos = self._detectar_poligonos(esq)
            circulos = []
            for pol in poligonos:
                c = self._calcular_circulo_inscrito(pol['contorno_original'])
                if c: circulos.append(c)
                
            if circulos:
                diam_px = [2*c['radio'] for c in circulos]
                diam_prom_px = np.mean(diam_px)
                diam_prom_m = calibrador.px_a_m(diam_prom_px)
            else: 
                diam_prom_px = 0; diam_prom_m = 0
                
            # 4. CÃ¡lculo de Centroide y Ãrea total paramÃ©trica
            cx, cy, area_m2 = 0, 0, 0
            M = cv2.moments(cnt)
            if M["m00"] > 0: 
                cx, cy = int(M["m10"]/M["m00"]), int(M["m01"]/M["m00"])
            area_px_real = np.sum(mask_r > 0)
            area_m2 = calibrador.area_px_a_m2(area_px_real) or 0
            
            severidad = self._clasificar_severidad(diam_prom_m)
            
            resultados.append({
                'tipo': self.NOMBRE_MTC, 'id': i+1, 'confianza': confianza,
                'espesor_px': diam_prom_px, 'espesor_mm': calibrador.px_a_mm(diam_prom_px) or 0,
                'diametro_px': diam_prom_px, 'diametro_mm': calibrador.px_a_mm(diam_prom_px) or 0,
                'longitud_px': 0, 'longitud_m': 0,
                'ancho_px': 0, 'ancho_m': 0,
                'area_px': area_px_real, 'area_m2': area_m2,
                'severidad': severidad, 'ubicacion_x': cx, 'ubicacion_y': cy,
                'poligonos': poligonos, 'circulos': circulos,
                'esqueleto': esq, 'mascara_roi': mask_r,
                'unidad': 'm2', 'n_celdas': len(circulos),
                'pasos': pasos, 'limpia': limpia,
            })
            
        return resultados

    def _pipeline_imagen(self, gris, mascara_roi):
        pasos = {}
        # Auto-calcular parametros optimos si no estan configurados manualmente
        self._auto_params(gris, mascara_roi)
        clahe = cv2.createCLAHE(clipLimit=self.params.get('clahe_clip', 4.0),
            tileGridSize=(self.params.get('clahe_tile', 8),)*2)
        mej = clahe.apply(gris); ecua = cv2.equalizeHist(mej)
        mej = cv2.addWeighted(mej, 0.7, ecua, 0.3, 0); pasos['contraste'] = mej
        if self.params.get('usar_frangi', True) and SKIMAGE_OK:
            try:
                norm = mej.astype(float)/255.0
                fr = frangi(norm, scale_range=(1,5), scale_step=1, black_ridges=True)
                fi = (fr*255).astype(np.uint8); mej = cv2.addWeighted(mej, 0.6, fi, 0.4, 0)
                pasos['frangi'] = fi
            except: pass
        suav = cv2.bilateralFilter(mej, self.params.get('bilateral_d', 9),
            self.params.get('bilateral_sigma_color', 75), self.params.get('bilateral_sigma_space', 75))
        pasos['suavizado'] = suav
        if self.params.get('usar_multiescala', True):
            umb = self._umb_multi(suav)
        else:
            bs = self.params.get('block_size', 23)
            if bs % 2 == 0: bs += 1
            umb = cv2.adaptiveThreshold(suav, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                cv2.THRESH_BINARY_INV, bs, self.params.get('C_umbral', 10))
        pasos['umbralizada'] = umb
        if mascara_roi is not None:
            umb = cv2.bitwise_and(umb, umb, mask=mascara_roi)
        k_ap = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (self.params.get('kernel_apertura',3),)*2)
        ab = cv2.morphologyEx(umb, cv2.MORPH_OPEN, k_ap)
        k_ci = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (self.params.get('kernel_cierre',6),)*2)
        ce = cv2.morphologyEx(ab, cv2.MORPH_CLOSE, k_ci, iterations=self.params.get('iteraciones_cierre', 2))
        pasos['morfologia'] = ce
        if SKIMAGE_OK:
            li = remove_small_objects(ce.astype(bool), min_size=self.params.get('min_area_objeto', 100)).astype(np.uint8)*255
        else: li = ce
        pasos['limpia'] = li
        return li, pasos

    def _auto_params(self, gris, mascara_roi):
        """Auto-calcula parametros optimos basado en analisis del histograma de la imagen"""
        if mascara_roi is not None:
            pixels = gris[mascara_roi > 0]
        else:
            pixels = gris[gris > 0]
        if len(pixels) < 100: return
        media = np.mean(pixels)
        std = np.std(pixels)
        # CLAHE clip: mayor contraste para imagenes con bajo contraste
        if std < 30:
            self.params['clahe_clip'] = 6.0  # Bajo contraste: clip alto
        elif std < 50:
            self.params['clahe_clip'] = 4.0  # Contraste medio
        else:
            self.params['clahe_clip'] = 2.5  # Alto contraste: clip bajo
        # CLAHE tile: ajustar segun tamaÃ±o de ROI
        n_pixels = len(pixels)
        if n_pixels > 100000:
            self.params['clahe_tile'] = 12
        elif n_pixels > 30000:
            self.params['clahe_tile'] = 8
        else:
            self.params['clahe_tile'] = 6
        # Bilateral: ajustar segun nivel de ruido
        if std < 25:
            self.params['bilateral_d'] = 7
            self.params['bilateral_sigma_color'] = 50
            self.params['bilateral_sigma_space'] = 50
        elif std < 45:
            self.params['bilateral_d'] = 9
            self.params['bilateral_sigma_color'] = 75
            self.params['bilateral_sigma_space'] = 75
        else:
            self.params['bilateral_d'] = 11
            self.params['bilateral_sigma_color'] = 100
            self.params['bilateral_sigma_space'] = 100
        # Block size y C para umbral adaptativo
        if media < 80:
            self.params['block_size'] = 19
            self.params['C_umbral'] = 7
        elif media < 140:
            self.params['block_size'] = 23
            self.params['C_umbral'] = 10
        else:
            self.params['block_size'] = 27
            self.params['C_umbral'] = 13
        # Kernel apertura/cierre
        if std < 30:
            self.params['kernel_apertura'] = 2
            self.params['kernel_cierre'] = 5
        else:
            self.params['kernel_apertura'] = 3
            self.params['kernel_cierre'] = 6

    def _umb_multi(self, gris):
        escalas = [{'bs': 15, 'C': 8, 'p': 0.3}, {'bs': 23, 'C': 10, 'p': 0.4}, {'bs': 33, 'C': 12, 'p': 0.3}]
        res = []
        for e in escalas:
            bs = e['bs']
            if bs % 2 == 0: bs += 1
            u = cv2.adaptiveThreshold(gris, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, bs, e['C'])
            res.append((u * e['p']).astype(np.float32))
        comb = np.clip(np.sum(res, axis=0), 0, 255).astype(np.uint8)
        _, f = cv2.threshold(comb, 127, 255, cv2.THRESH_BINARY); return f

    def _refinar_esqueleto(self, esq):
        nl, lb = cv2.connectedComponents(esq)
        ref = np.zeros_like(esq)
        for l in range(1, nl):
            c = (lb == l).astype(np.uint8)*255
            if np.sum(c == 255) >= self.params.get('min_longitud_rama', 30):
                ref = cv2.bitwise_or(ref, c)
        return ref

    def _cerrar_gaps(self, esq):
        g = self.params.get('max_gap_cierre', 20)
        k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (g, g))
        ce = cv2.morphologyEx(esq, cv2.MORPH_CLOSE, k, iterations=2)
        if SKIMAGE_OK: return (skeletonize(ce.astype(bool))*255).astype(np.uint8)
        return ce

    def _detectar_poligonos(self, esq):
        k = np.ones((3,3), np.uint8); ed = cv2.dilate(esq, k, iterations=1)
        cnts, _ = cv2.findContours(ed, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        pols = []
        for cnt in cnts:
            a = cv2.contourArea(cnt)
            if a < self.params.get('min_area_poligono', 300): continue
            p = cv2.arcLength(cnt, True)
            if p == 0: continue
            circ = 4*np.pi*a/(p**2)
            if circ < self.params.get('min_circularidad', 0.08): continue
            ap = cv2.approxPolyDP(cnt, 3, True); nv = len(ap)
            if nv < self.params.get('min_vertices', 4) or nv > self.params.get('max_vertices', 25): continue
            M = cv2.moments(cnt)
            if M["m00"] == 0: continue
            pols.append({'contorno': ap, 'contorno_original': cnt, 'area': a,
                'perimetro': p, 'circularidad': circ, 'vertices': nv,
                'centroide': (int(M["m10"]/M["m00"]), int(M["m01"]/M["m00"]))})
        return pols

    def _calcular_circulo_inscrito(self, cnt):
        x, y, w, h = cv2.boundingRect(cnt); m = 10
        mask = np.zeros((h+2*m, w+2*m), dtype=np.uint8)
        cv2.fillPoly(mask, [cnt - [x-m, y-m]], 255)
        dist = cv2.distanceTransform(mask, cv2.DIST_L2, 5)
        _, rm, _, cl = cv2.minMaxLoc(dist)
        r = int(rm * 0.85)
        if r < self.params.get('min_radio_circulo', 8): return None
        return {'centro': (int(cl[0]+x-m), int(cl[1]+y-m)), 'radio': r}

    @staticmethod
    def _clasificar_severidad(d_m):
        if d_m is None or d_m == 0: return 2
        if d_m > 0.50: return 1
        elif d_m >= 0.30: return 2
        else: return 3

    @staticmethod
    def dibujar(imagen, resultados, mostrar_mallas=True, mostrar_etiquetas=True, mostrar_numeros=True, escala_texto=1.0, mostrar_circulos=True, mostrar_poligonos=True, occupied_labels=None):
        vis = imagen.copy()
        colores = {1: (0,255,0), 2: (0,165,255), 3: (0,0,255)}
        etiquetas_ocupadas = occupied_labels if occupied_labels is not None else []
        for r in resultados:
            sev = r['severidad']
            d_mm = r.get('diametro_mm', 0)
            if d_mm == 0: sev = 2
            color = colores.get(sev, (255,255,255))
            if 'mascara_roi' in r:
                cnts, _ = cv2.findContours(r['mascara_roi'], cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                cv2.drawContours(vis, cnts, -1, color, 2)
            if mostrar_mallas:
                if 'esqueleto' in r:
                    k = np.ones((2,2), np.uint8); ev = cv2.dilate(r['esqueleto'], k, iterations=1)
                    vis[ev == 255] = (0,0,255)
                if mostrar_poligonos:
                    for pol in r.get('poligonos', []):
                        cv2.drawContours(vis, [pol['contorno']], -1, (0,255,255), 2)
            if mostrar_circulos:
                for circ in r.get('circulos', []):
                    cv2.circle(vis, circ['centro'], circ['radio'], (255,0,255), 2)
                    cv2.circle(vis, circ['centro'], 3, (0,255,0), -1)
            if mostrar_etiquetas:
                cx, cy = r['ubicacion_x'], r['ubicacion_y']
                detalle = f"A={r['area_m2']:.2f}m2 D={d_mm/1000:.3f}m" if mostrar_numeros else ""
                texto_visible_dos_lineas(
                    vis,
                    f"PIEL DE COCODRILO G:{sev}",
                    detalle,
                    (cx - 100, cy - 25),
                    escala_titulo=0.50,
                    escala_detalle=0.45,
                    color=color,
                    grosor_titulo=2,
                    grosor_detalle=1,
                    factor=escala_texto,
                )
        return vis

    def generar_pasos(self, mascara, imagen_original):
        """Genera mosaico de pasos intermedios para P.Cocodrilo"""
        mask_bin = (mascara > 0.5).astype(np.uint8) * 255
        h, w = imagen_original.shape[:2]
        mask_r = cv2.resize(mask_bin, (w, h)) if mask_bin.shape[:2] != (h, w) else mask_bin
        roi = cv2.bitwise_and(imagen_original, imagen_original, mask=mask_r)
        gris = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
        limpia, pasos_pipe = self._pipeline_imagen(gris, mask_r)
        all_pasos = {}
        for k, v in pasos_pipe.items():
            all_pasos[k] = v
        if SKIMAGE_OK:
            esq = skeletonize(limpia.astype(bool)); esq = (esq*255).astype(np.uint8)
            esq = self._refinar_esqueleto(esq); esq = self._cerrar_gaps(esq)
            all_pasos['esqueleto'] = esq
            poligonos = self._detectar_poligonos(esq)
            vis_pol = imagen_original.copy()
            vis_pol = cv2.bitwise_and(vis_pol, vis_pol, mask=mask_r)
            kd = np.ones((2,2), np.uint8); ev = cv2.dilate(esq, kd, iterations=1)
            vis_pol[ev == 255] = (0,0,255)
            for pol in poligonos:
                cv2.drawContours(vis_pol, [pol['contorno']], -1, (0,255,255), 2)
            circulos = []
            for pol in poligonos:
                c = self._calcular_circulo_inscrito(pol['contorno_original'])
                if c:
                    circulos.append(c)
                    cv2.circle(vis_pol, c['centro'], c['radio'], (255,0,255), 2)
            all_pasos['poligonos_circulos'] = vis_pol
        return all_pasos

# =============================================================================
# EXPORTADOR EXCEL
# =============================================================================

# =============================================================================
# ALIASES MTC
# =============================================================================

VentanaCalibracionIntegrada = VentanaCalibracion
ProcesadorBachesMTC = ProcesadorBaches
ProcesadorGrietasMTC = ProcesadorGrietas
ProcesadorParchesMTC = ProcesadorParches
ProcesadorPielCocodriloMTC = ProcesadorPielCocodrilo

class ProcesadorOjoDePescado:
    """
    Clase 0: BACHE â†’ OJO DE PESCADO
    Severidad VIZIR basada en diÃ¡metro y cantidad:
        1: D â‰¤ 300mm, Cantidad < 5
        2: D â‰¤ 300mm, Cantidad 5-10  Ã³  D â‰¤ 1000mm, Cantidad < 5
        3: D â‰¤ 300mm, Cantidad > 10  Ã³  D â‰¤ 1000mm, Cantidad 5-10
    Unidad: UNIDAD
    """

    NOMBRE_VIZIR = "OJO DE PESCADO"

    @staticmethod
    def procesar(mascara, calibrador, confianza, imagen_original):
        resultados = []
        mask_bin = (mascara > 0.5).astype(np.uint8)
        mask_full = mask_bin.copy()

        contornos, _ = cv2.findContours(
            mask_bin * 255, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        detecciones_temp = []
        for i, cnt in enumerate(contornos):
            area_px = cv2.contourArea(cnt)
            if area_px < 50:
                continue

            diametro_px = 2 * np.sqrt(area_px / np.pi)
            diametro_mm = calibrador.px_a_mm(diametro_px)
            area_m2 = calibrador.area_px_a_m2(area_px)

            M = cv2.moments(cnt)
            if M["m00"] == 0:
                continue
            cx = int(M["m10"] / M["m00"])
            cy = int(M["m01"] / M["m00"])

            mask_individual = np.zeros_like(mask_bin)
            cv2.fillPoly(mask_individual, [cnt], 1)

            detecciones_temp.append({
                'idx': i + 1,
                'diametro_px': diametro_px,
                'diametro_mm': diametro_mm if diametro_mm else 0,
                'area_px': area_px,
                'area_m2': area_m2 if area_m2 else 0,
                'cx': cx, 'cy': cy,
                'contorno': cnt,
                'mask': mask_individual,
            })

        # Clasificar severidad considerando CANTIDAD y DIÃMETRO (tabla VIZIR)
        cantidad_total = len(detecciones_temp)

        for det in detecciones_temp:
            severidad = ProcesadorOjoDePescado._clasificar_severidad(
                det['diametro_mm'], cantidad_total)

            resultados.append({
                'tipo': ProcesadorOjoDePescado.NOMBRE_VIZIR,
                'id': det['idx'],
                'confianza': confianza,
                'diametro_px': det['diametro_px'],
                'diametro_mm': det['diametro_mm'],
                'area_px': det['area_px'],
                'area_m2': det['area_m2'],
                'espesor_px': det['diametro_px'],
                'espesor_mm': det['diametro_mm'],
                'longitud_px': 0,
                'longitud_m': 0,
                'severidad': severidad,
                'ubicacion_x': det['cx'],
                'ubicacion_y': det['cy'],
                'contorno': det['contorno'],
                'mask': det['mask'],
                'mask_full': mask_full,
                'unidad': 'UNIDAD',
                'cantidad_en_imagen': cantidad_total,
            })

        return resultados

    @staticmethod
    def _clasificar_severidad(diametro_mm, cantidad):
        """
        Tabla de severidad VIZIR para Ojo de Pescado:
        â”Œâ”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”¬â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”¬â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”
        â”‚  DiÃ¡metro     â”‚ Cantidad â”‚ Severidad   â”‚
        â”œâ”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”¼â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”¼â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”¤
        â”‚ D â‰¤ 300mm     â”‚   < 5    â”‚     1       â”‚
        â”‚ D â‰¤ 300mm     â”‚  5 - 10  â”‚     2       â”‚
        â”‚ D â‰¤ 300mm     â”‚   > 10   â”‚     3       â”‚
        â”‚ D â‰¤ 1000mm    â”‚   < 5    â”‚     2       â”‚
        â”‚ D â‰¤ 1000mm    â”‚  5 - 10  â”‚     3       â”‚
        â”‚ D > 1000mm    â”‚    -     â”‚     3       â”‚
        â””â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”´â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”´â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”˜
        """
        if diametro_mm is None or diametro_mm <= 0:
            return 1

        if diametro_mm <= 300:
            if cantidad < 5:
                return 1
            elif cantidad <= 10:
                return 2
            else:
                return 3
        elif diametro_mm <= 1000:
            if cantidad < 5:
                return 2
            elif cantidad <= 10:
                return 3
            else:
                return 3
        else:
            return 3

    @staticmethod
    def dibujar(imagen, resultados, mostrar_mallas=True, mostrar_etiquetas=True, mostrar_numeros=True, escala_texto=1.0, mostrar_circulos=True, occupied_labels=None):
        vis = imagen.copy()
        colores_sev = {1: (0, 255, 0), 2: (0, 165, 255), 3: (0, 0, 255)}
        etiquetas_ocupadas = occupied_labels if occupied_labels is not None else []
        for r in resultados:
            cnt = r['contorno']
            sev = r['severidad']
            color = colores_sev.get(sev, (255, 255, 255))
            cv2.drawContours(vis, [cnt], -1, color, 2)
            cx, cy = r['ubicacion_x'], r['ubicacion_y']
            radio = int(r['diametro_px'] / 2)
            if mostrar_circulos:
                cv2.circle(vis, (cx, cy), max(radio, 5), color, 2)
            if mostrar_etiquetas:
                tipo_label = str(r.get('tipo', ProcesadorOjoDePescado.NOMBRE_VIZIR) or ProcesadorOjoDePescado.NOMBRE_VIZIR)
                titulo = f"{tipo_label} [Sev:{sev}]"
                detalle = f"D={r['diametro_mm']:.0f}mm" if mostrar_numeros else ""
                texto_visible_dos_lineas(
                    vis,
                    titulo,
                    detalle,
                    (cx - 80, cy - 25),
                    escala_titulo=0.45,
                    escala_detalle=0.42,
                    color=color,
                    grosor_titulo=2,
                    grosor_detalle=1,
                    factor=escala_texto,
                )
        return vis


class ProcesadorFisuras:
    """
    Clase 1: GRIETA â†’ FISURAS DE CONTRACCIÃ”N TÃ‰RMICA
                       o FISURAS LONGITUDINALES POR FATIGA
    ClasificaciÃ³n geomÃ©trica segÃºn Ã¡ngulo respecto al eje de la vÃ­a.
    Severidad VIZIR (longitudinales):
        1: ancho â‰¤ 10 mm
        2: 10 mm < ancho â‰¤ 20 mm
        3: ancho > 20 mm
    Unidad: metros (m)
    """

    NOMBRE_LONGITUDINAL = "FISURAS LONGITUDINALES POR FATIGA"
    NOMBRE_TRANSVERSAL = "FISURAS DE CONTRACCION TERMICA"

    @staticmethod
    def procesar(mascara, calibrador, confianza, imagen_original, merge_dist_px=30):
        resultados = []
        mask_bin = (mascara > 0.5).astype(np.uint8)
        mask_original = mask_bin.copy()

        # Unir segmentos cercanos de la misma orientacion antes de separar componentes.
        if merge_dist_px and merge_dist_px > 0:
            ks = int(2 * merge_dist_px + 1)
            k_union = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (ks, ks))
            mask_unida = cv2.dilate(mask_bin, k_union, iterations=1)
            mask_unida = cv2.morphologyEx(mask_unida, cv2.MORPH_CLOSE, k_union, iterations=1)
        else:
            mask_unida = mask_bin

        num_labels, labels = cv2.connectedComponents(mask_unida)

        for label_id in range(1, num_labels):
            comp_mask = cv2.bitwise_and(mask_original, (labels == label_id).astype(np.uint8))
            area_px = np.sum(comp_mask)
            if area_px < 30:
                continue

            # Transformada de distancia para espesor
            dist_map = cv2.distanceTransform(comp_mask, cv2.DIST_L2, 5)
            max_dist = np.max(dist_map)
            if max_dist > 0:
                # Mismo criterio que PCI: espesor maximo a partir de la zona roja del heatmap.
                umbral_rojo = 0.75 * max_dist
                zona_roja = (dist_map >= umbral_rojo).astype(np.uint8)
                if np.any(zona_roja):
                    dist_roja = cv2.distanceTransform(zona_roja, cv2.DIST_L2, 5)
                    espesor_px = 2 * np.max(dist_roja)
                else:
                    espesor_px = 2 * max_dist
            else:
                espesor_px = 0
            espesor_mm = calibrador.px_a_mm(espesor_px)

            # Punto de mÃ¡ximo espesor (centro de la zona mÃ¡s roja)
            max_loc = np.unravel_index(np.argmax(dist_map), dist_map.shape)
            punto_max = (max_loc[1], max_loc[0])

            # Longitud mediante esqueletizaciÃ³n
            esq = None
            try:
                if SKIMAGE_OK:
                    esq = skeletonize(comp_mask.astype(bool))
                    longitud_px = np.sum(esq)
                else:
                    cnts, _ = cv2.findContours(
                        comp_mask * 255, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    longitud_px = sum(cv2.arcLength(c, True) for c in cnts) / 2
            except Exception:
                cnts, _ = cv2.findContours(
                    comp_mask * 255, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                longitud_px = sum(cv2.arcLength(c, True) for c in cnts) / 2
            angulo_perp = ProcesadorGrietasPCI._calcular_angulo_perpendicular(esq, punto_max, comp_mask)

            longitud_m = calibrador.px_a_m(longitud_px)
            area_m2 = calibrador.area_px_a_m2(area_px)

            # ClasificaciÃ³n geomÃ©trica: longitudinal vs transversal
            angulo_grieta = ProcesadorFisuras._calcular_angulo_grieta(comp_mask)
            angulo_eje = calibrador.get_angulo_eje()
            es_longitudinal = ProcesadorFisuras._es_longitudinal(
                angulo_grieta, angulo_eje)

            if es_longitudinal:
                tipo = ProcesadorFisuras.NOMBRE_LONGITUDINAL
            else:
                tipo = ProcesadorFisuras.NOMBRE_TRANSVERSAL

            # Severidad VIZIR
            severidad = ProcesadorFisuras._clasificar_severidad(espesor_mm)

            resultados.append({
                'tipo': tipo,
                'id': label_id,
                'confianza': confianza,
                'espesor_px': espesor_px,
                'espesor_mm': espesor_mm if espesor_mm else 0,
                'longitud_px': longitud_px,
                'longitud_m': longitud_m if longitud_m else 0,
                'diametro_px': 0,
                'diametro_mm': 0,
                'area_px': area_px,
                'area_m2': area_m2 if area_m2 else 0,
                'severidad': severidad,
                'ubicacion_x': punto_max[0],
                'ubicacion_y': punto_max[1],
                'punto_max': punto_max,
                'distance_map': dist_map,
                'mask': comp_mask,
                'esqueleto': esq,
                'angulo_perp': angulo_perp,
                'angulo_grieta': angulo_grieta,
                'es_longitudinal': es_longitudinal,
                'unidad': 'm',
            })

        return resultados

    @staticmethod
    def _calcular_angulo_grieta(mask):
        """Calcula el Ã¡ngulo de orientaciÃ³n principal de la grieta usando PCA."""
        coords = np.column_stack(np.where(mask > 0))
        if len(coords) < 5:
            return 0.0
        # PCA para orientaciÃ³n principal
        mean = np.mean(coords, axis=0)
        centered = coords - mean
        cov = np.cov(centered.T)
        eigenvalues, eigenvectors = np.linalg.eigh(cov)
        # El eigenvector con mayor eigenvalue da la direcciÃ³n principal
        principal = eigenvectors[:, np.argmax(eigenvalues)]
        angulo = math.degrees(math.atan2(-principal[0], principal[1]))
        return angulo

    @staticmethod
    def _es_longitudinal(angulo_grieta, angulo_eje_via):
        """
        Determina si una grieta es longitudinal (paralela al eje)
        o transversal (perpendicular al eje).
        Umbral: Â±30Â° respecto al eje = longitudinal.
        """
        diff = abs(angulo_grieta - angulo_eje_via) % 180
        if diff > 90:
            diff = 180 - diff
        return diff <= 30  # Dentro de 30Â° del eje = longitudinal

    @staticmethod
    def _clasificar_severidad(espesor_mm):
        """
        Severidad VIZIR para fisuras longitudinales por fatiga:
            1: ancho â‰¤ 10 mm
            2: 10 mm < ancho â‰¤ 20 mm
            3: ancho > 20 mm
        """
        if espesor_mm is None or espesor_mm <= 0:
            return 1
        # VIZIR grietas: S1 <= 10 mm, S2 <= 35 mm, S3 > 35 mm.
        if espesor_mm <= 10:
            return 1
        elif espesor_mm <= 35:
            return 2
        else:
            return 3

    @staticmethod
    def dibujar(imagen, resultados, mostrar_mallas=True, mostrar_etiquetas=True, mostrar_numeros=True, escala_texto=1.0, mostrar_circulos=True, occupied_labels=None):
        vis = imagen.copy()
        colores_sev = {1: (0, 255, 0), 2: (0, 165, 255), 3: (0, 0, 255)}
        etiquetas_ocupadas = occupied_labels if occupied_labels is not None else []

        for r in resultados:
            mask = r['mask']
            dist_map = r['distance_map']
            punto = r['punto_max']
            sev = r['severidad']
            color_sev = colores_sev.get(sev, (255, 255, 255))

            # Mapa de calor sobre la grieta (Jet: Azul=min, Rojo=max)
            if mostrar_mallas and np.max(dist_map) > 0:
                dist_norm = (dist_map / np.max(dist_map) * 255).astype(np.uint8)
                heatmap = cv2.applyColorMap(dist_norm, cv2.COLORMAP_JET)
                mask_3ch = np.stack([mask, mask, mask], axis=-1)
                heatmap_masked = heatmap * mask_3ch
                alpha = 0.5
                vis = np.where(mask_3ch > 0,
                               cv2.addWeighted(vis, 1 - alpha, heatmap_masked, alpha, 0),
                               vis)

            radio_px = int(r['espesor_px'] / 2)
            if mostrar_circulos:
                radio_draw = max(radio_px, 3)
                angulo = r.get('angulo_perp', np.pi / 2)
                dx = int(radio_draw * np.cos(angulo))
                dy = int(radio_draw * np.sin(angulo))
                p1 = (punto[0] - dx, punto[1] - dy)
                p2 = (punto[0] + dx, punto[1] + dy)
                cv2.line(vis, p1, p2, color_sev, 2)
                marca_len = 6
                mx = int(marca_len * np.cos(angulo + np.pi / 2))
                my = int(marca_len * np.sin(angulo + np.pi / 2))
                cv2.line(vis, (p1[0] - mx, p1[1] - my), (p1[0] + mx, p1[1] + my), color_sev, 2)
                cv2.line(vis, (p2[0] - mx, p2[1] - my), (p2[0] + mx, p2[1] + my), color_sev, 2)

            if mostrar_etiquetas:
                tipo_label = str(
                    r.get(
                        'tipo',
                        ProcesadorFisuras.NOMBRE_LONGITUDINAL
                        if r.get('es_longitudinal', True)
                        else ProcesadorFisuras.NOMBRE_TRANSVERSAL,
                    ) or ""
                )
                titulo = f"{tipo_label} [Sev:{sev}]"
                detalle = ""
                if mostrar_numeros:
                    detalle = f"e={r['espesor_mm']:.1f}mm L={r['longitud_m']:.2f}m"
                radio_draw = max(radio_px, 3)
                texto_visible_dos_lineas(
                    vis,
                    titulo,
                    detalle,
                    (punto[0] - 80, punto[1] - radio_draw - 20),
                    escala_titulo=0.42,
                    escala_detalle=0.40,
                    color=color_sev,
                    grosor_titulo=2,
                    grosor_detalle=1,
                    factor=escala_texto,
                )

        return vis


class ProcesadorBachesYParches:
    """
    Clase 2: PARCHE â†’ BACHES Y PARCHES
    Severidad VIZIR basada en Ã¡rea:
        1: Ãrea â‰¤ 0.8 mÂ²
        2: 0.8 < Ãrea â‰¤ 1.5 mÂ²
        3: Ãrea > 1.5 mÂ²
    Unidad: mÂ²
    """

    NOMBRE_VIZIR = "BACHES Y PARCHES"

    @staticmethod
    def procesar(mascara, calibrador, confianza, imagen_original, params=None):
        resultados = []
        mask_bin = (mascara > 0.5).astype(np.uint8)

        contornos, _ = cv2.findContours(
            mask_bin * 255, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        for i, cnt in enumerate(contornos):
            area_px = cv2.contourArea(cnt)
            if area_px < 50:
                continue

            area_m2 = calibrador.area_px_a_m2(area_px)

            M = cv2.moments(cnt)
            if M["m00"] == 0:
                continue
            cx = int(M["m10"] / M["m00"])
            cy = int(M["m01"] / M["m00"])

            comp_mask = np.zeros(mascara.shape[:2], dtype=np.uint8)
            cv2.fillPoly(comp_mask, [cnt], 255)

            severidad, edges_int, ratio_fisuras = ProcesadorBachesYParches._clasificar_severidad(
                comp_mask, imagen_original, params=params
            )

            resultados.append({
                'tipo': ProcesadorBachesYParches.NOMBRE_VIZIR,
                'id': i + 1,
                'confianza': confianza,
                'espesor_px': 0,
                'espesor_mm': 0,
                'diametro_px': 0,
                'diametro_mm': 0,
                'longitud_px': 0,
                'longitud_m': 0,
                'area_px': area_px,
                'area_m2': area_m2 if area_m2 else 0,
                'severidad': severidad,
                'ubicacion_x': cx,
                'ubicacion_y': cy,
                'contorno': cnt,
                'mascara_parche': comp_mask,
                '_fisuras_mask': edges_int,
                'ratio_fisuras': ratio_fisuras,
                'unidad': 'm2',
            })

        return resultados

    @staticmethod
    def _clasificar_severidad(mask_parche, imagen, params=None):
        """
        Severidad VIZIR para Baches y Parches:
            1: Ãrea â‰¤ 0.8 mÂ²
            2: 0.8 < Ãrea â‰¤ 1.5 mÂ²
            3: Ãrea > 1.5 mÂ²
        """
        params = params or {}
        ratio_leve, ratio_moderado = normalizar_umbrales_parche(
            params.get("parche_ratio_leve_max", 0.08),
            params.get("parche_ratio_moderado_max", 0.18),
        )
        edges_masked, ratio = medir_relacion_fisuras_parche(mask_parche, imagen)
        severidad = clasificar_severidad_parche_por_ratio(
            ratio,
            metodo="VIZIR",
            ratio_leve_max=ratio_leve,
            ratio_moderado_max=ratio_moderado,
        )
        return severidad, edges_masked, ratio

    @staticmethod
    def dibujar(imagen, resultados, mostrar_mallas=True, mostrar_etiquetas=True, mostrar_numeros=True, escala_texto=1.0, mostrar_circulos=True, occupied_labels=None):
        vis = imagen.copy()
        colores_sev = {1: (0, 255, 0), 2: (0, 165, 255), 3: (0, 0, 255)}
        etiquetas_ocupadas = occupied_labels if occupied_labels is not None else []
        for r in resultados:
            cnt = r['contorno']
            sev = r['severidad']
            color = colores_sev.get(sev, (255, 255, 255))
            cv2.drawContours(vis, [cnt], -1, color, 2)
            cx, cy = r['ubicacion_x'], r['ubicacion_y']
            if mostrar_etiquetas:
                tipo_label = str(r.get('tipo', ProcesadorBachesYParches.NOMBRE_VIZIR) or ProcesadorBachesYParches.NOMBRE_VIZIR)
                titulo = f"{tipo_label} [Sev:{sev}]"
                detalle = f"A={r['area_m2']:.2f}m2" if mostrar_numeros else ""
                texto_visible_dos_lineas(
                    vis,
                    titulo,
                    detalle,
                    (cx - 60, cy - 20),
                    escala_titulo=0.45,
                    escala_detalle=0.42,
                    color=color,
                    grosor_titulo=2,
                    grosor_detalle=1,
                    factor=escala_texto,
                )
        return vis


class ProcesadorPielCocodrilo:
    """
    Clase 3: PIEL DE COCODRILO â†’ FISURA DE PIEL DE COCODRILO
    Severidad VIZIR basada en diÃ¡metro promedio de celdas:
        1: diÃ¡metro promedio > 500 mm
        2: 200 mm â‰¤ diÃ¡metro â‰¤ 500 mm
        3: diÃ¡metro < 200 mm
    Unidad: mÂ²
    """

    NOMBRE_VIZIR = "FISURA DE PIEL DE COCODRILO"

    def __init__(self, params=None):
        self.params = params or CONFIG_DEFAULT

    def procesar(self, mascara, calibrador, confianza, imagen_original):
        resultados = []
        mask_bin = (mascara > 0.5).astype(np.uint8) * 255

        h, w = imagen_original.shape[:2]
        mask_resized = cv2.resize(mask_bin, (w, h)) if mask_bin.shape[:2] != (h, w) else mask_bin

        roi = cv2.bitwise_and(imagen_original, imagen_original, mask=mask_resized)
        gris = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)

        limpia, pasos = self._pipeline_imagen(gris, mask_resized)
        pasos['roi_gris'] = gris

        # Esqueletizar
        if SKIMAGE_OK:
            esq_raw = skeletonize(limpia.astype(bool))
            esq_raw = (esq_raw * 255).astype(np.uint8)
            pasos['esqueleto_raw'] = esq_raw.copy()
            esq = self._refinar_esqueleto(esq_raw)
            pasos['esqueleto_refinado'] = esq.copy()
            esq = self._cerrar_gaps(esq)
            pasos['esqueleto_cerrado'] = esq.copy()
        else:
            esq = limpia

        # Detectar polÃ­gonos cerrados
        poligonos = self._detectar_poligonos(esq)

        # Calcular cÃ­rculos inscritos
        circulos = []
        for pol in poligonos:
            c = self._calcular_circulo_inscrito(pol['contorno_original'])
            if c:
                circulos.append(c)

        # DiÃ¡metro promedio de celdas
        if circulos:
            diametros_px = [2 * c['radio'] for c in circulos]
            diametro_promedio_px = np.mean(diametros_px)
            diametro_promedio_mm = calibrador.px_a_mm(diametro_promedio_px)
        else:
            diametro_promedio_px = 0
            diametro_promedio_mm = 0

        # Ãrea afectada
        area_px = np.sum(mask_resized > 0)
        area_m2 = calibrador.area_px_a_m2(area_px)

        # Severidad VIZIR
        severidad = self._clasificar_severidad(diametro_promedio_mm)

        # Centroide
        cnts_roi, _ = cv2.findContours(mask_resized, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cx, cy = 0, 0
        if cnts_roi:
            M = cv2.moments(cnts_roi[0])
            if M["m00"] > 0:
                cx = int(M["m10"] / M["m00"])
                cy = int(M["m01"] / M["m00"])

        resultados.append({
            'tipo': self.NOMBRE_VIZIR,
            'id': 1,
            'confianza': confianza,
            'espesor_px': diametro_promedio_px,
            'espesor_mm': diametro_promedio_mm if diametro_promedio_mm else 0,
            'diametro_px': diametro_promedio_px,
            'diametro_mm': diametro_promedio_mm if diametro_promedio_mm else 0,
            'diametro_promedio_mm': diametro_promedio_mm if diametro_promedio_mm else 0,
            'longitud_px': 0,
            'longitud_m': 0,
            'area_px': area_px,
            'area_m2': area_m2 if area_m2 else 0,
            'severidad': severidad,
            'ubicacion_x': cx,
            'ubicacion_y': cy,
            'poligonos': poligonos,
            'circulos': circulos,
            'esqueleto': esq,
            'mascara_roi': mask_resized,
            'unidad': 'm2',
            'n_celdas': len(circulos),
            'pasos': pasos,
            'limpia': limpia,
        })

        return resultados

    def _pipeline_imagen(self, gris, mascara_roi):
        pasos = {}

        clahe = cv2.createCLAHE(
            clipLimit=self.params.get('clahe_clip', 4.0),
            tileGridSize=(self.params.get('clahe_tile', 8),) * 2)
        mejorada = clahe.apply(gris)
        ecualizada = cv2.equalizeHist(mejorada)
        mejorada = cv2.addWeighted(mejorada, 0.7, ecualizada, 0.3, 0)
        pasos['contraste'] = mejorada

        if self.params.get('usar_frangi', True) and SKIMAGE_OK:
            try:
                norm = mejorada.astype(float) / 255.0
                fr = frangi(norm, scale_range=(1, 5), scale_step=1, black_ridges=True)
                frangi_img = (fr * 255).astype(np.uint8)
                mejorada = cv2.addWeighted(mejorada, 0.6, frangi_img, 0.4, 0)
                pasos['frangi'] = frangi_img
            except Exception:
                pass

        suavizada = cv2.bilateralFilter(
            mejorada,
            self.params.get('bilateral_d', 9),
            self.params.get('bilateral_sigma_color', 75),
            self.params.get('bilateral_sigma_space', 75))
        pasos['suavizado'] = suavizada

        if self.params.get('usar_multiescala', True):
            umbralizada = self._umbralizar_multiescala(suavizada)
        else:
            bs = self.params.get('block_size', 23)
            if bs % 2 == 0:
                bs += 1
            umbralizada = cv2.adaptiveThreshold(
                suavizada, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                cv2.THRESH_BINARY_INV, bs, self.params.get('C_umbral', 10))
        pasos['umbralizada'] = umbralizada

        if mascara_roi is not None:
            umbralizada = cv2.bitwise_and(umbralizada, umbralizada, mask=mascara_roi)

        k_ap = cv2.getStructuringElement(
            cv2.MORPH_ELLIPSE, (self.params.get('kernel_apertura', 3),) * 2)
        abierta = cv2.morphologyEx(umbralizada, cv2.MORPH_OPEN, k_ap)
        k_ci = cv2.getStructuringElement(
            cv2.MORPH_ELLIPSE, (self.params.get('kernel_cierre', 6),) * 2)
        cerrada = cv2.morphologyEx(abierta, cv2.MORPH_CLOSE, k_ci,
                                    iterations=self.params.get('iteraciones_cierre', 2))
        pasos['morfologia'] = cerrada

        if SKIMAGE_OK:
            limpia = remove_small_objects(
                cerrada.astype(bool),
                min_size=self.params.get('min_area_objeto', 100)
            ).astype(np.uint8) * 255
        else:
            limpia = cerrada
        pasos['limpia'] = limpia

        return limpia, pasos

    def _umbralizar_multiescala(self, gris):
        escalas = [
            {'block_size': 15, 'C': 8, 'peso': 0.3},
            {'block_size': 23, 'C': 10, 'peso': 0.4},
            {'block_size': 33, 'C': 12, 'peso': 0.3},
        ]
        resultados_e = []
        for esc in escalas:
            bs = esc['block_size']
            if bs % 2 == 0:
                bs += 1
            u = cv2.adaptiveThreshold(gris, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                       cv2.THRESH_BINARY_INV, bs, esc['C'])
            resultados_e.append((u * esc['peso']).astype(np.float32))
        combinado = np.clip(np.sum(resultados_e, axis=0), 0, 255).astype(np.uint8)
        _, final = cv2.threshold(combinado, 127, 255, cv2.THRESH_BINARY)
        return final

    def _refinar_esqueleto(self, esqueleto):
        num_labels, labels = cv2.connectedComponents(esqueleto)
        refinado = np.zeros_like(esqueleto)
        min_long = self.params.get('min_longitud_rama', 30)
        for label in range(1, num_labels):
            comp = (labels == label).astype(np.uint8) * 255
            if np.sum(comp == 255) >= min_long:
                refinado = cv2.bitwise_or(refinado, comp)
        return refinado

    def _cerrar_gaps(self, esqueleto):
        gap = self.params.get('max_gap_cierre', 20)
        kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (gap, gap))
        cerrado = cv2.morphologyEx(esqueleto, cv2.MORPH_CLOSE, kernel, iterations=2)
        if SKIMAGE_OK:
            final = skeletonize(cerrado.astype(bool))
            return (final * 255).astype(np.uint8)
        return cerrado

    def _detectar_poligonos(self, esqueleto):
        kernel = np.ones((3, 3), np.uint8)
        esq_dil = cv2.dilate(esqueleto, kernel, iterations=1)
        contornos, _ = cv2.findContours(esq_dil, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

        poligonos = []
        min_area = self.params.get('min_area_poligono', 300)
        min_circ = self.params.get('min_circularidad', 0.08)
        min_vert = self.params.get('min_vertices', 4)
        max_vert = self.params.get('max_vertices', 25)

        for cnt in contornos:
            area = cv2.contourArea(cnt)
            if area < min_area:
                continue
            per = cv2.arcLength(cnt, True)
            if per == 0:
                continue
            circ = 4 * np.pi * area / (per ** 2)
            if circ < min_circ:
                continue
            approx = cv2.approxPolyDP(cnt, 3, True)
            nv = len(approx)
            if nv < min_vert or nv > max_vert:
                continue
            M = cv2.moments(cnt)
            if M["m00"] == 0:
                continue
            cx = int(M["m10"] / M["m00"])
            cy = int(M["m01"] / M["m00"])
            poligonos.append({
                'contorno': approx,
                'contorno_original': cnt,
                'area': area,
                'perimetro': per,
                'circularidad': circ,
                'vertices': nv,
                'centroide': (cx, cy),
            })
        return poligonos

    def _calcular_circulo_inscrito(self, contorno):
        x, y, w, h = cv2.boundingRect(contorno)
        margen = 10
        mascara = np.zeros((h + 2 * margen, w + 2 * margen), dtype=np.uint8)
        cnt_local = contorno - [x - margen, y - margen]
        cv2.fillPoly(mascara, [cnt_local], 255)
        dist = cv2.distanceTransform(mascara, cv2.DIST_L2, 5)
        _, radio_max, _, centro_local = cv2.minMaxLoc(dist)
        cx_c = int(centro_local[0] + x - margen)
        cy_c = int(centro_local[1] + y - margen)
        radio = int(radio_max * 0.85)
        min_r = self.params.get('min_radio_circulo', 8)
        if radio < min_r:
            return None
        return {'centro': (cx_c, cy_c), 'radio': radio, 'area_circulo': np.pi * radio ** 2}

    @staticmethod
    def _clasificar_severidad(diametro_promedio_mm):
        """
        Severidad VIZIR para Piel de Cocodrilo:
            1: diÃ¡metro promedio > 500 mm
            2: 200 mm â‰¤ diÃ¡metro â‰¤ 500 mm
            3: diÃ¡metro < 200 mm
        """
        if diametro_promedio_mm is None or diametro_promedio_mm == 0:
            return 2  # Sin cÃ­rculos detectados = moderado
        if diametro_promedio_mm > 500:
            return 1
        elif diametro_promedio_mm >= 200:
            return 2
        else:
            return 3

    @staticmethod
    def dibujar(imagen, resultados, mostrar_mallas=True, mostrar_etiquetas=True, mostrar_numeros=True, escala_texto=1.0, mostrar_circulos=True, mostrar_poligonos=True, occupied_labels=None):
        vis = imagen.copy()
        colores_sev = {1: (0, 255, 0), 2: (0, 165, 255), 3: (0, 0, 255)}
        etiquetas_ocupadas = occupied_labels if occupied_labels is not None else []
        tipo_default = "FISURA DE PIEL DE COCODRILO"

        for r in resultados:
            sev = r['severidad']
            color_sev = colores_sev.get(sev, (255, 255, 255))

            # Contorno ROI
            if 'mascara_roi' in r:
                cnts, _ = cv2.findContours(
                    r['mascara_roi'], cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                cv2.drawContours(vis, cnts, -1, (255, 100, 0), 2)

            # Esqueleto (malla)
            if mostrar_mallas and 'esqueleto' in r:
                k = np.ones((2, 2), np.uint8)
                esq_vis = cv2.dilate(r['esqueleto'], k, iterations=1)
                vis[esq_vis == 255] = (0, 0, 255)

            # PolÃ­gonos cerrados
            if mostrar_mallas and mostrar_poligonos:
                for pol in r.get('poligonos', []):
                    cv2.drawContours(vis, [pol['contorno']], -1, (0, 255, 255), 2)

            # CÃ­rculos inscritos
            if mostrar_circulos:
                for circ in r.get('circulos', []):
                    cv2.circle(vis, circ['centro'], circ['radio'], (255, 0, 255), 2)
                    cv2.circle(vis, circ['centro'], 3, (0, 255, 0), -1)

            # Etiqueta
            if mostrar_etiquetas:
                cx, cy = r['ubicacion_x'], r['ubicacion_y']
                d_mm = r.get('diametro_promedio_mm', 0)
                tipo_label = str(r.get('tipo') or tipo_default)
                titulo = f"{tipo_label} [Sev:{sev}]"
                detalle = ""
                if mostrar_numeros:
                    detalle = f"D={d_mm:.0f}mm A={r['area_m2']:.2f}m2"
                texto_visible_dos_lineas(
                    vis,
                    titulo,
                    detalle,
                    (cx - 100, cy - 25),
                    escala_titulo=0.45,
                    escala_detalle=0.42,
                    color=color_sev,
                    grosor_titulo=2,
                    grosor_detalle=1,
                    factor=escala_texto,
                )

        return vis


# =============================================================================
# EXPORTADOR EXCEL - FORMATO VIZIR
# =============================================================================


# =============================================================================
# INTEGRACION AUTONOMA PCI + MTC + VIZIR
# =============================================================================

ProcesadorOjoDePescadoVIZIR = ProcesadorOjoDePescado
ProcesadorFisurasVIZIR = ProcesadorFisuras
ProcesadorBachesYParchesVIZIR = ProcesadorBachesYParches
ProcesadorPielCocodriloVIZIR = ProcesadorPielCocodrilo


def _extraer_geometria_piel_compartida(params, mascara, calibrador, confianza, imagen_original):
    base = ProcesadorPielCocodriloPCI(params)
    resultados = []
    mask_bin_full = (mascara > 0.5).astype(np.uint8) * 255

    h, w = imagen_original.shape[:2]
    if mask_bin_full.shape[:2] != (h, w):
        mask_bin_full = cv2.resize(mask_bin_full, (w, h))

    contornos, _ = cv2.findContours(mask_bin_full, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    for i, cnt in enumerate(contornos):
        area_contorno = cv2.contourArea(cnt)
        if area_contorno < 50:
            continue

        x, y, w_box, h_box = cv2.boundingRect(cnt)
        margen = 30
        x1 = max(0, x - margen)
        y1 = max(0, y - margen)
        x2 = min(w, x + w_box + margen)
        y2 = min(h, y + h_box + margen)

        mask_r_crop = np.zeros((y2 - y1, x2 - x1), dtype=np.uint8)
        cnt_crop = cnt.copy() - [x1, y1]
        cv2.fillPoly(mask_r_crop, [cnt_crop], 255)

        roi_crop = imagen_original[y1:y2, x1:x2]
        roi_crop = cv2.bitwise_and(roi_crop, roi_crop, mask=mask_r_crop)
        gris_crop = cv2.cvtColor(roi_crop, cv2.COLOR_BGR2GRAY)
        limpia_crop, pasos_crop = base._pipeline_imagen(gris_crop, mask_r_crop)

        if SKIMAGE_OK:
            esq_crop = skeletonize(limpia_crop.astype(bool))
            esq_crop = (esq_crop * 255).astype(np.uint8)
            if params.get("usar_refinamiento", True):
                esq_crop = base._refinar_esqueleto(esq_crop)
                esq_crop = base._cerrar_gaps(esq_crop)
        else:
            esq_crop = cv2.ximgproc.thinning(limpia_crop) if hasattr(cv2, "ximgproc") else limpia_crop

        mask_r = np.zeros((h, w), dtype=np.uint8)
        cv2.fillPoly(mask_r, [cnt], 255)
        area_px = int(np.sum(mask_r > 0))

        limpia = np.zeros((h, w), dtype=np.uint8)
        limpia[y1:y2, x1:x2] = limpia_crop

        esq = np.zeros((h, w), dtype=np.uint8)
        esq[y1:y2, x1:x2] = esq_crop

        pasos = {}
        for key, p_crop in pasos_crop.items():
            p_full = np.zeros((h, w), dtype=np.uint8)
            p_full[y1:y2, x1:x2] = p_crop
            pasos[key] = p_full

        poligonos = base._detectar_poligonos(esq)
        circulos = []
        for pol in poligonos:
            circulo = base._calcular_circulo_inscrito(pol["contorno_original"])
            if circulo:
                circulos.append(circulo)

        if circulos:
            diametros_px = [2 * c["radio"] for c in circulos]
            diametros_mm = [
                float(calibrador.px_a_mm(diametro_px) or 0)
                for diametro_px in diametros_px
            ]
            diametro_promedio_px = float(np.mean(diametros_px))
            diametro_min_px = float(np.min(diametros_px))
        else:
            diametros_mm = []
            diametro_promedio_px = 0.0
            diametro_min_px = 0.0

        diametro_promedio_mm = calibrador.px_a_mm(diametro_promedio_px) if diametro_promedio_px else 0
        diametro_promedio_m = calibrador.px_a_m(diametro_promedio_px) if diametro_promedio_px else 0
        diametro_min_mm = calibrador.px_a_mm(diametro_min_px) if diametro_min_px else 0
        diametro_min_m = calibrador.px_a_m(diametro_min_px) if diametro_min_px else 0

        area_m2 = calibrador.area_px_a_m2(area_px)
        cx, cy = 0, 0
        momentos = cv2.moments(cnt)
        if momentos["m00"] > 0:
            cx = int(momentos["m10"] / momentos["m00"])
            cy = int(momentos["m01"] / momentos["m00"])

        resultados.append({
            "id": i + 1,
            "confianza": confianza,
            "espesor_px": diametro_promedio_px,
            "espesor_mm": diametro_promedio_mm if diametro_promedio_mm else 0,
            "diametro_px": diametro_promedio_px,
            "diametro_mm": diametro_promedio_mm if diametro_promedio_mm else 0,
            "diametro_promedio_px": diametro_promedio_px,
            "diametro_promedio_mm": diametro_promedio_mm if diametro_promedio_mm else 0,
            "diametro_promedio_m": diametro_promedio_m if diametro_promedio_m else 0,
            "diametro_min_px": diametro_min_px,
            "diametro_min_mm": diametro_min_mm if diametro_min_mm else 0,
            "diametro_min_m": diametro_min_m if diametro_min_m else 0,
            "diametros_celdas_px": [float(d) for d in diametros_px] if circulos else [],
            "diametros_celdas_mm": diametros_mm,
            "area_px": area_px,
            "area_m2": area_m2 if area_m2 else 0,
            "ubicacion_x": cx,
            "ubicacion_y": cy,
            "poligonos": poligonos,
            "circulos": circulos,
            "esqueleto": esq,
            "mascara_roi": mask_r,
            "unidad": "m2",
            "n_celdas": len(circulos),
            "pasos": pasos,
            "limpia": limpia,
        })

    return resultados


def _procesar_piel_mtc_compartido(self, mascara, calibrador, confianza, imagen_original):
    resultados = []
    for base_res in _extraer_geometria_piel_compartida(self.params, mascara, calibrador, confianza, imagen_original):
        diametro_promedio_m = base_res.get("diametro_promedio_m", 0)
        diametro_promedio_mm = base_res.get("diametro_promedio_mm", 0)
        resultados.append({
            **base_res,
            "tipo": self.NOMBRE_MTC,
            "severidad": self._clasificar_severidad(diametro_promedio_m),
            "espesor_px": base_res.get("diametro_promedio_px", 0),
            "espesor_mm": diametro_promedio_mm or 0,
            "diametro_px": base_res.get("diametro_promedio_px", 0),
            "diametro_mm": diametro_promedio_mm or 0,
            "longitud_px": 0,
            "longitud_m": 0,
            "ancho_px": 0,
            "ancho_m": 0,
        })
    return resultados


def _procesar_piel_vizir_compartido(self, mascara, calibrador, confianza, imagen_original):
    resultados = []
    for base_res in _extraer_geometria_piel_compartida(self.params, mascara, calibrador, confianza, imagen_original):
        diametro_promedio_mm = base_res.get("diametro_promedio_mm", 0)
        resultados.append({
            **base_res,
            "tipo": self.NOMBRE_VIZIR,
            "severidad": self._clasificar_severidad(diametro_promedio_mm),
            "espesor_px": base_res.get("diametro_promedio_px", 0),
            "espesor_mm": diametro_promedio_mm or 0,
            "diametro_px": base_res.get("diametro_promedio_px", 0),
            "diametro_mm": diametro_promedio_mm or 0,
            "diametro_promedio_mm": diametro_promedio_mm or 0,
            "longitud_px": 0,
            "longitud_m": 0,
        })
    return resultados


ProcesadorPielCocodriloMTC.procesar = _procesar_piel_mtc_compartido
ProcesadorPielCocodriloVIZIR.procesar = _procesar_piel_vizir_compartido

def _calibrador_set_eje_via(self, valor_a, valor_b=None):
    if valor_b is not None and isinstance(valor_a, (tuple, list)) and isinstance(valor_b, (tuple, list)):
        self.calibrar_eje(tuple(valor_a), tuple(valor_b))
        return
    self.angulo_eje_via = 90.0 if valor_a is None else float(valor_a)


def _calibrador_get_angulo_eje(self):
    return 90.0 if getattr(self, 'angulo_eje_via', None) is None else float(self.angulo_eje_via)


Calibrador.set_eje_via = _calibrador_set_eje_via
Calibrador.get_angulo_eje = _calibrador_get_angulo_eje


class MotorIntegradoTresMetodos:
    def __init__(self, config=None):
        self.config = copy.deepcopy(CONFIG_DEFAULT)
        if config:
            self.config.update(config)
        self.config.setdefault('filtrar_baches', True)
        self.config.setdefault('filtrar_grietas', True)
        self.config.setdefault('filtrar_parches', True)
        self.config.setdefault('filtrar_piel', True)
        self.config.setdefault('merge_fisuras_px', 30)
        self.config.setdefault('usar_refinamiento', True)
        self.modelo = None
        self.modelo_cargado = False
        self.calibrador = Calibrador(self.config['ancho_via_real_m'])
        self.todos_resultados = {}

    def cargar_modelo(self, ruta):
        try:
            if not YOLO_OK:
                return False, 'ultralytics no instalado. pip install ultralytics'
            self.modelo = YOLO(ruta)
            self.modelo_cargado = True
            self.config['ruta_modelo'] = ruta
            return True, f"Modelo cargado: {Path(ruta).name}"
        except Exception as exc:
            self.modelo_cargado = False
            return False, f"Error cargando modelo: {exc}"

    def calibrar_imagen(self, imagen_cv2, usar_gui=False, parent=None):
        self.calibrador.ancho_via_real_m = self.config.get('ancho_via_real_m', 6.5)
        if usar_gui:
            ventana = VentanaCalibracionIntegrada(
                imagen_cv2,
                self.config['ancho_via_real_m'],
                parent=parent,
                borde_interno_m=self.config.get('borde_interno_m', 0.30),
                borde_externo_m=self.config.get('borde_externo_m', 0.30),
                usar_borde_berma_pci=self.config.get('usar_borde_berma_pci', True),
            )
            longitud, angulo_eje, pt_ini, pt_fin = ventana.ejecutar()
            if longitud:
                self.calibrador.calibrar_con_linea(longitud)
            else:
                self.calibrador.calibrar_con_ancho_imagen(imagen_cv2.shape[1])
            self.config['borde_interno_m'] = float(getattr(ventana, 'borde_interno_m', self.config.get('borde_interno_m', 0.30)))
            self.config['borde_externo_m'] = float(getattr(ventana, 'borde_externo_m', self.config.get('borde_externo_m', 0.30)))
            self.config['usar_borde_berma_pci'] = bool(getattr(ventana, 'usar_borde_berma_pci', self.config.get('usar_borde_berma_pci', True)))
            self.calibrador.set_linea_ancho(
                getattr(ventana, 'punto_inicio_ancho', None),
                getattr(ventana, 'punto_fin_ancho', None),
            )
            if pt_ini and pt_fin:
                self.calibrador.calibrar_eje(pt_ini, pt_fin)
            else:
                self.calibrador.set_eje_via(angulo_eje)
            return True
        alto, ancho = imagen_cv2.shape[:2]
        self.calibrador.calibrar_con_ancho_imagen(ancho)
        y_centro = max(0, min(alto - 1, alto // 2))
        self.calibrador.set_linea_ancho((0, y_centro), (max(ancho - 1, 0), y_centro))
        self.calibrador.set_eje_via(90.0)
        return True

    def procesar_imagen(self, ruta_imagen, callback_log=None, calibrar_gui=False, parent=None):
        def log(msg):
            if callback_log:
                callback_log(msg)

        imagen = cv2.imread(str(ruta_imagen))
        if imagen is None:
            log(f"ERROR: No se pudo cargar {ruta_imagen}")
            return None

        nombre = Path(ruta_imagen).name
        alto, ancho = imagen.shape[:2]
        log(f"Procesando integral: {nombre} ({ancho}x{alto})")

        if calibrar_gui:
            self.calibrar_imagen(imagen, usar_gui=True, parent=parent)
        elif self.calibrador.px_por_mm is None:
            self.calibrar_imagen(imagen, usar_gui=False)

        log(
            f"Calibracion activa: {self.calibrador.px_por_mm:.4f} px/mm | "
            f"Eje: {self.calibrador.get_angulo_eje():.1f} grados"
        )

        mascaras_raw = self._inferir_mascaras(imagen, log)
        resultados = {
            'PCI': self._procesar_pci_desde_mascaras(imagen, mascaras_raw),
            'MTC': self._procesar_mtc_desde_mascaras(imagen, mascaras_raw),
            'VIZIR': self._procesar_vizir_desde_mascaras(imagen, mascaras_raw),
        }

        self.todos_resultados[nombre] = {metodo: info['fallas'] for metodo, info in resultados.items()}

        for metodo, info in resultados.items():
            log(f"{metodo}: {len(info['fallas'])} falla(s) validas")

        return {
            'imagen': imagen,
            'nombre': nombre,
            'dimensiones': (ancho, alto),
            'calibracion_px_mm': self.calibrador.px_por_mm,
            'ancho_via_real_m': self.calibrador.ancho_via_real_m,
            'angulo_eje': self.calibrador.get_angulo_eje(),
            'raw_masks': mascaras_raw,
            'metodos': resultados,
        }

    def _inferir_mascaras(self, imagen, log):
        if not self.modelo_cargado or self.modelo is None:
            raise RuntimeError('Debe cargar el modelo YOLO antes de procesar.')
        alto, ancho = imagen.shape[:2]
        resultado = self.modelo(
            imagen,
            conf=self.config.get('confianza_min', 0.1),
            iou=self.config.get('iou_threshold', 0.45),
            verbose=False,
        )[0]
        if resultado.masks is None:
            log('YOLO: sin detecciones')
            return {}
        mascaras = {}
        for idx in range(len(resultado.boxes)):
            cls_id = int(resultado.boxes.cls[idx])
            conf = float(resultado.boxes.conf[idx])
            mask_raw = resultado.masks.data[idx].cpu().numpy()
            mask_resized = cv2.resize(mask_raw, (ancho, alto))
            mascaras.setdefault(cls_id, []).append((mask_resized, conf))
        for cls_id, lista in mascaras.items():
            clases_cfg = self.config.get("clases", {}) or {}
            nombre_clase = clases_cfg.get(cls_id)
            if nombre_clase is None:
                nombre_clase = clases_cfg.get(str(cls_id), f"CLASE_{cls_id}")
            log(f"YOLO {nombre_clase}: {len(lista)} deteccion(es)")
        return mascaras

    def _clone_masks(self, mascaras_por_clase):
        copia = {}
        for cls_id, lista in mascaras_por_clase.items():
            copia[cls_id] = [(mask.copy(), float(conf)) for mask, conf in lista]
        return copia

    def _separar_componentes_mascara(self, mask, min_pixels=30):
        mask_bin = (mask > 0.5).astype(np.uint8)
        num_labels, labels, stats, _ = cv2.connectedComponentsWithStats(mask_bin, 8)
        if num_labels <= 2:
            return [mask]
        componentes = []
        umbral = max(1, int(min_pixels))
        for label_id in range(1, num_labels):
            area = int(stats[label_id, cv2.CC_STAT_AREA])
            if area < umbral:
                continue
            comp_mask = (labels == label_id).astype(np.float32)
            componentes.append(comp_mask)
        return componentes or [mask]

    def _build_config(self):
        return copy.deepcopy(self.config)

    def _preclasificar_fisuras_por_orientacion(self, mascaras_raw, angle_fn=None, classifier_fn=None):
        mascaras_por_clase = self._clone_masks(mascaras_raw)
        if 1 not in mascaras_por_clase:
            return mascaras_por_clase

        lista_fisuras = mascaras_por_clase.pop(1)
        long_masks = []
        trans_masks = []
        if angle_fn is None:
            angle_fn = ProcesadorGrietasMTC._calcular_angulo
        if classifier_fn is None:
            classifier_fn = self.calibrador.es_longitudinal

        for mask_r, conf in lista_fisuras:
            mask_bin = (mask_r > 0.5).astype(np.uint8)
            if np.sum(mask_bin) < 5:
                continue
            try:
                angulo = float(angle_fn(mask_bin))
            except Exception:
                angulo = 0.0
            es_long = bool(classifier_fn(angulo))
            if es_long:
                long_masks.append((mask_r, conf))
            else:
                trans_masks.append((mask_r, conf))

        if long_masks:
            mascaras_por_clase['1_long'] = long_masks
        if trans_masks:
            mascaras_por_clase['1_trans'] = trans_masks
        return mascaras_por_clase

    def _clave_grupo_grieta(self, metodo, falla):
        if metodo == "PCI":
            es_long = falla.get("es_longitudinal")
            if es_long is None:
                angulo = falla.get("angulo_grieta", falla.get("angulo", 0.0))
                try:
                    es_long = self.calibrador.es_longitudinal(float(angulo))
                except Exception:
                    es_long = False
            return (metodo, bool(es_long))
        return (metodo, str(falla.get("tipo", "FISURA")))

    def _reprocesar_grieta_fusionada(self, metodo, mask_f, conf_f, imagen, cfg):
        if metodo == "PCI":
            return ProcesadorGrietasPCI.procesar(
                mask_f,
                self.calibrador,
                conf_f,
                imagen,
                merge_dist_px=cfg.get('merge_fisuras_px', 30),
            )
        if metodo == "MTC":
            return ProcesadorGrietasMTC.procesar(
                mask_f,
                self.calibrador,
                conf_f,
                imagen,
                merge_dist_px=cfg.get('merge_fisuras_px', 30),
            )
        return ProcesadorFisurasVIZIR.procesar(
            mask_f,
            self.calibrador,
            conf_f,
            imagen,
            merge_dist_px=cfg.get('merge_fisuras_px', 30),
        )

    def _fusionar_fallas_grieta_solapadas(self, fallas, metodo, imagen, cfg):
        if not fallas:
            return fallas

        otras = []
        grupos = {}
        for falla in fallas:
            if categorizar_tipo_falla(falla.get('tipo', '')) != 'grieta':
                otras.append(falla)
                continue
            mask = falla.get("mask")
            if mask is None:
                otras.append(falla)
                continue
            grupos.setdefault(self._clave_grupo_grieta(metodo, falla), []).append(falla)

        grietas_finales = []
        for _, grupo in grupos.items():
            if len(grupo) <= 1:
                grietas_finales.extend(grupo)
                continue

            lista_masks = []
            for falla in grupo:
                mask = np.asarray(falla.get("mask"), dtype=np.float32)
                if np.sum(mask > 0) < 1:
                    continue
                lista_masks.append((mask, float(falla.get("confianza", 0.0) or 0.0)))

            if len(lista_masks) <= 1:
                grietas_finales.extend(grupo)
                continue

            fusionadas = FusionadorMascaras.fusionar_por_clase(
                {0: lista_masks},
                iou_threshold=0.02,
                distancia_max_px=0,
            ).get(0, lista_masks)

            grupo_reprocesado = []
            for mask_f, conf_f in fusionadas:
                repro = self._reprocesar_grieta_fusionada(metodo, mask_f, conf_f, imagen, cfg)
                if repro:
                    grupo_reprocesado.extend(repro)

            if grupo_reprocesado:
                grietas_finales.extend(grupo_reprocesado)
            else:
                grietas_finales.extend(grupo)

        return otras + grietas_finales

    def _renumerar_fallas(self, fallas):
        conteo = {}
        for falla in fallas:
            tipo = falla.get('tipo', 'FALLA')
            conteo[tipo] = conteo.get(tipo, 0) + 1
            falla['id'] = conteo[tipo]
        return conteo

    def _resumen_severidad(self, fallas):
        resumen = {}
        for falla in fallas:
            sev = str(falla.get('severidad', '-'))
            resumen[sev] = resumen.get(sev, 0) + 1
        return resumen

    def _punto_representativo_falla(self, falla):
        mask = falla.get("mask")
        if mask is not None:
            try:
                mask_bin = (np.asarray(mask) > 0).astype(np.uint8)
                momentos = cv2.moments(mask_bin)
                if momentos["m00"] > 0:
                    return np.array(
                        [
                            momentos["m10"] / momentos["m00"],
                            momentos["m01"] / momentos["m00"],
                        ],
                        dtype=np.float32,
                    )
            except Exception:
                pass
        return np.array(
            [
                float(falla.get("ubicacion_x", 0) or 0),
                float(falla.get("ubicacion_y", 0) or 0),
            ],
            dtype=np.float32,
        )

    def _guardar_campos_base(self, falla, campos):
        for campo in campos:
            key = f"{campo}_base"
            if key not in falla:
                falla[key] = copy.deepcopy(falla.get(campo))

    def _restaurar_campos_base(self, falla, campos):
        for campo in campos:
            key = f"{campo}_base"
            if key in falla:
                falla[campo] = copy.deepcopy(falla.get(key))

    def _contexto_borde_berma(self, cfg):
        if not bool(cfg.get("usar_borde_berma_pci", True)):
            return None
        pt_a = getattr(self.calibrador, "punto_ancho_inicio", None)
        pt_b = getattr(self.calibrador, "punto_ancho_fin", None)
        px_por_mm = float(getattr(self.calibrador, "px_por_mm", 0.0) or 0.0)
        if pt_a is None or pt_b is None or px_por_mm <= 0:
            return None

        origen = np.array(pt_a, dtype=np.float32)
        destino = np.array(pt_b, dtype=np.float32)
        vector = destino - origen
        longitud_px = float(np.linalg.norm(vector))
        if longitud_px <= 1e-6:
            return None

        return {
            "origen": origen,
            "unitario": vector / longitud_px,
            "longitud_px": longitud_px,
            "borde_interno_px": float(cfg.get("borde_interno_m", 0.30) or 0.30) * 1000.0 * px_por_mm,
            "borde_externo_px": float(cfg.get("borde_externo_m", 0.30) or 0.30) * 1000.0 * px_por_mm,
        }

    def _proyeccion_falla_en_ancho(self, falla, ctx):
        punto = self._punto_representativo_falla(falla)
        return float(np.dot(punto - ctx["origen"], ctx["unitario"]))

    def _esta_en_berma(self, proyeccion, ctx):
        return (
            -ctx["borde_externo_px"] <= proyeccion <= 0
            or ctx["longitud_px"] <= proyeccion <= (ctx["longitud_px"] + ctx["borde_externo_px"])
        )

    def _esta_en_borde_berma(self, proyeccion, ctx):
        return (
            -ctx["borde_externo_px"] <= proyeccion <= ctx["borde_interno_px"]
            or (ctx["longitud_px"] - ctx["borde_interno_px"]) <= proyeccion <= (ctx["longitud_px"] + ctx["borde_externo_px"])
        )

    def _clasificar_grietas_borde_pci(self, fallas, cfg):
        ctx = self._contexto_borde_berma(cfg)
        if ctx is None:
            for falla in fallas:
                if falla.get("tipo") == "GRIETA DE BORDE":
                    falla["tipo"] = falla.get("tipo_base", ProcesadorGrietasPCI.NOMBRE_PCI)
                falla["es_borde"] = False
            return fallas

        for falla in fallas:
            if categorizar_tipo_falla(falla.get("tipo", "")) != "grieta":
                continue
            if not bool(falla.get("es_longitudinal", False)):
                if falla.get("tipo") == "GRIETA DE BORDE":
                    falla["tipo"] = falla.get("tipo_base", ProcesadorGrietasPCI.NOMBRE_PCI)
                falla["es_borde"] = False
                continue

            proyeccion = self._proyeccion_falla_en_ancho(falla, ctx)
            es_borde = self._esta_en_borde_berma(proyeccion, ctx)
            falla["es_borde"] = es_borde
            if es_borde:
                falla["tipo_base"] = falla.get("tipo", ProcesadorGrietasPCI.NOMBRE_PCI)
                falla["tipo"] = "GRIETA DE BORDE"
            elif falla.get("tipo") == "GRIETA DE BORDE":
                falla["tipo"] = falla.get("tipo_base", ProcesadorGrietasPCI.NOMBRE_PCI)
        return fallas

    def _clasificar_danos_puntuales_mtc(self, fallas, cfg):
        ctx = self._contexto_borde_berma(cfg)
        for falla in fallas:
            tipo_actual = str(falla.get("tipo", "") or "")
            es_candidato = ("BACHE" in tipo_actual.upper() or "HUECO" in tipo_actual.upper() or tipo_actual == "DAÑOS PUNTUALES")
            if not es_candidato:
                continue

            if ctx is None:
                if tipo_actual == "DAÑOS PUNTUALES":
                    self._restaurar_campos_base(falla, ["tipo", "severidad", "unidad", "diametro_px", "diametro_mm", "espesor_px", "espesor_mm"])
                falla["es_berma"] = False
                continue

            proyeccion = self._proyeccion_falla_en_ancho(falla, ctx)
            en_berma = self._esta_en_berma(proyeccion, ctx)
            falla["es_berma"] = bool(en_berma)
            if en_berma:
                self._guardar_campos_base(falla, ["tipo", "severidad", "unidad", "diametro_px", "diametro_mm", "espesor_px", "espesor_mm"])
                falla["tipo"] = "DAÑOS PUNTUALES"
                falla["severidad"] = 1
                falla["unidad"] = "m2"
                falla["diametro_px"] = 0
                falla["diametro_mm"] = 0
                falla["espesor_px"] = 0
                falla["espesor_mm"] = 0
            elif tipo_actual == "DAÑOS PUNTUALES":
                self._restaurar_campos_base(falla, ["tipo", "severidad", "unidad", "diametro_px", "diametro_mm", "espesor_px", "espesor_mm"])
        return fallas

    def _clasificar_fisuras_borde_vizir(self, fallas, cfg):
        ctx = self._contexto_borde_berma(cfg)
        for falla in fallas:
            tipo_actual = str(falla.get("tipo", "") or "")
            es_fisura = "FISURA" in tipo_actual.upper() and "COCODRILO" not in tipo_actual.upper()
            if not es_fisura:
                continue

            if ctx is None or not bool(falla.get("es_longitudinal", False)):
                if tipo_actual == "FISURAS DE BORDE":
                    self._restaurar_campos_base(falla, ["tipo", "unidad", "area_m2"])
                falla["es_borde"] = False
                continue

            proyeccion = self._proyeccion_falla_en_ancho(falla, ctx)
            es_borde = self._esta_en_borde_berma(proyeccion, ctx)
            falla["es_borde"] = bool(es_borde)
            if es_borde:
                self._guardar_campos_base(falla, ["tipo", "unidad", "area_m2"])
                falla["tipo"] = "FISURAS DE BORDE"
                falla["unidad"] = "m"
            elif tipo_actual == "FISURAS DE BORDE":
                self._restaurar_campos_base(falla, ["tipo", "unidad", "area_m2"])
        return fallas

    def _con_procesador_piel_pci(self, funcion):
        procesador_actual = globals().get('ProcesadorPielCocodrilo')
        globals()['ProcesadorPielCocodrilo'] = ProcesadorPielCocodriloPCI
        try:
            return funcion()
        finally:
            globals()['ProcesadorPielCocodrilo'] = procesador_actual

    def _procesar_pci_desde_mascaras(self, imagen, mascaras_raw):
        cfg = self._build_config()
        mascaras_por_clase = self._preclasificar_fisuras_por_orientacion(
            mascaras_raw,
            angle_fn=ProcesadorGrietasMTC._calcular_angulo,
            classifier_fn=self.calibrador.es_longitudinal,
        )
        fusionadas = FusionadorMascaras.fusionar_por_clase(
            mascaras_por_clase,
            iou_threshold=cfg.get('merge_iou_threshold', 0.10),
            distancia_max_px=cfg.get('merge_distancia_max_px', 50),
        )

        def _run():
            procesador_piel = ProcesadorPielCocodriloPCI(cfg)
            fallas_raw = []
            clases_fisura = {1, '1_long', '1_trans'}
            for cls_id, lista_fusionada in fusionadas.items():
                for mask_f, conf_f in lista_fusionada:
                    if cls_id == 0:
                        fallas_raw.extend(
                            ProcesadorBachesPCI.procesar(
                                mask_f,
                                self.calibrador,
                                conf_f,
                                imagen,
                                profundidad_asumida=cfg.get('profundidad_asumida_huecos', 'media'),
                            )
                        )
                    elif cls_id in clases_fisura:
                        mask_limpia = mask_f.copy()
                        for otro_cls, otra_lista in fusionadas.items():
                            if otro_cls in clases_fisura:
                                continue
                            for otra_mask, _ in otra_lista:
                                otra_bin = (otra_mask > 0.5).astype(np.uint8)
                                mask_limpia = mask_limpia * (1 - otra_bin)
                        if np.sum(mask_limpia > 0.5) > 0:
                            fallas_raw.extend(
                                ProcesadorGrietasPCI.procesar(
                                    mask_limpia,
                                    self.calibrador,
                                    conf_f,
                                    imagen,
                                    merge_dist_px=cfg.get('merge_fisuras_px', 30),
                                )
                            )
                    elif cls_id == 2:
                        fallas_raw.extend(ProcesadorParchesPCI.procesar(mask_f, self.calibrador, conf_f, imagen, cfg))
                    elif cls_id == 3:
                        fallas_raw.extend(procesador_piel.procesar(mask_f, self.calibrador, conf_f, imagen))
            return fallas_raw

        fallas_raw = self._con_procesador_piel_pci(_run)
        fallas_raw = self._fusionar_fallas_grieta_solapadas(fallas_raw, "PCI", imagen, cfg)
        fallas_raw = self._clasificar_grietas_borde_pci(fallas_raw, cfg)
        fallas = []
        for falla in fallas_raw:
            categoria = categorizar_tipo_falla(falla.get('tipo', ''))
            if categoria == 'hueco' and falla.get('diametro_mm', 0) < cfg.get('min_diametro_hueco_mm', 50.0):
                continue
            if categoria == 'grieta' and falla.get('longitud_m', 0) < cfg.get('min_longitud_grieta_m', 0.05):
                continue
            if categoria == 'parche' and falla.get('area_m2', 0) < cfg.get('min_area_parche_m2', 0.01):
                continue
            if categoria == 'piel' and falla.get('area_m2', 0) < cfg.get('min_area_piel_m2', 0.05):
                continue
            fallas.append(falla)
        self._renumerar_fallas(fallas)
        visual = self._con_procesador_piel_pci(lambda: self._dibujar_pci(imagen, fallas, cfg))
        return {'fallas': fallas, 'visual': visual, 'resumen_severidad': self._resumen_severidad(fallas)}

    def _procesar_mtc_desde_mascaras(self, imagen, mascaras_raw):
        cfg = self._build_config()
        mascaras_por_clase = self._preclasificar_fisuras_por_orientacion(
            mascaras_raw,
            angle_fn=ProcesadorGrietasMTC._calcular_angulo,
            classifier_fn=self.calibrador.es_longitudinal,
        )

        fusionadas = FusionadorMascaras.fusionar_por_clase(
            mascaras_por_clase,
            iou_threshold=cfg.get('merge_iou_threshold', 0.10),
            distancia_max_px=cfg.get('merge_distancia_max_px', 50),
        )

        procesador_piel = ProcesadorPielCocodriloMTC(cfg)
        fallas_raw = []
        clases_fisura = {1, '1_long', '1_trans'}
        for cls_id, lista_fusionada in fusionadas.items():
            for mask_f, conf_f in lista_fusionada:
                if cls_id == 0:
                    fallas_raw.extend(ProcesadorBachesMTC.procesar(mask_f, self.calibrador, conf_f, imagen))
                elif cls_id in clases_fisura:
                    mask_limpia = mask_f.copy()
                    for otro_cls, otra_lista in fusionadas.items():
                        if otro_cls in clases_fisura:
                            continue
                        for otra_mask, _ in otra_lista:
                            otra_bin = (otra_mask > 0.5).astype(np.uint8)
                            mask_limpia = mask_limpia * (1 - otra_bin)
                    if np.sum(mask_limpia > 0.5) <= 30:
                        continue
                    fallas_raw.extend(
                        ProcesadorGrietasMTC.procesar(
                            mask_limpia,
                            self.calibrador,
                            conf_f,
                            imagen,
                            merge_dist_px=cfg.get('merge_fisuras_px', 30),
                        )
                    )
                elif cls_id == 2:
                    fallas_raw.extend(ProcesadorParchesMTC.procesar(mask_f, self.calibrador, conf_f, imagen, cfg))
                elif cls_id == 3:
                    fallas_raw.extend(procesador_piel.procesar(mask_f, self.calibrador, conf_f, imagen))

        fallas_raw = self._fusionar_fallas_grieta_solapadas(fallas_raw, "MTC", imagen, cfg)
        fallas_raw = self._clasificar_danos_puntuales_mtc(fallas_raw, cfg)
        fallas = []
        for falla in fallas_raw:
            tipo = falla.get('tipo', '')
            skip = False
            if ('BACHE' in tipo or 'HUECO' in tipo) and cfg.get('filtrar_baches', True):
                skip = falla.get('diametro_mm', 0) < cfg.get('min_diametro_hueco_mm', 50)
            elif 'FISURA' in tipo and cfg.get('filtrar_grietas', True):
                skip = falla.get('longitud_m', 0) < cfg.get('min_longitud_grieta_m', 0.05)
            elif ('REPARACION' in tipo or 'PARCHADO' in tipo) and cfg.get('filtrar_parches', True):
                skip = falla.get('area_m2', 0) < cfg.get('min_area_parche_m2', 0.01)
            elif 'COCODRILO' in tipo and cfg.get('filtrar_piel', True):
                skip = falla.get('area_m2', 0) < cfg.get('min_area_piel_m2', 0.05)
            if not skip:
                fallas.append(falla)

        self._renumerar_fallas(fallas)
        visual = self._dibujar_mtc(imagen, fallas, cfg)
        return {'fallas': fallas, 'visual': visual, 'resumen_severidad': self._resumen_severidad(fallas)}

    def _procesar_vizir_desde_mascaras(self, imagen, mascaras_raw):
        cfg = self._build_config()
        mascaras_por_clase = self._preclasificar_fisuras_por_orientacion(
            mascaras_raw,
            angle_fn=ProcesadorFisurasVIZIR._calcular_angulo_grieta,
            classifier_fn=lambda ang: ProcesadorFisurasVIZIR._es_longitudinal(ang, self.calibrador.get_angulo_eje()),
        )
        fusionadas = FusionadorMascaras.fusionar_por_clase(
            mascaras_por_clase,
            iou_threshold=cfg.get('merge_iou_threshold', 0.10),
            distancia_max_px=cfg.get('merge_distancia_max_px', 50),
        )

        alto, ancho = imagen.shape[:2]
        mascara_cocodrilo_total = np.zeros((alto, ancho), dtype=np.uint8)
        for mask_f, _ in fusionadas.get(3, []):
            mascara_cocodrilo_total = np.maximum(mascara_cocodrilo_total, (mask_f > 0.5).astype(np.uint8))

        procesador_piel = ProcesadorPielCocodriloVIZIR(cfg)
        fallas_raw = []
        for cls_id, lista_fusionada in fusionadas.items():
            for mask_f, conf_f in lista_fusionada:
                if cls_id == 0:
                    fallas_raw.extend(ProcesadorOjoDePescadoVIZIR.procesar(mask_f, self.calibrador, conf_f, imagen))
                elif cls_id in (1, '1_long', '1_trans'):
                    mask_bin = (mask_f > 0.5).astype(np.uint8)
                    mask_bin_clean = mask_bin.copy()
                    mask_bin_clean[mascara_cocodrilo_total > 0] = 0
                    for other_cls in [0, 2]:
                        for other_mask, _ in fusionadas.get(other_cls, []):
                            mask_bin_clean[(other_mask > 0.5)] = 0
                    res_fisuras = []
                    if np.sum(mask_bin_clean) > 30:
                        res_fisuras = ProcesadorFisurasVIZIR.procesar(
                            mask_bin_clean.astype(np.float32),
                            self.calibrador,
                            conf_f,
                            imagen,
                            merge_dist_px=cfg.get('merge_fisuras_px', 30),
                        )
                    fallas_raw.extend(res_fisuras)
                elif cls_id == 2:
                    fallas_raw.extend(ProcesadorBachesYParchesVIZIR.procesar(mask_f, self.calibrador, conf_f, imagen, cfg))
                elif cls_id == 3:
                    componentes = self._separar_componentes_mascara(mask_f, min_pixels=cfg.get('min_area_objeto', 100))
                    for comp_mask in componentes:
                        fallas_raw.extend(procesador_piel.procesar(comp_mask, self.calibrador, conf_f, imagen))

        fallas_raw = self._fusionar_fallas_grieta_solapadas(fallas_raw, "VIZIR", imagen, cfg)
        fallas_raw = self._clasificar_fisuras_borde_vizir(fallas_raw, cfg)
        fallas = []
        for falla in fallas_raw:
            tipo = falla.get('tipo', '')
            skip = False
            if tipo == 'OJO DE PESCADO':
                skip = falla.get('diametro_mm', 0) < cfg.get('min_diametro_hueco_mm', 50.0)
            elif 'FISURA' in tipo and 'COCODRILO' not in tipo:
                skip = falla.get('longitud_m', 0) < cfg.get('min_longitud_grieta_m', 0.05)
            elif tipo == 'BACHES Y PARCHES':
                skip = falla.get('area_m2', 0) < cfg.get('min_area_parche_m2', 0.01)
            elif 'COCODRILO' in tipo:
                skip = falla.get('area_m2', 0) < cfg.get('min_area_piel_m2', 0.05)
            if not skip:
                fallas.append(falla)

        self._renumerar_fallas(fallas)
        visual = self._dibujar_vizir(imagen, fallas, cfg)
        return {'fallas': fallas, 'visual': visual, 'resumen_severidad': self._resumen_severidad(fallas)}

    def _dibujar_pci(self, imagen, fallas, cfg):
        vis = imagen.copy()
        ts = cfg.get('text_size', 0.45)
        sn = cfg.get('show_numeros', True)
        se = cfg.get('show_etiquetas', True)
        sm = cfg.get('show_mallas', True)
        sc = cfg.get('show_circulos', True)
        sp = cfg.get('show_poligonos', True)

        activas = [f for f in fallas if not f.get('excluida', False)]
        huecos = [f for f in activas if categorizar_tipo_falla(f.get('tipo')) == 'hueco']
        grietas = [f for f in activas if categorizar_tipo_falla(f.get('tipo')) == 'grieta']
        parches = [f for f in activas if categorizar_tipo_falla(f.get('tipo')) == 'parche']
        piel = [f for f in activas if categorizar_tipo_falla(f.get('tipo')) == 'piel']

        if huecos:
            vis = ProcesadorBachesPCI.dibujar(vis, huecos, text_size=ts, show_numeros=sn, show_etiquetas=se, show_circulos=sc)
        if grietas:
            vis = ProcesadorGrietasPCI.dibujar(vis, grietas, text_size=ts, show_numeros=sn, show_etiquetas=se)
        if parches:
            vis = ProcesadorParchesPCI.dibujar(vis, parches, text_size=ts, show_numeros=sn, show_etiquetas=se)
        if piel:
            vis = ProcesadorPielCocodriloPCI.dibujar(
                vis,
                piel,
                text_size=ts,
                show_numeros=sn,
                show_mallas=sm,
                show_circulos=sc,
                show_poligonos=sp,
                show_etiquetas=se,
            )

        y = 30
        cv2.putText(vis, 'ASTM D6433 | PCI', (10, y), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
        y += 25
        resumen = {}
        for falla in activas:
            tipo = falla['tipo']
            if tipo not in resumen:
                resumen[tipo] = {'total': 0, 'L': 0, 'M': 0, 'H': 0}
            resumen[tipo]['total'] += 1
            sev = severidad_ui(falla.get('severidad'))
            if sev in resumen[tipo]:
                resumen[tipo][sev] += 1
        for tipo, counts in resumen.items():
            txt = f"{tipo}: {counts['total']} (L:{counts['L']} M:{counts['M']} H:{counts['H']})"
            cv2.putText(vis, txt, (10, y), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (200, 200, 200), 1)
            y += 18
        return vis

    def _dibujar_mtc(self, imagen, fallas, cfg):
        vis = imagen.copy()
        et = 1.0
        mostrar_mallas = cfg.get('show_mallas', True)
        mostrar_etiquetas = cfg.get('show_etiquetas', True)
        mostrar_numeros = cfg.get('show_numeros', True)
        mostrar_circulos = cfg.get('show_circulos', True)
        mostrar_poligonos = cfg.get('show_poligonos', True)
        etiquetas_ocupadas = None
        if mostrar_etiquetas:
            lineas_panel = 1 + len({f.get('tipo', '') for f in fallas})
            alto_panel = 18 + lineas_panel * 22
            ancho_panel = min(vis.shape[1] - 6, 430)
            etiquetas_ocupadas = [(6, 6, ancho_panel, alto_panel)]

        ba = [f for f in fallas if 'BACHE' in f['tipo'] or 'HUECO' in f['tipo']]
        fi = [f for f in fallas if 'FISURA' in f['tipo']]
        pa = [f for f in fallas if 'REPARACION' in f['tipo'] or 'PARCHADO' in f['tipo']]
        pi = [f for f in fallas if 'COCODRILO' in f['tipo']]

        if ba:
            vis = ProcesadorBachesMTC.dibujar(vis, ba, mostrar_mallas=mostrar_mallas, mostrar_etiquetas=mostrar_etiquetas, mostrar_numeros=mostrar_numeros, escala_texto=et, mostrar_circulos=mostrar_circulos, occupied_labels=etiquetas_ocupadas)
        if fi:
            vis = ProcesadorGrietasMTC.dibujar(vis, fi, mostrar_etiquetas=mostrar_etiquetas, mostrar_numeros=mostrar_numeros, escala_texto=et, occupied_labels=etiquetas_ocupadas)
        if pa:
            vis = ProcesadorParchesMTC.dibujar(vis, pa, mostrar_mallas=mostrar_mallas, mostrar_etiquetas=mostrar_etiquetas, mostrar_numeros=mostrar_numeros, escala_texto=et, mostrar_circulos=mostrar_circulos, occupied_labels=etiquetas_ocupadas)
        if pi:
            vis = ProcesadorPielCocodriloMTC.dibujar(vis, pi, mostrar_mallas=mostrar_mallas, mostrar_etiquetas=mostrar_etiquetas, mostrar_numeros=mostrar_numeros, escala_texto=et, mostrar_circulos=mostrar_circulos, mostrar_poligonos=mostrar_poligonos, occupied_labels=etiquetas_ocupadas)

        y = 30
        texto_visible(vis, 'MTC 2018', (10, y), 0.60, (255, 255, 255), 2, factor=1.0)
        y += 25
        resumen = {}
        for falla in fallas:
            key = falla['tipo']
            if key not in resumen:
                resumen[key] = {'total': 0, 1: 0, 2: 0, 3: 0}
            resumen[key]['total'] += 1
            sev = falla.get('severidad')
            if sev in resumen[key]:
                resumen[key][sev] += 1
        for tipo, counts in resumen.items():
            texto_visible(vis, f"{tipo}: {counts['total']} (G1:{counts[1]} G2:{counts[2]} G3:{counts[3]})", (10, y), 0.48, (200, 200, 200), 1, factor=1.0)
            y += 22
        return vis

    def _dibujar_vizir(self, imagen, fallas, cfg):
        vis = imagen.copy()
        mostrar_mallas = cfg.get('show_mallas', True)
        mostrar_etiquetas = cfg.get('show_etiquetas', True)
        mostrar_numeros = cfg.get('show_numeros', True)
        mostrar_circulos = cfg.get('show_circulos', True)
        mostrar_poligonos = cfg.get('show_poligonos', True)
        etiquetas_ocupadas = None

        activas = [f for f in fallas if not f.get('excluida', False)]
        if mostrar_etiquetas:
            lineas_panel = 1 + len({f.get('tipo', '') for f in activas})
            alto_panel = 20 + lineas_panel * 18
            ancho_panel = min(vis.shape[1] - 6, 430)
            etiquetas_ocupadas = [(6, 6, ancho_panel, alto_panel)]
        ojo = [f for f in activas if f.get('tipo') == ProcesadorOjoDePescadoVIZIR.NOMBRE_VIZIR]
        fis = [f for f in activas if 'FISURA' in f.get('tipo', '') and 'COCODRILO' not in f.get('tipo', '')]
        bpp = [f for f in activas if f.get('tipo') == ProcesadorBachesYParchesVIZIR.NOMBRE_VIZIR]
        piel = [f for f in activas if 'COCODRILO' in f.get('tipo', '')]

        if ojo:
            vis = ProcesadorOjoDePescadoVIZIR.dibujar(
                vis, ojo,
                mostrar_mallas=mostrar_mallas,
                mostrar_etiquetas=mostrar_etiquetas,
                mostrar_numeros=mostrar_numeros,
                escala_texto=1.0,
                mostrar_circulos=mostrar_circulos,
                occupied_labels=etiquetas_ocupadas,
            )
        if fis:
            vis = ProcesadorFisurasVIZIR.dibujar(
                vis, fis,
                mostrar_mallas=mostrar_mallas,
                mostrar_etiquetas=mostrar_etiquetas,
                mostrar_numeros=mostrar_numeros,
                escala_texto=1.0,
                mostrar_circulos=mostrar_circulos,
                occupied_labels=etiquetas_ocupadas,
            )
        if bpp:
            vis = ProcesadorBachesYParchesVIZIR.dibujar(
                vis, bpp,
                mostrar_mallas=mostrar_mallas,
                mostrar_etiquetas=mostrar_etiquetas,
                mostrar_numeros=mostrar_numeros,
                escala_texto=1.0,
                mostrar_circulos=mostrar_circulos,
                occupied_labels=etiquetas_ocupadas,
            )
        if piel:
            vis = ProcesadorPielCocodriloVIZIR.dibujar(
                vis, piel,
                mostrar_mallas=mostrar_mallas,
                mostrar_etiquetas=mostrar_etiquetas,
                mostrar_numeros=mostrar_numeros,
                escala_texto=1.0,
                mostrar_circulos=mostrar_circulos,
                mostrar_poligonos=mostrar_poligonos,
                occupied_labels=etiquetas_ocupadas,
            )

        y = 30
        cv2.putText(vis, 'METODO VIZIR', (10, y), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
        y += 25
        resumen = {}
        for falla in activas:
            key = falla['tipo']
            if key not in resumen:
                resumen[key] = {'total': 0, 1: 0, 2: 0, 3: 0}
            resumen[key]['total'] += 1
            sev = falla.get('severidad')
            if sev in resumen[key]:
                resumen[key][sev] += 1
        for tipo, counts in resumen.items():
            txt = f"{tipo}: {counts['total']} (1:{counts[1]} 2:{counts[2]} 3:{counts[3]})"
            cv2.putText(vis, txt, (10, y), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (200, 200, 200), 1)
            y += 18
        return np.ascontiguousarray(vis)


class ExportadorExcelIntegrado:
    """Exporta un consolidado Excel con PCI, MTC y VIZIR."""

    NOMBRE_ARCHIVO = "RESULTADOS_INTEGRADOS_PAVIMENTOS.xlsx"

    @staticmethod
    def _valores_excel_celdas(falla):
        tipo = str(falla.get("tipo", "") or "").strip().upper()
        categoria = categorizar_tipo_falla(tipo)

        if categoria == "piel":
            return (
                round(float(falla.get("diametro_min_mm", 0.0) or 0.0), 2),
                round(float(falla.get("diametro_promedio_mm", 0.0) or 0.0), 2),
                0,
            )

        if categoria == "hueco" or tipo == "OJO DE PESCADO":
            return 0.0, 0.0, 1

        return (
            round(float(falla.get("diametro_min_mm", 0.0) or 0.0), 2),
            round(float(falla.get("diametro_promedio_mm", 0.0) or 0.0), 2),
            0,
        )

    @staticmethod
    def _fila_sin_fallas(nombre_img, metodo, cal_px_mm, ancho_via, angulo_eje):
        return {
            "Nombre_Imagen": nombre_img,
            "Metodo": metodo,
            "ID": f"{Path(nombre_img).stem}_{metodo}_SIN_0",
            "Tipo_Falla": "SIN FALLAS",
            "Unidad": "-",
            "Confianza_%": 0.0,
            "Severidad_Original": "-",
            "Severidad_UI": "-",
            "Espesor_Total_px": 0.0,
            "Espesor_Total_mm": 0.0,
            "Espesor_px": 0.0,
            "Espesor_mm": 0.0,
            "Longitud_px": 0.0,
            "Longitud_m": 0.0,
            "Diametro_px": 0.0,
            "Diametro_mm": 0.0,
            "Diam_Min_Celda_mm": 0.0,
            "Diam_Prom_Celda_mm": 0.0,
            "N_Celdas": 0,
            "Area_px": 0.0,
            "Area_m2": 0.0,
            "Prof_Asumida": "-",
            "Ubicacion_X": 0.0,
            "Ubicacion_Y": 0.0,
            "Calibracion_px_mm": cal_px_mm,
            "Ancho_Via_m": ancho_via,
            "Angulo_Eje_grados": angulo_eje,
        }

    @staticmethod
    def _iter_fallas(todos_resultados, calibrador):
        cal_px_mm_base = round(float(calibrador.px_por_mm or 0.0), 6)
        ancho_via_base = round(float(getattr(calibrador, "ancho_via_real_m", 0.0) or 0.0), 3)
        angulo_eje_base = round(float(getattr(calibrador, "angulo_eje_via", 0.0) or 0.0), 3)
        for nombre_img, payload in (todos_resultados or {}).items():
            if not isinstance(payload, dict):
                continue
            if "metodos" in payload:
                por_metodo = payload.get("metodos", {})
                cal_px_mm = round(float(payload.get("calibracion_px_mm", cal_px_mm_base) or 0.0), 6)
                ancho_via = round(float(payload.get("ancho_via_real_m", ancho_via_base) or 0.0), 3)
                angulo_eje = round(float(payload.get("angulo_eje", angulo_eje_base) or 0.0), 3)
            else:
                por_metodo = payload
                cal_px_mm = cal_px_mm_base
                ancho_via = ancho_via_base
                angulo_eje = angulo_eje_base
            for metodo, fallas in por_metodo.items():
                activas = [f for f in (fallas or []) if not f.get("excluida", False)]
                if not activas:
                    yield ExportadorExcelIntegrado._fila_sin_fallas(
                        nombre_img, metodo, cal_px_mm, ancho_via, angulo_eje
                    )
                    continue
                for falla in activas:
                    sev_raw = falla.get("severidad", "-")
                    sev_ui_val = severidad_ui(sev_raw)
                    diam_min_celda, diam_prom_celda, n_celdas = ExportadorExcelIntegrado._valores_excel_celdas(falla)
                    yield {
                        "Nombre_Imagen": nombre_img,
                        "Metodo": metodo,
                        "ID": f"{Path(nombre_img).stem}_{metodo}_{falla.get('tipo', 'FAL')[:3]}_{falla.get('id', 0)}",
                        "Tipo_Falla": falla.get("tipo", "FALLA"),
                        "Unidad": falla.get("unidad", "-"),
                        "Confianza_%": round(float(falla.get("confianza", 0.0)) * 100, 2),
                        "Severidad_Original": sev_raw,
                        "Severidad_UI": sev_ui_val,
                        "Espesor_Total_px": round(float(falla.get("espesor_total_px", falla.get("espesor_px", 0.0)) or 0.0), 2),
                        "Espesor_Total_mm": round(float(falla.get("espesor_total_mm", falla.get("espesor_mm", 0.0)) or 0.0), 2),
                        "Espesor_px": round(float(falla.get("espesor_px", 0.0) or 0.0), 2),
                        "Espesor_mm": round(float(falla.get("espesor_mm", 0.0) or 0.0), 2),
                        "Longitud_px": round(float(falla.get("longitud_px", 0.0) or 0.0), 2),
                        "Longitud_m": round(float(falla.get("longitud_m", 0.0) or 0.0), 4),
                        "Diametro_px": round(float(falla.get("diametro_px", 0.0) or 0.0), 2),
                        "Diametro_mm": round(float(falla.get("diametro_mm", 0.0) or 0.0), 2),
                        "Diam_Min_Celda_mm": diam_min_celda,
                        "Diam_Prom_Celda_mm": diam_prom_celda,
                        "N_Celdas": n_celdas,
                        "Area_px": round(float(falla.get("area_px", 0.0) or 0.0), 2),
                        "Area_m2": round(float(falla.get("area_m2", 0.0) or 0.0), 4),
                        "Prof_Asumida": falla.get("profundidad_asumida", "-"),
                        "Ubicacion_X": round(float(falla.get("ubicacion_x", 0.0) or 0.0), 2),
                        "Ubicacion_Y": round(float(falla.get("ubicacion_y", 0.0) or 0.0), 2),
                        "Calibracion_px_mm": cal_px_mm,
                        "Ancho_Via_m": ancho_via,
                        "Angulo_Eje_grados": angulo_eje,
                    }

    @staticmethod
    def _ajustar_excel(writer):
        if not OPENPYXL_OK:
            return
        wb = writer.book
        fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        font = Font(color="FFFFFF", bold=True, size=10)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row >= 1:
                for cell in ws[1]:
                    cell.fill = fill
                    cell.font = font
                    cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 34)

    @staticmethod
    def exportar(todos_resultados, ruta_salida, calibrador, nombre_archivo=None):
        if not PANDAS_OK:
            return None

        filas = list(ExportadorExcelIntegrado._iter_fallas(todos_resultados, calibrador))
        if not filas:
            return None

        archivo = Path(ruta_salida) / (nombre_archivo or ExportadorExcelIntegrado.NOMBRE_ARCHIVO)
        df = pd.DataFrame(filas)
        df["Hay_Falla"] = (df["Tipo_Falla"] != "SIN FALLAS").astype(int)
        columnas_detalle = [
            "Nombre_Imagen",
            "Metodo",
            "Ancho_Via_m",
            "Calibracion_px_mm",
            "Angulo_Eje_grados",
            "ID",
            "Tipo_Falla",
            "Unidad",
            "Confianza_%",
            "Severidad_Original",
            "Severidad_UI",
            "Espesor_Total_px",
            "Espesor_Total_mm",
            "Espesor_px",
            "Espesor_mm",
            "Longitud_px",
            "Longitud_m",
            "Diametro_px",
            "Diametro_mm",
            "Diam_Min_Celda_mm",
            "Diam_Prom_Celda_mm",
            "N_Celdas",
            "Area_px",
            "Area_m2",
            "Prof_Asumida",
            "Ubicacion_X",
            "Ubicacion_Y",
        ]
        columnas_extra = [col for col in df.columns if col not in columnas_detalle]
        df_detalle = df[[col for col in columnas_detalle if col in df.columns] + [col for col in columnas_extra if col != "Hay_Falla"]]

        stats = []
        df_stats_base = df[df["Hay_Falla"] == 1].copy()
        for (metodo, tipo), sub in df_stats_base.groupby(["Metodo", "Tipo_Falla"], dropna=False):
            stat = {
                "Metodo": metodo,
                "Tipo_Falla": tipo,
                "Unidad": sub["Unidad"].iloc[0] if len(sub) else "-",
                "Total": int(len(sub)),
                "Sev_L": int((sub["Severidad_UI"] == "L").sum()),
                "Sev_M": int((sub["Severidad_UI"] == "M").sum()),
                "Sev_H": int((sub["Severidad_UI"] == "H").sum()),
                "Sev_1": int((sub["Severidad_UI"] == "1").sum()),
                "Sev_2": int((sub["Severidad_UI"] == "2").sum()),
                "Sev_3": int((sub["Severidad_UI"] == "3").sum()),
            }
            for col, prefijo, ndigits in [
                ("Espesor_mm", "Espesor", 2),
                ("Longitud_m", "Longitud", 4),
                ("Area_m2", "Area", 4),
                ("Diametro_mm", "Diametro", 2),
                ("Diam_Prom_Celda_mm", "Diam_Prom_Celda", 2),
            ]:
                valores = pd.to_numeric(sub[col], errors="coerce").fillna(0.0)
                valores = valores[valores > 0]
                if len(valores) > 0:
                    stat[f"{prefijo}_Min"] = round(float(valores.min()), ndigits)
                    stat[f"{prefijo}_Max"] = round(float(valores.max()), ndigits)
                    stat[f"{prefijo}_Prom"] = round(float(valores.mean()), ndigits)
                    stat[f"{prefijo}_Total"] = round(float(valores.sum()), ndigits)
                else:
                    stat[f"{prefijo}_Min"] = 0
                    stat[f"{prefijo}_Max"] = 0
                    stat[f"{prefijo}_Prom"] = 0
                    stat[f"{prefijo}_Total"] = 0
            n_celdas = pd.to_numeric(sub["N_Celdas"], errors="coerce").fillna(0).astype(int)
            stat["N_Celdas_Total"] = int(n_celdas.sum())
            stats.append(stat)
        if stats:
            df_stats = pd.DataFrame(stats)
        else:
            df_stats = pd.DataFrame(columns=[
                "Metodo",
                "Tipo_Falla",
                "Unidad",
                "Total",
                "Sev_L",
                "Sev_M",
                "Sev_H",
                "Sev_1",
                "Sev_2",
                "Sev_3",
            ])

        resumen = (
            df.groupby(["Nombre_Imagen", "Metodo"], dropna=False)
            .agg({
                "Hay_Falla": "sum",
                "Ancho_Via_m": "first",
                "Confianza_%": "mean",
                "Espesor_mm": "mean",
                "Longitud_m": "sum",
                "Area_m2": "sum",
                "Diametro_mm": "mean",
                "N_Celdas": "sum",
            })
            .reset_index()
            .round(4)
            .rename(columns={
                "Hay_Falla": "Total_Fallas",
                "Confianza_%": "Confianza_Prom_%",
                "Espesor_mm": "Espesor_Prom_mm",
                "Longitud_m": "Longitud_Total_m",
                "Area_m2": "Area_Total_m2",
                "Diametro_mm": "Diametro_Prom_mm",
                "N_Celdas": "N_Celdas_Total",
            })
        )
        columnas_resumen = [
            "Nombre_Imagen",
            "Metodo",
            "Ancho_Via_m",
            "Total_Fallas",
            "Confianza_Prom_%",
            "Espesor_Prom_mm",
            "Longitud_Total_m",
            "Area_Total_m2",
            "Diametro_Prom_mm",
            "N_Celdas_Total",
        ]
        columnas_resumen_extra = [col for col in resumen.columns if col not in columnas_resumen]
        resumen = resumen[[col for col in columnas_resumen if col in resumen.columns] + columnas_resumen_extra]

        try:
            with pd.ExcelWriter(str(archivo), engine="openpyxl") as writer:
                df_detalle.to_excel(writer, sheet_name="Fallas_Detectadas", index=False)
                df_stats.to_excel(writer, sheet_name="Estadisticas", index=False)
                resumen.to_excel(writer, sheet_name="Resumen_por_Imagen", index=False)
                for metodo in sorted(df_detalle["Metodo"].unique()):
                    df_detalle[df_detalle["Metodo"] == metodo].to_excel(writer, sheet_name=f"{metodo}_Detalle"[:31], index=False)
                ExportadorExcelIntegrado._ajustar_excel(writer)
            return str(archivo)
        except Exception:
            csv_path = Path(ruta_salida) / Path(archivo).with_suffix(".csv").name
            df_detalle.to_csv(str(csv_path), index=False)
            return str(csv_path)


class GestorProyectoPavimentos:
    EXTENSION = ".pavproj"
    FORMATO = "proyecto_pavimentos_integrado"
    VERSION = 1

    class _Serializer:
        def __init__(self, zip_file):
            self.zip_file = zip_file
            self._array_idx = 0
            self._bytes_idx = 0

        def dump(self, value):
            if value is None or isinstance(value, (bool, int, str)):
                return value
            if isinstance(value, float):
                if math.isfinite(value):
                    return value
                return {"__kind__": "float", "value": repr(value)}
            if isinstance(value, np.generic):
                return self.dump(value.item())
            if isinstance(value, Path):
                return {"__kind__": "path", "value": str(value)}
            if isinstance(value, np.ndarray):
                return self._dump_array(value)
            if isinstance(value, bytes):
                return self._dump_bytes(value)
            if isinstance(value, tuple):
                return {"__kind__": "tuple", "items": [self.dump(v) for v in value]}
            if isinstance(value, list):
                return {"__kind__": "list", "items": [self.dump(v) for v in value]}
            if isinstance(value, set):
                return {"__kind__": "set", "items": [self.dump(v) for v in value]}
            if isinstance(value, dict):
                return {
                    "__kind__": "dict",
                    "items": [[self.dump(k), self.dump(v)] for k, v in value.items()],
                }
            raise TypeError(f"Tipo no serializable en proyecto: {type(value).__name__}")

        def _dump_array(self, value):
            self._array_idx += 1
            rel_path = f"arrays/arr_{self._array_idx:06d}.npy"
            buffer = io.BytesIO()
            np.save(buffer, value, allow_pickle=False)
            self.zip_file.writestr(rel_path, buffer.getvalue())
            return {"__kind__": "ndarray", "path": rel_path}

        def _dump_bytes(self, value):
            self._bytes_idx += 1
            rel_path = f"bin/blob_{self._bytes_idx:06d}.bin"
            self.zip_file.writestr(rel_path, value)
            return {"__kind__": "bytes", "path": rel_path}

    class _Deserializer:
        def __init__(self, base_dir, progress_callback=None, progress_start=0.0, progress_end=100.0, total_resources=0):
            self.base_dir = Path(base_dir)
            self.progress_callback = progress_callback
            self.progress_start = float(progress_start)
            self.progress_end = float(progress_end)
            self.total_resources = int(total_resources)
            self._resource_idx = 0

        def _report_resource(self, rel_path):
            self._resource_idx += 1
            if not self.progress_callback:
                return
            if self.total_resources > 0:
                frac = self._resource_idx / self.total_resources
                valor = self.progress_start + (self.progress_end - self.progress_start) * frac
            else:
                valor = self.progress_end
            self.progress_callback(valor, f"Restaurando datos: {Path(rel_path).name}")

        def load(self, value):
            if not isinstance(value, dict) or "__kind__" not in value:
                return value

            kind = value["__kind__"]
            if kind == "float":
                raw = value.get("value", "0.0")
                if raw == "nan":
                    return float("nan")
                if raw == "inf":
                    return float("inf")
                if raw == "-inf":
                    return float("-inf")
                return float(raw)
            if kind == "path":
                return value.get("value", "")
            if kind == "tuple":
                return tuple(self.load(v) for v in value.get("items", []))
            if kind == "list":
                return [self.load(v) for v in value.get("items", [])]
            if kind == "set":
                return set(self.load(v) for v in value.get("items", []))
            if kind == "dict":
                return {self.load(k): self.load(v) for k, v in value.get("items", [])}
            if kind == "ndarray":
                rel_path = value.get("path", "")
                self._report_resource(rel_path)
                with (self.base_dir / rel_path).open("rb") as fh:
                    return np.load(fh, allow_pickle=False)
            if kind == "bytes":
                rel_path = value.get("path", "")
                self._report_resource(rel_path)
                return (self.base_dir / rel_path).read_bytes()
            raise ValueError(f"Tipo de dato desconocido en proyecto: {kind}")

    @classmethod
    def _count_external_resources(cls, value):
        if isinstance(value, dict):
            kind = value.get("__kind__")
            if kind in {"ndarray", "bytes"}:
                return 1
            return sum(cls._count_external_resources(v) for v in value.values())
        if isinstance(value, list):
            return sum(cls._count_external_resources(v) for v in value)
        return 0

    @staticmethod
    def _nombre_seguro(nombre, fallback="imagen"):
        nombre = str(nombre or fallback)
        stem = Path(nombre).stem.encode("ascii", "ignore").decode("ascii")
        stem = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in stem).strip("._")
        stem = stem or fallback
        suffix = Path(nombre).suffix.lower() or ".png"
        return f"{stem}{suffix}"

    @classmethod
    def guardar(cls, ruta_archivo, estado):
        ruta_archivo = Path(ruta_archivo)
        ruta_archivo.parent.mkdir(parents=True, exist_ok=True)

        with zipfile.ZipFile(str(ruta_archivo), "w", compression=zipfile.ZIP_DEFLATED) as zip_file:
            for item in estado.get("imagenes", []):
                origen = item.get("source_path")
                destino = str(item.get("archive_path", "")).replace("\\", "/").lstrip("/")
                if not origen or not Path(origen).is_file():
                    raise FileNotFoundError(f"No se encontro la imagen a empaquetar: {origen}")
                if not destino:
                    raise ValueError("Imagen sin ruta interna en el proyecto.")
                zip_file.write(str(origen), arcname=destino)

            serializer = cls._Serializer(zip_file)
            manifest = {
                "format": cls.FORMATO,
                "version": cls.VERSION,
                "saved_at": datetime.now().isoformat(timespec="seconds"),
                "state": serializer.dump(estado),
            }
            zip_file.writestr(
                "manifest.json",
                json.dumps(manifest, ensure_ascii=False, indent=2),
            )

        return str(ruta_archivo)

    @classmethod
    def cargar(cls, ruta_archivo, progress_callback=None):
        ruta_archivo = Path(ruta_archivo)
        workspace = Path(tempfile.mkdtemp(prefix="pavproj_"))
        try:
            with zipfile.ZipFile(str(ruta_archivo), "r") as zip_file:
                infos = zip_file.infolist()
                total_infos = max(len(infos), 1)
                if progress_callback:
                    progress_callback(2, f"Abriendo proyecto: {ruta_archivo.name}")
                for idx, info in enumerate(infos, start=1):
                    zip_file.extract(info, workspace)
                    if progress_callback:
                        valor = 5 + (50 * idx / total_infos)
                        progress_callback(valor, f"Extrayendo: {Path(info.filename).name}")

            manifest_path = workspace / "manifest.json"
            if not manifest_path.exists():
                raise ValueError("El archivo no contiene manifest.json")

            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            if manifest.get("format") != cls.FORMATO:
                raise ValueError("El archivo seleccionado no es un proyecto valido de pavimentos.")
            version = int(manifest.get("version", 0) or 0)
            if version > cls.VERSION:
                raise ValueError("La version del proyecto es mas nueva que esta aplicacion.")

            state_data = manifest.get("state")
            total_resources = cls._count_external_resources(state_data)
            if progress_callback:
                progress_callback(58, "Leyendo manifiesto del proyecto...")
            deserializer = cls._Deserializer(
                workspace,
                progress_callback=progress_callback,
                progress_start=60.0,
                progress_end=82.0,
                total_resources=total_resources,
            )
            state = deserializer.load(state_data)
            if progress_callback:
                progress_callback(82, "Datos del proyecto restaurados.")
            return {
                "workspace": str(workspace),
                "manifest": manifest,
                "state": state,
            }
        except Exception:
            shutil.rmtree(workspace, ignore_errors=True)
            raise


class AplicacionIntegradaTresMetodos(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('Analisis de Pavimentos | PCI + MTC + VIZIR | Universidad Nacional del Altiplano')
        self.configure(bg=EstiloUI.BG_DARK)
        try:
            self.state('zoomed')
        except Exception:
            self.geometry('1500x900')
        self.minsize(1280, 760)

        self.motor = MotorIntegradoTresMetodos()
        self.imagenes_cargadas = []
        self.imagen_actual_idx = -1
        self.resultado_actual = None
        self.procesando = False
        self._photo_refs = {}
        self._imagenes_render = {}
        self.vars = {}

        self._crear_estilos()
        self._crear_interfaz()
        self._verificar_dependencias()

    def _crear_estilos(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Accent.TButton', background=EstiloUI.BG_BUTTON, foreground='white', font=EstiloUI.FONT_SUBTITLE, padding=(10, 5))
        style.map('Accent.TButton', background=[('active', EstiloUI.BG_BUTTON_HOVER)])
        style.configure('Secondary.TButton', background=EstiloUI.BG_BUTTON_SECONDARY, foreground=EstiloUI.FG_PRIMARY, font=EstiloUI.FONT_BODY, padding=(8, 4))
        style.map('Secondary.TButton',
                  background=[('active', EstiloUI.BG_ACCENT)],
                  foreground=[('active', EstiloUI.FG_PRIMARY)])
        style.configure('ViewMode.TButton', background=EstiloUI.BG_BUTTON_SECONDARY, foreground=EstiloUI.FG_PRIMARY, font=("Segoe UI", 10, "bold"), padding=(8, 2))
        style.map('ViewMode.TButton',
                  background=[('active', EstiloUI.BG_ACCENT)],
                  foreground=[('active', EstiloUI.FG_PRIMARY)])
        style.configure('ViewModeActive.TButton', background=EstiloUI.BG_BUTTON, foreground='white', font=("Segoe UI", 10, "bold"), padding=(8, 2))
        style.map('ViewModeActive.TButton', background=[('active', EstiloUI.BG_BUTTON_HOVER)])
        style.configure('View.TNotebook', background=EstiloUI.BG_DARK, borderwidth=0)
        style.configure('View.TNotebook.Tab', background=EstiloUI.BG_BUTTON_SECONDARY, foreground=EstiloUI.FG_PRIMARY, padding=(14, 8), font=EstiloUI.FONT_BODY)
        style.map('View.TNotebook.Tab',
                  background=[('selected', EstiloUI.BG_PANEL), ('active', EstiloUI.BG_ACCENT)],
                  foreground=[('selected', EstiloUI.FG_ACCENT), ('active', EstiloUI.FG_PRIMARY)])

    def _crear_interfaz(self):
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)
        self._crear_barra_superior()

        self.paned = tk.PanedWindow(self, orient=tk.HORIZONTAL, bg=EstiloUI.BG_DARK, sashwidth=6)
        self.paned.grid(row=1, column=0, sticky='nsew')

        self.frame_config = tk.Frame(self.paned, bg=EstiloUI.BG_PANEL, width=280)
        self.frame_visor = tk.Frame(self.paned, bg=EstiloUI.BG_DARK)
        self.frame_resultados = tk.Frame(self.paned, bg=EstiloUI.BG_PANEL, width=340)

        self.paned.add(self.frame_config, minsize=240, width=280)
        self.paned.add(self.frame_visor, minsize=520)
        self.paned.add(self.frame_resultados, minsize=300, width=340)

        self._crear_panel_config()
        self._crear_panel_visor()
        self._crear_panel_resultados()
        self._crear_barra_estado()

    def _crear_barra_superior(self):
        barra = tk.Frame(self, bg=EstiloUI.BG_PANEL, height=58)
        barra.grid(row=0, column=0, sticky='ew')
        barra.grid_propagate(False)

        tk.Label(
            barra,
            text='ANALISIS DE PAVIMENTOS | PCI + MTC + VIZIR',
            font=EstiloUI.FONT_TITLE,
            bg=EstiloUI.BG_PANEL,
            fg=EstiloUI.FG_PRIMARY,
        ).pack(side='left', padx=15)

        botones = tk.Frame(barra, bg=EstiloUI.BG_PANEL)
        botones.pack(side='right', padx=10, pady=8)

        ttk.Button(botones, text='Guardar Resultados', style='Accent.TButton', command=self._guardar_resultados).pack(side='right', padx=3)
        ttk.Button(botones, text='Procesar Actual', style='Accent.TButton', command=self._procesar_actual).pack(side='right', padx=3)
        ttk.Button(botones, text='Calibrar Actual', style='Secondary.TButton', command=self._calibrar_actual).pack(side='right', padx=3)
        ttk.Button(botones, text='Cargar Imagenes', style='Secondary.TButton', command=self._cargar_imagenes).pack(side='right', padx=3)
        ttk.Button(botones, text='Cargar Modelo', style='Secondary.TButton', command=self._cargar_modelo).pack(side='right', padx=3)

    def _crear_panel_config(self):
        self.frame_config.grid_rowconfigure(0, weight=1)
        self.frame_config.grid_columnconfigure(0, weight=1)
        canvas = tk.Canvas(self.frame_config, bg=EstiloUI.BG_PANEL, highlightthickness=0)
        scroll = ttk.Scrollbar(self.frame_config, orient='vertical', command=canvas.yview)
        inner = tk.Frame(canvas, bg=EstiloUI.BG_PANEL)
        inner.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=inner, anchor='nw')
        canvas.configure(yscrollcommand=scroll.set)
        canvas.grid(row=0, column=0, sticky='nsew')
        scroll.grid(row=0, column=1, sticky='ns')

        self._sec(inner, 'CONFIGURACION GENERAL')
        self._slider(inner, 'ancho_via_real_m', 'Ancho de via (m)', 3.0, 12.0, self.motor.config.get('ancho_via_real_m', 6.5), 0.1)
        self._slider(inner, 'confianza_min', 'Confianza minima', 0.05, 0.95, self.motor.config.get('confianza_min', 0.1), 0.05)
        self._slider(inner, 'iou_threshold', 'IoU threshold', 0.10, 0.95, self.motor.config.get('iou_threshold', 0.45), 0.05)

        self._sec(inner, 'VISUALIZACION')
        for key, text, default in [
            ('show_etiquetas', 'Mostrar etiquetas', True),
            ('show_numeros', 'Mostrar valores numericos', True),
            ('show_mallas', 'Mostrar mallas y heatmaps', True),
            ('show_circulos', 'Mostrar circulos auxiliares', True),
        ]:
            var = tk.BooleanVar(value=self.motor.config.get(key, default))
            self.vars[key] = var
            tk.Checkbutton(
                inner,
                text=text,
                variable=var,
                bg=EstiloUI.BG_PANEL,
                fg=EstiloUI.FG_PRIMARY,
                selectcolor=EstiloUI.BG_DARK,
                activebackground=EstiloUI.BG_PANEL,
                font=EstiloUI.FONT_BODY,
                command=self._re_render_resultado,
            ).pack(anchor='w', padx=10, pady=2)

        self._sec(inner, 'RUTAS')
        self.lbl_modelo_path = tk.Label(inner, text='Modelo: no cargado', justify='left', wraplength=240, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_SECONDARY, font=EstiloUI.FONT_SMALL)
        self.lbl_modelo_path.pack(fill='x', padx=10, pady=2)
        self.lbl_calibracion = tk.Label(inner, text='Calibracion: automatica por ancho de imagen', justify='left', wraplength=240, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_SECONDARY, font=EstiloUI.FONT_SMALL)
        self.lbl_calibracion.pack(fill='x', padx=10, pady=2)

    def _crear_panel_visor(self):
        self.frame_visor.grid_rowconfigure(2, weight=1)
        self.frame_visor.grid_columnconfigure(0, weight=1)

        nav = tk.Frame(self.frame_visor, bg=EstiloUI.BG_CARD, height=40)
        nav.grid(row=0, column=0, sticky='ew', pady=(0, 2))
        nav.grid_propagate(False)
        ttk.Button(nav, text='< Anterior', style='Secondary.TButton', command=self._imagen_anterior).pack(side='left', padx=5, pady=4)
        ttk.Button(nav, text='Siguiente >', style='Secondary.TButton', command=self._imagen_siguiente).pack(side='left', padx=5, pady=4)
        self.lbl_nav = tk.Label(nav, text='Sin imagenes cargadas', bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_PRIMARY, font=EstiloUI.FONT_BODY)
        self.lbl_nav.pack(side='left', padx=12)

        info = tk.Frame(self.frame_visor, bg=EstiloUI.BG_PANEL, height=28)
        info.grid(row=1, column=0, sticky='ew', pady=(0, 2))
        info.grid_propagate(False)
        self.lbl_info_img = tk.Label(info, text='--', bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT, font=EstiloUI.FONT_SMALL)
        self.lbl_info_img.pack(side='left', padx=8)

        notebook = ttk.Notebook(self.frame_visor, style='View.TNotebook')
        notebook.grid(row=2, column=0, sticky='nsew')
        self.notebook_vistas = notebook
        self.view_labels = {}

        for key in ['Original', 'PCI', 'MTC', 'VIZIR']:
            frame = tk.Frame(notebook, bg=EstiloUI.BG_INPUT)
            frame.pack_propagate(False)
            lbl = tk.Label(frame, text='Sin imagen', bg=EstiloUI.BG_INPUT, fg=EstiloUI.FG_SECONDARY)
            lbl.pack(fill='both', expand=True)
            lbl.bind('<Configure>', lambda e: self._render_all_views())
            notebook.add(frame, text=key)
            self.view_labels[key] = lbl

    def _crear_panel_resultados(self):
        self.frame_resultados.grid_rowconfigure(1, weight=1)
        self.frame_resultados.grid_rowconfigure(3, weight=1)
        self.frame_resultados.grid_columnconfigure(0, weight=1)

        head = tk.Frame(self.frame_resultados, bg=EstiloUI.BG_CARD)
        head.grid(row=0, column=0, sticky='ew', padx=5, pady=(5, 2))
        tk.Label(head, text='RESUMENES', bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_ACCENT, font=EstiloUI.FONT_SUBTITLE).pack(anchor='w', padx=10, pady=6)

        nb = ttk.Notebook(self.frame_resultados, style='View.TNotebook')
        nb.grid(row=1, column=0, sticky='nsew', padx=5, pady=(0, 5))
        self.summary_texts = {}
        for key in ['General', 'PCI', 'MTC', 'VIZIR']:
            frame = tk.Frame(nb, bg=EstiloUI.BG_CARD)
            txt = scrolledtext.ScrolledText(
                frame,
                wrap=tk.WORD,
                font=EstiloUI.FONT_MONO,
                bg=EstiloUI.BG_INPUT,
                fg=EstiloUI.FG_PRIMARY,
                insertbackground=EstiloUI.FG_PRIMARY,
                relief='flat',
            )
            txt.pack(fill='both', expand=True)
            nb.add(frame, text=key)
            self.summary_texts[key] = txt

        log_head = tk.Frame(self.frame_resultados, bg=EstiloUI.BG_CARD)
        log_head.grid(row=2, column=0, sticky='ew', padx=5, pady=(2, 2))
        tk.Label(log_head, text='LOG', bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_ACCENT, font=EstiloUI.FONT_SUBTITLE).pack(anchor='w', padx=10, pady=6)

        self.log_text = scrolledtext.ScrolledText(
            self.frame_resultados,
            wrap=tk.WORD,
            font=EstiloUI.FONT_MONO,
            bg=EstiloUI.BG_INPUT,
            fg=EstiloUI.FG_LOG,
            insertbackground=EstiloUI.FG_LOG,
            relief='flat',
        )
        self.log_text.grid(row=3, column=0, sticky='nsew', padx=5, pady=(0, 5))

    def _crear_barra_estado(self):
        barra = tk.Frame(self, bg=EstiloUI.BG_CARD, height=34)
        barra.grid(row=2, column=0, sticky='ew')
        barra.grid_propagate(False)
        self.lbl_estado = tk.Label(barra, text='Listo', bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_SECONDARY, font=EstiloUI.FONT_SMALL)
        self.lbl_estado.pack(side='left', padx=8)
        self.project_progress_frame = tk.Frame(barra, bg=EstiloUI.BG_CARD)
        self.project_progress_label = tk.Label(
            self.project_progress_frame,
            text='Cargando proyecto...',
            bg=EstiloUI.BG_CARD,
            fg=EstiloUI.FG_HIGHLIGHT,
            font=EstiloUI.FONT_SMALL,
        )
        self.project_progress_label.pack(side='left', padx=(0, 6))
        self.project_progress = ttk.Progressbar(
            self.project_progress_frame,
            orient='horizontal',
            mode='determinate',
            maximum=100,
            length=180,
        )
        self.project_progress.pack(side='left')
        self.lbl_autores = tk.Label(
            barra,
            text='Bach. Miguel Bernardino Quispe Arias | Bach. Briza Edith Catachura Aycaya',
            bg=EstiloUI.BG_CARD,
            fg=EstiloUI.FG_PRIMARY,
            font=EstiloUI.FONT_SMALL,
        )
        self.lbl_autores.place(relx=0.5, rely=0.5, anchor='center')
        self.lbl_modelo = tk.Label(barra, text='Modelo: no cargado', bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_HIGHLIGHT, font=EstiloUI.FONT_SMALL)
        self.lbl_modelo.pack(side='right', padx=8)

    def _sec(self, parent, texto):
        tk.Frame(parent, bg=EstiloUI.BG_ACCENT, height=2).pack(fill='x', padx=10, pady=(10, 0))
        tk.Label(parent, text=texto, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_ACCENT, font=EstiloUI.FONT_LABEL).pack(anchor='w', padx=10, pady=(2, 4))

    def _slider(self, parent, key, label, from_, to, default, resolution):
        frame = tk.Frame(parent, bg=EstiloUI.BG_PANEL)
        frame.pack(fill='x', padx=10, pady=2)
        row = tk.Frame(frame, bg=EstiloUI.BG_PANEL)
        row.pack(fill='x')
        tk.Label(row, text=label, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_SECONDARY, font=EstiloUI.FONT_SMALL).pack(side='left')
        value_lbl = tk.Label(row, text=str(default), bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT, font=EstiloUI.FONT_LABEL)
        value_lbl.pack(side='right')
        var = tk.DoubleVar(value=default)
        self.vars[key] = var
        scale = tk.Scale(frame, from_=from_, to=to, orient='horizontal', resolution=resolution, variable=var, showvalue=False, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, troughcolor=EstiloUI.BG_DARK, highlightthickness=0, length=220)
        scale.pack(fill='x')
        var.trace_add('write', lambda *a, v=var, lbl=value_lbl: lbl.config(text=f"{v.get():.2f}"))

    def _general_config_keys(self):
        return ("ancho_via_real_m", "confianza_min", "iou_threshold")

    def _on_config_general_scope_change(self):
        if hasattr(self, "var_config_general_scope"):
            self._config_general_scope = str(self.var_config_general_scope.get() or "actual")
        self._actualizar_scope_general_ui()

    def _actualizar_scope_general_ui(self):
        if hasattr(self, "var_config_general_scope"):
            valor = str(self.var_config_general_scope.get() or "actual")
            self._config_general_scope = valor
        else:
            valor = getattr(self, "_config_general_scope", "actual")
        lbl = getattr(self, "lbl_scope_general", None)
        if lbl is None:
            return
        if valor == "todas":
            txt = "Al aplicar, ancho de via, confianza e IoU se copiaran a todas las imagenes."
        else:
            txt = "Al aplicar, ancho de via, confianza e IoU solo se guardaran para la imagen actual."
        lbl.config(text=txt)

    def _snapshot_config_general(self):
        self._sync_config()
        return {key: copy.deepcopy(self.motor.config.get(key)) for key in self._general_config_keys()}

    def _aplicar_config_general_desde_panel(self):
        scope = getattr(self, "_config_general_scope", "actual")
        valores = self._snapshot_config_general()
        if scope == "todas":
            if not self._config_global_base:
                self._config_global_base = self._snapshot_config_avanzada()
            for key, value in valores.items():
                self._config_global_base[key] = copy.deepcopy(value)
            for idx in list(self._config_por_imagen.keys()):
                cfg_img = copy.deepcopy(self._config_por_imagen.get(idx, {}))
                for key, value in valores.items():
                    cfg_img[key] = copy.deepcopy(value)
                self._config_por_imagen[idx] = cfg_img
            self._log("Configuracion general aplicada a todas las imagenes. Procese de nuevo para recalcular.")
            self._estado("Configuracion general aplicada a todas las imagenes")
        else:
            idx = self.imagen_actual_idx
            if idx is None or idx < 0:
                self._log("No hay una imagen activa para guardar la configuracion general individual.")
                self._estado("Sin imagen activa")
                return
            cfg_img = copy.deepcopy(self._config_por_imagen.get(idx, self._config_actual_para_ui()))
            for key, value in valores.items():
                cfg_img[key] = copy.deepcopy(value)
            self._config_por_imagen[idx] = cfg_img
            self._log("Configuracion general aplicada solo a la imagen actual. Procese de nuevo para recalcular.")
            self._estado("Configuracion general aplicada a la imagen actual")
        self._actualizar_scope_general_ui()

    def _sync_config(self):
        for key, var in self.vars.items():
            if isinstance(var, tk.BooleanVar):
                self.motor.config[key] = bool(var.get())
            else:
                self.motor.config[key] = float(var.get())
        self.motor.calibrador.ancho_via_real_m = self.motor.config.get('ancho_via_real_m', 6.5)

    def _verificar_dependencias(self):
        faltantes = []
        if not YOLO_OK:
            faltantes.append('ultralytics')
        if not PANDAS_OK:
            faltantes.append('pandas (solo exportaciones Excel)')
        if faltantes:
            self._log('Dependencias opcionales faltantes: ' + ', '.join(faltantes))

    def _cargar_modelo(self):
        ruta = filedialog.askopenfilename(
            title='Seleccionar modelo YOLO',
            filetypes=[('Modelo PyTorch', '*.pt'), ('Todos', '*.*')],
            initialdir=_resolver_directorio_dialogo('modelos', self.motor.config.get('ruta_modelo')),
        )
        if not ruta:
            return
        ok, msg = self.motor.cargar_modelo(ruta)
        self.lbl_modelo.config(text=f"Modelo: {Path(ruta).name}" if ok else 'Modelo: error')
        self.lbl_modelo_path.config(text=f"Modelo: {ruta}")
        self._estado(msg)
        self._log(msg)

    def _cargar_imagenes(self):
        rutas = filedialog.askopenfilenames(
            title='Seleccionar imagenes',
            filetypes=[('Imagenes', '*.jpg *.jpeg *.png *.bmp *.tif *.tiff'), ('Todos', '*.*')],
            initialdir=self.motor.config.get('ruta_imagenes', '.'),
        )
        if not rutas:
            return
        self.imagenes_cargadas = list(rutas)
        self.imagen_actual_idx = 0
        self.resultado_actual = None
        self._actualizar_nav()
        self._mostrar_original_actual()
        self._estado(f"{len(rutas)} imagen(es) cargada(s)")
        self._log(f"Se cargaron {len(rutas)} imagen(es).")

    def _imagen_anterior(self):
        if not self.imagenes_cargadas:
            return
        self.imagen_actual_idx = (self.imagen_actual_idx - 1) % len(self.imagenes_cargadas)
        self.resultado_actual = None
        self._actualizar_nav()
        self._mostrar_original_actual()

    def _imagen_siguiente(self):
        if not self.imagenes_cargadas:
            return
        self.imagen_actual_idx = (self.imagen_actual_idx + 1) % len(self.imagenes_cargadas)
        self.resultado_actual = None
        self._actualizar_nav()
        self._mostrar_original_actual()

    def _ruta_actual(self):
        if not self.imagenes_cargadas or self.imagen_actual_idx < 0:
            return None
        return self.imagenes_cargadas[self.imagen_actual_idx]

    def _actualizar_nav(self):
        if not self.imagenes_cargadas:
            self.lbl_nav.config(text='Sin imagenes cargadas')
            self.lbl_info_img.config(text='--')
            return
        ruta = Path(self.imagenes_cargadas[self.imagen_actual_idx])
        self.lbl_nav.config(text=f"{self.imagen_actual_idx + 1}/{len(self.imagenes_cargadas)} | {ruta.name}")
        self.lbl_info_img.config(text=str(ruta))

    def _mostrar_original_actual(self, reset_modos=True):
        ruta = self._ruta_actual()
        if not ruta:
            return
        img = cv2.imread(str(ruta))
        if img is None:
            self._log(f"No se pudo leer: {ruta}")
            return
        self._imagenes_render = {'Original': img, 'PCI': None, 'MTC': None, 'VIZIR': None}
        self._render_all_views()
        self._actualizar_resumenes_vacios(ruta, img)

    def _actualizar_resumenes_vacios(self, ruta, img):
        h, w = img.shape[:2]
        self._set_text('General', f"Archivo: {Path(ruta).name}\nDimensiones: {w}x{h}\n\nProcese la imagen para obtener PCI, MTC y VIZIR.")
        for key in ['PCI', 'MTC', 'VIZIR']:
            self._set_text(key, 'Sin resultados aun.')

    def _calibrar_actual(self):
        ruta = self._ruta_actual()
        if not ruta:
            messagebox.showinfo('Info', 'Cargue una imagen primero.')
            return
        img = cv2.imread(str(ruta))
        if img is None:
            messagebox.showerror('Error', 'No se pudo cargar la imagen actual.')
            return
        self._sync_config()
        self.motor.calibrar_imagen(img, usar_gui=True, parent=self)
        self.lbl_calibracion.config(
            text=f"Calibracion manual: {self.motor.calibrador.px_por_mm:.4f} px/mm | Eje {self.motor.calibrador.get_angulo_eje():.1f} grados"
        )
        self._log('Calibracion manual actualizada.')
        self._estado('Calibracion manual guardada')

    def _procesar_actual(self):
        ruta = self._ruta_actual()
        if not ruta:
            messagebox.showinfo('Info', 'Cargue una imagen primero.')
            return
        if not self.motor.modelo_cargado:
            messagebox.showinfo('Info', 'Cargue el modelo YOLO primero.')
            return
        if self.procesando:
            return
        self._sync_config()
        self.procesando = True
        self._estado('Procesando...')
        self._log(f"\n{'=' * 60}\nProcesando {Path(ruta).name}")

        def worker():
            try:
                resultado = self.motor.procesar_imagen(ruta, callback_log=self._log_safe)
                self.after(0, lambda: self._on_procesado(resultado))
            except Exception as exc:
                tb = traceback.format_exc()
                self.after(0, lambda: self._on_error(exc, tb))

        Thread(target=worker, daemon=True).start()

    def _on_procesado(self, resultado):
        self.procesando = False
        if resultado is None:
            self._estado('Error en procesamiento')
            return
        self.resultado_actual = resultado
        self._imagenes_render = {
            'Original': resultado['imagen'],
            'PCI': resultado['metodos']['PCI']['visual'],
            'MTC': resultado['metodos']['MTC']['visual'],
            'VIZIR': resultado['metodos']['VIZIR']['visual'],
        }
        self._render_all_views()
        self._actualizar_resumenes(resultado)
        self.lbl_calibracion.config(
            text=f"Calibracion activa: {resultado['calibracion_px_mm']:.4f} px/mm | Eje {resultado['angulo_eje']:.1f} grados"
        )
        self._estado('Procesamiento completado')

    def _on_error(self, error, detalle):
        self.procesando = False
        self._estado('Error en procesamiento')
        self._log(str(error))
        self._log(detalle.rstrip())

    def _actualizar_resumenes(self, resultado):
        ancho, alto = resultado['dimensiones']
        general = [
            f"Archivo: {resultado['nombre']}",
            f"Dimensiones: {ancho}x{alto}",
            f"Calibracion: {resultado['calibracion_px_mm']:.4f} px/mm",
            f"Eje de via: {resultado['angulo_eje']:.1f} grados",
            '',
        ]
        for metodo in ['PCI', 'MTC', 'VIZIR']:
            total = len(resultado['metodos'][metodo]['fallas'])
            general.append(f"{metodo}: {total} falla(s)")
        self._set_text('General', '\n'.join(general))

        for metodo in ['PCI', 'MTC', 'VIZIR']:
            data = resultado['metodos'][metodo]
            lineas = [f"Metodo: {metodo}", f"Total fallas: {len(data['fallas'])}", '']
            if data['resumen_severidad']:
                lineas.append('Severidad:')
                for sev, cantidad in sorted(data['resumen_severidad'].items(), key=lambda item: item[0]):
                    lineas.append(f"  {sev}: {cantidad}")
                lineas.append('')
            conteo_tipo = {}
            for falla in data['fallas']:
                tipo = falla.get('tipo', 'FALLA')
                conteo_tipo[tipo] = conteo_tipo.get(tipo, 0) + 1
            if conteo_tipo:
                lineas.append('Tipos detectados:')
                for tipo, cantidad in conteo_tipo.items():
                    lineas.append(f"  {tipo}: {cantidad}")
                lineas.append('')
            if data['fallas']:
                lineas.append('Detalle:')
                for falla in data['fallas']:
                    lineas.append(self._describir_falla(falla))
            else:
                lineas.append('Sin fallas detectadas.')
            self._set_text(metodo, '\n'.join(lineas))

    def _describir_falla(self, falla):
        base = (
            f"- {falla.get('tipo', 'FALLA')} #{falla.get('id', '-')}: "
            f"Sev {falla.get('severidad', '-')} | Conf {falla.get('confianza', 0) * 100:.1f}%"
        )
        extras = []
        if falla.get('diametro_mm', 0):
            extras.append(f"D={falla.get('diametro_mm', 0):.1f} mm")
        if falla.get('espesor_mm', 0):
            extras.append(f"e={falla.get('espesor_mm', 0):.1f} mm")
        if falla.get('longitud_m', 0):
            extras.append(f"L={falla.get('longitud_m', 0):.3f} m")
        if falla.get('area_m2', 0):
            extras.append(f"A={falla.get('area_m2', 0):.4f} m2")
        return base + (" | " + " | ".join(extras) if extras else '')

    def _set_text(self, key, content):
        txt = self.summary_texts[key]
        txt.config(state='normal')
        txt.delete('1.0', tk.END)
        txt.insert(tk.END, content)
        txt.config(state='disabled')

    def _mostrar_cv2_en_label(self, key, img_cv):
        lbl = self.view_labels[key]
        if img_cv is None:
            lbl.config(image='', text='Sin imagen')
            lbl.image = None
            return
        h, w = img_cv.shape[:2]
        max_w = max(lbl.winfo_width() - 12, 320)
        max_h = max(lbl.winfo_height() - 12, 240)
        escala = min(max_w / w, max_h / h, 1.0)
        if escala <= 0:
            escala = 1.0
        if escala != 1.0:
            img_show = cv2.resize(img_cv, (max(int(w * escala), 1), max(int(h * escala), 1)))
        else:
            img_show = img_cv
        rgb = cv2.cvtColor(img_show, cv2.COLOR_BGR2RGB)
        photo = ImageTk.PhotoImage(Image.fromarray(rgb))
        lbl.config(image=photo, text='')
        lbl.image = photo
        self._photo_refs[key] = photo

    def _render_all_views(self):
        for key in ['Original', 'PCI', 'MTC', 'VIZIR']:
            self._mostrar_cv2_en_label(key, self._imagenes_render.get(key))

    def _re_render_resultado(self):
        if not self.resultado_actual:
            return
        self._sync_config()
        imagen = self.resultado_actual['imagen']
        raw_masks = self.resultado_actual.get('raw_masks')
        if raw_masks is None:
            return
        pci = self.motor._procesar_pci_desde_mascaras(imagen, raw_masks)
        mtc = self.motor._procesar_mtc_desde_mascaras(imagen, raw_masks)
        vizir = self.motor._procesar_vizir_desde_mascaras(imagen, raw_masks)
        self.resultado_actual['metodos']['PCI'] = pci
        self.resultado_actual['metodos']['MTC'] = mtc
        self.resultado_actual['metodos']['VIZIR'] = vizir
        self._imagenes_render['PCI'] = pci['visual']
        self._imagenes_render['MTC'] = mtc['visual']
        self._imagenes_render['VIZIR'] = vizir['visual']
        self._render_all_views()
        self._actualizar_resumenes(self.resultado_actual)

    def _guardar_resultados(self):
        if not self.resultado_actual:
            messagebox.showinfo('Info', 'Procese una imagen primero.')
            return
        carpeta = filedialog.askdirectory(title='Seleccionar carpeta de salida', initialdir=self.motor.config.get('ruta_salida', '.'))
        if not carpeta:
            return
        carpeta = Path(carpeta)
        carpeta.mkdir(parents=True, exist_ok=True)
        nombre_base = Path(self.resultado_actual['nombre']).stem
        guardados = []
        for key in ['Original', 'PCI', 'MTC', 'VIZIR']:
            img = self._imagenes_render.get(key)
            if img is None:
                continue
            ruta = carpeta / f"{nombre_base}_{key}.png"
            cv2.imwrite(str(ruta), img)
            guardados.append(ruta.name)
        resumen = carpeta / f"{nombre_base}_resumen.txt"
        contenido = []
        for key in ['General', 'PCI', 'MTC', 'VIZIR']:
            contenido.append(f"[{key}]")
            contenido.append(self.summary_texts[key].get('1.0', tk.END).strip())
            contenido.append('')
        resumen.write_text('\n'.join(contenido), encoding='utf-8')
        guardados.append(resumen.name)
        self._log('Guardado: ' + ', '.join(guardados))
        self._estado(f"Resultados guardados en {carpeta}")

    def _log(self, msg):
        self.log_text.insert(tk.END, msg + '\n')
        self.log_text.see(tk.END)

    def _log_safe(self, msg):
        self.after(0, lambda m=msg: self._log(m))

    def _estado(self, msg):
        self.lbl_estado.config(text=msg)


class AplicacionIntegradaTresMetodosAvanzada(AplicacionIntegradaTresMetodos):
    METODOS = ("PCI", "MTC", "VIZIR")
    CLAVES_AVANZADAS = [
        "ancho_via_real_m",
        "confianza_min",
        "iou_threshold",
        "filtrar_baches",
        "min_diametro_hueco_mm",
        "profundidad_asumida_huecos",
        "filtrar_grietas",
        "min_longitud_grieta_m",
        "usar_borde_berma_pci",
        "borde_interno_m",
        "borde_externo_m",
        "filtrar_parches",
        "min_area_parche_m2",
        "parche_ratio_leve_max",
        "parche_ratio_moderado_max",
        "filtrar_piel",
        "min_area_piel_m2",
        "merge_iou_threshold",
        "merge_distancia_max_px",
        "merge_fisuras_px",
        "clahe_clip",
        "clahe_tile",
        "bilateral_d",
        "bilateral_sigma_color",
        "bilateral_sigma_space",
        "block_size",
        "C_umbral",
        "usar_frangi",
        "usar_refinamiento",
        "usar_multiescala",
        "kernel_apertura",
        "kernel_cierre",
        "iteraciones_cierre",
        "min_area_poligono",
        "min_circularidad",
        "min_vertices",
        "max_vertices",
        "min_radio_circulo",
        "min_longitud_rama",
        "min_area_objeto",
        "max_gap_cierre",
    ]

    def __init__(self):
        self._tiempo_inicio = None
        self._crono_after_id = None
        self._crono_elapsed = 0.0
        self.lbl_cronometro = None
        self.btn_detener = None
        self._detener_flag = False
        self._tab_metodo_por_id = {}
        self._modo_vista_por_metodo = {metodo: "resultado" for metodo in self.METODOS}
        self._vista_btns_por_metodo = {}
        self._view_state = {}
        self._drag_pan_state = {}
        self._config_por_imagen = {}
        self._config_global_base = {}
        self._render_meta = {}
        self._active_canvas_key = None
        self._mousewheel_global_bound = False
        self._canvas_widgets = {}
        self._result_canvas_por_metodo = {}
        self._pasos_canvas_por_metodo = {}
        self._panel_metodo = {}
        self._falla_vars = {metodo: [] for metodo in self.METODOS}
        self._summary_cache = {"General": ""}
        self._resultados_batch = {}
        self._calibracion_unica_guardada = None
        self._config_general_scope = "actual"
        self.calibrar_cada_imagen = False
        self.lbl_cal_estado = None
        self._proyecto_workspace = None
        self._proyecto_activo_path = None
        self._rutas_origen_imagen = {}
        super().__init__()
        self._init_vars_avanzadas()
        if not self._config_global_base:
            self._sync_config()
            self._config_global_base = self._snapshot_config_avanzada()
        try:
            self.geometry("1850x1040")
        except Exception:
            pass
        try:
            self.minsize(1500, 860)
        except Exception:
            pass
        try:
            self.paned.paneconfigure(self.frame_config, minsize=220, width=250)
            self.paned.paneconfigure(self.frame_visor, minsize=900)
            self.paned.paneconfigure(self.frame_resultados, minsize=360, width=400)
        except Exception:
            pass
        self.after_idle(self._ajustar_layout_inicial)
        self._crear_menu_principal()

    def _crear_barra_superior(self):
        barra = tk.Frame(self, bg=EstiloUI.BG_PANEL, height=58)
        barra.grid(row=0, column=0, sticky="ew")
        barra.grid_propagate(False)

        tk.Label(
            barra,
            text="ANALISIS DE PAVIMENTOS | PCI + MTC + VIZIR",
            font=EstiloUI.FONT_TITLE,
            bg=EstiloUI.BG_PANEL,
            fg=EstiloUI.FG_PRIMARY,
        ).pack(side="left", padx=15)

        botones = tk.Frame(barra, bg=EstiloUI.BG_PANEL)
        botones.pack(side="right", padx=10, pady=8)

        ttk.Button(botones, text="Guardar Resultados", style="Accent.TButton", command=self._guardar_resultados).pack(side="right", padx=3)
        ttk.Button(botones, text="Exportar Solo Excel", style="Accent.TButton", command=self._exportar_solo_excel).pack(side="right", padx=3)
        ttk.Button(botones, text="Procesar Todo", style="Accent.TButton", command=self._procesar_todo).pack(side="right", padx=3)
        ttk.Button(botones, text="Procesar Actual", style="Accent.TButton", command=self._procesar_actual).pack(side="right", padx=3)
        self.lbl_cronometro = tk.Label(
            botones,
            text="--",
            width=8,
            bg=EstiloUI.BG_PANEL,
            fg=EstiloUI.FG_HIGHLIGHT,
            font=("Consolas", 10, "bold"),
        )
        self.lbl_cronometro.pack(side="right", padx=4)
        self.btn_detener = ttk.Button(botones, text="Detener", style="Secondary.TButton", command=self._detener_procesamiento, state="disabled")
        self.btn_detener.pack(side="right", padx=3)
        ttk.Button(botones, text="Cargar Imagenes", style="Secondary.TButton", command=self._cargar_imagenes).pack(side="right", padx=3)
        ttk.Button(botones, text="Cargar Modelo", style="Secondary.TButton", command=self._cargar_modelo).pack(side="right", padx=3)

    def _crear_menu_principal(self):
        menu_bar = tk.Menu(self, tearoff=False)
        menu_archivo = tk.Menu(menu_bar, tearoff=False)
        menu_archivo.add_command(label="Nuevo Proyecto", accelerator="Ctrl+N", command=self._nuevo_proyecto)
        menu_archivo.add_command(label="Abrir Proyecto...", accelerator="Ctrl+O", command=self._abrir_proyecto)
        menu_archivo.add_command(label="Guardar Proyecto", accelerator="Ctrl+S", command=self._guardar_proyecto)
        menu_bar.add_cascade(label="Archivo", menu=menu_archivo)

        menu_imagenes = tk.Menu(menu_bar, tearoff=False)
        menu_imagenes.add_command(label="Cargar Imagenes...", command=self._cargar_imagenes)
        menu_imagenes.add_command(label="Anadir Mas Imagenes...", command=self._anadir_imagenes)
        menu_bar.add_cascade(label="Imagenes", menu=menu_imagenes)

        self.config(menu=menu_bar)
        self._menu_principal = menu_bar
        self._menu_archivo = menu_archivo
        self._menu_imagenes = menu_imagenes
        self.bind_all("<Control-n>", self._shortcut_nuevo_proyecto, add="+")
        self.bind_all("<Control-N>", self._shortcut_nuevo_proyecto, add="+")
        self.bind_all("<Control-o>", self._shortcut_abrir_proyecto, add="+")
        self.bind_all("<Control-O>", self._shortcut_abrir_proyecto, add="+")
        self.bind_all("<Control-s>", self._shortcut_guardar_proyecto, add="+")
        self.bind_all("<Control-S>", self._shortcut_guardar_proyecto, add="+")

    def _shortcut_nuevo_proyecto(self, event=None):
        self._nuevo_proyecto()
        return "break"

    def _shortcut_abrir_proyecto(self, event=None):
        self._abrir_proyecto()
        return "break"

    def _shortcut_guardar_proyecto(self, event=None):
        self._guardar_proyecto()
        return "break"

    def _actualizar_ui_modelo(self):
        ruta_modelo = str(self.motor.config.get("ruta_modelo", "") or "").strip()
        if self.motor.modelo_cargado and self.motor.modelo is not None and ruta_modelo:
            self.lbl_modelo.config(text=f"Modelo: {Path(ruta_modelo).name}", fg=EstiloUI.BG_SUCCESS)
            self.lbl_modelo_path.config(text=f"Modelo: {ruta_modelo}")
            return
        self.lbl_modelo.config(text="Modelo: no cargado", fg=EstiloUI.FG_HIGHLIGHT)
        self.lbl_modelo_path.config(text="Modelo: no cargado")

    def _mostrar_progreso_proyecto(self, texto="Cargando proyecto...", valor=0.0):
        if not hasattr(self, "project_progress_frame"):
            return
        self.project_progress_label.config(text=texto)
        self.project_progress.configure(mode="determinate", maximum=100)
        self.project_progress["value"] = max(0.0, min(100.0, float(valor)))
        if not self.project_progress_frame.winfo_manager():
            self.project_progress_frame.pack(side='left', padx=(4, 8))
        self.update_idletasks()

    def _actualizar_progreso_proyecto(self, valor=None, texto=None):
        if not hasattr(self, "project_progress_frame"):
            return
        if texto is not None:
            self.project_progress_label.config(text=texto)
        if valor is not None:
            self.project_progress["value"] = max(0.0, min(100.0, float(valor)))
        self.update_idletasks()

    def _ocultar_progreso_proyecto(self):
        if not hasattr(self, "project_progress_frame"):
            return
        self.project_progress.stop()
        self.project_progress["value"] = 0
        try:
            self.project_progress_frame.pack_forget()
        except Exception:
            pass
        self.update_idletasks()

    def _limpiar_ui_proyecto(self):
        self._summary_cache = {"General": "Sin proyecto cargado."}
        self._set_text("General", self._summary_cache["General"])
        for metodo in self.METODOS:
            self._summary_cache[metodo] = "Sin resultados aun."
            panel = self._panel_metodo[metodo]
            panel["summary_label"].config(text=self._summary_cache[metodo])
            self._reconstruir_lista_fallas(metodo, [])
        self._actualizar_datos_imagen_panel()
        self._ajustar_vistas_resultado()
        self._imagenes_render["Original"] = None
        self._render_single_view("Original")
        for metodo in self.METODOS:
            self._imagenes_render[metodo] = None
            self._imagenes_render[f"{metodo}_Pasos"] = None
            self._cambiar_modo_vista_metodo(metodo, "resultado")
            self._render_single_view(metodo)
        self._seleccionar_metodo_en_ui("Original")

    def _hay_estado_proyecto(self):
        return bool(
            self.imagenes_cargadas
            or self._resultados_batch
            or self.resultado_actual
            or self._proyecto_activo_path
        )

    def _nuevo_proyecto(self):
        if self.procesando:
            messagebox.showinfo("Proyecto", "Espere a que termine el procesamiento antes de crear un proyecto nuevo.")
            return
        if self._hay_estado_proyecto():
            continuar = messagebox.askyesno(
                "Nuevo proyecto",
                "Se limpiara el proyecto actual y los cambios no guardados se perderan.\n\nDesea continuar?",
            )
            if not continuar:
                return

        modelo_actual = self.motor.modelo if self.motor.modelo_cargado else None
        modelo_cargado = bool(self.motor.modelo_cargado and modelo_actual is not None)
        ruta_modelo = str(self.motor.config.get("ruta_modelo", "") or "")

        self._liberar_proyecto_activo()
        self.motor = MotorIntegradoTresMetodos()
        if modelo_cargado:
            self.motor.modelo = modelo_actual
            self.motor.modelo_cargado = True
            if ruta_modelo:
                self.motor.config["ruta_modelo"] = ruta_modelo

        self.imagenes_cargadas = []
        self.imagen_actual_idx = -1
        self.resultado_actual = None
        self._resultados_batch = {}
        self._rutas_origen_imagen = {}
        self._config_por_imagen.clear()
        self._config_global_base = {}
        self._config_general_scope = "actual"
        if hasattr(self, "var_config_general_scope"):
            self.var_config_general_scope.set(self._config_general_scope)
            self._actualizar_scope_general_ui()
        self._view_state.clear()
        self._drag_pan_state.clear()
        self._calibracion_unica_guardada = None
        self._detener_flag = False
        self.procesando = False
        self._detener_cronometro_integrado()
        self._set_detener_enabled(False)

        if isinstance(self.vars.get("modo_calibracion"), tk.StringVar):
            self.vars["modo_calibracion"].set("automatica")
        self._aplicar_config_en_vars(copy.deepcopy(self.motor.config))
        self._sync_config()
        self._config_global_base = self._snapshot_config_avanzada()
        self.motor.todos_resultados = {}
        self._actualizar_nav()
        self._actualizar_ui_modelo()
        self._actualizar_estado_calibracion_ui()
        self._ocultar_progreso_proyecto()
        self._limpiar_ui_proyecto()
        self._estado("Proyecto nuevo listo")
        self._log("Proyecto nuevo creado.")

    def destroy(self):
        try:
            self._liberar_proyecto_activo()
        finally:
            super().destroy()

    def _set_detener_enabled(self, enabled):
        if self.btn_detener is None:
            return
        try:
            self.btn_detener.config(state="normal" if enabled else "disabled")
        except Exception:
            pass

    def _detener_procesamiento(self):
        if not self.procesando:
            self._set_detener_enabled(False)
            return
        self._detener_flag = True
        self._set_detener_enabled(False)
        self._estado("Deteniendo...")
        self._log("Solicitud de detencion registrada. Se detendra al finalizar la imagen en curso.")

    def _ajustar_layout_inicial(self):
        if not hasattr(self, "paned"):
            return
        try:
            total = max(self.winfo_width(), 1850)
            left = 250
            right = max(total - 400, left + 850)
            self.paned.sash_place(0, left, 1)
            self.paned.sash_place(1, right, 1)
        except Exception:
            pass

    def _formato_tiempo_integrado(self, elapsed):
        mins = int(elapsed // 60)
        secs = elapsed % 60
        if mins > 0:
            return f"{mins}:{secs:04.1f}"
        return f"{secs:.1f}s"

    def _iniciar_cronometro_integrado(self):
        self._tiempo_inicio = time.time()
        self._crono_elapsed = 0.0
        if self.lbl_cronometro is not None:
            self.lbl_cronometro.config(text="0.0s")
        self._actualizar_cronometro_integrado()

    def _actualizar_cronometro_integrado(self):
        if not self.procesando or not self._tiempo_inicio or self.lbl_cronometro is None:
            return
        self._crono_elapsed = time.time() - self._tiempo_inicio
        self.lbl_cronometro.config(text=self._formato_tiempo_integrado(self._crono_elapsed))
        self._crono_after_id = self.after(100, self._actualizar_cronometro_integrado)

    def _detener_cronometro_integrado(self):
        if self._crono_after_id:
            self.after_cancel(self._crono_after_id)
            self._crono_after_id = None
        if self._tiempo_inicio:
            self._crono_elapsed = time.time() - self._tiempo_inicio
        if self.lbl_cronometro is not None:
            self.lbl_cronometro.config(text=self._formato_tiempo_integrado(self._crono_elapsed) if self._crono_elapsed else "--")
        self._tiempo_inicio = None
        return self._crono_elapsed

    def _crear_var(self, key, default):
        actual = self.vars.get(key)
        if isinstance(default, bool):
            if not isinstance(actual, tk.BooleanVar):
                actual = tk.BooleanVar(value=default)
                self.vars[key] = actual
        elif isinstance(default, int) and not isinstance(default, bool):
            if not isinstance(actual, tk.IntVar):
                actual = tk.IntVar(value=int(default))
                self.vars[key] = actual
        elif isinstance(default, str):
            if not isinstance(actual, tk.StringVar):
                actual = tk.StringVar(value=default)
                self.vars[key] = actual
        else:
            if not isinstance(actual, tk.DoubleVar):
                actual = tk.DoubleVar(value=float(default))
                self.vars[key] = actual
        return actual

    def _init_vars_avanzadas(self):
        defaults = {
            "filtrar_baches": True,
            "min_diametro_hueco_mm": 50.0,
            "filtrar_grietas": True,
            "min_longitud_grieta_m": 0.05,
            "usar_borde_berma_pci": True,
            "borde_interno_m": 0.30,
            "borde_externo_m": 0.30,
            "filtrar_parches": True,
            "min_area_parche_m2": 0.01,
            "parche_ratio_leve_max": 0.08,
            "parche_ratio_moderado_max": 0.18,
            "filtrar_piel": True,
            "min_area_piel_m2": 0.05,
            "merge_iou_threshold": 0.10,
            "merge_distancia_max_px": 50,
            "merge_fisuras_px": 30,
            "clahe_clip": 4.0,
            "clahe_tile": 8,
            "bilateral_d": 9,
            "bilateral_sigma_color": 75,
            "bilateral_sigma_space": 75,
            "block_size": 23,
            "C_umbral": 10,
            "usar_frangi": True,
            "usar_refinamiento": True,
            "usar_multiescala": True,
            "kernel_apertura": 3,
            "kernel_cierre": 6,
            "iteraciones_cierre": 2,
            "min_area_poligono": 300,
            "min_circularidad": 0.08,
            "min_vertices": 4,
            "max_vertices": 25,
            "min_radio_circulo": 8,
            "min_longitud_rama": 30,
            "min_area_objeto": 100,
            "max_gap_cierre": 20,
        }
        for key, default in defaults.items():
            self._crear_var(key, self.motor.config.get(key, default))
        self._crear_var("profundidad_asumida_huecos", self.motor.config.get("profundidad_asumida_huecos", "media (25-50mm)"))

    def _crear_panel_config(self):
        self.frame_config.grid_rowconfigure(0, weight=1)
        self.frame_config.grid_columnconfigure(0, weight=1)
        canvas = tk.Canvas(self.frame_config, bg=EstiloUI.BG_PANEL, highlightthickness=0)
        scroll = ttk.Scrollbar(self.frame_config, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg=EstiloUI.BG_PANEL)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scroll.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        scroll.grid(row=0, column=1, sticky="ns")

        self._sec(inner, "CONFIGURACION GENERAL")
        self._slider(inner, "ancho_via_real_m", "Ancho de via (m)", 3.0, 12.0, self.motor.config.get("ancho_via_real_m", 6.5), 0.1)
        self._slider(inner, "confianza_min", "Confianza minima", 0.05, 0.95, self.motor.config.get("confianza_min", 0.1), 0.05)
        self._slider(inner, "iou_threshold", "IoU threshold", 0.10, 0.95, self.motor.config.get("iou_threshold", 0.45), 0.05)
        scope_box = tk.Frame(inner, bg=EstiloUI.BG_PANEL)
        scope_box.pack(fill="x", padx=10, pady=(2, 2))
        tk.Label(
            scope_box,
            text="Aplicar configuracion general a:",
            bg=EstiloUI.BG_PANEL,
            fg=EstiloUI.FG_PRIMARY,
            font=EstiloUI.FONT_SMALL,
        ).pack(anchor="w")
        self.var_config_general_scope = tk.StringVar(value=self._config_general_scope)
        for txt, val in [
            ("Solo esta imagen", "actual"),
            ("Todas las imagenes", "todas"),
        ]:
            tk.Radiobutton(
                scope_box,
                text=txt,
                variable=self.var_config_general_scope,
                value=val,
                bg=EstiloUI.BG_PANEL,
                fg=EstiloUI.FG_PRIMARY,
                selectcolor=EstiloUI.BG_DARK,
                activebackground=EstiloUI.BG_PANEL,
                font=EstiloUI.FONT_SMALL,
                command=self._on_config_general_scope_change,
            ).pack(anchor="w", pady=1)
        ttk.Button(
            inner,
            text="Aplicar configuracion general",
            style="Secondary.TButton",
            command=self._aplicar_config_general_desde_panel,
        ).pack(fill="x", padx=10, pady=(0, 4))
        self.lbl_scope_general = tk.Label(
            inner,
            text="",
            justify="left",
            wraplength=255,
            bg=EstiloUI.BG_PANEL,
            fg=EstiloUI.FG_HIGHLIGHT,
            font=EstiloUI.FONT_SMALL,
        )
        self.lbl_scope_general.pack(fill="x", padx=10, pady=(0, 2))
        self._actualizar_scope_general_ui()

        self._sec(inner, "CALIBRACION")
        modo_var = self.vars.get("modo_calibracion")
        if not isinstance(modo_var, tk.StringVar):
            modo_var = tk.StringVar(value="automatica")
            self.vars["modo_calibracion"] = modo_var
        for texto, valor in [
            ("Automatica (ancho imagen)", "automatica"),
            ("Calibrar cada imagen (OpenCV)", "cada_imagen"),
            ("Calibracion unica (1 ejemplar)", "unica"),
        ]:
            tk.Radiobutton(
                inner,
                text=texto,
                variable=modo_var,
                value=valor,
                bg=EstiloUI.BG_PANEL,
                fg=EstiloUI.FG_PRIMARY,
                selectcolor=EstiloUI.BG_DARK,
                activebackground=EstiloUI.BG_PANEL,
                font=EstiloUI.FONT_SMALL,
                command=self._actualizar_estado_calibracion_ui,
            ).pack(anchor="w", padx=10, pady=1)
        btn_cal = tk.Frame(inner, bg=EstiloUI.BG_PANEL)
        btn_cal.pack(fill="x", padx=10, pady=(3, 2))
        ttk.Button(
            btn_cal,
            text="Calibrar ahora (1 imagen)",
            style="Secondary.TButton",
            command=self._calibrar_unica,
        ).pack(fill="x")
        self.lbl_cal_estado = tk.Label(
            inner,
            text="Sin calibracion unica",
            justify="left",
            wraplength=255,
            bg=EstiloUI.BG_PANEL,
            fg=EstiloUI.FG_SECONDARY,
            font=EstiloUI.FONT_SMALL,
        )
        self.lbl_cal_estado.pack(fill="x", padx=10, pady=(1, 2))

        self._sec(inner, "VISUALIZACION")
        for key, text, default in [
            ("show_etiquetas", "Mostrar etiquetas", True),
            ("show_numeros", "Mostrar valores numericos", True),
            ("show_mallas", "Mostrar mallas / heatmaps", True),
            ("show_circulos", "Mostrar circulos auxiliares", True),
            ("show_poligonos", "Mostrar poligonos piel cocodrilo", True),
        ]:
            var = self.vars.get(key)
            if not isinstance(var, tk.BooleanVar):
                var = tk.BooleanVar(value=self.motor.config.get(key, default))
                self.vars[key] = var
            tk.Checkbutton(
                inner,
                text=text,
                variable=var,
                bg=EstiloUI.BG_PANEL,
                fg=EstiloUI.FG_PRIMARY,
                selectcolor=EstiloUI.BG_DARK,
                activebackground=EstiloUI.BG_PANEL,
                font=EstiloUI.FONT_BODY,
                command=self._redibujar_todas_las_vistas,
            ).pack(anchor="w", padx=10, pady=2)

        self._sec(inner, "CONFIGURACION AVANZADA")
        ttk.Button(inner, text="Abrir configuracion avanzada", style="Accent.TButton", command=self._abrir_config_avanzada).pack(fill="x", padx=10, pady=4)
        for texto, cmd in [
            ("Filtros tamano minimo", self._abrir_filtros),
            ("Fusion solapamientos", self._abrir_fusion),
            ("Piel de cocodrilo", self._abrir_piel_config),
            ("Deteccion poligonos", self._abrir_poligonos),
            ("Perfiles rapidos", self._abrir_perfiles),
        ]:
            ttk.Button(inner, text=texto, style="Secondary.TButton", command=cmd).pack(fill="x", padx=10, pady=2)

        tk.Label(
            inner,
            text="Click sobre una falla en PCI, MTC o VIZIR para ocultarla o volverla a mostrar.",
            justify="left",
            wraplength=255,
            bg=EstiloUI.BG_PANEL,
            fg=EstiloUI.FG_HIGHLIGHT,
            font=EstiloUI.FONT_SMALL,
        ).pack(fill="x", padx=10, pady=(8, 4))

        self._sec(inner, "RUTAS")
        self.lbl_modelo_path = tk.Label(inner, text="Modelo: no cargado", justify="left", wraplength=255, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_SECONDARY, font=EstiloUI.FONT_SMALL)
        self.lbl_modelo_path.pack(fill="x", padx=10, pady=2)
        self.lbl_calibracion = tk.Label(inner, text="Calibracion: automatica por ancho de imagen", justify="left", wraplength=255, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_SECONDARY, font=EstiloUI.FONT_SMALL)
        self.lbl_calibracion.pack(fill="x", padx=10, pady=2)
        self._actualizar_estado_calibracion_ui()

    def _crear_panel_visor(self):
        self.frame_visor.grid_rowconfigure(2, weight=1)
        self.frame_visor.grid_columnconfigure(0, weight=1)

        nav = tk.Frame(self.frame_visor, bg=EstiloUI.BG_CARD, height=46)
        nav.grid(row=0, column=0, sticky="ew", pady=(0, 2))
        nav.grid_propagate(False)
        ttk.Button(nav, text="< Anterior", style="Secondary.TButton", command=self._imagen_anterior).pack(side="left", padx=5, pady=4)
        ttk.Button(nav, text="Siguiente >", style="Secondary.TButton", command=self._imagen_siguiente).pack(side="left", padx=5, pady=4)
        self.lbl_nav = tk.Label(nav, text="Sin imagenes cargadas", bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_PRIMARY, font=EstiloUI.FONT_BODY)
        self.lbl_nav.pack(side="left", padx=12, fill="x", expand=True)
        zoom_box = tk.Frame(nav, bg=EstiloUI.BG_CARD)
        zoom_box.pack(side="right", padx=8, pady=4)
        tk.Label(zoom_box, text="Zoom", bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_SECONDARY, font=EstiloUI.FONT_SMALL).pack(side="left", padx=(0, 4))
        ttk.Button(zoom_box, text="-", style="Secondary.TButton", width=3, command=self._zoom_out_current_view).pack(side="left", padx=(0, 2))
        self.zoom_scale_var = tk.DoubleVar(value=100.0)
        self.zoom_scale = ttk.Scale(
            zoom_box,
            from_=5.0,
            to=1200.0,
            orient="horizontal",
            variable=self.zoom_scale_var,
            command=self._on_zoom_scale_change,
            length=170,
        )
        self.zoom_scale.pack(side="left", padx=2)
        ttk.Button(zoom_box, text="+", style="Secondary.TButton", width=3, command=self._zoom_in_current_view).pack(side="left", padx=2)
        ttk.Button(zoom_box, text="Ajustar", style="Secondary.TButton", command=self._zoom_fit_current_view).pack(side="left", padx=(4, 2))
        ttk.Button(zoom_box, text="1:1", style="Secondary.TButton", command=self._zoom_100_current_view).pack(side="left", padx=(0, 4))
        self.lbl_zoom_pct = tk.Label(zoom_box, text="100%", bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_HIGHLIGHT, font=EstiloUI.FONT_LABEL, width=7, anchor="e")
        self.lbl_zoom_pct.pack(side="left", padx=(4, 0))
        self._zoom_ui_updating = False

        info = tk.Frame(self.frame_visor, bg=EstiloUI.BG_PANEL, height=26)
        info.grid(row=1, column=0, sticky="ew", pady=(0, 2))
        info.grid_propagate(False)
        self.lbl_info_img = tk.Label(info, text="--", bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT, font=EstiloUI.FONT_SMALL, anchor="w")
        self.lbl_info_img.pack(side="left", padx=8, fill="x", expand=True)

        notebook = ttk.Notebook(self.frame_visor, style="View.TNotebook")
        notebook.grid(row=2, column=0, sticky="nsew")
        self.notebook_vistas = notebook
        self._tab_metodo_por_id = {}
        self.view_mode_bar = tk.Frame(self.frame_visor, bg=EstiloUI.BG_DARK)
        self.btn_view_resultado = ttk.Button(
            self.view_mode_bar,
            text="Resultado",
            style="ViewMode.TButton",
            width=10,
            command=lambda: self._cambiar_modo_vista_actual("resultado"),
        )
        self.btn_view_resultado.pack(side="left", padx=(0, 3))
        self.btn_view_pasos = ttk.Button(
            self.view_mode_bar,
            text="Pasos",
            style="ViewMode.TButton",
            width=10,
            command=lambda: self._cambiar_modo_vista_actual("pasos"),
        )
        self.btn_view_pasos.pack(side="left")

        frame_original = tk.Frame(notebook, bg=EstiloUI.BG_INPUT)
        frame_original.grid_rowconfigure(0, weight=1)
        frame_original.grid_columnconfigure(0, weight=1)
        canvas_original = tk.Canvas(frame_original, bg=EstiloUI.BG_INPUT, highlightthickness=0)
        canvas_original.grid(row=0, column=0, sticky="nsew")
        canvas_original.bind("<Configure>", lambda e, key="Original": self._render_single_view(key))
        self._bind_canvas_zoom_pan(canvas_original, "Original")
        notebook.add(frame_original, text="Original")
        self._canvas_widgets["Original"] = canvas_original
        self._tab_metodo_por_id[str(frame_original)] = "Original"

        for metodo in self.METODOS:
            frame = tk.Frame(notebook, bg=EstiloUI.BG_DARK)
            frame.grid_rowconfigure(0, weight=1)
            frame.grid_columnconfigure(0, weight=1)

            canvas_res = tk.Canvas(frame, bg=EstiloUI.BG_INPUT, highlightthickness=0)
            canvas_res.grid(row=0, column=0, sticky="nsew")
            canvas_res.bind("<Configure>", lambda e, key=metodo: self._render_single_view(key))
            canvas_res.bind("<Button-1>", lambda e, m=metodo: self._on_result_canvas_click(e, m))
            self._bind_canvas_zoom_pan(canvas_res, metodo)

            notebook.add(frame, text=metodo)

            self._canvas_widgets[metodo] = canvas_res
            self._result_canvas_por_metodo[metodo] = canvas_res
            self._tab_metodo_por_id[str(frame)] = metodo
            self._cambiar_modo_vista_metodo(metodo, self._modo_vista_por_metodo.get(metodo, "resultado"))

        notebook.bind("<<NotebookTabChanged>>", self._on_view_tab_changed)
        self.after_idle(self._actualizar_barra_modo_vista)

    def _crear_panel_resultados(self):
        panel_padx = (14, 6)
        self.frame_resultados.grid_rowconfigure(2, weight=3)
        self.frame_resultados.grid_rowconfigure(4, weight=2)
        self.frame_resultados.grid_columnconfigure(0, weight=1)

        info_img = tk.Frame(self.frame_resultados, bg=EstiloUI.BG_CARD)
        info_img.grid(row=0, column=0, sticky="ew", padx=panel_padx, pady=(5, 2))
        tk.Label(info_img, text="DATOS DE LA IMAGEN", bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_ACCENT, font=EstiloUI.FONT_SUBTITLE).pack(anchor="w", padx=10, pady=(6, 2))
        self.lbl_datos_imagen = tk.Label(
            info_img,
            text="Sin imagen cargada.",
            justify="left",
            anchor="nw",
            wraplength=330,
            bg=EstiloUI.BG_CARD,
            fg=EstiloUI.FG_PRIMARY,
            font=EstiloUI.FONT_SMALL,
        )
        self.lbl_datos_imagen.pack(fill="x", padx=10, pady=(0, 8))

        head = tk.Frame(self.frame_resultados, bg=EstiloUI.BG_CARD)
        head.grid(row=1, column=0, sticky="ew", padx=panel_padx, pady=(2, 2))
        tk.Label(head, text="RESUMENES", bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_ACCENT, font=EstiloUI.FONT_SUBTITLE).pack(anchor="w", padx=10, pady=6)

        nb = ttk.Notebook(self.frame_resultados, style="View.TNotebook")
        nb.grid(row=2, column=0, sticky="nsew", padx=panel_padx, pady=(0, 5))
        self.summary_notebook = nb
        self._summary_tab_by_key = {}
        self.summary_texts = {}

        general_frame = tk.Frame(nb, bg=EstiloUI.BG_CARD)
        txt = scrolledtext.ScrolledText(
            general_frame,
            wrap=tk.WORD,
            font=EstiloUI.FONT_MONO,
            bg=EstiloUI.BG_INPUT,
            fg=EstiloUI.FG_PRIMARY,
            insertbackground=EstiloUI.FG_PRIMARY,
            relief="flat",
        )
        txt.pack(fill="both", expand=True)
        self._bind_text_mousewheel(txt)
        nb.add(general_frame, text="General")
        self.summary_texts["General"] = txt
        self._summary_tab_by_key["Original"] = general_frame
        self._summary_tab_by_key["General"] = general_frame

        for metodo in self.METODOS:
            frame = tk.Frame(nb, bg=EstiloUI.BG_CARD)
            frame.grid_rowconfigure(2, weight=1)
            frame.grid_columnconfigure(0, weight=1)

            resumen_lbl = tk.Label(frame, text="Sin resultados aun.", justify="left", anchor="nw", wraplength=330, bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_PRIMARY, font=EstiloUI.FONT_SMALL)
            resumen_lbl.grid(row=0, column=0, sticky="ew", padx=8, pady=(8, 4))

            controles = tk.Frame(frame, bg=EstiloUI.BG_CARD)
            controles.grid(row=1, column=0, sticky="ew", padx=8, pady=(0, 4))
            ttk.Button(controles, text="Marcar todo", style="Secondary.TButton", command=lambda m=metodo: self._set_method_all(m, True)).pack(side="left", padx=(0, 4))
            ttk.Button(controles, text="Quitar todo", style="Secondary.TButton", command=lambda m=metodo: self._set_method_all(m, False)).pack(side="left", padx=4)
            count_lbl = tk.Label(controles, text="0/0 activas", bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_HIGHLIGHT, font=EstiloUI.FONT_LABEL)
            count_lbl.pack(side="right")

            list_host = tk.Frame(
                frame,
                bg=EstiloUI.BG_CARD,
                highlightbackground=EstiloUI.BORDER,
                highlightthickness=1,
                bd=0,
            )
            list_host.grid(row=2, column=0, sticky="nsew", padx=6, pady=(0, 6))
            list_host.grid_rowconfigure(0, weight=1)
            list_host.grid_columnconfigure(0, weight=1)
            list_canvas = tk.Canvas(list_host, bg=EstiloUI.BG_CARD, highlightthickness=0)
            list_scroll = ttk.Scrollbar(list_host, orient="vertical", command=list_canvas.yview)
            list_scroll_x = ttk.Scrollbar(list_host, orient="horizontal", command=list_canvas.xview)
            list_inner = tk.Frame(list_canvas, bg=EstiloUI.BG_CARD)
            list_inner.bind("<Configure>", lambda e, c=list_canvas: c.configure(scrollregion=c.bbox("all")))
            list_canvas.create_window((0, 0), window=list_inner, anchor="nw")
            list_canvas.configure(yscrollcommand=list_scroll.set, xscrollcommand=list_scroll_x.set)
            list_canvas.grid(row=0, column=0, sticky="nsew")
            list_scroll.grid(row=0, column=1, sticky="ns")
            list_scroll_x.grid(row=1, column=0, sticky="ew")
            empty_lbl = tk.Label(list_inner, text="Sin fallas detectadas.", bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_SECONDARY, font=EstiloUI.FONT_SMALL)
            empty_lbl.pack(anchor="w", padx=6, pady=6)
            self._bind_summary_mousewheel(frame, list_canvas)
            self._bind_summary_mousewheel(resumen_lbl, list_canvas)
            self._bind_summary_mousewheel(controles, list_canvas)
            self._bind_summary_mousewheel(count_lbl, list_canvas)
            self._bind_summary_mousewheel(list_canvas, list_canvas)
            self._bind_summary_mousewheel(list_inner, list_canvas)
            self._bind_summary_mousewheel(empty_lbl, list_canvas)

            nb.add(frame, text=metodo)
            self._summary_tab_by_key[metodo] = frame
            self._panel_metodo[metodo] = {
                "summary_label": resumen_lbl,
                "count_label": count_lbl,
                "list_canvas": list_canvas,
                "list_inner": list_inner,
                "list_hscroll": list_scroll_x,
                "empty_label": empty_lbl,
                "rows": [],
            }

        log_head = tk.Frame(self.frame_resultados, bg=EstiloUI.BG_CARD)
        log_head.grid(row=3, column=0, sticky="ew", padx=panel_padx, pady=(2, 2))
        tk.Label(log_head, text="LOG", bg=EstiloUI.BG_CARD, fg=EstiloUI.FG_ACCENT, font=EstiloUI.FONT_SUBTITLE).pack(anchor="w", padx=10, pady=6)

        self.log_text = scrolledtext.ScrolledText(
            self.frame_resultados,
            wrap=tk.WORD,
            font=EstiloUI.FONT_MONO,
            bg=EstiloUI.BG_INPUT,
            fg=EstiloUI.FG_LOG,
            insertbackground=EstiloUI.FG_LOG,
            relief="flat",
        )
        self.log_text.grid(row=4, column=0, sticky="nsew", padx=panel_padx, pady=(0, 5))
        self._bind_text_mousewheel(self.log_text)

    def _mousewheel_units(self, event):
        delta = getattr(event, "delta", 0)
        if delta > 0:
            return -1
        if delta < 0:
            return 1
        num = getattr(event, "num", None)
        if num == 4:
            return -1
        if num == 5:
            return 1
        return 0

    def _bind_mousewheel_handler(self, widget, handler):
        widget.bind("<MouseWheel>", handler, add="+")
        widget.bind("<Button-4>", handler, add="+")
        widget.bind("<Button-5>", handler, add="+")

    def _bind_text_mousewheel(self, widget):
        def _on_scroll(event, target=widget):
            units = self._mousewheel_units(event)
            if units:
                target.yview_scroll(units, "units")
            return "break"

        self._bind_mousewheel_handler(widget, _on_scroll)

    def _bind_summary_mousewheel(self, widget, canvas):
        def _on_scroll(event, target=canvas):
            units = self._mousewheel_units(event)
            if units:
                if bool(getattr(event, 'state', 0) & 0x0001):
                    target.xview_scroll(units, "units")
                else:
                    target.yview_scroll(units, "units")
            return "break"

        self._bind_mousewheel_handler(widget, _on_scroll)

    def _seccion_titulo(self, parent, texto):
        self._sec(parent, texto)

    def _sync_config(self):
        for key, var in self.vars.items():
            if key not in self.motor.config and key != "profundidad_asumida_huecos":
                continue
            if isinstance(var, tk.BooleanVar):
                self.motor.config[key] = bool(var.get())
            elif isinstance(var, tk.IntVar):
                self.motor.config[key] = int(var.get())
            elif isinstance(var, tk.StringVar):
                if key == "profundidad_asumida_huecos":
                    texto = str(var.get()).lower()
                    if "baja" in texto:
                        self.motor.config[key] = "baja"
                    elif "alta" in texto:
                        self.motor.config[key] = "alta"
                    else:
                        self.motor.config[key] = "media"
                else:
                    self.motor.config[key] = str(var.get())
            else:
                base = self.motor.config.get(key)
                valor = float(var.get())
                if isinstance(base, int) and not isinstance(base, bool):
                    self.motor.config[key] = int(round(valor))
                else:
                    self.motor.config[key] = valor

        bs = int(self.motor.config.get("block_size", 23))
        bs = max(11, min(51, bs))
        if bs % 2 == 0:
            bs += 1 if bs < 51 else -1
        self.motor.config["block_size"] = bs

        self.motor.config["clahe_tile"] = max(4, min(16, int(self.motor.config.get("clahe_tile", 8))))
        self.motor.config["bilateral_d"] = max(3, min(15, int(self.motor.config.get("bilateral_d", 9))))
        self.motor.config["kernel_apertura"] = max(2, min(7, int(self.motor.config.get("kernel_apertura", 3))))
        self.motor.config["kernel_cierre"] = max(3, min(12, int(self.motor.config.get("kernel_cierre", 6))))
        self.motor.config["iteraciones_cierre"] = max(1, min(5, int(self.motor.config.get("iteraciones_cierre", 2))))
        self.motor.config["min_area_poligono"] = max(50, int(self.motor.config.get("min_area_poligono", 300)))
        self.motor.config["min_area_objeto"] = max(20, int(self.motor.config.get("min_area_objeto", 100)))
        self.motor.config["min_longitud_rama"] = max(10, int(self.motor.config.get("min_longitud_rama", 30)))
        self.motor.config["min_vertices"] = max(3, int(self.motor.config.get("min_vertices", 4)))
        self.motor.config["max_vertices"] = max(self.motor.config["min_vertices"], int(self.motor.config.get("max_vertices", 25)))
        self.motor.config["borde_interno_m"] = max(0.0, min(3.0, float(self.motor.config.get("borde_interno_m", 0.30))))
        self.motor.config["borde_externo_m"] = max(0.0, min(3.0, float(self.motor.config.get("borde_externo_m", 0.30))))
        ratio_leve, ratio_moderado = normalizar_umbrales_parche(
            self.motor.config.get("parche_ratio_leve_max", 0.08),
            self.motor.config.get("parche_ratio_moderado_max", 0.18),
        )
        self.motor.config["parche_ratio_leve_max"] = ratio_leve
        self.motor.config["parche_ratio_moderado_max"] = ratio_moderado

        self.motor.calibrador.ancho_via_real_m = self.motor.config.get("ancho_via_real_m", 6.5)

    def _bind_canvas_zoom_pan(self, canvas, key):
        canvas.bind("<Enter>", lambda e, c=canvas, k=key: self._activar_canvas_zoom(c, k))
        canvas.bind("<Motion>", lambda e, c=canvas, k=key: self._activar_canvas_zoom(c, k))
        canvas.bind("<Leave>", lambda e, k=key: self._desactivar_canvas_zoom(k))
        canvas.bind("<MouseWheel>", lambda e, k=key: self._on_canvas_wheel(e, k))
        canvas.bind("<Button-4>", lambda e, k=key: self._on_canvas_wheel_linux_local(e, k, True))
        canvas.bind("<Button-5>", lambda e, k=key: self._on_canvas_wheel_linux_local(e, k, False))
        canvas.bind("<ButtonPress-2>", lambda e, k=key: self._on_canvas_pan_start(e, k))
        canvas.bind("<B2-Motion>", lambda e, k=key: self._on_canvas_pan_move(e, k))
        canvas.bind("<ButtonRelease-2>", lambda e, k=key: self._on_canvas_pan_end(e, k))
        canvas.bind("<ButtonPress-3>", lambda e, k=key: self._on_canvas_pan_start(e, k))
        canvas.bind("<B3-Motion>", lambda e, k=key: self._on_canvas_pan_move(e, k))
        canvas.bind("<ButtonRelease-3>", lambda e, k=key: self._on_canvas_pan_end(e, k))
        canvas.bind("<Shift-ButtonPress-1>", lambda e, k=key: self._on_canvas_pan_start(e, k))
        canvas.bind("<Shift-B1-Motion>", lambda e, k=key: self._on_canvas_pan_move(e, k))
        canvas.bind("<Shift-ButtonRelease-1>", lambda e, k=key: self._on_canvas_pan_end(e, k))
        canvas.bind("<Control-ButtonPress-1>", lambda e, k=key: self._on_canvas_pan_start(e, k))
        canvas.bind("<Control-B1-Motion>", lambda e, k=key: self._on_canvas_pan_move(e, k))
        canvas.bind("<Control-ButtonRelease-1>", lambda e, k=key: self._on_canvas_pan_end(e, k))
        canvas.bind("<Double-Button-2>", lambda e, k=key: self._reset_canvas_zoom(k))
        if not self._mousewheel_global_bound:
            self.bind_all("<MouseWheel>", self._on_canvas_wheel_global, add="+")
            self.bind_all("<Button-4>", lambda e: self._on_canvas_wheel_global(e, zoom_in=True), add="+")
            self.bind_all("<Button-5>", lambda e: self._on_canvas_wheel_global(e, zoom_in=False), add="+")
            self._mousewheel_global_bound = True

    def _on_canvas_wheel_linux_local(self, event, key, zoom_in):
        class _Evt:
            pass
        evt = _Evt()
        evt.x = getattr(event, "x", 0)
        evt.y = getattr(event, "y", 0)
        evt.state = getattr(event, "state", 0)
        evt.delta = 120 if zoom_in else -120
        return self._on_canvas_wheel(evt, key)

    def _activar_canvas_zoom(self, canvas, key):
        self._active_canvas_key = key
        try:
            canvas.focus_set()
            if not self._drag_pan_state.get(self._state_key_for_canvas(key)):
                canvas.configure(cursor="crosshair")
        except Exception:
            pass

    def _desactivar_canvas_zoom(self, key):
        if self._active_canvas_key == key:
            self._active_canvas_key = None
        canvas = self._canvas_widgets.get(key)
        if canvas is not None:
            try:
                canvas.configure(cursor="")
            except Exception:
                pass

    def _on_canvas_wheel_global(self, event, zoom_in=None):
        widget = None
        try:
            widget = self.winfo_containing(self.winfo_pointerx(), self.winfo_pointery())
        except Exception:
            widget = getattr(event, "widget", None)
        found_key = None
        while widget is not None:
            for candidate_key, canvas in self._canvas_widgets.items():
                if widget is canvas:
                    found_key = candidate_key
                    break
            if found_key is not None:
                break
            try:
                widget = widget.master
            except Exception:
                widget = None
        if found_key is None:
            return
        key = found_key
        self._active_canvas_key = found_key
        if zoom_in is None:
            delta = getattr(event, "delta", 0)
            if delta == 0:
                return
            self._on_canvas_wheel(event, key)
        else:
            class _Evt:
                pass
            evt = _Evt()
            evt.x = getattr(event, "x", 0)
            evt.y = getattr(event, "y", 0)
            evt.state = getattr(event, "state", 0)
            evt.delta = 120 if zoom_in else -120
            self._on_canvas_wheel(evt, key)
        return "break"

    def _get_view_state(self, key):
        return self._view_state.setdefault(
            key,
            {
                "zoom_level": None,
                "zoom_min": 0.05,
                "zoom_max": 12.0,
                "pan_offset_x": 0.0,
                "pan_offset_y": 0.0,
                "image_size": None,
            },
        )

    def _state_key_for_canvas(self, key):
        if key in self.METODOS and self._modo_vista_por_metodo.get(key, "resultado") == "pasos":
            return f"{key}_Pasos"
        return key

    def _current_view_key(self):
        if not hasattr(self, "notebook_vistas"):
            return "Original"
        try:
            tab_id = self.notebook_vistas.select()
            return self._tab_metodo_por_id.get(str(tab_id)) or self._tab_metodo_por_id.get(tab_id) or "Original"
        except Exception:
            return "Original"

    def _canvas_pointer_coords(self, key, event=None):
        canvas = self._canvas_widgets.get(key)
        if canvas is None:
            return None, None
        try:
            x = canvas.winfo_pointerx() - canvas.winfo_rootx()
            y = canvas.winfo_pointery() - canvas.winfo_rooty()
            if 0 <= x <= max(canvas.winfo_width(), 1) and 0 <= y <= max(canvas.winfo_height(), 1):
                return x, y
        except Exception:
            pass
        if event is not None:
            return getattr(event, "x", None), getattr(event, "y", None)
        return None, None

    def _update_zoom_ui(self, key=None):
        if not hasattr(self, "lbl_zoom_pct"):
            return
        if key is None:
            key = self._current_view_key()
        state_key = self._state_key_for_canvas(key)
        estado = self._get_view_state(state_key)
        zoom_level = estado.get("zoom_level")
        if zoom_level is None:
            zoom_level = 1.0
        pct = max(1, int(round(float(zoom_level) * 100)))
        self._zoom_ui_updating = True
        try:
            self.zoom_scale_var.set(float(pct))
        finally:
            self._zoom_ui_updating = False
        self.lbl_zoom_pct.config(text=f"{pct}%")

    def _on_zoom_scale_change(self, _value=None):
        if getattr(self, "_zoom_ui_updating", False):
            return
        key = self._current_view_key()
        state_key = self._state_key_for_canvas(key)
        if state_key not in self._render_meta and self._imagenes_render.get(key) is None and self._imagenes_render.get(state_key) is None:
            return
        self._apply_zoom_canvas(key, float(self.zoom_scale_var.get()) / 100.0)

    def _zoom_in_current_view(self):
        key = self._current_view_key()
        state_key = self._state_key_for_canvas(key)
        meta = self._render_meta.get(state_key)
        estado = self._get_view_state(state_key)
        zoom_actual = float(estado.get("zoom_level") or (meta or {}).get("zoom_level") or 1.0)
        self._apply_zoom_canvas(key, zoom_actual * 1.15)

    def _zoom_out_current_view(self):
        key = self._current_view_key()
        state_key = self._state_key_for_canvas(key)
        meta = self._render_meta.get(state_key)
        estado = self._get_view_state(state_key)
        zoom_actual = float(estado.get("zoom_level") or (meta or {}).get("zoom_level") or 1.0)
        self._apply_zoom_canvas(key, zoom_actual / 1.15)

    def _zoom_fit_current_view(self):
        self._zoom_fit_canvas(self._current_view_key(), render=True)

    def _zoom_100_current_view(self):
        key = self._current_view_key()
        self._apply_zoom_canvas(key, 1.0)

    def _zoom_fit_canvas(self, key, render=True):
        state_key = self._state_key_for_canvas(key)
        canvas = self._canvas_widgets.get(key)
        img_cv = self._imagenes_render.get(state_key) if state_key.endswith("_Pasos") else self._imagenes_render.get(key)
        if canvas is None or img_cv is None:
            return
        img_cv = self._to_bgr(img_cv)
        if img_cv is None:
            return
        h, w = img_cv.shape[:2]
        cw = max(canvas.winfo_width(), 100)
        ch = max(canvas.winfo_height(), 100)
        scale_x = cw / max(w, 1)
        scale_y = ch / max(h, 1)
        estado = self._get_view_state(state_key)
        zoom_fit = min(scale_x, scale_y)
        zoom_fit = max(estado["zoom_min"], min(estado["zoom_max"], zoom_fit))
        estado["zoom_level"] = zoom_fit
        estado["pan_offset_x"] = int((cw - w * zoom_fit) / 2)
        estado["pan_offset_y"] = int((ch - h * zoom_fit) / 2)
        estado["image_size"] = (w, h)
        self._drag_pan_state.pop(state_key, None)
        if render:
            self._render_single_view(key)
        else:
            self._update_zoom_ui(key)

    def _clamp_view_state(self, key, estado=None, img_cv=None):
        state_key = self._state_key_for_canvas(key)
        if estado is None:
            estado = self._get_view_state(state_key)
        canvas = self._canvas_widgets.get(key)
        if canvas is None:
            return estado
        if img_cv is None:
            img_cv = self._imagenes_render.get(state_key) if state_key.endswith("_Pasos") else self._imagenes_render.get(key)
        img_cv = self._to_bgr(img_cv)
        if img_cv is None:
            return estado
        h, w = img_cv.shape[:2]
        zoom_level = float(estado.get("zoom_level") or 1.0)
        cw = max(canvas.winfo_width(), 100)
        ch = max(canvas.winfo_height(), 100)
        w_scaled = w * zoom_level
        h_scaled = h * zoom_level

        visible_min_x = min(max(48.0, min(cw, w_scaled) * 0.25), w_scaled) if w_scaled > 0 else 0.0
        visible_min_y = min(max(48.0, min(ch, h_scaled) * 0.25), h_scaled) if h_scaled > 0 else 0.0
        min_x = int(visible_min_x - w_scaled)
        max_x = int(cw - visible_min_x)
        min_y = int(visible_min_y - h_scaled)
        max_y = int(ch - visible_min_y)

        estado["pan_offset_x"] = int(min(max(float(estado.get("pan_offset_x", 0.0)), min_x), max_x))
        estado["pan_offset_y"] = int(min(max(float(estado.get("pan_offset_y", 0.0)), min_y), max_y))

        estado["image_size"] = (w, h)
        return estado

    def _apply_zoom_canvas(self, key, new_zoom, center_x=None, center_y=None):
        state_key = self._state_key_for_canvas(key)
        meta = self._render_meta.get(state_key)
        if not meta:
            self._render_single_view(key)
            meta = self._render_meta.get(state_key)
            if not meta:
                return
        estado = self._get_view_state(state_key)
        old_zoom = float(estado.get("zoom_level") or meta.get("zoom_level") or 1.0)
        new_zoom = max(estado["zoom_min"], min(estado["zoom_max"], float(new_zoom)))
        if abs(new_zoom - old_zoom) < 1e-6:
            return
        if center_x is None or center_y is None:
            center_x = meta.get("canvas_width", 0) / 2
            center_y = meta.get("canvas_height", 0) / 2
        img_x = (center_x - float(estado.get("pan_offset_x", 0.0))) / old_zoom
        img_y = (center_y - float(estado.get("pan_offset_y", 0.0))) / old_zoom
        estado["zoom_level"] = new_zoom
        estado["pan_offset_x"] = int(center_x - img_x * new_zoom)
        estado["pan_offset_y"] = int(center_y - img_y * new_zoom)
        self._clamp_view_state(key, estado)
        self._render_single_view(key)

    def _cambiar_modo_vista_metodo(self, metodo, modo):
        if metodo not in self.METODOS or modo not in ("resultado", "pasos"):
            return
        self._modo_vista_por_metodo[metodo] = modo
        self._render_single_view(metodo)
        self._actualizar_barra_modo_vista()
        self._update_zoom_ui(metodo)

    def _cambiar_modo_vista_actual(self, modo):
        metodo = self._current_view_key()
        if metodo in self.METODOS:
            self._cambiar_modo_vista_metodo(metodo, modo)

    def _on_canvas_wheel(self, event, key):
        self._active_canvas_key = key
        state_key = self._state_key_for_canvas(key)
        meta = self._render_meta.get(state_key)
        if not meta:
            self._render_single_view(key)
            meta = self._render_meta.get(state_key)
            if not meta:
                return "break"
        estado = self._get_view_state(state_key)
        ctrl_presionado = bool(event.state & 0x0004)
        factor_base = 1.30 if ctrl_presionado else 1.15
        factor = factor_base if event.delta > 0 else 1 / factor_base
        zoom_actual = float(estado.get("zoom_level") or meta.get("zoom_level") or 1.0)
        center_x, center_y = self._canvas_pointer_coords(key, event)
        self._apply_zoom_canvas(key, zoom_actual * factor, center_x, center_y)
        return "break"

    def _on_canvas_pan_start(self, event, key):
        state_key = self._state_key_for_canvas(key)
        self._activar_canvas_zoom(event.widget, key)
        try:
            event.widget.configure(cursor="fleur")
        except Exception:
            pass
        self._drag_pan_state[state_key] = {
            "x": event.x,
            "y": event.y,
            "pan_x": self._get_view_state(state_key)["pan_offset_x"],
            "pan_y": self._get_view_state(state_key)["pan_offset_y"],
        }
        return "break"

    def _on_canvas_pan_move(self, event, key):
        state_key = self._state_key_for_canvas(key)
        drag = self._drag_pan_state.get(state_key)
        if not drag:
            return
        estado = self._get_view_state(state_key)
        estado["pan_offset_x"] = drag["pan_x"] + (event.x - drag["x"])
        estado["pan_offset_y"] = drag["pan_y"] + (event.y - drag["y"])
        self._clamp_view_state(key, estado)
        self._render_single_view(key)
        return "break"

    def _on_canvas_pan_end(self, event, key):
        state_key = self._state_key_for_canvas(key)
        self._drag_pan_state.pop(state_key, None)
        canvas = self._canvas_widgets.get(key)
        if canvas is not None:
            try:
                canvas.configure(cursor="crosshair")
            except Exception:
                pass
        return "break"

    def _reset_canvas_zoom(self, key, render=True):
        state_key = self._state_key_for_canvas(key)
        estado = self._get_view_state(state_key)
        estado["zoom_level"] = None
        estado["pan_offset_x"] = 0.0
        estado["pan_offset_y"] = 0.0
        estado["image_size"] = None
        self._drag_pan_state.pop(state_key, None)
        if render:
            self._render_single_view(key)
        self._update_zoom_ui(key)

    def _on_view_tab_changed(self, event=None):
        if not hasattr(self, "notebook_vistas"):
            return
        tab_id = self.notebook_vistas.select()
        vista = self._tab_metodo_por_id.get(str(tab_id)) or self._tab_metodo_por_id.get(tab_id)
        self._actualizar_barra_modo_vista(vista)
        if vista in self.METODOS:
            self._render_single_view(vista)
        self._update_zoom_ui(vista)
        if hasattr(self, "summary_notebook") and vista in getattr(self, "_summary_tab_by_key", {}):
            try:
                self.summary_notebook.select(self._summary_tab_by_key[vista])
            except Exception:
                pass

    def _seleccionar_metodo_en_ui(self, vista):
        if not hasattr(self, "notebook_vistas"):
            return
        try:
            for tab_id in self.notebook_vistas.tabs():
                if self.notebook_vistas.tab(tab_id, "text") == vista:
                    self.notebook_vistas.select(tab_id)
                    break
        except Exception:
            pass
        if hasattr(self, "summary_notebook") and vista in getattr(self, "_summary_tab_by_key", {}):
            try:
                self.summary_notebook.select(self._summary_tab_by_key[vista])
            except Exception:
                pass
        self._actualizar_barra_modo_vista(vista)
        self._update_zoom_ui(vista)

    def _actualizar_barra_modo_vista(self, vista=None):
        barra = getattr(self, "view_mode_bar", None)
        if barra is None or not hasattr(self, "notebook_vistas"):
            return
        if vista is None:
            vista = self._current_view_key()
        if vista not in self.METODOS:
            try:
                barra.place_forget()
            except Exception:
                pass
            return
        try:
            barra.place(in_=self.notebook_vistas, relx=1.0, x=-8, y=2, anchor="ne")
        except Exception:
            return
        modo = self._modo_vista_por_metodo.get(vista, "resultado")
        if hasattr(self, "btn_view_resultado"):
            self.btn_view_resultado.configure(style="ViewModeActive.TButton" if modo == "resultado" else "ViewMode.TButton")
        if hasattr(self, "btn_view_pasos"):
            self.btn_view_pasos.configure(style="ViewModeActive.TButton" if modo == "pasos" else "ViewMode.TButton")

    def _resolver_valor_config(self, key, default=None):
        idx = self.imagen_actual_idx
        if idx in self._config_por_imagen and key in self._config_por_imagen[idx]:
            return self._config_por_imagen[idx][key]
        if self._config_global_base and key in self._config_global_base:
            return self._config_global_base[key]
        if key in self.motor.config:
            return self.motor.config[key]
        return default

    def _snapshot_config_avanzada(self):
        return {key: copy.deepcopy(self.motor.config[key]) for key in self.CLAVES_AVANZADAS if key in self.motor.config}

    def _guardar_config_imagen(self, idx=None):
        if idx is None:
            idx = self.imagen_actual_idx
        if idx is None or idx < 0:
            return
        self._config_por_imagen[idx] = self._snapshot_config_avanzada()

    def _guardar_config_global(self):
        self._config_global_base = self._snapshot_config_avanzada()

    def _registrar_config_procesada(self, idx=None, config=None):
        if idx is None:
            idx = self.imagen_actual_idx
        if idx is None or idx < 0:
            return
        if config is None:
            config = self._snapshot_config_avanzada()
        self._config_por_imagen[idx] = copy.deepcopy(config)

    def _config_actual_para_ui(self):
        cfg = copy.deepcopy(self._config_global_base) if self._config_global_base else self._snapshot_config_avanzada()
        idx = self.imagen_actual_idx
        if idx in self._config_por_imagen:
            cfg.update(self._config_por_imagen[idx])
        return cfg

    def _aplicar_config_en_vars(self, config):
        if not config:
            return
        for key, valor in config.items():
            if key not in self.vars:
                continue
            var = self.vars[key]
            if isinstance(var, tk.BooleanVar):
                var.set(bool(valor))
            elif isinstance(var, tk.IntVar):
                var.set(int(valor))
            elif isinstance(var, tk.StringVar):
                if key == "profundidad_asumida_huecos":
                    if valor == "baja":
                        var.set("baja (13-25mm)")
                    elif valor == "alta":
                        var.set("alta (>50mm)")
                    elif "media" not in str(valor).lower():
                        var.set(str(valor))
                    else:
                        var.set("media (25-50mm)")
                else:
                    var.set(str(valor))
            else:
                var.set(float(valor))

    def _crear_ventana_modal(self, titulo):
        win = tk.Toplevel(self)
        win.title(titulo)
        win.configure(bg=EstiloUI.BG_DARK)
        win.transient(self)
        win.grab_set()
        win.bind("<Escape>", lambda e: win.destroy())
        return win

    def _sl_modal(self, parent, key, label, from_, to, default, step=1, resolution=None):
        frame = tk.Frame(parent, bg=EstiloUI.BG_PANEL)
        frame.pack(fill="x", padx=10, pady=2)
        top = tk.Frame(frame, bg=EstiloUI.BG_PANEL)
        top.pack(fill="x")
        tk.Label(top, text=label, font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY).pack(side="left")
        valor_inicial = self._resolver_valor_config(key, default)
        var = self._crear_var(key, valor_inicial)
        val_label = tk.Label(top, text="", font=EstiloUI.FONT_LABEL, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT, width=7, anchor="e")
        val_label.pack(side="right")
        res = resolution if resolution is not None else step
        slider = tk.Scale(frame, from_=from_, to=to, orient="horizontal", variable=var, resolution=res, showvalue=False, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT, troughcolor=EstiloUI.BG_DARK, highlightthickness=0, activebackground=EstiloUI.BG_BUTTON, length=280, sliderlength=15, bd=0, relief="flat")
        slider.pack(fill="x", pady=(0, 2))
        trace_holder = {"id": None}

        def _upd(*_):
            if not val_label.winfo_exists():
                trace_id = trace_holder.get("id")
                if trace_id:
                    try:
                        var.trace_remove("write", trace_id)
                    except Exception:
                        pass
                    trace_holder["id"] = None
                return
            valor = var.get()
            try:
                if isinstance(var, tk.IntVar):
                    val_label.config(text=str(int(valor)))
                else:
                    val_label.config(text=f"{float(valor):.2f}")
            except tk.TclError:
                trace_id = trace_holder.get("id")
                if trace_id:
                    try:
                        var.trace_remove("write", trace_id)
                    except Exception:
                        pass
                    trace_holder["id"] = None

        def _cleanup_trace(event=None):
            if event is not None and event.widget is not frame:
                return
            trace_id = trace_holder.get("id")
            if trace_id:
                try:
                    var.trace_remove("write", trace_id)
                except Exception:
                    pass
                trace_holder["id"] = None

        trace_holder["id"] = var.trace_add("write", _upd)
        frame.bind("<Destroy>", _cleanup_trace, add="+")
        _upd()

    def _aplicar_config_individual_y_cerrar(self, win):
        self._sync_config()
        self._guardar_config_imagen()
        self._log("Configuracion avanzada guardada para la imagen actual. Procese de nuevo para recalcular.")
        win.destroy()

    def _aplicar_config_todas_y_cerrar(self, win):
        self._sync_config()
        self._guardar_config_global()
        self._config_por_imagen.clear()
        self._log("Configuracion avanzada guardada para todas las imagenes. Procese de nuevo para recalcular.")
        win.destroy()

    def _aplicar_perfil(self, valores):
        for key, valor in valores.items():
            if key in self.vars:
                self.vars[key].set(valor)

    def _perfil_piel_estandar(self):
        self._aplicar_perfil({"clahe_clip": 3.0, "clahe_tile": 8, "bilateral_d": 9, "bilateral_sigma_color": 75, "bilateral_sigma_space": 75, "block_size": 25, "C_umbral": 12, "kernel_apertura": 3, "kernel_cierre": 5, "iteraciones_cierre": 2, "usar_frangi": False, "usar_multiescala": False, "min_area_poligono": 300, "min_circularidad": 0.08, "min_vertices": 4, "max_vertices": 25, "min_radio_circulo": 8, "min_longitud_rama": 30, "min_area_objeto": 100, "max_gap_cierre": 20})
        self._log("Perfil ESTANDAR de piel de cocodrilo aplicado.")

    def _perfil_piel_agresivo(self):
        self._aplicar_perfil({"clahe_clip": 4.5, "clahe_tile": 6, "bilateral_d": 7, "bilateral_sigma_color": 60, "bilateral_sigma_space": 60, "block_size": 19, "C_umbral": 8, "kernel_apertura": 2, "kernel_cierre": 7, "iteraciones_cierre": 3, "usar_frangi": True, "usar_multiescala": True, "min_area_poligono": 150, "min_circularidad": 0.05, "min_vertices": 3, "max_vertices": 35, "min_radio_circulo": 5, "min_longitud_rama": 15, "min_area_objeto": 50, "max_gap_cierre": 30})
        self._log("Perfil AGRESIVO de piel de cocodrilo aplicado.")

    def _perfil_piel_suave(self):
        self._aplicar_perfil({"clahe_clip": 2.5, "clahe_tile": 10, "bilateral_d": 11, "bilateral_sigma_color": 90, "bilateral_sigma_space": 90, "block_size": 31, "C_umbral": 15, "kernel_apertura": 4, "kernel_cierre": 4, "iteraciones_cierre": 1, "usar_frangi": False, "usar_multiescala": False, "min_area_poligono": 500, "min_circularidad": 0.12, "min_vertices": 5, "max_vertices": 20, "min_radio_circulo": 12, "min_longitud_rama": 50, "min_area_objeto": 200, "max_gap_cierre": 15})
        self._log("Perfil SUAVE de piel de cocodrilo aplicado.")

    def _abrir_filtros(self):
        self._abrir_config_avanzada("filtros")

    def _abrir_fusion(self):
        self._abrir_config_avanzada("fusion")

    def _abrir_piel_config(self):
        self._abrir_config_avanzada("piel")

    def _abrir_poligonos(self):
        self._abrir_config_avanzada("poligonos")

    def _abrir_perfiles(self):
        self._abrir_config_avanzada("perfiles")

    def _abrir_config_avanzada(self, seccion=None):
        titulos = {None: "Configuracion Avanzada", "filtros": "Filtros Tamano Minimo", "fusion": "Fusion Solapamientos", "piel": "Piel de Cocodrilo", "poligonos": "Deteccion Poligonos", "perfiles": "Perfiles Rapidos"}
        win = self._crear_ventana_modal(titulos.get(seccion, "Configuracion Avanzada"))
        canvas = tk.Canvas(win, bg=EstiloUI.BG_PANEL, highlightthickness=0)
        vsb = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
        hsb = ttk.Scrollbar(win, orient="horizontal", command=canvas.xview)
        sf = tk.Frame(canvas, bg=EstiloUI.BG_PANEL)
        sf.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=sf, anchor="nw")
        canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        win.grid_rowconfigure(0, weight=1)
        win.grid_columnconfigure(0, weight=1)
        canvas.bind("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))

        if seccion in (None, "filtros"):
            self._seccion_titulo(sf, "FILTROS TAMANO MINIMO")
            tk.Checkbutton(sf, text="Filtrar huecos / baches", variable=self.vars["filtrar_baches"], bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK, font=EstiloUI.FONT_SMALL).pack(anchor="w", padx=10, pady=1)
            self._sl_modal(sf, "min_diametro_hueco_mm", "Min diametro hueco (mm)", 0, 200, 50, step=10)
            prof_frame = tk.Frame(sf, bg=EstiloUI.BG_PANEL)
            prof_frame.pack(fill="x", padx=10, pady=2)
            tk.Label(prof_frame, text="Prof. asumida huecos:", font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_SECONDARY).pack(side="left")
            prof_val = self._resolver_valor_config("profundidad_asumida_huecos", "media")
            inicial = "media (25-50mm)"
            if prof_val == "baja":
                inicial = "baja (13-25mm)"
            elif prof_val == "alta":
                inicial = "alta (>50mm)"
            self.vars["profundidad_asumida_huecos"].set(inicial)
            ttk.Combobox(prof_frame, textvariable=self.vars["profundidad_asumida_huecos"], values=["baja (13-25mm)", "media (25-50mm)", "alta (>50mm)"], state="readonly", width=16).pack(side="right")
            tk.Checkbutton(sf, text="Filtrar grietas / fisuras", variable=self.vars["filtrar_grietas"], bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK, font=EstiloUI.FONT_SMALL).pack(anchor="w", padx=10, pady=1)
            self._sl_modal(sf, "min_longitud_grieta_m", "Min longitud grieta (m)", 0.0, 1.0, 0.05, resolution=0.01)
            tk.Checkbutton(sf, text="Usar clasificacion de borde interno y berma", variable=self.vars["usar_borde_berma_pci"], bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK, font=EstiloUI.FONT_SMALL).pack(anchor="w", padx=10, pady=1)
            tk.Label(sf, text="Aplica a: PCI / MTC y VIZIR", justify="left", wraplength=320, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT, font=EstiloUI.FONT_SMALL).pack(fill="x", padx=10, pady=(4, 1))
            self._sl_modal(sf, "borde_interno_m", "Borde interno (m)", 0.00, 3.00, 0.30, resolution=0.01)
            self._sl_modal(sf, "borde_externo_m", "Berma (m)", 0.00, 3.00, 0.30, resolution=0.01)
            tk.Checkbutton(sf, text="Filtrar parches", variable=self.vars["filtrar_parches"], bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK, font=EstiloUI.FONT_SMALL).pack(anchor="w", padx=10, pady=1)
            self._sl_modal(sf, "min_area_parche_m2", "Min area parche (m2)", 0.0, 1.0, 0.01, resolution=0.005)
            self._sl_modal(sf, "parche_ratio_leve_max", "Rel. fisurada max Sev 1", 0.00, 1.00, 0.08, resolution=0.01)
            self._sl_modal(sf, "parche_ratio_moderado_max", "Rel. fisurada max Sev 2", 0.00, 1.00, 0.18, resolution=0.01)
            tk.Checkbutton(sf, text="Filtrar piel de cocodrilo", variable=self.vars["filtrar_piel"], bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK, font=EstiloUI.FONT_SMALL).pack(anchor="w", padx=10, pady=1)
            self._sl_modal(sf, "min_area_piel_m2", "Min area piel (m2)", 0.0, 2.0, 0.05, resolution=0.01)
            self._sl_modal(sf, "merge_fisuras_px", "Dist union fisuras (px)", 0, 120, 30, step=5)

        if seccion in (None, "fusion"):
            self._seccion_titulo(sf, "FUSION SOLAPAMIENTOS")
            self._sl_modal(sf, "merge_iou_threshold", "IoU fusion", 0.0, 0.5, 0.10, resolution=0.02)
            self._sl_modal(sf, "merge_distancia_max_px", "Dist max fusion (px)", 0, 200, 50, step=10)

        if seccion in (None, "piel"):
            self._seccion_titulo(sf, "PIEL DE COCODRILO")
            ttk.Button(sf, text="Auto-calibrar zona (seleccionar muestra)", style="Accent.TButton", command=self._auto_calibrar_piel).pack(fill="x", padx=10, pady=4)
            self._sl_modal(sf, "clahe_clip", "CLAHE clip", 1.0, 8.0, 4.0, resolution=0.5)
            self._sl_modal(sf, "clahe_tile", "CLAHE tile", 4, 16, 8)
            self._sl_modal(sf, "bilateral_d", "Bilateral d", 3, 15, 9)
            self._sl_modal(sf, "bilateral_sigma_color", "Sigma color", 20, 150, 75)
            self._sl_modal(sf, "bilateral_sigma_space", "Sigma espacio", 20, 150, 75)
            self._sl_modal(sf, "block_size", "Block size", 11, 51, 23, step=2)
            self._sl_modal(sf, "C_umbral", "Constante C", 3, 25, 10)
            tk.Checkbutton(sf, text="Filtro Frangi", variable=self.vars["usar_frangi"], bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK, font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)
            tk.Checkbutton(sf, text="Refinar esqueleto y cerrar gaps", variable=self.vars["usar_refinamiento"], bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK, font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)
            tk.Checkbutton(sf, text="Multi-escala", variable=self.vars["usar_multiescala"], bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK, font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)
            self._sl_modal(sf, "kernel_apertura", "Kernel apertura", 2, 7, 3)
            self._sl_modal(sf, "kernel_cierre", "Kernel cierre", 3, 12, 6)
            self._sl_modal(sf, "iteraciones_cierre", "Iteraciones cierre", 1, 5, 2)

        if seccion in (None, "poligonos"):
            self._seccion_titulo(sf, "DETECCION POLIGONOS")
            self._sl_modal(sf, "min_area_poligono", "Area min poligono", 50, 1000, 300, step=50)
            self._sl_modal(sf, "min_circularidad", "Circularidad min", 0.01, 0.5, 0.08, resolution=0.01)
            self._sl_modal(sf, "min_vertices", "Vertices min", 3, 8, 4)
            self._sl_modal(sf, "max_vertices", "Vertices max", 10, 50, 25)
            self._sl_modal(sf, "min_radio_circulo", "Radio min circulo", 3, 30, 8)
            self._sl_modal(sf, "min_longitud_rama", "Long min rama", 10, 100, 30, step=5)
            self._sl_modal(sf, "min_area_objeto", "Area min objeto", 20, 500, 100, step=10)
            self._sl_modal(sf, "max_gap_cierre", "Max gap cierre", 5, 50, 20)

        if seccion in (None, "perfiles"):
            self._seccion_titulo(sf, "PERFILES RAPIDOS")
            for nombre, cmd in [("Estandar", self._perfil_piel_estandar), ("Agresivo", self._perfil_piel_agresivo), ("Suave", self._perfil_piel_suave)]:
                ttk.Button(sf, text=nombre, style="Secondary.TButton", command=cmd).pack(fill="x", padx=10, pady=2)

        tk.Frame(sf, bg=EstiloUI.BG_ACCENT, height=2).pack(fill="x", padx=10, pady=(15, 5))
        btn_f = tk.Frame(sf, bg=EstiloUI.BG_PANEL)
        btn_f.pack(fill="x", padx=10, pady=5)
        ttk.Button(btn_f, text="Guardar para imagen actual y cerrar", style="Accent.TButton", command=lambda: self._aplicar_config_individual_y_cerrar(win)).pack(fill="x", pady=2)
        ttk.Button(btn_f, text="Guardar para todas las imagenes y cerrar", style="Accent.TButton", command=lambda: self._aplicar_config_todas_y_cerrar(win)).pack(fill="x", pady=2)

        win.update_idletasks()
        req_w = sf.winfo_reqwidth() + vsb.winfo_reqwidth() + 28
        req_h = sf.winfo_reqheight() + hsb.winfo_reqheight() + 28
        max_h = int(self.winfo_screenheight() * 0.86)
        max_w = int(self.winfo_screenwidth() * 0.82)
        w = min(req_w, max_w)
        h = min(req_h, max_h)
        x = self.winfo_x() + max((self.winfo_width() - w) // 2, 10)
        y = self.winfo_y() + max((self.winfo_height() - h) // 2, 10)
        win.geometry(f"{w}x{h}+{x}+{y}")

    def _auto_calibrar_piel(self):
        ruta = self._ruta_actual()
        if not ruta:
            messagebox.showinfo("Info", "Cargue una imagen primero.")
            return
        imagen = cv2.imread(str(ruta))
        if imagen is None:
            messagebox.showerror("Error", "No se pudo cargar la imagen actual.")
            return
        win = VentanaAutoCalibrarPiel(self, imagen)
        if hasattr(win, "resultado") and win.resultado:
            self._aplicar_perfil(win.resultado)
            self._log(f"Auto-calibracion de piel aplicada sobre {Path(ruta).name}")
        else:
            self._log("Auto-calibracion de piel cancelada.")

    def _normalizar_config_piel(self, cfg):
        cfg = copy.deepcopy(cfg)
        bs = int(cfg.get("block_size", 23))
        bs = max(11, min(51, bs))
        if bs % 2 == 0:
            bs += 1 if bs < 51 else -1
        cfg["block_size"] = bs
        cfg["clahe_tile"] = max(4, min(16, int(cfg.get("clahe_tile", 8))))
        cfg["bilateral_d"] = max(3, min(15, int(cfg.get("bilateral_d", 9))))
        cfg["kernel_apertura"] = max(2, min(7, int(cfg.get("kernel_apertura", 3))))
        cfg["kernel_cierre"] = max(3, min(12, int(cfg.get("kernel_cierre", 6))))
        cfg["iteraciones_cierre"] = max(1, min(5, int(cfg.get("iteraciones_cierre", 2))))
        cfg["min_area_poligono"] = max(50, int(cfg.get("min_area_poligono", 300)))
        cfg["min_area_objeto"] = max(20, int(cfg.get("min_area_objeto", 100)))
        cfg["min_longitud_rama"] = max(10, int(cfg.get("min_longitud_rama", 30)))
        cfg["min_vertices"] = max(3, int(cfg.get("min_vertices", 4)))
        cfg["max_vertices"] = max(cfg["min_vertices"], int(cfg.get("max_vertices", 25)))
        return cfg

    def _normalizar_config_parche(self, cfg):
        cfg = copy.deepcopy(cfg or {})
        ratio_leve, ratio_moderado = normalizar_umbrales_parche(
            cfg.get("parche_ratio_leve_max", self.motor.config.get("parche_ratio_leve_max", 0.08)),
            cfg.get("parche_ratio_moderado_max", self.motor.config.get("parche_ratio_moderado_max", 0.18)),
        )
        cfg["parche_ratio_leve_max"] = ratio_leve
        cfg["parche_ratio_moderado_max"] = ratio_moderado
        return cfg

    def _resolver_mascara_parche_falla(self, falla):
        mask = falla.get("mascara_parche")
        if isinstance(mask, np.ndarray) and mask.size > 0:
            return (mask > 0).astype(np.uint8) * 255
        if not self.resultado_actual or self.resultado_actual.get("imagen") is None:
            return None
        if "contorno" not in falla:
            return None
        h, w = self.resultado_actual["imagen"].shape[:2]
        mask = np.zeros((h, w), dtype=np.uint8)
        cv2.fillPoly(mask, [falla["contorno"]], 255)
        return mask

    def _configurar_falla_individual(self, metodo, idx):
        if not self.resultado_actual:
            return
        fallas = self.resultado_actual["metodos"].get(metodo, {}).get("fallas", [])
        if idx < 0 or idx >= len(fallas):
            return
        categoria = categorizar_tipo_falla(fallas[idx].get("tipo"))
        if categoria == "piel":
            self._configurar_piel_falla(metodo, idx)
            return
        if categoria == "parche":
            self._configurar_parche_falla(metodo, idx)
            return
        messagebox.showinfo("Info", "La configuracion individual solo aplica a piel de cocodrilo y parcheos.")

    def _configurar_parche_falla(self, metodo, idx):
        if not self.resultado_actual:
            return
        fallas = self.resultado_actual["metodos"].get(metodo, {}).get("fallas", [])
        if idx < 0 or idx >= len(fallas):
            return

        falla = fallas[idx]
        if categorizar_tipo_falla(falla.get("tipo")) != "parche":
            messagebox.showinfo("Info", "La configuracion individual solo aplica a parcheos.")
            return

        mask = self._resolver_mascara_parche_falla(falla)
        if mask is None:
            messagebox.showinfo("Info", "Esta deteccion no tiene mascara disponible para recalcular severidad.")
            return

        base_cfg = self._normalizar_config_parche(
            falla.get("config_parche_personalizada", self.motor.config.copy())
        )
        ratio_actual = float(falla.get("ratio_fisuras", 0.0) or 0.0)
        if ratio_actual <= 0.0 and self.resultado_actual.get("imagen") is not None:
            _, ratio_actual = medir_relacion_fisuras_parche(mask, self.resultado_actual["imagen"])

        win = tk.Toplevel(self)
        win.title(f"{metodo} | Configurar Parche #{falla.get('id', idx + 1)}")
        win.configure(bg=EstiloUI.BG_DARK)
        win.transient(self)
        win.grab_set()
        win.bind("<Escape>", lambda e: win.destroy())

        sf = tk.Frame(win, bg=EstiloUI.BG_PANEL)
        sf.pack(fill="both", expand=True, padx=8, pady=8)
        self._seccion_titulo(sf, f"CONFIGURACION INDIVIDUAL - {metodo} - PARCHE #{falla.get('id', idx + 1)}")

        ratio_leve_var = tk.DoubleVar(value=base_cfg.get("parche_ratio_leve_max", 0.08))
        ratio_moderado_var = tk.DoubleVar(value=base_cfg.get("parche_ratio_moderado_max", 0.18))

        def _sl_local(parent, var, label, default):
            frame = tk.Frame(parent, bg=EstiloUI.BG_PANEL)
            frame.pack(fill="x", padx=10, pady=2)
            hdr = tk.Frame(frame, bg=EstiloUI.BG_PANEL)
            hdr.pack(fill="x")
            tk.Label(hdr, text=label, font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY).pack(side="left")
            val_lbl = tk.Label(hdr, font=("Segoe UI", 9, "bold"), bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT)
            val_lbl.pack(side="right")
            slider = tk.Scale(
                frame,
                from_=0.00,
                to=1.00,
                orient="horizontal",
                variable=var,
                resolution=0.01,
                showvalue=False,
                bg=EstiloUI.BG_PANEL,
                fg=EstiloUI.FG_HIGHLIGHT,
                troughcolor=EstiloUI.BG_DARK,
                highlightthickness=0,
                activebackground=EstiloUI.BG_BUTTON,
                length=320,
                sliderlength=15,
                bd=0,
                relief="flat",
            )
            slider.pack(fill="x", pady=(0, 2))

            def _upd(*_args):
                val_lbl.config(text=f"{float(var.get()):.2f}")

            var.trace_add("write", _upd)
            if default is not None:
                var.set(default)
            _upd()

        tk.Label(
            sf,
            text=f"Relacion actual de area fisurada: {ratio_actual:.3f}",
            font=EstiloUI.FONT_SMALL,
            bg=EstiloUI.BG_PANEL,
            fg=EstiloUI.FG_HIGHLIGHT,
        ).pack(anchor="w", padx=10, pady=(2, 6))
        _sl_local(sf, ratio_leve_var, "Rel. fisurada max Sev 1", base_cfg.get("parche_ratio_leve_max", 0.08))
        _sl_local(sf, ratio_moderado_var, "Rel. fisurada max Sev 2", base_cfg.get("parche_ratio_moderado_max", 0.18))

        btns = tk.Frame(sf, bg=EstiloUI.BG_PANEL)
        btns.pack(fill="x", padx=10, pady=(10, 2))
        ttk.Button(
            btns,
            text="Aplicar a esta deteccion",
            style="Accent.TButton",
            command=lambda: self._aplicar_parche_individual(
                win,
                metodo,
                idx,
                {
                    "parche_ratio_leve_max": ratio_leve_var,
                    "parche_ratio_moderado_max": ratio_moderado_var,
                },
            ),
        ).pack(fill="x", pady=2)
        ttk.Button(btns, text="Cancelar", style="Secondary.TButton", command=win.destroy).pack(fill="x", pady=2)

        win.update_idletasks()
        width = max(460, sf.winfo_reqwidth() + 20)
        height = min(int(self.winfo_screenheight() * 0.75), sf.winfo_reqheight() + 30)
        x = self.winfo_rootx() + max(20, (self.winfo_width() - width) // 2)
        y = self.winfo_rooty() + max(20, (self.winfo_height() - height) // 2)
        win.geometry(f"{width}x{height}+{x}+{y}")

    def _aplicar_parche_individual(self, win, metodo, idx, local_vars):
        if not self.resultado_actual:
            return
        fallas = self.resultado_actual["metodos"].get(metodo, {}).get("fallas", [])
        if idx < 0 or idx >= len(fallas):
            return

        falla = fallas[idx]
        mask = self._resolver_mascara_parche_falla(falla)
        imagen = self.resultado_actual.get("imagen")
        if mask is None or imagen is None:
            self._log("Mascara de parche no disponible. Procese de nuevo la imagen.")
            win.destroy()
            return

        custom_cfg = self._normalizar_config_parche({
            "parche_ratio_leve_max": float(local_vars["parche_ratio_leve_max"].get()),
            "parche_ratio_moderado_max": float(local_vars["parche_ratio_moderado_max"].get()),
        })
        edges_masked, ratio = medir_relacion_fisuras_parche(mask, imagen)
        severidad = clasificar_severidad_parche_por_ratio(
            ratio,
            metodo=metodo,
            ratio_leve_max=custom_cfg["parche_ratio_leve_max"],
            ratio_moderado_max=custom_cfg["parche_ratio_moderado_max"],
        )

        falla["mascara_parche"] = mask
        falla["_fisuras_mask"] = edges_masked
        falla["ratio_fisuras"] = ratio
        falla["severidad"] = severidad
        falla["config_parche_personalizada"] = custom_cfg
        if metodo == "MTC":
            falla["esqueleto"] = edges_masked

        self._cachear_resultado_actual()
        self._refresh_method_views(metodo)
        self._actualizar_resumenes(self.resultado_actual)
        self._estado(f"{metodo}: parche #{falla.get('id', idx + 1)} reclasificado")
        self._log(
            f"{metodo}: {falla.get('tipo', 'PARCHE')} #{falla.get('id', idx + 1)} "
            f"| Rel={ratio:.3f} | Sev={severidad}"
        )
        win.destroy()

    def _configurar_piel_falla(self, metodo, idx):
        if not self.resultado_actual:
            return
        fallas = self.resultado_actual["metodos"].get(metodo, {}).get("fallas", [])
        if idx < 0 or idx >= len(fallas):
            return

        falla = fallas[idx]
        if categorizar_tipo_falla(falla.get("tipo")) != "piel":
            messagebox.showinfo("Info", "La configuracion individual solo aplica a piel de cocodrilo.")
            return
        if "mascara_roi" not in falla:
            messagebox.showinfo("Info", "Esta deteccion no tiene mascara disponible para reprocesar.")
            return

        win = tk.Toplevel(self)
        win.title(f"{metodo} | Configurar Piel de Cocodrilo #{falla.get('id', idx + 1)}")
        win.configure(bg=EstiloUI.BG_DARK)
        win.transient(self)
        win.grab_set()
        win.bind("<Escape>", lambda e: win.destroy())

        sf = tk.Frame(win, bg=EstiloUI.BG_PANEL)
        sf.pack(fill="both", expand=True, padx=8, pady=8)
        self._seccion_titulo(sf, f"CONFIGURACION INDIVIDUAL - {metodo} - FALLA #{falla.get('id', idx + 1)}")

        local_vars = {}
        base_cfg = self._normalizar_config_piel(falla.get("config_personalizada", self.motor.config.copy()))

        for k, v in [
            ("clahe_clip", 4.0),
            ("clahe_tile", 8),
            ("bilateral_d", 9),
            ("bilateral_sigma_color", 75),
            ("bilateral_sigma_space", 75),
            ("block_size", 23),
            ("C_umbral", 10),
            ("kernel_apertura", 3),
            ("kernel_cierre", 6),
            ("iteraciones_cierre", 2),
        ]:
            val = base_cfg.get(k, v)
            local_vars[k] = tk.DoubleVar(value=val) if isinstance(v, float) else tk.IntVar(value=val)

        local_vars["usar_frangi"] = tk.BooleanVar(value=base_cfg.get("usar_frangi", True))
        local_vars["usar_multiescala"] = tk.BooleanVar(value=base_cfg.get("usar_multiescala", True))
        local_vars["usar_refinamiento"] = tk.BooleanVar(value=base_cfg.get("usar_refinamiento", True))

        def _sl_local(parent, key, label, from_, to, step=1, resolution=None):
            frame = tk.Frame(parent, bg=EstiloUI.BG_PANEL)
            frame.pack(fill="x", padx=10, pady=2)
            hdr = tk.Frame(frame, bg=EstiloUI.BG_PANEL)
            hdr.pack(fill="x")
            tk.Label(hdr, text=label, font=EstiloUI.FONT_SMALL, bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY).pack(side="left")
            val_lbl = tk.Label(hdr, font=("Segoe UI", 9, "bold"), bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_HIGHLIGHT)
            val_lbl.pack(side="right")
            var = local_vars[key]
            slider = tk.Scale(
                frame,
                from_=from_,
                to=to,
                orient="horizontal",
                variable=var,
                resolution=(resolution if resolution is not None else step),
                showvalue=False,
                bg=EstiloUI.BG_PANEL,
                fg=EstiloUI.FG_HIGHLIGHT,
                troughcolor=EstiloUI.BG_DARK,
                highlightthickness=0,
                activebackground=EstiloUI.BG_BUTTON,
                length=320,
                sliderlength=15,
                bd=0,
                relief="flat",
            )
            slider.pack(fill="x", pady=(0, 2))

            def _upd(*_args):
                value = var.get()
                if resolution is not None and resolution < 1:
                    val_lbl.config(text=f"{float(value):.2f}")
                else:
                    val_lbl.config(text=str(int(value)))

            var.trace_add("write", _upd)
            _upd()

        _sl_local(sf, "clahe_clip", "CLAHE Clip", 1.0, 8.0, resolution=0.5)
        _sl_local(sf, "block_size", "Block Size", 11, 51, step=2)
        _sl_local(sf, "C_umbral", "Constante C", 3, 25)
        _sl_local(sf, "kernel_cierre", "Kernel Cierre", 3, 12)
        _sl_local(sf, "iteraciones_cierre", "Iter. Cierre", 1, 5)

        tk.Checkbutton(sf, text="Filtro Frangi", variable=local_vars["usar_frangi"], bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK, font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)
        tk.Checkbutton(sf, text="Refinar esqueleto y cerrar gaps", variable=local_vars["usar_refinamiento"], bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK, font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)
        tk.Checkbutton(sf, text="Multi-escala", variable=local_vars["usar_multiescala"], bg=EstiloUI.BG_PANEL, fg=EstiloUI.FG_PRIMARY, selectcolor=EstiloUI.BG_DARK, font=EstiloUI.FONT_BODY).pack(anchor="w", padx=10, pady=1)

        btns = tk.Frame(sf, bg=EstiloUI.BG_PANEL)
        btns.pack(fill="x", padx=10, pady=(10, 2))
        ttk.Button(btns, text="Aplicar a esta deteccion", style="Accent.TButton", command=lambda: self._aplicar_piel_individual(win, metodo, idx, local_vars)).pack(fill="x", pady=2)
        ttk.Button(btns, text="Cancelar", style="Secondary.TButton", command=win.destroy).pack(fill="x", pady=2)

        win.update_idletasks()
        width = max(460, sf.winfo_reqwidth() + 20)
        height = min(int(self.winfo_screenheight() * 0.85), sf.winfo_reqheight() + 30)
        x = self.winfo_rootx() + max(20, (self.winfo_width() - width) // 2)
        y = self.winfo_rooty() + max(20, (self.winfo_height() - height) // 2)
        win.geometry(f"{width}x{height}+{x}+{y}")

    def _aplicar_piel_individual(self, win, metodo, idx, local_vars):
        if not self.resultado_actual:
            return
        fallas = self.resultado_actual["metodos"].get(metodo, {}).get("fallas", [])
        if idx < 0 or idx >= len(fallas):
            return

        falla_vieja = fallas[idx]
        if "mascara_roi" not in falla_vieja:
            self._log("Mascara no disponible. Procese de nuevo la imagen.")
            win.destroy()
            return

        custom_cfg = self.motor.config.copy()
        for k, var in local_vars.items():
            if isinstance(var, tk.BooleanVar):
                custom_cfg[k] = var.get()
            elif isinstance(var, tk.DoubleVar):
                custom_cfg[k] = float(var.get())
            else:
                custom_cfg[k] = int(var.get())
        custom_cfg = self._normalizar_config_piel(custom_cfg)

        self._estado(f"Reprocesando piel de cocodrilo {metodo} #{falla_vieja.get('id', idx + 1)}...")

        def _reproc():
            try:
                if metodo == "PCI":
                    proc = ProcesadorPielCocodriloPCI(custom_cfg)
                    nuevo_res = self.motor._con_procesador_piel_pci(
                        lambda: proc.procesar(
                            falla_vieja["mascara_roi"],
                            self.motor.calibrador,
                            falla_vieja["confianza"],
                            self.resultado_actual["imagen"],
                        )
                    )
                elif metodo == "MTC":
                    proc = ProcesadorPielCocodriloMTC(custom_cfg)
                    nuevo_res = proc.procesar(
                        falla_vieja["mascara_roi"],
                        self.motor.calibrador,
                        falla_vieja["confianza"],
                        self.resultado_actual["imagen"],
                    )
                else:
                    proc = ProcesadorPielCocodriloVIZIR(custom_cfg)
                    nuevo_res = proc.procesar(
                        falla_vieja["mascara_roi"],
                        self.motor.calibrador,
                        falla_vieja["confianza"],
                        self.resultado_actual["imagen"],
                    )

                if nuevo_res:
                    nueva_falla = max(
                        nuevo_res,
                        key=lambda f: (f.get("area_m2", 0.0), f.get("area_px", 0.0), f.get("diametro_mm", 0.0)),
                    )
                    nueva_falla["config_personalizada"] = custom_cfg
                    nueva_falla["id"] = falla_vieja.get("id", idx + 1)
                    nueva_falla["excluida"] = falla_vieja.get("excluida", False)
                    self.after(0, lambda: self._finalizar_reprocesado_piel(win, metodo, idx, nueva_falla, "Falla reprocesada con exito."))
                else:
                    self.after(0, lambda: self._finalizar_reprocesado_piel(win, metodo, idx, None, "Reprocesamiento sin resultados."))
            except Exception as e:
                detalle = traceback.format_exc()
                self.after(0, lambda err=str(e), det=detalle: self._finalizar_reprocesado_piel(win, metodo, idx, None, f"Error: {err}", det))

        Thread(target=_reproc, daemon=True).start()

    def _finalizar_reprocesado_piel(self, win, metodo, idx, nueva_falla, msg, detalle=None):
        if nueva_falla is not None and self.resultado_actual:
            self.resultado_actual["metodos"][metodo]["fallas"][idx] = nueva_falla
            self._cachear_resultado_actual()
            self._refresh_method_views(metodo)
            self._actualizar_resumenes(self.resultado_actual)
        self._estado(msg)
        self._log(f"{metodo}: {msg}")
        if detalle:
            self._log(detalle.rstrip())
        try:
            win.destroy()
        except Exception:
            pass

    def _modo_calibracion_actual(self):
        var = self.vars.get("modo_calibracion")
        if isinstance(var, tk.StringVar):
            return var.get()
        return "automatica"

    def _actualizar_estado_calibracion_ui(self, resultado=None):
        modo = self._modo_calibracion_actual()

        if self.lbl_cal_estado is not None:
            if self._calibracion_unica_guardada:
                self.lbl_cal_estado.config(
                    text=(
                        f"Calibracion unica lista: "
                        f"{self._calibracion_unica_guardada['px_por_mm']:.4f} px/mm | "
                        f"Eje {self._calibracion_unica_guardada['angulo_eje_via']:.1f} grados"
                    ),
                    fg=EstiloUI.BG_SUCCESS,
                )
            elif modo == "unica":
                self.lbl_cal_estado.config(
                    text="Modo unica seleccionado. Falta calibrar una imagen de referencia.",
                    fg=EstiloUI.BG_WARNING,
                )
            elif modo == "cada_imagen":
                self.lbl_cal_estado.config(
                    text="Se abrira la calibracion manual al procesar cada imagen.",
                    fg=EstiloUI.FG_HIGHLIGHT,
                )
            else:
                self.lbl_cal_estado.config(text="Sin calibracion unica", fg=EstiloUI.FG_SECONDARY)

        if self.lbl_calibracion is None:
            return
        if resultado and resultado.get("calibracion_px_mm"):
            self.lbl_calibracion.config(
                text=(
                    f"Calibracion activa: {resultado['calibracion_px_mm']:.4f} px/mm | "
                    f"Eje {resultado.get('angulo_eje', 90.0):.1f} grados"
                )
            )
            return
        if modo == "unica" and self._calibracion_unica_guardada:
            self.lbl_calibracion.config(
                text=(
                    f"Calibracion base unica: {self._calibracion_unica_guardada['px_por_mm']:.4f} px/mm | "
                    f"Eje {self._calibracion_unica_guardada['angulo_eje_via']:.1f} grados"
                )
            )
        elif modo == "cada_imagen":
            self.lbl_calibracion.config(text="Calibracion: manual para cada imagen antes de procesar")
        else:
            self.lbl_calibracion.config(text="Calibracion: automatica por ancho de imagen")

    def _cachear_resultado_actual(self):
        if self.resultado_actual is None or self.imagen_actual_idx < 0:
            return
        self._resultados_batch[self.imagen_actual_idx] = self.resultado_actual
        self._actualizar_nav()

    def _ajustar_vistas_resultado(self):
        claves = ["Original"]
        for metodo in self.METODOS:
            claves.append(metodo)
            claves.append(f"{metodo}_Pasos")
        for key in claves:
            estado = self._get_view_state(key)
            estado["zoom_level"] = None
            estado["pan_offset_x"] = 0.0
            estado["pan_offset_y"] = 0.0
            estado["image_size"] = None
            self._drag_pan_state.pop(key, None)

    def _restaurar_resultado_guardado(self, resultado, preservar_vistas=False):
        self.resultado_actual = resultado
        self._normalizar_fallas_resultado(resultado)
        self.motor.calibrador.px_por_mm = resultado.get("calibracion_px_mm")
        self.motor.calibrador.angulo_eje_via = resultado.get("angulo_eje", 90.0)
        self.motor.calibrador.ancho_via_real_m = resultado.get(
            "ancho_via_real_m",
            self.motor.config.get("ancho_via_real_m", 6.5),
        )
        if not preservar_vistas:
            self._ajustar_vistas_resultado()
        self._redibujar_todas_las_vistas()
        self._actualizar_resumenes(resultado)
        self._actualizar_datos_imagen_panel(resultado=resultado)
        self._actualizar_estado_calibracion_ui(resultado)
        if not preservar_vistas:
            self.after_idle(self._zoom_fit_current_view)
        self._estado("Resultado restaurado")

    def _mostrar_estado_imagen_actual(self):
        idx = self.imagen_actual_idx
        if idx in self._resultados_batch:
            self._restaurar_resultado_guardado(self._resultados_batch[idx])
        else:
            self.resultado_actual = None
            self._mostrar_original_actual()

    def _calibrar_unica(self, auto_procesar=False):
        if not self.imagenes_cargadas:
            ruta = filedialog.askopenfilename(
                title="Seleccionar imagen para calibrar",
                initialdir=self.motor.config.get("ruta_imagenes", "."),
                filetypes=[("Imagenes", "*.jpg *.jpeg *.png *.bmp *.tif *.tiff"), ("Todos", "*.*")],
            )
            if not ruta:
                return
        else:
            ruta = self.imagenes_cargadas[self.imagen_actual_idx]
        imagen = cv2.imread(str(ruta))
        if imagen is None:
            self._log("ERROR: No se pudo cargar la imagen para calibrar.")
            return
        self._sync_config()
        self.motor.calibrador.ancho_via_real_m = self.motor.config.get("ancho_via_real_m", 6.5)
        self._log(f"Calibrando imagen de referencia: {Path(ruta).name}")
        self.motor.calibrar_imagen(imagen, usar_gui=True, parent=self)
        self._aplicar_config_en_vars({
            "usar_borde_berma_pci": self.motor.config.get("usar_borde_berma_pci", True),
            "borde_interno_m": self.motor.config.get("borde_interno_m", 0.30),
            "borde_externo_m": self.motor.config.get("borde_externo_m", 0.30),
        })
        if self.motor.calibrador.px_por_mm:
            linea_ancho_rel = None
            pt_ini_ancho = getattr(self.motor.calibrador, "punto_ancho_inicio", None)
            pt_fin_ancho = getattr(self.motor.calibrador, "punto_ancho_fin", None)
            alto_img, ancho_img = imagen.shape[:2]
            if pt_ini_ancho and pt_fin_ancho and ancho_img > 0 and alto_img > 0:
                linea_ancho_rel = (
                    (float(pt_ini_ancho[0]) / float(ancho_img), float(pt_ini_ancho[1]) / float(alto_img)),
                    (float(pt_fin_ancho[0]) / float(ancho_img), float(pt_fin_ancho[1]) / float(alto_img)),
                )
            self._calibracion_unica_guardada = {
                "px_por_mm": self.motor.calibrador.px_por_mm,
                "angulo_eje_via": self.motor.calibrador.get_angulo_eje(),
                "linea_ancho_rel": linea_ancho_rel,
            }
            if isinstance(self.vars.get("modo_calibracion"), tk.StringVar):
                self.vars["modo_calibracion"].set("unica")
            self._actualizar_estado_calibracion_ui()
            self._estado("Calibracion unica guardada")
            self._log(
                f"Calibracion unica guardada: {self.motor.calibrador.px_por_mm:.4f} px/mm | "
                f"Eje {self.motor.calibrador.get_angulo_eje():.1f} grados"
            )
            if auto_procesar:
                self.after(100, self._procesar_actual)
        else:
            self._actualizar_estado_calibracion_ui()
            self._estado("Calibracion cancelada")
            self._log("Calibracion unica cancelada o fallida.")

    def _calibrar_actual(self):
        modo = self._modo_calibracion_actual()
        if modo == "unica":
            self._calibrar_unica(auto_procesar=False)
            return
        ruta = self._ruta_actual()
        if not ruta:
            messagebox.showinfo("Info", "Cargue una imagen primero.")
            return
        img = cv2.imread(str(ruta))
        if img is None:
            messagebox.showerror("Error", "No se pudo cargar la imagen actual.")
            return
        self._sync_config()
        self.motor.calibrador.ancho_via_real_m = self.motor.config.get("ancho_via_real_m", 6.5)
        self.motor.calibrar_imagen(img, usar_gui=True, parent=self)
        self._aplicar_config_en_vars({
            "usar_borde_berma_pci": self.motor.config.get("usar_borde_berma_pci", True),
            "borde_interno_m": self.motor.config.get("borde_interno_m", 0.30),
            "borde_externo_m": self.motor.config.get("borde_externo_m", 0.30),
        })
        if self.motor.calibrador.px_por_mm:
            info = {
                "calibracion_px_mm": self.motor.calibrador.px_por_mm,
                "angulo_eje": self.motor.calibrador.get_angulo_eje(),
            }
            self._actualizar_estado_calibracion_ui(info)
            self._log(
                f"Calibracion manual actualizada: {self.motor.calibrador.px_por_mm:.4f} px/mm | "
                f"Eje {self.motor.calibrador.get_angulo_eje():.1f} grados"
            )
            self._estado("Calibracion manual guardada")

    def _preparar_calibracion_para_procesar(self, ruta):
        modo = self._modo_calibracion_actual()
        self.calibrar_cada_imagen = (modo == "cada_imagen")
        self.motor.calibrador.ancho_via_real_m = self.motor.config.get("ancho_via_real_m", 6.5)

        if modo == "unica":
            if not self._calibracion_unica_guardada:
                if messagebox.askyesno(
                    "Calibracion unica",
                    "No hay calibracion unica guardada.\nDesea calibrar ahora con una imagen?",
                ):
                    self._calibrar_unica(auto_procesar=True)
                else:
                    self._estado("Procesamiento cancelado")
                return False
            self.motor.calibrador.px_por_mm = self._calibracion_unica_guardada["px_por_mm"]
            self.motor.calibrador.angulo_eje_via = self._calibracion_unica_guardada["angulo_eje_via"]
            self.motor.calibrador.set_linea_ancho(None, None)
            linea_ancho_rel = self._calibracion_unica_guardada.get("linea_ancho_rel")
            if linea_ancho_rel:
                imagen = cv2.imread(str(ruta))
                if imagen is not None:
                    alto_img, ancho_img = imagen.shape[:2]
                    try:
                        pt_ini_ancho = (
                            int(float(linea_ancho_rel[0][0]) * ancho_img),
                            int(float(linea_ancho_rel[0][1]) * alto_img),
                        )
                        pt_fin_ancho = (
                            int(float(linea_ancho_rel[1][0]) * ancho_img),
                            int(float(linea_ancho_rel[1][1]) * alto_img),
                        )
                        self.motor.calibrador.set_linea_ancho(pt_ini_ancho, pt_fin_ancho)
                    except Exception:
                        self.motor.calibrador.set_linea_ancho(None, None)
            self._actualizar_estado_calibracion_ui()
            self._log(
                f"Usando calibracion unica: {self.motor.calibrador.px_por_mm:.4f} px/mm | "
                f"Eje {self.motor.calibrador.get_angulo_eje():.1f} grados"
            )
            return True

        if modo == "cada_imagen":
            imagen = cv2.imread(str(ruta))
            if imagen is None:
                messagebox.showerror("Error", "No se pudo cargar la imagen actual para calibrar.")
                return False
            self._log(f"Calibrando imagen actual antes de procesar: {Path(ruta).name}")
            self.motor.calibrar_imagen(imagen, usar_gui=True, parent=self)
            info = {
                "calibracion_px_mm": self.motor.calibrador.px_por_mm,
                "angulo_eje": self.motor.calibrador.get_angulo_eje(),
            }
            self._actualizar_estado_calibracion_ui(info)
            return True

        self.motor.calibrador.px_por_mm = None
        self.motor.calibrador.linea_px = None
        self.motor.calibrador.set_linea_ancho(None, None)
        self.motor.calibrador.angulo_eje_via = 90.0
        self._actualizar_estado_calibracion_ui()
        return True

    def _liberar_proyecto_activo(self, reset_path=True):
        workspace = getattr(self, "_proyecto_workspace", None)
        if workspace:
            shutil.rmtree(workspace, ignore_errors=True)
        self._proyecto_workspace = None
        if reset_path:
            self._proyecto_activo_path = None

    def _ruta_interna_proyecto(self, idx, ruta):
        nombre = GestorProyectoPavimentos._nombre_seguro(Path(ruta).name, fallback=f"imagen_{idx + 1:03d}")
        return f"images/{idx:04d}_{nombre}"

    def _serializar_resultado_proyecto(self, resultado):
        payload = {
            "nombre": resultado.get("nombre"),
            "dimensiones": tuple(resultado.get("dimensiones", (0, 0))),
            "calibracion_px_mm": resultado.get("calibracion_px_mm"),
            "ancho_via_real_m": resultado.get("ancho_via_real_m"),
            "angulo_eje": resultado.get("angulo_eje", 90.0),
            "tiempo_procesamiento_s": resultado.get("tiempo_procesamiento_s"),
            "raw_masks": resultado.get("raw_masks"),
            "metodos": {},
        }
        for metodo in self.METODOS:
            data_metodo = resultado.get("metodos", {}).get(metodo, {}) or {}
            payload["metodos"][metodo] = {
                "fallas": data_metodo.get("fallas", []),
            }
        return payload

    def _normalizar_fallas_importadas_proyecto(self, metodo, fallas):
        if metodo != "VIZIR":
            return fallas
        for falla in fallas:
            tipo = str(falla.get("tipo", "") or "").upper()
            # Migrar nombre antiguo al nuevo
            if tipo == "FISURA TRANSVERSAL DE JUNTA DE CONSTRUCCION":
                falla["tipo"] = ProcesadorFisuras.NOMBRE_TRANSVERSAL
                tipo = falla["tipo"]
            if "FISURA" in tipo and "COCODRILO" not in tipo:
                falla["severidad"] = ProcesadorFisurasVIZIR._clasificar_severidad(
                    falla.get("espesor_mm")
                )
        return fallas

    def _deserializar_resultado_proyecto(self, payload, ruta_imagen):
        imagen = cv2.imread(str(ruta_imagen))
        if imagen is None:
            raise ValueError(f"No se pudo restaurar la imagen del proyecto: {ruta_imagen}")

        dims = payload.get("dimensiones") or (imagen.shape[1], imagen.shape[0])
        resultado = {
            "imagen": imagen,
            "nombre": payload.get("nombre") or Path(ruta_imagen).name,
            "dimensiones": tuple(dims),
            "calibracion_px_mm": payload.get("calibracion_px_mm"),
            "ancho_via_real_m": payload.get("ancho_via_real_m", self.motor.config.get("ancho_via_real_m", 6.5)),
            "angulo_eje": payload.get("angulo_eje", 90.0),
            "raw_masks": payload.get("raw_masks"),
            "metodos": {},
        }
        if payload.get("tiempo_procesamiento_s") is not None:
            resultado["tiempo_procesamiento_s"] = payload.get("tiempo_procesamiento_s")

        for metodo in self.METODOS:
            data_metodo = payload.get("metodos", {}).get(metodo, {}) or {}
            fallas = self._normalizar_fallas_importadas_proyecto(
                metodo,
                data_metodo.get("fallas", []),
            )
            resultado["metodos"][metodo] = {
                "fallas": fallas,
                "visual": None,
                "pasos_visual": None,
                "resumen_severidad": self.motor._resumen_severidad(fallas),
            }
        return resultado

    def _consolidar_estado_pendiente_proyecto(self):
        self._sync_config()
        idx = self.imagen_actual_idx
        if idx is not None and idx >= 0:
            self._guardar_config_imagen(idx)

        if getattr(self, "_config_general_scope", "actual") != "todas":
            return

        valores_generales = self._snapshot_config_general()
        if not self._config_global_base:
            self._config_global_base = self._snapshot_config_avanzada()
        for key, value in valores_generales.items():
            self._config_global_base[key] = copy.deepcopy(value)

        for idx_cfg, config_img in list(self._config_por_imagen.items()):
            cfg_actual = copy.deepcopy(config_img or {})
            for key, value in valores_generales.items():
                cfg_actual[key] = copy.deepcopy(value)
            self._config_por_imagen[idx_cfg] = cfg_actual

    def _construir_estado_proyecto(self):
        self._consolidar_estado_pendiente_proyecto()
        self._cachear_resultado_actual()

        imagenes = []
        for idx, ruta in enumerate(self.imagenes_cargadas):
            imagenes.append({
                "index": idx,
                "name": Path(ruta).name,
                "source_path": str(ruta),
                "original_path": self._rutas_origen_imagen.get(idx, str(ruta)),
                "archive_path": self._ruta_interna_proyecto(idx, ruta),
            })

        resultados_batch = {}
        for idx, resultado in sorted(self._resultados_batch.items()):
            if resultado is None:
                continue
            resultados_batch[idx] = self._serializar_resultado_proyecto(resultado)

        return {
            "app": "AplicacionIntegradaTresMetodosAvanzada",
            "config": copy.deepcopy(self.motor.config),
            "config_global_base": copy.deepcopy(self._config_global_base),
            "config_por_imagen": copy.deepcopy(self._config_por_imagen),
            "config_ui_actual": self._config_actual_para_ui(),
            "config_general_scope": getattr(self, "_config_general_scope", "actual"),
            "modo_calibracion": self._modo_calibracion_actual(),
            "modo_vista_por_metodo": copy.deepcopy(self._modo_vista_por_metodo),
            "vista_actual": self._current_view_key(),
            "view_state": copy.deepcopy(self._view_state),
            "calibracion_unica_guardada": copy.deepcopy(self._calibracion_unica_guardada),
            "imagen_actual_idx": int(self.imagen_actual_idx if self.imagen_actual_idx >= 0 else 0),
            "imagenes": imagenes,
            "resultados_batch": resultados_batch,
        }

    def _config_desde_proyecto(self, payload):
        config = copy.deepcopy(CONFIG_DEFAULT)
        config.update(payload.get("config", {}) or {})
        config.setdefault("filtrar_baches", True)
        config.setdefault("filtrar_grietas", True)
        config.setdefault("usar_borde_berma_pci", True)
        config.setdefault("borde_interno_m", 0.30)
        config.setdefault("borde_externo_m", 0.30)
        config.setdefault("filtrar_parches", True)
        config.setdefault("parche_ratio_leve_max", 0.08)
        config.setdefault("parche_ratio_moderado_max", 0.18)
        config.setdefault("filtrar_piel", True)
        config.setdefault("merge_fisuras_px", 30)
        config.setdefault("usar_refinamiento", True)
        return config

    def _restaurar_estado_modelo_proyecto(self):
        ruta_modelo = str(self.motor.config.get("ruta_modelo", "") or "").strip()
        self.motor.modelo = None
        self.motor.modelo_cargado = False

        if ruta_modelo and YOLO_OK and Path(ruta_modelo).is_file():
            ok, msg = self.motor.cargar_modelo(ruta_modelo)
            if ok:
                self.lbl_modelo.config(text=f"Modelo: {Path(ruta_modelo).name}", fg=EstiloUI.BG_SUCCESS)
                self.lbl_modelo_path.config(text=f"Modelo: {ruta_modelo}")
                self._log(f"Modelo restaurado automaticamente: {Path(ruta_modelo).name}")
                return
            self._log(f"No se pudo restaurar automaticamente el modelo: {msg}")

        if ruta_modelo:
            self.lbl_modelo.config(text="Modelo: pendiente", fg=EstiloUI.FG_HIGHLIGHT)
            self.lbl_modelo_path.config(text=f"Modelo guardado en proyecto: {ruta_modelo}")
            if not YOLO_OK:
                self._log("Proyecto abierto sin auto-cargar modelo: ultralytics no disponible.")
            elif not Path(ruta_modelo).is_file():
                self._log(f"Modelo pendiente de carga manual. Ruta guardada: {ruta_modelo}")
        else:
            self.lbl_modelo.config(text="Modelo: no cargado", fg=EstiloUI.FG_HIGHLIGHT)
            self.lbl_modelo_path.config(text="Modelo: no cargado")

    def _reconstruir_todos_resultados_motor(self):
        self.motor.todos_resultados = {}
        for _, resultado in sorted(self._resultados_batch.items()):
            if not resultado:
                continue
            nombre = resultado.get("nombre") or "imagen"
            self.motor.todos_resultados[nombre] = {
                metodo: resultado.get("metodos", {}).get(metodo, {}).get("fallas", [])
                for metodo in self.METODOS
            }

    def _aplicar_estado_proyecto(self, payload, workspace, ruta_proyecto, progress_callback=None):
        config = self._config_desde_proyecto(payload)
        config_global_base = copy.deepcopy(payload.get("config_global_base", {}) or {})
        config_por_imagen = copy.deepcopy(payload.get("config_por_imagen", {}) or {})
        calibracion_unica = copy.deepcopy(payload.get("calibracion_unica_guardada"))
        config_general_scope = str(payload.get("config_general_scope", "actual") or "actual")
        if config_general_scope not in {"actual", "todas"}:
            config_general_scope = "actual"
        modo_vista = {metodo: "resultado" for metodo in self.METODOS}
        modo_vista.update(payload.get("modo_vista_por_metodo", {}) or {})
        vista_actual = str(payload.get("vista_actual", "Original") or "Original")
        if vista_actual not in ("Original", *self.METODOS):
            vista_actual = "Original"
        view_state = copy.deepcopy(payload.get("view_state", {}) or {})
        if not isinstance(view_state, dict):
            view_state = {}

        def report(valor, texto):
            if progress_callback:
                progress_callback(valor, texto)

        imagenes_info = sorted(payload.get("imagenes", []) or [], key=lambda item: item.get("index", 0))
        imagenes_cargadas = []
        rutas_origen = {}
        total_imagenes = max(len(imagenes_info), 1)
        for pos, info in enumerate(imagenes_info):
            ruta_extraida = Path(workspace) / str(info.get("archive_path", ""))
            if not ruta_extraida.exists():
                raise FileNotFoundError(f"Falta una imagen dentro del proyecto: {info.get('archive_path')}")
            imagenes_cargadas.append(str(ruta_extraida))
            rutas_origen[pos] = info.get("original_path", info.get("source_path", str(ruta_extraida)))
            report(82 + (6 * (pos + 1) / total_imagenes), f"Restaurando imagen {pos + 1}/{len(imagenes_info)}")

        resultados_batch = {}
        resultados_guardados = payload.get("resultados_batch", {}) or {}
        total_resultados = max(len(resultados_guardados), 1)
        for idx, data in sorted(resultados_guardados.items()):
            idx_int = int(idx)
            if idx_int < 0 or idx_int >= len(imagenes_cargadas):
                continue
            resultados_batch[idx_int] = self._deserializar_resultado_proyecto(data, imagenes_cargadas[idx_int])
            report(88 + (8 * len(resultados_batch) / total_resultados), f"Reconstruyendo resultados {len(resultados_batch)}/{len(resultados_guardados)}")

        imagen_actual_idx = int(payload.get("imagen_actual_idx", 0) or 0)
        if imagenes_cargadas:
            imagen_actual_idx = max(0, min(imagen_actual_idx, len(imagenes_cargadas) - 1))
        else:
            imagen_actual_idx = -1

        config_ui = copy.deepcopy(config)
        config_ui.update(payload.get("config_ui_actual") or {})

        self._proyecto_workspace = str(workspace)
        self._proyecto_activo_path = str(ruta_proyecto)
        self.motor.config = config
        self.motor.calibrador.ancho_via_real_m = self.motor.config.get("ancho_via_real_m", 6.5)
        self._config_global_base = config_global_base
        self._config_por_imagen = config_por_imagen
        self._config_general_scope = config_general_scope
        self._calibracion_unica_guardada = calibracion_unica
        self._modo_vista_por_metodo = modo_vista
        self._summary_cache = {"General": ""}
        self._resultados_batch = resultados_batch
        self._rutas_origen_imagen = rutas_origen
        self.imagenes_cargadas = imagenes_cargadas
        self.imagen_actual_idx = imagen_actual_idx
        self._view_state = view_state
        self._drag_pan_state.clear()

        if isinstance(self.vars.get("modo_calibracion"), tk.StringVar):
            self.vars["modo_calibracion"].set(payload.get("modo_calibracion", "automatica"))
        if hasattr(self, "var_config_general_scope"):
            self.var_config_general_scope.set(self._config_general_scope)
            self._actualizar_scope_general_ui()

        self._aplicar_config_en_vars(config_ui)
        self._sync_config()
        report(97, "Restaurando configuracion visual...")
        self._restaurar_estado_modelo_proyecto()
        self._reconstruir_todos_resultados_motor()

        if self.imagenes_cargadas:
            self._actualizar_nav()
            if self.imagen_actual_idx in self._resultados_batch:
                self._restaurar_resultado_guardado(self._resultados_batch[self.imagen_actual_idx], preservar_vistas=True)
            else:
                self.resultado_actual = None
                self._mostrar_original_actual(reset_modos=False)
            self._seleccionar_metodo_en_ui(vista_actual)
        else:
            self.resultado_actual = None
            self._actualizar_nav()
            self._actualizar_datos_imagen_panel()
            self._actualizar_estado_calibracion_ui()

        report(100, "Proyecto cargado.")
        self._estado(f"Proyecto cargado: {Path(ruta_proyecto).name}")
        self._log(f"Proyecto abierto: {ruta_proyecto}")

    def _guardar_proyecto(self, forzar_dialogo=False):
        if self.procesando:
            messagebox.showinfo("Proyecto", "Espere a que termine el procesamiento antes de guardar el proyecto.")
            return
        if not self.imagenes_cargadas and not self._resultados_batch and not self.resultado_actual:
            messagebox.showinfo("Proyecto", "No hay imagenes ni resultados para guardar.")
            return

        initialdir = _resolver_directorio_dialogo("proyectos", self.motor.config.get("ruta_proyectos"))
        if self._proyecto_activo_path:
            initialdir = str(Path(self._proyecto_activo_path).parent)

        nombre_base = "Proyecto_Pavimentos"
        ruta_actual = self._ruta_actual()
        if ruta_actual:
            nombre_base = f"{Path(ruta_actual).stem}_Proyecto"

        ruta = self._proyecto_activo_path if self._proyecto_activo_path and not forzar_dialogo else None
        if not ruta:
            ruta = filedialog.asksaveasfilename(
                title="Guardar proyecto de pavimentos",
                defaultextension=GestorProyectoPavimentos.EXTENSION,
                filetypes=[("Proyecto Pavimentos", f"*{GestorProyectoPavimentos.EXTENSION}"), ("Todos", "*.*")],
                initialdir=initialdir,
                initialfile=f"{nombre_base}{GestorProyectoPavimentos.EXTENSION}",
            )
            if not ruta:
                return

        try:
            estado = self._construir_estado_proyecto()
            guardado = GestorProyectoPavimentos.guardar(ruta, estado)
            self._proyecto_activo_path = guardado
            self.motor.config["ruta_proyectos"] = str(Path(guardado).parent)
            self._estado(f"Proyecto guardado: {Path(guardado).name}")
            self._log(f"Proyecto guardado: {guardado}")
            messagebox.showinfo("Proyecto", f"Guardado de proyecto concluido.\nArchivo: {guardado}")
        except Exception as exc:
            messagebox.showerror("Proyecto", f"No se pudo guardar el proyecto:\n{exc}")
            self._estado("Error al guardar proyecto")
            self._log(f"ERROR guardando proyecto: {exc}")

    def _abrir_proyecto(self):
        if self.procesando:
            messagebox.showinfo("Proyecto", "Espere a que termine el procesamiento antes de abrir otro proyecto.")
            return
        initialdir = _resolver_directorio_dialogo("proyectos", self.motor.config.get("ruta_proyectos"))
        if self._proyecto_activo_path:
            initialdir = str(Path(self._proyecto_activo_path).parent)

        ruta = filedialog.askopenfilename(
            title="Abrir proyecto de pavimentos",
            filetypes=[("Proyecto Pavimentos", f"*{GestorProyectoPavimentos.EXTENSION}"), ("Todos", "*.*")],
            initialdir=initialdir,
        )
        if not ruta:
            return
        self.motor.config["ruta_proyectos"] = str(Path(ruta).parent)

        cargado = None
        mensaje_exito = None
        workspace_anterior = self._proyecto_workspace
        proyecto_anterior = self._proyecto_activo_path
        self._mostrar_progreso_proyecto("Abriendo proyecto...", 0)
        self._estado("Abriendo proyecto...")

        def on_progress(valor, texto):
            self._actualizar_progreso_proyecto(valor, texto)

        try:
            cargado = GestorProyectoPavimentos.cargar(ruta, progress_callback=on_progress)
            self._aplicar_estado_proyecto(cargado["state"], cargado["workspace"], ruta, progress_callback=on_progress)
            if workspace_anterior and workspace_anterior != self._proyecto_workspace:
                shutil.rmtree(workspace_anterior, ignore_errors=True)
            mensaje_exito = f"Importacion de proyecto concluida.\nArchivo: {ruta}"
        except Exception as exc:
            if cargado and cargado.get("workspace"):
                shutil.rmtree(cargado["workspace"], ignore_errors=True)
            self._proyecto_workspace = workspace_anterior
            self._proyecto_activo_path = proyecto_anterior
            messagebox.showerror("Proyecto", f"No se pudo abrir el proyecto:\n{exc}")
            self._estado("Error al abrir proyecto")
            self._log(f"ERROR abriendo proyecto: {exc}")
        finally:
            self._ocultar_progreso_proyecto()
        if mensaje_exito:
            messagebox.showinfo("Proyecto", mensaje_exito)

    def _actualizar_nav(self):
        if not self.imagenes_cargadas:
            self.lbl_nav.config(text="Sin imagenes cargadas")
            self.lbl_info_img.config(text="--")
            return
        ruta = Path(self.imagenes_cargadas[self.imagen_actual_idx])
        procesada = "OK" if self.imagen_actual_idx in self._resultados_batch else "--"
        self.lbl_nav.config(text=f"{self.imagen_actual_idx + 1}/{len(self.imagenes_cargadas)} | {ruta.name} [{procesada}]")
        self.lbl_info_img.config(text=str(ruta))

    def _actualizar_datos_imagen_panel(self, ruta=None, img=None, resultado=None):
        if not hasattr(self, "lbl_datos_imagen"):
            return
        if resultado is not None:
            ancho, alto = resultado["dimensiones"]
            lineas = [
                f"Archivo: {resultado['nombre']}",
                f"Dimensiones: {ancho}x{alto}",
                f"Calibracion: {resultado['calibracion_px_mm']:.4f} px/mm",
                f"Eje de via: {resultado['angulo_eje']:.1f} grados",
            ]
            tiempo = resultado.get("tiempo_procesamiento_s")
            if tiempo:
                lineas.append(f"Tiempo de procesamiento: {self._formato_tiempo_integrado(tiempo)}")
            self.lbl_datos_imagen.config(text="\n".join(lineas))
            return
        if ruta is not None and img is not None:
            h, w = img.shape[:2]
            self.lbl_datos_imagen.config(text=f"Archivo: {Path(ruta).name}\nDimensiones: {w}x{h}")
            return
        self.lbl_datos_imagen.config(text="Sin imagen cargada.")

    def _cargar_imagenes(self):
        rutas = filedialog.askopenfilenames(
            title="Seleccionar imagenes",
            filetypes=[("Imagenes", "*.jpg *.jpeg *.png *.bmp *.tif *.tiff"), ("Todos", "*.*")],
            initialdir=_resolver_directorio_dialogo("imagenes", self.motor.config.get("ruta_imagenes")),
        )
        if not rutas:
            return
        self._liberar_proyecto_activo()
        self.motor.config["ruta_imagenes"] = str(Path(rutas[0]).parent)
        self.imagenes_cargadas = list(rutas)
        self._rutas_origen_imagen = {idx: str(ruta) for idx, ruta in enumerate(self.imagenes_cargadas)}
        self.imagen_actual_idx = 0
        self.resultado_actual = None
        self._resultados_batch = {}
        self.motor.todos_resultados = {}
        self._config_por_imagen.clear()
        self._view_state.clear()
        self._drag_pan_state.clear()
        self._proyecto_activo_path = None
        if hasattr(self, "var_config_general_scope"):
            self.var_config_general_scope.set(self._config_general_scope)
            self._actualizar_scope_general_ui()
        self._aplicar_config_en_vars(self._config_actual_para_ui())
        self._sync_config()
        self._actualizar_nav()
        self._mostrar_estado_imagen_actual()
        self._actualizar_estado_calibracion_ui()
        self._estado(f"{len(rutas)} imagen(es) cargada(s)")
        self._log(f"Se cargaron {len(rutas)} imagen(es).")

    def _anadir_imagenes(self):
        if self.procesando:
            messagebox.showinfo("Imagenes", "Espere a que termine el procesamiento antes de agregar mas imagenes.")
            return

        rutas = filedialog.askopenfilenames(
            title="Agregar mas imagenes",
            filetypes=[("Imagenes", "*.jpg *.jpeg *.png *.bmp *.tif *.tiff"), ("Todos", "*.*")],
            initialdir=_resolver_directorio_dialogo("imagenes", self.motor.config.get("ruta_imagenes")),
        )
        if not rutas:
            return
        self.motor.config["ruta_imagenes"] = str(Path(rutas[0]).parent)

        tenia_imagenes = bool(self.imagenes_cargadas)
        existentes = {os.path.normcase(os.path.normpath(str(ruta))) for ruta in self.imagenes_cargadas}
        nuevas = []
        duplicadas = 0
        for ruta in rutas:
            clave = os.path.normcase(os.path.normpath(str(ruta)))
            if clave in existentes:
                duplicadas += 1
                continue
            existentes.add(clave)
            nuevas.append(str(ruta))

        if not nuevas:
            msg = "Las imagenes seleccionadas ya estaban cargadas."
            if duplicadas:
                msg = f"No se agregaron imagenes nuevas. Duplicadas omitidas: {duplicadas}."
            messagebox.showinfo("Imagenes", msg)
            self._estado("Sin imagenes nuevas para agregar")
            self._log(msg)
            return

        idx_inicio = len(self.imagenes_cargadas)
        self.imagenes_cargadas.extend(nuevas)
        for offset, ruta in enumerate(nuevas):
            self._rutas_origen_imagen[idx_inicio + offset] = str(ruta)

        if not tenia_imagenes:
            self.imagen_actual_idx = 0
            self.resultado_actual = None
            self._aplicar_config_en_vars(self._config_actual_para_ui())
            self._sync_config()
            self._mostrar_estado_imagen_actual()
            self._actualizar_estado_calibracion_ui()

        self._actualizar_nav()
        total = len(self.imagenes_cargadas)
        msg = f"Se agregaron {len(nuevas)} imagen(es). Total cargadas: {total}."
        if duplicadas:
            msg += f" Duplicadas omitidas: {duplicadas}."
        self._estado(msg)
        self._log(msg)

    def _procesar_actual(self):
        ruta = self._ruta_actual()
        if not ruta:
            messagebox.showinfo("Info", "Cargue una imagen primero.")
            return
        if not self.motor.modelo_cargado:
            messagebox.showinfo("Info", "Cargue el modelo YOLO primero.")
            return
        if self.procesando:
            return
        self._sync_config()
        if not self._preparar_calibracion_para_procesar(ruta):
            return
        self._detener_flag = False
        self.procesando = True
        self._set_detener_enabled(True)
        self._iniciar_cronometro_integrado()
        self._estado("Procesando...")
        self._log(f"\n{'=' * 60}\nProcesando {Path(ruta).name}")

        def worker():
            try:
                resultado = self.motor.procesar_imagen(ruta, callback_log=self._log_safe, calibrar_gui=False, parent=self)
                self.after(0, lambda: self._on_procesado(resultado))
            except Exception as exc:
                tb = traceback.format_exc()
                self.after(0, lambda: self._on_error(exc, tb))

        Thread(target=worker, daemon=True).start()

    def _procesar_todo(self):
        if not self.imagenes_cargadas:
            messagebox.showinfo("Info", "Cargue imagenes primero.")
            return
        if not self.motor.modelo_cargado:
            messagebox.showinfo("Info", "Cargue el modelo YOLO primero.")
            return
        if self.procesando:
            return

        modo = self._modo_calibracion_actual()
        if modo == "unica" and not self._calibracion_unica_guardada:
            if messagebox.askyesno(
                "Calibracion unica",
                "No hay calibracion unica guardada.\nDesea calibrar ahora con una imagen?",
            ):
                self._calibrar_unica(auto_procesar=False)
            if not self._calibracion_unica_guardada:
                self._estado("Procesamiento por lote cancelado")
                return
        elif modo == "cada_imagen":
            continuar = messagebox.askyesno(
                "Procesar todo",
                "Se abrira la calibracion manual para cada imagen del lote.\nDesea continuar?",
            )
            if not continuar:
                return

        self._sync_config()
        self._detener_flag = False
        self.procesando = True
        self._set_detener_enabled(True)
        self._iniciar_cronometro_integrado()
        self._batch_ctx = {
            "index": 0,
            "total": len(self.imagenes_cargadas),
            "ok": 0,
            "errores": 0,
            "config_base": copy.deepcopy(self._config_global_base) if self._config_global_base else self._snapshot_config_avanzada(),
            "modo_calibracion": modo,
        }
        self._estado("Procesando lote...")
        self._log(f"\n{'=' * 60}\nProcesando lote completo ({len(self.imagenes_cargadas)} imagenes)")
        self.after(10, self._procesar_siguiente_lote)

    def _config_para_indice(self, idx, config_base):
        config_idx = copy.deepcopy(config_base)
        if idx in self._config_por_imagen:
            config_idx.update(self._config_por_imagen[idx])
        return config_idx

    def _procesar_siguiente_lote(self):
        ctx = getattr(self, "_batch_ctx", None)
        if not ctx:
            return
        if self._detener_flag:
            self._finalizar_procesar_todo(detenido=True)
            return
        if ctx["index"] >= ctx["total"]:
            self._finalizar_procesar_todo()
            return

        idx = ctx["index"]
        ruta = self.imagenes_cargadas[idx]
        self.imagen_actual_idx = idx
        config_idx = self._config_para_indice(idx, ctx["config_base"])
        self._aplicar_config_en_vars(config_idx)
        self._sync_config()
        self._actualizar_nav()
        self._mostrar_original_actual()

        if not self._preparar_calibracion_para_procesar(ruta):
            if not self.procesando:
                return
            self._batch_ctx["errores"] += 1
            self._batch_ctx["index"] += 1
            self.after(10, self._procesar_siguiente_lote)
            return

        self._batch_ctx["item_start"] = time.time()
        self._log(f"[{idx + 1}/{ctx['total']}] Procesando {Path(ruta).name}")

        def worker():
            try:
                resultado = self.motor.procesar_imagen(ruta, callback_log=self._log_safe, calibrar_gui=False, parent=self)
                self.after(0, lambda i=idx, r=resultado: self._on_procesado_lote(i, r))
            except Exception as exc:
                tb = traceback.format_exc()
                self.after(0, lambda i=idx, err=exc, det=tb: self._on_error_lote(i, err, det))

        Thread(target=worker, daemon=True).start()

    def _on_procesado_lote(self, idx, resultado):
        ctx = getattr(self, "_batch_ctx", None)
        if not ctx:
            return
        elapsed = time.time() - ctx.get("item_start", time.time())
        if resultado is not None:
            resultado["tiempo_procesamiento_s"] = elapsed
            self.resultado_actual = resultado
            self._normalizar_fallas_resultado(resultado)
            self._registrar_config_procesada(idx)
            self._cachear_resultado_actual()
            self._restaurar_resultado_guardado(resultado)
            ctx["ok"] += 1
            self._log(f"Tiempo imagen: {self._formato_tiempo_integrado(elapsed)}")
        else:
            ctx["errores"] += 1
            self._log("Imagen sin resultado procesable.")
        ctx["index"] += 1
        if self._detener_flag:
            self._finalizar_procesar_todo(detenido=True)
            return
        self.after(10, self._procesar_siguiente_lote)

    def _on_error_lote(self, idx, error, detalle):
        ctx = getattr(self, "_batch_ctx", None)
        if ctx:
            ctx["errores"] += 1
        nombre = Path(self.imagenes_cargadas[idx]).name if 0 <= idx < len(self.imagenes_cargadas) else f"#{idx}"
        self._log(f"Error en lote ({nombre}): {error}")
        self._log(detalle.rstrip())
        if ctx:
            ctx["index"] += 1
            self.after(10, self._procesar_siguiente_lote)
        else:
            self._on_error(error, detalle)

    def _finalizar_procesar_todo(self, detenido=False):
        ctx = getattr(self, "_batch_ctx", None)
        self._batch_ctx = None
        self.procesando = False
        total_elapsed = self._detener_cronometro_integrado()
        self._set_detener_enabled(False)
        total = ctx["total"] if ctx else 0
        ok = ctx["ok"] if ctx else 0
        errores = ctx["errores"] if ctx else 0
        self._estado(f"{'Lote detenido' if detenido else 'Lote completado'} ({ok}/{total})")
        self._log(f"\n{'=' * 60}")
        self._log(f"{'Lote detenido' if detenido else 'Lote completado'}: {ok}/{total} imagen(es) procesadas")
        if errores:
            self._log(f"Errores en lote: {errores}")
        self._log(f"Tiempo total lote: {self._formato_tiempo_integrado(total_elapsed)}")
        self._actualizar_nav()
        self._mostrar_estado_imagen_actual()
        self._seleccionar_metodo_en_ui("PCI")
        self._detener_flag = False
        if not detenido:
            msg = (
                f"Procesamiento concluido.\n"
                f"Imagenes procesadas: {ok}/{total}\n"
                f"Errores: {errores}"
            )
            self._log(msg)
            messagebox.showinfo("Procesar Todo", msg)

    def _on_error(self, error, detalle):
        self.procesando = False
        t_elapsed = self._detener_cronometro_integrado()
        self._set_detener_enabled(False)
        self._detener_flag = False
        self._estado("Error en procesamiento")
        if t_elapsed:
            self._log(f"Tiempo de procesamiento hasta el error: {self._formato_tiempo_integrado(t_elapsed)}")
        self._log(str(error))
        self._log(detalle.rstrip())

    def _imagen_anterior(self):
        if not self.imagenes_cargadas:
            return
        self._sync_config()
        self._guardar_config_imagen()
        self.imagen_actual_idx = (self.imagen_actual_idx - 1) % len(self.imagenes_cargadas)
        self._aplicar_config_en_vars(self._config_actual_para_ui())
        self._sync_config()
        self._actualizar_nav()
        self._mostrar_estado_imagen_actual()

    def _imagen_siguiente(self):
        if not self.imagenes_cargadas:
            return
        self._sync_config()
        self._guardar_config_imagen()
        self.imagen_actual_idx = (self.imagen_actual_idx + 1) % len(self.imagenes_cargadas)
        self._aplicar_config_en_vars(self._config_actual_para_ui())
        self._sync_config()
        self._actualizar_nav()
        self._mostrar_estado_imagen_actual()

    def _mostrar_original_actual(self, reset_modos=True):
        ruta = self._ruta_actual()
        if not ruta:
            return
        img = cv2.imread(str(ruta))
        if img is None:
            self._log(f"No se pudo leer: {ruta}")
            return
        self._imagenes_render = {"Original": img}
        if reset_modos:
            self._reset_canvas_zoom("Original", render=False)
        for metodo in self.METODOS:
            self._imagenes_render[metodo] = None
            self._imagenes_render[f"{metodo}_Pasos"] = None
            if reset_modos:
                self._modo_vista_por_metodo[metodo] = "resultado"
                self._cambiar_modo_vista_metodo(metodo, "resultado")
        self._render_single_view("Original")
        for metodo in self.METODOS:
            self._render_single_view(metodo)
        self._actualizar_resumenes_vacios(ruta, img)
        self._actualizar_datos_imagen_panel(ruta=ruta, img=img)
        self._actualizar_estado_calibracion_ui()

    def _actualizar_resumenes_vacios(self, ruta, img):
        h, w = img.shape[:2]
        self._summary_cache["General"] = f"Archivo: {Path(ruta).name}\nDimensiones: {w}x{h}\n\nProcese la imagen para obtener PCI, MTC y VIZIR."
        self._set_text("General", self._summary_cache["General"])
        self._actualizar_datos_imagen_panel(ruta=ruta, img=img)
        for metodo in self.METODOS:
            self._summary_cache[metodo] = "Sin resultados aun."
            panel = self._panel_metodo[metodo]
            panel["summary_label"].config(text=self._summary_cache[metodo])
            panel["count_label"].config(text="0/0 activas")
            self._reconstruir_lista_fallas(metodo, [])

    def _normalizar_fallas_resultado(self, resultado):
        for metodo in self.METODOS:
            for falla in resultado["metodos"].get(metodo, {}).get("fallas", []):
                falla.setdefault("excluida", False)

    def _on_procesado(self, resultado):
        self.procesando = False
        t_elapsed = self._detener_cronometro_integrado()
        self._set_detener_enabled(False)
        if resultado is None:
            self._detener_flag = False
            self._estado("Error en procesamiento")
            return
        self.resultado_actual = resultado
        self.resultado_actual["tiempo_procesamiento_s"] = t_elapsed
        self._normalizar_fallas_resultado(resultado)
        self._registrar_config_procesada(self.imagen_actual_idx)
        for metodo in self.METODOS:
            self._cambiar_modo_vista_metodo(metodo, "resultado")
        self._ajustar_vistas_resultado()
        self._cachear_resultado_actual()
        self._actualizar_estado_calibracion_ui(resultado)
        self._redibujar_todas_las_vistas()
        self._actualizar_resumenes(resultado)
        self._seleccionar_metodo_en_ui("PCI")
        self.after_idle(self._zoom_fit_current_view)
        if self._detener_flag:
            self._estado(f"Procesamiento detenido tras finalizar imagen actual ({self._formato_tiempo_integrado(t_elapsed)})")
            self._log(f"Detencion aplicada tras finalizar la imagen actual. Tiempo: {self._formato_tiempo_integrado(t_elapsed)}")
        else:
            self._estado(f"Procesamiento completado ({self._formato_tiempo_integrado(t_elapsed)})")
            self._log(f"Tiempo de procesamiento: {self._formato_tiempo_integrado(t_elapsed)}")
        self._detener_flag = False

    def _build_general_summary(self, resultado):
        ancho, alto = resultado["dimensiones"]
        lineas = [
            f"Archivo: {resultado['nombre']}",
            f"Dimensiones: {ancho}x{alto}",
            f"Calibracion: {resultado['calibracion_px_mm']:.4f} px/mm",
            f"Eje de via: {resultado['angulo_eje']:.1f} grados",
            "",
        ]
        tiempo_proc = resultado.get("tiempo_procesamiento_s")
        if tiempo_proc:
            lineas.insert(4, f"Tiempo proc.: {self._formato_tiempo_integrado(tiempo_proc)}")
        for metodo in self.METODOS:
            total = len(resultado["metodos"][metodo]["fallas"])
            activas = len(self._get_fallas_activas(metodo))
            lineas.append(f"{metodo}: {activas}/{total} fallas activas")
        lineas.extend([
            "",
            "Control manual:",
            "- Click en la imagen para ocultar o reactivar una falla.",
            "- Use los checks del panel derecho para palomear o despalomear fallas.",
        ])
        return "\n".join(lineas)

    def _build_method_summary(self, metodo, fallas):
        total = len(fallas)
        activas = [f for f in fallas if not f.get("excluida", False)]
        lineas = [f"Metodo: {metodo}", f"Activas: {len(activas)}/{total}"]
        sev_counts = {}
        tipo_counts = {}
        for falla in activas:
            sev = self._severity_display(metodo, falla.get("severidad"))
            sev_counts[sev] = sev_counts.get(sev, 0) + 1
            tipo = falla.get("tipo", "FALLA")
            tipo_counts[tipo] = tipo_counts.get(tipo, 0) + 1
        if sev_counts:
            orden = ["1", "2", "3", "-"] if metodo in ("MTC", "VIZIR") else ["L", "M", "H", "-"]
            partes = [f"{sev}:{sev_counts[sev]}" for sev in orden if sev in sev_counts]
            extras = [f"{sev}:{cant}" for sev, cant in sev_counts.items() if sev not in orden]
            lineas.append("Severidad: " + ", ".join(partes + extras))
        else:
            lineas.append("Severidad: sin fallas activas")
        if tipo_counts:
            lineas.append("Tipos:")
            for tipo, cant in tipo_counts.items():
                lineas.append(f"  {tipo}: {cant}")
        else:
            lineas.append("Tipos: sin fallas activas")
        lineas.append("")
        lineas.append("Pasos: seleccione el metodo y luego pulse el boton Pasos.")
        return "\n".join(lineas)

    def _severity_color(self, sev):
        return {"L": "#00d084", "M": "#9a6700", "H": "#ff5c5c"}.get(severidad_ui(sev), EstiloUI.FG_PRIMARY)

    def _severity_display(self, metodo, severidad):
        if metodo in ("MTC", "VIZIR"):
            if isinstance(severidad, (int, np.integer)):
                return str(int(severidad))
            if isinstance(severidad, (float, np.floating)):
                valor = float(severidad)
                return str(int(valor)) if valor.is_integer() else f"{valor:.2f}".rstrip("0").rstrip(".")
            sev_txt = str(severidad).strip().upper()
            if not sev_txt or sev_txt == "NONE":
                return "-"
            return {"L": "1", "M": "2", "H": "3"}.get(sev_txt, sev_txt)
        return severidad_ui(severidad)

    def _build_diameter_extras(self, falla):
        tipo = str(falla.get("tipo", "") or "").strip().upper()
        if categorizar_tipo_falla(tipo) in {"grieta", "parche"}:
            return []
        if tipo == "DAÑOS PUNTUALES":
            return []
        diametro = falla.get("diametro_promedio_mm", None)
        if diametro is None:
            diametro = falla.get("diametro_mm", 0)
        try:
            diametro = float(diametro or 0)
        except (TypeError, ValueError):
            diametro = 0.0
        return [f"D={diametro:.1f} mm"]

    def _show_area_in_summary(self, metodo, falla):
        tipo = str(falla.get("tipo", "") or "").strip().upper()
        if metodo == "PCI" and categorizar_tipo_falla(tipo) == "grieta":
            return False
        if metodo == "VIZIR" and tipo in {
            "FISURAS LONGITUDINALES POR FATIGA",
            "FISURAS DE CONTRACCION TERMICA",
            "FISURAS DE BORDE",
        }:
            return False
        return True

    def _show_thickness_in_summary(self, falla):
        tipo = str(falla.get("tipo", "") or "").strip().upper()
        if categorizar_tipo_falla(tipo) == "hueco":
            return False
        if tipo == "DAÑOS PUNTUALES":
            return False
        if tipo == "OJO DE PESCADO":
            return False
        return True

    def _build_fault_label(self, metodo, falla):
        partes = [f"{falla.get('tipo', 'FALLA')} #{falla.get('id', '-')}"]
        sev = self._severity_display(metodo, falla.get("severidad"))
        if sev != "-":
            partes.append(f"S{sev}")
        partes.append(f"{falla.get('confianza', 0) * 100:.1f}%")
        extras = self._build_diameter_extras(falla)
        if (
            falla.get("espesor_mm", 0)
            and not falla.get("diametro_promedio_mm", 0)
            and self._show_thickness_in_summary(falla)
        ):
            extras.append(f"e={falla.get('espesor_mm', 0):.1f} mm")
        if falla.get("longitud_m", 0):
            extras.append(f"L={falla.get('longitud_m', 0):.3f} m")
        if falla.get("area_m2", 0) and self._show_area_in_summary(metodo, falla):
            extras.append(f"A={falla.get('area_m2', 0):.4f} m2")
        base = " | ".join(partes)
        if extras:
            base += " | " + " | ".join(extras)
        return base

    def _actualizar_resumenes(self, resultado):
        self._actualizar_datos_imagen_panel(resultado=resultado)
        self._summary_cache["General"] = self._build_general_summary(resultado)
        self._set_text("General", self._summary_cache["General"])
        for metodo in self.METODOS:
            fallas = resultado["metodos"][metodo]["fallas"]
            texto = self._build_method_summary(metodo, fallas)
            self._summary_cache[metodo] = texto
            panel = self._panel_metodo[metodo]
            panel["summary_label"].config(text=texto)
            self._reconstruir_lista_fallas(metodo, fallas)

    def _reconstruir_lista_fallas(self, metodo, fallas):
        panel = self._panel_metodo[metodo]
        for row in panel["rows"]:
            row.destroy()
        panel["rows"].clear()
        self._falla_vars[metodo] = []
        if not fallas:
            panel["empty_label"].config(text="Sin fallas detectadas.")
            panel["empty_label"].pack(anchor="w", padx=6, pady=6)
            self._bind_summary_mousewheel(panel["empty_label"], panel["list_canvas"])
            panel["count_label"].config(text=f"0/{len(fallas)} activas")
            panel["list_canvas"].update_idletasks()
            panel["list_canvas"].configure(scrollregion=panel["list_canvas"].bbox("all"))
            return
        panel["empty_label"].pack_forget()
        activas = 0
        for idx, falla in enumerate(fallas):
            var = tk.BooleanVar(value=not falla.get("excluida", False))
            if var.get():
                activas += 1
            row = tk.Frame(panel["list_inner"], bg=EstiloUI.BG_CARD)
            row.pack(anchor="w", padx=3, pady=1)

            cb = tk.Checkbutton(
                row,
                text=self._build_fault_label(metodo, falla),
                variable=var,
                bg=EstiloUI.BG_CARD,
                fg=self._severity_color(falla.get("severidad")),
                selectcolor=EstiloUI.BG_BUTTON_SECONDARY,
                activebackground=EstiloUI.BG_CARD,
                activeforeground=self._severity_color(falla.get("severidad")),
                font=EstiloUI.FONT_SMALL,
                anchor="w",
                justify="left",
                wraplength=0,
                command=lambda m=metodo, f=falla, v=var: self._on_falla_toggle(m, f, v),
            )
            cb.pack(side="left", anchor="w")
            self._bind_summary_mousewheel(row, panel["list_canvas"])
            self._bind_summary_mousewheel(cb, panel["list_canvas"])

            if categorizar_tipo_falla(falla.get("tipo")) in {"piel", "parche"}:
                btn_cfg = tk.Button(
                    row,
                    text="\u2699",
                    font=("Segoe UI", 9, "bold"),
                    bg=EstiloUI.BG_PANEL,
                    fg=EstiloUI.FG_HIGHLIGHT,
                    activebackground=EstiloUI.BG_ACCENT,
                    activeforeground=EstiloUI.FG_PRIMARY,
                    relief="groove",
                    bd=1,
                    padx=6,
                    command=lambda m=metodo, i=idx: self._configurar_falla_individual(m, i),
                )
                btn_cfg.pack(side="right", padx=(4, 0))
                self._bind_summary_mousewheel(btn_cfg, panel["list_canvas"])

            panel["rows"].append(row)
            self._falla_vars[metodo].append((falla, var))
        panel["count_label"].config(text=f"{activas}/{len(fallas)} activas")
        panel["list_canvas"].update_idletasks()
        panel["list_canvas"].configure(scrollregion=panel["list_canvas"].bbox("all"))

    def _set_method_all(self, metodo, incluir):
        if not self.resultado_actual:
            return
        for falla in self.resultado_actual["metodos"].get(metodo, {}).get("fallas", []):
            falla["excluida"] = not incluir
        self._cachear_resultado_actual()
        self._refresh_method_views(metodo)
        self._actualizar_resumenes(self.resultado_actual)

    def _on_falla_toggle(self, metodo, falla, var):
        falla["excluida"] = not bool(var.get())
        estado = "visible" if var.get() else "oculta"
        self._log(f"{metodo}: {falla.get('tipo', 'FALLA')} #{falla.get('id', '-')} -> {estado}")
        self._cachear_resultado_actual()
        self._refresh_method_views(metodo)
        self._actualizar_resumenes(self.resultado_actual)

    def _get_fallas_activas(self, metodo):
        if not self.resultado_actual:
            return []
        fallas = self.resultado_actual["metodos"].get(metodo, {}).get("fallas", [])
        return [falla for falla in fallas if not falla.get("excluida", False)]

    def _dibujar_metodo(self, metodo, imagen, fallas):
        cfg = copy.deepcopy(self.motor.config)
        if metodo == "PCI":
            return self.motor._con_procesador_piel_pci(lambda: self.motor._dibujar_pci(imagen, fallas, cfg))
        if metodo == "MTC":
            return self.motor._dibujar_mtc(imagen, fallas, cfg)
        return self.motor._dibujar_vizir(imagen, fallas, cfg)

    def _to_bgr(self, imagen):
        if imagen is None:
            return None
        arr = np.array(imagen)
        if arr.dtype == np.bool_:
            arr = arr.astype(np.uint8) * 255
        if arr.ndim == 2:
            if arr.dtype != np.uint8:
                if np.max(arr) <= 1.0:
                    arr = (arr * 255).astype(np.uint8)
                else:
                    arr = np.clip(arr, 0, 255).astype(np.uint8)
            return cv2.cvtColor(arr, cv2.COLOR_GRAY2BGR)
        if arr.ndim == 3 and arr.shape[2] == 3:
            if arr.dtype != np.uint8:
                if np.max(arr) <= 1.0:
                    arr = (arr * 255).astype(np.uint8)
                else:
                    arr = np.clip(arr, 0, 255).astype(np.uint8)
            return np.ascontiguousarray(arr)
        return None

    def _append_step_image(self, items, titulo, imagen):
        img = self._to_bgr(imagen)
        if img is not None:
            items.append((titulo, img))

    def _formatear_titulo_paso(self, clave):
        return str(clave).replace("_", " ").title()

    def _extraer_contorno_poligono(self, poly):
        valor = poly
        if isinstance(poly, dict):
            for key in ("contorno", "contorno_original", "puntos", "points"):
                if poly.get(key) is not None:
                    valor = poly[key]
                    break
            else:
                return None
        try:
            arr = np.asarray(valor, dtype=np.int32)
        except (TypeError, ValueError):
            return None
        if arr.ndim == 2 and arr.shape[1] == 2 and len(arr) >= 3:
            return arr.reshape((-1, 1, 2))
        if arr.ndim == 3 and arr.shape[1] == 1 and arr.shape[2] == 2 and len(arr) >= 3:
            return arr
        return None

    def _resolver_mascara_visual(self, falla, image_shape):
        mask = falla.get("mask")
        if isinstance(mask, np.ndarray) and mask.shape[:2] == image_shape[:2]:
            return (mask > 0).astype(np.uint8) * 255
        mask = falla.get("mascara_parche")
        if isinstance(mask, np.ndarray) and mask.shape[:2] == image_shape[:2]:
            return (mask > 0).astype(np.uint8) * 255
        mask = falla.get("mascara_roi")
        if isinstance(mask, np.ndarray) and mask.shape[:2] == image_shape[:2]:
            return (mask > 0).astype(np.uint8) * 255
        cnt = falla.get("contorno")
        if cnt is not None:
            mask = np.zeros(image_shape[:2], dtype=np.uint8)
            cv2.fillPoly(mask, [cnt], 255)
            return mask
        return None

    def _overlay_mask_on_image(self, imagen, mask, dim_color=(22, 22, 30)):
        if imagen is None:
            return None
        base = self._to_bgr(imagen).copy()
        if mask is not None and mask.shape[:2] == base.shape[:2]:
            mask_bin = (mask > 0).astype(np.uint8)
            overlay = base.copy()
            overlay[mask_bin == 0] = dim_color
            base = overlay
        return base

    def _crear_visual_piel(self, falla, imagen):
        base = self._overlay_mask_on_image(imagen, falla.get("mascara_roi"))
        if base is None:
            return None
        for poly in falla.get("poligonos", []):
            contorno = self._extraer_contorno_poligono(poly)
            if contorno is not None:
                cv2.drawContours(base, [contorno], -1, (255, 255, 0), 2)
        for circ in falla.get("circulos", []):
            centro = tuple(int(v) for v in circ.get("centro", (0, 0)))
            radio = int(circ.get("radio", 0))
            if radio > 0:
                cv2.circle(base, centro, radio, (255, 0, 255), 2)
                cv2.circle(base, centro, 3, (0, 255, 0), -1)
        return base

    def _crear_visual_piel_medicion(self, falla, imagen):
        base = self._crear_visual_piel(falla, imagen)
        if base is None:
            return None
        cx = int(falla.get("ubicacion_x", base.shape[1] // 2))
        cy = int(falla.get("ubicacion_y", base.shape[0] // 2))
        cv2.circle(base, (cx, cy), 5, (255, 255, 255), -1)
        cv2.putText(base, f"Dprom={float(falla.get('diametro_promedio_mm', 0) or 0):.1f} mm", (max(8, cx - 90), max(20, cy - 18)), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (255, 255, 255), 1)
        cv2.putText(base, f"Celdas={int(falla.get('n_celdas', 0) or 0)} A={float(falla.get('area_m2', 0) or 0):.3f} m2", (max(8, cx - 90), max(38, cy + 2)), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (220, 220, 220), 1)
        return base

    def _crear_visual_esqueleto(self, falla, imagen):
        if imagen is None:
            return None
        mask = self._resolver_mascara_visual(falla, imagen.shape)
        base = self._overlay_mask_on_image(imagen, mask)
        if base is None:
            return None
        esq = falla.get("esqueleto")
        if esq is None and mask is not None and SKIMAGE_OK:
            try:
                esq = (skeletonize((mask > 0).astype(bool)) * 255).astype(np.uint8)
            except Exception:
                esq = None
        if esq is None:
            return None
        esq_u8 = np.asarray(esq)
        if esq_u8.ndim == 3:
            esq_u8 = cv2.cvtColor(self._to_bgr(esq_u8), cv2.COLOR_BGR2GRAY)
        base[esq_u8 > 0] = (0, 255, 0)
        punto = falla.get("punto_max")
        if punto is not None:
            cv2.circle(base, tuple(int(v) for v in punto), 4, (255, 255, 255), -1)
        return base

    def _annotate_distance_map(self, falla):
        dist_map = falla.get("distance_map")
        if dist_map is None:
            return None
        max_dist = np.max(dist_map)
        if max_dist <= 0:
            return None
        dist_norm = (dist_map / max_dist * 255).astype(np.uint8)
        dist_color = cv2.applyColorMap(dist_norm, cv2.COLORMAP_JET)
        punto = falla.get("punto_max")
        espesor_px = falla.get("espesor_px", 0)
        angulo = falla.get("angulo_perp")
        if angulo is None and falla.get("angulo") is not None:
            try:
                angulo = math.radians(float(falla.get("angulo"))) + np.pi / 2
            except Exception:
                angulo = None
        if angulo is None:
            angulo = np.pi / 2
        punto_total = falla.get("punto_max_global")
        zona_max = (dist_map >= (0.75 * max_dist)).astype(np.uint8)
        if np.any(zona_max):
            overlay = dist_color.copy()
            overlay[zona_max > 0] = (0, 0, 255)
            dist_color = cv2.addWeighted(dist_color, 0.75, overlay, 0.25, 0)
            cnts, _ = cv2.findContours((zona_max * 255).astype(np.uint8), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            cv2.drawContours(dist_color, cnts, -1, (255, 255, 255), 1)
        if punto_total is not None:
            cv2.circle(dist_color, tuple(int(v) for v in punto_total), 3, (180, 180, 180), -1)
        if punto is not None and espesor_px > 0:
            radio = max(int(espesor_px / 2), 3)
            dx = int(radio * np.cos(angulo))
            dy = int(radio * np.sin(angulo))
            p1 = (punto[0] - dx, punto[1] - dy)
            p2 = (punto[0] + dx, punto[1] + dy)
            cv2.line(dist_color, p1, p2, (255, 255, 255), 2)
            cv2.circle(dist_color, punto, 3, (255, 255, 255), -1)
            tick = 5
            mx = int(tick * np.cos(angulo + np.pi / 2))
            my = int(tick * np.sin(angulo + np.pi / 2))
            cv2.line(dist_color, (p1[0] - mx, p1[1] - my), (p1[0] + mx, p1[1] + my), (255, 255, 255), 2)
            cv2.line(dist_color, (p2[0] - mx, p2[1] - my), (p2[0] + mx, p2[1] + my), (255, 255, 255), 2)
            cv2.putText(dist_color, f"e zona={falla.get('espesor_mm', 0):.1f} mm", (punto[0] + 8, punto[1] - 8), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (255, 255, 255), 1)
            if falla.get("espesor_total_mm", 0):
                cv2.putText(dist_color, f"e total={falla.get('espesor_total_mm', 0):.1f} mm", (punto[0] + 8, punto[1] + 10), cv2.FONT_HERSHEY_SIMPLEX, 0.38, (220, 220, 220), 1)
            cv2.putText(dist_color, "Zona roja >= 75% del max", (8, 18), cv2.FONT_HERSHEY_SIMPLEX, 0.42, (255, 255, 255), 1)
        return dist_color

    def _annotate_peak_zone(self, falla, imagen):
        dist_map = falla.get("distance_map")
        if dist_map is None or imagen is None:
            return None
        max_dist = np.max(dist_map)
        if max_dist <= 0:
            return None
        mask = self._resolver_mascara_visual(falla, imagen.shape)
        base = self._overlay_mask_on_image(imagen, mask)
        if base is None:
            return None
        zona_max = (dist_map >= (0.75 * max_dist)).astype(np.uint8)
        overlay = base.copy()
        overlay[zona_max > 0] = (0, 0, 255)
        base = cv2.addWeighted(base, 0.70, overlay, 0.30, 0)
        cnts, _ = cv2.findContours((zona_max * 255).astype(np.uint8), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cv2.drawContours(base, cnts, -1, (255, 255, 255), 1)
        punto = falla.get("punto_max")
        if punto is not None:
            cv2.circle(base, tuple(int(v) for v in punto), 4, (255, 255, 255), -1)
        cv2.putText(base, "Zona usada para medir el espesor maximo", (8, 18), cv2.FONT_HERSHEY_SIMPLEX, 0.42, (255, 255, 255), 1)
        return base

    def _crear_visual_hueco_medicion(self, falla, imagen):
        if imagen is None:
            return None
        mask = self._resolver_mascara_visual(falla, imagen.shape)
        base = self._overlay_mask_on_image(imagen, mask)
        if base is None:
            return None
        cnt = falla.get("contorno")
        if cnt is not None:
            cv2.drawContours(base, [cnt], -1, (0, 255, 255), 2)
        cx = int(falla.get("ubicacion_x", base.shape[1] // 2))
        cy = int(falla.get("ubicacion_y", base.shape[0] // 2))
        radio = max(int(float(falla.get("diametro_px", 0) or 0) / 2), 3)
        if radio > 0:
            cv2.circle(base, (cx, cy), radio, (255, 255, 255), 2)
            cv2.line(base, (cx - radio, cy), (cx + radio, cy), (255, 255, 255), 2)
            cv2.circle(base, (cx, cy), 4, (255, 255, 255), -1)
        cv2.putText(base, f"D={float(falla.get('diametro_mm', 0) or 0):.1f} mm", (max(8, cx - 80), max(18, cy - radio - 8)), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (255, 255, 255), 1)
        return base

    def _crear_visual_parche_medicion(self, falla, imagen):
        if imagen is None:
            return None
        mask = self._resolver_mascara_visual(falla, imagen.shape)
        base = self._overlay_mask_on_image(imagen, mask)
        if base is None:
            return None
        cnt = falla.get("contorno")
        if cnt is not None:
            cv2.drawContours(base, [cnt], -1, (0, 255, 255), 2)
        fisuras = falla.get("_fisuras_mask", falla.get("esqueleto"))
        if isinstance(fisuras, np.ndarray):
            if fisuras.ndim == 3:
                fisuras = cv2.cvtColor(self._to_bgr(fisuras), cv2.COLOR_BGR2GRAY)
            base[fisuras > 0] = (0, 0, 255)
        cfg_parche = self._normalizar_config_parche(falla.get("config_parche_personalizada", self.motor.config))
        ratio = float(falla.get("ratio_fisuras", 0.0) or 0.0)
        cv2.putText(base, f"Rel={ratio:.3f}", (8, 18), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (255, 255, 255), 1)
        cv2.putText(base, f"S1<={cfg_parche['parche_ratio_leve_max']:.2f} S2<={cfg_parche['parche_ratio_moderado_max']:.2f}", (8, 36), cv2.FONT_HERSHEY_SIMPLEX, 0.38, (220, 220, 220), 1)
        return base

    def _generar_mosaico_pasos_metodo(self, metodo, data, imagen, nombre):
        fallas = data.get("fallas", [])
        if not fallas or imagen is None:
            return None
        fuente = [f for f in fallas if not f.get("excluida", False)]
        if not fuente:
            return None
        items = []
        for falla in fuente:
            tipo = str(falla.get("tipo", "FALLA"))
            prefijo = f"{metodo} | {tipo} #{falla.get('id', '-')}"
            categoria = categorizar_tipo_falla(tipo)
            pasos = falla.get("pasos", {})
            if categoria == "piel":
                orden = [
                    ("Contraste CLAHE", pasos.get("contraste")),
                    ("Frangi", pasos.get("frangi")),
                    ("Suavizado bilateral", pasos.get("suavizado")),
                    ("Umbralizacion", pasos.get("umbralizada")),
                    ("Morfologia", pasos.get("morfologia")),
                    ("Limpieza", pasos.get("limpia")),
                    ("Esqueleto", falla.get("esqueleto") if falla.get("esqueleto") is not None else pasos.get("esqueleto")),
                    ("ROI YOLO", falla.get("mascara_roi")),
                    ("Poligonos y circulos", self._crear_visual_piel(falla, imagen)),
                    ("Medicion y severidad", self._crear_visual_piel_medicion(falla, imagen)),
                ]
                for titulo, img in orden:
                    self._append_step_image(items, f"{prefijo} | {titulo}", img)
            else:
                for clave, img_paso in pasos.items():
                    self._append_step_image(items, f"{prefijo} | {self._formatear_titulo_paso(clave)}", img_paso)
                if categoria == "grieta" or "FISURA" in tipo.upper():
                    self._append_step_image(items, f"{prefijo} | Mascara", falla.get("mask"))
                    self._append_step_image(items, f"{prefijo} | Zona espesor maximo", self._annotate_peak_zone(falla, imagen))
                    self._append_step_image(items, f"{prefijo} | Distancia / espesor", self._annotate_distance_map(falla))
                    self._append_step_image(items, f"{prefijo} | Esqueleto / longitud", self._crear_visual_esqueleto(falla, imagen))
                elif categoria == "hueco" or tipo == "OJO DE PESCADO":
                    self._append_step_image(items, f"{prefijo} | Mascara", falla.get("mask"))
                    self._append_step_image(items, f"{prefijo} | Mascara completa", falla.get("mask_full"))
                    self._append_step_image(items, f"{prefijo} | Medicion diametro", self._crear_visual_hueco_medicion(falla, imagen))
                elif categoria == "parche" or "PARCH" in tipo.upper():
                    mask_vis = None
                    cnt = falla.get("contorno")
                    if cnt is not None:
                        mask_vis = np.zeros(imagen.shape[:2], dtype=np.uint8)
                        cv2.fillPoly(mask_vis, [cnt], 255)
                    self._append_step_image(items, f"{prefijo} | Mascara", mask_vis)
                    self._append_step_image(items, f"{prefijo} | Fisuras internas", falla.get("_fisuras_mask"))
                    self._append_step_image(items, f"{prefijo} | Esqueleto / bordes", falla.get("esqueleto"))
                    self._append_step_image(items, f"{prefijo} | Relacion / severidad", self._crear_visual_parche_medicion(falla, imagen))
        if not items:
            return None
        cell_w = 360
        cell_h = 250
        celdas = []
        for titulo, img in items:
            img_r = cv2.resize(self._to_bgr(img), (cell_w, cell_h))
            cv2.rectangle(img_r, (0, 0), (cell_w, 30), (22, 22, 28), -1)
            cv2.putText(img_r, titulo[:72], (8, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.48, (0, 255, 255), 1)
            celdas.append(img_r)
        n_cols = min(2, len(celdas))
        n_rows = (len(celdas) + n_cols - 1) // n_cols
        while len(celdas) < n_rows * n_cols:
            celdas.append(np.zeros((cell_h, cell_w, 3), dtype=np.uint8))
        filas = []
        for idx in range(n_rows):
            filas.append(np.hstack(celdas[idx * n_cols:(idx + 1) * n_cols]))
        mosaico = np.vstack(filas)
        header = np.zeros((38, mosaico.shape[1], 3), dtype=np.uint8)
        cv2.putText(header, f"PASOS {metodo} | {nombre}", (10, 26), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
        return np.vstack([header, mosaico])

    def _refresh_method_views(self, metodo):
        if not self.resultado_actual:
            return
        imagen = self.resultado_actual["imagen"]
        activas = self._get_fallas_activas(metodo)
        visual = self._dibujar_metodo(metodo, imagen, activas)
        self.resultado_actual["metodos"][metodo]["visual"] = visual
        self._imagenes_render[metodo] = visual
        self.resultado_actual["metodos"][metodo]["pasos_visual"] = None
        self._imagenes_render[f"{metodo}_Pasos"] = None
        self._render_single_view(metodo)

    def _redibujar_todas_las_vistas(self):
        if not self.resultado_actual:
            self._render_single_view("Original")
            return
        self._sync_config()
        self._imagenes_render["Original"] = self.resultado_actual["imagen"]
        self._render_single_view("Original")
        for metodo in self.METODOS:
            self._refresh_method_views(metodo)

    def _render_single_view(self, key):
        if key in self.METODOS:
            if self._modo_vista_por_metodo.get(key, "resultado") == "pasos":
                if self.resultado_actual:
                    cached = self.resultado_actual["metodos"].get(key, {}).get("pasos_visual")
                    if cached is None:
                        cached = self._generar_mosaico_pasos_metodo(
                            key,
                            self.resultado_actual["metodos"][key],
                            self.resultado_actual["imagen"],
                            self.resultado_actual["nombre"],
                        )
                        self.resultado_actual["metodos"][key]["pasos_visual"] = cached
                    self._imagenes_render[f"{key}_Pasos"] = cached
                    self._render_image_on_canvas(key, cached)
                else:
                    self._render_image_on_canvas(key, None)
                return
            self._render_image_on_canvas(key, self._imagenes_render.get(key))
            return
        if key.endswith("_Pasos"):
            metodo = key.replace("_Pasos", "")
            if self.resultado_actual:
                cached = self.resultado_actual["metodos"].get(metodo, {}).get("pasos_visual")
                if cached is None:
                    cached = self._generar_mosaico_pasos_metodo(
                        metodo,
                        self.resultado_actual["metodos"][metodo],
                        self.resultado_actual["imagen"],
                        self.resultado_actual["nombre"],
                    )
                    self.resultado_actual["metodos"][metodo]["pasos_visual"] = cached
                self._imagenes_render[key] = cached
            else:
                self._imagenes_render[key] = None
        self._render_image_on_canvas(key, self._imagenes_render.get(key))

    def _render_image_on_canvas(self, key, img_cv):
        canvas = self._canvas_widgets.get(key)
        if canvas is None:
            return
        state_key = self._state_key_for_canvas(key)
        canvas.delete("all")
        if img_cv is None:
            cw = max(canvas.winfo_width(), 320)
            ch = max(canvas.winfo_height(), 240)
            canvas.create_text(cw // 2, ch // 2, text="Sin imagen", fill=EstiloUI.FG_SECONDARY, font=EstiloUI.FONT_SUBTITLE)
            self._photo_refs.pop(state_key, None)
            self._render_meta.pop(state_key, None)
            try:
                canvas.configure(scrollregion=(0, 0, cw, ch))
            except Exception:
                pass
            return
        img_cv = self._to_bgr(img_cv)
        if img_cv is None:
            return
        h, w = img_cv.shape[:2]
        cw = max(canvas.winfo_width(), 320)
        ch = max(canvas.winfo_height(), 240)
        estado = self._get_view_state(state_key)
        if estado.get("zoom_level") is None or estado.get("image_size") != (w, h):
            self._zoom_fit_canvas(key, render=False)
            estado = self._get_view_state(state_key)
        self._clamp_view_state(key, estado, img_cv)

        zoom_level = float(estado.get("zoom_level") or 1.0)
        pan_offset_x = float(estado.get("pan_offset_x", 0.0))
        pan_offset_y = float(estado.get("pan_offset_y", 0.0))

        w_scaled = int(w * zoom_level)
        h_scaled = int(h * zoom_level)
        x_start_scaled = max(0, -pan_offset_x)
        y_start_scaled = max(0, -pan_offset_y)
        x_end_scaled = min(w_scaled, cw - pan_offset_x)
        y_end_scaled = min(h_scaled, ch - pan_offset_y)

        if x_end_scaled <= x_start_scaled or y_end_scaled <= y_start_scaled:
            canvas.create_text(cw // 2, ch // 2, text="Fuera de vista", fill=EstiloUI.FG_SECONDARY, font=EstiloUI.FONT_SUBTITLE)
            self._photo_refs.pop(state_key, None)
            self._render_meta.pop(state_key, None)
            return

        x1 = max(0, int(x_start_scaled / zoom_level))
        y1 = max(0, int(y_start_scaled / zoom_level))
        x2 = min(w, int(np.ceil(x_end_scaled / zoom_level)))
        y2 = min(h, int(np.ceil(y_end_scaled / zoom_level)))
        if x2 <= x1 or y2 <= y1:
            return

        vista = img_cv[y1:y2, x1:x2]
        vista_w = int((x2 - x1) * zoom_level)
        vista_h = int((y2 - y1) * zoom_level)
        if vista_w < 1 or vista_h < 1:
            return
        interp = cv2.INTER_NEAREST if zoom_level > 2.0 else cv2.INTER_LINEAR
        vista = cv2.resize(vista, (vista_w, vista_h), interpolation=interp)
        draw_x = int(max(0, pan_offset_x))
        draw_y = int(max(0, pan_offset_y))
        rgb = cv2.cvtColor(vista, cv2.COLOR_BGR2RGB)
        photo = ImageTk.PhotoImage(Image.fromarray(rgb))
        canvas.create_image(draw_x, draw_y, anchor="nw", image=photo)
        try:
            canvas.configure(scrollregion=(0, 0, max(cw, w_scaled), max(ch, h_scaled)))
        except Exception:
            pass
        self._photo_refs[state_key] = photo
        self._render_meta[state_key] = {
            "zoom_level": zoom_level,
            "pan_offset_x": pan_offset_x,
            "pan_offset_y": pan_offset_y,
            "width": w,
            "height": h,
            "canvas_width": cw,
            "canvas_height": ch,
        }
        self._update_zoom_ui(key)

    def _mask_hit(self, mask, x, y):
        if mask is None:
            return False
        arr = np.array(mask)
        if arr.ndim != 2:
            return False
        xi = int(round(x))
        yi = int(round(y))
        if yi < 0 or xi < 0 or yi >= arr.shape[0] or xi >= arr.shape[1]:
            return False
        return bool(arr[yi, xi] > 0)

    def _falla_hit(self, falla, x, y):
        contorno = falla.get("contorno")
        if contorno is not None:
            try:
                if cv2.pointPolygonTest(contorno, (float(x), float(y)), False) >= 0:
                    return True
            except Exception:
                pass
        for key in ("mask", "mask_full", "mascara_roi", "mascara_parche"):
            if self._mask_hit(falla.get(key), x, y):
                return True
        return False

    def _pick_falla_at(self, metodo, x, y):
        if not self.resultado_actual:
            return None
        fallas = self.resultado_actual["metodos"][metodo]["fallas"]
        candidatos = []
        for falla in fallas:
            if self._falla_hit(falla, x, y):
                area = falla.get("area_px", 0) or 0
                if area <= 0 and falla.get("contorno") is not None:
                    try:
                        area = cv2.contourArea(falla["contorno"])
                    except Exception:
                        area = 0
                cx = falla.get("ubicacion_x", x)
                cy = falla.get("ubicacion_y", y)
                candidatos.append((area if area > 0 else 10**12, math.hypot(x - cx, y - cy), falla))
        if candidatos:
            candidatos.sort(key=lambda item: (item[0], item[1]))
            return candidatos[0][2]
        mejor = None
        mejor_dist = float("inf")
        for falla in fallas:
            cx = falla.get("ubicacion_x")
            cy = falla.get("ubicacion_y")
            if cx is None or cy is None:
                continue
            dist = math.hypot(x - cx, y - cy)
            if dist < 70 and dist < mejor_dist:
                mejor_dist = dist
                mejor = falla
        return mejor

    def _on_result_canvas_click(self, event, metodo):
        if not self.resultado_actual:
            return
        if self._modo_vista_por_metodo.get(metodo, "resultado") != "resultado":
            return
        meta = self._render_meta.get(self._state_key_for_canvas(metodo))
        if not meta:
            return
        zoom_level = float(meta.get("zoom_level") or 1.0)
        x = (event.x - float(meta.get("pan_offset_x", 0.0))) / zoom_level
        y = (event.y - float(meta.get("pan_offset_y", 0.0))) / zoom_level
        if x < 0 or y < 0 or x >= meta["width"] or y >= meta["height"]:
            return
        falla = self._pick_falla_at(metodo, x, y)
        if falla is None:
            return
        falla["excluida"] = not falla.get("excluida", False)
        estado = "oculta" if falla["excluida"] else "visible"
        self._log(f"{metodo}: click sobre {falla.get('tipo', 'FALLA')} #{falla.get('id', '-')} -> {estado}")
        self._cachear_resultado_actual()
        self._refresh_method_views(metodo)
        self._actualizar_resumenes(self.resultado_actual)

    def _iter_resultados_exportables(self):
        vistos = set()
        for idx in sorted(self._resultados_batch):
            resultado = self._resultados_batch.get(idx)
            if not resultado:
                continue
            nombre = resultado.get("nombre") or f"imagen_{idx}"
            if nombre in vistos:
                continue
            vistos.add(nombre)
            yield idx, resultado
        if self.resultado_actual:
            nombre = self.resultado_actual.get("nombre") or "imagen_actual"
            if nombre not in vistos:
                yield self.imagen_actual_idx, self.resultado_actual

    def _dataset_exportacion_integrado(self):
        dataset = {}
        for _, resultado in self._iter_resultados_exportables():
            self._normalizar_fallas_resultado(resultado)
            dataset[resultado["nombre"]] = {
                "metodos": {
                    metodo: resultado["metodos"].get(metodo, {}).get("fallas", [])
                    for metodo in self.METODOS
                },
                "calibracion_px_mm": resultado.get("calibracion_px_mm", self.motor.calibrador.px_por_mm),
                "angulo_eje": resultado.get("angulo_eje", getattr(self.motor.calibrador, "angulo_eje_via", 90.0)),
                "ancho_via_real_m": resultado.get(
                    "ancho_via_real_m",
                    getattr(self.motor.calibrador, "ancho_via_real_m", self.motor.config.get("ancho_via_real_m", 0.0)),
                ),
            }
        return dataset

    def _anotar_imagen_sin_fallas(self, imagen, metodo):
        if imagen is None:
            return None
        vis = imagen.copy()
        h, w = vis.shape[:2]
        overlay = vis.copy()
        box_h = max(72, int(h * 0.11))
        y1 = max(10, (h - box_h) // 2)
        y2 = min(h - 10, y1 + box_h)
        cv2.rectangle(overlay, (20, y1), (max(40, w - 20), y2), (18, 24, 38), -1)
        vis = cv2.addWeighted(overlay, 0.45, vis, 0.55, 0)
        titulo = f"{metodo}: SIN FALLAS DETECTADAS"
        detalle = "Exportado con valores en 0"
        texto_visible_dos_lineas(
            vis,
            titulo,
            detalle,
            (30, y1 + 28),
            escala_titulo=0.75,
            escala_detalle=0.55,
            color=(0, 255, 255),
            grosor_titulo=2,
            grosor_detalle=1,
        )
        return vis

    def _render_resultado_exportable(self, resultado, metodo):
        imagen = resultado.get("imagen")
        if imagen is None:
            return None
        fallas = resultado.get("metodos", {}).get(metodo, {}).get("fallas", [])
        activas = [f for f in fallas if not f.get("excluida", False)]
        visual = self._dibujar_metodo(metodo, imagen, activas)
        if not activas:
            visual = self._anotar_imagen_sin_fallas(visual, metodo)
        return visual

    def _render_pasos_exportable(self, resultado, metodo):
        if not resultado:
            return None
        data_metodo = resultado.get("metodos", {}).get(metodo, {})
        pasos = data_metodo.get("pasos_visual")
        if pasos is None:
            pasos = self._generar_mosaico_pasos_metodo(
                metodo,
                data_metodo,
                resultado.get("imagen"),
                resultado.get("nombre", metodo),
            )
            data_metodo["pasos_visual"] = pasos
        return pasos

    def _guardar_resultados(self):
        self._cachear_resultado_actual()
        dataset = self._dataset_exportacion_integrado()
        if not dataset:
            messagebox.showinfo("Info", "Procese al menos una imagen primero.")
            return
        carpeta_base = filedialog.askdirectory(
            title="Seleccionar carpeta de salida",
            initialdir=_resolver_directorio_dialogo("resultados", self.motor.config.get("ruta_salida")),
        )
        if not carpeta_base:
            return
        self.motor.config["ruta_salida"] = str(carpeta_base)
        carpeta_base = Path(carpeta_base)
        carpeta_resultados = carpeta_base / "Resultados"
        carpeta_pasos = carpeta_base / "Pasos"
        carpeta_resultados.mkdir(parents=True, exist_ok=True)
        carpeta_pasos.mkdir(parents=True, exist_ok=True)
        carpetas_resultados_metodo = {}
        carpetas_pasos_metodo = {}
        for metodo in self.METODOS:
            carpeta_m_res = carpeta_resultados / metodo
            carpeta_m_pas = carpeta_pasos / metodo
            carpeta_m_res.mkdir(parents=True, exist_ok=True)
            carpeta_m_pas.mkdir(parents=True, exist_ok=True)
            carpetas_resultados_metodo[metodo] = carpeta_m_res
            carpetas_pasos_metodo[metodo] = carpeta_m_pas

        guardados = []
        for _, resultado in self._iter_resultados_exportables():
            nombre_base = Path(resultado["nombre"]).stem
            for metodo in self.METODOS:
                visual = self._render_resultado_exportable(resultado, metodo)
                if visual is None:
                    continue
                ruta = carpetas_resultados_metodo[metodo] / f"{nombre_base}_{metodo}.png"
                if cv2.imwrite(str(ruta), self._to_bgr(visual)):
                    guardados.append(str(ruta.relative_to(carpeta_base)))
                pasos = self._render_pasos_exportable(resultado, metodo)
                if pasos is not None:
                    ruta_pasos = carpetas_pasos_metodo[metodo] / f"{nombre_base}_{metodo}_Pasos.png"
                    if cv2.imwrite(str(ruta_pasos), self._to_bgr(pasos)):
                        guardados.append(str(ruta_pasos.relative_to(carpeta_base)))

        excel_path = ExportadorExcelIntegrado.exportar(dataset, carpeta_base, self.motor.calibrador)
        if excel_path:
            guardados.append(Path(excel_path).name)
            self._log(f"Excel exportado: {excel_path}")
        elif not PANDAS_OK:
            self._log("Excel no generado: pandas no disponible.")
        else:
            self._log("Excel no generado: no hay fallas activas palomeadas para exportar.")

        if guardados:
            self._log("Guardado: " + ", ".join(guardados))
        self._log(f"Carpeta Resultados: {carpeta_resultados}")
        self._log(f"Carpeta Pasos: {carpeta_pasos}")
        for metodo in self.METODOS:
            self._log(f"  {metodo} Resultados: {carpetas_resultados_metodo[metodo]}")
            self._log(f"  {metodo} Pasos: {carpetas_pasos_metodo[metodo]}")
        self._estado(f"Resultados guardados en {carpeta_base}")
        msg = f"Guardado de resultados concluido.\nCarpeta base: {carpeta_base}"
        self._log(msg)
        messagebox.showinfo("Guardar Resultados", msg)

    def _exportar_solo_excel(self):
        self._cachear_resultado_actual()
        dataset = self._dataset_exportacion_integrado()
        if not dataset:
            messagebox.showinfo("Info", "Procese al menos una imagen primero.")
            return
        carpeta = filedialog.askdirectory(
            title="Seleccionar carpeta para exportar Excel",
            initialdir=_resolver_directorio_dialogo("resultados", self.motor.config.get("ruta_salida")),
        )
        if not carpeta:
            return
        self.motor.config["ruta_salida"] = str(carpeta)
        excel_path = ExportadorExcelIntegrado.exportar(dataset, carpeta, self.motor.calibrador)
        if excel_path:
            self._log(f"Excel exportado: {excel_path}")
            self._estado(f"Excel exportado en {carpeta}")
            msg = f"Exportacion de Excel concluida.\nArchivo: {excel_path}"
            self._log(msg)
            messagebox.showinfo("Exportar Solo Excel", msg)
            return
        if not PANDAS_OK:
            messagebox.showwarning("Excel", "No se pudo exportar porque pandas/openpyxl no estan disponibles.")
            self._log("Excel no generado: pandas/openpyxl no disponible.")
        else:
            messagebox.showinfo("Excel", "No hay fallas palomeadas para exportar.")
            self._log("Excel no generado: no hay fallas activas palomeadas.")


AplicacionIntegradaTresMetodos = AplicacionIntegradaTresMetodosAvanzada


# =============================================================================
# EJECUCION
# =============================================================================

def main():
    app = AplicacionIntegradaTresMetodos()
    app.mainloop()


if __name__ == "__main__":
    main()

