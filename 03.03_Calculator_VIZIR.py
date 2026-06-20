#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════╗
║        SISTEMA DE ANÁLISIS VIZIR - PAVIMENTOS FLEXIBLES            ║
║  Evaluación superficial según norma INVIAS / ASTM D6433            ║
║                                                                      ║
║  Autores: Bach. Miguel Bernardino Quispe Arias                       ║
║           Bach. Briza Edith Catachura Aycaya                         ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
import os
import math
import re
import unicodedata
from collections import defaultdict

import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

# ══════════════════════════════════════════════
#  PALETA DE COLORES - TEMA CLARO TIPO MAC
# ══════════════════════════════════════════════
C = {
    "bg_main":      "#eceef2",
    "bg_panel":     "#dfe3e9",
    "bg_card":      "#f7f8fa",
    "bg_input":     "#ffffff",
    "bg_header":    "#f5f5f7",
    "bg_row_even":  "#fafbfc",
    "bg_row_odd":   "#f2f4f7",
    "bg_row_type_a":"#edf4ff",
    "bg_row_type_b":"#f3f6fa",
    "accent":       "#556070",
    "accent2":      "#0a84ff",
    "accent_alt":   "#8d96a3",
    "accent_dim":   "#d3d8e0",
    "gold":         "#d6a23a",
    "green":        "#34c759",
    "green_bg":     "#e8f7ec",
    "orange":       "#ff9f0a",
    "orange_bg":    "#fff3df",
    "red":          "#ff5f57",
    "red_bg":       "#ffe9e7",
    "txt_primary":  "#1d1d1f",
    "txt_second":   "#4b5563",
    "txt_dim":      "#8a8f98",
    "border":       "#d5dae2",
    "border_hl":    "#c5ccd6",
    "white":        "#ffffff",
}

# ══════════════════════════════════════════════
#  CONSTANTES Y TABLAS DEL MÉTODO VIZIR
# ══════════════════════════════════════════════

FALLAS_TIPO_A = {
    "AH":  {"nombre": "Ahuellamiento", "unidad": "m²"},
    "DL":  {"nombre": "Depresiones o hundimientos longitudinales", "unidad": "m²"},
    "DT":  {"nombre": "Depresiones o hundimientos transversales", "unidad": "m²"},
    "FLT": {"nombre": "Fisuras longitudinales por fatiga", "unidad": "m"},
    "FPC": {"nombre": "Fisuras piel de cocodrilo", "unidad": "m²"},
    "B":   {"nombre": "Baches y parches", "unidad": "m²"},
}

FALLAS_TIPO_B = {
    "FIJ": {"nombre": "Fisura longitudinal de junta de construcción", "unidad": "m"},
    "FTJ": {"nombre": "Fisura transversal de junta de construcción", "unidad": "m"},
    "FCT": {"nombre": "Fisuras de contracción térmica", "unidad": "m"},
    "FP":  {"nombre": "Fisuras parabólicas", "unidad": "m"},
    "FB":  {"nombre": "Fisura de borde", "unidad": "m"},
    "O":   {"nombre": "Ojos de pescado", "unidad": "unidad (1)"},
    "DM":  {"nombre": "Desplazamiento o abultamiento", "unidad": "m"},
    "PL":  {"nombre": "Pérdida de película ligante", "unidad": "m²"},
    "PA":  {"nombre": "Pérdida de agregados", "unidad": "m²"},
    "D":   {"nombre": "Descascaramiento", "unidad": "m²"},
    "PU":  {"nombre": "Pulimiento de agregados", "unidad": "m"},
    "EX":  {"nombre": "Exudación", "unidad": "m²"},
    "AM":  {"nombre": "Afloramiento de mortero", "unidad": "m²"},
    "AA":  {"nombre": "Afloramiento de agua", "unidad": "m²"},
    "DB":  {"nombre": "Desintegración de bordes del pavimento", "unidad": "m²"},
    "ECB": {"nombre": "Escalonamiento entre calzada y berma", "unidad": "m"},
    "EB":  {"nombre": "Erosión de las bermas", "unidad": "m"},
    "S":   {"nombre": "Segregación", "unidad": "m²"},
}

TABLA_IF = {1: {1: 1, 2: 2, 3: 3}, 2: {1: 2, 2: 3, 3: 4}, 3: {1: 3, 2: 4, 3: 5}}
TABLA_ID = {1: {1: 1, 2: 2, 3: 3}, 2: {1: 2, 2: 3, 3: 4}, 3: {1: 3, 2: 4, 3: 5}}

TABLA_IS_PRIMER = {
    0: {0: 1, 1: 2, 2: 2, 3: 3, 4: 4, 5: 4},
    1: {0: 3, 1: 3, 2: 3, 3: 4, 4: 5, 5: 5},
    2: {0: 3, 1: 3, 2: 3, 3: 4, 4: 5, 5: 5},
    3: {0: 4, 1: 5, 2: 5, 3: 5, 4: 6, 5: 6},
    4: {0: 5, 1: 6, 2: 6, 3: 7, 4: 7, 5: 7},
    5: {0: 5, 1: 6, 2: 6, 3: 7, 4: 7, 5: 7},
}

TABLA_CORRECCION = {1: {1: 0, 2: 0, 3: 0}, 2: {1: 0, 2: 0, 3: 1}, 3: {1: 0, 2: 1, 3: 1}}


def clasificar_is(v):
    if v <= 2:
        return "Bueno", C["green"], C["green_bg"]
    elif v <= 4:
        return "Marginal", C["orange"], C["orange_bg"]
    else:
        return "Deficiente", C["red"], C["red_bg"]


def ext_col(pct):
    e = pct * 100 if pct <= 1 else pct
    if e <= 10:
        return 1
    elif e <= 50:
        return 2
    return 3


def ext_label(pct):
    e = pct * 100 if pct <= 1 else pct
    if e <= 10:
        return "0 - 10 %"
    elif e <= 50:
        return "10 - 50 %"
    return "> 50 %"


def gravedad_ponderada(fallas):
    """Calcula la gravedad representativa ponderando por la suma de cantidades.

    G = (s1*1 + s2*2 + s3*3) / (s1 + s2 + s3)

    donde s1, s2 y s3 son las sumas de la cantidad/total de fallas
    con gravedad 1, 2 y 3 respectivamente.

    Criterios:
        G < 1.5 -> 1
        1.5 <= G < 2.5 -> 2
        G >= 2.5 -> 3

    Retorna: (g_representativa, g_promedio, s1, s2, s3)
    """
    s1 = sum(float(f.get("total", 0.0) or 0.0) for f in fallas if f.get("gravedad") == 1)
    s2 = sum(float(f.get("total", 0.0) or 0.0) for f in fallas if f.get("gravedad") == 2)
    s3 = sum(float(f.get("total", 0.0) or 0.0) for f in fallas if f.get("gravedad") == 3)
    total = s1 + s2 + s3
    if total == 0:
        return 0, 0.0, 0, 0, 0

    g_promedio = (s1 * 1 + s2 * 2 + s3 * 3) / total
    if g_promedio < 1.5:
        g_rep = 1
    elif g_promedio < 2.5:
        g_rep = 2
    else:
        g_rep = 3
    return g_rep, g_promedio, s1, s2, s3


def normalizar_unidad_medida(unidad):
    return str(unidad or "").strip().lower().replace("Â²", "2").replace("²", "2")


def es_unidad_contable(unidad):
    return bool(re.search(r"\b(?:und|unidad|unidades)\b", normalizar_unidad_medida(unidad)))


def unidad_canonica_vizir(codigo, unidad):
    if str(codigo or "").strip().upper() == "O":
        return "unidad (1)"
    return str(unidad or "").strip()


def fmt_total(total, unidad):
    """Formatea el total: entero si la unidad es de conteo, 2 decimales en otro caso."""
    if es_unidad_contable(unidad):
        return int(total)
    return round(total, 2)


def fmt_total_str(total, unidad):
    """Formatea el total como string para mostrar en interfaz."""
    if es_unidad_contable(unidad):
        return str(int(total))
    return f"{total:.2f}"


# ══════════════════════════════════════════════
#  WIDGETS HELPER
# ══════════════════════════════════════════════

PROG_RANGO_RE = re.compile(r"(\d+)\+(\d{1,3})_(\d+)\+(\d{1,3})", re.IGNORECASE)
PROG_PUNTO_RE = re.compile(r"(\d+)\+(\d{1,3})", re.IGNORECASE)

_SIN_FALLAS_KEYWORDS = ("SIN FALLA", "SIN DETERIORO", "SIN DAÑO", "SIN DANO",
                         "NO PRESENTA", "SIN PATOLOGIA")


def es_sin_fallas(nombre):
    """Detecta si el nombre de falla indica ausencia de deterioros."""
    if not nombre:
        return False
    n = nombre.strip().upper()
    return any(kw in n for kw in _SIN_FALLAS_KEYWORDS)


def prog_a_metros(valor):
    if valor is None:
        return None
    s = str(valor).strip().upper()
    if not s:
        return None
    m = PROG_PUNTO_RE.search(s)
    if m:
        return int(m.group(1)) * 1000 + int(m.group(2))
    s2 = "".join(ch for ch in s if ch.isdigit())
    if not s2:
        return None
    return int(s2)


def metros_a_prog(metros):
    if metros is None:
        return ""
    metros = int(round(float(metros)))
    km, m = divmod(metros, 1000)
    return f"{km:02d}+{m:03d}"


def extraer_rango_progresiva(texto):
    if texto is None:
        return None
    m = PROG_RANGO_RE.search(str(texto))
    if not m:
        return None
    ini = int(m.group(1)) * 1000 + int(m.group(2))
    fin = int(m.group(3)) * 1000 + int(m.group(4))
    if fin < ini:
        ini, fin = fin, ini
    return ini, fin


def inferir_texto_progresiva(entry):
    if not isinstance(entry, dict):
        return ""
    preferidos = ("archivo_progresiva", "observaciones", "falla")
    revisados = set()
    for key in preferidos:
        txt = entry.get(key)
        if txt is None:
            continue
        txt = str(txt).strip()
        if not txt:
            continue
        revisados.add(txt)
        if extraer_rango_progresiva(txt):
            return txt
    for key, value in entry.items():
        if key in ("abscisa_ini", "abscisa_fin", "prog_tramo", "prog_ini_m", "prog_fin_m"):
            continue
        if value is None:
            continue
        txt = str(value).strip()
        if not txt or txt in revisados:
            continue
        if extraer_rango_progresiva(txt):
            return txt
    return ""


def normalizar_cabecera_vizir(texto):
    s = str(texto or "").strip()
    if ":" in s:
        s = s.split(":", 1)[1]
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower().replace("²", "2").replace("%", " pct ")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return " ".join(s.split())


def puntuar_cabecera_vizir(field_label, header):
    alias = {
        "Nombre": "Nombre de Falla",
        "Tipo": "Tipo (A / B)",
        "Área": "Área (m²)",
        "Area": "Área (m²)",
        "Longitud": "Longitud (m)",
    }
    campo = alias.get(field_label, field_label)
    h = normalizar_cabecera_vizir(header)
    score = 0

    if campo == "Archivo / Imagen":
        if h == "nombre imagen":
            return 160
        if "nombre imagen" in h:
            score += 150
        if "imagen" in h:
            score += 120
        if "archivo" in h or "ruta" in h or "path" in h:
            score += 100
        if "tif" in h:
            score += 80
        if "falla" in h:
            score -= 40
    elif campo == "Nombre de Falla":
        if h == "tipo falla":
            return 160
        if "tipo falla" in h or "nombre falla" in h:
            score += 140
        if "falla" in h or "deterioro" in h:
            score += 100
        if "imagen" in h:
            score -= 80
    elif campo == "Gravedad":
        if h == "severidad original":
            return 160
        if "severidad original" in h:
            score += 150
        if h == "gravedad":
            score += 140
        if "gravedad" in h or "severidad" in h:
            score += 110
        if "ui" in h:
            score -= 30
    elif campo == "Unidad":
        if h == "unidad":
            return 160
        if re.search(r"\b(?:cantidad|numero|unidades)\b", h) or re.search(r"\bunidad\b\s*1\b", h):
            score -= 120
        if re.search(r"\bunidad\b", h):
            score += 120
    elif campo == "Longitud (m)":
        if h == "longitud m":
            return 160
        if "longitud" in h and "m" in h:
            score += 140
        if "longitud" in h:
            score += 70
        if "px" in h:
            score -= 90
    elif campo == "Área (m²)":
        if h == "area m2":
            return 160
        if "area" in h and "m2" in h:
            score += 140
        if "area" in h:
            score += 70
        if "px" in h:
            score -= 90
    elif campo == "Unidades (und)":
        if h == "unidades":
            return 150
        if h in ("unidad 1", "unidades 1"):
            return 160
        if re.search(r"\bunidad\b\s*1\b", h):
            score += 150
        if "unidades" in h or "cantidad" in h or "numero" in h:
            score += 120
    elif campo == "Extensión (%)":
        if h == "extension pct":
            return 150
        if "extension" in h:
            score += 120
    elif campo == "Tipo (A / B)":
        if h == "tipo":
            return 130
        if "tipo" in h and "falla" not in h:
            score += 90
    elif campo == "Código":
        if h == "codigo":
            return 150
        if "codigo" in h:
            score += 120
        if h == "id":
            score -= 80
    elif campo == "Abscisa Inicial":
        if "abscisa ini" in h or "progresiva ini" in h or "pk ini" in h:
            score += 130
    elif campo == "Abscisa Final":
        if "abscisa fin" in h or "progresiva fin" in h or "pk fin" in h:
            score += 130
    elif campo == "Observaciones":
        if "observ" in h or "coment" in h or "nota" in h:
            score += 120

    return score


def autodetect_header_vizir(headers, field_label):
    best_header = None
    best_score = 0
    for header in headers or []:
        score = puntuar_cabecera_vizir(field_label, header)
        if score > best_score:
            best_header = header
            best_score = score
    return best_header, best_score


def completar_progresiva_entry(entry):
    item = dict(entry)
    ini = prog_a_metros(item.get("abscisa_ini"))
    fin = prog_a_metros(item.get("abscisa_fin"))
    texto_prog = item.get("archivo_progresiva") or inferir_texto_progresiva(item)
    if texto_prog and not item.get("archivo_progresiva"):
        item["archivo_progresiva"] = str(texto_prog).strip()
    if (ini is None or fin is None) and texto_prog:
        rg = extraer_rango_progresiva(texto_prog)
        if rg:
            if ini is None:
                ini = rg[0]
            if fin is None:
                fin = rg[1]
    if ini is not None and fin is not None and fin < ini:
        ini, fin = fin, ini
    if not item.get("abscisa_ini") and ini is not None:
        item["abscisa_ini"] = metros_a_prog(ini)
    if not item.get("abscisa_fin") and fin is not None:
        item["abscisa_fin"] = metros_a_prog(fin)
    item["prog_ini_m"] = ini
    item["prog_fin_m"] = fin
    item["prog_tramo"] = (
        f"{metros_a_prog(ini)} - {metros_a_prog(fin)}"
        if ini is not None and fin is not None else ""
    )
    return item


class AnalizadorProgresivasVIZIR:
    @staticmethod
    def preparar_registros(registros):
        return [completar_progresiva_entry(r) for r in registros]

    @classmethod
    def filtrar_registros(cls, registros, prog_ini=None, prog_fin=None):
        base = cls.preparar_registros(registros)
        filtrados = []
        for item in base:
            ini = item.get("prog_ini_m")
            fin = item.get("prog_fin_m")
            if ini is None or fin is None:
                continue
            if prog_ini is not None and ini < prog_ini:
                continue
            if prog_fin is not None and fin > prog_fin:
                continue
            filtrados.append(item)
        return filtrados

    @classmethod
    def bloques_disponibles(cls, registros, tam_bloque, prog_ini=None, prog_fin=None):
        if tam_bloque <= 0:
            raise ValueError("El tamano del bloque debe ser mayor que cero.")
        base = cls.filtrar_registros(registros, prog_ini, prog_fin)
        if not base:
            return [], base
        inicio = prog_ini if prog_ini is not None else min(r["prog_ini_m"] for r in base)
        grupos = defaultdict(list)
        for item in base:
            bini = inicio + (int(item["prog_ini_m"] - inicio) // tam_bloque) * tam_bloque
            if item["prog_fin_m"] <= bini + tam_bloque:
                grupos[bini].append(item)
        bloques = []
        for bini in sorted(grupos):
            bfin = bini + tam_bloque
            blk = grupos[bini]
            if not blk:
                continue
            bloques.append({
                "ini": bini,
                "fin": bfin,
                "n_reg": len(blk),
                "n_tramos": len({r.get("prog_tramo", "") for r in blk if r.get("prog_tramo")}),
                "label": f"{metros_a_prog(bini)} - {metros_a_prog(bfin)}  |  {len(blk)} reg",
            })
        return bloques, base

    @classmethod
    def tramos_disponibles(cls, registros, prog_ini=None, prog_fin=None):
        base = cls.filtrar_registros(registros, prog_ini, prog_fin)
        if not base:
            return [], base
        tramos_map = defaultdict(list)
        for item in base:
            key = (item.get("prog_ini_m"), item.get("prog_fin_m"))
            if key[0] is None or key[1] is None:
                continue
            tramos_map[key].append(item)
        tramos = []
        for (ini, fin) in sorted(tramos_map):
            tramos.append({
                "ini": ini,
                "fin": fin,
                "n_reg": len(tramos_map[(ini, fin)]),
                "label": f"{metros_a_prog(ini)} - {metros_a_prog(fin)}",
            })
        return tramos, base


def make_label(parent, text, size=10, bold=False, color=None, anchor="w"):
    f = ("Segoe UI", size, "bold") if bold else ("Segoe UI", size)
    lbl = tk.Label(parent, text=text, font=f,
                   fg=color or C["txt_primary"], bg=parent["bg"],
                   anchor=anchor)
    return lbl


def make_entry(parent, textvariable=None, width=25, **kw):
    e = tk.Entry(parent, textvariable=textvariable, width=width,
                 font=("Segoe UI", 10),
                 bg=C["bg_input"], fg=C["txt_primary"],
                 disabledbackground=C["bg_panel"],
                 disabledforeground=C["txt_dim"],
                 insertbackground=C["accent2"],
                 selectbackground=C["accent2"],
                 selectforeground=C["white"],
                 relief="flat", bd=0, highlightthickness=1,
                 highlightbackground=C["border"],
                 highlightcolor=C["accent2"], **kw)
    return e


def make_button(parent, text, command, style="normal", compact=False, state="normal"):
    colors = {
        "normal":  (C["bg_panel"],   C["txt_primary"], C["border_hl"]),
        "gold":    (C["gold"],       C["txt_primary"], "#e2b14c"),
        "green":   (C["green"],      C["white"],       "#28b64a"),
        "danger":  (C["red"],        C["white"],       "#ff453a"),
        "dim":     (C["accent_dim"], C["txt_primary"], C["border_hl"]),
        "warning": (C["orange"],     C["txt_primary"], "#ffb340"),
    }
    bg, fg, hover_bg = colors.get(style, colors["normal"])
    font_size = 9 if compact else 10
    padx = 10 if compact else 16
    pady = 3 if compact else 6
    btn = tk.Button(parent, text=text, command=command,
                    font=("Segoe UI", font_size, "bold"),
                    bg=bg, fg=fg, activebackground=hover_bg,
                    activeforeground=fg, disabledforeground=C["txt_dim"],
                    relief="flat", padx=padx, pady=pady, cursor="hand2",
                    borderwidth=0, bd=0, state=state)

    def _on_enter(_event):
        if str(btn.cget("state")) != "disabled":
            btn.config(bg=hover_bg)

    def _on_leave(_event):
        if str(btn.cget("state")) != "disabled":
            btn.config(bg=bg)

    btn.bind("<Enter>", _on_enter)
    btn.bind("<Leave>", _on_leave)
    return btn


def make_combo(parent, textvariable, values, width=28):
    cb = ttk.Combobox(parent, textvariable=textvariable,
                      values=values, state="readonly",
                      width=width, font=("Segoe UI", 10),
                      style="Mac.TCombobox")
    return cb


def make_card(parent, title=""):
    card = tk.Frame(parent, bg=C["bg_card"],
                    bd=0,
                    highlightbackground=C["border"],
                    highlightcolor=C["border"],
                    highlightthickness=1)
    if title:
        hdr = tk.Frame(card, bg=C["bg_header"], height=36, bd=0)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text=f"  {title}", font=("Segoe UI", 11, "bold"),
                 fg=C["txt_primary"], bg=C["bg_header"],
                 anchor="w").pack(fill="x", padx=8, pady=6)
    body = tk.Frame(card, bg=C["bg_card"], bd=0)
    body.pack(fill="both", expand=True, padx=12, pady=10)
    return card, body


def make_separator(parent):
    return tk.Frame(parent, bg=C["border"], height=1)


# ══════════════════════════════════════════════
#  SCROLLABLE FRAME
# ══════════════════════════════════════════════

class PanelProgresivasVIZIR(tk.Frame):
    MODOS = {
        "Todo el archivo": "todo",
        "Rango manual": "manual",
        "Tramos detectados": "tramos",
        "Bloques automaticos": "bloques",
    }

    def __init__(self, parent, apply_cb):
        super().__init__(parent, bg=C["bg_card"])
        self.apply_cb = apply_cb
        self.registros = []
        self.tramos = []
        self.bloques = []
        self._build()

    def _build(self):
        row1 = tk.Frame(self, bg=C["bg_card"])
        row1.pack(fill="x", pady=(0, 6))
        tk.Label(row1, text="Modo:", font=("Segoe UI", 10),
                 fg=C["txt_second"], bg=C["bg_card"]).pack(side="left")
        self.v_modo = tk.StringVar(value="Todo el archivo")
        self.cmb_modo = make_combo(row1, self.v_modo, list(self.MODOS.keys()), width=20)
        self.cmb_modo.pack(side="left", padx=(6, 18))

        tk.Label(row1, text="Desde:", font=("Segoe UI", 10),
                 fg=C["txt_second"], bg=C["bg_card"]).pack(side="left")
        self.v_ini = tk.StringVar()
        make_entry(row1, textvariable=self.v_ini, width=12).pack(side="left", padx=(6, 12))

        tk.Label(row1, text="Hasta:", font=("Segoe UI", 10),
                 fg=C["txt_second"], bg=C["bg_card"]).pack(side="left")
        self.v_fin = tk.StringVar()
        make_entry(row1, textvariable=self.v_fin, width=12).pack(side="left", padx=(6, 12))

        tk.Label(row1, text="Bloque (m):", font=("Segoe UI", 10),
                 fg=C["txt_second"], bg=C["bg_card"]).pack(side="left")
        self.v_bloque = tk.StringVar(value="100")
        make_entry(row1, textvariable=self.v_bloque, width=10).pack(side="left", padx=(6, 0))

        row2 = tk.Frame(self, bg=C["bg_card"])
        row2.pack(fill="x")
        make_button(row2, "Detectar tramos / bloques", self._detectar_bloques, style="dim").pack(side="left")
        self.v_bloque_sel = tk.StringVar()
        self.cmb_bloque = make_combo(row2, self.v_bloque_sel, [], width=38)
        self.cmb_bloque.pack(side="left", padx=(10, 10))
        make_button(row2, "Aplicar filtro", self.apply_cb, style="green").pack(side="left")

        row3 = tk.Frame(self, bg=C["bg_card"])
        row3.pack(fill="x", pady=(8, 0))
        tk.Label(row3, text="Tramo inicial:", font=("Segoe UI", 10),
                 fg=C["txt_second"], bg=C["bg_card"]).pack(side="left")
        self.v_tramo_ini = tk.StringVar()
        self.cmb_tramo_ini = make_combo(row3, self.v_tramo_ini, [], width=22)
        self.cmb_tramo_ini.pack(side="left", padx=(6, 12))

        tk.Label(row3, text="Tramo final:", font=("Segoe UI", 10),
                 fg=C["txt_second"], bg=C["bg_card"]).pack(side="left")
        self.v_tramo_fin = tk.StringVar()
        self.cmb_tramo_fin = make_combo(row3, self.v_tramo_fin, [], width=22)
        self.cmb_tramo_fin.pack(side="left", padx=(6, 0))

        self.lbl_info = tk.Label(self, text="Sin filtro de progresiva.",
                                 font=("Segoe UI", 9), fg=C["txt_dim"],
                                 bg=C["bg_card"])
        self.lbl_info.pack(fill="x", pady=(8, 0))

    def resetear(self):
        self.registros = []
        self.tramos = []
        self.bloques = []
        self.v_modo.set("Todo el archivo")
        self.v_ini.set("")
        self.v_fin.set("")
        self.v_bloque.set("100")
        self.v_tramo_ini.set("")
        self.v_tramo_fin.set("")
        self.v_bloque_sel.set("")
        self.cmb_tramo_ini["values"] = []
        self.cmb_tramo_fin["values"] = []
        self.cmb_bloque["values"] = []
        self.lbl_info.config(text="Sin filtro de progresiva.", fg=C["txt_dim"])

    def configurar(self, registros):
        self.registros = list(registros or [])
        self._detectar_bloques(silent=True)

    def _parse_limites(self):
        ini = prog_a_metros(self.v_ini.get().strip()) if self.v_ini.get().strip() else None
        fin = prog_a_metros(self.v_fin.get().strip()) if self.v_fin.get().strip() else None
        if ini is not None and fin is not None and fin <= ini:
            raise ValueError("La progresiva final debe ser mayor que la inicial.")
        return ini, fin

    def _parse_bloque(self):
        txt = self.v_bloque.get().strip() or "100"
        tam = int(float(txt.replace(",", ".")))
        if tam <= 0:
            raise ValueError("El bloque debe ser mayor que cero.")
        return tam

    def _detectar_bloques(self, silent=False):
        self.tramos = []
        self.bloques = []
        self.v_tramo_ini.set("")
        self.v_tramo_fin.set("")
        self.v_bloque_sel.set("")
        self.cmb_tramo_ini["values"] = []
        self.cmb_tramo_fin["values"] = []
        self.cmb_bloque["values"] = []
        if not self.registros:
            self.lbl_info.config(text="Importe datos para detectar progresivas.", fg=C["txt_dim"])
            return
        try:
            ini, fin = self._parse_limites()
            tam = self._parse_bloque()
            tramos, _ = AnalizadorProgresivasVIZIR.tramos_disponibles(self.registros, ini, fin)
            bloques, base = AnalizadorProgresivasVIZIR.bloques_disponibles(self.registros, tam, ini, fin)
        except Exception as e:
            if not silent:
                messagebox.showerror("Progresivas", str(e), parent=self.winfo_toplevel())
            self.lbl_info.config(text=str(e), fg=C["red"])
            return
        if not base:
            preparados = AnalizadorProgresivasVIZIR.preparar_registros(self.registros)
            con_archivo = sum(1 for r in preparados if str(r.get("archivo_progresiva", "")).strip())
            con_rango = sum(1 for r in preparados if r.get("prog_ini_m") is not None and r.get("prog_fin_m") is not None)
            self.lbl_info.config(
                text=(
                    "No se detectaron progresivas validas. "
                    f"Registros: {len(preparados)}  |  Con texto de archivo: {con_archivo}  |  "
                    f"Con rango detectado: {con_rango}"
                ),
                fg=C["orange"],
            )
            return
        self.tramos = tramos
        tramo_labels = [t["label"] for t in tramos]
        self.cmb_tramo_ini["values"] = tramo_labels
        self.cmb_tramo_fin["values"] = tramo_labels
        if tramo_labels:
            self.v_tramo_ini.set(tramo_labels[0])
            self.v_tramo_fin.set(tramo_labels[-1])
        self.bloques = bloques
        labels = [b["label"] for b in bloques]
        self.cmb_bloque["values"] = labels
        if labels:
            self.v_bloque_sel.set(labels[0])
        n_tramos = len(tramo_labels)
        self.lbl_info.config(
            text=f"Tramos detectados: {n_tramos}  |  Bloques: {len(labels)}",
            fg=C["accent"],
        )

    def _buscar_tramo(self, label):
        if not label:
            return None
        for item in self.tramos:
            if item["label"] == label:
                return item
        return None

    def obtener_config(self):
        ini, fin = self._parse_limites()
        tam = self._parse_bloque()
        modo = self.MODOS.get(self.v_modo.get(), "todo")
        bloque = None
        if self.v_bloque_sel.get():
            for item in self.bloques:
                if item["label"] == self.v_bloque_sel.get():
                    bloque = item
                    break
        return {
            "modo": modo,
            "ini": ini,
            "fin": fin,
            "tam_bloque": tam,
            "bloque": bloque,
            "tramo_ini": self._buscar_tramo(self.v_tramo_ini.get()),
            "tramo_fin": self._buscar_tramo(self.v_tramo_fin.get()),
        }


class ScrollFrame(tk.Frame):
    def __init__(self, parent, bg=None):
        super().__init__(parent, bg=bg or C["bg_main"])
        self.canvas = tk.Canvas(self, bg=bg or C["bg_main"],
                                highlightthickness=0, borderwidth=0)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical",
                                        command=self.canvas.yview)
        self.inner = tk.Frame(self.canvas, bg=bg or C["bg_main"])

        self.inner.bind("<Configure>",
                        lambda e: self.canvas.configure(
                            scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0, 0), window=self.inner,
                                   anchor="nw", tags="inner")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.inner.bind("<Enter>", self._bind_wheel)
        self.inner.bind("<Leave>", self._unbind_wheel)

    def _on_canvas_configure(self, event):
        self.canvas.itemconfig("inner", width=event.width)

    def _bind_wheel(self, event):
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind_all("<Button-4>", self._on_mousewheel_linux)
        self.canvas.bind_all("<Button-5>", self._on_mousewheel_linux)

    def _unbind_wheel(self, event):
        self.canvas.unbind_all("<MouseWheel>")
        self.canvas.unbind_all("<Button-4>")
        self.canvas.unbind_all("<Button-5>")

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _on_mousewheel_linux(self, event):
        if event.num == 4:
            self.canvas.yview_scroll(-1, "units")
        elif event.num == 5:
            self.canvas.yview_scroll(1, "units")


# ══════════════════════════════════════════════
#  TREEVIEW ESTILO CLARO
# ══════════════════════════════════════════════

def configure_dark_treeview(tree, columns, widths=None):
    style_name = f"Dark{id(tree)}.Treeview"
    s = ttk.Style()
    s.configure(style_name,
                background=C["bg_input"],
                foreground=C["txt_primary"],
                fieldbackground=C["bg_input"],
                font=("Segoe UI", 9),
                rowheight=26,
                borderwidth=0,
                relief="flat")
    s.configure(f"{style_name}.Heading",
                background=C["bg_header"],
                foreground=C["txt_second"],
                font=("Segoe UI", 9, "bold"),
                borderwidth=0,
                relief="flat")
    s.map(style_name,
          background=[("selected", C["accent2"])],
          foreground=[("selected", C["white"])])
    s.map(f"{style_name}.Heading",
          background=[("active", C["bg_header"])],
          foreground=[("active", C["txt_primary"])])

    tree.configure(style=style_name)
    tree["columns"] = columns
    tree["show"] = "headings"
    for i, col in enumerate(columns):
        w = widths[i] if widths and i < len(widths) else 110
        tree.heading(col, text=col)
        tree.column(col, width=w, minwidth=50, anchor="center")
    tree.tag_configure("even", background=C["bg_row_even"])
    tree.tag_configure("odd", background=C["bg_row_odd"])
    tree.tag_configure("tipo_a", background=C["bg_row_type_a"])
    tree.tag_configure("tipo_b", background=C["bg_row_type_b"])
    tree.tag_configure("sin_fallas", background=C["green_bg"], foreground=C["green"])


# ══════════════════════════════════════════════
#  DIÁLOGO DE IMPORTACIÓN EXCEL CON MAPEO
# ══════════════════════════════════════════════

class VIZIRImportDialog(tk.Toplevel):
    """Ventana secundaria modal para importar datos desde Excel con mapeo de columnas VIZIR"""

    def __init__(self, parent, filepath):
        super().__init__(parent)
        self.title("📥 Importar Datos desde Excel — Mapeo de Columnas VIZIR")
        self.geometry("1100x820")
        self.minsize(960, 700)
        self.configure(bg=C["bg_main"])
        self.transient(parent)
        self.grab_set()

        self.filepath = filepath
        self.wb = None
        self.result = None  # Will hold processed data dict on success
        self.headers = []
        self.raw_data = []

        # Config vars
        self.selected_sheet = tk.StringVar()
        self.var_hdr_row = tk.IntVar(value=1)
        self.var_row_start = tk.IntVar(value=2)
        self.var_row_end = tk.IntVar(value=100)

        # Column mapping vars
        self.col_falla = tk.StringVar(value="Ninguna")
        self.col_tipo = tk.StringVar(value="Ninguna")
        self.col_codigo = tk.StringVar(value="Ninguna")
        self.col_gravedad = tk.StringVar(value="Ninguna")
        self.col_unidad = tk.StringVar(value="Ninguna")
        self.col_longitud = tk.StringVar(value="Ninguna")
        self.col_area_val = tk.StringVar(value="Ninguna")
        self.col_unidades = tk.StringVar(value="Ninguna")
        self.col_extension = tk.StringVar(value="Ninguna")
        self.col_cant_ini = tk.StringVar(value="Ninguna")
        self.col_cant_fin = tk.StringVar(value="Ninguna")
        self.col_abscisa_ini = tk.StringVar(value="Ninguna")
        self.col_abscisa_fin = tk.StringVar(value="Ninguna")
        self.col_observaciones = tk.StringVar(value="Ninguna")

        self.map_combos = {}
        self.verify_vars = {}  # per-record: {idx: {"name": var, "grav": var}}
        self.datos_procesados = []
        self.nombre_map = {}

        self._build_ui()
        self._load_workbook()
        self.protocol("WM_DELETE_WINDOW", self._cancel)
        self.focus_force()

    def _load_workbook(self):
        try:
            self.wb = openpyxl.load_workbook(self.filepath, data_only=True)
            self.cmb_sheet["values"] = self.wb.sheetnames
            if self.wb.sheetnames:
                self.cmb_sheet.current(0)
                self._on_sheet_change(None)
                self._load_preview()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir:\n{e}", parent=self)
            self._cancel()

    def _build_ui(self):
        # ── Scrollable container ──
        outer = tk.Frame(self, bg=C["bg_main"])
        outer.pack(fill="both", expand=True)

        self._canvas = tk.Canvas(outer, bg=C["bg_main"], highlightthickness=0)
        v_scroll = ttk.Scrollbar(outer, orient="vertical", command=self._canvas.yview)
        h_scroll = ttk.Scrollbar(outer, orient="horizontal", command=self._canvas.xview)
        self._canvas.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

        v_scroll.pack(side="right", fill="y")
        h_scroll.pack(side="bottom", fill="x")
        self._canvas.pack(side="left", fill="both", expand=True)

        main = tk.Frame(self._canvas, bg=C["bg_main"])
        self._canvas.create_window((0, 0), window=main, anchor="nw", tags="inner")
        main.bind("<Configure>",
                  lambda e: self._canvas.configure(scrollregion=self._canvas.bbox("all")))

        def _on_mousewheel(event):
            self._canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        def _on_shift_mousewheel(event):
            self._canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")
        self._canvas.bind_all("<MouseWheel>", _on_mousewheel)
        self._canvas.bind_all("<Shift-MouseWheel>", _on_shift_mousewheel)

        def _on_canvas_cfg(event):
            self._canvas.itemconfig("inner", width=event.width)
        self._canvas.bind("<Configure>", _on_canvas_cfg)

        # ── Title ──
        title_f = tk.Frame(main, bg=C["bg_header"], height=44)
        title_f.pack(fill="x")
        title_f.pack_propagate(False)
        tk.Label(title_f, text="📥  IMPORTAR DATOS DESDE EXCEL — VIZIR",
                 font=("Segoe UI", 13, "bold"),
                 fg=C["txt_primary"], bg=C["bg_header"]).pack(side="left", padx=15)
        tk.Label(title_f, text=os.path.basename(self.filepath),
                 font=("Segoe UI", 10),
                 fg=C["accent"], bg=C["bg_header"]).pack(side="left", padx=10)

        # ── Sheet + row config ──
        card1, body1 = make_card(main, "Configuración de Hoja y Filas")
        card1.pack(fill="x", padx=12, pady=(10, 6))

        row_cfg = tk.Frame(body1, bg=C["bg_card"])
        row_cfg.pack(fill="x")

        tk.Label(row_cfg, text="Hoja:", font=("Segoe UI", 10),
                 fg=C["txt_second"], bg=C["bg_card"]).pack(side="left")
        self.cmb_sheet = make_combo(row_cfg, self.selected_sheet, [], width=22)
        self.cmb_sheet.pack(side="left", padx=8)
        self.cmb_sheet.bind("<<ComboboxSelected>>", self._on_sheet_change)

        tk.Label(row_cfg, text="Fila encabezados:", font=("Segoe UI", 10),
                 fg=C["txt_second"], bg=C["bg_card"]).pack(side="left", padx=(18, 4))
        make_entry(row_cfg, textvariable=self.var_hdr_row, width=5).pack(side="left")

        tk.Label(row_cfg, text="Fila inicio:", font=("Segoe UI", 10),
                 fg=C["txt_second"], bg=C["bg_card"]).pack(side="left", padx=(18, 4))
        make_entry(row_cfg, textvariable=self.var_row_start, width=5).pack(side="left")

        tk.Label(row_cfg, text="Fila fin:", font=("Segoe UI", 10),
                 fg=C["txt_second"], bg=C["bg_card"]).pack(side="left", padx=(18, 4))
        make_entry(row_cfg, textvariable=self.var_row_end, width=5).pack(side="left")

        make_button(row_cfg, "↺  Actualizar", self._load_preview, style="dim").pack(
            side="left", padx=18)

        # ── Preview ──
        card2, body2 = make_card(main, "Vista Previa del Excel (primeras 40 filas)")
        card2.pack(fill="x", padx=12, pady=6)

        tree_frame = tk.Frame(body2, bg=C["bg_panel"])
        tree_frame.pack(fill="both", expand=True, pady=4)
        self.tree_preview = ttk.Treeview(tree_frame, height=5)
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_preview.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree_preview.xview)
        self.tree_preview.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree_preview.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)

        # ── Column Mapping ──
        card3, body3 = make_card(main, "Mapeo de Columnas (Filtro)")
        card3.pack(fill="x", padx=12, pady=6)

        tk.Label(body3,
                 text="Asigne cada columna del archivo a su campo. Los campos con (*) son obligatorios.",
                 font=("Segoe UI", 9), fg=C["txt_dim"],
                 bg=C["bg_card"]).pack(anchor="w", pady=(0, 8))

        map_grid = tk.Frame(body3, bg=C["bg_card"])
        map_grid.pack(fill="x")

        map_fields = [
            ("(*) Nombre de Falla:", self.col_falla, "Nombre descriptivo del daño"),
            ("(*) Gravedad:", self.col_gravedad, "1 = Bajo, 2 = Regular, 3 = Alto"),
            ("Tipo (A / B):", self.col_tipo, "Estructural (A) o Funcional (B)"),
            ("Código:", self.col_codigo, "Código VIZIR (FPC, AH, etc.)"),
            ("Unidad:", self.col_unidad, "m, m² o unidad (1) (si está en una columna)"),
            ("Longitud (m):", self.col_longitud, "Cantidad en metros lineales"),
            ("Área (m²):", self.col_area_val, "Cantidad en metros cuadrados"),
            ("Unidades (und):", self.col_unidades, "Cantidad en unidades"),
            ("Extensión (%):", self.col_extension, "Se calcula si no existe"),
            ("Abscisa Inicial:", self.col_abscisa_ini, "Progresiva de inicio"),
            ("Abscisa Final:", self.col_abscisa_fin, "Progresiva de fin"),
            ("Observaciones:", self.col_observaciones, "Notas adicionales"),
        ]
        for i, (lbl, var, tip) in enumerate(map_fields):
            is_req = lbl.startswith("(*)")
            fg = C["gold"] if is_req else C["txt_primary"]
            tk.Label(map_grid, text=lbl, font=("Segoe UI", 10, "bold"),
                     fg=fg, bg=C["bg_card"]).grid(
                row=i, column=0, sticky="e", padx=(0, 8), pady=3)
            cb = make_combo(map_grid, var, ["Ninguna"], width=32)
            cb.grid(row=i, column=1, pady=3, sticky="w")
            tk.Label(map_grid, text=tip, font=("Segoe UI", 9),
                     fg=C["txt_dim"], bg=C["bg_card"]).grid(
                row=i, column=2, sticky="w", padx=12, pady=3)
            self.map_combos[lbl] = cb

        # Partial quantities range
        cant_row = tk.Frame(body3, bg=C["bg_card"])
        cant_row.pack(fill="x", pady=(8, 0))
        tk.Label(cant_row, text="Cantidades parciales →  Desde:",
                 font=("Segoe UI", 10), fg=C["txt_second"],
                 bg=C["bg_card"]).pack(side="left")
        self.cmb_cant_ini = make_combo(cant_row, self.col_cant_ini, ["Ninguna"], width=18)
        self.cmb_cant_ini.pack(side="left", padx=6)
        tk.Label(cant_row, text="Hasta:", font=("Segoe UI", 10),
                 fg=C["txt_second"], bg=C["bg_card"]).pack(side="left", padx=(12, 0))
        self.cmb_cant_fin = make_combo(cant_row, self.col_cant_fin, ["Ninguna"], width=18)
        self.cmb_cant_fin.pack(side="left", padx=6)

        btn_map = tk.Frame(body3, bg=C["bg_card"])
        btn_map.pack(fill="x", pady=(12, 0))
        make_button(btn_map, "🔍  Autodetectar", self._autodetect, style="dim").pack(
            side="left")
        make_button(btn_map, "✅  Procesar", self._process_and_import,
                    style="green").pack(side="left", padx=12)
        self.lbl_status = tk.Label(btn_map, text="",
                                    font=("Segoe UI", 10),
                                    fg=C["txt_dim"], bg=C["bg_card"])
        self.lbl_status.pack(side="left", padx=12)

        # Inline verification for names (populated after processing)
        self.verify_name_frame = tk.LabelFrame(body3,
            text="  📋  Nombres de Falla encontrados  ",
            bg=C["bg_card"], fg=C["accent"],
            font=("Segoe UI", 10, "bold"), padx=10, pady=6)
        # Not packed yet — shown after processing

        # Inline verification for severity (populated after processing)
        self.verify_grav_frame = tk.LabelFrame(body3,
            text="  📋  Gravedades encontradas  ",
            bg=C["bg_card"], fg=C["accent"],
            font=("Segoe UI", 10, "bold"), padx=10, pady=6)
        # Not packed yet

        # Confirm button (shown after processing)
        self.btn_confirm_frame = tk.Frame(body3, bg=C["bg_card"])
        # Not packed yet

        # ── Cancel button ──
        btn_cancel = tk.Frame(main, bg=C["bg_main"])
        btn_cancel.pack(fill="x", padx=12, pady=(10, 15))
        make_button(btn_cancel, "✖  Cancelar",
                    self._cancel, style="danger").pack(side="left")

    # ── Helpers ──

    def _on_sheet_change(self, _):
        if not self.wb:
            return
        ws = self.wb[self.selected_sheet.get()]
        self.var_row_end.set(ws.max_row)


    def _load_preview(self):
        if not self.wb:
            return
        ws = self.wb[self.selected_sheet.get()]
        hr = self.var_hdr_row.get()
        rs = self.var_row_start.get()
        re_ = self.var_row_end.get()

        self.headers = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=hr, column=c).value
            letter = openpyxl.utils.get_column_letter(c)
            self.headers.append(f"{letter}: {v}" if v else f"{letter}: (Col {c})")

        self.raw_data = []
        for r in range(rs, re_ + 1):
            row = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                row.append(v if v is not None else "")
            if any(str(x).strip() for x in row):
                self.raw_data.append(row)

        # Update preview tree
        configure_dark_treeview(self.tree_preview,
                                self.headers,
                                [max(80, len(str(h)) * 7) for h in self.headers])
        self.tree_preview.delete(*self.tree_preview.get_children())
        for i, row in enumerate(self.raw_data[:40]):
            disp = [str(v)[:25] if v else "" for v in row]
            while len(disp) < len(self.headers):
                disp.append("")
            tag = "even" if i % 2 == 0 else "odd"
            self.tree_preview.insert("", "end", values=disp, tags=(tag,))

        # Update combos
        opts = ["Ninguna"] + self.headers
        for cb in self.map_combos.values():
            cb["values"] = opts
        self.cmb_cant_ini["values"] = opts
        self.cmb_cant_fin["values"] = opts
        self._autodetect(silent=True, only_empty=True)

        self.lbl_status.config(
            text=f"{len(self.raw_data)} filas × {len(self.headers)} columnas", fg=C["accent"])

    def _autodetect(self, silent=False, only_empty=False):
        if not self.headers:
            messagebox.showwarning("Aviso", "Cargue la vista previa primero.", parent=self)
            return
        kw = {
            "Nombre": self.col_falla,
            "Gravedad": self.col_gravedad,
            "Tipo": self.col_tipo,
            "Código": self.col_codigo,
            "Unidad": self.col_unidad,
            "Longitud": self.col_longitud,
            "Área": self.col_area_val,
            "Unidades": self.col_unidades,
            "Extensión": self.col_extension,
            "Abscisa Inicial": self.col_abscisa_ini,
            "Abscisa Final": self.col_abscisa_fin,
            "Observaciones": self.col_observaciones,
        }
        if hasattr(self, "col_archivo"):
            kw["Archivo / Imagen"] = self.col_archivo
        found = 0
        for field_label, var in kw.items():
            if only_empty and var.get() not in ("", "Ninguna"):
                continue
            header, score = autodetect_header_vizir(self.headers, field_label)
            if header and score > 0:
                var.set(header)
                found += 1
        if not silent:
            messagebox.showinfo("Autodetección", f"Se detectaron {found} columnas.", parent=self)

    def _col_idx(self, var):
        v = var.get()
        if v == "Ninguna" or not v:
            return None
        try:
            return self.headers.index(v)
        except ValueError:
            return None

    def _process_and_import(self):
        """Process data and populate verification, then allow confirmation."""
        i_falla = self._col_idx(self.col_falla)
        i_tipo = self._col_idx(self.col_tipo)
        i_cod = self._col_idx(self.col_codigo)
        i_grav = self._col_idx(self.col_gravedad)
        i_ext = self._col_idx(self.col_extension)
        i_ci = self._col_idx(self.col_cant_ini)
        i_cf = self._col_idx(self.col_cant_fin)
        i_abs_i = self._col_idx(self.col_abscisa_ini)
        i_abs_f = self._col_idx(self.col_abscisa_fin)
        i_obs = self._col_idx(self.col_observaciones)
        i_arch = self._col_idx(self.col_archivo) if hasattr(self, "col_archivo") else None

        # Quantity column indices: Longitud, Área, Unidades
        i_und = self._col_idx(self.col_unidad)
        i_long = self._col_idx(self.col_longitud)
        i_area = self._col_idx(self.col_area_val)
        i_unid = self._col_idx(self.col_unidades)

        # Mandatory: Nombre (falla) + Gravedad
        if i_falla is None or i_grav is None:
            messagebox.showerror("Error",
                                 "'Nombre de Falla' y 'Gravedad' son obligatorios.", parent=self)
            return

        self.datos_procesados = []

        for row in self.raw_data:
            def safe(idx):
                if idx is None or idx >= len(row):
                    return ""
                return row[idx] if row[idx] is not None else ""

            nombre = str(safe(i_falla)).strip()
            if not nombre:
                continue

            # Progresivas sin fallas: preservar la progresiva con registro marcador
            if es_sin_fallas(nombre):
                entry = {
                    "falla": nombre,
                    "tipo": "",
                    "codigo": "",
                    "unidad": "",
                    "gravedad": 0,
                    "total": 0.0,
                    "extension": 0.0,
                    "extension_raw": 0.0,
                    "sin_fallas": True,
                    "archivo_progresiva": str(safe(i_arch)).strip(),
                    "abscisa_ini": str(safe(i_abs_i)).strip(),
                    "abscisa_fin": str(safe(i_abs_f)).strip(),
                    "observaciones": str(safe(i_obs)).strip(),
                }
                self.datos_procesados.append(completar_progresiva_entry(entry))
                continue

            grav = 0
            try:
                grav = int(float(str(safe(i_grav))))
            except (ValueError, TypeError):
                pass
            if grav <= 0:
                continue

            cod = str(safe(i_cod)).strip().upper() if i_cod is not None else ""

            ext = 0.0
            try:
                ext = float(str(safe(i_ext)))
            except (ValueError, TypeError):
                pass

            # Read quantity from the column that matches the unit of measurement
            total = 0.0
            unidad = ""
            # Try Unidad text column first
            if i_und is not None:
                unidad = str(safe(i_und)).strip()

            # Determine expected unit from fault catalog if not in column
            if not unidad and cod:
                all_fallas = {**FALLAS_TIPO_A, **FALLAS_TIPO_B}
                if cod in all_fallas:
                    unidad = all_fallas[cod].get("unidad", "")

            unidad = unidad_canonica_vizir(cod, unidad)

            # Map unit to the corresponding quantity column only
            unidad_lower = normalizar_unidad_medida(unidad)
            if "m2" in unidad_lower:
                # Área (m²) → only read from Área column
                if i_area is not None:
                    try:
                        val = float(str(safe(i_area)))
                        if val > 0:
                            total = val
                    except (ValueError, TypeError):
                        pass
            elif es_unidad_contable(unidad):
                # Unidades → only read from Unidades column (enteros, unidad=1)
                if i_unid is not None:
                    try:
                        val = int(float(str(safe(i_unid))))
                        if val > 0:
                            total = val
                    except (ValueError, TypeError):
                        pass
            elif "m" in unidad_lower:
                # Longitud (m) → only read from Longitud column
                if i_long is not None:
                    try:
                        val = float(str(safe(i_long)))
                        if val > 0:
                            total = val
                    except (ValueError, TypeError):
                        pass
            else:
                # Unknown unit: try each column individually, pick the first with value
                for idx_q, uname in [(i_long, "m"), (i_area, "m²"), (i_unid, "unidad (1)")]:
                    if idx_q is not None:
                        try:
                            val = float(str(safe(idx_q)))
                            if val > 0:
                                total = int(val) if es_unidad_contable(uname) else val
                                if not unidad:
                                    unidad = uname
                                break
                        except (ValueError, TypeError):
                            pass

            # Sum partial quantities if still 0
            if total == 0.0 and i_ci is not None and i_cf is not None:
                parcial_sum = 0.0
                for ci in range(i_ci, i_cf + 1):
                    if ci < len(row) and row[ci]:
                        try:
                            parcial_sum += float(row[ci])
                        except (ValueError, TypeError):
                            pass
                # Entero si la unidad es de conteo
                if es_unidad_contable(unidad):
                    total = int(parcial_sum)
                else:
                    total = parcial_sum

            entry = {
                "falla": nombre,
                "tipo": str(safe(i_tipo)).strip().upper(),
                "codigo": cod,
                "unidad": unidad_canonica_vizir(cod, unidad),
                "gravedad": min(grav, 3),
                "total": total,
                "extension": ext,
                "extension_raw": ext,
                "archivo_progresiva": str(safe(i_arch)).strip(),
                "abscisa_ini": str(safe(i_abs_i)).strip(),
                "abscisa_fin": str(safe(i_abs_f)).strip(),
                "observaciones": str(safe(i_obs)).strip(),
            }
            self.datos_procesados.append(completar_progresiva_entry(entry))

        if not self.datos_procesados:
            messagebox.showwarning("Aviso", "No se encontraron registros válidos.", parent=self)
            return

        # Populate verification with per-record name + severity correction
        self._populate_verification()

        self.lbl_status.config(
            text=f"✅ {len(self.datos_procesados)} registros procesados. "
                 f"Corrija nombres/severidad y pulse 'Confirmar e Importar'.",
            fg=C["green"])

    def _populate_verification(self):
        """Populate inline verification frames for names and severity."""
        # Clear previous content
        for w in self.verify_name_frame.winfo_children():
            w.destroy()
        for w in self.verify_grav_frame.winfo_children():
            w.destroy()
        for w in self.btn_confirm_frame.winfo_children():
            w.destroy()

        # Show frames (pack in correct order)
        self.verify_name_frame.pack(fill="x", pady=(10, 0))
        self.verify_grav_frame.pack(fill="x", pady=(6, 0))
        self.btn_confirm_frame.pack(fill="x", pady=(10, 0))

        all_fallas = {**FALLAS_TIPO_A, **FALLAS_TIPO_B}
        name_options = ["(Sin cambio)"] + \
                       [f"{k} — {v['nombre']}" for k, v in FALLAS_TIPO_A.items()] + \
                       [f"{k} — {v['nombre']}" for k, v in FALLAS_TIPO_B.items()]

        # ── Unique names with correction combos ──
        unique_names = {}
        n_sin_fallas = 0
        for d in self.datos_procesados:
            if d.get("sin_fallas"):
                n_sin_fallas += 1
                continue
            key = d["falla"]
            if key not in unique_names:
                unique_names[key] = {"codigo": d["codigo"], "count": 0,
                                     "total": 0.0, "unidad": d.get("unidad", "")}
            unique_names[key]["count"] += 1
            unique_names[key]["total"] += d.get("total", 0.0)

        info_txt = f"{len(unique_names)} nombres únicos. Corrija si es necesario:"
        if n_sin_fallas:
            info_txt += f"  ({n_sin_fallas} progresivas sin fallas detectadas)"
        tk.Label(self.verify_name_frame,
                 text=info_txt,
                 font=("Segoe UI", 9), fg=C["txt_dim"],
                 bg=C["bg_card"]).pack(anchor="w", pady=(0, 4))

        self.verify_name_vars = {}
        for i, (name, info) in enumerate(unique_names.items()):
            matched = info["codigo"] in all_fallas
            bg = C["bg_row_even"] if i % 2 == 0 else C["bg_row_odd"]
            rf = tk.Frame(self.verify_name_frame, bg=bg)
            rf.pack(fill="x")

            st = "✅" if matched else "⚠️"
            lbl_txt = f'{st}  "{name[:28]}"  (Cód: {info["codigo"] or "?"}' \
                      f'  ×{info["count"]}  ∑{fmt_total_str(info["total"], info["unidad"])} {info["unidad"]})'
            tk.Label(rf, text=lbl_txt, font=("Segoe UI", 9),
                     fg=C["txt_primary"], bg=bg, anchor="w").pack(
                side="left", padx=6, pady=3)

            var = tk.StringVar(value="(Sin cambio)")
            cb = make_combo(rf, var, name_options, width=36)
            cb.pack(side="right", padx=6, pady=3)
            self.verify_name_vars[name] = var

        # ── Unique severities (excluir registros sin fallas) ──
        unique_gravs = set()
        for d in self.datos_procesados:
            if d.get("sin_fallas"):
                continue
            unique_gravs.add(d["gravedad"])

        grav_text = ", ".join(str(g) for g in sorted(unique_gravs))
        valid = all(1 <= g <= 3 for g in unique_gravs) if unique_gravs else True
        st_txt = "✅ Válidas" if valid else "⚠️ Valores fuera de rango (deben ser 1, 2 o 3)"
        st_fg = C["green"] if valid else C["orange"]
        tk.Label(self.verify_grav_frame,
                 text=f"Valores encontrados: {grav_text}  —  {st_txt}",
                 font=("Segoe UI", 10), fg=st_fg,
                 bg=C["bg_card"]).pack(anchor="w", pady=2)

        # Confirm button
        make_button(self.btn_confirm_frame, "✅  Confirmar e Importar",
                    self._confirm_and_close, style="green").pack(side="left")

    def _confirm_and_close(self):
        """Apply name corrections by unique name and return result."""
        # Apply name corrections
        for orig_name, var in self.verify_name_vars.items():
            val = var.get()
            if val != "(Sin cambio)" and " — " in val:
                parts = val.split(" — ", 1)
                new_cod = parts[0].strip()
                new_name = parts[1].strip()
                # Determine tipo
                new_tipo = ""
                if new_cod in FALLAS_TIPO_A:
                    new_tipo = "A"
                elif new_cod in FALLAS_TIPO_B:
                    new_tipo = "B"
                # Apply to ALL records with this original name
                for d in self.datos_procesados:
                    if d["falla"] == orig_name:
                        d["codigo"] = new_cod
                        d["falla"] = new_name
                        if new_tipo:
                            d["tipo"] = new_tipo

        fallas_tipo_a = [d for d in self.datos_procesados if d["tipo"] == "A"]
        fallas_tipo_b = [d for d in self.datos_procesados if d["tipo"] == "B"]

        self.result = {
            "datos_procesados": self.datos_procesados,
            "fallas_tipo_a": fallas_tipo_a,
            "fallas_tipo_b": fallas_tipo_b,
            "nombre_map": {},
            "headers": self.headers,
            "raw_data": self.raw_data,
            "wb": self.wb,
        }

        self._cleanup()
        self.destroy()

    def _cancel(self):
        self.result = None
        self._cleanup()
        self.destroy()

    def _cleanup(self):
        try:
            self._canvas.unbind_all("<MouseWheel>")
            self._canvas.unbind_all("<Shift-MouseWheel>")
        except Exception:
            pass
        self.grab_release()


# ══════════════════════════════════════════════
#  DIÁLOGO DE CORRECCIONES MANUALES
# ══════════════════════════════════════════════

class VIZIRCorrectionDialog(tk.Toplevel):
    """Ventana para corregir manualmente nombres de fallas y severidad."""

    def __init__(self, parent, datos_procesados):
        super().__init__(parent)
        self.title("✏️ Corrección Manual de Nombres y Severidad")
        self.geometry("1050x680")
        self.minsize(900, 500)
        self.configure(bg=C["bg_main"])
        self.transient(parent)
        self.grab_set()

        self.datos = [dict(d) for d in datos_procesados]  # deep copy
        self.result = None
        self.row_widgets = []

        all_fallas = {**FALLAS_TIPO_A, **FALLAS_TIPO_B}
        self.catalog_options = ["(Sin cambio)"] + \
            [f"{k} — {v['nombre']}" for k, v in FALLAS_TIPO_A.items()] + \
            [f"{k} — {v['nombre']}" for k, v in FALLAS_TIPO_B.items()]
        self.grav_options = ["(Sin cambio)", "1", "2", "3"]

        self._build_ui()
        self.protocol("WM_DELETE_WINDOW", self._cancel)
        self.focus_force()

    def _build_ui(self):
        # Title
        title_f = tk.Frame(self, bg=C["bg_header"], height=40)
        title_f.pack(fill="x")
        title_f.pack_propagate(False)
        tk.Label(title_f, text="✏️  CORRECCIÓN MANUAL — Nombres y Severidad",
                 font=("Segoe UI", 12, "bold"),
                 fg=C["txt_primary"], bg=C["bg_header"]).pack(side="left", padx=15)

        # Scrollable table
        outer = tk.Frame(self, bg=C["bg_main"])
        outer.pack(fill="both", expand=True, padx=10, pady=8)

        canvas = tk.Canvas(outer, bg=C["bg_main"], highlightthickness=0)
        vsb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        self._inner = tk.Frame(canvas, bg=C["bg_main"])
        canvas.create_window((0, 0), window=self._inner, anchor="nw", tags="inner")
        self._inner.bind("<Configure>",
                         lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
                    lambda e: canvas.itemconfig("inner", width=e.width))
        canvas.bind_all("<MouseWheel>",
                        lambda e: canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))
        self._canvas = canvas

        # Header row
        hdr = tk.Frame(self._inner, bg=C["bg_header"])
        hdr.pack(fill="x", pady=(0, 2))
        for txt, w in [("#", 4), ("Nombre actual", 28), ("Código", 7),
                        ("Grav.", 5), ("Nuevo Nombre / Código", 38),
                        ("Nueva Grav.", 10)]:
            tk.Label(hdr, text=txt, font=("Segoe UI", 9, "bold"),
                     fg=C["accent"], bg=C["bg_header"], width=w,
                     anchor="w").pack(side="left", padx=4, pady=4)

        # Data rows
        for i, d in enumerate(self.datos):
            bg = C["bg_row_even"] if i % 2 == 0 else C["bg_row_odd"]
            rf = tk.Frame(self._inner, bg=bg)
            rf.pack(fill="x")

            tk.Label(rf, text=str(i+1), font=("Segoe UI", 9),
                     fg=C["txt_dim"], bg=bg, width=4).pack(side="left", padx=4, pady=2)
            tk.Label(rf, text=d["falla"][:30], font=("Segoe UI", 9),
                     fg=C["txt_primary"], bg=bg, width=28,
                     anchor="w").pack(side="left", padx=4)
            tk.Label(rf, text=d.get("codigo", ""), font=("Segoe UI", 9, "bold"),
                     fg=C["txt_second"], bg=bg, width=7,
                     anchor="w").pack(side="left", padx=4)
            tk.Label(rf, text=str(d["gravedad"]), font=("Segoe UI", 9),
                     fg=C["txt_primary"], bg=bg, width=5).pack(side="left", padx=4)

            var_name = tk.StringVar(value="(Sin cambio)")
            cb_name = make_combo(rf, var_name, self.catalog_options, width=38)
            cb_name.pack(side="left", padx=4, pady=2)

            var_grav = tk.StringVar(value="(Sin cambio)")
            cb_grav = make_combo(rf, var_grav, self.grav_options, width=10)
            cb_grav.pack(side="left", padx=4, pady=2)

            self.row_widgets.append({"idx": i, "var_name": var_name, "var_grav": var_grav})

        # Action buttons
        btn_f = tk.Frame(self, bg=C["bg_main"])
        btn_f.pack(fill="x", padx=10, pady=(6, 12))
        make_button(btn_f, "✅  Aplicar Correcciones",
                    self._apply, style="green").pack(side="left", padx=(0, 10))
        make_button(btn_f, "✖  Cancelar",
                    self._cancel, style="danger").pack(side="left")

    def _apply(self):
        changed = 0
        for rw in self.row_widgets:
            i = rw["idx"]
            vn = rw["var_name"].get()
            vg = rw["var_grav"].get()
            if vn != "(Sin cambio)" and " — " in vn:
                parts = vn.split(" — ", 1)
                self.datos[i]["codigo"] = parts[0].strip()
                self.datos[i]["falla"] = parts[1].strip()
                changed += 1
            if vg != "(Sin cambio)":
                try:
                    self.datos[i]["gravedad"] = int(vg)
                    changed += 1
                except ValueError:
                    pass
        self.result = self.datos
        self._cleanup()
        self.destroy()
        messagebox.showinfo("Correcciones", f"Se aplicaron {changed} correcciones.")

    def _cancel(self):
        self.result = None
        self._cleanup()
        self.destroy()

    def _cleanup(self):
        try:
            self._canvas.unbind_all("<MouseWheel>")
        except Exception:
            pass
        self.grab_release()


# ══════════════════════════════════════════════
#  DIÁLOGO UNIFICADO: IMPORTAR + MAPEO (3 col)
# ══════════════════════════════════════════════

class VIZIRColumnMapDialog(tk.Toplevel):
    """Ventana unificada de importación y mapeo de columnas con layout de 3 columnas,
    vista previa del Excel, autodetectar y botones de corrección."""

    def __init__(self, parent, filepath=None, headers=None, datos_procesados=None):
        super().__init__(parent)
        self.geometry("1120x850")
        self.minsize(980, 700)
        self.configure(bg=C["bg_main"])
        self.transient(parent)
        self.grab_set()

        self.filepath = filepath
        self.wb = None
        self.result = None
        self.headers = headers or []
        self.raw_data = []
        self.datos_procesados = list(datos_procesados) if datos_procesados else []

        # Mode: "import" if filepath given, "remap" if datos_procesados given
        self.mode = "import" if filepath else "remap"
        self.title("📥  Importar y Mapear Columnas — VIZIR" if self.mode == "import"
                   else "🔧  Mapeo de Columnas — VIZIR")

        # Config vars
        self.selected_sheet = tk.StringVar()
        self.var_hdr_row = tk.IntVar(value=1)
        self.var_row_start = tk.IntVar(value=2)
        self.var_row_end = tk.IntVar(value=100)

        # Column mapping vars
        self.col_vars = {}
        self.map_combos = {}

        # Partial quantities
        self.col_cant_ini = tk.StringVar(value="Ninguna")
        self.col_cant_fin = tk.StringVar(value="Ninguna")

        self._build_ui()
        if self.mode == "import":
            self._load_workbook()
        else:
            self._update_combos_from_headers()
            self._update_summary()
        self.protocol("WM_DELETE_WINDOW", self._cancel)
        self.focus_force()

    # ──────────── LOAD WORKBOOK ────────────
    def _load_workbook(self):
        try:
            self.wb = openpyxl.load_workbook(self.filepath, data_only=True)
            self.cmb_sheet["values"] = self.wb.sheetnames
            if self.wb.sheetnames:
                self.cmb_sheet.current(0)
                self._on_sheet_change(None)
                self._load_preview()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir:\n{e}", parent=self)
            self._cancel()

    # ──────────── BUILD UI ────────────
    def _build_ui(self):
        # ── Scrollable outer ──
        outer = tk.Frame(self, bg=C["bg_main"])
        outer.pack(fill="both", expand=True)

        self._canvas = tk.Canvas(outer, bg=C["bg_main"], highlightthickness=0)
        v_scroll = ttk.Scrollbar(outer, orient="vertical", command=self._canvas.yview)
        h_scroll = ttk.Scrollbar(outer, orient="horizontal", command=self._canvas.xview)
        self._canvas.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
        v_scroll.pack(side="right", fill="y")
        h_scroll.pack(side="bottom", fill="x")
        self._canvas.pack(side="left", fill="both", expand=True)

        main = tk.Frame(self._canvas, bg=C["bg_main"])
        self._canvas.create_window((0, 0), window=main, anchor="nw", tags="inner")
        main.bind("<Configure>",
                  lambda e: self._canvas.configure(scrollregion=self._canvas.bbox("all")))

        def _mw(event):
            self._canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        def _mw_shift(event):
            self._canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")
        self._canvas.bind_all("<MouseWheel>", _mw)
        self._canvas.bind_all("<Shift-MouseWheel>", _mw_shift)
        self._canvas.bind("<Configure>",
                          lambda e: self._canvas.itemconfig("inner", width=e.width))

        # ── Title ──
        title_f = tk.Frame(main, bg=C["bg_header"], height=44)
        title_f.pack(fill="x")
        title_f.pack_propagate(False)
        title_txt = ("📥  IMPORTAR Y MAPEAR COLUMNAS — VIZIR" if self.mode == "import"
                     else "🔧  MAPEO DE COLUMNAS — VIZIR")
        tk.Label(title_f, text=title_txt,
                 font=("Segoe UI", 13, "bold"),
                 fg=C["txt_primary"], bg=C["bg_header"]).pack(side="left", padx=15)
        if self.filepath:
            tk.Label(title_f, text=os.path.basename(self.filepath),
                     font=("Segoe UI", 10),
                     fg=C["accent"], bg=C["bg_header"]).pack(side="left", padx=10)

        # ══════════════════════════════════════════
        #  SECCIÓN 1: Configuración de Hoja y Filas
        # ══════════════════════════════════════════
        if self.mode == "import":
            card1, body1 = make_card(main, "Configuración de Hoja y Filas")
            card1.pack(fill="x", padx=12, pady=(10, 6))

            row_cfg = tk.Frame(body1, bg=C["bg_card"])
            row_cfg.pack(fill="x")

            tk.Label(row_cfg, text="Hoja:", font=("Segoe UI", 10),
                     fg=C["txt_second"], bg=C["bg_card"]).pack(side="left")
            self.cmb_sheet = make_combo(row_cfg, self.selected_sheet, [], width=22)
            self.cmb_sheet.pack(side="left", padx=8)
            self.cmb_sheet.bind("<<ComboboxSelected>>", self._on_sheet_change)

            tk.Label(row_cfg, text="Fila encabezados:", font=("Segoe UI", 10),
                     fg=C["txt_second"], bg=C["bg_card"]).pack(side="left", padx=(18, 4))
            make_entry(row_cfg, textvariable=self.var_hdr_row, width=5).pack(side="left")

            tk.Label(row_cfg, text="Fila inicio:", font=("Segoe UI", 10),
                     fg=C["txt_second"], bg=C["bg_card"]).pack(side="left", padx=(18, 4))
            make_entry(row_cfg, textvariable=self.var_row_start, width=5).pack(side="left")

            tk.Label(row_cfg, text="Fila fin:", font=("Segoe UI", 10),
                     fg=C["txt_second"], bg=C["bg_card"]).pack(side="left", padx=(18, 4))
            make_entry(row_cfg, textvariable=self.var_row_end, width=5).pack(side="left")

            make_button(row_cfg, "↺  Actualizar", self._load_preview,
                        style="dim").pack(side="left", padx=18)

        # ══════════════════════════════════════════
        #  SECCIÓN 2: Vista Previa del Excel
        # ══════════════════════════════════════════
        if self.mode == "import":
            card2, body2 = make_card(main, "Vista Previa del Excel (primeras 40 filas)")
            card2.pack(fill="x", padx=12, pady=6)

            tree_frame = tk.Frame(body2, bg=C["bg_panel"])
            tree_frame.pack(fill="both", expand=True, pady=4)
            self.tree_preview = ttk.Treeview(tree_frame, height=5)
            vsb_t = ttk.Scrollbar(tree_frame, orient="vertical",
                                   command=self.tree_preview.yview)
            hsb_t = ttk.Scrollbar(tree_frame, orient="horizontal",
                                   command=self.tree_preview.xview)
            self.tree_preview.configure(yscrollcommand=vsb_t.set,
                                         xscrollcommand=hsb_t.set)
            self.tree_preview.grid(row=0, column=0, sticky="nsew")
            vsb_t.grid(row=0, column=1, sticky="ns")
            hsb_t.grid(row=1, column=0, sticky="ew")
            tree_frame.grid_columnconfigure(0, weight=1)
            tree_frame.grid_rowconfigure(0, weight=1)

        # ══════════════════════════════════════════
        #  SECCIÓN 3: Mapeo de Columnas (3 columnas)
        # ══════════════════════════════════════════
        card3, body3 = make_card(main, "Mapeo de Columnas")
        card3.pack(fill="x", padx=12, pady=6)

        tk.Label(body3,
                 text="Asigne cada columna del archivo a su campo. "
                      "Los campos con (*) son obligatorios.",
                 font=("Segoe UI", 9), fg=C["txt_dim"],
                 bg=C["bg_card"]).pack(anchor="w", pady=(0, 6))

        # ── 3-column header ──
        hdr_f = tk.Frame(body3, bg=C["bg_header"])
        hdr_f.pack(fill="x", pady=(0, 4))
        tk.Label(hdr_f, text="Campo Requerido", font=("Segoe UI", 10, "bold"),
                 fg=C["accent"], bg=C["bg_header"], width=24,
                 anchor="center").pack(side="left", padx=2, pady=6)
        tk.Label(hdr_f, text="Columna en el Excel", font=("Segoe UI", 10, "bold"),
                 fg=C["accent"], bg=C["bg_header"], width=34,
                 anchor="center").pack(side="left", padx=2, pady=6)
        tk.Label(hdr_f, text="Descripción", font=("Segoe UI", 10, "bold"),
                 fg=C["accent"], bg=C["bg_header"],
                 anchor="center").pack(side="left", fill="x", expand=True, padx=2, pady=6)

        # ── Field rows ──
        fields_info = [
            ("Nombre de Falla",  True,  "Nombre descriptivo del daño",     "nombre"),
            ("Gravedad",         True,  "1 = Bajo, 2 = Regular, 3 = Alto", "gravedad"),
            ("Tipo (A / B)",     False, "Estructural (A) o Funcional (B)", None),
            ("Código",           False, "Código VIZIR (FPC, AH, etc.)",    None),
            ("Unidad",           False, "m, m² o unidad (1)",             None),
            ("Longitud (m)",     False, "Cantidad en metros lineales",     None),
            ("Área (m²)",        False, "Cantidad en metros cuadrados",    None),
            ("Unidades (und)",   False, "Cantidad en unidades",            None),
            ("Extensión (%)",    False, "Se calcula si no existe",         None),
            ("Archivo / Imagen", False, "Nombre del tif para extraer progresivas", None),
            ("Abscisa Inicial",  False, "Progresiva de inicio",            None),
            ("Abscisa Final",    False, "Progresiva de fin",               None),
            ("Observaciones",    False, "Notas adicionales",               None),
        ]

        opts = ["Ninguna"] + self.headers
        self._correction_btns = {}
        for i, (label, required, desc, corr_type) in enumerate(fields_info):
            bg = C["bg_row_even"] if i % 2 == 0 else C["bg_row_odd"]
            rf = tk.Frame(body3, bg=bg)
            rf.pack(fill="x")

            # Col 1: Campo Requerido
            prefix = "(*) " if required else "      "
            fg = C["gold"] if required else C["txt_primary"]
            tk.Label(rf, text=f"{prefix}{label}", font=("Segoe UI", 10, "bold"),
                     fg=fg, bg=bg, width=24, anchor="w").pack(side="left", padx=6, pady=4)

            # Col 2: Combobox
            var = tk.StringVar(value="Ninguna")
            cb = make_combo(rf, var, opts, width=32)
            cb.pack(side="left", padx=6, pady=4)
            self.col_vars[label] = var
            self.map_combos[label] = cb

            # Col 3: Descripción + botón ⚠
            desc_f = tk.Frame(rf, bg=bg)
            desc_f.pack(side="left", fill="x", expand=True, padx=6)

            tk.Label(desc_f, text=desc, font=("Segoe UI", 9),
                     fg=C["txt_dim"], bg=bg, anchor="w").pack(side="left")

            if corr_type:
                cmd = (self._open_name_correction if corr_type == "nombre"
                       else self._open_grav_correction)
                lbl_btn = ("⚠  Corregir Nombres" if corr_type == "nombre"
                           else "⚠  Corregir Gravedad")
                btn = make_button(
                    desc_f, lbl_btn, cmd,
                    style="warning",
                    compact=True,
                    state="normal" if self.datos_procesados else "disabled",
                )
                btn.pack(side="right", padx=4)
                self._correction_btns[corr_type] = btn

        # ── Partial quantities range ──
        cant_row = tk.Frame(body3, bg=C["bg_card"])
        cant_row.pack(fill="x", pady=(8, 0))
        tk.Label(cant_row, text="Cantidades parciales →  Desde:",
                 font=("Segoe UI", 10), fg=C["txt_second"],
                 bg=C["bg_card"]).pack(side="left")
        self.cmb_cant_ini = make_combo(cant_row, self.col_cant_ini, opts, width=18)
        self.cmb_cant_ini.pack(side="left", padx=6)
        tk.Label(cant_row, text="Hasta:", font=("Segoe UI", 10),
                 fg=C["txt_second"], bg=C["bg_card"]).pack(side="left", padx=(12, 0))
        self.cmb_cant_fin = make_combo(cant_row, self.col_cant_fin, opts, width=18)
        self.cmb_cant_fin.pack(side="left", padx=6)

        # ── Buttons: Autodetectar + Procesar ──
        btn_map = tk.Frame(body3, bg=C["bg_card"])
        btn_map.pack(fill="x", pady=(12, 0))
        make_button(btn_map, "🔍  Autodetectar", self._autodetect,
                    style="dim").pack(side="left")
        make_button(btn_map, "✅  Procesar", self._process_data,
                    style="green").pack(side="left", padx=12)
        self.lbl_status = tk.Label(btn_map, text="",
                                    font=("Segoe UI", 10),
                                    fg=C["txt_dim"], bg=C["bg_card"])
        self.lbl_status.pack(side="left", padx=12)

        # ── Summary frame (populated after processing) ──
        self.summary_frame = tk.Frame(body3, bg=C["bg_card"])
        # Not packed yet

        # ── Confirm button frame (shown after processing) ──
        self.btn_confirm_frame = tk.Frame(body3, bg=C["bg_card"])
        # Not packed yet

        # ── Cancel button ──
        btn_cancel = tk.Frame(main, bg=C["bg_main"])
        btn_cancel.pack(fill="x", padx=12, pady=(10, 15))
        make_button(btn_cancel, "✖  Cancelar",
                    self._cancel, style="danger").pack(side="left")

    # ──────────── SHEET / PREVIEW ────────────

    def _on_sheet_change(self, _):
        if not self.wb:
            return
        ws = self.wb[self.selected_sheet.get()]
        self.var_row_end.set(ws.max_row)

    def _load_preview(self):
        if not self.wb:
            return
        ws = self.wb[self.selected_sheet.get()]
        hr = self.var_hdr_row.get()
        rs = self.var_row_start.get()
        re_ = self.var_row_end.get()

        self.headers = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=hr, column=c).value
            letter = openpyxl.utils.get_column_letter(c)
            self.headers.append(f"{letter}: {v}" if v else f"{letter}: (Col {c})")

        self.raw_data = []
        for r in range(rs, re_ + 1):
            row = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                row.append(v if v is not None else "")
            if any(str(x).strip() for x in row):
                self.raw_data.append(row)

        # Update preview tree
        configure_dark_treeview(self.tree_preview,
                                self.headers,
                                [max(80, len(str(h)) * 7) for h in self.headers])
        self.tree_preview.delete(*self.tree_preview.get_children())
        for i, row in enumerate(self.raw_data[:40]):
            disp = [str(v)[:25] if v else "" for v in row]
            while len(disp) < len(self.headers):
                disp.append("")
            tag = "even" if i % 2 == 0 else "odd"
            self.tree_preview.insert("", "end", values=disp, tags=(tag,))

        # Update all combos
        self._update_combos_from_headers()
        self._autodetect(silent=True, only_empty=True)

        self.lbl_status.config(
            text=f"{len(self.raw_data)} filas × {len(self.headers)} columnas",
            fg=C["accent"])

    def _update_combos_from_headers(self):
        opts = ["Ninguna"] + self.headers
        for cb in self.map_combos.values():
            cb["values"] = opts
        self.cmb_cant_ini["values"] = opts
        self.cmb_cant_fin["values"] = opts

    # ──────────── AUTODETECT ────────────

    def _autodetect(self, silent=False, only_empty=False):
        if not self.headers:
            messagebox.showwarning("Aviso", "Cargue la vista previa primero.", parent=self)
            return
        kw = [
            "Nombre de Falla",
            "Gravedad",
            "Tipo (A / B)",
            "Código",
            "Unidad",
            "Longitud (m)",
            "Área (m²)",
            "Unidades (und)",
            "Extensión (%)",
            "Archivo / Imagen",
            "Abscisa Inicial",
            "Abscisa Final",
            "Observaciones",
        ]
        found = 0
        for field_label in kw:
            var = self.col_vars.get(field_label)
            if not var:
                continue
            if only_empty and var.get() not in ("", "Ninguna"):
                continue
            header, score = autodetect_header_vizir(self.headers, field_label)
            if header and score > 0:
                var.set(header)
                found += 1
        if not silent:
            messagebox.showinfo("Autodetección",
                                f"Se detectaron {found} columnas.", parent=self)

    # ──────────── COL INDEX HELPER ────────────

    def _col_idx(self, label):
        var = self.col_vars.get(label)
        if not var:
            return None
        v = var.get()
        if v == "Ninguna" or not v:
            return None
        try:
            return self.headers.index(v)
        except ValueError:
            return None

    # ──────────── PROCESS DATA ────────────

    def _process_data(self):
        """Process raw data using the column mapping and populate summary."""
        if self.mode == "import" and not self.raw_data:
            messagebox.showwarning("Aviso",
                                   "Cargue la vista previa primero.", parent=self)
            return

        # If in remap mode without raw_data, just update summary
        if self.mode == "remap" and not self.raw_data:
            self._update_summary()
            return

        i_falla = self._col_idx("Nombre de Falla")
        i_grav = self._col_idx("Gravedad")
        i_tipo = self._col_idx("Tipo (A / B)")
        i_cod = self._col_idx("Código")
        i_ext = self._col_idx("Extensión (%)")
        i_und = self._col_idx("Unidad")
        i_long = self._col_idx("Longitud (m)")
        i_area = self._col_idx("Área (m²)")
        i_unid = self._col_idx("Unidades (und)")
        i_arch = self._col_idx("Archivo / Imagen")
        i_abs_i = self._col_idx("Abscisa Inicial")
        i_abs_f = self._col_idx("Abscisa Final")
        i_obs = self._col_idx("Observaciones")

        i_ci_v = self.col_cant_ini.get()
        i_cf_v = self.col_cant_fin.get()
        i_ci = self.headers.index(i_ci_v) if i_ci_v != "Ninguna" and i_ci_v in self.headers else None
        i_cf = self.headers.index(i_cf_v) if i_cf_v != "Ninguna" and i_cf_v in self.headers else None

        if i_falla is None or i_grav is None:
            messagebox.showerror("Error",
                                 "'Nombre de Falla' y 'Gravedad' son obligatorios.",
                                 parent=self)
            return

        self.datos_procesados = []

        for row in self.raw_data:
            def safe(idx):
                if idx is None or idx >= len(row):
                    return ""
                return row[idx] if row[idx] is not None else ""

            nombre = str(safe(i_falla)).strip()
            if not nombre:
                continue

            # Progresivas sin fallas: preservar la progresiva con registro marcador
            if es_sin_fallas(nombre):
                entry = {
                    "falla": nombre,
                    "tipo": "",
                    "codigo": "",
                    "unidad": "",
                    "gravedad": 0,
                    "total": 0.0,
                    "extension": 0.0,
                    "extension_raw": 0.0,
                    "sin_fallas": True,
                    "archivo_progresiva": str(safe(i_arch)).strip(),
                    "abscisa_ini": str(safe(i_abs_i)).strip(),
                    "abscisa_fin": str(safe(i_abs_f)).strip(),
                    "observaciones": str(safe(i_obs)).strip(),
                }
                self.datos_procesados.append(completar_progresiva_entry(entry))
                continue

            grav = 0
            try:
                grav = int(float(str(safe(i_grav))))
            except (ValueError, TypeError):
                pass
            if grav <= 0:
                continue

            cod = str(safe(i_cod)).strip().upper() if i_cod is not None else ""

            ext = 0.0
            try:
                ext = float(str(safe(i_ext)))
            except (ValueError, TypeError):
                pass

            # Read quantity filtered by unit
            total = 0.0
            unidad = ""
            if i_und is not None:
                unidad = str(safe(i_und)).strip()

            if not unidad and cod:
                all_f = {**FALLAS_TIPO_A, **FALLAS_TIPO_B}
                if cod in all_f:
                    unidad = all_f[cod].get("unidad", "")

            unidad = unidad_canonica_vizir(cod, unidad)
            unidad_lower = normalizar_unidad_medida(unidad)
            if "m2" in unidad_lower:
                if i_area is not None:
                    try:
                        val = float(str(safe(i_area)))
                        if val > 0:
                            total = val
                    except (ValueError, TypeError):
                        pass
            elif es_unidad_contable(unidad):
                # Unidades enteras (unidad=1) para ojos de pescado, etc.
                if i_unid is not None:
                    try:
                        val = int(float(str(safe(i_unid))))
                        if val > 0:
                            total = val
                    except (ValueError, TypeError):
                        pass
            elif "m" in unidad_lower:
                if i_long is not None:
                    try:
                        val = float(str(safe(i_long)))
                        if val > 0:
                            total = val
                    except (ValueError, TypeError):
                        pass
            else:
                for idx_q, uname in [(i_long, "m"), (i_area, "m²"), (i_unid, "unidad (1)")]:
                    if idx_q is not None:
                        try:
                            val = float(str(safe(idx_q)))
                            if val > 0:
                                total = int(val) if es_unidad_contable(uname) else val
                                if not unidad:
                                    unidad = uname
                                break
                        except (ValueError, TypeError):
                            pass

            if total == 0.0 and i_ci is not None and i_cf is not None:
                parcial_sum = 0.0
                for ci in range(i_ci, i_cf + 1):
                    if ci < len(row) and row[ci]:
                        try:
                            parcial_sum += float(row[ci])
                        except (ValueError, TypeError):
                            pass
                # Entero si la unidad es de conteo
                if es_unidad_contable(unidad):
                    total = int(parcial_sum)
                else:
                    total = parcial_sum

            entry = {
                "falla": nombre,
                "tipo": str(safe(i_tipo)).strip().upper(),
                "codigo": cod,
                "unidad": unidad_canonica_vizir(cod, unidad),
                "gravedad": min(grav, 3),
                "total": total,
                "extension": ext,
                "extension_raw": ext,
                "archivo_progresiva": str(safe(i_arch)).strip(),
                "abscisa_ini": str(safe(i_abs_i)).strip(),
                "abscisa_fin": str(safe(i_abs_f)).strip(),
                "observaciones": str(safe(i_obs)).strip(),
            }
            self.datos_procesados.append(completar_progresiva_entry(entry))

        if not self.datos_procesados:
            messagebox.showwarning("Aviso",
                                   "No se encontraron registros válidos.", parent=self)
            return

        # Enable correction buttons
        for btn in self._correction_btns.values():
            btn.config(state="normal")

        self._update_summary()

        self.lbl_status.config(
            text=f"✅ {len(self.datos_procesados)} registros procesados.",
            fg=C["green"])

    # ──────────── SUMMARY ────────────

    def _update_summary(self):
        """Populate summary and confirm button after processing."""
        for w in self.summary_frame.winfo_children():
            w.destroy()
        for w in self.btn_confirm_frame.winfo_children():
            w.destroy()

        if not self.datos_procesados:
            return

        self.summary_frame.pack(fill="x", pady=(10, 0))
        self.btn_confirm_frame.pack(fill="x", pady=(10, 0))

        all_fallas = {**FALLAS_TIPO_A, **FALLAS_TIPO_B}
        datos_con_falla = [d for d in self.datos_procesados if not d.get("sin_fallas")]
        n_sin_fallas = len(self.datos_procesados) - len(datos_con_falla)
        n = len(self.datos_procesados)
        n_a = sum(1 for d in datos_con_falla if d.get("tipo") == "A")
        n_b = sum(1 for d in datos_con_falla if d.get("tipo") == "B")
        unique_names = set(d["falla"] for d in datos_con_falla)
        unique_gravs = sorted(set(d["gravedad"] for d in datos_con_falla))
        unique_codes = set(d.get("codigo", "") for d in datos_con_falla
                           if d.get("codigo"))

        # Summary title
        tk.Label(self.summary_frame, text="📋  Resumen de Datos Procesados",
                 font=("Segoe UI", 10, "bold"), fg=C["accent"],
                 bg=C["bg_card"]).pack(anchor="w", pady=(6, 4))

        summary_txt = f"Total: {n} registros   |   Tipo A: {n_a}   |   Tipo B: {n_b}"
        if n_sin_fallas:
            summary_txt += f"   |   Sin fallas: {n_sin_fallas}"
        tk.Label(self.summary_frame,
                 text=summary_txt,
                 font=("Segoe UI", 10), fg=C["txt_primary"],
                 bg=C["bg_card"]).pack(anchor="w", pady=1)
        tk.Label(self.summary_frame,
                 text=f"Nombres únicos: {len(unique_names)}   |   "
                      f"Códigos: {', '.join(sorted(unique_codes)) or '—'}   |   "
                      f"Gravedades: {', '.join(str(g) for g in unique_gravs)}",
                 font=("Segoe UI", 10), fg=C["txt_second"],
                 bg=C["bg_card"]).pack(anchor="w", pady=1)

        # Warnings (excluir registros sin fallas)
        unmatched = [nm for nm in unique_names
                     if not any(d.get("codigo") in all_fallas
                                for d in datos_con_falla if d["falla"] == nm)]
        if unmatched:
            tk.Label(self.summary_frame,
                     text=f"⚠  {len(unmatched)} nombre(s) sin código VIZIR válido — "
                          f"Use 'Corregir Nombres'.",
                     font=("Segoe UI", 10, "bold"), fg=C["orange"],
                     bg=C["bg_card"]).pack(anchor="w", pady=(4, 0))

        valid_gravs = all(1 <= g <= 3 for g in unique_gravs) if unique_gravs else True
        if not valid_gravs:
            tk.Label(self.summary_frame,
                     text="⚠  Gravedades fuera de rango (1–3) — Use 'Corregir Gravedad'.",
                     font=("Segoe UI", 10, "bold"), fg=C["orange"],
                     bg=C["bg_card"]).pack(anchor="w", pady=(4, 0))

        # Confirm button
        make_button(self.btn_confirm_frame, "✅  Confirmar e Importar",
                    self._confirm, style="green").pack(side="left")

    # ──────────── CORRECTIONS ────────────

    def _open_name_correction(self):
        if not self.datos_procesados:
            messagebox.showwarning("Aviso", "Procese los datos primero.", parent=self)
            return
        dlg = None
        try:
            dlg = VIZIRNameCorrectionSubDialog(self, self.datos_procesados)
            self.wait_window(dlg)
            if dlg.result is not None:
                self.datos_procesados = dlg.result
                self._update_summary()
        except Exception as e:
            try:
                if dlg is not None and dlg.winfo_exists():
                    dlg.destroy()
            except Exception:
                pass
            messagebox.showerror("CorrecciÃ³n de nombres",
                                 f"No se pudo abrir la correcciÃ³n de nombres.\n\n{e}",
                                 parent=self)

    def _open_grav_correction(self):
        if not self.datos_procesados:
            messagebox.showwarning("Aviso", "Procese los datos primero.", parent=self)
            return
        dlg = None
        try:
            dlg = VIZIRGravCorrectionSubDialog(self, self.datos_procesados)
            self.wait_window(dlg)
            if dlg.result is not None:
                self.datos_procesados = dlg.result
                self._update_summary()
        except Exception as e:
            try:
                if dlg is not None and dlg.winfo_exists():
                    dlg.destroy()
            except Exception:
                pass
            messagebox.showerror("CorrecciÃ³n de gravedad",
                                 f"No se pudo abrir la correcciÃ³n de gravedad.\n\n{e}",
                                 parent=self)

    # ──────────── CONFIRM / CANCEL ────────────

    def _confirm(self):
        fallas_tipo_a = [d for d in self.datos_procesados if d["tipo"] == "A"]
        fallas_tipo_b = [d for d in self.datos_procesados if d["tipo"] == "B"]

        self.result = {
            "datos_procesados": self.datos_procesados,
            "fallas_tipo_a": fallas_tipo_a,
            "fallas_tipo_b": fallas_tipo_b,
            "nombre_map": {},
            "headers": self.headers,
            "raw_data": self.raw_data,
            "wb": self.wb,
        }
        self._cleanup()
        self.destroy()

    def _cancel(self):
        self.result = None
        self._cleanup()
        self.destroy()

    def _cleanup(self):
        try:
            self._canvas.unbind_all("<MouseWheel>")
            self._canvas.unbind_all("<Shift-MouseWheel>")
        except Exception:
            pass
        self.grab_release()


# ══════════════════════════════════════════════
#  SUB-DIÁLOGO: CORRECCIÓN DE NOMBRES
# ══════════════════════════════════════════════

class VIZIRNameCorrectionSubDialog(tk.Toplevel):
    """Sub-ventana para corregir nombres de fallas agrupados por nombre único."""

    # ── Autodetección de nombre por coincidencia ──
    @staticmethod
    def _best_match(name, code, catalog):
        """Busca la mejor coincidencia entre el nombre/código del Excel y el catálogo.
        Retorna (opción formateada 'COD — nombre', porcentaje) o (None, 0)."""

        # 1) Match exacto por código
        if code and code in catalog:
            return f"{code} — {catalog[code]['nombre']}", 100

        name_lower = name.lower().strip()

        # 2) Match exacto por nombre completo
        for cod, info in catalog.items():
            if info["nombre"].lower() == name_lower:
                return f"{cod} — {info['nombre']}", 100

        # 3) Match por palabras clave con cálculo de porcentaje
        stopwords = {"de", "del", "la", "las", "los", "el", "en", "por", "o", "y", "a"}
        name_words = [w for w in name_lower.split() if w not in stopwords and len(w) > 2]

        if not name_words:
            return None, 0

        best_cod = None
        best_pct = 0

        for cod, info in catalog.items():
            cat_lower = info["nombre"].lower()
            cat_words = [w for w in cat_lower.split() if w not in stopwords and len(w) > 2]

            if not cat_words:
                continue

            # Contar palabras del Excel que coinciden con el catálogo
            matched_excel = 0
            for nw in name_words:
                for cw in cat_words:
                    if nw in cw or cw in nw:
                        matched_excel += 1
                        break

            # Contar palabras del catálogo que coinciden con el Excel
            matched_cat = 0
            for cw in cat_words:
                for nw in name_words:
                    if cw in nw or nw in cw:
                        matched_cat += 1
                        break

            # Porcentaje bidireccional: promedio de ambas direcciones
            pct_excel = (matched_excel / len(name_words)) * 100 if name_words else 0
            pct_cat = (matched_cat / len(cat_words)) * 100 if cat_words else 0
            pct = (pct_excel + pct_cat) / 2

            # Bonus: si uno contiene al otro completamente
            if name_lower in cat_lower or cat_lower in name_lower:
                pct = min(pct + 30, 100)

            if pct > best_pct:
                best_pct = pct
                best_cod = cod

        # Exigir al menos 20% de coincidencia
        if best_pct >= 20 and best_cod:
            return f"{best_cod} — {catalog[best_cod]['nombre']}", round(best_pct)

        return None, 0

    def __init__(self, parent, datos):
        super().__init__(parent)
        self.title("⚠  Corrección de Nombres de Falla")
        self.geometry("880x520")
        self.minsize(750, 400)
        self.configure(bg=C["bg_main"])
        self.transient(parent)
        self.grab_set()

        self.datos = [dict(d) for d in datos]
        self.result = None
        self.name_vars = {}

        self._build_ui()
        self.protocol("WM_DELETE_WINDOW", self._cancel)
        self.focus_force()

    def _build_ui(self):
        # Title
        title_f = tk.Frame(self, bg=C["bg_header"], height=40)
        title_f.pack(fill="x")
        title_f.pack_propagate(False)
        tk.Label(title_f, text="⚠  CORRECCIÓN DE NOMBRES DE FALLA",
                 font=("Segoe UI", 12, "bold"),
                 fg=C["txt_primary"], bg=C["bg_header"]).pack(side="left", padx=15)

        tk.Label(self,
                 text="Nombres autodetectados por coincidencia con el catálogo VIZIR. "
                      "Verifique y corrija si es necesario.",
                 font=("Segoe UI", 9), fg=C["txt_dim"],
                 bg=C["bg_main"]).pack(anchor="w", padx=15, pady=(8, 4))

        # Scrollable table
        outer = tk.Frame(self, bg=C["bg_main"])
        outer.pack(fill="both", expand=True, padx=10, pady=4)

        canvas = tk.Canvas(outer, bg=C["bg_main"], highlightthickness=0)
        vsb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        inner = tk.Frame(canvas, bg=C["bg_main"])
        canvas.create_window((0, 0), window=inner, anchor="nw", tags="inner")
        inner.bind("<Configure>",
                   lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
                    lambda e: canvas.itemconfig("inner", width=e.width))
        canvas.bind_all("<MouseWheel>",
                        lambda e: canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))
        self._canvas = canvas

        # Header
        hdr = tk.Frame(inner, bg=C["bg_header"])
        hdr.pack(fill="x", pady=(0, 2))
        for txt, w in [("Nombre Actual", 26), ("Cód.", 5), ("×", 3),
                        ("% Coinc.", 8), ("Nuevo Nombre / Código VIZIR", 38)]:
            tk.Label(hdr, text=txt, font=("Segoe UI", 9, "bold"),
                     fg=C["accent"], bg=C["bg_header"], width=w,
                     anchor="w").pack(side="left", padx=3, pady=4)

        # Build catalog options
        all_fallas = {**FALLAS_TIPO_A, **FALLAS_TIPO_B}
        name_options = ["(Sin cambio)"] + \
            [f"{k} — {v['nombre']}" for k, v in FALLAS_TIPO_A.items()] + \
            [f"{k} — {v['nombre']}" for k, v in FALLAS_TIPO_B.items()]

        # Group by unique name
        unique_names = {}
        for d in self.datos:
            key = d["falla"]
            if key not in unique_names:
                unique_names[key] = {"codigo": d.get("codigo", ""), "count": 0}
            unique_names[key]["count"] += 1

        auto_count = 0
        for i, (name, info) in enumerate(unique_names.items()):
            matched = info["codigo"] in all_fallas
            bg = C["bg_row_even"] if i % 2 == 0 else C["bg_row_odd"]
            rf = tk.Frame(inner, bg=bg)
            rf.pack(fill="x")

            # Autodetect best match with percentage
            auto_match, match_pct = self._best_match(name, info["codigo"], all_fallas)
            if matched:
                auto_val = "(Sin cambio)"
                st_icon = "✅"
                fg_name = C["txt_primary"]
                match_pct = 100
            elif auto_match:
                auto_val = auto_match
                st_icon = "🔄"
                fg_name = C["gold"]
                auto_count += 1
            else:
                auto_val = "(Sin cambio)"
                st_icon = "⚠ "
                fg_name = C["orange"]
                match_pct = 0

            # Color del porcentaje según nivel
            if match_pct >= 80:
                pct_fg = C["green"]
            elif match_pct >= 50:
                pct_fg = C["gold"]
            elif match_pct > 0:
                pct_fg = C["orange"]
            else:
                pct_fg = C["red"]

            tk.Label(rf, text=f"{st_icon} {name[:26]}", font=("Segoe UI", 9),
                     fg=fg_name, bg=bg, width=26,
                     anchor="w").pack(side="left", padx=3, pady=3)
            tk.Label(rf, text=info["codigo"] or "?", font=("Segoe UI", 9, "bold"),
                     fg=C["txt_second"], bg=bg, width=5,
                     anchor="w").pack(side="left", padx=3)
            tk.Label(rf, text=str(info["count"]), font=("Segoe UI", 9),
                     fg=C["txt_dim"], bg=bg, width=3,
                     anchor="center").pack(side="left", padx=3)

            # Porcentaje de coincidencia
            pct_text = f"{match_pct}%" if match_pct > 0 else "—"
            tk.Label(rf, text=pct_text, font=("Segoe UI", 9, "bold"),
                     fg=pct_fg, bg=bg, width=8,
                     anchor="center").pack(side="left", padx=3)

            var = tk.StringVar(value=auto_val)
            cb = make_combo(rf, var, name_options, width=38)
            cb.pack(side="left", padx=3, pady=3)
            self.name_vars[name] = var

        # Info label about auto-detection
        if auto_count > 0:
            tk.Label(self,
                     text=f"🔄 Se autodetectaron {auto_count} coincidencia(s). "
                          f"Verifique antes de aplicar.",
                     font=("Segoe UI", 10, "bold"), fg=C["gold"],
                     bg=C["bg_main"]).pack(anchor="w", padx=15, pady=(4, 0))

        # Buttons
        btn_f = tk.Frame(self, bg=C["bg_main"])
        btn_f.pack(fill="x", padx=10, pady=(6, 12))
        make_button(btn_f, "✅  Aplicar Correcciones",
                    self._apply, style="green").pack(side="left", padx=4)
        make_button(btn_f, "✖  Cancelar",
                    self._cancel, style="danger").pack(side="left", padx=4)

    def _apply(self):
        changed = 0
        for orig_name, var in self.name_vars.items():
            val = var.get()
            if val != "(Sin cambio)" and " — " in val:
                parts = val.split(" — ", 1)
                new_cod = parts[0].strip()
                new_name = parts[1].strip()
                new_tipo = ""
                if new_cod in FALLAS_TIPO_A:
                    new_tipo = "A"
                elif new_cod in FALLAS_TIPO_B:
                    new_tipo = "B"
                for d in self.datos:
                    if d["falla"] == orig_name:
                        d["codigo"] = new_cod
                        d["falla"] = new_name
                        if new_tipo:
                            d["tipo"] = new_tipo
                        changed += 1
        self.result = self.datos
        self._cleanup()
        self.destroy()
        if changed:
            messagebox.showinfo("Correcciones",
                                f"Se corrigieron {changed} registro(s).")

    def _cancel(self):
        self.result = None
        self._cleanup()
        self.destroy()

    def _cleanup(self):
        try:
            self._canvas.unbind_all("<MouseWheel>")
        except Exception:
            pass
        self.grab_release()


# ══════════════════════════════════════════════
#  SUB-DIÁLOGO: CORRECCIÓN DE GRAVEDAD
# ══════════════════════════════════════════════

class VIZIRGravCorrectionSubDialog(tk.Toplevel):
    """Sub-ventana para corregir gravedades agrupadas por nombre de falla."""

    # Mapeo de texto de gravedad a valor numérico
    _GRAV_KEYWORDS = {
        "bajo": 1, "baja": 1, "leve": 1, "low": 1, "ligero": 1, "ligera": 1,
        "medio": 2, "media": 2, "regular": 2, "moderado": 2, "moderada": 2,
        "medium": 2, "mod": 2,
        "alto": 3, "alta": 3, "severo": 3, "severa": 3, "high": 3, "grave": 3,
    }

    @staticmethod
    def _parse_grav_int(raw_val):
        if raw_val is None:
            return None
        try:
            return int(float(str(raw_val).strip().replace(",", ".")))
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _autodetect_grav(raw_val):
        """Intenta convertir un valor de gravedad crudo a 1, 2 o 3.
        Retorna (valor_corregido, porcentaje_coincidencia) o (None, 0).
        Acepta: números, texto descriptivo (bajo/medio/alto), rangos."""
        if isinstance(raw_val, (int, float)):
            v = int(raw_val)
            if 1 <= v <= 3:
                return v, 100
            # Clamp valores fuera de rango — menor confianza
            return max(1, min(v, 3)), 60

        s = str(raw_val).strip().lower()

        # Intentar como número
        try:
            v = int(float(s))
            if 1 <= v <= 3:
                return v, 100
            return max(1, min(v, 3)), 60
        except (ValueError, TypeError):
            pass

        # Buscar por palabras clave — coincidencia exacta de keyword
        for keyword, val in VIZIRGravCorrectionSubDialog._GRAV_KEYWORDS.items():
            if keyword == s:
                return val, 100
        # Buscar por coincidencia parcial
        for keyword, val in VIZIRGravCorrectionSubDialog._GRAV_KEYWORDS.items():
            if keyword in s:
                return val, 75

        return None, 0

    def __init__(self, parent, datos):
        super().__init__(parent)
        self.title("⚠  Corrección de Gravedad")
        self.geometry("780x480")
        self.minsize(650, 380)
        self.configure(bg=C["bg_main"])
        self.transient(parent)
        self._parent = parent
        self._canvas = None

        self.datos = [dict(d) for d in datos]
        self.result = None
        self.grav_vars = []
        self.grav_groups = []

        self._build_ui()
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self._cancel)
        self.focus_force()

    def _build_grav_groups(self):
        groups = {}
        for idx, d in enumerate(self.datos):
            falla = str(d.get("falla", "") or "")
            codigo = str(d.get("codigo", "") or "")
            grav_raw = d.get("gravedad")
            key = (falla, codigo, grav_raw)
            if key not in groups:
                groups[key] = {
                    "falla": falla,
                    "codigo": codigo,
                    "gravedad": grav_raw,
                    "indices": [],
                }
            groups[key]["indices"].append(idx)
        return list(groups.values())

    def _build_ui(self):
        # Title
        title_f = tk.Frame(self, bg=C["bg_header"], height=40)
        title_f.pack(fill="x")
        title_f.pack_propagate(False)
        tk.Label(title_f, text="⚠  CORRECCIÓN DE GRAVEDAD",
                 font=("Segoe UI", 12, "bold"),
                 fg=C["txt_primary"], bg=C["bg_header"]).pack(side="left", padx=15)

        tk.Label(self,
                 text="Gravedades autodetectadas. Valores válidos: 1 (Bajo), 2 (Regular), 3 (Alto).",
                 font=("Segoe UI", 9), fg=C["txt_dim"],
                 bg=C["bg_main"]).pack(anchor="w", padx=15, pady=(8, 4))

        # Scrollable table
        outer = tk.Frame(self, bg=C["bg_main"])
        outer.pack(fill="both", expand=True, padx=10, pady=4)

        canvas = tk.Canvas(outer, bg=C["bg_main"], highlightthickness=0)
        vsb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        inner = tk.Frame(canvas, bg=C["bg_main"])
        canvas.create_window((0, 0), window=inner, anchor="nw", tags="inner")
        inner.bind("<Configure>",
                   lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
                    lambda e: canvas.itemconfig("inner", width=e.width))
        def _mw(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        canvas.bind("<MouseWheel>", _mw)
        inner.bind("<MouseWheel>", _mw)
        self._canvas = canvas

        # Header
        hdr = tk.Frame(inner, bg=C["bg_header"])
        hdr.pack(fill="x", pady=(0, 2))
        for txt, w in [("#", 4), ("Nombre", 24), ("Código", 6), ("Regs", 6),
                        ("Grav. Actual", 11), ("% Coinc.", 8),
                        ("Nueva Grav.", 11)]:
            tk.Label(hdr, text=txt, font=("Segoe UI", 9, "bold"),
                     fg=C["accent"], bg=C["bg_header"], width=w,
                     anchor="w").pack(side="left", padx=3, pady=4)

        grav_options = ["(Sin cambio)", "1", "2", "3"]
        auto_count = 0
        self.grav_groups = self._build_grav_groups()

        for i, grp in enumerate(self.grav_groups):
            bg = C["bg_row_even"] if i % 2 == 0 else C["bg_row_odd"]
            rf = tk.Frame(inner, bg=bg)
            rf.pack(fill="x")

            grav_raw = grp.get("gravedad")
            grav_num = self._parse_grav_int(grav_raw)
            grav_valid = grav_num is not None and 1 <= grav_num <= 3

            # Auto-detect correction with percentage
            auto_val = "(Sin cambio)"
            match_pct = 100 if grav_valid else 0
            if not grav_valid:
                detected, match_pct = self._autodetect_grav(grav_raw)
                if detected is not None:
                    auto_val = str(detected)
                    auto_count += 1
                else:
                    match_pct = 0

            fg_grav = C["green"] if grav_valid else C["red"]
            st_icon = "✅" if grav_valid else ("🔄" if auto_val != "(Sin cambio)" else "⚠ ")

            # Color del porcentaje según nivel
            if match_pct >= 80:
                pct_fg = C["green"]
            elif match_pct >= 50:
                pct_fg = C["gold"]
            elif match_pct > 0:
                pct_fg = C["orange"]
            else:
                pct_fg = C["red"]

            tk.Label(rf, text=str(i + 1), font=("Segoe UI", 9),
                     fg=C["txt_dim"], bg=bg, width=4).pack(side="left", padx=3, pady=2)
            tk.Label(rf, text=grp["falla"][:24], font=("Segoe UI", 9),
                     fg=C["txt_primary"], bg=bg, width=24,
                     anchor="w").pack(side="left", padx=3)
            tk.Label(rf, text=grp.get("codigo", ""), font=("Segoe UI", 9, "bold"),
                     fg=C["txt_second"], bg=bg, width=6,
                     anchor="w").pack(side="left", padx=3)
            tk.Label(rf, text=str(len(grp["indices"])), font=("Segoe UI", 9),
                     fg=C["txt_dim"], bg=bg, width=6,
                     anchor="center").pack(side="left", padx=3)
            tk.Label(rf, text=f"{st_icon} {grav_raw}", font=("Segoe UI", 9, "bold"),
                     fg=fg_grav, bg=bg, width=11,
                     anchor="center").pack(side="left", padx=3)

            # Porcentaje de coincidencia
            pct_text = f"{match_pct}%" if match_pct > 0 else "—"
            tk.Label(rf, text=pct_text, font=("Segoe UI", 9, "bold"),
                     fg=pct_fg, bg=bg, width=8,
                     anchor="center").pack(side="left", padx=3)

            var = tk.StringVar(value=auto_val)
            cb = make_combo(rf, var, grav_options, width=11)
            cb.pack(side="left", padx=3, pady=2)
            self.grav_vars.append({"indices": grp["indices"], "var": var})

        # Info about auto-detection
        if auto_count > 0:
            tk.Label(self,
                     text=f"🔄 Se autocorrigieron {auto_count} gravedad(es) fuera de rango. "
                          f"Verifique antes de aplicar.",
                     font=("Segoe UI", 10, "bold"), fg=C["gold"],
                     bg=C["bg_main"]).pack(anchor="w", padx=15, pady=(4, 0))

        # Buttons
        btn_f = tk.Frame(self, bg=C["bg_main"])
        btn_f.pack(fill="x", padx=10, pady=(6, 12))
        make_button(btn_f, "✅  Aplicar Correcciones",
                    self._apply, style="green").pack(side="left", padx=4)
        make_button(btn_f, "✖  Cancelar",
                    self._cancel, style="danger").pack(side="left", padx=4)

    def _apply(self):
        changed = 0
        for rw in self.grav_vars:
            vg = rw["var"].get()
            if vg != "(Sin cambio)":
                try:
                    nuevo = int(vg)
                    for idx in rw["indices"]:
                        self.datos[idx]["gravedad"] = nuevo
                        changed += 1
                except ValueError:
                    pass
        self.result = self.datos
        self._cleanup()
        self.destroy()
        if changed:
            messagebox.showinfo("Correcciones",
                                f"Se corrigieron {changed} gravedad(es).",
                                parent=self._parent)

    def _cancel(self):
        self.result = None
        self._cleanup()
        self.destroy()

    def _cleanup(self):
        try:
            if self._canvas is not None:
                self._canvas.unbind("<MouseWheel>")
        except Exception:
            pass
        try:
            self.grab_release()
        except Exception:
            pass


# ══════════════════════════════════════════════
#  CLASE PRINCIPAL
# ══════════════════════════════════════════════

class VIZIRApp:
    def __init__(self, root):
        self.root = root
        self.root.title("VIZIR — Análisis de Pavimentos Flexibles")
        self.root.geometry("1340x850")
        self.root.minsize(1100, 700)
        self.root.configure(bg=C["bg_main"])

        # Data vars
        self.filepath = None
        self.wb = None
        self.headers = []
        self.raw_data = []
        self.datos_procesados_base = []
        self.datos_procesados = []
        self.fallas_tipo_a_data = []
        self.fallas_tipo_b_data = []
        self.resultados = {}
        self.resultados_bloques = []
        self.resultados_tramos = []
        self.nombre_map = {}
        self.v_resultado_sel = tk.StringVar()
        self.resultado_selector_map = {}
        self.resultado_selector_widgets = []
        self.resultado_is_widgets = []

        # Project info vars
        self.var_proyecto = tk.StringVar()
        self.var_via = tk.StringVar()
        self.var_evaluador = tk.StringVar()
        self.var_fecha = tk.StringVar()
        self.var_prog_ini = tk.StringVar()
        self.var_prog_fin = tk.StringVar()
        self.var_tramo = tk.StringVar()
        self.var_um = tk.StringVar()
        self.var_largo = tk.DoubleVar(value=0.0)
        self.var_ancho = tk.DoubleVar(value=0.0)
        self.var_area = tk.DoubleVar(value=650.0)
        self.prog_ctx = {"modo": "todo", "desc": "Todo el archivo", "ini": None, "fin": None}
        self.area_calculo_activa = self._get_float_var(self.var_area, 0.0)


        s = ttk.Style()
        s.theme_use("clam")
        self.root.option_add("*TCombobox*Listbox.background", C["bg_input"])
        self.root.option_add("*TCombobox*Listbox.foreground", C["txt_primary"])
        self.root.option_add("*TCombobox*Listbox.selectBackground", C["accent2"])
        self.root.option_add("*TCombobox*Listbox.selectForeground", C["white"])
        self.root.option_add("*TCombobox*Listbox.font", "{Segoe UI} 10")
        s.configure("Dark.TNotebook", background=C["bg_main"],
                    borderwidth=0, tabmargins=[10, 8, 10, 0])
        # Tabs inactivas: pequeñas y tenues
        s.configure("Dark.TNotebook.Tab",
                    background=C["bg_panel"],
                    foreground=C["txt_second"],
                    font=("Segoe UI", 10),
                    padding=[18, 8],
                    borderwidth=0)
        # Tab activa: grande, resaltante, bold
        s.map("Dark.TNotebook.Tab",
              background=[("selected", C["bg_input"]),
                          ("active", C["bg_card"]),
                          ("!selected", C["bg_panel"])],
              foreground=[("selected", C["txt_primary"]),
                          ("active", C["txt_primary"]),
                          ("!selected", C["txt_second"])])
        s.configure("Mac.TCombobox",
                    fieldbackground=C["bg_input"],
                    background=C["bg_input"],
                    foreground=C["txt_primary"],
                    bordercolor=C["border"],
                    arrowcolor=C["accent"],
                    lightcolor=C["bg_input"],
                    darkcolor=C["bg_input"],
                    relief="flat")
        s.map("Mac.TCombobox",
              fieldbackground=[("readonly", C["bg_input"])],
              background=[("readonly", C["bg_input"])],
              foreground=[("readonly", C["txt_primary"])],
              arrowcolor=[("active", C["accent2"]), ("readonly", C["accent"])])
        for scrollbar_style in (
            "Vertical.TScrollbar",
            "Horizontal.TScrollbar",
            "Dark.Vertical.TScrollbar",
            "Dark.Horizontal.TScrollbar",
        ):
            s.configure(scrollbar_style,
                        background=C["bg_panel"],
                        troughcolor=C["bg_main"],
                        bordercolor=C["bg_main"],
                        arrowcolor=C["txt_dim"],
                        darkcolor=C["bg_panel"],
                        lightcolor=C["bg_panel"],
                        relief="flat",
                        borderwidth=0,
                        arrowsize=10)
            s.map(scrollbar_style,
                  background=[("active", C["border_hl"])])

        self._build_header()
        self._build_notebook()

    # ──────────────── HEADER ────────────────
    def _build_header(self):
        hdr = tk.Frame(self.root, bg=C["bg_header"], height=78)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)

        tk.Label(hdr, text="🛣️  SISTEMA DE ANÁLISIS VIZIR",
                 font=("Segoe UI", 15, "bold"),
                 fg=C["txt_primary"], bg=C["bg_header"]).pack(side="left", padx=20)
        tk.Label(hdr, text="Evaluación Superficial de Pavimentos Flexibles",
                 font=("Segoe UI", 10),
                 fg=C["txt_second"], bg=C["bg_header"]).pack(side="left", padx=10)
        right_header = tk.Frame(hdr, bg=C["bg_header"])
        right_header.pack(side="right", fill="y", padx=(10, 18))
        tk.Label(right_header, text="Bach. Miguel Bernardino Quispe Arias  |  Bach. Briza Edith Catachura Aycaya",
                 font=("Segoe UI", 9),
                 fg=C["txt_second"], bg=C["bg_header"]).pack(anchor="e", pady=(7, 2))

        quick_result = tk.Frame(right_header, bg=C["bg_header"])
        quick_result.pack(anchor="e", pady=(2, 6))
        tk.Label(quick_result, text="Resultado:",
                 font=("Segoe UI", 10),
                 fg=C["txt_second"], bg=C["bg_header"]).pack(side="left", padx=(0, 6))
        self.cmb_header_resultado_sel = make_combo(quick_result, self.v_resultado_sel, [], width=30)
        self.cmb_header_resultado_sel.pack(side="left")
        self.cmb_header_resultado_sel.bind("<<ComboboxSelected>>", lambda _e: self._on_resultado_selected())
        self.resultado_selector_widgets.append(self.cmb_header_resultado_sel)
        self._register_resultado_is_badge(quick_result, side="left")

    # ──────────────── NOTEBOOK ────────────────
    def _build_notebook(self):
        self.nb = ttk.Notebook(self.root, style="Dark.TNotebook")
        self.nb.pack(fill="both", expand=True, padx=0, pady=0)

        tabs = [
            ("  Entrada de Datos  ", self._build_tab_entrada),
            ("  Cálculos VIZIR  ",   self._build_tab_calculos),
            ("  Procedimientos  ",   self._build_tab_procedimientos),
            ("  Resultados  ",       self._build_tab_resultados),
        ]
        self.tab_frames = {}
        for label, builder in tabs:
            f = tk.Frame(self.nb, bg=C["bg_main"])
            self.nb.add(f, text=label)
            self.tab_frames[label.strip()] = f
            builder(f)

    def _register_resultado_selector(self, parent, width=36):
        tk.Label(parent, text="Resultado:",
                 font=("Segoe UI", 10),
                 fg=C["txt_second"], bg=C["bg_main"]).pack(side="left", padx=(0, 6))
        combo = make_combo(parent, self.v_resultado_sel, [], width=width)
        combo.pack(side="left")
        combo.bind("<<ComboboxSelected>>", lambda _e: self._on_resultado_selected())
        self.resultado_selector_widgets = [
            w for w in self.resultado_selector_widgets
            if w.winfo_exists()
        ]
        self.resultado_selector_widgets.append(combo)
        self._register_resultado_is_badge(parent)
        return combo

    def _register_resultado_is_badge(self, parent, side="right"):
        self.resultado_is_widgets = [
            w for w in self.resultado_is_widgets
            if w.winfo_exists()
        ]
        badge = tk.Label(parent, text="Is: --",
                         font=("Segoe UI", 11, "bold"),
                         fg=C["txt_dim"], bg=C["bg_card"],
                         padx=12, pady=4)
        badge.pack(side=side, padx=(10, 0))
        self.resultado_is_widgets.append(badge)
        return badge

    def _update_resultado_is_badges(self):
        self.resultado_is_widgets = [
            w for w in self.resultado_is_widgets
            if w.winfo_exists()
        ]
        resultado = self.resultado_selector_map.get(self.v_resultado_sel.get())
        if not resultado:
            text = "Is: --"
            fg = C["txt_dim"]
            bg = C["bg_card"]
        else:
            is_final = resultado.get("is_final", "")
            clasif = str(resultado.get("clasificacion", "") or "").upper()
            text = f"Is: {is_final}  {clasif}".strip()
            try:
                _clasif, fg, bg = clasificar_is(int(is_final))
            except (TypeError, ValueError):
                fg = C["txt_dim"]
                bg = C["bg_card"]

        for badge in self.resultado_is_widgets:
            badge.config(text=text, fg=fg, bg=bg)

    def _descripcion_gravedad_representativa(self, g_promedio):
        if g_promedio < 1.5:
            return "G < 1.5, por lo tanto toma el valor 1"
        if g_promedio < 2.5:
            return "1.5 <= G < 2.5, por lo tanto toma el valor 2"
        return "G >= 2.5, por lo tanto toma el valor 3"

    def _categoria_extension_info(self, extension_pct):
        if extension_pct <= 0:
            return 0, "Sin extension"
        base = extension_pct / 100 if extension_pct <= 100 else extension_pct
        return ext_col(base), ext_label(base)

    def _gravedad_formula_lines(self, grav_rep, grav_prom, s1, s2, s3, fallas=None):
        if fallas:
            terminos_num = []
            terminos_den = []
            numerador = 0.0
            total = 0.0
            for f in fallas:
                cantidad = float(f.get("total", 0.0) or 0.0)
                gravedad = int(f.get("gravedad", 0) or 0)
                if cantidad <= 0 or gravedad <= 0:
                    continue
                terminos_num.append(f"{gravedad}*{cantidad:.2f}")
                terminos_den.append(f"{cantidad:.2f}")
                numerador += gravedad * cantidad
                total += cantidad
            if total > 0:
                return [
                    f"Numerador = {' + '.join(terminos_num)} = {numerador:.2f}",
                    f"Denominador = {' + '.join(terminos_den)} = {total:.2f}",
                    f"G = ({' + '.join(terminos_num)}) / ({' + '.join(terminos_den)}) = {grav_prom:.2f}",
                    f"Gravedad representativa = {grav_rep} porque {self._descripcion_gravedad_representativa(grav_prom)}.",
                ]

        total = s1 + s2 + s3
        if total == 0:
            return [
                "No se registraron fallas en este grupo.",
                "G = 0.00 y la gravedad representativa es 0.",
            ]
        numerador = s1 * 1 + s2 * 2 + s3 * 3
        return [
            f"Suma por gravedad: S1 = {s1:.2f}, S2 = {s2:.2f}, S3 = {s3:.2f}, Total = {total:.2f}",
            f"G = (1*{s1:.2f} + 2*{s2:.2f} + 3*{s3:.2f}) / {total:.2f} = {numerador:.2f}/{total:.2f} = {grav_prom:.2f}",
            f"Gravedad representativa = {grav_rep} porque {self._descripcion_gravedad_representativa(grav_prom)}.",
        ]

    def _render_resultado_context(self, parent, resultado):
        card, body = make_card(parent, "Resultado Seleccionado")
        card.pack(fill="x", pady=6)

        scope = self.v_resultado_sel.get() or f"Actual: {resultado.get('progresiva', 'Todo el archivo')}"
        lineas = [
            f"Vista activa: {scope}",
            f"Progresiva analizada: {resultado.get('progresiva', 'Todo el archivo')}",
            f"Registros analizados: {resultado.get('n_reg', 0)}   |   Tramos: {resultado.get('n_tramos', 0)}",
            f"Area de muestra: {resultado.get('area', 0.0):.2f} m2   |   U.M.: {self.var_um.get() or '-'}",
        ]
        for linea in lineas:
            tk.Label(body, text=linea,
                     font=("Segoe UI", 10),
                     fg=C["txt_second"], bg=C["bg_card"]).pack(anchor="w", pady=2)

    def _agrupar_resumen_fallas(self, datos):
        grupos = {}
        for d in datos:
            if d.get("sin_fallas"):
                key = ("-", "-", "SIN FALLAS", "-", "")
                grupos.setdefault(key, {
                    "codigo": "-",
                    "tipo": "-",
                    "falla": "SIN FALLAS",
                    "gravedad": "-",
                    "unidad": "",
                    "total": 0.0,
                    "extension_pct": 0.0,
                })
                continue
            key = (
                d.get("codigo", ""),
                d.get("tipo", ""),
                d.get("falla", ""),
                d.get("gravedad", ""),
                d.get("unidad", ""),
            )
            grp = grupos.setdefault(key, {
                "codigo": d.get("codigo", ""),
                "tipo": d.get("tipo", ""),
                "falla": d.get("falla", ""),
                "gravedad": d.get("gravedad", ""),
                "unidad": d.get("unidad", ""),
                "total": 0.0,
                "extension_pct": 0.0,
            })
            grp["total"] += float(d.get("total", 0.0) or 0.0)
            grp["extension_pct"] += d["extension"] * 100 if d["extension"] <= 1 else d["extension"]
        resumen = list(grupos.values())
        resumen.sort(key=lambda x: (
            x["tipo"] == "-",
            str(x["tipo"]),
            str(x["codigo"]),
            str(x["gravedad"]),
            str(x["falla"]),
        ))
        return resumen

    def _procedimiento_sections(self, r, scope_label=None):
        scope = scope_label or self.v_resultado_sel.get() or f"Actual: {r.get('progresiva', 'Todo el archivo')}"
        ec_f, lbl_f = self._categoria_extension_info(r.get("ext_fis", 0.0))
        ec_d, lbl_d = self._categoria_extension_info(r.get("ext_def", 0.0))
        ec_r, lbl_r = self._categoria_extension_info(r.get("ext_rep", 0.0))
        fisuras = [d for d in r.get("fallas_a", []) if d.get("codigo") in ("FPC", "FLT")]
        deforms = [d for d in r.get("fallas_a", []) if d.get("codigo") in ("AH", "DL", "DT")]
        reps = [d for d in r.get("fallas_a", []) if d.get("codigo") == "B"]
        sections = [
            ("Resultado Visible", [
                f"Vista activa: {scope}",
                f"Unidad de Muestreo: {self.var_um.get() or '-'}",
                f"Area de muestra: {r.get('area', 0.0):.2f} m2",
                f"Progresiva analizada: {r.get('progresiva', 'Todo el archivo')}",
                f"Registros analizados: {r.get('n_reg', 0)}   |   Tramos: {r.get('n_tramos', 0)}",
            ]),
            ("Formulas Aplicadas", [
                "Promedio ponderado de gravedad: G = sum(gravedad_i * cantidad_i) / sum(cantidad_i)",
                "Cada termino usa la gravedad de la fila multiplicada por la cantidad total de esa falla.",
                "Gravedad representativa: G < 1.5 -> 1, 1.5 <= G < 2.5 -> 2, G >= 2.5 -> 3",
                "Extension acumulada: E = suma de extensiones del grupo",
                "If = TABLA_IF[Gravedad representativa][Categoria de extension]",
                "Id = TABLA_ID[Gravedad representativa][Categoria de extension]",
                "Primer Is = TABLA_IS_PRIMER[Id][If]",
                "Correccion = TABLA_CORRECCION[Gravedad representativa de B][Categoria de extension de B]",
                "Is final = min(Primer Is + Correccion, 7)",
            ]),
            ("Fisuracion (If)", [
                *self._gravedad_formula_lines(
                    r.get("grav_fis", 0), r.get("grav_prom_fis", 0.0),
                    r.get("s1_fis", 0), r.get("s2_fis", 0), r.get("s3_fis", 0),
                    fallas=fisuras
                ),
                f"Extension acumulada = {r.get('ext_fis', 0.0):.2f}% -> categoria {ec_f} ({lbl_f})",
                f"If = TABLA_IF[G={r.get('grav_fis', 0)}][E={ec_f}] = {r.get('if', 0)}",
            ]),
            ("Deformacion (Id)", [
                *self._gravedad_formula_lines(
                    r.get("grav_def", 0), r.get("grav_prom_def", 0.0),
                    r.get("s1_def", 0), r.get("s2_def", 0), r.get("s3_def", 0),
                    fallas=deforms
                ),
                f"Extension acumulada = {r.get('ext_def', 0.0):.2f}% -> categoria {ec_d} ({lbl_d})",
                f"Id = TABLA_ID[G={r.get('grav_def', 0)}][E={ec_d}] = {r.get('id', 0)}",
            ]),
            ("Primer Is", [
                f"Primer Is = TABLA_IS_PRIMER[Id={r.get('id', 0)}][If={r.get('if', 0)}] = {r.get('is_primer', 0)}",
            ]),
        ]
        if r.get("s1_rep", 0) + r.get("s2_rep", 0) + r.get("s3_rep", 0) > 0:
            rep_lines = [
                *self._gravedad_formula_lines(
                    r.get("grav_rep", 0), r.get("grav_prom_rep", 0.0),
                    r.get("s1_rep", 0), r.get("s2_rep", 0), r.get("s3_rep", 0),
                    fallas=reps
                ),
                f"Extension acumulada = {r.get('ext_rep', 0.0):.2f}% -> categoria {ec_r} ({lbl_r})",
                f"Correccion = TABLA_CORRECCION[G={r.get('grav_rep', 0)}][E={ec_r}] = {r.get('correccion', 0)}",
            ]
        else:
            rep_lines = ["No se registraron baches y parches (B), por lo tanto la correccion es 0."]
        sections.append(("Correccion por Reparacion", rep_lines))
        sections.append(("Resultado Final", [
            f"Is final = min({r.get('is_primer', 0)} + {r.get('correccion', 0)}, 7) = {r.get('is_final', 0)}",
            f"Clasificacion = {r.get('clasificacion', '').upper()}",
        ]))
        return sections

    # ══════════════════════════════════════════════
    #  TAB 1: ENTRADA DE DATOS
    # ══════════════════════════════════════════════
    def _build_tab_entrada(self, parent):
        sf = ScrollFrame(parent, bg=C["bg_main"])
        sf.pack(fill="both", expand=True)
        frame = sf.inner

        # ── SECCIÓN: Información del Proyecto ──
        card, body = make_card(frame, "Información del Proyecto")
        card.pack(fill="x", padx=15, pady=(12, 6))

        fields_left = [
            ("Proyecto:", self.var_proyecto, 55),
            ("Nombre de la vía:", self.var_via, 55),
            ("Evaluado por:", self.var_evaluador, 55),
            ("Fecha:", self.var_fecha, 20),
        ]
        fields_right = [
            ("Progresiva inicial:", self.var_prog_ini, 18),
            ("Progresiva final:", self.var_prog_fin, 18),
            ("Tramo:", self.var_tramo, 10),
            ("Unidad de muestreo:", self.var_um, 10),
        ]

        grid = tk.Frame(body, bg=C["bg_card"])
        grid.pack(fill="x")

        for i, (lbl, var, w) in enumerate(fields_left):
            tk.Label(grid, text=lbl, font=("Segoe UI", 10),
                     fg=C["txt_second"], bg=C["bg_card"]).grid(
                row=i, column=0, sticky="e", padx=(0, 8), pady=4)
            make_entry(grid, textvariable=var, width=w).grid(
                row=i, column=1, sticky="w", pady=4)

        for i, (lbl, var, w) in enumerate(fields_right):
            tk.Label(grid, text=lbl, font=("Segoe UI", 10),
                     fg=C["txt_second"], bg=C["bg_card"]).grid(
                row=i, column=2, sticky="e", padx=(25, 8), pady=4)
            make_entry(grid, textvariable=var, width=w).grid(
                row=i, column=3, sticky="w", pady=4)

        # Largo, Ancho y Área de muestra
        dim_row = tk.Frame(body, bg=C["bg_card"])
        dim_row.pack(fill="x", pady=(8, 0))

        tk.Label(dim_row, text="Largo (m):",
                 font=("Segoe UI", 10), fg=C["txt_second"],
                 bg=C["bg_card"]).pack(side="left")
        e_largo = make_entry(dim_row, textvariable=self.var_largo, width=10)
        e_largo.pack(side="left", padx=(6, 15))

        tk.Label(dim_row, text="Ancho (m):",
                 font=("Segoe UI", 10), fg=C["txt_second"],
                 bg=C["bg_card"]).pack(side="left")
        e_ancho = make_entry(dim_row, textvariable=self.var_ancho, width=10)
        e_ancho.pack(side="left", padx=(6, 15))

        tk.Label(dim_row, text="Área (m²):",
                 font=("Segoe UI", 10, "bold"),
                 fg=C["gold"], bg=C["bg_card"]).pack(side="left")
        make_entry(dim_row, textvariable=self.var_area, width=12).pack(
            side="left", padx=6)

        def _auto_calc_area(*_):
            l = self._get_float_var(self.var_largo, 0.0)
            a = self._get_float_var(self.var_ancho, 0.0)
            area_calc = round(l * a, 2) if l > 0 and a > 0 else 0.0
            self.var_area.set(area_calc)
        self.var_largo.trace_add("write", _auto_calc_area)
        self.var_ancho.trace_add("write", _auto_calc_area)

        # ── SECCIÓN: Importar archivo Excel ──
        card2, body2 = make_card(frame, "Importar Datos desde Excel")
        card2.pack(fill="x", padx=15, pady=6)

        row_file = tk.Frame(body2, bg=C["bg_card"])
        row_file.pack(fill="x", pady=4)

        self.lbl_file = tk.Label(row_file, text="Ningún archivo seleccionado",
                                  font=("Segoe UI", 10), fg=C["txt_dim"],
                                  bg=C["bg_card"])
        self.lbl_file.pack(side="left", padx=(0, 15))
        make_button(row_file, "📂  Abrir Excel e Importar", self._open_file).pack(side="left")
        make_button(row_file, "🔧  Mapear Columnas",
                    self._open_column_map, style="dim").pack(side="left", padx=8)

        # Import status
        self.lbl_import_status = tk.Label(body2, text="",
                                           font=("Segoe UI", 10),
                                           fg=C["txt_dim"], bg=C["bg_card"])
        self.lbl_import_status.pack(fill="x", pady=(4, 0))

        card_prog, body_prog = make_card(frame, "Filtro por Progresivas")
        card_prog.pack(fill="x", padx=15, pady=6)
        self.panel_prog = PanelProgresivasVIZIR(body_prog, self._apply_progressive_filter)
        self.panel_prog.pack(fill="x")

        # ── SECCIÓN: Vista Previa de Datos Procesados ──
        card3, body3 = make_card(frame, "Vista Previa de Datos Procesados")
        card3.pack(fill="x", padx=15, pady=6)

        prev_frame = tk.Frame(body3, bg=C["bg_panel"])
        prev_frame.pack(fill="both", expand=True, pady=4)
        self.tree_processed = ttk.Treeview(prev_frame, height=8)
        p_cols = ("Código", "Nombre", "Tipo", "Gravedad", "Total", "Extensión", "Unidad")
        p_widths = (70, 220, 50, 70, 90, 90, 60)
        configure_dark_treeview(self.tree_processed, p_cols, p_widths)
        configure_dark_treeview(
            self.tree_processed,
            ("Codigo", "Nombre", "Tipo", "Gravedad", "Total", "Extension", "Prog. Ini", "Prog. Fin", "Is Final"),
            (70, 220, 50, 70, 90, 90, 90, 90, 70),
        )
        vsb_p = ttk.Scrollbar(prev_frame, orient="vertical",
                               command=self.tree_processed.yview)
        hsb_p = ttk.Scrollbar(prev_frame, orient="horizontal",
                               command=self.tree_processed.xview)
        self.tree_processed.configure(yscrollcommand=vsb_p.set,
                                       xscrollcommand=hsb_p.set)
        self.tree_processed.grid(row=0, column=0, sticky="nsew")
        vsb_p.grid(row=0, column=1, sticky="ns")
        hsb_p.grid(row=1, column=0, sticky="ew")
        prev_frame.grid_columnconfigure(0, weight=1)
        prev_frame.grid_rowconfigure(0, weight=1)

        self.lbl_preview_count = tk.Label(body3, text="Sin datos importados.",
                                           font=("Segoe UI", 9),
                                           fg=C["txt_dim"], bg=C["bg_card"])
        self.lbl_preview_count.pack(anchor="w", pady=(4, 0))


        # ── SECCIÓN: Catálogo de Fallas VIZIR ──
        card5, body5 = make_card(frame, "Catálogo de Fallas VIZIR")
        card5.pack(fill="x", padx=15, pady=(6, 15))

        cat_nb = ttk.Notebook(body5, style="Dark.TNotebook")
        cat_nb.pack(fill="both", expand=True)

        # Tab A
        fa = tk.Frame(cat_nb, bg=C["bg_card"])
        cat_nb.add(fa, text="  Tipo A — Estructurales  ")
        tree_a = ttk.Treeview(fa, height=6)
        cols_a = ("Código", "Nombre", "Unidad")
        configure_dark_treeview(tree_a, cols_a, [70, 380, 60])
        for i, (cod, info) in enumerate(FALLAS_TIPO_A.items()):
            tag = "even" if i % 2 == 0 else "odd"
            tree_a.insert("", "end", values=(cod, info["nombre"], info["unidad"]),
                          tags=(tag,))
        tree_a.pack(fill="both", expand=True, padx=4, pady=4)

        # Tab B
        fb = tk.Frame(cat_nb, bg=C["bg_card"])
        cat_nb.add(fb, text="  Tipo B — Funcionales  ")
        tree_b = ttk.Treeview(fb, height=8)
        configure_dark_treeview(tree_b, cols_a, [70, 380, 60])
        for i, (cod, info) in enumerate(FALLAS_TIPO_B.items()):
            tag = "even" if i % 2 == 0 else "odd"
            tree_b.insert("", "end", values=(cod, info["nombre"], info["unidad"]),
                          tags=(tag,))
        tree_b.pack(fill="both", expand=True, padx=4, pady=4)

    # ── helpers tab entrada ──

    def _open_file(self):
        fp = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")])
        if not fp:
            return
        self.filepath = fp
        self.lbl_file.config(text=os.path.basename(fp), fg=C["accent"])

        # Open unified import + mapping dialog
        dlg = VIZIRColumnMapDialog(self.root, filepath=fp)
        self.root.wait_window(dlg)

        if dlg.result is not None:
            self._apply_import_result(dlg.result)
        else:
            self.lbl_import_status.config(
                text="Importación cancelada.", fg=C["orange"])

    def _apply_import_result(self, result):
        """Apply result from VIZIRColumnMapDialog to main app."""
        self.datos_procesados_base = [completar_progresiva_entry(d) for d in result["datos_procesados"]]
        self.datos_procesados = []
        self.fallas_tipo_a_data = []
        self.fallas_tipo_b_data = []
        self.resultados = {}
        self.resultados_bloques = []
        self.resultados_tramos = []
        self.nombre_map = result.get("nombre_map", {})
        self.headers = result.get("headers", [])
        self.raw_data = result.get("raw_data", [])
        self.wb = result.get("wb")

        # Try auto-read project info
        if self.wb:
            ws = self.wb[self.wb.sheetnames[0]]
            self._try_read_project_info(ws)

        self.panel_prog.resetear()
        self.panel_prog.configurar(self.datos_procesados_base)
        self._apply_progressive_filter()

        # Calculate extension if missing
        area = self.area_calculo_activa
        for d in self.datos_procesados:
            if d["extension"] == 0.0 and d["total"] > 0 and area > 0:
                d["extension"] = d["total"] / area

        n_a = len(self.fallas_tipo_a_data)
        n_b = len(self.fallas_tipo_b_data)
        total = len(self.datos_procesados)

        self.lbl_import_status.config(
            text=f"Activos {total} registros  |  Tipo A: {n_a}  |  Tipo B: {n_b}  |  {self.prog_ctx.get('desc', 'Todo el archivo')}",
            fg=C["green"])

        self._populate_preview()

        messagebox.showinfo("Importación Exitosa",
                            f"Total: {total} registros\n"
                            f"Tipo A: {n_a}\nTipo B: {n_b}")

    def _build_prog_ctx(self, modo, ini, fin, datos, desc=None):
        return {
            "modo": modo,
            "ini": ini,
            "fin": fin,
            "desc": desc or (f"{metros_a_prog(ini)} - {metros_a_prog(fin)}" if ini is not None and fin is not None else "Todo el archivo"),
            "n_reg": len(datos),
            "n_tramos": len({d.get("prog_tramo", "") for d in datos if d.get("prog_tramo")}),
        }

    def _get_float_var(self, var, default=0.0):
        try:
            value = var.get()
        except (tk.TclError, ValueError):
            try:
                value = self.root.tk.globalgetvar(var._name)
            except (AttributeError, tk.TclError):
                return default
        txt = str(value).strip().replace(",", ".")
        if not txt:
            return default
        try:
            return float(txt)
        except (TypeError, ValueError):
            return default

    def _get_area_calculo(self, ctx):
        area_base = max(0.0, self._get_float_var(self.var_area, 0.0))
        largo_base = max(0.0, self._get_float_var(self.var_largo, 0.0))
        ancho = max(0.0, self._get_float_var(self.var_ancho, 0.0))
        if area_base > 0:
            return round(area_base, 2)
        if largo_base > 0 and ancho > 0:
            return round(largo_base * ancho, 2)
        return 0.0

    def _build_active_dataset(self, datos, ctx):
        area = self._get_area_calculo(ctx)
        activos = []
        for d in datos:
            item = completar_progresiva_entry(d)
            total_raw = float(item.get("total", 0.0) or 0.0)
            # Preservar entero para unidades de conteo (ojo de pescado, etc.)
            unidad = item.get("unidad", "")
            if unidad and es_unidad_contable(unidad):
                total = int(total_raw)
            else:
                total = total_raw
            item["total"] = total
            ext_raw = float(item.get("extension_raw", item.get("extension", 0.0)) or 0.0)
            if ext_raw > 0:
                item["extension"] = ext_raw
            elif total > 0 and area > 0:
                item["extension"] = total / area
            else:
                item["extension"] = 0.0
            activos.append(item)
        fallas_a = [d for d in activos if d.get("tipo") == "A"]
        fallas_b = [d for d in activos if d.get("tipo") == "B"]
        return activos, fallas_a, fallas_b, area

    def _set_active_dataset(self, datos, ctx):
        activos, fallas_a, fallas_b, area = self._build_active_dataset(datos, ctx)
        self.datos_procesados = activos
        self.fallas_tipo_a_data = fallas_a
        self.fallas_tipo_b_data = fallas_b
        self.prog_ctx = ctx
        self.area_calculo_activa = area

    def _calcular_resultados_por_progresiva(self, datos):
        grupos = defaultdict(list)
        for item in datos:
            ini = item.get("prog_ini_m")
            fin = item.get("prog_fin_m")
            if ini is None or fin is None:
                continue
            grupos[(ini, fin)].append(item)

        detalle = []
        for ini, fin in sorted(grupos):
            datos_tramo = grupos[(ini, fin)]
            ctx = self._build_prog_ctx("tramo_detalle", ini, fin, datos_tramo)
            res = self._calcular_resultados_vizir(datos_tramo, ctx)
            res["selector_label"] = f"Tramo: {res['progresiva']}"
            detalle.append(res)
        return detalle

    def _render_calc_detalle_progresiva(self, parent, resultados_detalle):
        if not resultados_detalle:
            return

        card, body = make_card(parent, "Detalle por Progresiva")
        card.pack(fill="x", pady=(0, 12))

        tk.Label(
            body,
            text=(
                f"Progresivas detectadas en el rango activo: {len(resultados_detalle)}"
            ),
            font=("Segoe UI", 10),
            fg=C["txt_second"],
            bg=C["bg_card"],
        ).pack(anchor="w", pady=(0, 6))

        tree = ttk.Treeview(body, height=min(12, len(resultados_detalle)))
        cols = ("Progresiva", "Regs", "Area", "If", "Id", "Primer Is", "Corr.", "Is", "Clasificacion")
        widths = (220, 55, 80, 45, 45, 70, 55, 45, 110)
        configure_dark_treeview(tree, cols, widths)

        row_map = {}
        for i, res in enumerate(resultados_detalle):
            tag = "even" if i % 2 == 0 else "odd"
            item_id = f"tramo_det_{i}"
            row_map[item_id] = res.get("selector_label", "")
            tree.insert(
                "",
                "end",
                iid=item_id,
                values=(
                    res["progresiva"],
                    res["n_reg"],
                    f"{res['area']:.2f}",
                    res["if"],
                    res["id"],
                    res["is_primer"],
                    res["correccion"],
                    res["is_final"],
                    res["clasificacion"],
                ),
                tags=(tag,),
            )

        def _on_select(_event):
            sel = tree.selection()
            if not sel:
                return
            label = row_map.get(sel[0], "")
            if label:
                self.v_resultado_sel.set(label)
                self._on_resultado_selected()

        tree.bind("<<TreeviewSelect>>", _on_select)
        tree.pack(fill="x", pady=4)

    def _calcular_resultados_vizir(self, datos, ctx):
        activos, fallas_a, fallas_b, area = self._build_active_dataset(datos, ctx)

        fisuras = [d for d in fallas_a if d["codigo"] in ("FPC", "FLT")]
        if fisuras:
            grav_rep_f, grav_prom_f, s1_f, s2_f, s3_f = gravedad_ponderada(fisuras)
            total_ext_f = sum((f["extension"] * 100 if f["extension"] <= 1 else f["extension"]) for f in fisuras)
            ec_f = ext_col(total_ext_f / 100 if total_ext_f <= 100 else total_ext_f)
            if_val = TABLA_IF.get(min(grav_rep_f, 3), {}).get(ec_f, 0)
        else:
            grav_rep_f = 0
            grav_prom_f = 0.0
            s1_f = s2_f = s3_f = 0
            total_ext_f = 0
            if_val = 0

        deforms = [d for d in fallas_a if d["codigo"] in ("AH", "DL", "DT")]
        if deforms:
            grav_rep_d, grav_prom_d, s1_d, s2_d, s3_d = gravedad_ponderada(deforms)
            total_ext_d = sum((f["extension"] * 100 if f["extension"] <= 1 else f["extension"]) for f in deforms)
            ec_d = ext_col(total_ext_d / 100 if total_ext_d <= 100 else total_ext_d)
            id_val = TABLA_ID.get(min(grav_rep_d, 3), {}).get(ec_d, 0)
        else:
            grav_rep_d = 0
            grav_prom_d = 0.0
            s1_d = s2_d = s3_d = 0
            total_ext_d = 0
            id_val = 0

        is_primer = TABLA_IS_PRIMER.get(min(id_val, 5), {}).get(min(if_val, 5), 1)

        reps = [d for d in fallas_a if d["codigo"] == "B"]
        if reps:
            grav_rep_r, grav_prom_r, s1_r, s2_r, s3_r = gravedad_ponderada(reps)
            total_ext_r = sum((r["extension"] * 100 if r["extension"] <= 1 else r["extension"]) for r in reps)
            ec_r = ext_col(total_ext_r / 100 if total_ext_r <= 100 else total_ext_r)
            corr = TABLA_CORRECCION.get(min(grav_rep_r, 3), {}).get(ec_r, 0)
        else:
            grav_rep_r = 0
            grav_prom_r = 0.0
            s1_r = s2_r = s3_r = 0
            total_ext_r = 0
            corr = 0

        is_final = min(is_primer + corr, 7)
        clasif, clr, _ = clasificar_is(is_final)
        return {
            "if": if_val,
            "id": id_val,
            "is_primer": is_primer,
            "correccion": corr,
            "is_final": is_final,
            "clasificacion": clasif,
            "color": clr,
            "datos": activos,
            "fallas_a": fallas_a,
            "fallas_b": fallas_b,
            "grav_fis": grav_rep_f,
            "ext_fis": total_ext_f,
            "grav_prom_fis": grav_prom_f,
            "s1_fis": s1_f, "s2_fis": s2_f, "s3_fis": s3_f,
            "grav_def": grav_rep_d,
            "ext_def": total_ext_d,
            "grav_prom_def": grav_prom_d,
            "s1_def": s1_d, "s2_def": s2_d, "s3_def": s3_d,
            "grav_rep": grav_rep_r,
            "ext_rep": total_ext_r,
            "grav_prom_rep": grav_prom_r,
            "s1_rep": s1_r, "s2_rep": s2_r, "s3_rep": s3_r,
            "area": area,
            "progresiva": ctx.get("desc", "Todo el archivo"),
            "prog_ini": ctx.get("ini"),
            "prog_fin": ctx.get("fin"),
            "n_reg": len(activos),
            "n_tramos": len({d.get("prog_tramo", "") for d in activos if d.get("prog_tramo")}),
        }

    def _apply_progressive_filter(self):
        if not self.datos_procesados_base:
            self.datos_procesados = []
            self.fallas_tipo_a_data = []
            self.fallas_tipo_b_data = []
            self.resultados_bloques = []
            self.resultados_tramos = []
            self.area_calculo_activa = self._get_float_var(self.var_area, 0.0)
            self.prog_ctx = {"modo": "todo", "desc": "Todo el archivo", "ini": None, "fin": None}
            if hasattr(self, "calc_batch_frame"):
                for w in self.calc_batch_frame.winfo_children():
                    w.destroy()
            self._populate_preview()
            return

        try:
            cfg = self.panel_prog.obtener_config()
            if cfg["modo"] == "todo":
                datos = AnalizadorProgresivasVIZIR.preparar_registros(self.datos_procesados_base)
                ctx = self._build_prog_ctx("todo", None, None, datos, "Todo el archivo")
            elif cfg["modo"] == "manual":
                if cfg["ini"] is None or cfg["fin"] is None:
                    raise ValueError("Ingrese la progresiva inicial y final para el rango manual.")
                datos = AnalizadorProgresivasVIZIR.filtrar_registros(self.datos_procesados_base, cfg["ini"], cfg["fin"])
                if not datos:
                    raise ValueError("No se encontraron registros dentro del rango de progresivas indicado.")
                ctx = self._build_prog_ctx("manual", cfg["ini"], cfg["fin"], datos)
            elif cfg["modo"] == "tramos":
                tramo_ini = cfg.get("tramo_ini")
                tramo_fin = cfg.get("tramo_fin")
                if not tramo_ini or not tramo_fin:
                    raise ValueError("Seleccione el tramo inicial y final detectados.")
                ini = tramo_ini["ini"]
                fin = tramo_fin["fin"]
                if fin <= ini:
                    raise ValueError("El tramo final debe quedar despues del tramo inicial.")
                datos = AnalizadorProgresivasVIZIR.filtrar_registros(self.datos_procesados_base, ini, fin)
                if not datos:
                    raise ValueError("No se encontraron registros dentro del rango de tramos seleccionado.")
                ctx = self._build_prog_ctx("tramos", ini, fin, datos)
            else:
                if not cfg["bloque"]:
                    self.panel_prog._detectar_bloques(silent=True)
                    cfg = self.panel_prog.obtener_config()
                bloque = cfg.get("bloque")
                if not bloque:
                    raise ValueError("Detecte y seleccione un bloque automatico antes de aplicar.")
                datos = AnalizadorProgresivasVIZIR.filtrar_registros(self.datos_procesados_base, bloque["ini"], bloque["fin"])
                if not datos:
                    raise ValueError("El bloque seleccionado no contiene registros analizables.")
                ctx = self._build_prog_ctx("bloques", bloque["ini"], bloque["fin"], datos, bloque["label"])

            self._set_active_dataset(datos, ctx)

            total_base = len(self.datos_procesados_base)
            total_act = len(self.datos_procesados)
            self.lbl_import_status.config(
                text=f"Activos {total_act} de {total_base} registros  |  Progresiva: {ctx['desc']}",
                fg=C["green"] if total_act else C["orange"],
            )
            self.resultados_bloques = []
            self.resultados_tramos = []
            if hasattr(self, "calc_batch_frame"):
                for w in self.calc_batch_frame.winfo_children():
                    w.destroy()
            self._populate_preview()
        except Exception as e:
            messagebox.showerror("Progresivas", str(e), parent=self.root)

    def _populate_preview(self):
        """Fill the processed data preview Treeview with all records."""
        self.tree_processed.delete(*self.tree_processed.get_children())
        resultados_por_tramo = self._calcular_resultados_preview_por_tramo()
        self._populate_is_preview(resultados_por_tramo)
        for i, d in enumerate(self.datos_procesados):
            resultado_tramo = resultados_por_tramo.get((d.get("prog_ini_m"), d.get("prog_fin_m")))
            is_final = resultado_tramo["is_final"] if resultado_tramo else ""
            if d.get("sin_fallas"):
                self.tree_processed.insert("", "end", values=(
                    "-", d["falla"][:35], "-", "-", "-", "-",
                    d.get("abscisa_ini", ""),
                    d.get("abscisa_fin", ""),
                    is_final,
                ), tags=("sin_fallas",))
                continue
            ep = d["extension"] * 100 if d["extension"] <= 1 else d["extension"]
            tag = "even" if i % 2 == 0 else "odd"
            self.tree_processed.insert("", "end", values=(
                d.get("codigo", ""),
                d["falla"][:35],
                d.get("tipo", ""),
                d["gravedad"],
                fmt_total_str(d['total'], d.get('unidad', '')),
                f"{ep:.1f}%",
                d.get("abscisa_ini", ""),
                d.get("abscisa_fin", ""),
                is_final,
            ), tags=(tag,))
        n = len(self.datos_procesados)
        total_base = len(self.datos_procesados_base) if self.datos_procesados_base else n
        txt = f"{n} registros activos"
        if total_base != n:
            txt += f" de {total_base} importados"
        if self.prog_ctx.get("modo") != "todo":
            txt += f"  |  {self.prog_ctx.get('desc', '')}"
        txt += f"  |  Area de calculo: {self.area_calculo_activa:.2f} m²"
        self.lbl_preview_count.config(
            text=txt,
            fg=C["accent"] if n else C["txt_dim"])

    def _calcular_resultados_preview_por_tramo(self):
        """Return VIZIR results grouped by detected progressive range."""
        grupos = defaultdict(list)
        for item in self.datos_procesados:
            ini = item.get("prog_ini_m")
            fin = item.get("prog_fin_m")
            if ini is None or fin is None:
                continue
            grupos[(ini, fin)].append(item)

        resultados_por_tramo = {}
        for ini, fin in sorted(grupos):
            datos_tramo = grupos[(ini, fin)]
            ctx = self._build_prog_ctx("preview_tramo", ini, fin, datos_tramo)
            resultados_por_tramo[(ini, fin)] = self._calcular_resultados_vizir(datos_tramo, ctx)
        return resultados_por_tramo

    def _populate_is_preview(self, resultados_por_tramo):
        if not hasattr(self, "tree_is_preview"):
            return
        if not self.tree_is_preview.winfo_exists():
            return

        self.tree_is_preview.delete(*self.tree_is_preview.get_children())
        for res in resultados_por_tramo.values():
            is_final = res.get("is_final", 0)
            if is_final <= 2:
                tag = "is_bueno"
            elif is_final <= 4:
                tag = "is_marginal"
            else:
                tag = "is_deficiente"
            self.tree_is_preview.insert(
                "",
                "end",
                values=(
                    res.get("progresiva", ""),
                    res.get("n_reg", 0),
                    f"{res.get('area', 0.0):.2f}",
                    res.get("if", 0),
                    res.get("id", 0),
                    is_final,
                    res.get("clasificacion", ""),
                ),
                tags=(tag,),
            )

        if hasattr(self, "lbl_is_preview_count") and self.lbl_is_preview_count.winfo_exists():
            n = len(resultados_por_tramo)
            self.lbl_is_preview_count.config(
                text=f"{n} progresivas con Is final calculado." if n else "Sin progresivas calculadas.",
                fg=C["accent"] if n else C["txt_dim"],
            )

    def _open_column_map(self):
        """Abre el diálogo de mapeo de columnas para correcciones posteriores."""
        if not self.datos_procesados_base:
            messagebox.showwarning("Aviso",
                                   "Primero importe un archivo Excel.")
            return
        dlg = VIZIRColumnMapDialog(self.root,
                                    headers=self.headers,
                                    datos_procesados=self.datos_procesados_base)
        self.root.wait_window(dlg)
        if dlg.result is not None:
            self._apply_import_result(dlg.result)

    def _open_correction_dialog(self):
        if not self.datos_procesados_base:
            messagebox.showwarning("Aviso", "No hay datos importados para corregir.")
            return
        dlg = VIZIRCorrectionDialog(self.root, self.datos_procesados_base)
        self.root.wait_window(dlg)
        if dlg.result is not None:
            self.datos_procesados_base = [completar_progresiva_entry(d) for d in dlg.result]
            self.panel_prog.configurar(self.datos_procesados_base)
            self._apply_progressive_filter()

    def _try_read_project_info(self, ws):
        """Intenta leer información del proyecto desde el Excel"""
        def cell(r, c):
            v = ws.cell(row=r, column=c).value
            return str(v).strip() if v else ""

        for r in range(1, min(15, ws.max_row + 1)):
            for c in range(1, min(6, ws.max_column + 1)):
                v = cell(r, c)
                vl = v.lower()
                if "proyecto" in vl and c < ws.max_column:
                    val = cell(r, c + 1) or cell(r, c + 2)
                    if val and val.lower() != "proyecto:":
                        self.var_proyecto.set(val)
                elif "nombre de la v" in vl or "vía" in vl.replace(":", ""):
                    val = cell(r, c + 1) or cell(r, c + 2)
                    if val:
                        self.var_via.set(val)
                elif "evaluad" in vl:
                    val = cell(r, c + 1) or cell(r, c + 2)
                    if val:
                        self.var_evaluador.set(val)
                elif "fecha" in vl:
                    val = cell(r, c + 1) or cell(r, c + 2)
                    if val:
                        self.var_fecha.set(val)
                elif "progresiva ini" in vl:
                    val = cell(r, c + 1) or cell(r, c + 2)
                    if val:
                        self.var_prog_ini.set(val)
                elif "progresiva fin" in vl:
                    val = cell(r, c + 1) or cell(r, c + 2)
                    if val:
                        self.var_prog_fin.set(val)
                elif "tramo" in vl:
                    val = cell(r, c + 1) or cell(r, c + 2)
                    if val:
                        self.var_tramo.set(val)
                elif "unidad de muestr" in vl:
                    val = cell(r, c + 1) or cell(r, c + 2)
                    if val:
                        self.var_um.set(val)
                elif "rea de muestra" in vl:
                    val = cell(r, c + 1) or cell(r, c + 2)
                    if val:
                        try:
                            self.var_area.set(float(val))
                        except ValueError:
                            pass

    # ══════════════════════════════════════════════
    #  TAB 2: CÁLCULOS VIZIR
    # ══════════════════════════════════════════════
    def _build_tab_calculos(self, parent):
        sf = ScrollFrame(parent, bg=C["bg_main"])
        sf.pack(fill="both", expand=True)
        self.calc_inner = sf.inner

        top_bar = tk.Frame(self.calc_inner, bg=C["bg_main"])
        top_bar.pack(fill="x", padx=15, pady=12)
        make_button(top_bar, "Ejecutar C\u00e1lculo VIZIR",
                    self._run_calc, style="gold").pack(side="left")
        make_button(top_bar, "Calcular Todo por Bloques",
                    self._run_calc_blocks, style="dim").pack(side="left", padx=8)
        self.lbl_calc_status = tk.Label(top_bar, text="",
                                         font=("Segoe UI", 10),
                                         fg=C["txt_dim"], bg=C["bg_main"])
        self.lbl_calc_status.pack(side="left", padx=20)

        sel_bar = tk.Frame(self.calc_inner, bg=C["bg_main"])
        sel_bar.pack(fill="x", padx=15, pady=(0, 8))
        self._register_resultado_selector(sel_bar)

        self.calc_results_frame = tk.Frame(self.calc_inner, bg=C["bg_main"])
        self.calc_results_frame.pack(fill="both", expand=True, padx=15)
        self.calc_batch_frame = tk.Frame(self.calc_inner, bg=C["bg_main"])
        self.calc_batch_frame.pack(fill="x", padx=15, pady=(0, 15))

    def _run_calc(self):
        if not self.datos_procesados:
            messagebox.showwarning("Aviso", "No hay datos. Complete la Entrada de Datos.")
            return

        self._set_active_dataset(self.datos_procesados, self.prog_ctx)
        self._populate_preview()
        self.resultados_bloques = []
        self.resultados_tramos = self._calcular_resultados_por_progresiva(self.datos_procesados)
        self.resultados = self._calcular_resultados_vizir(self.datos_procesados, self.prog_ctx)

        for w in self.calc_batch_frame.winfo_children():
            w.destroy()

        self._refresh_resultado_selector(
            prefer=f"Actual: {self.resultados.get('progresiva', 'Todo el archivo')}"
        )
        self._populate_calculos()
        self._populate_procedimientos()
        self._populate_resultados()
        self.lbl_calc_status.config(text="C\u00e1lculo completado", fg=C["green"])
        return

        for w in self.calc_results_frame.winfo_children():
            w.destroy()

        area = self.area_calculo_activa
        frame = self.calc_results_frame

        # ── Resumen de fallas ──
        card_r, body_r = make_card(frame, "Resumen de Fallas Registradas")
        card_r.pack(fill="x", pady=6)

        tree_r = ttk.Treeview(body_r, height=min(8, len(self.datos_procesados)))
        cols_r = ("C\u00f3digo", "Tipo", "Falla", "Gravedad", "Total", "Extensi\u00f3n (%)", "Rango Ext.")
        widths_r = (70, 50, 220, 70, 90, 100, 100)
        configure_dark_treeview(tree_r, cols_r, widths_r)
        for i, d in enumerate(self.datos_procesados):
            if d.get("sin_fallas"):
                tree_r.insert("", "end", values=(
                    "-", "-", d["falla"][:30],
                    "-", "-", "-", "SIN FALLAS"
                ), tags=("sin_fallas",))
                continue
            ep = d["extension"] * 100 if d["extension"] <= 1 else d["extension"]
            tag = "tipo_a" if d["tipo"] == "A" else "tipo_b"
            tree_r.insert("", "end", values=(
                d["codigo"], d["tipo"], d["falla"][:30],
                d["gravedad"], fmt_total_str(d['total'], d.get('unidad', '')),
                f"{ep:.2f} %", ext_label(d["extension"])
            ), tags=(tag,))
        tree_r.pack(fill="x", pady=4)

        # ── CÁLCULO If ──
        fisuras = [d for d in self.fallas_tipo_a_data
                   if d["codigo"] in ("FPC", "FLT")]
        if fisuras:
            max_grav_f = max(f["gravedad"] for f in fisuras)
            total_ext_f = sum(
                (f["extension"] * 100 if f["extension"] <= 1 else f["extension"])
                for f in fisuras)
            ec_f = ext_col(total_ext_f / 100 if total_ext_f <= 100 else total_ext_f)
            if_val = TABLA_IF.get(min(max_grav_f, 3), {}).get(ec_f, 0)
        else:
            max_grav_f = 0
            total_ext_f = 0
            if_val = 0

        self._make_index_card(frame, "\u00cdndice de Fisuraci\u00f3n (If)",
                              fisuras, max_grav_f, total_ext_f, if_val,
                              C["accent"], "FPC, FLT", compact=True)

        # ── CÁLCULO Id ──
        deforms = [d for d in self.fallas_tipo_a_data
                   if d["codigo"] in ("AH", "DL", "DT")]
        if deforms:
            max_grav_d = max(f["gravedad"] for f in deforms)
            total_ext_d = sum(
                (f["extension"] * 100 if f["extension"] <= 1 else f["extension"])
                for f in deforms)
            ec_d = ext_col(total_ext_d / 100 if total_ext_d <= 100 else total_ext_d)
            id_val = TABLA_ID.get(min(max_grav_d, 3), {}).get(ec_d, 0)
        else:
            max_grav_d = 0
            total_ext_d = 0
            id_val = 0

        self._make_index_card(frame, "\u00cdndice de Deformaci\u00f3n (Id)",
                              deforms, max_grav_d, total_ext_d, id_val,
                              C["accent_alt"], "AH, DL, DT")

        # ── PRIMER Is ──
        is_primer = TABLA_IS_PRIMER.get(min(id_val, 5), {}).get(min(if_val, 5), 1)

        card_is, body_is = make_card(frame, "Primer Valor de Is")
        card_is.pack(fill="x", pady=6)
        tk.Label(body_is,
                 text=f"If = {if_val}   |   Id = {id_val}   ->   Primer Is = {is_primer}",
                 font=("Segoe UI", 13, "bold"),
                 fg=C["gold"], bg=C["bg_card"]).pack(anchor="w", pady=4)

        # ── CORRECCIÓN ──
        reps = [d for d in self.fallas_tipo_a_data if d["codigo"] == "B"]
        if reps:
            max_grav_r = max(r["gravedad"] for r in reps)
            total_ext_r = sum(
                (r["extension"] * 100 if r["extension"] <= 1 else r["extension"])
                for r in reps)
            ec_r = ext_col(total_ext_r / 100 if total_ext_r <= 100 else total_ext_r)
            corr = TABLA_CORRECCION.get(min(max_grav_r, 3), {}).get(ec_r, 0)
        else:
            max_grav_r = 0
            total_ext_r = 0
            corr = 0

        card_c, body_c = make_card(frame, "Correcci\u00f3n por Reparaci\u00f3n (Baches y Parches)")
        card_c.pack(fill="x", pady=6)
        if reps:
            tk.Label(body_c,
                     text=f"Gravedad: {max_grav_r}   |   Extensi\u00f3n: {total_ext_r:.2f}%   ->   Correcci\u00f3n = {corr}",
                     font=("Segoe UI", 11),
                     fg=C["txt_primary"], bg=C["bg_card"]).pack(anchor="w", pady=4)
        else:
            tk.Label(body_c, text="No se registraron baches y parches (B). Correcci\u00f3n = 0",
                     font=("Segoe UI", 11), fg=C["txt_dim"],
                     bg=C["bg_card"]).pack(anchor="w", pady=4)

        # ── Is FINAL ──
        is_final = min(is_primer + corr, 7)
        clasif, clr, clr_bg = clasificar_is(is_final)

        card_f, body_f = make_card(frame, "\u00cdNDICE DE DETERIORO SUPERFICIAL (Is) FINAL")
        card_f.pack(fill="x", pady=(10, 15))

        tk.Label(body_f,
                 text=f"Is  =  Primer Is + Correcci\u00f3n  =  {is_primer} + {corr}  =  {is_final}",
                 font=("Segoe UI", 12),
                 fg=C["txt_primary"], bg=C["bg_card"]).pack(anchor="w", pady=(4, 8))

        result_box = tk.Frame(body_f, bg=clr_bg,
                               highlightbackground=clr, highlightthickness=2)
        result_box.pack(fill="x", pady=6, ipady=12)
        tk.Label(result_box,
                 text=f"Is = {is_final}     ->     {clasif.upper()}",
                 font=("Segoe UI", 22, "bold"),
                 fg=clr, bg=clr_bg).pack()

        # Scale
        scale_f = tk.Frame(body_f, bg=C["bg_card"])
        scale_f.pack(fill="x", pady=(8, 4))
        for rng, (lab, col_s, col_bg_s) in [
            ((1, 2), ("Bueno", C["green"], C["green_bg"])),
            ((3, 4), ("Marginal", C["orange"], C["orange_bg"])),
            ((5, 6, 7), ("Deficiente", C["red"], C["red_bg"])),
        ]:
            is_active = is_final in rng
            rng_txt = f"{min(rng)}–{max(rng)}"
            marker = "  ◄" if is_active else ""
            bw = 2 if is_active else 0
            lbl = tk.Label(scale_f,
                           text=f"  {rng_txt}: {lab}{marker}  ",
                           font=("Segoe UI", 10, "bold" if is_active else ""),
                           fg=col_s, bg=col_bg_s if is_active else C["bg_card"],
                           highlightbackground=col_s, highlightthickness=bw,
                           padx=14, pady=5)
            lbl.pack(side="left", padx=4)

        self.resultados_bloques = []
        self.resultados_tramos = self._calcular_resultados_por_progresiva(self.datos_procesados)
        self._render_calc_detalle_progresiva(frame, self.resultados_tramos)

        self.resultados = self._calcular_resultados_vizir(self.datos_procesados, self.prog_ctx)
        self._refresh_resultado_selector(
            prefer=f"Actual: {self.resultados.get('progresiva', 'Todo el archivo')}"
        )

        self._populate_procedimientos()
        self._populate_resultados()
        self.lbl_calc_status.config(text="C\u00e1lculo completado", fg=C["green"])

    def _populate_calculos(self):
        for w in self.calc_results_frame.winfo_children():
            w.destroy()

        if not self._has_resultados_disponibles():
            return

        r = self._get_resultado_visible()
        if not r:
            return

        frame = self.calc_results_frame
        self._render_resultado_context(frame, r)

        card_r, body_r = make_card(frame, "Resumen de Fallas Registradas")
        card_r.pack(fill="x", pady=6)

        datos_visibles = r.get("datos", [])
        resumen_fallas = self._agrupar_resumen_fallas(datos_visibles)
        tree_wrap = tk.Frame(body_r, bg=C["bg_card"])
        tree_wrap.pack(fill="both", expand=True, pady=4)

        tree_r = ttk.Treeview(tree_wrap, height=min(10, max(len(resumen_fallas), 1)))
        cols_r = ("C\u00f3digo", "Tipo", "Falla", "Gravedad", "Total", "Extensi\u00f3n (%)", "Rango Ext.")
        widths_r = (70, 50, 220, 70, 90, 100, 100)
        configure_dark_treeview(tree_r, cols_r, widths_r)
        tree_scroll = ttk.Scrollbar(tree_wrap, orient="vertical", command=tree_r.yview, style="Dark.Vertical.TScrollbar")
        tree_r.configure(yscrollcommand=tree_scroll.set)

        for item in resumen_fallas:
            if item.get("falla") == "SIN FALLAS":
                tree_r.insert("", "end", values=("-", "-", "SIN FALLAS", "-", "-", "-", "SIN FALLAS"),
                              tags=("sin_fallas",))
                continue
            ep = item.get("extension_pct", 0.0)
            tag = "tipo_a" if item.get("tipo") == "A" else "tipo_b"
            tree_r.insert("", "end", values=(
                item.get("codigo", ""),
                item.get("tipo", ""),
                item.get("falla", "")[:30],
                item.get("gravedad", ""),
                fmt_total_str(item.get("total", 0.0), item.get("unidad", "")),
                f"{ep:.2f} %",
                ext_label(ep)
            ), tags=(tag,))
        tree_r.pack(side="left", fill="both", expand=True)
        tree_scroll.pack(side="right", fill="y")

        fisuras = [d for d in r.get("fallas_a", []) if d.get("codigo") in ("FPC", "FLT")]
        self._make_index_card(
            frame, "\u00cdndice de Fisuraci\u00f3n (If)",
            fisuras, r.get("grav_fis", 0), r.get("ext_fis", 0.0), r.get("if", 0),
            C["accent"], "FPC, FLT", compact=True,
            grav_prom=r.get("grav_prom_fis", 0.0),
            n1=r.get("s1_fis", 0), n2=r.get("s2_fis", 0), n3=r.get("s3_fis", 0)
        )

        deforms = [d for d in r.get("fallas_a", []) if d.get("codigo") in ("AH", "DL", "DT")]
        self._make_index_card(
            frame, "\u00cdndice de Deformaci\u00f3n (Id)",
            deforms, r.get("grav_def", 0), r.get("ext_def", 0.0), r.get("id", 0),
            C["accent_alt"], "AH, DL, DT",
            grav_prom=r.get("grav_prom_def", 0.0),
            n1=r.get("s1_def", 0), n2=r.get("s2_def", 0), n3=r.get("s3_def", 0)
        )

        card_is, body_is = make_card(frame, "Primer Valor de Is")
        card_is.pack(fill="x", pady=6)
        tk.Label(
            body_is,
            text=f"Primer Is = TABLA_IS_PRIMER[Id={r.get('id', 0)}][If={r.get('if', 0)}] = {r.get('is_primer', 0)}",
            font=("Segoe UI", 13, "bold"),
            fg=C["gold"], bg=C["bg_card"]
        ).pack(anchor="w", pady=4)

        reps = [d for d in r.get("fallas_a", []) if d.get("codigo") == "B"]
        card_c, body_c = make_card(frame, "Correcci\u00f3n por Reparaci\u00f3n (Baches y Parches)")
        card_c.pack(fill="x", pady=6)
        if reps:
            for line in self._gravedad_formula_lines(
                r.get("grav_rep", 0), r.get("grav_prom_rep", 0.0),
                r.get("s1_rep", 0), r.get("s2_rep", 0), r.get("s3_rep", 0),
                fallas=reps
            ):
                tk.Label(body_c, text=line,
                         font=("Segoe UI", 10),
                         fg=C["txt_second"], bg=C["bg_card"]).pack(anchor="w", pady=2)
            ec_r, lbl_r = self._categoria_extension_info(r.get("ext_rep", 0.0))
            tk.Label(
                body_c,
                text=f"Extensi\u00f3n acumulada = {r.get('ext_rep', 0.0):.2f}%  ->  categoria {ec_r} ({lbl_r})",
                font=("Segoe UI", 10),
                fg=C["txt_primary"], bg=C["bg_card"]
            ).pack(anchor="w", pady=(4, 2))
            tk.Label(
                body_c,
                text=f"Correcci\u00f3n = TABLA_CORRECCION[G={r.get('grav_rep', 0)}][E={ec_r}] = {r.get('correccion', 0)}",
                font=("Segoe UI", 11, "bold"),
                fg=C["red"], bg=C["bg_card"]
            ).pack(anchor="w", pady=(4, 2))
        else:
            tk.Label(
                body_c,
                text="No se registraron baches y parches (B). Correccion = 0.",
                font=("Segoe UI", 11),
                fg=C["txt_dim"], bg=C["bg_card"]
            ).pack(anchor="w", pady=4)

        is_final = r.get("is_final", 0)
        clasif, clr, clr_bg = clasificar_is(is_final) if is_final else ("Sin clasificar", C["txt_dim"], C["bg_input"])
        card_f, body_f = make_card(frame, "\u00cdNDICE DE DETERIORO SUPERFICIAL (Is) FINAL")
        card_f.pack(fill="x", pady=(10, 15))

        tk.Label(
            body_f,
            text=f"Is = min({r.get('is_primer', 0)} + {r.get('correccion', 0)}, 7) = {is_final}",
            font=("Segoe UI", 12),
            fg=C["txt_primary"], bg=C["bg_card"]
        ).pack(anchor="w", pady=(4, 8))

        result_box = tk.Frame(body_f, bg=clr_bg,
                               highlightbackground=clr, highlightthickness=2)
        result_box.pack(fill="x", pady=6, ipady=12)
        tk.Label(
            result_box,
            text=f"Is = {is_final}     ->     {r.get('clasificacion', clasif).upper()}",
            font=("Segoe UI", 22, "bold"),
            fg=clr, bg=clr_bg
        ).pack()

        scale_f = tk.Frame(body_f, bg=C["bg_card"])
        scale_f.pack(fill="x", pady=(8, 4))
        for rng, (lab, col_s, col_bg_s) in [
            ((1, 2), ("Bueno", C["green"], C["green_bg"])),
            ((3, 4), ("Marginal", C["orange"], C["orange_bg"])),
            ((5, 6, 7), ("Deficiente", C["red"], C["red_bg"])),
        ]:
            is_active = is_final in rng
            rng_txt = f"{min(rng)}-{max(rng)}"
            marker = "  <" if is_active else ""
            bw = 2 if is_active else 0
            tk.Label(
                scale_f,
                text=f"  {rng_txt}: {lab}{marker}  ",
                font=("Segoe UI", 10, "bold" if is_active else ""),
                fg=col_s, bg=col_bg_s if is_active else C["bg_card"],
                highlightbackground=col_s, highlightthickness=bw,
                padx=14, pady=5
            ).pack(side="left", padx=4)

        self._render_calc_detalle_progresiva(frame, self.resultados_tramos)

    def _resolver_limites_bloques(self):
        cfg = self.panel_prog.obtener_config()
        ini = cfg.get("ini")
        fin = cfg.get("fin")
        if cfg.get("modo") == "tramos":
            tramo_ini = cfg.get("tramo_ini")
            tramo_fin = cfg.get("tramo_fin")
            if tramo_ini and tramo_fin:
                ini = tramo_ini["ini"]
                fin = tramo_fin["fin"]
        return ini, fin, cfg.get("tam_bloque", 100)

    def _run_calc_blocks(self):
        if not self.datos_procesados_base:
            messagebox.showwarning("Aviso", "No hay datos importados para analizar por bloques.")
            return

        for w in self.calc_batch_frame.winfo_children():
            w.destroy()

        ini, fin, tam_bloque = self._resolver_limites_bloques()
        bloques, _ = AnalizadorProgresivasVIZIR.bloques_disponibles(
            self.datos_procesados_base, tam_bloque, ini, fin
        )
        if not bloques:
            messagebox.showwarning("Aviso", "No se detectaron bloques analizables con el rango actual.")
            return

        self.resultados_tramos = []
        self.resultados_bloques = []
        for bloque in bloques:
            datos = AnalizadorProgresivasVIZIR.filtrar_registros(
                self.datos_procesados_base, bloque["ini"], bloque["fin"]
            )
            if not datos:
                continue
            ctx = self._build_prog_ctx("bloques", bloque["ini"], bloque["fin"], datos, bloque["label"])
            resultado = self._calcular_resultados_vizir(datos, ctx)
            resultado["n_reg_bloque"] = bloque["n_reg"]
            resultado["n_tramos_bloque"] = bloque["n_tramos"]
            self.resultados_bloques.append(resultado)

        if not self.resultados_bloques:
            messagebox.showwarning("Aviso", "Los bloques detectados no contienen datos analizables.")
            return

        card, body = make_card(self.calc_batch_frame, f"Resumen por Bloques de {tam_bloque} m")
        card.pack(fill="x", pady=(0, 8))
        tk.Label(
            body,
            text=(
                f"Bloques analizados: {len(self.resultados_bloques)}  |  "
                f"Rango: {metros_a_prog(self.resultados_bloques[0]['prog_ini'])} - "
                f"{metros_a_prog(self.resultados_bloques[-1]['prog_fin'])}"
            ),
            font=("Segoe UI", 10),
            fg=C["txt_second"],
            bg=C["bg_card"],
        ).pack(anchor="w", pady=(0, 6))

        tree = ttk.Treeview(body, height=min(12, len(self.resultados_bloques)))
        cols = ("Bloque", "Regs", "Tramos", "Area", "If", "Id", "Is", "Clasificacion")
        widths = (230, 60, 70, 80, 50, 50, 50, 120)
        configure_dark_treeview(tree, cols, widths)
        row_map = {}
        for i, r_blk in enumerate(self.resultados_bloques):
            tag = "even" if i % 2 == 0 else "odd"
            item_id = f"bloque_{i}"
            row_map[item_id] = r_blk["progresiva"]
            tree.insert("", "end", iid=item_id, values=(
                r_blk["progresiva"],
                r_blk["n_reg"],
                r_blk["n_tramos"],
                f"{r_blk['area']:.2f}",
                r_blk["if"],
                r_blk["id"],
                r_blk["is_final"],
                r_blk["clasificacion"],
            ), tags=(tag,))

        def _on_block_select(_event):
            sel = tree.selection()
            if not sel:
                return
            label = row_map.get(sel[0], "")
            if label:
                self.v_resultado_sel.set(label)
                self._on_resultado_selected()

        tree.bind("<<TreeviewSelect>>", _on_block_select)
        tree.pack(fill="x", pady=4)
        self._refresh_resultado_selector(prefer=self.resultados_bloques[0].get("progresiva"))
        self._on_resultado_selected()
        self.lbl_calc_status.config(
            text=f"Calculo por bloques completado ({len(self.resultados_bloques)} bloques de {tam_bloque} m)",
            fg=C["green"],
        )

    def _make_index_card(self, parent, title, fallas, grav, ext, value, color, codes,
                         compact=False, grav_prom=0.0, n1=0, n2=0, n3=0):
        card, body = make_card(parent, title)
        card.pack(fill="x", pady=6)

        if fallas:
            if compact:
                items = []
                for f in fallas:
                    ep = f["extension"] * 100 if f["extension"] <= 1 else f["extension"]
                    items.append(f"{f['codigo']} | G:{f['gravedad']} | E:{ep:.2f}%")
                tk.Label(body,
                         text="   ||   ".join(items),
                         font=("Segoe UI", 10), fg=C["txt_second"],
                         bg=C["bg_card"], justify="left",
                         wraplength=900).pack(anchor="w")
            else:
                for f in fallas:
                    ep = f["extension"] * 100 if f["extension"] <= 1 else f["extension"]
                    tk.Label(body,
                             text=f"  -  {f['codigo']}  |  Gravedad: {f['gravedad']},  Extensi\u00f3n: {ep:.2f}%",
                             font=("Segoe UI", 10), fg=C["txt_second"],
                             bg=C["bg_card"]).pack(anchor="w")
            for line in self._gravedad_formula_lines(grav, grav_prom, n1, n2, n3, fallas=fallas):
                tk.Label(body, text=line,
                         font=("Segoe UI", 10),
                         fg=C["txt_second"], bg=C["bg_card"]).pack(anchor="w", pady=2)
            ec, lbl = self._categoria_extension_info(ext)
            tk.Label(body,
                     text=f"Extensi\u00f3n acumulada = {ext:.2f}%  ->  categoria {ec} ({lbl})",
                     font=("Segoe UI", 10), fg=C["txt_primary"],
                     bg=C["bg_card"]).pack(anchor="w", pady=(6, 2))
        else:
            tk.Label(body,
                     text=f"No se encontraron fallas ({codes}).",
                     font=("Segoe UI", 10), fg=C["txt_dim"],
                     bg=C["bg_card"]).pack(anchor="w")

        tk.Label(body,
                 text=f"->  {title.split('(')[1].split(')')[0]} = {value}",
                 font=("Segoe UI", 13, "bold"), fg=color,
                 bg=C["bg_card"]).pack(anchor="w", pady=(6, 0))

    # ══════════════════════════════════════════════
    #  TAB 3: PROCEDIMIENTOS
    # ══════════════════════════════════════════════
    def _build_tab_procedimientos(self, parent):
        sf = ScrollFrame(parent, bg=C["bg_main"])
        sf.pack(fill="both", expand=True)
        self.proc_inner = sf.inner
        proc_sel_bar = tk.Frame(self.proc_inner, bg=C["bg_main"])
        proc_sel_bar.pack(fill="x", padx=15, pady=(12, 0))
        self._register_resultado_selector(proc_sel_bar)
        self.proc_dynamic_frame = tk.Frame(self.proc_inner, bg=C["bg_main"])
        self.proc_dynamic_frame.pack(fill="x", padx=15, pady=(6, 6))

        # Static content — method steps and tables
        self._build_static_procedure_info()

    def _build_static_procedure_info(self):
        frame = self.proc_inner

        # ── Descripción del método ──
        card0, body0 = make_card(frame, "Método VIZIR — Descripción General")
        card0.pack(fill="x", padx=15, pady=(12, 6))

        desc = (
            "El método VIZIR (Visión e Inspección de Zonas e Itinerarios en Riesgo) "
            "es un sistema de evaluación visual para pavimentos flexibles desarrollado "
            "por el Laboratorio Central de Puentes y Caminos de Francia (LCPC).\n\n"
            "Clasifica los deterioros en dos categorías:\n"
            "   •  Tipo A — Degradaciones estructurales: vinculadas a la capacidad estructural.\n"
            "   •  Tipo B — Degradaciones funcionales: relacionadas con la condición superficial.\n\n"
            "El resultado final es el Índice de Deterioro Superficial (Is), "
            "que varía de 1 a 7 y clasifica el pavimento como Bueno (1–2), Marginal (3–4) o Deficiente (5–7)."
        )
        tk.Label(body0, text=desc, font=("Segoe UI", 10), fg=C["txt_second"],
                 bg=C["bg_card"], justify="left", wraplength=900).pack(anchor="w")

        # ── Pasos del procedimiento ──
        card1, body1 = make_card(frame, "Pasos del Procedimiento de Cálculo")
        card1.pack(fill="x", padx=15, pady=6)

        pasos = [
            ("1. Registrar Fallas",
             "Identificar y registrar cada deterioro con su código, tipo (A/B), gravedad (1–3) y extensión."),
            ("2. Calcular Índice de Fisuración (If)",
             "Se evalúan las fallas FPC y FLT. Se calcula G = sum(gravedad_i * cantidad_i) / sum(cantidad_i), "
             "se obtiene la gravedad representativa y luego la extensión acumulada. "
             "Con la tabla de If se obtiene el valor (0–5)."),
            ("3. Calcular Índice de Deformación (Id)",
             "Se evalúan las fallas AH, DL y DT. Se usa el mismo promedio ponderado para la gravedad representativa "
             "y la misma lógica de extensión que en If, "
             "utilizando la tabla de Id (0–5)."),
            ("4. Obtener Primer Valor de Is",
             "Con If e Id se ingresa a la tabla de primer valor de Is para obtener un valor inicial."),
            ("5. Corrección por Reparación",
             "Si existen baches y parches (B), se obtiene un factor de corrección (0 o 1) "
             "según la gravedad representativa y la extensión de las reparaciones."),
            ("6. Calcular Is Final",
             "Is = Primer Is + Corrección. El valor máximo es 7. "
             "Se clasifica: 1–2 = Bueno, 3–4 = Marginal, 5–7 = Deficiente."),
        ]

        for titulo, desc in pasos:
            pf = tk.Frame(body1, bg=C["bg_card"])
            pf.pack(fill="x", pady=4)
            tk.Label(pf, text=titulo, font=("Segoe UI", 11, "bold"),
                     fg=C["accent"], bg=C["bg_card"]).pack(anchor="w")
            tk.Label(pf, text=desc, font=("Segoe UI", 10),
                     fg=C["txt_second"], bg=C["bg_card"],
                     wraplength=850, justify="left").pack(anchor="w", padx=(16, 0), pady=(2, 6))

        # ── TABLAS ──
        self._build_table_card(frame, "Tabla: Índice de Fisuración (If)", TABLA_IF)
        self._build_table_card(frame, "Tabla: Índice de Deformación (Id)", TABLA_ID)
        self._build_is_primer_table(frame)
        self._build_table_card(frame, "Tabla: Corrección por Reparación", TABLA_CORRECCION)
        self._build_classification_card(frame)

        # ── Calculated values (dynamic) ──
        if not hasattr(self, "proc_dynamic_frame"):
            self.proc_dynamic_frame = tk.Frame(frame, bg=C["bg_main"])
            self.proc_dynamic_frame.pack(fill="x", padx=15, pady=(6, 15))

    def _build_table_card(self, parent, title, tabla):
        card, body = make_card(parent, title)
        card.pack(fill="x", padx=15, pady=6)

        # Header
        hdr = tk.Frame(body, bg=C["bg_header"])
        hdr.pack(fill="x")
        tk.Label(hdr, text="Gravedad \\ Extensión", font=("Segoe UI", 10, "bold"),
                 fg=C["accent"], bg=C["bg_header"], width=22,
                 anchor="center").grid(row=0, column=0, padx=2, pady=5)
        for j, lbl in enumerate(["0 – 10 %", "10 – 50 %", "> 50 %"]):
            tk.Label(hdr, text=lbl, font=("Segoe UI", 10, "bold"),
                     fg=C["accent"], bg=C["bg_header"], width=14,
                     anchor="center").grid(row=0, column=j+1, padx=2, pady=5)

        for g in [1, 2, 3]:
            bg = C["bg_row_even"] if g % 2 == 1 else C["bg_row_odd"]
            rf = tk.Frame(body, bg=bg)
            rf.pack(fill="x")
            grav_names = {1: "1 (Bajo)", 2: "2 (Regular)", 3: "3 (Alto)"}
            tk.Label(rf, text=grav_names[g], font=("Segoe UI", 10, "bold"),
                     fg=C["txt_primary"], bg=bg, width=22,
                     anchor="center").grid(row=0, column=0, padx=2, pady=4)
            for e in [1, 2, 3]:
                val = tabla[g][e]
                tk.Label(rf, text=str(val), font=("Segoe UI", 11, "bold"),
                         fg=C["gold"], bg=bg, width=14,
                         anchor="center").grid(row=0, column=e, padx=2, pady=4)

    def _build_is_primer_table(self, parent):
        card, body = make_card(parent, "Tabla: Primer Valor de Is (If vs Id)")
        card.pack(fill="x", padx=15, pady=6)

        hdr = tk.Frame(body, bg=C["bg_header"])
        hdr.pack(fill="x")
        tk.Label(hdr, text="Id \\ If", font=("Segoe UI", 10, "bold"),
                 fg=C["accent"], bg=C["bg_header"], width=10,
                 anchor="center").grid(row=0, column=0, padx=2, pady=5)
        for if_v in range(6):
            tk.Label(hdr, text=str(if_v), font=("Segoe UI", 10, "bold"),
                     fg=C["accent"], bg=C["bg_header"], width=7,
                     anchor="center").grid(row=0, column=if_v+1, padx=2, pady=5)

        for id_v in range(6):
            bg = C["bg_row_even"] if id_v % 2 == 0 else C["bg_row_odd"]
            rf = tk.Frame(body, bg=bg)
            rf.pack(fill="x")
            tk.Label(rf, text=str(id_v), font=("Segoe UI", 10, "bold"),
                     fg=C["txt_primary"], bg=bg, width=10,
                     anchor="center").grid(row=0, column=0, padx=2, pady=3)
            for if_v in range(6):
                val = TABLA_IS_PRIMER[id_v][if_v]
                _, clr, _ = clasificar_is(val)
                tk.Label(rf, text=str(val), font=("Segoe UI", 10, "bold"),
                         fg=clr, bg=bg, width=7,
                         anchor="center").grid(row=0, column=if_v+1, padx=2, pady=3)

    def _build_classification_card(self, parent):
        card, body = make_card(parent, "Clasificación del Estado del Pavimento")
        card.pack(fill="x", padx=15, pady=6)

        data = [
            ("1 – 2", "Bueno", C["green"], C["green_bg"],
             "Pavimento en buen estado. Necesita mantenimiento rutinario."),
            ("3 – 4", "Marginal", C["orange"], C["orange_bg"],
             "Pavimento con deterioro moderado. Requiere mantenimiento preventivo o correctivo."),
            ("5 – 6 – 7", "Deficiente", C["red"], C["red_bg"],
             "Pavimento con deterioro severo. Necesita rehabilitación o reconstrucción."),
        ]
        for rng, clasif, clr, bgc, desc in data:
            rf = tk.Frame(body, bg=bgc, highlightbackground=clr,
                          highlightthickness=1)
            rf.pack(fill="x", pady=3, ipady=6)
            tk.Label(rf, text=f"  Is = {rng}", font=("Segoe UI", 12, "bold"),
                     fg=clr, bg=bgc, width=14,
                     anchor="w").pack(side="left", padx=(8, 0))
            tk.Label(rf, text=clasif, font=("Segoe UI", 12, "bold"),
                     fg=clr, bg=bgc, width=12,
                     anchor="center").pack(side="left")
            tk.Label(rf, text=desc, font=("Segoe UI", 10),
                     fg=C["txt_primary"], bg=bgc,
                     anchor="w").pack(side="left", padx=15)

    def _populate_procedimientos(self):
        """Adds dynamic calculated values to procedures tab"""
        for w in self.proc_dynamic_frame.winfo_children():
            w.destroy()

        if not self._has_resultados_disponibles():
            return

        r = self._get_resultado_visible()
        if not r:
            messagebox.showwarning("Aviso", "No hay resultados para exportar.")
            return
        def add_section(parent, title, lines, color=C["accent"]):
            box = tk.Frame(parent, bg=C["bg_card"])
            box.pack(fill="x", pady=4)
            tk.Label(box, text=title,
                     font=("Segoe UI", 11, "bold"),
                     fg=color, bg=C["bg_card"]).pack(anchor="w")
            for line in lines:
                tk.Label(box, text=line,
                         font=("Segoe UI", 10),
                         fg=C["txt_second"], bg=C["bg_card"],
                         wraplength=860, justify="left").pack(anchor="w", padx=(16, 0), pady=(2, 0))

        colors = {
            "Resultado Visible": C["accent"],
            "Formulas Aplicadas": C["accent"],
            "Fisuracion (If)": C["accent"],
            "Deformacion (Id)": C["accent_alt"],
            "Primer Is": C["gold"],
            "Correccion por Reparacion": C["red"],
            "Resultado Final": C["gold"],
        }
        card_num = None
        body_num = None
        for title, lines in self._procedimiento_sections(r):
            if title in ("Resultado Visible", "Formulas Aplicadas"):
                card, body = make_card(self.proc_dynamic_frame, title)
                card.pack(fill="x", pady=6)
                for line in lines:
                    tk.Label(body, text=line,
                             font=("Segoe UI", 10),
                             fg=C["txt_second"], bg=C["bg_card"],
                             wraplength=880, justify="left").pack(anchor="w", pady=2)
            else:
                if body_num is None:
                    card_num, body_num = make_card(self.proc_dynamic_frame, "Desarrollo Numerico")
                    card_num.pack(fill="x", pady=6)
                add_section(body_num, title, lines, color=colors.get(title, C["accent"]))
        return

        vista_activa = self.v_resultado_sel.get() or f"Actual: {r.get('progresiva', 'Todo el archivo')}"

        card_r, body_r = make_card(self.proc_dynamic_frame, "Resultado Visible")
        card_r.pack(fill="x", pady=6)
        for line in [
            f"Vista activa: {vista_activa}",
            f"Unidad de Muestreo: {self.var_um.get() or '-'}",
            f"Area de muestra: {r.get('area', 0.0):.2f} m2",
            f"Progresiva analizada: {r.get('progresiva', 'Todo el archivo')}",
            f"Registros analizados: {r.get('n_reg', 0)}   |   Tramos: {r.get('n_tramos', 0)}",
        ]:
            tk.Label(body_r, text=line,
                     font=("Segoe UI", 10),
                     fg=C["txt_second"], bg=C["bg_card"]).pack(anchor="w", pady=2)

        card_fm, body_fm = make_card(self.proc_dynamic_frame, "Formulas Aplicadas")
        card_fm.pack(fill="x", pady=6)
        for line in [
            "Promedio ponderado de gravedad: G = sum(gravedad_i * cantidad_i) / sum(cantidad_i)",
            "Cada termino usa la gravedad de la fila multiplicada por la cantidad total de esa falla.",
            "Gravedad representativa: G < 1.5 -> 1, 1.5 <= G < 2.5 -> 2, G >= 2.5 -> 3",
            "Extension acumulada: E = suma de extensiones del grupo",
            "If = TABLA_IF[Gravedad representativa][Categoria de extension]",
            "Id = TABLA_ID[Gravedad representativa][Categoria de extension]",
            "Primer Is = TABLA_IS_PRIMER[Id][If]",
            "Correccion = TABLA_CORRECCION[Gravedad representativa de B][Categoria de extension de B]",
            "Is final = min(Primer Is + Correccion, 7)",
        ]:
            tk.Label(body_fm, text=line,
                     font=("Segoe UI", 10),
                     fg=C["txt_second"], bg=C["bg_card"],
                     wraplength=880, justify="left").pack(anchor="w", pady=2)

        card_num, body_num = make_card(self.proc_dynamic_frame, "Desarrollo Numerico")
        card_num.pack(fill="x", pady=6)

        ec_f, lbl_f = self._categoria_extension_info(r.get("ext_fis", 0.0))
        ec_d, lbl_d = self._categoria_extension_info(r.get("ext_def", 0.0))
        ec_r, lbl_r = self._categoria_extension_info(r.get("ext_rep", 0.0))

        add_section(body_num, "Fisuracion (If)", [
            *self._gravedad_formula_lines(
                r.get("grav_fis", 0), r.get("grav_prom_fis", 0.0),
                r.get("s1_fis", 0), r.get("s2_fis", 0), r.get("s3_fis", 0),
                fallas=[d for d in r.get("fallas_a", []) if d.get("codigo") in ("FPC", "FLT")]
            ),
            f"Extension acumulada = {r.get('ext_fis', 0.0):.2f}% -> categoria {ec_f} ({lbl_f})",
            f"If = TABLA_IF[G={r.get('grav_fis', 0)}][E={ec_f}] = {r.get('if', 0)}",
        ])

        add_section(body_num, "Deformacion (Id)", [
            *self._gravedad_formula_lines(
                r.get("grav_def", 0), r.get("grav_prom_def", 0.0),
                r.get("s1_def", 0), r.get("s2_def", 0), r.get("s3_def", 0),
                fallas=[d for d in r.get("fallas_a", []) if d.get("codigo") in ("AH", "DL", "DT")]
            ),
            f"Extension acumulada = {r.get('ext_def', 0.0):.2f}% -> categoria {ec_d} ({lbl_d})",
            f"Id = TABLA_ID[G={r.get('grav_def', 0)}][E={ec_d}] = {r.get('id', 0)}",
        ], color=C["accent_alt"])

        add_section(body_num, "Primer Is", [
            f"Primer Is = TABLA_IS_PRIMER[Id={r.get('id', 0)}][If={r.get('if', 0)}] = {r.get('is_primer', 0)}",
        ], color=C["gold"])

        if r.get("s1_rep", 0) + r.get("s2_rep", 0) + r.get("s3_rep", 0) > 0:
            rep_lines = [
                *self._gravedad_formula_lines(
                    r.get("grav_rep", 0), r.get("grav_prom_rep", 0.0),
                    r.get("s1_rep", 0), r.get("s2_rep", 0), r.get("s3_rep", 0),
                    fallas=[d for d in r.get("fallas_a", []) if d.get("codigo") == "B"]
                ),
                f"Extension acumulada = {r.get('ext_rep', 0.0):.2f}% -> categoria {ec_r} ({lbl_r})",
                f"Correccion = TABLA_CORRECCION[G={r.get('grav_rep', 0)}][E={ec_r}] = {r.get('correccion', 0)}",
            ]
        else:
            rep_lines = ["No se registraron baches y parches (B), por lo tanto la correccion es 0."]
        add_section(body_num, "Correccion por Reparacion", rep_lines, color=C["red"])

        add_section(body_num, "Resultado Final", [
            f"Is final = min({r.get('is_primer', 0)} + {r.get('correccion', 0)}, 7) = {r.get('is_final', 0)}",
            f"Clasificacion = {r.get('clasificacion', '').upper()}",
        ], color=C["gold"])
        return
        card, body = make_card(self.proc_dynamic_frame,
                                "Valores Calculados para esta Unidad de Muestreo")
        card.pack(fill="x", pady=6)

        info = [
            f"Unidad de Muestreo: {self.var_um.get() or '—'}",
            f"Área de muestra: {r['area']} m²",
            f"Progresiva analizada: {r.get('progresiva', 'Todo el archivo')}",
            "",
            f"Gravedad de fisuración: {r['grav_fis']}   |   Extensión: {r['ext_fis']:.2f}%   →   If = {r['if']}",
            f"Gravedad de deformación: {r['grav_def']}   |   Extensión: {r['ext_def']:.2f}%   →   Id = {r['id']}",
            "",
            f"Primer Is (de tabla If={r['if']}, Id={r['id']}): {r['is_primer']}",
            f"Corrección por reparación: {r['correccion']}",
            f"Is Final = {r['is_primer']} + {r['correccion']} = {r['is_final']}   →   {r['clasificacion'].upper()}",
        ]

        for line in info:
            if not line:
                tk.Frame(body, bg=C["bg_card"], height=6).pack()
                continue
            bold = "→" in line or "If =" in line or "Id =" in line or "Is Final" in line
            clr = C["gold"] if "Is Final" in line else (C["accent"] if bold else C["txt_second"])
            tk.Label(body, text=line,
                     font=("Segoe UI", 11 if bold else 10, "bold" if bold else ""),
                     fg=clr, bg=C["bg_card"]).pack(anchor="w", pady=2)

    # ══════════════════════════════════════════════
    #  TAB 4: RESULTADOS
    # ══════════════════════════════════════════════
    def _build_tab_resultados(self, parent):
        sf = ScrollFrame(parent, bg=C["bg_main"])
        sf.pack(fill="both", expand=True)
        self.res_inner = sf.inner

        # Export bar
        self.res_export_bar = tk.Frame(self.res_inner, bg=C["bg_main"])
        self.res_export_bar.pack(fill="x", padx=15, pady=12)

        make_button(self.res_export_bar, "💾  Exportar Gráficos (Color)",
                    lambda: self._export_charts(True), style="normal").pack(side="left", padx=4)

        make_button(self.res_export_bar, "📊  Exportar Resumen Excel",
                    self._export_excel, style="green").pack(side="left", padx=15)

        tk.Frame(self.res_export_bar, bg=C["bg_main"], width=10).pack(side="left")
        self.cmb_resultado_sel = self._register_resultado_selector(self.res_export_bar)

        self.res_content = tk.Frame(self.res_inner, bg=C["bg_main"])
        self.res_content.pack(fill="both", expand=True, padx=15)

    def _refresh_resultado_selector(self, prefer=None):
        opciones = []
        mapping = {}
        if self.resultados:
            label_actual = f"Actual: {self.resultados.get('progresiva', 'Todo el archivo')}"
            opciones.append(label_actual)
            mapping[label_actual] = self.resultados
        for tramo in self.resultados_tramos:
            label = tramo.get("selector_label") or f"Tramo: {tramo.get('progresiva', '')}"
            if not tramo.get("progresiva"):
                continue
            opciones.append(label)
            mapping[label] = tramo
        for blk in self.resultados_bloques:
            label = blk.get("progresiva", "")
            if not label:
                continue
            opciones.append(label)
            mapping[label] = blk
        self.resultado_selector_map = mapping
        self.resultado_selector_widgets = [
            w for w in self.resultado_selector_widgets
            if w.winfo_exists()
        ]
        for combo in self.resultado_selector_widgets:
            combo["values"] = opciones
        actual = prefer or self.v_resultado_sel.get()
        if actual not in mapping and opciones:
            actual = opciones[0]
        self.v_resultado_sel.set(actual if opciones else "")
        self._update_resultado_is_badges()

    def _has_resultados_disponibles(self):
        return bool(self.resultados or self.resultados_bloques)

    def _get_resultado_visible(self):
        self._refresh_resultado_selector()
        visible = self.resultado_selector_map.get(self.v_resultado_sel.get())
        if visible:
            return visible
        if self.resultados:
            return self.resultados
        return self.resultados_bloques[0] if self.resultados_bloques else None

    def _obtener_paquetes_exportacion(self):
        if self.resultados_bloques:
            return "bloques", list(self.resultados_bloques)

        cfg = {}
        try:
            cfg = self.panel_prog.obtener_config()
        except Exception:
            cfg = {}

        if cfg.get("modo") == "bloques" and self.datos_procesados_base:
            ini, fin, tam_bloque = self._resolver_limites_bloques()
            try:
                bloques, _ = AnalizadorProgresivasVIZIR.bloques_disponibles(
                    self.datos_procesados_base, tam_bloque, ini, fin
                )
            except Exception:
                bloques = []

            paquetes = []
            for bloque in bloques:
                datos = AnalizadorProgresivasVIZIR.filtrar_registros(
                    self.datos_procesados_base, bloque["ini"], bloque["fin"]
                )
                if not datos:
                    continue
                ctx = self._build_prog_ctx("bloques", bloque["ini"], bloque["fin"], datos, bloque["label"])
                resultado = self._calcular_resultados_vizir(datos, ctx)
                resultado["n_reg_bloque"] = bloque["n_reg"]
                resultado["n_tramos_bloque"] = bloque["n_tramos"]
                paquetes.append(resultado)
            if paquetes:
                return "bloques", paquetes

        if self.resultados_tramos:
            return "tramos", list(self.resultados_tramos)

        r = self._get_resultado_visible()
        if r:
            return "actual", [r]
        return "actual", []

    def _on_resultado_selected(self):
        self._update_resultado_is_badges()
        self._populate_calculos()
        self._populate_procedimientos()
        self._populate_resultados()

    def _populate_resultados(self):
        for w in self.res_content.winfo_children():
            w.destroy()

        if not self._has_resultados_disponibles():
            return

        r = self._get_resultado_visible()
        if not r:
            return

        # ── Summary card ──
        card_s, body_s = make_card(self.res_content, "Resumen Final")
        card_s.pack(fill="x", pady=6)

        info_row = tk.Frame(body_s, bg=C["bg_card"])
        info_row.pack(fill="x", pady=4)

        items = [
            ("If", str(r["if"]), C["accent"]),
            ("Id", str(r["id"]), C["accent_alt"]),
            ("Primer Is", str(r["is_primer"]), C["gold"]),
            ("Corrección", str(r["correccion"]), C["red"]),
        ]
        for i, (lbl, val, clr) in enumerate(items):
            f_item = tk.Frame(info_row, bg=C["bg_input"],
                               highlightbackground=C["border"], highlightthickness=1)
            f_item.pack(side="left", padx=6, ipadx=14, ipady=6)
            tk.Label(f_item, text=lbl, font=("Segoe UI", 9),
                     fg=C["txt_dim"], bg=C["bg_input"]).pack()
            tk.Label(f_item, text=val, font=("Segoe UI", 18, "bold"),
                     fg=clr, bg=C["bg_input"]).pack()

        # Final Is
        is_f = r["is_final"]
        clasif, clr, clr_bg = clasificar_is(is_f)
        result_box = tk.Frame(body_s, bg=clr_bg,
                               highlightbackground=clr, highlightthickness=2)
        result_box.pack(fill="x", pady=10, ipady=14)
        tk.Label(result_box,
                 text=f"ÍNDICE DE DETERIORO SUPERFICIAL (Is) = {is_f}   →   {clasif.upper()}",
                 font=("Segoe UI", 18, "bold"),
                 fg=clr, bg=clr_bg).pack()

        # Project info
        proj_info = tk.Frame(body_s, bg=C["bg_card"])
        proj_info.pack(fill="x", pady=4)
        pi_items = [
            ("Vía", self.var_via.get()),
            ("Tramo", self.var_tramo.get()),
            ("U.M.", self.var_um.get()),
            ("Área", f"{r['area']:.2f} m²"),
            ("Progresiva", r.get("progresiva", "Todo el archivo")),
        ]
        for lbl, val in pi_items:
            if val:
                tk.Label(proj_info, text=f"{lbl}: {val}",
                         font=("Segoe UI", 10), fg=C["txt_second"],
                         bg=C["bg_card"]).pack(side="left", padx=12)

        # ── Charts ──
        card_ch, body_ch = make_card(self.res_content, "Diagramas de Torta")
        card_ch.pack(fill="x", pady=6)

        self.current_fig = self._create_pie_charts(body_ch, color_mode=True)

    def _create_pie_charts(self, parent, color_mode=True):
        r = self._get_resultado_visible()
        if not r:
            return None
        fig = Figure(figsize=(14, 5), dpi=100, facecolor=C["bg_card"])

        # ── Pie 1: Tipo A vs B ──
        ax1 = fig.add_subplot(131)
        n_a = len(r["fallas_a"])
        n_b = len(r["fallas_b"])
        if n_a + n_b > 0:
            if color_mode:
                colors1 = ['#ef4444', '#3b82f6']
            else:
                colors1 = ['#888888', '#cccccc']
            wedges1, _, auto1 = ax1.pie(
                [n_a, n_b],
                labels=[f'Tipo A\n({n_a})', f'Tipo B\n({n_b})'],
                colors=colors1, autopct='%1.1f%%', startangle=90,
                textprops={'fontsize': 9, 'color': C["txt_primary"]},
                pctdistance=0.6)
            for at in auto1:
                at.set_fontweight('bold')
            if not color_mode:
                hatches = ['///', '...']
                for i, w in enumerate(wedges1):
                    w.set_hatch(hatches[i])
                    w.set_edgecolor('white')
        ax1.set_title('Por Tipo de Deterioro', fontsize=11,
                      fontweight='bold', color=C["txt_primary"], pad=12)
        ax1.set_facecolor(C["bg_card"])

        # ── Pie 2: Gravedad ──
        ax2 = fig.add_subplot(132)
        sev = defaultdict(int)
        for d in r["datos"]:
            sev[d["gravedad"]] += 1
        if sev:
            sev_names = {1: 'Bajo', 2: 'Regular', 3: 'Alto'}
            labels2 = [f"{sev_names.get(s, s)} ({sev[s]})" for s in sorted(sev)]
            sizes2 = [sev[s] for s in sorted(sev)]
            if color_mode:
                colors2 = ['#22c55e', '#f59e0b', '#ef4444'][:len(sizes2)]
            else:
                colors2 = ['#666666', '#999999', '#cccccc'][:len(sizes2)]
            wedges2, _, auto2 = ax2.pie(
                sizes2, labels=labels2, colors=colors2,
                autopct='%1.1f%%', startangle=90,
                textprops={'fontsize': 9, 'color': C["txt_primary"]},
                pctdistance=0.6)
            for at in auto2:
                at.set_fontweight('bold')
            if not color_mode:
                hats = ['///', '\\\\\\', 'xxx']
                for i, w in enumerate(wedges2):
                    w.set_hatch(hats[i % len(hats)])
                    w.set_edgecolor('white')
        ax2.set_title('Por Nivel de Gravedad', fontsize=11,
                      fontweight='bold', color=C["txt_primary"], pad=12)
        ax2.set_facecolor(C["bg_card"])

        # ── Pie 3: Por código ──
        ax3 = fig.add_subplot(133)
        all_fallas_chart = {**FALLAS_TIPO_A, **FALLAS_TIPO_B}
        code_t = defaultdict(float)
        for d in r["datos"]:
            code_t[d["codigo"]] += d["total"]
        if code_t:
            sorted_c = sorted(code_t.items(), key=lambda x: x[1], reverse=True)
            labels3 = [f"{c} ({fmt_total_str(v, all_fallas_chart.get(c, {}).get('unidad', ''))})"
                       for c, v in sorted_c]
            sizes3 = [v for _, v in sorted_c]
            if color_mode:
                cmap = plt.cm.Set2
                colors3 = [cmap(i / max(len(sorted_c), 1)) for i in range(len(sorted_c))]
            else:
                colors3 = [f"#{hex(60 + i * 35)[2:]*3}" if (60 + i*35) < 256
                           else '#dddddd' for i in range(len(sorted_c))]
                # Better grayscale
                n = len(sorted_c)
                colors3 = [f'#{hex(min(80 + i * (140 // max(n, 1)), 220))[2:]*3}'
                           for i in range(n)]
            wedges3, _, auto3 = ax3.pie(
                sizes3, labels=labels3, colors=colors3,
                autopct='%1.1f%%', startangle=90,
                textprops={'fontsize': 8, 'color': C["txt_primary"]},
                pctdistance=0.7)
            for at in auto3:
                at.set_fontweight('bold')
                at.set_fontsize(8)
            if not color_mode:
                hats = ['///', '\\\\\\', 'xxx', '...', '+++', 'ooo', '---', '|||']
                for i, w in enumerate(wedges3):
                    w.set_hatch(hats[i % len(hats)])
                    w.set_edgecolor('white')
        ax3.set_title('Por Tipo de Falla (Cantidad)', fontsize=11,
                      fontweight='bold', color=C["txt_primary"], pad=12)
        ax3.set_facecolor(C["bg_card"])

        fig.tight_layout(pad=2.5)

        canvas = FigureCanvasTkAgg(fig, master=parent)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True, pady=6)

        return fig

    def _export_charts(self, color_mode):
        if not self._has_resultados_disponibles():
            messagebox.showwarning("Aviso", "Ejecute el cálculo primero.")
            return

        suffix = "color" if color_mode else "bn"
        fp = filedialog.asksaveasfilename(
            title="Guardar gráficos",
            defaultextension=".png",
            filetypes=[("PNG", "*.png"), ("JPEG", "*.jpg"),
                       ("PDF", "*.pdf"), ("SVG", "*.svg")],
            initialfile=f"VIZIR_graficos_{suffix}")
        if not fp:
            return

        # Create figure for export with white or dark bg
        export_fig = Figure(figsize=(14, 5), dpi=200)

        if color_mode:
            export_fig.set_facecolor('white')
            txt_color = '#222222'
        else:
            export_fig.set_facecolor('white')
            txt_color = '#111111'

        r = self._get_resultado_visible()
        if not r:
            messagebox.showwarning("Aviso", "No hay resultados para exportar.")
            return

        # Pie 1
        ax1 = export_fig.add_subplot(131)
        n_a = len(r["fallas_a"])
        n_b = len(r["fallas_b"])
        if n_a + n_b > 0:
            if color_mode:
                c1 = ['#ef4444', '#3b82f6']
            else:
                c1 = ['#888888', '#cccccc']
            w1, _, a1 = ax1.pie([n_a, n_b],
                                 labels=[f'Tipo A ({n_a})', f'Tipo B ({n_b})'],
                                 colors=c1, autopct='%1.1f%%', startangle=90,
                                 textprops={'fontsize': 9, 'color': txt_color},
                                 pctdistance=0.6)
            for at in a1:
                at.set_fontweight('bold')
            if not color_mode:
                for i, w in enumerate(w1):
                    w.set_hatch(['///', '...'][i])
                    w.set_edgecolor('black')
        ax1.set_title('Por Tipo de Deterioro', fontsize=11,
                      fontweight='bold', color=txt_color, pad=12)

        # Pie 2
        ax2 = export_fig.add_subplot(132)
        sev = defaultdict(int)
        for d in r["datos"]:
            sev[d["gravedad"]] += 1
        if sev:
            sev_names = {1: 'Bajo', 2: 'Regular', 3: 'Alto'}
            lb2 = [f"{sev_names.get(s, s)} ({sev[s]})" for s in sorted(sev)]
            sz2 = [sev[s] for s in sorted(sev)]
            if color_mode:
                c2 = ['#22c55e', '#f59e0b', '#ef4444'][:len(sz2)]
            else:
                c2 = ['#666666', '#999999', '#cccccc'][:len(sz2)]
            w2, _, a2 = ax2.pie(sz2, labels=lb2, colors=c2,
                                 autopct='%1.1f%%', startangle=90,
                                 textprops={'fontsize': 9, 'color': txt_color},
                                 pctdistance=0.6)
            for at in a2:
                at.set_fontweight('bold')
            if not color_mode:
                hats = ['///', '\\\\\\', 'xxx']
                for i, w in enumerate(w2):
                    w.set_hatch(hats[i % len(hats)])
                    w.set_edgecolor('black')
        ax2.set_title('Por Nivel de Gravedad', fontsize=11,
                      fontweight='bold', color=txt_color, pad=12)

        # Pie 3
        ax3 = export_fig.add_subplot(133)
        all_fallas_chart2 = {**FALLAS_TIPO_A, **FALLAS_TIPO_B}
        code_t = defaultdict(float)
        for d in r["datos"]:
            code_t[d["codigo"]] += d["total"]
        if code_t:
            sc = sorted(code_t.items(), key=lambda x: x[1], reverse=True)
            lb3 = [f"{c} ({fmt_total_str(v, all_fallas_chart2.get(c, {}).get('unidad', ''))})"
                   for c, v in sc]
            sz3 = [v for _, v in sc]
            if color_mode:
                cmap = plt.cm.Set2
                c3 = [cmap(i / max(len(sc), 1)) for i in range(len(sc))]
            else:
                n = len(sc)
                c3 = [f'#{hex(min(80 + i * (140 // max(n, 1)), 220))[2:]*3}'
                      for i in range(n)]
            w3, _, a3 = ax3.pie(sz3, labels=lb3, colors=c3,
                                 autopct='%1.1f%%', startangle=90,
                                 textprops={'fontsize': 8, 'color': txt_color},
                                 pctdistance=0.7)
            for at in a3:
                at.set_fontweight('bold')
                at.set_fontsize(8)
            if not color_mode:
                hats = ['///', '\\\\\\', 'xxx', '...', '+++', 'ooo']
                for i, w in enumerate(w3):
                    w.set_hatch(hats[i % len(hats)])
                    w.set_edgecolor('black')
        ax3.set_title('Por Tipo de Falla', fontsize=11,
                      fontweight='bold', color=txt_color, pad=12)

        export_fig.tight_layout(pad=2.5)
        export_fig.savefig(fp, dpi=200, bbox_inches='tight',
                            facecolor='white', edgecolor='none')
        plt.close(export_fig)
        messagebox.showinfo("Exportado", f"Gráficos guardados en:\n{fp}")

    def _export_excel_legacy(self):
        if not self._has_resultados_disponibles():
            messagebox.showwarning("Aviso", "Ejecute el cálculo primero.")
            return

        fp = filedialog.asksaveasfilename(
            title="Guardar resultados",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="VIZIR_Resultados")
        if not fp:
            return

        r = self._get_resultado_visible()
        if not r:
            messagebox.showwarning("Aviso", "No hay resultados para exportar.")
            return
        wb = openpyxl.Workbook()
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        bold_font = Font(bold=True, size=11)
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="1a2747", end_color="1a2747", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))

        # Sheet 1: Resumen
        ws = wb.active
        ws.title = "Resumen VIZIR"
        ws.append(["RESULTADOS DEL ANÁLISIS VIZIR"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])
        ws.append(["Proyecto:", self.var_proyecto.get()])
        ws.append(["Vía:", self.var_via.get()])
        ws.append(["Evaluado por:", self.var_evaluador.get()])
        ws.append(["Tramo:", self.var_tramo.get()])
        ws.append(["Unidad de Muestreo:", self.var_um.get()])
        ws.append(["Largo (m):", self._get_float_var(self.var_largo, 0.0)])
        ws.append(["Ancho (m):", self._get_float_var(self.var_ancho, 0.0)])
        ws.append(["Área de Muestra (m²):", r["area"]])
        ws.append(["Progresiva Analizada:", r.get("progresiva", "Todo el archivo")])
        ws.append([])
        ws.append(["Parámetro", "Valor"])
        ws["A12"].font = bold_font
        ws["B12"].font = bold_font
        ws.append(["Índice de Fisuración (If)", r["if"]])
        ws.append(["Índice de Deformación (Id)", r["id"]])
        ws.append(["Primer Is", r["is_primer"]])
        ws.append(["Corrección por Reparación", r["correccion"]])
        ws.append(["Is Final", r["is_final"]])
        ws.append(["Clasificación", r["clasificacion"]])

        # Sheet 2: Detalle
        ws2 = wb.create_sheet("Detalle Fallas")
        headers = ["Falla", "Tipo", "Código", "Unidad", "Gravedad", "Total", "Extensión (%)"]
        ws2.append(headers)
        for i, h in enumerate(headers):
            cell = ws2.cell(row=1, column=i+1)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        for d in r["datos"]:
            ep = d['extension'] * 100 if d['extension'] <= 1 else d['extension']
            ws2.append([d["falla"], d["tipo"], d["codigo"], d["unidad"],
                        d["gravedad"], fmt_total(d["total"], d.get("unidad", "")),
                        round(ep, 4)])

        # Sheet 3: Tablas
        ws3 = wb.create_sheet("Tablas VIZIR")
        ws3.append(["TABLA DE ÍNDICE DE FISURACIÓN (If)"])
        ws3["A1"].font = bold_font
        ws3.append(["Gravedad", "0 a 10%", "10 a 50%", "> 50%"])
        for g in [1, 2, 3]:
            ws3.append([g, TABLA_IF[g][1], TABLA_IF[g][2], TABLA_IF[g][3]])
        ws3.append([])
        ws3.append(["TABLA DE ÍNDICE DE DEFORMACIÓN (Id)"])
        ws3.append(["Gravedad", "0 a 10%", "10 a 50%", "> 50%"])
        for g in [1, 2, 3]:
            ws3.append([g, TABLA_ID[g][1], TABLA_ID[g][2], TABLA_ID[g][3]])
        ws3.append([])
        ws3.append(["TABLA PRIMER VALOR Is"])
        ws3.append(["Id \\ If", 0, 1, 2, 3, 4, 5])
        for id_v in range(6):
            ws3.append([id_v] + [TABLA_IS_PRIMER[id_v][f] for f in range(6)])
        ws3.append([])
        ws3.append(["TABLA DE CORRECCIÓN POR REPARACIÓN"])
        ws3.append(["Gravedad", "0 a 10%", "10 a 50%", "> 50%"])
        for g in [1, 2, 3]:
            ws3.append([g, TABLA_CORRECCION[g][1], TABLA_CORRECCION[g][2],
                        TABLA_CORRECCION[g][3]])

        if self.resultados_bloques:
            ws4 = wb.create_sheet("Bloques Progresiva")
            headers4 = ["Bloque", "Prog. Ini", "Prog. Fin", "N Reg", "N Tramos", "Area (m2)", "If", "Id", "Primer Is", "Correccion", "Is Final", "Clasificacion"]
            ws4.append(headers4)
            for i, h in enumerate(headers4):
                cell = ws4.cell(row=1, column=i+1)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            for blk in self.resultados_bloques:
                ws4.append([
                    blk.get("progresiva", ""),
                    metros_a_prog(blk.get("prog_ini")),
                    metros_a_prog(blk.get("prog_fin")),
                    blk.get("n_reg", 0),
                    blk.get("n_tramos", 0),
                    round(float(blk.get("area", 0.0)), 2),
                    blk.get("if", 0),
                    blk.get("id", 0),
                    blk.get("is_primer", 0),
                    blk.get("correccion", 0),
                    blk.get("is_final", 0),
                    blk.get("clasificacion", ""),
                ])

        # Adjust widths
        hojas = [ws, ws2, ws3]
        if self.resultados_bloques:
            hojas.append(ws4)
        for ws_s in hojas:
            for col in ws_s.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws_s.column_dimensions[col_letter].width = min(max_len + 4, 50)

        wb.save(fp)
        messagebox.showinfo("Exportado", f"Resultados guardados en:\n{fp}")

    def _export_excel(self):
        if not self._has_resultados_disponibles():
            messagebox.showwarning("Aviso", "Ejecute el calculo primero.")
            return

        fp = filedialog.asksaveasfilename(
            title="Guardar resultados",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="VIZIR_Resultados")
        if not fp:
            return

        modo_export, paquetes_export = self._obtener_paquetes_exportacion()
        if not paquetes_export:
            messagebox.showwarning("Aviso", "No hay resultados para exportar.")
            return
        r_visible = self._get_resultado_visible()
        r = r_visible if r_visible else paquetes_export[0]

        wb = openpyxl.Workbook()
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        bold_font = Font(bold=True, size=11)
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="1a2747", end_color="1a2747", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))

        # Sheet 1: Resumen
        ws = wb.active
        ws.title = "Resumen VIZIR"
        ws.append(["RESULTADOS DEL ANALISIS VIZIR"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])
        ws.append(["Proyecto:", self.var_proyecto.get()])
        ws.append(["Via:", self.var_via.get()])
        ws.append(["Evaluado por:", self.var_evaluador.get()])
        ws.append(["Tramo:", self.var_tramo.get()])
        ws.append(["Unidad de Muestreo:", self.var_um.get()])
        ws.append(["Largo (m):", self._get_float_var(self.var_largo, 0.0)])
        ws.append(["Ancho (m):", self._get_float_var(self.var_ancho, 0.0)])
        ws.append(["Area de Muestra (m2):", r["area"]])
        ws.append(["Progresiva Analizada:", r.get("progresiva", "Todo el archivo")])
        etiqueta_export = {
            "bloques": "Bloques exportados:",
            "tramos": "Tramos exportados:",
        }.get(modo_export, "Resultados exportados:")
        ws.append([etiqueta_export, len(paquetes_export)])
        if len(paquetes_export) > 1:
            ini_export = paquetes_export[0].get("prog_ini")
            fin_export = paquetes_export[-1].get("prog_fin")
            if ini_export is not None or fin_export is not None:
                ws.append([
                    "Rango exportado:",
                    f"{metros_a_prog(ini_export)} - {metros_a_prog(fin_export)}",
                ])
        ws.append([])
        ws.append(["Parametro", "Valor"])
        row_header = ws.max_row
        ws.cell(row=row_header, column=1).font = bold_font
        ws.cell(row=row_header, column=2).font = bold_font
        ws.append(["Indice de Fisuracion (If)", r["if"]])
        ws.append(["Indice de Deformacion (Id)", r["id"]])
        ws.append(["Primer Is", r["is_primer"]])
        ws.append(["Correccion por Reparacion", r["correccion"]])
        ws.append(["Is Final", r["is_final"]])
        ws.append(["Clasificacion", r["clasificacion"]])

        # Sheet 2: Detalle
        ws2 = wb.create_sheet("Detalle Fallas")
        headers = [
            "Bloque/Tramo", "Prog. Ini", "Prog. Fin",
            "Falla", "Tipo", "Codigo", "Unidad",
            "Gravedad", "Total", "Extension (%)"
        ]
        ws2.append(headers)
        for i, _h in enumerate(headers):
            cell = ws2.cell(row=1, column=i + 1)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        for paquete in paquetes_export:
            prog_desc = paquete.get("progresiva", "")
            prog_ini = metros_a_prog(paquete.get("prog_ini"))
            prog_fin = metros_a_prog(paquete.get("prog_fin"))
            for d in paquete.get("datos", []):
                ep = d['extension'] * 100 if d['extension'] <= 1 else d['extension']
                ws2.append([
                    prog_desc, prog_ini, prog_fin,
                    d["falla"], d["tipo"], d["codigo"], d["unidad"],
                    d["gravedad"], fmt_total(d["total"], d.get("unidad", "")),
                    round(ep, 4)
                ])

        # Sheet 3: Tablas
        ws3 = wb.create_sheet("Tablas VIZIR")
        ws3.append(["TABLA DE INDICE DE FISURACION (If)"])
        ws3["A1"].font = bold_font
        ws3.append(["Gravedad", "0 a 10%", "10 a 50%", "> 50%"])
        for g in [1, 2, 3]:
            ws3.append([g, TABLA_IF[g][1], TABLA_IF[g][2], TABLA_IF[g][3]])
        ws3.append([])
        ws3.append(["TABLA DE INDICE DE DEFORMACION (Id)"])
        ws3.append(["Gravedad", "0 a 10%", "10 a 50%", "> 50%"])
        for g in [1, 2, 3]:
            ws3.append([g, TABLA_ID[g][1], TABLA_ID[g][2], TABLA_ID[g][3]])
        ws3.append([])
        ws3.append(["TABLA PRIMER VALOR Is"])
        ws3.append(["Id \\ If", 0, 1, 2, 3, 4, 5])
        for id_v in range(6):
            ws3.append([id_v] + [TABLA_IS_PRIMER[id_v][f] for f in range(6)])
        ws3.append([])
        ws3.append(["TABLA DE CORRECCION POR REPARACION"])
        ws3.append(["Gravedad", "0 a 10%", "10 a 50%", "> 50%"])
        for g in [1, 2, 3]:
            ws3.append([g, TABLA_CORRECCION[g][1], TABLA_CORRECCION[g][2],
                        TABLA_CORRECCION[g][3]])

        ws4 = None
        if modo_export in ("bloques", "tramos") and paquetes_export:
            nombre_hoja = "Bloques Progresiva" if modo_export == "bloques" else "Tramos Progresiva"
            etiqueta = "Bloque" if modo_export == "bloques" else "Tramo"
            ws4 = wb.create_sheet(nombre_hoja)
            headers4 = [
                etiqueta, "Prog. Ini", "Prog. Fin", "N Reg", "N Tramos",
                "Area (m2)", "If", "Id", "Primer Is", "Correccion", "Is Final", "Clasificacion"
            ]
            ws4.append(headers4)
            for i, _h in enumerate(headers4):
                cell = ws4.cell(row=1, column=i + 1)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            for item in paquetes_export:
                ws4.append([
                    item.get("progresiva", ""),
                    metros_a_prog(item.get("prog_ini")),
                    metros_a_prog(item.get("prog_fin")),
                    item.get("n_reg_bloque", item.get("n_reg", 0)),
                    item.get("n_tramos_bloque", item.get("n_tramos", 0)),
                    round(float(item.get("area", 0.0)), 2),
                    item.get("if", 0),
                    item.get("id", 0),
                    item.get("is_primer", 0),
                    item.get("correccion", 0),
                    item.get("is_final", 0),
                    item.get("clasificacion", ""),
                ])

        ws5 = wb.create_sheet("Procedimientos")
        headers5 = ["Unidad", "Seccion", "Paso", "Detalle"]
        ws5.append(headers5)
        for i, _h in enumerate(headers5):
            cell = ws5.cell(row=1, column=i + 1)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        for paquete in paquetes_export:
            unidad = paquete.get("progresiva", "Todo el archivo")
            for seccion, lineas in self._procedimiento_sections(paquete, scope_label=unidad):
                for paso, detalle in enumerate(lineas, start=1):
                    ws5.append([unidad, seccion, paso, detalle])
            ws5.append([])

        # Adjust widths
        hojas = [ws, ws2, ws3, ws5]
        if ws4 is not None:
            hojas.append(ws4)
        for ws_s in hojas:
            for col in ws_s.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws_s.column_dimensions[col_letter].width = min(max_len + 4, 50)

        wb.save(fp)
        messagebox.showinfo("Exportado", f"Resultados guardados en:\n{fp}")


# ══════════════════════════════════════════════
#  PUNTO DE ENTRADA
# ══════════════════════════════════════════════
if __name__ == "__main__":
    root = tk.Tk()
    app = VIZIRApp(root)
    root.mainloop()
